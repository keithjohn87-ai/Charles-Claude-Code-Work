#!/usr/bin/env python3
"""
Build ContrPro GC Bid Estimator (XLSX) — GC suite (Complete tier).

Prime-contract bid roll-up for a General Contractor. The GC's job at bid time
is to assemble subcontractor quotes by CSI division, add self-perform (if any),
add Division 01 General Conditions, then layer overhead + profit + bond + tax
to arrive at the prime bid number to the Owner.

This is the trade-agnostic cousin of Steel Erection's Bid Estimator: where the
Steel Estimator computes from-scratch (tons × hours × rate), the GC Estimator
ROLLS UP sub quotes for trade work and self-performs Division 01 in detail.

Tabs:
  1. Instructions
  2. Project Info             (project + delivery method + schedule + markups)
  3. Trade Bids               (per CSI division — sub quotes + markup)
  4. Self-Perform             (if any — labor/material/equipment by CSI)
  5. General Conditions       (Division 01 — supervision, facilities, utilities)
  6. Allowances & Contingencies
  7. Bid Summary              (roll-up → OH + profit + bond + tax → final bid)
  8. CSI Reference            (hidden)

Run:
    /Users/home/charles/.venv/bin/python3 \\
        /Users/home/charles/contrpro/scripts/build_gc_bid_estimator_xlsx.py

Output:
    /Users/home/charles/contrpro/files/packages/complete/gc/GC_Bid_Estimator.xlsx
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
ACCENT_GOLD = "C9A227"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
YELLOW_FILL = "FFEB9C"
YELLOW_FONT = "9C5700"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_SUMMARY_LABEL = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW_FILL)
FILL_GOLD = PatternFill("solid", fgColor=ACCENT_GOLD)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H1 = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_BIG_NUMBER = Font(name="Calibri", size=18, bold=True, color=BRAND_BLUE)
FONT_GREY_ITALIC = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_USD0 = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
FMT_PCT = "0.0%"
FMT_INT = "0"
FMT_NUM2 = "0.00"
FMT_DATE = "yyyy-mm-dd"

# Full CSI MasterFormat divisions (Div 01-49 active)
DIVISIONS = [
    ("01", "General Requirements"),
    ("02", "Existing Conditions"),
    ("03", "Concrete"),
    ("04", "Masonry"),
    ("05", "Metals"),
    ("06", "Wood, Plastics, and Composites"),
    ("07", "Thermal and Moisture Protection"),
    ("08", "Openings"),
    ("09", "Finishes"),
    ("10", "Specialties"),
    ("11", "Equipment"),
    ("12", "Furnishings"),
    ("13", "Special Construction"),
    ("14", "Conveying Equipment"),
    ("21", "Fire Suppression"),
    ("22", "Plumbing"),
    ("23", "HVAC"),
    ("25", "Integrated Automation"),
    ("26", "Electrical"),
    ("27", "Communications"),
    ("28", "Electronic Safety and Security"),
    ("31", "Earthwork"),
    ("32", "Exterior Improvements"),
    ("33", "Utilities"),
    ("34", "Transportation"),
    ("35", "Waterway and Marine Construction"),
]


def set_col_widths(ws, widths):
    for col, w in widths:
        ws.column_dimensions[col].width = w


def style_header_row(ws, row, cols, start_col=1):
    for c in range(start_col, start_col + cols):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


def apply_body_style(ws, row, cols, start_col=1):
    for c in range(start_col, start_col + cols):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_BODY
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.border = BORDER


# ---------------------------------------------------------------------------
# Tab 1: Instructions
# ---------------------------------------------------------------------------

def build_instructions(ws):
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 110)])

    ws["A1"] = "GC BID ESTIMATOR — Instructions"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 36

    rows = [
        "",
        "WHAT THIS WORKBOOK IS",
        "Prime-contract bid roll-up for a General Contractor. Use to assemble subcontractor quotes by CSI "
        "division, layer General Conditions (Div 01), then apply overhead, profit, bond, and tax to land "
        "the final bid number you submit to the Owner.",
        "",
        "Unlike trade-specific bid estimators (Steel Erection, Electrical, etc.) which compute labor + "
        "material + equipment from scratch, the GC Estimator rolls up sub quotes you received from your "
        "trade partners. The detail estimating happens inside the trade-specific workbooks; this workbook "
        "is the consolidation layer.",
        "",
        "WORKFLOW",
        "1. Project Info: project, owner, architect, delivery method, schedule, markup defaults.",
        "2. Trade Bids: one row per CSI division. Enter the lowest qualifying sub quote per trade. Workbook "
        "computes GC markup on subs (typically 5-10%) plus optional contingency per trade.",
        "3. Self-Perform: any work the GC self-performs (rare on true GC; common on design-build / CMAR). "
        "Enter as labor + material + equipment cost. Marked up with self-perform OH + profit.",
        "4. General Conditions: Division 01 — supervision, project office, dumpsters, port-a-john, signage, "
        "site security, temporary utilities, etc. Itemized monthly/lump-sum.",
        "5. Allowances & Contingencies: explicit allowances called out in the bid + project-level "
        "contingency.",
        "6. Bid Summary: roll-up → corporate overhead + profit + payment-and-performance bond + sales tax "
        "→ final bid number.",
        "",
        "MARKUP STRUCTURE (LOCK IT IN ONE WAY OR THE OTHER)",
        "Two conventions exist. Pick one and zero out the other to avoid double-counting:",
        "  (A) Line-level markup ONLY: each trade gets its GC markup at the line item; no corporate OH+P at "
        "summary. Defaults: Sub Markup 7%, Self-Perform OH 10% + Profit 8%. Summary OH and Profit = 0%.",
        "  (B) Summary-level markup ONLY: zero markup on each line; corporate OH (8-12%) + Profit (5-10%) "
        "added at the Bid Summary. Defaults: Sub Markup 0%, Self-Perform OH 0% + Profit 0%. Summary "
        "OH = 10%, Profit = 8%.",
        "Default values are set for (A). Adjust Project Info to switch to (B).",
        "",
        "GENERAL CONDITIONS — TYPICAL RUN-RATE",
        "Division 01 General Conditions typically run 5-10% of the trade-cost subtotal on commercial work; "
        "tight industrial may run 8-12%; complex multi-site or fast-track can hit 15%. The pack lets you "
        "itemize OR enter as a lump sum. Itemizing produces a better bid; lump-sum is faster.",
        "",
        "BOND, INSURANCE, TAX",
        "Bond default: 1.0% of bid (varies 0.5-1.5% by surety + project profile).",
        "Sales tax: project-state-specific; on materials, not on labor/equipment in most states. Confirm "
        "with your CPA before relying on the default.",
        "Builder's Risk: typically Owner-supplied. Confirm before adding to bid.",
        "",
        "CONTINGENCY CONVENTION",
        "GC contingency lives in TWO PLACES:",
        "  - Per-trade contingency on the Trade Bids tab (covers sub quote risk on that specific trade).",
        "  - Project contingency on the Allowances tab (covers unknowns + estimating risk overall).",
        "Don't apply both at the SAME % on the same scope or you're double-buffering. Industry-typical: "
        "1-3% per-trade for risky scopes, plus 2-4% project-level. Total typically 3-6% combined.",
        "",
        "THIS WORKBOOK + GC APPLICATION FOR PAYMENT",
        "The Trade Bids + Self-Perform totals from this workbook become the Schedule of Values in the GC "
        "Application for Payment workbook. After award, copy the line-item values from Trade Bids into "
        "the SOV Setup tab of GC_Application_for_Payment.xlsx and bill against them monthly.",
        "",
        "WHAT THIS WORKBOOK IS NOT",
        "  - Not a takeoff tool. Use the trade-specific bid estimators (Steel Erection, Electrical, etc.) "
        "for from-scratch quantity-based estimating, or buy sub quotes.",
        "  - Not a CPM scheduling tool. Use GC_Project_Operations.xlsx for the Schedule tab.",
        "  - Not a job-cost accounting system. Use Job_Costing_Spreadsheet.xlsx (Business tier) for that.",
    ]

    for i, text in enumerate(rows, start=2):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if text and text == text.upper() and len(text) < 70:
            cell.font = FONT_H2
        else:
            cell.font = FONT_BODY


# ---------------------------------------------------------------------------
# Tab 2: Project Info
# ---------------------------------------------------------------------------

def build_project_info(ws):
    ws.title = "Project Info"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 42), ("C", 36)])

    ws["B1"] = "PROJECT INFORMATION — GC Bid"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:C1")
    ws.row_dimensions[1].height = 32

    info_rows = [
        ("Project Name", ""),
        ("Project Address", ""),
        ("Owner", ""),
        ("Architect", ""),
        ("Engineer (Structural / MEP)", ""),
        ("Bid Due Date", ""),
        ("Project Type", "Commercial — New Construction"),
        ("Delivery Method", "Design-Bid-Build"),
        ("Construction Schedule (months)", 0),
        ("Anticipated NTP Date", ""),
        ("Anticipated Substantial Completion", ""),
        ("Project State (sales-tax rule)", ""),
        ("Davis-Bacon / Prevailing Wage?", "No"),
        ("", ""),
        ("MARKUP CONVENTION (SEE INSTRUCTIONS — PICK ONE)", ""),
        ("Sub Markup % (per-trade line-level)", 0.07),
        ("Self-Perform OH %", 0.10),
        ("Self-Perform Profit %", 0.08),
        ("Corporate / Summary OH %", 0.00),
        ("Corporate / Summary Profit %", 0.00),
        ("", ""),
        ("BOND, INSURANCE, TAX", ""),
        ("P&P Bond %", 0.010),
        ("Sales Tax on Materials %", 0.07),
        ("Builder's Risk in bid? (Owner usually supplies)", "No"),
        ("", ""),
        ("CONTINGENCY", ""),
        ("Per-Trade Contingency Default %", 0.02),
        ("Project-Level Contingency %", 0.03),
        ("", ""),
        ("BIDDER OF RECORD", ""),
        ("Lead Estimator", ""),
        ("Project Executive", ""),
        ("Authorized Signatory", ""),
    ]

    name_map = {
        "ProjectName": "$C$3",
        "ProjectState": "$C$14",
        "IsPrevailingWage": "$C$15",
        "SubMarkup": "$C$18",
        "SelfPerfOH": "$C$19",
        "SelfPerfProfit": "$C$20",
        "CorpOH": "$C$21",
        "CorpProfit": "$C$22",
        "BondPct": "$C$25",
        "SalesTax": "$C$26",
        "BuildersRisk": "$C$27",
        "TradeContingency": "$C$30",
        "ProjectContingency": "$C$31",
    }
    for nm, ref in name_map.items():
        ws.parent.defined_names[nm] = DefinedName(name=nm, attr_text=f"'Project Info'!{ref}")

    for i, (label, val) in enumerate(info_rows, start=3):
        is_section = (label and label == label.upper() and not label.startswith(("Project ", "Owner", "Architect", "Engineer", "Bid ", "Anticipated", "Lead ", "Authorized"))) and len(label) < 70
        if is_section:
            ws.cell(row=i, column=2, value=label).font = FONT_H2
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
            continue
        if label:
            ws.cell(row=i, column=2, value=label).font = FONT_BODY_BOLD
            ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
            ws.cell(row=i, column=2).border = BORDER
            c = ws.cell(row=i, column=3, value=val)
            c.font = FONT_BODY
            c.border = BORDER
            if isinstance(val, float):
                c.number_format = FMT_PCT if val < 1 and val != 0 else FMT_NUM2
            if "Date" in label or "Completion" in label or "NTP" in label:
                c.number_format = FMT_DATE
            if isinstance(val, int) and "months" in label:
                c.number_format = FMT_INT

    dv_yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    dv_yn.add("C15")
    dv_yn.add("C27")
    ws.add_data_validation(dv_yn)

    dv_delivery = DataValidation(
        type="list",
        formula1='"Design-Bid-Build,Design-Build,GMP,CM at Risk,IPD,Other"',
        allow_blank=True,
    )
    dv_delivery.add("C10")
    ws.add_data_validation(dv_delivery)


# ---------------------------------------------------------------------------
# Tab 3: Trade Bids
# ---------------------------------------------------------------------------

TRADE_COLS = [
    ("A", "CSI Div", 9),
    ("B", "Trade Name", 32),
    ("C", "Sub (low qualifying bidder)", 26),
    ("D", "Quote Date", 12),
    ("E", "Sub Quote ($)", 16),
    ("F", "Quote Includes Bond?", 14),
    ("G", "Quote Includes Sales Tax?", 16),
    ("H", "GC Markup %", 12),
    ("I", "Per-Trade Contingency %", 16),
    ("J", "Line Total ($)", 16),
    ("K", "Notes / Exclusions", 30),
]


def build_trade_bids(ws):
    ws.title = "Trade Bids"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in TRADE_COLS])

    ws["A1"] = "TRADE BIDS — Subcontractor Quote Roll-Up"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:K1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "One row per CSI division. Enter the lowest qualifying sub quote per trade. The GC Markup column "
        "defaults to the SubMarkup named range (set on Project Info — default 7%); override per trade for "
        "thin-margin or high-risk scopes. Line Total = SubQuote × (1 + Markup) × (1 + Contingency)."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:K2")
    ws.row_dimensions[2].height = 32

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(TRADE_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(TRADE_COLS))

    # Seed all 26 CSI divisions
    for i, (code, name) in enumerate(DIVISIONS, start=HEADER_ROW + 1):
        ws.cell(row=i, column=1, value=code).font = FONT_BODY
        ws.cell(row=i, column=2, value=name).font = FONT_BODY
        ws.cell(row=i, column=4).number_format = FMT_DATE
        ws.cell(row=i, column=5).number_format = FMT_USD
        # GC Markup default from named range
        ws.cell(row=i, column=8, value="=SubMarkup").number_format = FMT_PCT
        # Contingency default from named range
        ws.cell(row=i, column=9, value="=TradeContingency").number_format = FMT_PCT
        # Line Total = Quote × (1 + markup) × (1 + contingency)
        ws.cell(
            row=i,
            column=10,
            value=f'=IFERROR(E{i}*(1+H{i})*(1+I{i}),0)',
        ).number_format = FMT_USD
        apply_body_style(ws, i, len(TRADE_COLS))

    # Validations on Y/N columns
    dv_yn = DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=True)
    last_data_row = HEADER_ROW + len(DIVISIONS)
    dv_yn.add(f"F{HEADER_ROW+1}:F{last_data_row}")
    dv_yn.add(f"G{HEADER_ROW+1}:G{last_data_row}")
    ws.add_data_validation(dv_yn)

    # Totals row
    TOTAL_ROW = last_data_row + 1
    ws.cell(row=TOTAL_ROW, column=2, value="TRADE BIDS TOTAL").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=2).fill = FILL_SUBHEADER
    for col_idx, col_letter in [(5, "E"), (10, "J")]:
        c = ws.cell(
            row=TOTAL_ROW,
            column=col_idx,
            value=f"=SUM({col_letter}{HEADER_ROW+1}:{col_letter}{last_data_row})",
        )
        c.font = FONT_BODY_BOLD
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_USD
        c.border = BORDER


# ---------------------------------------------------------------------------
# Tab 4: Self-Perform
# ---------------------------------------------------------------------------

SELFPERF_COLS = [
    ("A", "CSI Div", 9),
    ("B", "Activity", 32),
    ("C", "Labor Cost ($)", 16),
    ("D", "Material Cost ($)", 16),
    ("E", "Equipment Cost ($)", 16),
    ("F", "Direct Cost ($)", 16),
    ("G", "OH %", 9),
    ("H", "Profit %", 9),
    ("I", "Line Total ($)", 16),
    ("J", "Notes", 22),
]


def build_selfperform(ws):
    ws.title = "Self-Perform"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in SELFPERF_COLS])

    ws["A1"] = "SELF-PERFORM WORK — GC Crew"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Work the GC self-performs with its own crews (rare on true GC; common on Design-Build / CM-at-Risk). "
        "Common self-perform scopes: site work, concrete, carpentry, general labor, demolition. Direct Cost "
        "auto-sums Labor+Material+Equipment; Line Total applies SelfPerfOH + SelfPerfProfit defaults."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:J2")
    ws.row_dimensions[2].height = 32

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(SELFPERF_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(SELFPERF_COLS))

    # Seed with empty rows + a few common self-perform examples
    examples = [
        ("01", "Project Supervision (allocated)"),
        ("02", "Demolition"),
        ("03", "Concrete — flatwork"),
        ("06", "Rough carpentry"),
        ("31", "Site work / earthwork"),
        ("", ""),
        ("", ""),
        ("", ""),
    ]
    for i, (csi, activity) in enumerate(examples, start=HEADER_ROW + 1):
        ws.cell(row=i, column=1, value=csi)
        ws.cell(row=i, column=2, value=activity)
        ws.cell(row=i, column=3).number_format = FMT_USD
        ws.cell(row=i, column=4).number_format = FMT_USD
        ws.cell(row=i, column=5).number_format = FMT_USD
        # Direct cost = labor + material + equipment
        ws.cell(row=i, column=6, value=f'=IFERROR(SUM(C{i}:E{i}),0)').number_format = FMT_USD
        # OH + Profit from named ranges
        ws.cell(row=i, column=7, value="=SelfPerfOH").number_format = FMT_PCT
        ws.cell(row=i, column=8, value="=SelfPerfProfit").number_format = FMT_PCT
        # Line Total = direct × (1 + OH) × (1 + Profit)
        ws.cell(
            row=i,
            column=9,
            value=f'=IFERROR(F{i}*(1+G{i})*(1+H{i}),0)',
        ).number_format = FMT_USD
        apply_body_style(ws, i, len(SELFPERF_COLS))

    last_data_row = HEADER_ROW + len(examples)
    TOTAL_ROW = last_data_row + 1
    ws.cell(row=TOTAL_ROW, column=2, value="SELF-PERFORM TOTAL").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=2).fill = FILL_SUBHEADER
    for col_idx, col_letter in [(6, "F"), (9, "I")]:
        c = ws.cell(
            row=TOTAL_ROW,
            column=col_idx,
            value=f"=SUM({col_letter}{HEADER_ROW+1}:{col_letter}{last_data_row})",
        )
        c.font = FONT_BODY_BOLD
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_USD
        c.border = BORDER

    # CSI dropdown
    csi_list = ",".join(c for c, _ in DIVISIONS)
    dv_csi = DataValidation(type="list", formula1=f'"{csi_list}"', allow_blank=True)
    dv_csi.add(f"A{HEADER_ROW+1}:A{last_data_row}")
    ws.add_data_validation(dv_csi)


# ---------------------------------------------------------------------------
# Tab 5: General Conditions (Div 01)
# ---------------------------------------------------------------------------

GC_COLS = [
    ("A", "Category", 28),
    ("B", "Item", 32),
    ("C", "Basis", 16),
    ("D", "Quantity", 12),
    ("E", "Unit Cost ($)", 14),
    ("F", "Duration (months)", 13),
    ("G", "Line Total ($)", 14),
    ("H", "Notes", 22),
]


def build_general_conditions(ws):
    ws.title = "General Conditions"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in GC_COLS])

    ws["A1"] = "GENERAL CONDITIONS — Division 01"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Division 01 line items not part of any sub trade. Basis column drives the math: "
        "MONTHLY = Qty × Unit × Duration, LUMP = Qty × Unit, PER-SF = Qty × Unit (where Qty = project SF). "
        "Typically 5-10% of trade-cost subtotal on commercial work."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 32

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(GC_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(GC_COLS))

    # Seed with comprehensive Division 01 line items
    seed = [
        ("Supervision",      "Project Manager (on-site allocation)",  "MONTHLY", 1, 0, 0),
        ("Supervision",      "Superintendent",                         "MONTHLY", 1, 0, 0),
        ("Supervision",      "Assistant Superintendent",               "MONTHLY", 0, 0, 0),
        ("Supervision",      "Project Engineer",                       "MONTHLY", 0, 0, 0),
        ("Supervision",      "Safety Officer",                         "MONTHLY", 0, 0, 0),
        ("Field Office",     "Trailer rental",                         "MONTHLY", 1, 0, 0),
        ("Field Office",     "Trailer set-up + tear-down",             "LUMP",    1, 0, 0),
        ("Field Office",     "Power / utilities to trailer",           "MONTHLY", 1, 0, 0),
        ("Field Office",     "Phone / internet",                       "MONTHLY", 1, 0, 0),
        ("Field Office",     "Office supplies + copies",               "MONTHLY", 1, 0, 0),
        ("Site Facilities",  "Port-a-John (per unit per month)",       "MONTHLY", 1, 0, 0),
        ("Site Facilities",  "Dumpsters (rolling)",                    "MONTHLY", 1, 0, 0),
        ("Site Facilities",  "Site fencing + gates",                   "LUMP",    1, 0, 0),
        ("Site Facilities",  "Site signage + safety signage",          "LUMP",    1, 0, 0),
        ("Site Facilities",  "Site security (off-hours)",              "MONTHLY", 0, 0, 0),
        ("Temp Utilities",   "Temp power (service + monthly)",         "MONTHLY", 1, 0, 0),
        ("Temp Utilities",   "Temp water / hose bibs",                 "MONTHLY", 1, 0, 0),
        ("Temp Utilities",   "Temp lighting (interior)",               "MONTHLY", 1, 0, 0),
        ("Temp Utilities",   "Temp heat (winter weeks)",               "MONTHLY", 0, 0, 0),
        ("Logistics",        "Material hoisting / forklift rental",    "MONTHLY", 0, 0, 0),
        ("Logistics",        "Crane mobilization for misc lifts",      "LUMP",    1, 0, 0),
        ("Logistics",        "Equipment moves (between phases)",       "LUMP",    1, 0, 0),
        ("Logistics",        "Traffic control / flagger (urban)",      "MONTHLY", 0, 0, 0),
        ("Logistics",        "Snow removal (winter)",                  "MONTHLY", 0, 0, 0),
        ("Closeout",         "Final cleaning (lump sum)",              "LUMP",    1, 0, 0),
        ("Closeout",         "As-built drawings (CAD service)",        "LUMP",    1, 0, 0),
        ("Closeout",         "O&M manuals + training",                 "LUMP",    1, 0, 0),
        ("Closeout",         "Substantial completion punch-list closeout", "LUMP", 1, 0, 0),
        ("Office OH",        "Project accounting allocation",          "MONTHLY", 1, 0, 0),
        ("Office OH",        "Estimating allocation (post-bid)",       "LUMP",    1, 0, 0),
        ("Office OH",        "Permit + inspection fees (allowance)",   "LUMP",    1, 0, 0),
        ("", "", "", "", "", ""),
        ("", "", "", "", "", ""),
    ]

    for i, (cat, item, basis, qty, ucost, dur) in enumerate(seed, start=HEADER_ROW + 1):
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2, value=item)
        ws.cell(row=i, column=3, value=basis)
        ws.cell(row=i, column=4, value=qty if qty else 0).number_format = FMT_NUM2
        ws.cell(row=i, column=5, value=ucost if ucost else 0).number_format = FMT_USD
        ws.cell(row=i, column=6, value=dur if dur else 0).number_format = FMT_NUM2
        # Line total = qty × unit × (duration if MONTHLY else 1)
        ws.cell(
            row=i,
            column=7,
            value=(
                f'=IFERROR('
                f'IF(C{i}="MONTHLY",D{i}*E{i}*F{i},D{i}*E{i}),0)'
            ),
        ).number_format = FMT_USD
        apply_body_style(ws, i, len(GC_COLS))

    last_data_row = HEADER_ROW + len(seed)
    TOTAL_ROW = last_data_row + 1
    ws.cell(row=TOTAL_ROW, column=2, value="GENERAL CONDITIONS TOTAL").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=2).fill = FILL_SUBHEADER
    c = ws.cell(
        row=TOTAL_ROW,
        column=7,
        value=f"=SUM(G{HEADER_ROW+1}:G{last_data_row})",
    )
    c.font = FONT_BODY_BOLD
    c.fill = FILL_SUBHEADER
    c.number_format = FMT_USD
    c.border = BORDER

    dv_basis = DataValidation(
        type="list",
        formula1='"MONTHLY,LUMP,PER-SF,PER-LF"',
        allow_blank=True,
    )
    dv_basis.add(f"C{HEADER_ROW+1}:C{last_data_row}")
    ws.add_data_validation(dv_basis)


# ---------------------------------------------------------------------------
# Tab 6: Allowances & Contingencies
# ---------------------------------------------------------------------------

ALLOW_COLS = [
    ("A", "Type", 14),
    ("B", "Description", 38),
    ("C", "Amount ($)", 16),
    ("D", "Basis / Notes", 30),
]


def build_allowances(ws):
    ws.title = "Allowances"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in ALLOW_COLS])

    ws["A1"] = "ALLOWANCES & CONTINGENCIES"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Explicit allowances (called out in the bid for Owner-directed scope where price isn't yet defined) "
        "plus project-level contingency. Per-trade contingency is on the Trade Bids tab."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:D2")

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(ALLOW_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(ALLOW_COLS))

    seed = [
        ("Allowance",   "Finish hardware (Owner specifying TBD)",  0, "Owner selects from approved list"),
        ("Allowance",   "Appliances (Owner specifying TBD)",        0, "Owner-directed late selection"),
        ("Allowance",   "Specialty doors / glazing (TBD design)",   0, "Pending design refinement"),
        ("Allowance",   "Paint color selection (TBD)",              0, "Color premium if non-standard"),
        ("Allowance",   "Landscaping allowance",                    0, "Owner-selected design"),
        ("Allowance",   "Permit + inspection fees actual",          0, "If actuals exceed Div 01 allowance"),
        ("",            "",                                          0, ""),
        ("Contingency", "Project-level contingency (unknowns)",     0, "=CostOfWork × ProjectContingency"),
        ("",            "",                                          0, ""),
        ("",            "",                                          0, ""),
    ]
    for i, (typ, desc, amt, notes) in enumerate(seed, start=HEADER_ROW + 1):
        ws.cell(row=i, column=1, value=typ)
        ws.cell(row=i, column=2, value=desc)
        ws.cell(row=i, column=3, value=amt).number_format = FMT_USD
        ws.cell(row=i, column=4, value=notes)
        apply_body_style(ws, i, len(ALLOW_COLS))

    last_data_row = HEADER_ROW + len(seed)
    TOTAL_ROW = last_data_row + 1
    ws.cell(row=TOTAL_ROW, column=2, value="ALLOWANCES + CONTINGENCY TOTAL").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=2).fill = FILL_SUBHEADER
    c = ws.cell(
        row=TOTAL_ROW,
        column=3,
        value=f"=SUM(C{HEADER_ROW+1}:C{last_data_row})",
    )
    c.font = FONT_BODY_BOLD
    c.fill = FILL_SUBHEADER
    c.number_format = FMT_USD
    c.border = BORDER

    dv_type = DataValidation(
        type="list",
        formula1='"Allowance,Contingency"',
        allow_blank=True,
    )
    dv_type.add(f"A{HEADER_ROW+1}:A{last_data_row}")
    ws.add_data_validation(dv_type)


# ---------------------------------------------------------------------------
# Tab 7: Bid Summary
# ---------------------------------------------------------------------------

def build_bid_summary(ws):
    ws.title = "Bid Summary"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 48), ("C", 22), ("D", 16), ("E", 28)])

    ws["B1"] = "BID SUMMARY — Roll-Up to Final Prime Bid"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:E1")
    ws.row_dimensions[1].height = 32

    ws["B2"] = "Project: =ProjectName"
    ws["B2"].font = FONT_BODY_BOLD
    ws.merge_cells("B2:E2")

    # === Cost roll-up ===
    ws["B4"] = "COST OF WORK"
    ws["B4"].font = FONT_H2

    # Trade Bids total: row 31 (4 header + 26 divisions + 1 totals = row 31)
    # Self-Perform total: row 4 header + 8 examples + 1 = row 13
    # General Conditions total: row 4 header + 33 seed + 1 = row 38
    # Allowances total: row 4 header + 10 seed + 1 = row 15
    rows = [
        ("Trade Bids (from Trade Bids tab)",          "='Trade Bids'!J31"),
        ("Self-Perform (from Self-Perform tab)",      "='Self-Perform'!I13"),
        ("General Conditions (Div 01)",               "='General Conditions'!G38"),
        ("Allowances + Contingency",                   "='Allowances'!C15"),
    ]

    start_row = 5
    for i, (label, formula) in enumerate(rows, start=start_row):
        ws.cell(row=i, column=2, value=label).font = FONT_BODY_BOLD
        ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
        ws.cell(row=i, column=2).border = BORDER
        c = ws.cell(row=i, column=3, value=formula)
        c.font = FONT_BODY
        c.number_format = FMT_USD
        c.border = BORDER

    cost_row = start_row + len(rows)
    ws.cell(row=cost_row, column=2, value="COST OF WORK SUBTOTAL").font = FONT_BODY_BOLD
    ws.cell(row=cost_row, column=2).fill = FILL_SUBHEADER
    ws.cell(row=cost_row, column=2).border = BORDER
    c = ws.cell(row=cost_row, column=3, value=f"=SUM(C{start_row}:C{cost_row-1})")
    c.font = FONT_BODY_BOLD
    c.fill = FILL_SUBHEADER
    c.number_format = FMT_USD
    c.border = BORDER

    # === Corporate OH + Profit (summary-level — typically 0 if line-level markup chosen) ===
    cur = cost_row + 2
    ws.cell(row=cur, column=2, value="Corporate Overhead").font = FONT_BODY_BOLD
    ws.cell(row=cur, column=2).fill = FILL_SUMMARY_LABEL
    ws.cell(row=cur, column=2).border = BORDER
    ws.cell(row=cur, column=3, value=f"=C{cost_row}*CorpOH").number_format = FMT_USD
    ws.cell(row=cur, column=3).border = BORDER
    ws.cell(row=cur, column=4, value="=CorpOH").number_format = FMT_PCT
    ws.cell(row=cur, column=4).border = BORDER
    cur += 1
    ws.cell(row=cur, column=2, value="Corporate Profit").font = FONT_BODY_BOLD
    ws.cell(row=cur, column=2).fill = FILL_SUMMARY_LABEL
    ws.cell(row=cur, column=2).border = BORDER
    ws.cell(row=cur, column=3, value=f"=C{cost_row}*CorpProfit").number_format = FMT_USD
    ws.cell(row=cur, column=3).border = BORDER
    ws.cell(row=cur, column=4, value="=CorpProfit").number_format = FMT_PCT
    ws.cell(row=cur, column=4).border = BORDER
    cur += 1

    # === Sales tax on materials (rough — applied to ~50% of trade cost as material content proxy) ===
    ws.cell(row=cur, column=2, value="Sales Tax (on est. material content)").font = FONT_BODY_BOLD
    ws.cell(row=cur, column=2).fill = FILL_SUMMARY_LABEL
    ws.cell(row=cur, column=2).border = BORDER
    ws.cell(
        row=cur,
        column=3,
        value=f"=C{cost_row}*0.5*SalesTax",
    ).number_format = FMT_USD
    ws.cell(row=cur, column=3).border = BORDER
    ws.cell(row=cur, column=4, value="=SalesTax").number_format = FMT_PCT
    ws.cell(row=cur, column=4).border = BORDER
    ws.cell(row=cur, column=5, value="Material is ~50% of trade cost — adjust per project").font = FONT_GREY_ITALIC
    cur += 1

    # Pre-bond total
    pre_bond_row = cur
    ws.cell(row=pre_bond_row, column=2, value="PRE-BOND SUBTOTAL").font = FONT_BODY_BOLD
    ws.cell(row=pre_bond_row, column=2).fill = FILL_SUBHEADER
    ws.cell(row=pre_bond_row, column=2).border = BORDER
    c = ws.cell(
        row=pre_bond_row,
        column=3,
        value=f"=C{cost_row}+C{cost_row+2}+C{cost_row+3}+C{cost_row+4}",
    )
    c.font = FONT_BODY_BOLD
    c.fill = FILL_SUBHEADER
    c.number_format = FMT_USD
    c.border = BORDER
    cur += 1

    # Bond
    ws.cell(row=cur, column=2, value="Performance + Payment Bond").font = FONT_BODY_BOLD
    ws.cell(row=cur, column=2).fill = FILL_SUMMARY_LABEL
    ws.cell(row=cur, column=2).border = BORDER
    # Bond formula handles its own markup so the bond cost itself is bonded
    ws.cell(
        row=cur,
        column=3,
        value=f"=C{pre_bond_row}*BondPct/(1-BondPct)",
    ).number_format = FMT_USD
    ws.cell(row=cur, column=3).border = BORDER
    ws.cell(row=cur, column=4, value="=BondPct").number_format = FMT_PCT
    ws.cell(row=cur, column=4).border = BORDER
    bond_row = cur
    cur += 2

    # FINAL BID
    final_row = cur
    ws.cell(row=final_row, column=2, value="FINAL PRIME BID").font = FONT_TITLE
    ws.cell(row=final_row, column=2).fill = FILL_GOLD
    ws.cell(row=final_row, column=2).border = BORDER
    c = ws.cell(row=final_row, column=3, value=f"=C{pre_bond_row}+C{bond_row}")
    c.font = FONT_TITLE
    c.fill = FILL_GOLD
    c.number_format = FMT_USD0
    c.border = BORDER
    ws.row_dimensions[final_row].height = 40

    # === Sanity checks ===
    sanity = cur + 3
    ws.cell(row=sanity, column=2, value="SANITY CHECKS").font = FONT_H2

    checks = [
        ("Trade % of total bid (sanity: 75-90% commercial)", f"=IFERROR('Trade Bids'!J31/C{final_row},0)"),
        ("Gen Conditions % of trade total (sanity: 5-12%)", f"=IFERROR('General Conditions'!G38/'Trade Bids'!J31,0)"),
        ("Self-Perform % of total (varies — 0% true GC, 20-40% CMAR)", f"=IFERROR('Self-Perform'!I13/C{final_row},0)"),
        ("Allowance % of total (sanity: typically <5%)", f"=IFERROR('Allowances'!C15/C{final_row},0)"),
        ("Effective markup over Cost of Work", f"=IFERROR((C{final_row}-C{cost_row})/C{cost_row},0)"),
    ]
    for i, (label, formula) in enumerate(checks, start=sanity + 1):
        ws.cell(row=i, column=2, value=label).font = FONT_BODY
        ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
        ws.cell(row=i, column=2).border = BORDER
        c = ws.cell(row=i, column=3, value=formula)
        c.font = FONT_BODY
        c.number_format = FMT_PCT
        c.border = BORDER


# ---------------------------------------------------------------------------
# Tab 8: CSI Reference (hidden)
# ---------------------------------------------------------------------------

def build_csi_reference(ws):
    ws.title = "CSI Reference"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 10), ("B", 60)])
    ws["A1"] = "CSI MasterFormat Divisions"
    ws["A1"].font = FONT_H2
    for i, (code, name) in enumerate(DIVISIONS, start=3):
        ws.cell(row=i, column=1, value=code).font = FONT_BODY
        ws.cell(row=i, column=2, value=name).font = FONT_BODY
    ws.sheet_state = "hidden"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    wb = Workbook()
    build_instructions(wb.active)
    build_project_info(wb.create_sheet())
    build_trade_bids(wb.create_sheet())
    build_selfperform(wb.create_sheet())
    build_general_conditions(wb.create_sheet())
    build_allowances(wb.create_sheet())
    build_bid_summary(wb.create_sheet())
    build_csi_reference(wb.create_sheet())

    out_dir = "/Users/home/charles/contrpro/files/packages/complete/gc"
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "GC_Bid_Estimator.xlsx")
    wb.save(out_path)
    sz = os.path.getsize(out_path)
    print(f"Wrote {out_path} ({sz//1024} KB)")


if __name__ == "__main__":
    main()
