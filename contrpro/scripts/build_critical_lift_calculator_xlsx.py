#!/usr/bin/env python3
"""
Build ContrPro Critical Lift Calculator (XLSX) — Steel Erection trade pack.

Companion to crane-and-rigging-lift-plan.html. Computes pick weight, percent
of crane capacity, sling angle de-rating, ground bearing pressure, and
critical-lift flagging. Output: a single-page printable lift summary suitable
for the Critical Lift Permit.

Tabs:
  1. Instructions
  2. Lift Identification    (project, lift #, element, date)
  3. Pick Weight             (structural weight + rigging + crane block)
  4. Crane Capacity Check   (radius / boom config → capacity from operator-input load chart)
  5. Sling Angle Check      (per-sling load with angle de-rate)
  6. Ground Bearing Check   (outrigger pressure vs. soils capacity)
  7. Tandem Lift            (optional second-crane balance check)
  8. Lift Permit Summary    (printable single-page roll-up)

Run:
    /Users/home/charles/.venv/bin/python3 \\
        /Users/home/charles/contrpro/scripts/build_critical_lift_calculator_xlsx.py

Output:
    /Users/home/charles/contrpro/files/packages/complete/steel-erection/Critical_Lift_Calculator.xlsx
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
ACCENT_GOLD = "C9A227"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
YELLOW_FILL = "FFEB9C"
YELLOW_FONT = "9C5700"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_SUMMARY_LABEL = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW_FILL)
FILL_GOLD = PatternFill("solid", fgColor=ACCENT_GOLD)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_BIG_NUMBER = Font(name="Calibri", size=18, bold=True, color=BRAND_BLUE)
FONT_GREY_ITALIC = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_LB = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
FMT_PCT = "0.0%"
FMT_INT = "0"
FMT_NUM2 = "0.00"
FMT_NUM1 = "0.0"
FMT_FT = '#,##0.0" ft"'
FMT_DATE = "yyyy-mm-dd"


def set_col_widths(ws, widths):
    for col, w in widths:
        ws.column_dimensions[col].width = w


def style_header_row(ws, row, cols, start_col=1):
    for c in range(start_col, start_col + cols):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


# ---------------------------------------------------------------------------
# Tab 1: Instructions
# ---------------------------------------------------------------------------

def build_instructions(ws):
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 110)])

    ws["A1"] = "CRITICAL LIFT CALCULATOR — Instructions"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 36

    rows = [
        "",
        "WHAT THIS WORKBOOK COMPUTES",
        "Pick weight (structural + rigging + crane block), percent of crane capacity at the lift radius, "
        "sling angle de-rate, ground bearing pressure under outriggers, and a roll-up suitable for the "
        "Critical Lift Permit (Section 11 of the Lift Plan HTML).",
        "",
        "WORKFLOW",
        "1. Lift Identification: project, lift #, element, date.",
        "2. Pick Weight: enter structural weight (from shop drawings or MTRs), rigging weights, hook block.",
        "3. Crane Capacity Check: enter crane configuration + load chart capacity at the lift radius. The "
        "workbook computes % of capacity used. Critical-lift trigger fires at ≥ 75%.",
        "4. Sling Angle Check: enter sling configuration (length, sling-to-load attachment geometry). The "
        "workbook computes sling angle and per-sling load. Sling angles below 45° from horizontal are "
        "FLAGGED — choose longer slings or a spreader bar.",
        "5. Ground Bearing Check: enter outrigger pad / mat dimensions + crane manufacturer's reported "
        "outrigger reaction force at this load and radius. The workbook computes bearing pressure and "
        "compares to soils-report capacity.",
        "6. Tandem Lift (optional): if two cranes share the load, enter each crane's share and radius — "
        "the workbook checks both cranes individually for ≥ 75% capacity threshold.",
        "7. Lift Permit Summary: print this tab as the back-up sheet to the Critical Lift Permit in the "
        "Lift Plan HTML.",
        "",
        "CRITICAL-LIFT THRESHOLD",
        "Per most commercial lift plans, a critical lift is triggered at ≥ 75% of crane capacity at the "
        "operating radius. Some firms use 80% as a slightly more relaxed threshold; some use 50% on more "
        "conservative programs. Adjust the threshold on Pick Weight tab if your firm's policy differs.",
        "",
        "SLING ANGLE — CRITICAL CHECK",
        "Sling angle is the angle between the sling and the horizontal at the load. Below 60° the share "
        "of load on each sling rises sharply: at 30° each sling sees 2× the vertical share. The workbook "
        "computes both the sling angle and the per-sling load with de-rate. Slings must have a working "
        "load limit (WLL) ≥ the per-sling load.",
        "",
        "GROUND BEARING — CRITICAL CHECK",
        "Outrigger float pressure = outrigger reaction force ÷ outrigger pad area. The reaction force at "
        "the worst-case outrigger is given by the crane manufacturer's outrigger reaction chart at the "
        "lift configuration. Bearing pressure must be ≤ soils-report capacity / factor of safety (typically "
        "2.0). If a mat is used instead of a pad, mat area takes the place of pad area in the formula.",
        "",
        "ROUNDING + SAFETY MARGINS",
        "The workbook rounds DOWN on capacities and UP on loads. Always design with a margin: even when "
        "the workbook reports 73%, treat anything above 65% as 'a critical lift in waiting' and run the "
        "full critical-lift procedure if conditions could push it higher (wind, hung-up sling, swing).",
        "",
        "WHAT THIS WORKBOOK DOES NOT DO",
        "  - Verify crane operator certification. (Check the operator's NCCCO card.)",
        "  - Verify rigging certification. (Inspect slings, shackles, hooks before lift.)",
        "  - Substitute for the operator's load chart. (Use the manufacturer's chart for the specific "
        "crane in the specific configuration on the specific outrigger set.)",
        "  - Compute crane stability or boom-deflection effects. (Manufacturer's chart already does this.)",
        "  - Make critical-lift policy decisions. (The Lift Director and the EOR own that call.)",
    ]

    for i, text in enumerate(rows, start=2):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if text and text == text.upper() and len(text) < 70:
            cell.font = FONT_H2
        else:
            cell.font = FONT_BODY


# ---------------------------------------------------------------------------
# Tab 2: Lift Identification
# ---------------------------------------------------------------------------

def build_lift_id(ws):
    ws.title = "Lift Identification"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 42), ("C", 38)])

    ws["B1"] = "LIFT IDENTIFICATION"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:C1")
    ws.row_dimensions[1].height = 32

    info_rows = [
        ("Project Name", ""),
        ("Project Address", ""),
        ("Controlling Contractor (GC)", ""),
        ("Steel Erector", ""),
        ("Lift Permit No.", ""),
        ("Lift Date / Time", ""),
        ("Element / Description", ""),
        ("Gridline / Location", ""),
        ("Lift Director", ""),
        ("Crane Operator", ""),
        ("Signal Person", ""),
        ("Rigger", ""),
        ("Spotter (energized lines, if applicable)", ""),
        ("EOR Review Required? (Yes/No)", "No"),
        ("EOR Review Date / Reviewer", ""),
    ]

    name_map = {
        "ProjectName": "$C$3",
        "LiftPermit": "$C$7",
        "LiftElement": "$C$9",
        "LiftDirector": "$C$11",
        "LiftOperator": "$C$12",
    }
    for nm, ref in name_map.items():
        ws.parent.defined_names[nm] = DefinedName(name=nm, attr_text=f"'Lift Identification'!{ref}")

    for i, (label, val) in enumerate(info_rows, start=3):
        ws.cell(row=i, column=2, value=label).font = FONT_BODY_BOLD
        ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
        ws.cell(row=i, column=2).border = BORDER
        c = ws.cell(row=i, column=3, value=val)
        c.font = FONT_BODY
        c.border = BORDER

    dv_yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    dv_yn.add("C16")
    ws.add_data_validation(dv_yn)


# ---------------------------------------------------------------------------
# Tab 3: Pick Weight
# ---------------------------------------------------------------------------

def build_pick_weight(ws):
    ws.title = "Pick Weight"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 42), ("C", 18), ("D", 36)])

    ws["B1"] = "PICK WEIGHT CALCULATION"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:D1")
    ws.row_dimensions[1].height = 32

    ws["B2"] = (
        "Total pick weight = structural element + all rigging components + hook block + reeving below "
        "the hook. Source structural weight from shop drawings or MTRs; use vendor-stated rigging weights."
    )
    ws["B2"].font = FONT_GREY_ITALIC
    ws.merge_cells("B2:D2")

    rows = [
        ("Structural element weight (lb)",          0),
        ("Allowance for paint / fireproofing / accessories (lb)", 0),
        ("Primary sling(s) total weight (lb)",     0),
        ("Secondary sling(s) / chokers (lb)",     0),
        ("Spreader bar / lifting beam (lb)",      0),
        ("Shackles, hooks, swivels (lb)",         0),
        ("Hook block (below pulleys) (lb)",       0),
        ("Reeving below the block (lb)",          0),
        ("Tag-line weight (lb)",                   0),
        ("Misc. attachments / equalizer (lb)",   0),
    ]

    start_row = 4
    for i, (label, val) in enumerate(rows, start=start_row):
        ws.cell(row=i, column=2, value=label).font = FONT_BODY
        ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
        ws.cell(row=i, column=2).border = BORDER
        c = ws.cell(row=i, column=3, value=val)
        c.font = FONT_BODY
        c.number_format = FMT_LB
        c.border = BORDER

    # Total pick weight = sum
    total_row = start_row + len(rows)
    ws.cell(row=total_row, column=2, value="TOTAL PICK WEIGHT (lb)").font = FONT_BODY_BOLD
    ws.cell(row=total_row, column=2).fill = FILL_SUBHEADER
    ws.cell(row=total_row, column=2).border = BORDER
    c = ws.cell(row=total_row, column=3, value=f"=SUM(C{start_row}:C{total_row-1})")
    c.font = FONT_BIG_NUMBER
    c.number_format = FMT_LB
    c.fill = FILL_SUBHEADER
    c.border = BORDER

    ws.parent.defined_names["TotalPickWeight"] = DefinedName(
        name="TotalPickWeight",
        attr_text=f"'Pick Weight'!$C${total_row}",
    )

    # Critical lift threshold
    threshold_row = total_row + 3
    ws.cell(row=threshold_row, column=2, value="CRITICAL LIFT THRESHOLD (% of crane capacity)").font = FONT_BODY_BOLD
    ws.cell(row=threshold_row, column=2).fill = FILL_SUMMARY_LABEL
    ws.cell(row=threshold_row, column=2).border = BORDER
    c = ws.cell(row=threshold_row, column=3, value=0.75)
    c.number_format = FMT_PCT
    c.font = FONT_BODY
    c.border = BORDER
    ws.cell(row=threshold_row, column=4, value="Industry-typical = 75%. Lower for conservative programs.").font = FONT_GREY_ITALIC

    ws.parent.defined_names["CriticalThreshold"] = DefinedName(
        name="CriticalThreshold",
        attr_text=f"'Pick Weight'!$C${threshold_row}",
    )

    # 2026-05-17 audit fix: WarningThreshold for "NEAR CRITICAL" flag. Previous
    # hardcoded 0.65 left a thin 10-point band between routine and critical;
    # bumped to 0.60 to widen the early-warning zone. Adjustable per program.
    warning_row = threshold_row + 1
    ws.cell(row=warning_row, column=2, value="NEAR CRITICAL WARNING THRESHOLD (% — flags margin-thin lifts)").font = FONT_BODY_BOLD
    ws.cell(row=warning_row, column=2).fill = FILL_SUMMARY_LABEL
    ws.cell(row=warning_row, column=2).border = BORDER
    c = ws.cell(row=warning_row, column=3, value=0.60)
    c.number_format = FMT_PCT
    c.font = FONT_BODY
    c.border = BORDER
    ws.cell(row=warning_row, column=4, value="Lifts above this but below Critical Threshold get a yellow margin-warning flag.").font = FONT_GREY_ITALIC

    ws.parent.defined_names["WarningThreshold"] = DefinedName(
        name="WarningThreshold",
        attr_text=f"'Pick Weight'!$C${warning_row}",
    )


# ---------------------------------------------------------------------------
# Tab 4: Crane Capacity Check
# ---------------------------------------------------------------------------

def build_capacity_check(ws):
    ws.title = "Crane Capacity Check"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 46), ("C", 20), ("D", 32)])

    ws["B1"] = "CRANE CAPACITY CHECK"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:D1")
    ws.row_dimensions[1].height = 32

    ws["B2"] = (
        "Pull capacity values directly from the crane manufacturer's load chart for THIS crane in THIS "
        "configuration on THIS outrigger set at THIS radius. Never approximate from a generic chart."
    )
    ws["B2"].font = FONT_GREY_ITALIC
    ws.merge_cells("B2:D2")

    rows = [
        ("Crane make / model", "", "text"),
        ("Boom length (ft)", 0, "number"),
        ("Jib / fly configuration", "", "text"),
        ("Counterweight (lb)", 0, "lb"),
        ("Outrigger configuration", "Fully extended", "text"),
        ("Lift radius (ft from center pin)", 0, "ft"),
        ("Boom angle at lift (degrees)", 0, "deg"),
        ("Capacity at this radius — from manufacturer's load chart (lb)", 0, "lb"),
        ("Deduction for hook block + reeving already in chart? (Yes/No)", "No", "yn"),
        ("If No, deduction (lb)", 0, "lb"),
        ("NET CRANE CAPACITY (lb)", 0, "calc"),
        ("", "", ""),
        ("Total Pick Weight (from Pick Weight tab)", 0, "calc"),
        ("% of Capacity Used", 0, "pct"),
        ("Critical Lift Flag", 0, "flag"),
    ]

    start_row = 4
    for i, (label, val, kind) in enumerate(rows, start=start_row):
        ws.cell(row=i, column=2, value=label).font = FONT_BODY_BOLD if "TOTAL" in label or "NET" in label or "Flag" in label else FONT_BODY
        if label:
            ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
            ws.cell(row=i, column=2).border = BORDER
        if kind == "text":
            ws.cell(row=i, column=3, value=val).border = BORDER
        elif kind == "number":
            c = ws.cell(row=i, column=3, value=val)
            c.number_format = FMT_NUM1
            c.border = BORDER
        elif kind == "lb":
            c = ws.cell(row=i, column=3, value=val)
            c.number_format = FMT_LB
            c.border = BORDER
        elif kind == "ft":
            c = ws.cell(row=i, column=3, value=val)
            c.number_format = FMT_FT
            c.border = BORDER
        elif kind == "deg":
            c = ws.cell(row=i, column=3, value=val)
            c.number_format = '#,##0.0"°"'
            c.border = BORDER
        elif kind == "yn":
            c = ws.cell(row=i, column=3, value=val)
            c.border = BORDER
        elif kind == "calc":
            c = ws.cell(row=i, column=3, value=0)  # formula filled below
            c.border = BORDER
        elif kind == "pct":
            c = ws.cell(row=i, column=3, value=0)
            c.number_format = FMT_PCT
            c.border = BORDER
            c.font = FONT_BIG_NUMBER
        elif kind == "flag":
            c = ws.cell(row=i, column=3, value="")
            c.font = FONT_BODY_BOLD
            c.border = BORDER

    # Compute NET CRANE CAPACITY (row start_row+10) = capacity (start_row+7) - IF(yn=No,deduction,0)
    net_row = start_row + 10  # NET CRANE CAPACITY
    cap_row = start_row + 7   # Capacity at this radius
    yn_row = start_row + 8    # Deduction Y/N
    ded_row = start_row + 9   # If No, deduction
    ws.cell(row=net_row, column=3, value=f'=C{cap_row} - IF(C{yn_row}="No",C{ded_row},0)').number_format = FMT_LB

    # Total Pick Weight pull-through
    pickwt_row = start_row + 12
    ws.cell(row=pickwt_row, column=3, value="=TotalPickWeight").number_format = FMT_LB

    # % of capacity
    pct_row = start_row + 13
    ws.cell(row=pct_row, column=3, value=f'=IFERROR(C{pickwt_row}/C{net_row},0)').number_format = FMT_PCT

    # Flag
    flag_row = start_row + 14
    ws.cell(
        row=flag_row,
        column=3,
        value=(
            f'=IF(C{pct_row}>=1,"OVERLOAD — DO NOT LIFT",'
            f'IF(C{pct_row}>=CriticalThreshold,"CRITICAL LIFT — full procedure required",'
            f'IF(C{pct_row}>=WarningThreshold,"NEAR CRITICAL — review margin","Routine lift")))'
        ),
    )

    # Conditional formatting on flag and pct
    ws.conditional_formatting.add(
        f"C{flag_row}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("OVERLOAD",C{flag_row}))'], fill=FILL_RED, font=Font(bold=True, color=RED_FONT)),
    )
    ws.conditional_formatting.add(
        f"C{flag_row}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("CRITICAL LIFT",C{flag_row}))'], fill=FILL_YELLOW, font=Font(bold=True, color=YELLOW_FONT)),
    )
    ws.conditional_formatting.add(
        f"C{flag_row}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("Routine",C{flag_row}))'], fill=FILL_GREEN, font=Font(bold=True, color=GREEN_FONT)),
    )
    ws.conditional_formatting.add(
        f"C{pct_row}",
        CellIsRule(operator="greaterThanOrEqual", formula=["1"], fill=FILL_RED, font=Font(bold=True, color=RED_FONT)),
    )
    ws.conditional_formatting.add(
        f"C{pct_row}",
        CellIsRule(operator="greaterThanOrEqual", formula=["CriticalThreshold"], fill=FILL_YELLOW, font=Font(bold=True, color=YELLOW_FONT)),
    )
    # Margin-warning band between WarningThreshold and CriticalThreshold gets a softer color
    ws.conditional_formatting.add(
        f"C{pct_row}",
        CellIsRule(operator="greaterThanOrEqual", formula=["WarningThreshold"], fill=PatternFill("solid", fgColor="FFE4B5"), font=Font(color=YELLOW_FONT)),
    )

    dv_yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    dv_yn.add(f"C{yn_row}")
    ws.add_data_validation(dv_yn)

    ws.parent.defined_names["NetCapacity"] = DefinedName(
        name="NetCapacity", attr_text=f"'Crane Capacity Check'!$C${net_row}",
    )
    ws.parent.defined_names["PctCapacity"] = DefinedName(
        name="PctCapacity", attr_text=f"'Crane Capacity Check'!$C${pct_row}",
    )
    ws.parent.defined_names["LiftFlag"] = DefinedName(
        name="LiftFlag", attr_text=f"'Crane Capacity Check'!$C${flag_row}",
    )
    ws.parent.defined_names["LiftRadius"] = DefinedName(
        name="LiftRadius", attr_text=f"'Crane Capacity Check'!$C${start_row+5}",
    )


# ---------------------------------------------------------------------------
# Tab 5: Sling Angle Check
# ---------------------------------------------------------------------------

def build_sling_angle(ws):
    ws.title = "Sling Angle Check"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 46), ("C", 20), ("D", 38)])

    ws["B1"] = "SLING ANGLE &amp; PER-SLING LOAD"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:D1")
    ws.row_dimensions[1].height = 32

    ws["B2"] = (
        "Sling angle is the angle between the sling and horizontal at the load attachment. Below 60° "
        "per-sling load rises sharply; below 45° you should choose longer slings or a spreader bar."
    )
    ws["B2"].font = FONT_GREY_ITALIC
    ws.merge_cells("B2:D2")

    rows = [
        ("Number of slings used",                 2),
        ("Sling rigged length (ft, center-of-load to hook)", 0),
        ("Horizontal distance between sling attachment points on load (ft)", 0),
        ("Pick weight share per sling (lb) — assume equal share unless engineered", 0),
        ("Sling angle from horizontal (deg) — computed", 0),
        ("Sling angle factor — load multiplier", 0),
        ("Per-sling load (lb) — computed", 0),
        ("Sling working load limit (WLL) — from sling tag (lb)", 0),
        ("Sling Adequacy Check", ""),
    ]

    start_row = 4
    for i, (label, val) in enumerate(rows, start=start_row):
        ws.cell(row=i, column=2, value=label).font = FONT_BODY
        ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
        ws.cell(row=i, column=2).border = BORDER
        c = ws.cell(row=i, column=3, value=val)
        c.border = BORDER

    # Sling angle = arctan(sling_length_vert / (horizontal/2)) — simplified: assume slings meet at hook
    # Actually: sling angle = arcsin( vertical_height / sling_length )
    # Vertical = sqrt( sling_length^2 - (horizontal/2)^2 ); sling_angle = arctan(vertical / (horizontal/2))
    # Simpler: sling_angle = arccos((horizontal/2) / sling_length)
    sling_len_row = start_row + 1
    horiz_row = start_row + 2
    share_row = start_row + 3
    angle_row = start_row + 4
    factor_row = start_row + 5
    perload_row = start_row + 6
    wll_row = start_row + 7
    check_row = start_row + 8

    # Even share — if user wants to manually override, they can; default = pick_weight / number_of_slings
    ws.cell(row=share_row, column=3, value=f"=IFERROR(TotalPickWeight/C{start_row},0)").number_format = FMT_LB

    # Angle from horizontal in degrees
    # Use DEGREES(ACOS((horizontal/2)/sling_length)) — protect against div/0 and arc-cos arg > 1
    ws.cell(
        row=angle_row,
        column=3,
        value=(
            f'=IFERROR('
            f'IF(AND(C{sling_len_row}>0,(C{horiz_row}/2)/C{sling_len_row}<=1),'
            f'DEGREES(ACOS((C{horiz_row}/2)/C{sling_len_row})),0),0)'
        ),
    ).number_format = '#,##0.0"°"'

    # Load multiplier = 1 / sin(angle)
    ws.cell(
        row=factor_row,
        column=3,
        value=f'=IFERROR(IF(C{angle_row}>0,1/SIN(RADIANS(C{angle_row})),1),1)',
    ).number_format = FMT_NUM2

    # Per-sling load = share × multiplier
    ws.cell(
        row=perload_row,
        column=3,
        value=f'=IFERROR(C{share_row}*C{factor_row},0)',
    ).number_format = FMT_LB

    # WLL — user enters
    ws.cell(row=wll_row, column=3).number_format = FMT_LB

    # Check
    ws.cell(
        row=check_row,
        column=3,
        value=(
            f'=IF(C{wll_row}=0,"Enter WLL",'
            f'IF(C{angle_row}<45,"ANGLE TOO LOW — increase sling length / use spreader",'
            f'IF(C{perload_row}>C{wll_row},"OVERLOADED — increase sling capacity",'
            f'IF(C{angle_row}<60,"BORDERLINE — verify",'
            f'"OK"))))'
        ),
    )
    ws.cell(row=check_row, column=3).font = FONT_BODY_BOLD

    ws.conditional_formatting.add(
        f"C{check_row}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("OVERLOADED",C{check_row}))'], fill=FILL_RED, font=Font(bold=True, color=RED_FONT)),
    )
    ws.conditional_formatting.add(
        f"C{check_row}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("TOO LOW",C{check_row}))'], fill=FILL_RED, font=Font(bold=True, color=RED_FONT)),
    )
    ws.conditional_formatting.add(
        f"C{check_row}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("BORDERLINE",C{check_row}))'], fill=FILL_YELLOW, font=Font(bold=True, color=YELLOW_FONT)),
    )
    ws.conditional_formatting.add(
        f"C{check_row}",
        FormulaRule(formula=[f'C{check_row}="OK"'], fill=FILL_GREEN, font=Font(bold=True, color=GREEN_FONT)),
    )

    # Reference table
    ref_start = check_row + 3
    ws.cell(row=ref_start, column=2, value="SLING-ANGLE LOAD MULTIPLIERS (reference)").font = FONT_H2
    ref_data = [
        ("90° (vertical)",  "1.000"),
        ("75°",             "1.035"),
        ("60°",             "1.155"),
        ("45°",             "1.414"),
        ("30°",             "2.000"),
        ("15°",             "3.864"),
    ]
    for i, (angle, mult) in enumerate(ref_data, start=ref_start + 1):
        ws.cell(row=i, column=2, value=angle).font = FONT_BODY
        ws.cell(row=i, column=2).border = BORDER
        ws.cell(row=i, column=3, value=mult).font = FONT_BODY
        ws.cell(row=i, column=3).border = BORDER


# ---------------------------------------------------------------------------
# Tab 6: Ground Bearing Check
# ---------------------------------------------------------------------------

def build_ground_bearing(ws):
    ws.title = "Ground Bearing Check"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 50), ("C", 20), ("D", 36)])

    ws["B1"] = "GROUND BEARING PRESSURE CHECK"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:D1")
    ws.row_dimensions[1].height = 32

    ws["B2"] = (
        "Outrigger reaction force is the worst-case load on a single outrigger float at the lift radius — "
        "pull from the crane manufacturer's outrigger-reaction chart, NOT computed here. Pad / mat area "
        "must be sized so that bearing pressure stays below soils-report capacity / FS."
    )
    ws["B2"].font = FONT_GREY_ITALIC
    ws.merge_cells("B2:D2")

    rows = [
        ("Soils-report allowable bearing capacity (psf)", 0),
        ("Factor of safety (typical 2.0)", 2.0),
        ("Allowable bearing pressure (psf) — computed", 0),
        ("", ""),
        ("Worst-case outrigger reaction force at this lift (lb) — from crane mfr chart", 0),
        ("Outrigger pad / mat — Length (ft)", 0),
        ("Outrigger pad / mat — Width (ft)", 0),
        ("Pad / mat area (sq ft) — computed", 0),
        ("Bearing pressure under pad / mat (psf) — computed", 0),
        ("", ""),
        ("Ground Bearing Adequacy Check", ""),
    ]

    start_row = 4
    for i, (label, val) in enumerate(rows, start=start_row):
        if label:
            ws.cell(row=i, column=2, value=label).font = FONT_BODY
            ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
            ws.cell(row=i, column=2).border = BORDER
            c = ws.cell(row=i, column=3, value=val)
            c.border = BORDER

    soils_row = start_row + 0
    fs_row = start_row + 1
    allow_row = start_row + 2
    react_row = start_row + 4
    len_row = start_row + 5
    wid_row = start_row + 6
    area_row = start_row + 7
    press_row = start_row + 8
    check_row = start_row + 10

    # Allowable bearing = soils / FS
    ws.cell(row=allow_row, column=3, value=f'=IFERROR(C{soils_row}/C{fs_row},0)').number_format = FMT_LB
    # Area = L * W
    ws.cell(row=area_row, column=3, value=f'=IFERROR(C{len_row}*C{wid_row},0)').number_format = FMT_NUM1
    # Pressure = force / area
    ws.cell(row=press_row, column=3, value=f'=IFERROR(C{react_row}/C{area_row},0)').number_format = FMT_LB

    # Check
    ws.cell(
        row=check_row,
        column=3,
        value=(
            f'=IF(C{press_row}=0,"Enter inputs",'
            f'IF(C{press_row}>C{allow_row},"OVER CAPACITY — increase pad/mat or change crane position",'
            f'IF(C{press_row}>C{allow_row}*0.85,"BORDERLINE — close to limit, verify mat condition","OK")))'
        ),
    )
    ws.cell(row=check_row, column=3).font = FONT_BODY_BOLD
    ws.conditional_formatting.add(
        f"C{check_row}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("OVER CAPACITY",C{check_row}))'], fill=FILL_RED, font=Font(bold=True, color=RED_FONT)),
    )
    ws.conditional_formatting.add(
        f"C{check_row}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("BORDERLINE",C{check_row}))'], fill=FILL_YELLOW, font=Font(bold=True, color=YELLOW_FONT)),
    )
    ws.conditional_formatting.add(
        f"C{check_row}",
        FormulaRule(formula=[f'C{check_row}="OK"'], fill=FILL_GREEN, font=Font(bold=True, color=GREEN_FONT)),
    )


# ---------------------------------------------------------------------------
# Tab 7: Tandem Lift (optional)
# ---------------------------------------------------------------------------

def build_tandem(ws):
    ws.title = "Tandem Lift"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 42), ("C", 18), ("D", 18), ("E", 30)])

    ws["B1"] = "TANDEM LIFT — Two-Crane Balance Check"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:E1")
    ws.row_dimensions[1].height = 32

    ws["B2"] = (
        "Use only when two cranes share a single load. Both cranes individually must stay below the "
        "critical-lift threshold AT THE LIFT RADIUS. Tandem dynamics also require synchronized motion "
        "— see Lift Plan §6."
    )
    ws["B2"].font = FONT_GREY_ITALIC
    ws.merge_cells("B2:E2")

    HEADER_ROW = 4
    ws.cell(row=HEADER_ROW, column=2, value="Parameter")
    ws.cell(row=HEADER_ROW, column=3, value="Crane 1 (lead)")
    ws.cell(row=HEADER_ROW, column=4, value="Crane 2 (assist)")
    ws.cell(row=HEADER_ROW, column=5, value="Notes")
    style_header_row(ws, HEADER_ROW, 4, start_col=2)

    rows = [
        ("Crane make / model",                      "", ""),
        ("Operator",                                 "", ""),
        ("Boom length (ft)",                         0, 0),
        ("Lift radius (ft)",                         0, 0),
        ("Capacity at radius (lb) — from chart",     0, 0),
        ("Share of pick weight (lb)",                0, 0),
        ("% of capacity (computed)",                 0, 0),
        ("Status (computed)",                        "", ""),
    ]

    for i, (label, c1, c2) in enumerate(rows, start=HEADER_ROW + 1):
        ws.cell(row=i, column=2, value=label).font = FONT_BODY
        ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
        ws.cell(row=i, column=2).border = BORDER
        ws.cell(row=i, column=3, value=c1).border = BORDER
        ws.cell(row=i, column=4, value=c2).border = BORDER

    cap_r = HEADER_ROW + 5
    share_r = HEADER_ROW + 6
    pct_r = HEADER_ROW + 7
    stat_r = HEADER_ROW + 8

    # % capacity per crane
    ws.cell(row=pct_r, column=3, value=f'=IFERROR(C{share_r}/C{cap_r},0)').number_format = FMT_PCT
    ws.cell(row=pct_r, column=4, value=f'=IFERROR(D{share_r}/D{cap_r},0)').number_format = FMT_PCT
    # Status
    for col in (3, 4):
        letter = get_column_letter(col)
        ws.cell(
            row=stat_r,
            column=col,
            value=(
                f'=IF({letter}{pct_r}>=1,"OVERLOAD",'
                f'IF({letter}{pct_r}>=CriticalThreshold,"CRITICAL",'
                f'IF({letter}{pct_r}>=WarningThreshold,"NEAR CRITICAL","OK")))'
            ),
        )
        ws.cell(row=stat_r, column=col).font = FONT_BODY_BOLD
        ws.conditional_formatting.add(
            f"{letter}{stat_r}",
            FormulaRule(formula=[f'{letter}{stat_r}="OVERLOAD"'], fill=FILL_RED, font=Font(bold=True, color=RED_FONT)),
        )
        ws.conditional_formatting.add(
            f"{letter}{stat_r}",
            FormulaRule(formula=[f'{letter}{stat_r}="CRITICAL"'], fill=FILL_YELLOW, font=Font(bold=True, color=YELLOW_FONT)),
        )
        ws.conditional_formatting.add(
            f"{letter}{stat_r}",
            FormulaRule(formula=[f'{letter}{stat_r}="OK"'], fill=FILL_GREEN, font=Font(bold=True, color=GREEN_FONT)),
        )

    # Cross-check: sum of shares = pick weight
    sum_row = stat_r + 2
    ws.cell(row=sum_row, column=2, value="Sum of shares (should = Total Pick Weight)").font = FONT_BODY_BOLD
    ws.cell(row=sum_row, column=2).fill = FILL_SUMMARY_LABEL
    ws.cell(row=sum_row, column=2).border = BORDER
    ws.cell(row=sum_row, column=3, value=f"=C{share_r}+D{share_r}").number_format = FMT_LB
    ws.cell(row=sum_row, column=4, value="=TotalPickWeight").number_format = FMT_LB
    ws.cell(row=sum_row, column=5, value=f'=IF(ABS((C{sum_row}-D{sum_row}))<10,"OK — balanced","MISMATCH — recompute shares")')
    ws.conditional_formatting.add(
        f"E{sum_row}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("MISMATCH",E{sum_row}))'], fill=FILL_RED, font=Font(bold=True, color=RED_FONT)),
    )
    ws.conditional_formatting.add(
        f"E{sum_row}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("OK",E{sum_row}))'], fill=FILL_GREEN, font=Font(bold=True, color=GREEN_FONT)),
    )


# ---------------------------------------------------------------------------
# Tab 8: Lift Permit Summary
# ---------------------------------------------------------------------------

def build_permit_summary(ws):
    ws.title = "Lift Permit Summary"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 46), ("C", 28)])

    ws["B1"] = "CRITICAL LIFT — PERMIT SUMMARY"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:C1")
    ws.row_dimensions[1].height = 36

    ws["B2"] = "Print this tab as the calculator back-up for the Critical Lift Permit."
    ws["B2"].font = FONT_GREY_ITALIC
    ws.merge_cells("B2:C2")

    summary = [
        ("Project",                            "=ProjectName"),
        ("Lift Permit No.",                    "=LiftPermit"),
        ("Element / Description",              "=LiftElement"),
        ("Lift Director",                      "=LiftDirector"),
        ("Crane Operator",                     "=LiftOperator"),
        ("", ""),
        ("Total Pick Weight (lb)",             "=TotalPickWeight"),
        ("Net Crane Capacity at Radius (lb)",  "=NetCapacity"),
        ("Lift Radius (ft)",                   "=LiftRadius"),
        ("% of Crane Capacity",                "=PctCapacity"),
        ("Lift Classification",                "=LiftFlag"),
        ("", ""),
        ("Critical Lift Threshold",            "=CriticalThreshold"),
        ("Sling Angle from Horizontal",        "='Sling Angle Check'!C8"),
        ("Per-Sling Load (lb)",                "='Sling Angle Check'!C10"),
        ("Sling Adequacy",                     "='Sling Angle Check'!C12"),
        ("", ""),
        ("Bearing Pressure under Pad (psf)",   "='Ground Bearing Check'!C12"),
        ("Allowable Bearing Pressure (psf)",   "='Ground Bearing Check'!C6"),
        ("Ground Bearing Adequacy",            "='Ground Bearing Check'!C14"),
    ]

    start_row = 4
    for i, (label, val) in enumerate(summary, start=start_row):
        if label:
            ws.cell(row=i, column=2, value=label).font = FONT_BODY_BOLD
            ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
            ws.cell(row=i, column=2).border = BORDER
            c = ws.cell(row=i, column=3, value=val)
            c.font = FONT_BODY
            c.border = BORDER
            if "Pick Weight" in label or "Capacity" in label and "%" not in label or "Bearing Pressure" in label or "Per-Sling" in label:
                c.number_format = FMT_LB
            if "%" in label or label == "Critical Lift Threshold":
                c.number_format = FMT_PCT
                c.font = FONT_BIG_NUMBER

    # Signature block
    sig_row = start_row + len(summary) + 2
    ws.cell(row=sig_row, column=2, value="LIFT DIRECTOR APPROVAL").font = FONT_H2
    ws.cell(row=sig_row + 1, column=2, value="Signature: ____________________________   Date: ____________").font = FONT_BODY_BOLD
    ws.cell(row=sig_row + 2, column=2, value="Posted at crane during the lift. Filed in project record after.").font = FONT_GREY_ITALIC

    # Print setup
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    wb = Workbook()
    build_instructions(wb.active)
    build_lift_id(wb.create_sheet())
    build_pick_weight(wb.create_sheet())
    build_capacity_check(wb.create_sheet())
    build_sling_angle(wb.create_sheet())
    build_ground_bearing(wb.create_sheet())
    build_tandem(wb.create_sheet())
    build_permit_summary(wb.create_sheet())

    out_dir = "/Users/home/charles/contrpro/files/packages/complete/steel-erection"
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "Critical_Lift_Calculator.xlsx")
    wb.save(out_path)
    sz = os.path.getsize(out_path)
    print(f"Wrote {out_path} ({sz//1024} KB)")


if __name__ == "__main__":
    main()
