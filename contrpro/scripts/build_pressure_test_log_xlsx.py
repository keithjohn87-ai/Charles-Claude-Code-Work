#!/usr/bin/env python3
"""
Build ContrPro Pressure Test Log (XLSX) — Plumbing trade pack.

Companion to pipe-installation-and-test-guide.html.

Tabs:
  1. Instructions
  2. Test Schedule (planned tests by system + area)
  3. Hydrostatic Test Log (water systems)
  4. DWV Test Log (sanitary + storm)
  5. Gas Pressure Test Log
  6. Failed-Test Investigation Log
  7. Reporting Summary

Output:
    /Users/home/charles/contrpro/files/packages/complete/plumbing/Pressure_Test_Log.xlsx
"""
from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUB = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=BRAND_BLUE)
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

OUT = "/Users/home/charles/contrpro/files/packages/complete/plumbing/Pressure_Test_Log.xlsx"


def widths(ws, cols):
    for col, w in cols:
        ws.column_dimensions[col].width = w


def title(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def header(ws, row, cols):
    for i, h in enumerate(cols):
        c = ws.cell(row=row, column=1 + i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 28


def blanks(ws, start, n_rows, n_cols, fmts=None):
    fmts = fmts or {}
    for r in range(start, start + n_rows):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c, value="")
            cell.border = BORDER
            cell.font = FONT_BODY
            cell.alignment = CENTER if c > 1 else LEFT
            if c in fmts:
                cell.number_format = fmts[c]


def build_instructions(ws):
    ws.title = "Instructions"
    widths(ws, [("A", 110)])
    ws["A1"] = "PRESSURE TEST LOG — INSTRUCTIONS"
    ws["A1"].font = FONT_TITLE
    body = [
        "",
        "PURPOSE",
        "Captures every pressure / leak test performed on the project — hydrostatic water,",
        "DWV (sanitary + storm), and natural gas. Submission-ready format for the AHJ",
        "plumbing inspector and the project closeout binder.",
        "",
        "WHEN TO LOG A TEST",
        "  • Underground / in-slab — BEFORE backfill or slab pour",
        "  • Above-ground rough — BEFORE drywall closes the assembly",
        "  • Top-out stack test — DWV system pressurized for final",
        "  • Gas pressure test — BEFORE any appliance is connected + before AHJ witness",
        "  • Re-test after any repair or modification to a previously-tested assembly",
        "",
        "WHAT TO RECORD",
        "  • Test date + time + duration",
        "  • System tested + area (floor / building / wing)",
        "  • Test pressure + medium (water / air / nitrogen)",
        "  • Start + finish gauge readings",
        "  • Gauge serial # + calibration date",
        "  • Tester name + signature",
        "  • Witness (AHJ inspector / GC superintendent) — name + signature if applicable",
        "  • Pass / fail + remarks",
        "",
        "FAILED-TEST PROTOCOL",
        "Document failure in the Failed-Test Investigation Log tab. Identify the leak,",
        "perform the repair, re-test, and re-document. Don't 'erase' a failed test —",
        "the inspection trail is part of the QC system.",
        "",
        "GAUGE CALIBRATION",
        "Test gauge must read in the middle third of its scale at the test pressure (e.g.",
        "use a 0-100 psi gauge for a 50 psi test, not a 0-1000 psi gauge). Gauge must have",
        "calibration within prior 12 months — record cal date + cal lab on every test row.",
        "",
        "DOCUMENT VERSION",
        "Pressure_Test_Log.xlsx v1.0 — 2026-05-19",
        "ContrPro Complete Tier · Plumbing Trade Pack",
    ]
    for i, line in enumerate(body, start=3):
        c = ws.cell(row=i, column=1, value=line)
        if line.isupper() and line.strip() and not line.startswith(" "):
            c.font = FONT_H2
        else:
            c.font = FONT_BODY
        c.alignment = LEFT


def build_schedule(ws):
    ws.title = "Test Schedule"
    widths(ws, [("A", 26), ("B", 22), ("C", 14), ("D", 14), ("E", 18), ("F", 18), ("G", 22)])
    title(ws, 1, "TEST SCHEDULE (planned)", 7)
    header(ws, 3, ["System", "Area / Floor", "Planned Date", "Test Type", "Pressure", "Medium", "Status"])
    blanks(ws, 4, 25, 7, {3: "yyyy-mm-dd"})


def build_hydrostatic(ws):
    ws.title = "Hydrostatic Test Log"
    widths(ws, [("A", 14), ("B", 24), ("C", 22), ("D", 14), ("E", 14),
                ("F", 14), ("G", 14), ("H", 14), ("I", 14), ("J", 18), ("K", 14), ("L", 14), ("M", 22)])
    title(ws, 1, "HYDROSTATIC TEST LOG — DOMESTIC WATER + STORM", 13)
    header(ws, 3, ["Test ID", "System (CW / HW / Recirc / Storm)", "Area / Floor",
                   "Test Date", "Test Time",
                   "Test Pressure (psi)", "Start (psi)", "Finish (psi)", "Hold (min)",
                   "Gauge Serial / Cal Date", "Tester", "Pass / Fail", "Notes"])
    blanks(ws, 4, 40, 13, {4: "yyyy-mm-dd", 5: "hh:mm", 6: "0", 7: "0", 8: "0", 9: "0"})


def build_dwv(ws):
    ws.title = "DWV Test Log"
    widths(ws, [("A", 14), ("B", 22), ("C", 22), ("D", 14), ("E", 14),
                ("F", 18), ("G", 14), ("H", 18), ("I", 22), ("J", 14), ("K", 14), ("L", 22)])
    title(ws, 1, "DWV TEST LOG — SANITARY + STORM (Air or Water Head)", 12)
    header(ws, 3, ["Test ID", "System (Sanitary / Storm)", "Area / Floor",
                   "Test Date", "Test Time",
                   "Method (Water Head / 5 psi Air)", "Hold (min)",
                   "Gauge Serial / Cal Date", "Cap Failures Noted", "Tester",
                   "Pass / Fail", "Notes"])
    blanks(ws, 4, 40, 12, {4: "yyyy-mm-dd", 5: "hh:mm", 7: "0"})


def build_gas(ws):
    ws.title = "Gas Pressure Test"
    widths(ws, [("A", 14), ("B", 22), ("C", 22), ("D", 14), ("E", 14),
                ("F", 16), ("G", 14), ("H", 14), ("I", 14), ("J", 22), ("K", 22), ("L", 14), ("M", 22)])
    title(ws, 1, "GAS PRESSURE TEST LOG (IFGC 406 / NFPA 54)", 13)
    header(ws, 3, ["Test ID", "System (NG / LP)", "Area / Floor",
                   "Test Date", "Test Time",
                   "Test Pressure (psi)", "Start (psi)", "Finish (psi)", "Hold (min)",
                   "Gauge Serial / Cal Date", "Test Medium (Air / N2)", "Tester / License #",
                   "Pass / Fail + Notes"])
    blanks(ws, 4, 30, 13, {4: "yyyy-mm-dd", 5: "hh:mm", 6: "0.0", 7: "0.0", 8: "0.0", 9: "0"})


def build_failed(ws):
    ws.title = "Failed-Test Investigation"
    widths(ws, [("A", 14), ("B", 18), ("C", 22), ("D", 22), ("E", 22), ("F", 18), ("G", 14), ("H", 18)])
    title(ws, 1, "FAILED-TEST INVESTIGATION LOG", 8)
    header(ws, 3, ["Test ID (failed)", "Date", "Failure Mode", "Leak Location Found",
                   "Cause", "Repair Performed", "Re-test Date", "Re-test Result"])
    blanks(ws, 4, 25, 8, {2: "yyyy-mm-dd", 7: "yyyy-mm-dd"})


def build_summary(ws):
    ws.title = "Reporting Summary"
    widths(ws, [("A", 28), ("B", 18), ("C", 18), ("D", 18), ("E", 20)])
    title(ws, 1, "REPORTING SUMMARY (informational)", 5)
    header(ws, 3, ["System Category", "Tests Planned", "Tests Performed", "Tests Passed", "% Complete"])
    rows = [
        ("Hydrostatic — Domestic Water",),
        ("Hydrostatic — Storm",),
        ("DWV — Sanitary",),
        ("DWV — Storm",),
        ("Gas Pressure",),
        ("Other",),
    ]
    for i, (cat,) in enumerate(rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=cat).alignment = LEFT
        # COUNTA-based status — counts non-empty Test ID cells per log tab matching the category
        # (kept simple — tester fills in manually if cross-tab match is complex)
        ws.cell(row=r, column=2, value=0)
        ws.cell(row=r, column=3, value=0)
        ws.cell(row=r, column=4, value=0)
        ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/B{r},\"\")").number_format = "0.0%"
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).font = FONT_BODY
            ws.cell(row=r, column=c).alignment = CENTER if c > 1 else LEFT


def main():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb = Workbook()
    build_instructions(wb.active)
    build_schedule(wb.create_sheet("Test Schedule"))
    build_hydrostatic(wb.create_sheet("Hydrostatic Test Log"))
    build_dwv(wb.create_sheet("DWV Test Log"))
    build_gas(wb.create_sheet("Gas Pressure Test"))
    build_failed(wb.create_sheet("Failed-Test Investigation"))
    build_summary(wb.create_sheet("Reporting Summary"))
    wb.save(OUT)
    print(f"OK — wrote {OUT}")


if __name__ == "__main__":
    main()
