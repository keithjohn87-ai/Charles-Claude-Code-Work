#!/usr/bin/env python3
"""
Build ContrPro Subcontractor Tracker (XLSX) - Professional (entry) tier.

The streamlined sub master list - who they are, what trade, basic compliance
(license, COI, W9). The Pro version (separate file) adds prequal, performance
scoring, work history, hourly rates, etc.

Produces a production-grade workbook with:
  - Instructions tab (purpose, what to track, when to graduate to Pro tier)
  - Subcontractor Master List (CSI-coded, dropdowns, conditional formatting)
  - Trade Summary (COUNTIFS roll-up by trade)
  - Dashboard (totals, compliance counts, risk flags, top trades)
  - CSI Reference (hidden, named ranges for VLOOKUP / dropdowns)

Run:
    /Users/home/charles/.venv/bin/python3 \
        /Users/home/charles/contrpro/scripts/build_subcontractor_tracker_xlsx.py

Output:
    /Users/home/charles/contrpro/files/packages/business/Subcontractor_Tracker.xlsx
"""

from __future__ import annotations

import os
from typing import Tuple

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# ---------------------------------------------------------------------------
# Brand & styling (matches AR/AP/Change Order Log/Job Costing)
# ---------------------------------------------------------------------------

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
ACCENT_GOLD = "C9A227"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
YELLOW_FILL = "FFEB9C"
YELLOW_FONT = "9C5700"
ORANGE_FILL = "FFD8A8"
ORANGE_FONT = "8A4B00"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_SUMMARY_LABEL = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW_FILL)
FILL_ORANGE = PatternFill("solid", fgColor=ORANGE_FILL)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H1 = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_BIG_NUMBER = Font(name="Calibri", size=18, bold=True, color=BRAND_BLUE)
FONT_GREY_ITALIC = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_PCT = "0.0%"
FMT_INT = "0"
FMT_DATE = "yyyy-mm-dd"


# ---------------------------------------------------------------------------
# CSI MasterFormat division reference (canonical, same as sibling trackers)
# ---------------------------------------------------------------------------

DIVISIONS: list[tuple[str, str]] = [
    ("01", "General Requirements"),
    ("02", "Existing Conditions"),
    ("03", "Concrete"),
    ("04", "Masonry"),
    ("05", "Metals"),
    ("06", "Wood, Plastics, and Composites"),
    ("07", "Thermal and Moisture Protection"),
    ("08", "Openings"),
    ("09", "Finishes"),
    ("10", "Specialties"),
    ("11", "Equipment"),
    ("12", "Furnishings"),
    ("13", "Special Construction"),
    ("14", "Conveying Equipment"),
    ("21", "Fire Suppression"),
    ("22", "Plumbing"),
    ("23", "HVAC"),
    ("25", "Integrated Automation"),
    ("26", "Electrical"),
    ("27", "Communications"),
    ("28", "Electronic Safety and Security"),
    ("31", "Earthwork"),
    ("32", "Exterior Improvements"),
    ("33", "Utilities"),
    ("34", "Transportation"),
    ("35", "Waterway and Marine Construction"),
    ("40", "Process Interconnections"),
    ("41", "Material Processing and Handling Equipment"),
    ("42", "Process Heating, Cooling, and Drying Equipment"),
    ("43", "Process Gas and Liquid Handling, Purification, and Storage Equipment"),
    ("44", "Pollution and Waste Control Equipment"),
    ("45", "Industry-Specific Manufacturing Equipment"),
    ("46", "Water and Wastewater Equipment"),
    ("48", "Electrical Power Generation"),
    ("49", "Reserved"),
]


# ---------------------------------------------------------------------------
# Dropdown vocabularies
# ---------------------------------------------------------------------------

TRADES = [
    "General",
    "Mechanical",
    "Electrical",
    "Plumbing",
    "Structural Steel",
    "HVAC",
    "Roofing",
    "Concrete",
    "Masonry",
    "Drywall",
    "Flooring",
    "Painting",
    "Other",
]

LICENSE_STATUSES = ["Active", "Expired", "Pending", "Unknown"]
COI_STATES = ["Yes", "No", "Expired"]
W9_STATES = ["Yes", "No"]
PREFERRED_STATES = ["Yes", "No", "Reserved"]

# Trades that get aggregated in the Trade Summary tab. We surface every value
# from the dropdown so the user can see counts of each trade even when zero.
TRADE_SUMMARY_ORDER = TRADES


# ---------------------------------------------------------------------------
# Pre-populated examples (8 subs across the 5 trade categories called out).
# (company, contact, phone, email, address, trade, csi_div, csi_code,
#  license_num, license_state, license_status, coi, w9, preferred,
#  last_used, notes)
# ---------------------------------------------------------------------------

EXAMPLE_SUBS: list[tuple] = [
    ("EXAMPLE - Pacific Steel Erectors",
     "Marcus Reyes", "(555) 213-8842", "marcus@pacificsteel.example",
     "1240 Industrial Way, Tacoma WA 98421",
     "Structural Steel", "05", "05 12 00",
     "WA-CTR-104882", "WA", "Active",
     "Yes", "Yes", "Yes",
     "2026-04-12",
     "Solid steel erection crew. Used on Riverside Office Bldg. "
     "COI thru 2026-09-30. Preferred for steel >50 tons."),

    ("EXAMPLE - Apex Plumbing & Mechanical",
     "Dana Whitfield", "(555) 661-0034", "dana@apexpm.example",
     "5612 Service Rd, Portland OR 97211",
     "Plumbing", "22", "22 10 00",
     "OR-PLB-77231", "OR", "Active",
     "Yes", "Yes", "Yes",
     "2026-05-02",
     "Apex handles plumbing rough-in and fixtures. Estimator: Dana. "
     "COI thru 2026-12-31. Pay app cadence is monthly."),

    ("EXAMPLE - Bayside Electric LLC",
     "Renee Hu", "(555) 408-2917", "renee@baysideelec.example",
     "888 Harbor Blvd, Oakland CA 94606",
     "Electrical", "26", "26 50 00",
     "CA-C10-1004821", "CA", "Active",
     "Expired", "Yes", "No",
     "2026-02-12",
     "RISK: COI expired 2026-04-30. Pending renewed cert before "
     "any new awards. Quality is fine, paperwork is slow."),

    ("EXAMPLE - North Cascade HVAC",
     "Sam Petrov", "(555) 920-4477", "sam@ncascadehvac.example",
     "201 Mountain View Dr, Bellingham WA 98225",
     "HVAC", "23", "23 30 00",
     "WA-MEC-55619", "WA", "Active",
     "Yes", "Yes", "Yes",
     "2026-03-22",
     "Preferred HVAC. Strong on VAV controls. "
     "Owner-rep approved on Riverside and Pinnacle."),

    ("EXAMPLE - Granite West Concrete",
     "Tomas Okafor", "(555) 717-3360", "tomas@granitewest.example",
     "3300 Plant Rd, Spokane WA 99202",
     "Concrete", "03", "03 30 00",
     "WA-CON-88914", "WA", "Active",
     "Yes", "Yes", "No",
     "2026-01-18",
     "Used on small slab pours. Bigger pours we move to GraniteCore. "
     "Pricing is fair, scheduling is the friction."),

    ("EXAMPLE - Cornerstone Masonry",
     "Lila Bennett", "(555) 332-5511", "lila@cornermason.example",
     "44 Brick Yard Ln, Boise ID 83702",
     "Masonry", "04", "04 20 00",
     "ID-MSN-21908", "ID", "Pending",
     "Yes", "No", "Reserved",
     "",
     "License renewal pending with state - confirm before award. "
     "W9 not received. New relationship, no work assigned yet."),

    ("EXAMPLE - Skyline Roofing Co",
     "Jorge Alvarez", "(555) 224-6611", "jorge@skylineroof.example",
     "1101 Industrial Pkwy, Reno NV 89502",
     "Roofing", "07", "07 50 00",
     "NV-ROOF-44520", "NV", "Active",
     "Yes", "Yes", "Yes",
     "2026-04-19",
     "TPO and modified bit. Did the wind-event membrane repair on "
     "Riverside. Responsive in emergencies, premium pricing."),

    ("EXAMPLE - Allied Drywall & Finishes",
     "Kim Park", "(555) 559-2218", "kim@allieddrywall.example",
     "78 Trade Center Dr, Seattle WA 98108",
     "Drywall", "09", "09 29 00",
     "WA-CTR-203847", "WA", "Unknown",
     "No", "Yes", "No",
     "2025-11-04",
     "Used 6 months ago. License status unknown - need to verify with "
     "L&I before next award. No current COI on file."),
]

EMPTY_USER_ROWS = 25


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def set_col_widths(ws, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_header_row(ws, row: int, last_col: int) -> None:
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 28


# ---------------------------------------------------------------------------
# Hidden tab - CSI Reference (named ranges for VLOOKUP + dropdowns)
# ---------------------------------------------------------------------------

def build_csi_reference(wb: Workbook) -> str:
    ws = wb.create_sheet("CSI Reference")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "CSI MasterFormat Division Reference"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:B1")

    ws["A2"] = "Canonical list - do not edit. Feeds dropdowns and VLOOKUPs in Subcontractor Master List."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)

    ws.cell(row=3, column=1, value="CSI Division")
    ws.cell(row=3, column=2, value="Division Name")
    style_header_row(ws, 3, 2)

    for i, (code, name) in enumerate(DIVISIONS, start=4):
        ws.cell(row=i, column=1, value=code).number_format = "@"
        ws.cell(row=i, column=2, value=name)
        for c in (1, 2):
            ws.cell(row=i, column=c).border = BORDER
            ws.cell(row=i, column=c).font = FONT_BODY

    set_col_widths(ws, {"A": 14, "B": 60})

    first = 4
    last = 4 + len(DIVISIONS) - 1
    div_range = f"'CSI Reference'!$A${first}:$A${last}"
    full_range = f"'CSI Reference'!$A${first}:$B${last}"
    wb.defined_names["CSI_Divisions"] = DefinedName("CSI_Divisions", attr_text=div_range)
    wb.defined_names["CSI_Table"] = DefinedName("CSI_Table", attr_text=full_range)
    return ws.title


# ---------------------------------------------------------------------------
# Tab 1 - Instructions
# ---------------------------------------------------------------------------

def build_instructions(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {"A": 4, "B": 110})

    ws["B2"] = "ContrPro - Subcontractor Tracker"
    ws["B2"].font = FONT_TITLE
    ws.row_dimensions[2].height = 32

    ws["B3"] = ("Your master list of subs - who they are, what they do, "
                "and whether their paperwork is current. Professional tier (entry).")
    ws["B3"].font = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

    sections: list[tuple[str, list[str]]] = [
        (
            "What this workbook is",
            [
                "Four tabs you use, one hidden lookup. Subcontractor Master List is where every sub lives - "
                "contact info, trade, license, COI, W9. Trade Summary rolls them up by trade. Dashboard shows "
                "totals and compliance flags at a glance. CSI Reference is a hidden canonical list - leave it alone.",
                "This is the LEAN version. Track who you work with and confirm their paperwork is current. "
                "Nothing more. The Pro version handles prequalification, performance scoring, and work history.",
            ],
        ),
        (
            "What to track (minimum)",
            [
                "Contact - Company, primary contact, phone, email, address. Email and phone are the two you "
                "actually use, but address matters for W9 / 1099 prep at year-end.",
                "Trade - The trade dropdown (General, Mechanical, Electrical, etc.) is your filter. "
                "CSI Division pulls the canonical 2-digit code; Division Name auto-fills via VLOOKUP.",
                "Compliance - License #, License State, License Status (Active/Expired/Pending/Unknown), "
                "COI on File, W9 on File. These are the three rows you check before the sub steps on site.",
                "Preferred - Yes / No / Reserved. 'Reserved' means in your rotation but only for specific scopes.",
                "Last Used Date - When you last sent them a PO or pay app. Anyone with no entry inside 12 months "
                "should be re-verified before re-awarding.",
            ],
        ),
        (
            "Before you award - the 3-check rule",
            [
                "1. License Status = Active and License State matches the project's state of work.",
                "2. COI on File = Yes and the cert is current. If your tracker says Yes but the PDF on file is "
                "expired, that is still a No - update this tracker the same day the cert expires.",
                "3. W9 on File = Yes. No W9 means you cannot issue a 1099 in January. Get it before the first PO.",
                "Three greens or you do not award. This tracker is the receipt that says you checked.",
            ],
        ),
        (
            "When to graduate to the Pro tracker",
            [
                "Move to Subcontractor_Tracker_Pro.xlsx when you need any of these:",
                "  - Performance scoring (rating subs after each job)",
                "  - Work history (which jobs each sub has been on, $$ run through them)",
                "  - Pricing / hourly rate / markup data for estimating",
                "  - Prequalification questionnaire and capacity tracking",
                "  - Insurance limits, bonding capacity, EMR (experience mod rate)",
                "If you are awarding more than ~25 subs a year or running multiple PMs picking subs, you have "
                "outgrown this tracker. Pro is where you go next.",
            ],
        ),
        (
            "Tab-by-tab quick start",
            [
                "1. Subcontractor Master List - one row per sub. Fill in company, contact, phone, email, address, "
                "Trade (dropdown), CSI Division (dropdown), CSI Code (text like 22 10 00), license info, "
                "COI / W9 / Preferred (dropdowns), Last Used Date, Notes.",
                "2. Trade Summary - DO NOT edit. Rolls up counts by Trade automatically.",
                "3. Dashboard - DO NOT edit. Read-only totals and risk flags.",
                "4. CSI Reference - hidden by default. Unhide only if you need to inspect the canonical list.",
            ],
        ),
        (
            "Color coding cheat sheet",
            [
                "License Status: green = Active, yellow = Pending, red = Expired or Unknown.",
                "COI on File: green = Yes, red = No or Expired.",
                "W9 on File: green = Yes, red = No.",
                "Preferred: green when Yes. No styling for No or Reserved.",
                "Dashboard risk panel: counts in red = subs missing a required record.",
            ],
        ),
        (
            "Maintenance cadence",
            [
                "Monthly: scan License Status and COI columns. Anything red gets a follow-up email today.",
                "Quarterly: re-run the 3-check rule on every Preferred = Yes sub. A Preferred sub with stale paper "
                "is worse than an unknown sub.",
                "At year-end: confirm every sub you paid this year has W9 on File = Yes. The 1099 deadline is "
                "January 31 and IRS penalties are per-form.",
            ],
        ),
        (
            "Tabs in this workbook",
            [
                "Instructions  -  this tab",
                "Subcontractor Master List  -  the main worksheet",
                "Trade Summary  -  count roll-up by trade",
                "Dashboard  -  totals and risk flags",
                "CSI Reference  -  canonical division list (hidden)",
            ],
        ),
    ]

    row = 5
    for heading, paragraphs in sections:
        ws.cell(row=row, column=2, value=heading)
        ws.cell(row=row, column=2).font = FONT_H2
        ws.cell(row=row, column=2).fill = FILL_SUBHEADER
        ws.cell(row=row, column=2).alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[row].height = 22
        row += 1
        for para in paragraphs:
            ws.cell(row=row, column=2, value=para)
            ws.cell(row=row, column=2).font = FONT_BODY
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = max(18, 16 * (1 + len(para) // 110))
            row += 1
        row += 1

    ws.cell(row=row + 1, column=2,
            value="ContrPro - built for builders. Questions: support@contrpro.com")
    ws.cell(row=row + 1, column=2).font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)


# ---------------------------------------------------------------------------
# Tab 2 - Subcontractor Master List (main worksheet)
# ---------------------------------------------------------------------------

SUB_HEADERS = [
    "Company Name", "Primary Contact", "Phone", "Email", "Address",
    "Trade", "CSI Division", "CSI Code", "Division Name",
    "License Number", "License State", "License Status",
    "COI on File?", "W9 on File?", "Preferred?",
    "Last Used Date", "Notes",
]


def build_master_list(wb: Workbook) -> Tuple[str, int, int]:
    ws = wb.create_sheet("Subcontractor Master List")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Subcontractor Master List"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(SUB_HEADERS))

    # Header row at row 3
    for i, h in enumerate(SUB_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(SUB_HEADERS))

    data_start = 4
    n_pre = len(EXAMPLE_SUBS)
    n_total = n_pre + EMPTY_USER_ROWS
    data_end = data_start + n_total - 1

    COL = {h: get_column_letter(i + 1) for i, h in enumerate(SUB_HEADERS)}
    col_div = COL["CSI Division"]       # G
    col_divname = COL["Division Name"]  # I

    # Pre-populated rows
    for idx, row_data in enumerate(EXAMPLE_SUBS):
        r = data_start + idx
        (company, contact, phone, email, address, trade, csi_div, csi_code,
         lic_num, lic_state, lic_status, coi, w9, preferred,
         last_used, notes) = row_data

        ws.cell(row=r, column=1, value=company)
        ws.cell(row=r, column=2, value=contact)
        ws.cell(row=r, column=3, value=phone)
        ws.cell(row=r, column=4, value=email)
        ws.cell(row=r, column=5, value=address)
        ws.cell(row=r, column=6, value=trade)
        ws.cell(row=r, column=7, value=csi_div)
        ws.cell(row=r, column=8, value=csi_code)
        # Division Name auto-filled via VLOOKUP
        ws.cell(
            row=r, column=9,
            value=f'=IFERROR(VLOOKUP({col_div}{r},CSI_Table,2,FALSE),"")',
        )
        ws.cell(row=r, column=10, value=lic_num)
        ws.cell(row=r, column=11, value=lic_state)
        ws.cell(row=r, column=12, value=lic_status)
        ws.cell(row=r, column=13, value=coi)
        ws.cell(row=r, column=14, value=w9)
        ws.cell(row=r, column=15, value=preferred)
        ws.cell(row=r, column=16, value=last_used if last_used else None)
        ws.cell(row=r, column=17, value=notes)

    # Empty user-fillable rows (VLOOKUP pre-seeded)
    for i in range(EMPTY_USER_ROWS):
        r = data_start + n_pre + i
        ws.cell(
            row=r, column=9,
            value=f'=IFERROR(VLOOKUP({col_div}{r},CSI_Table,2,FALSE),"")',
        )

    # Per-row formatting
    for r in range(data_start, data_end + 1):
        for c in range(1, len(SUB_HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(c in (1, 5, 9, 17)),
            )
        # CSI Division and CSI Code as text so leading zeros stick
        ws.cell(row=r, column=7).number_format = "@"
        ws.cell(row=r, column=8).number_format = "@"
        # Last Used Date
        ws.cell(row=r, column=16).number_format = FMT_DATE
        ws.row_dimensions[r].height = 32

    # --- Conditional formatting ---

    # License Status (L): green Active, yellow Pending, red Expired/Unknown
    lic_range = f"L{data_start}:L{data_end}"
    ws.conditional_formatting.add(
        lic_range,
        FormulaRule(formula=[f'$L{data_start}="Active"'], stopIfTrue=False,
                    fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True)),
    )
    ws.conditional_formatting.add(
        lic_range,
        FormulaRule(formula=[f'$L{data_start}="Pending"'], stopIfTrue=False,
                    fill=FILL_YELLOW, font=Font(color=YELLOW_FONT, bold=True)),
    )
    ws.conditional_formatting.add(
        lic_range,
        FormulaRule(formula=[f'OR($L{data_start}="Expired",$L{data_start}="Unknown")'],
                    stopIfTrue=False,
                    fill=FILL_RED, font=Font(color=RED_FONT, bold=True)),
    )

    # COI on File (M): green Yes, red No/Expired
    coi_range = f"M{data_start}:M{data_end}"
    ws.conditional_formatting.add(
        coi_range,
        FormulaRule(formula=[f'$M{data_start}="Yes"'], stopIfTrue=False,
                    fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True)),
    )
    ws.conditional_formatting.add(
        coi_range,
        FormulaRule(formula=[f'OR($M{data_start}="No",$M{data_start}="Expired")'],
                    stopIfTrue=False,
                    fill=FILL_RED, font=Font(color=RED_FONT, bold=True)),
    )

    # W9 on File (N): green Yes, red No
    w9_range = f"N{data_start}:N{data_end}"
    ws.conditional_formatting.add(
        w9_range,
        FormulaRule(formula=[f'$N{data_start}="Yes"'], stopIfTrue=False,
                    fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True)),
    )
    ws.conditional_formatting.add(
        w9_range,
        FormulaRule(formula=[f'$N{data_start}="No"'], stopIfTrue=False,
                    fill=FILL_RED, font=Font(color=RED_FONT, bold=True)),
    )

    # Preferred (O): green Yes (highlight only)
    pref_range = f"O{data_start}:O{data_end}"
    ws.conditional_formatting.add(
        pref_range,
        FormulaRule(formula=[f'$O{data_start}="Yes"'], stopIfTrue=False,
                    fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True)),
    )

    # --- Data validation (dropdowns) ---

    # Trade (F)
    dv_trade = DataValidation(
        type="list", formula1='"' + ",".join(TRADES) + '"', allow_blank=True,
    )
    dv_trade.error = "Pick a trade from the dropdown."
    dv_trade.errorTitle = "Invalid trade"
    ws.add_data_validation(dv_trade)
    dv_trade.add(f"F{data_start}:F{data_end}")

    # CSI Division (G) - uses named range
    dv_div = DataValidation(
        type="list", formula1="=CSI_Divisions", allow_blank=True,
    )
    dv_div.error = "Pick a CSI Division from the dropdown."
    dv_div.errorTitle = "Invalid CSI Division"
    ws.add_data_validation(dv_div)
    dv_div.add(f"G{data_start}:G{data_end}")

    # License Status (L)
    dv_lic = DataValidation(
        type="list", formula1='"' + ",".join(LICENSE_STATUSES) + '"', allow_blank=True,
    )
    dv_lic.error = "Pick a license status from the dropdown."
    dv_lic.errorTitle = "Invalid license status"
    ws.add_data_validation(dv_lic)
    dv_lic.add(f"L{data_start}:L{data_end}")

    # COI (M)
    dv_coi = DataValidation(
        type="list", formula1='"' + ",".join(COI_STATES) + '"', allow_blank=True,
    )
    dv_coi.error = "Pick Yes / No / Expired."
    dv_coi.errorTitle = "Invalid COI value"
    ws.add_data_validation(dv_coi)
    dv_coi.add(f"M{data_start}:M{data_end}")

    # W9 (N)
    dv_w9 = DataValidation(
        type="list", formula1='"' + ",".join(W9_STATES) + '"', allow_blank=True,
    )
    dv_w9.error = "Pick Yes / No."
    dv_w9.errorTitle = "Invalid W9 value"
    ws.add_data_validation(dv_w9)
    dv_w9.add(f"N{data_start}:N{data_end}")

    # Preferred (O)
    dv_pref = DataValidation(
        type="list", formula1='"' + ",".join(PREFERRED_STATES) + '"', allow_blank=True,
    )
    dv_pref.error = "Pick Yes / No / Reserved."
    dv_pref.errorTitle = "Invalid preferred value"
    ws.add_data_validation(dv_pref)
    dv_pref.add(f"O{data_start}:O{data_end}")

    # Column widths
    set_col_widths(ws, {
        "A": 32, "B": 22, "C": 16, "D": 28, "E": 30,
        "F": 18, "G": 12, "H": 12, "I": 26,
        "J": 18, "K": 12, "L": 14,
        "M": 12, "N": 12, "O": 12,
        "P": 14, "Q": 38,
    })

    ws.freeze_panes = "B4"

    return ws.title, data_start, data_end


# ---------------------------------------------------------------------------
# Tab 3 - Trade Summary
# ---------------------------------------------------------------------------

TRADE_SUMMARY_HEADERS = [
    "Trade", "Count", "Preferred Count",
    "With COI Count", "Active License Count",
]


def build_trade_summary(wb: Workbook, sub_sheet: str, sub_start: int, sub_end: int) -> None:
    ws = wb.create_sheet("Trade Summary")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Trade Summary (auto roll-up)"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(TRADE_SUMMARY_HEADERS))

    ws["A2"] = ("Reads from Subcontractor Master List via COUNTIFS - do not edit. "
                "One row per trade in the canonical dropdown list.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)

    for i, h in enumerate(TRADE_SUMMARY_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(TRADE_SUMMARY_HEADERS))

    trade_col = f"'{sub_sheet}'!$F${sub_start}:$F${sub_end}"
    pref_col = f"'{sub_sheet}'!$O${sub_start}:$O${sub_end}"
    coi_col = f"'{sub_sheet}'!$M${sub_start}:$M${sub_end}"
    lic_col = f"'{sub_sheet}'!$L${sub_start}:$L${sub_end}"

    start_row = 4
    for idx, trade in enumerate(TRADE_SUMMARY_ORDER):
        r = start_row + idx
        ws.cell(row=r, column=1, value=trade)
        # Count
        ws.cell(row=r, column=2,
                value=f'=COUNTIF({trade_col},A{r})')
        # Preferred Count
        ws.cell(row=r, column=3,
                value=f'=COUNTIFS({trade_col},A{r},{pref_col},"Yes")')
        # With COI Count
        ws.cell(row=r, column=4,
                value=f'=COUNTIFS({trade_col},A{r},{coi_col},"Yes")')
        # Active License Count
        ws.cell(row=r, column=5,
                value=f'=COUNTIFS({trade_col},A{r},{lic_col},"Active")')

        for c in range(2, 6):
            ws.cell(row=r, column=c).number_format = FMT_INT
        for c in range(1, len(TRADE_SUMMARY_HEADERS) + 1):
            ws.cell(row=r, column=c).border = BORDER
        ws.row_dimensions[r].height = 22

    end_row = start_row + len(TRADE_SUMMARY_ORDER) - 1

    # Totals row
    tr = end_row + 2
    ws.cell(row=tr, column=1, value="TOTALS").font = FONT_BODY_BOLD
    ws.cell(row=tr, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=tr, column=1).fill = FILL_SUBHEADER
    ws.cell(row=tr, column=1).border = BORDER
    for col_letter in ("B", "C", "D", "E"):
        cell = ws.cell(row=tr, column=ord(col_letter) - ord("A") + 1,
                       value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
        cell.font = FONT_BODY_BOLD
        cell.fill = FILL_SUBHEADER
        cell.border = BORDER
        cell.number_format = FMT_INT

    # Conditional formatting - highlight cells > 0 in green
    for col_letter in ("B", "C", "D", "E"):
        ws.conditional_formatting.add(
            f"{col_letter}{start_row}:{col_letter}{end_row}",
            CellIsRule(operator="greaterThan", formula=["0"],
                       fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True)),
        )

    set_col_widths(ws, {
        "A": 22, "B": 12, "C": 18, "D": 18, "E": 22,
    })

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Tab 4 - Dashboard
# ---------------------------------------------------------------------------

def build_dashboard(wb: Workbook, sub_sheet: str, sub_start: int, sub_end: int) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Subcontractor Dashboard"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:F1")

    ws["A2"] = ("Auto-calculated from Subcontractor Master List. "
                "Update that tab - this one stays in sync.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)
    ws.merge_cells("A2:F2")

    company_col = f"'{sub_sheet}'!$A${sub_start}:$A${sub_end}"
    trade_col = f"'{sub_sheet}'!$F${sub_start}:$F${sub_end}"
    lic_col = f"'{sub_sheet}'!$L${sub_start}:$L${sub_end}"
    coi_col = f"'{sub_sheet}'!$M${sub_start}:$M${sub_end}"
    w9_col = f"'{sub_sheet}'!$N${sub_start}:$N${sub_end}"
    pref_col = f"'{sub_sheet}'!$O${sub_start}:$O${sub_end}"

    # ---------------- Top panel: master counts ----------------
    panel_start = 4
    ws.cell(row=panel_start, column=1, value="Roster Snapshot").font = FONT_H1
    ws.cell(row=panel_start, column=1).fill = FILL_HEADER
    ws.cell(row=panel_start, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=panel_start, start_column=1,
                   end_row=panel_start, end_column=6)
    ws.row_dimensions[panel_start].height = 26

    # (label, formula, fmt)
    metrics = [
        ("Total Subs in List",
         f'=COUNTA({company_col})', FMT_INT),
        ("Total Preferred Subs",
         f'=COUNTIF({pref_col},"Yes")', FMT_INT),
        ("Fully Compliant (COI + W9 + Active License)",
         f'=SUMPRODUCT(({coi_col}="Yes")*({w9_col}="Yes")*({lic_col}="Active"))',
         FMT_INT),
        ("Compliance Rate",
         f'=IFERROR(SUMPRODUCT(({coi_col}="Yes")*({w9_col}="Yes")*({lic_col}="Active"))'
         f'/COUNTA({company_col}),0)',
         FMT_PCT),
    ]

    row = panel_start + 1
    for label, formula, fmt in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = FONT_BODY_BOLD
        ws.cell(row=row, column=1).fill = FILL_SUMMARY_LABEL
        ws.cell(row=row, column=1).alignment = Alignment(vertical="center", indent=1)
        ws.cell(row=row, column=1).border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

        ws.cell(row=row, column=4, value=formula)
        ws.cell(row=row, column=4).font = FONT_BIG_NUMBER
        ws.cell(row=row, column=4).number_format = fmt
        ws.cell(row=row, column=4).alignment = Alignment(
            horizontal="right", vertical="center", indent=1,
        )
        ws.cell(row=row, column=4).border = BORDER
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 28
        row += 1

    # ---------------- Risk panel ----------------
    risk_row = row + 1
    ws.cell(row=risk_row, column=1, value="Risk Flags").font = FONT_H1
    ws.cell(row=risk_row, column=1).fill = FILL_HEADER
    ws.cell(row=risk_row, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=risk_row, start_column=1,
                   end_row=risk_row, end_column=6)
    ws.row_dimensions[risk_row].height = 26

    risk_metrics = [
        ("Missing W9",
         f'=COUNTIF({w9_col},"No")'),
        ("Missing COI",
         f'=COUNTIF({coi_col},"No")'),
        ("COI Expired",
         f'=COUNTIF({coi_col},"Expired")'),
        ("License Expired",
         f'=COUNTIF({lic_col},"Expired")'),
        ("License Pending",
         f'=COUNTIF({lic_col},"Pending")'),
        ("License Status Unknown",
         f'=COUNTIF({lic_col},"Unknown")'),
    ]

    r = risk_row + 1
    for label, formula in risk_metrics:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=1).font = FONT_BODY_BOLD
        ws.cell(row=r, column=1).fill = FILL_SUMMARY_LABEL
        ws.cell(row=r, column=1).alignment = Alignment(vertical="center", indent=1)
        ws.cell(row=r, column=1).border = BORDER
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

        ws.cell(row=r, column=4, value=formula)
        ws.cell(row=r, column=4).font = FONT_BIG_NUMBER
        ws.cell(row=r, column=4).number_format = FMT_INT
        ws.cell(row=r, column=4).alignment = Alignment(
            horizontal="right", vertical="center", indent=1,
        )
        ws.cell(row=r, column=4).border = BORDER
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 26

        # Red fill if count > 0 (any risk is bad)
        ws.conditional_formatting.add(
            f"D{r}",
            CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=False,
                       fill=FILL_RED, font=Font(color=RED_FONT, bold=True, size=18)),
        )
        ws.conditional_formatting.add(
            f"D{r}",
            CellIsRule(operator="equal", formula=["0"], stopIfTrue=False,
                       fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True, size=18)),
        )
        r += 1

    # ---------------- Top trades panel ----------------
    top_row = r + 1
    ws.cell(row=top_row, column=1, value="Top Trades by Count").font = FONT_H1
    ws.cell(row=top_row, column=1).fill = FILL_HEADER
    ws.cell(row=top_row, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=top_row, start_column=1,
                   end_row=top_row, end_column=6)
    ws.row_dimensions[top_row].height = 26

    headers = ["Rank", "Trade", "Count"]
    hr = top_row + 1
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hr, column=i, value=h)
    style_header_row(ws, hr, len(headers))
    # Merge Trade to span two cols, Count to span three
    ws.merge_cells(start_row=hr, start_column=2, end_row=hr, end_column=3)
    ws.merge_cells(start_row=hr, start_column=4, end_row=hr, end_column=6)
    # but column index 3 is "Count" header now. Re-set properly:
    ws.cell(row=hr, column=1, value="Rank")
    ws.cell(row=hr, column=2, value="Trade")
    ws.cell(row=hr, column=4, value="Count")

    # Top 5 by COUNTIF over the trade dropdown list.
    # Approach: use LARGE on an array of COUNTIFs - but openpyxl doesn't
    # write array formulas easily. Instead, compute the counts on a hidden
    # helper area at the right of the dashboard and pull the top 5 via INDEX/MATCH.
    helper_col_letter = "H"
    helper_label_col = "I"
    helper_start = top_row + 1
    for idx, trade in enumerate(TRADE_SUMMARY_ORDER):
        hr_ = helper_start + idx
        ws.cell(row=hr_, column=ord(helper_col_letter) - ord("A") + 1,
                value=f'=COUNTIF({trade_col},"{trade}")')
        ws.cell(row=hr_, column=ord(helper_label_col) - ord("A") + 1, value=trade)
    helper_end = helper_start + len(TRADE_SUMMARY_ORDER) - 1

    helper_count_range = f"${helper_col_letter}${helper_start}:${helper_col_letter}${helper_end}"
    helper_label_range = f"${helper_label_col}${helper_start}:${helper_label_col}${helper_end}"

    # Hide helper columns (visually)
    ws.column_dimensions[helper_col_letter].hidden = True
    ws.column_dimensions[helper_label_col].hidden = True

    # Display top 5
    sr = hr + 1
    for i in range(5):
        rr = sr + i
        rank = i + 1
        ws.cell(row=rr, column=1, value=rank)
        ws.cell(row=rr, column=1).font = FONT_BODY_BOLD
        ws.cell(row=rr, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=rr, column=1).border = BORDER

        # Trade = INDEX(label_range, MATCH(LARGE(count_range, i+1), count_range, 0))
        ws.cell(
            row=rr, column=2,
            value=(f'=IFERROR(INDEX({helper_label_range},'
                   f'MATCH(LARGE({helper_count_range},{rank}),{helper_count_range},0)),"")'),
        )
        ws.cell(row=rr, column=2).font = FONT_BODY_BOLD
        ws.cell(row=rr, column=2).alignment = Alignment(vertical="center", indent=1)
        ws.cell(row=rr, column=2).border = BORDER
        ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=3)

        ws.cell(
            row=rr, column=4,
            value=f'=IFERROR(LARGE({helper_count_range},{rank}),0)',
        )
        ws.cell(row=rr, column=4).font = FONT_BODY_BOLD
        ws.cell(row=rr, column=4).number_format = FMT_INT
        ws.cell(row=rr, column=4).alignment = Alignment(
            horizontal="right", vertical="center", indent=1,
        )
        ws.cell(row=rr, column=4).border = BORDER
        ws.merge_cells(start_row=rr, start_column=4, end_row=rr, end_column=6)
        ws.row_dimensions[rr].height = 22

    set_col_widths(ws, {
        "A": 22, "B": 12, "C": 14, "D": 18, "E": 14, "F": 10,
    })


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

OUTPUT_PATH = "/Users/home/charles/contrpro/files/packages/business/Subcontractor_Tracker.xlsx"


def build() -> str:
    wb = Workbook()

    # Build CSI Reference first so named ranges resolve, then move to end.
    build_instructions(wb)            # Active sheet renamed to Instructions
    build_csi_reference(wb)           # Adds named ranges CSI_Divisions, CSI_Table
    sub_sheet, sub_start, sub_end = build_master_list(wb)
    build_trade_summary(wb, sub_sheet, sub_start, sub_end)
    build_dashboard(wb, sub_sheet, sub_start, sub_end)

    desired_order = [
        "Instructions",
        "Subcontractor Master List",
        "Trade Summary",
        "Dashboard",
        "CSI Reference",
    ]
    wb._sheets = [wb[name] for name in desired_order]

    # Hide CSI Reference
    wb["CSI Reference"].sheet_state = "hidden"

    # Active tab = Dashboard on open
    wb.active = wb.sheetnames.index("Dashboard")

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
