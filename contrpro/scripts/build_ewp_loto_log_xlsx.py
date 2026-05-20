#!/usr/bin/env python3
"""
Build Energized Work Permit + LOTO Log (XLSX) — Electrical trade pack.

Tabs:
  1. Instructions
  2. LOTO Log (daily issuance + close-out)
  3. Energized Work Permit Log
  4. Voltage-Test Verification Log
  5. Arc-Flash PPE Inventory + Cal Tracking
  6. Reporting Summary
"""
from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=BRAND_BLUE)
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

OUT = "/Users/home/charles/contrpro/files/packages/complete/electrical/EWP_and_LOTO_Log.xlsx"


def widths(ws, cols):
    for col, w in cols:
        ws.column_dimensions[col].width = w


def title(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def header(ws, row, cols):
    for i, h in enumerate(cols):
        c = ws.cell(row=row, column=1 + i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 28


def blanks(ws, start, n_rows, n_cols, fmts=None):
    fmts = fmts or {}
    for r in range(start, start + n_rows):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c, value="")
            cell.border = BORDER
            cell.font = FONT_BODY
            cell.alignment = CENTER if c > 1 else LEFT
            if c in fmts:
                cell.number_format = fmts[c]


def build_instructions(ws):
    ws.title = "Instructions"
    widths(ws, [("A", 110)])
    ws["A1"] = "EWP + LOTO LOG — INSTRUCTIONS"
    ws["A1"].font = FONT_TITLE
    body = [
        "",
        "PURPOSE",
        "Single running log of every Lockout/Tagout event, every Energized Work Permit, every",
        "voltage-test verification, and the arc-flash PPE inventory + calibration tracking.",
        "Submission-ready format for OSHA inquiries + NFPA 70E audits.",
        "",
        "WORKFLOW",
        "  • LOTO Log — one row per LOTO event. Open at lockout, close at lock removal.",
        "  • EWP Log — one row per energized work permit. Includes incident energy + PPE cat.",
        "  • Voltage-Test Log — one row per test-the-tester verification (before + after).",
        "  • PPE Inventory — current inventory, cert dates, last inspection.",
        "",
        "OSHA + NFPA 70E REQUIRE",
        "  • Training records — qualified persons documented annually.",
        "  • LOTO events — logged + retained per company program.",
        "  • EWPs — retained per company program; some jurisdictions require 3+ years.",
        "  • Rubber-glove dielectric testing — every 6 months per ASTM F496.",
        "  • Meter calibration — per manufacturer; verify before high-energy work.",
        "",
        "DOCUMENT VERSION",
        "EWP_and_LOTO_Log.xlsx v1.0 — 2026-05-19",
        "ContrPro Complete Tier · Electrical Trade Pack",
    ]
    for i, line in enumerate(body, start=3):
        c = ws.cell(row=i, column=1, value=line)
        if line.isupper() and line.strip() and not line.startswith(" "):
            c.font = FONT_H2
        else:
            c.font = FONT_BODY
        c.alignment = LEFT


def build_loto(ws):
    ws.title = "LOTO Log"
    widths(ws, [("A", 8), ("B", 14), ("C", 22), ("D", 22), ("E", 22), ("F", 14), ("G", 14), ("H", 22)])
    title(ws, 1, "LOTO LOG", 8)
    header(ws, 3, ["#", "Date", "Equipment / Circuit", "Energy Sources Isolated",
                   "Locked By (Worker)", "Lock Applied", "Lock Removed", "Notes"])
    blanks(ws, 4, 60, 8, {2: "yyyy-mm-dd", 6: "hh:mm", 7: "hh:mm"})


def build_ewp(ws):
    ws.title = "EWP Log"
    widths(ws, [("A", 8), ("B", 14), ("C", 22), ("D", 14), ("E", 16), ("F", 14),
                ("G", 14), ("H", 14), ("I", 22), ("J", 22), ("K", 22)])
    title(ws, 1, "ENERGIZED WORK PERMIT LOG", 11)
    header(ws, 3, ["EWP #", "Date", "Equipment", "Voltage (V)", "Fault Current (kA)",
                   "Incident Energy (cal/cm²)", "PPE Cat", "Arc-Flash Bdy (in)",
                   "Qualified Worker(s)", "Supervisor Approver", "Notes / Justification"])
    blanks(ws, 4, 40, 11, {2: "yyyy-mm-dd", 4: "0", 5: "0.0", 6: "0.0", 8: "0"})


def build_voltage(ws):
    ws.title = "Voltage Test Log"
    widths(ws, [("A", 8), ("B", 14), ("C", 22), ("D", 12), ("E", 14), ("F", 14), ("G", 14), ("H", 22)])
    title(ws, 1, "VOLTAGE TEST VERIFICATION (test-the-tester)", 8)
    header(ws, 3, ["#", "Date", "Equipment / Circuit", "Voltage Class",
                   "Meter Serial", "Pre-Test OK?", "Post-Test OK?", "Tester"])
    blanks(ws, 4, 60, 8, {2: "yyyy-mm-dd"})


def build_ppe(ws):
    ws.title = "PPE Inventory"
    widths(ws, [("A", 8), ("B", 22), ("C", 14), ("D", 18), ("E", 14), ("F", 14), ("G", 18)])
    title(ws, 1, "ARC-FLASH PPE INVENTORY + CALIBRATION", 7)
    header(ws, 3, ["#", "Item", "Cat / Rating (cal/cm²)", "Serial / ID",
                   "Last Inspection", "Next Cal Due", "Assigned To"])
    blanks(ws, 4, 40, 7, {5: "yyyy-mm-dd", 6: "yyyy-mm-dd"})


def build_summary(ws):
    ws.title = "Reporting Summary"
    widths(ws, [("A", 30), ("B", 18), ("C", 18), ("D", 22)])
    title(ws, 1, "REPORTING SUMMARY", 4)
    header(ws, 3, ["Category", "Count (this project)", "Issues / Failures", "Notes"])
    rows = [
        "LOTO events",
        "EWPs issued",
        "Voltage-test verifications",
        "PPE items in inventory",
        "PPE items past cal date",
        "Training certifications current",
    ]
    for i, cat in enumerate(rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=cat).alignment = LEFT
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).font = FONT_BODY


def main():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb = Workbook()
    build_instructions(wb.active)
    build_loto(wb.create_sheet("LOTO Log"))
    build_ewp(wb.create_sheet("EWP Log"))
    build_voltage(wb.create_sheet("Voltage Test Log"))
    build_ppe(wb.create_sheet("PPE Inventory"))
    build_summary(wb.create_sheet("Reporting Summary"))
    wb.save(OUT)
    print(f"OK — wrote {OUT}")


if __name__ == "__main__":
    main()
