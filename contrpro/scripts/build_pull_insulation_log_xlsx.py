#!/usr/bin/env python3
"""Pull + Insulation Test Log (XLSX) — Electrical trade pack."""
from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BRAND_BLUE = "1E3A5F"
FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=BRAND_BLUE)
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

OUT = "/Users/home/charles/contrpro/files/packages/complete/electrical/Pull_and_Insulation_Test_Log.xlsx"


def widths(ws, cols):
    for col, w in cols:
        ws.column_dimensions[col].width = w


def title(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def header(ws, row, cols):
    for i, h in enumerate(cols):
        c = ws.cell(row=row, column=1 + i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 28


def blanks(ws, start, n_rows, n_cols, fmts=None):
    fmts = fmts or {}
    for r in range(start, start + n_rows):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c, value="")
            cell.border = BORDER
            cell.font = FONT_BODY
            cell.alignment = CENTER if c > 1 else LEFT
            if c in fmts:
                cell.number_format = fmts[c]


def build_instructions(ws):
    ws.title = "Instructions"
    widths(ws, [("A", 110)])
    ws["A1"] = "PULL + INSULATION TEST LOG — INSTRUCTIONS"
    ws["A1"].font = FONT_TITLE
    body = [
        "",
        "PURPOSE",
        "Captures every feeder/branch pull + megger insulation test. AHJ + NETA-style acceptance",
        "test format. Submission-ready for AHJ pre-energization inspection.",
        "",
        "WORKFLOW",
        "  • Cable Pull Log — one row per pull. Captures circuit ID, raceway, conductor size,",
        "    pull tension (if engineered), pulling lube, length, completion date.",
        "  • Insulation Test Log — one row per megger test. Includes test voltage, duration,",
        "    result, NETA-style polarization index where required.",
        "  • Acceptance Sign-off — single record per feeder/system once all tests pass.",
        "",
        "TYPICAL ACCEPTANCE THRESHOLDS",
        "  • 500 VDC test for ≤600V circuits — typical ≥100 MΩ.",
        "  • 1000 VDC test for medium-voltage — per spec.",
        "  • Polarization Index (PI) ≥ 2.0 for critical feeders (10 min ÷ 1 min reading).",
        "  • Disconnect equipment + neutrals before testing.",
        "",
        "DOCUMENT VERSION",
        "Pull_and_Insulation_Test_Log.xlsx v1.0 — 2026-05-19",
        "ContrPro Complete Tier · Electrical Trade Pack",
    ]
    for i, line in enumerate(body, start=3):
        c = ws.cell(row=i, column=1, value=line)
        if line.isupper() and line.strip() and not line.startswith(" "):
            c.font = FONT_H2
        else:
            c.font = FONT_BODY
        c.alignment = LEFT


def build_pull(ws):
    ws.title = "Cable Pull Log"
    widths(ws, [("A", 8), ("B", 14), ("C", 20), ("D", 14), ("E", 16), ("F", 12), ("G", 14), ("H", 14), ("I", 22)])
    title(ws, 1, "CABLE PULL LOG", 9)
    header(ws, 3, ["#", "Date", "Circuit / Feeder ID", "Raceway Type/Size",
                   "Conductor Size + Count", "Length (ft)", "Pull Tension (lbs)",
                   "Lube Used", "Pulled By / Notes"])
    blanks(ws, 4, 60, 9, {2: "yyyy-mm-dd", 6: "0", 7: "0"})


def build_insulation(ws):
    ws.title = "Insulation Test Log"
    widths(ws, [("A", 8), ("B", 14), ("C", 20), ("D", 14), ("E", 14), ("F", 14), ("G", 14), ("H", 14), ("I", 14), ("J", 14), ("K", 14), ("L", 22)])
    title(ws, 1, "INSULATION (MEGGER) TEST LOG", 12)
    header(ws, 3, ["#", "Date", "Circuit / Feeder ID", "Test Voltage (VDC)",
                   "Duration (s)", "A-G (MΩ)", "B-G (MΩ)", "C-G (MΩ)",
                   "PI (10/1)", "Pass/Fail", "Tester", "Notes"])
    blanks(ws, 4, 60, 12, {2: "yyyy-mm-dd", 4: "0", 5: "0", 6: "0.0", 7: "0.0", 8: "0.0", 9: "0.00"})


def build_acceptance(ws):
    ws.title = "Acceptance Sign-Off"
    widths(ws, [("A", 8), ("B", 14), ("C", 24), ("D", 14), ("E", 18), ("F", 22), ("G", 22)])
    title(ws, 1, "ACCEPTANCE SIGN-OFF", 7)
    header(ws, 3, ["#", "Date", "Feeder / System", "Result (Pass/Fail)",
                   "Inspector Witness", "Contractor Sign", "Notes"])
    blanks(ws, 4, 30, 7, {2: "yyyy-mm-dd"})


def build_summary(ws):
    ws.title = "Reporting Summary"
    widths(ws, [("A", 28), ("B", 18), ("C", 18), ("D", 22)])
    title(ws, 1, "REPORTING SUMMARY", 4)
    header(ws, 3, ["Category", "Total", "Passed", "Notes"])
    rows = [
        "Cable pulls completed",
        "Insulation tests completed",
        "Insulation tests passed first-time",
        "Insulation tests re-tested after repair",
        "Feeders accepted",
    ]
    for i, cat in enumerate(rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=cat).alignment = LEFT
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).font = FONT_BODY


def main():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb = Workbook()
    build_instructions(wb.active)
    build_pull(wb.create_sheet("Cable Pull Log"))
    build_insulation(wb.create_sheet("Insulation Test Log"))
    build_acceptance(wb.create_sheet("Acceptance Sign-Off"))
    build_summary(wb.create_sheet("Reporting Summary"))
    wb.save(OUT)
    print(f"OK — wrote {OUT}")


if __name__ == "__main__":
    main()
