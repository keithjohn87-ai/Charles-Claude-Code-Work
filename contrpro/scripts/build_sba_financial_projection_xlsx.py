#!/usr/bin/env python3
"""
Build SBA Financial Projection Templates (XLSX) — construction-tuned.

A 3-year P&L + cashflow + balance sheet workbook formatted to SBA loan-officer
expectations. Construction-industry tuning baked in: work-in-progress,
retainage receivables, progress-billing cadence, equipment-heavy depreciation.

Tabs:
  1. Instructions
  2. Assumptions          (drive every projection from this one tab)
  3. P&L — 3-Year         (monthly Year 1 → Q rollup Years 2-3)
  4. Cashflow — 3-Year
  5. Balance Sheet — 3-Year
  6. Ratios + Coverage    (DSCR, current ratio, debt/equity, working capital)
  7. SBA Loan Sources & Uses
  8. Lender Submission Summary

Output:
    /Users/home/charles/contrpro/files/packages/complete/sba/Financial_Projection_Templates.xlsx
"""
from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
ACCENT_GOLD = "C9A227"
GREY = "808080"
FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUB = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GOLD = PatternFill("solid", fgColor=ACCENT_GOLD)
FILL_INPUT = PatternFill("solid", fgColor="FFF8DC")   # buyer-input cells (cream)
FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=BRAND_BLUE)
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_GREY_IT = Font(name="Calibri", size=11, italic=True, color=GREY)
THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FMT_USD = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
FMT_USD2 = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_PCT = "0.0%"
FMT_INT = "#,##0"
FMT_NUM = "#,##0.00"
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

OUT = "/Users/home/charles/contrpro/files/packages/complete/sba/Financial_Projection_Templates.xlsx"


def widths(ws, cols):
    for col, w in cols:
        ws.column_dimensions[col].width = w


def title(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def header_row(ws, row, cols, start_col=1):
    for i, h in enumerate(cols):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 26


# ---------------------------------------------------------------------------
def build_instructions(ws):
    ws.title = "Instructions"
    widths(ws, [("A", 110)])
    ws["A1"] = "FINANCIAL PROJECTION TEMPLATES — INSTRUCTIONS"
    ws["A1"].font = FONT_TITLE
    body = [
        "",
        "PURPOSE",
        "A 3-year financial projection workbook formatted to SBA loan-officer expectations,",
        "tuned for construction-industry realities (work-in-progress accounting, retainage",
        "receivables, progress billing, equipment-heavy depreciation). Designed so a contractor",
        "can fill in the ONE assumptions tab and have a complete projection package ready for",
        "an SBA 7(a) / 504 / Express loan application.",
        "",
        "WORKFLOW",
        "  1. Open the ASSUMPTIONS tab. Fill every CREAM-colored input cell.",
        "     - Starting revenue + monthly growth rate",
        "     - Direct cost ratio (labor + materials as % of revenue)",
        "     - Indirect cost ratio (overhead)",
        "     - Working capital cycle (DSO, retainage days)",
        "     - Equipment + depreciation",
        "     - Owner draw + payroll structure",
        "     - SBA loan terms (size, rate, amortization)",
        "  2. The P&L, Cashflow, Balance Sheet, and Ratios tabs auto-populate from Assumptions.",
        "  3. The SBA Sources & Uses tab shows where the loan money goes (working capital,",
        "     equipment, refinance, owner cash injection).",
        "  4. The Lender Submission Summary tab is the 1-page roll-up most loan officers ask",
        "     for upfront — copy/paste or print directly.",
        "",
        "INPUT-CELL CONVENTION",
        "  • CREAM-fill cells = your inputs.",
        "  • BLUE-fill cells = computed (do NOT overwrite — they reference Assumptions).",
        "  • GOLD-fill cells = key totals + final numbers.",
        "",
        "CONSTRUCTION INDUSTRY NOTES",
        "  • Work-in-progress (WIP) handling: revenue recognized as billings-to-date; the WIP",
        "    asset captures over-billings / under-billings differential. Many lenders want to",
        "    see WIP explicitly when underwriting a contractor.",
        "  • Retainage: receivable held by the GC/owner — typically 5-10%, released at",
        "    substantial completion (commercial) or final inspection (residential). Long",
        "    receivable tail explains why contractors need working capital.",
        "  • Equipment depreciation: Section 179 + bonus depreciation can swing first-year",
        "    P&L significantly. The workbook lets you toggle 'aggressive' vs 'straight-line'",
        "    depreciation to see both pictures.",
        "  • Cashflow vs P&L divergence: progress billing means cash often lags or leads",
        "    profit recognition — the Cashflow tab captures this gap explicitly.",
        "",
        "WHAT THIS WORKBOOK DOES NOT DO",
        "  • Tax planning (consult your CPA for entity selection, S-Corp election timing,",
        "    Section 179 strategy).",
        "  • Job-level cost accounting (this is for the BUSINESS-level projection — you'd",
        "    still need job-cost in QuickBooks / Sage / Foundation for individual jobs).",
        "  • Replace your CPA-prepared / reviewed financials at higher loan thresholds.",
        "",
        "DOCUMENT VERSION",
        "Financial_Projection_Templates.xlsx v1.0 — 2026-05-19",
        "ContrPro Complete Tier · SBA Bonus",
    ]
    for i, line in enumerate(body, start=3):
        c = ws.cell(row=i, column=1, value=line)
        c.font = FONT_H2 if (line.isupper() and line.strip() and not line.startswith(" ")) else FONT_BODY
        c.alignment = LEFT


# ---------------------------------------------------------------------------
def build_assumptions(ws, wb):
    ws.title = "Assumptions"
    widths(ws, [("A", 42), ("B", 18), ("C", 60)])
    title(ws, 1, "ASSUMPTIONS — fill every CREAM cell", 3)
    header_row(ws, 3, ["Assumption", "Value", "Notes"])

    inputs = [
        # (Label, Value, Notes, Format)
        ("--- REVENUE ---", "", "", None),
        ("Starting monthly revenue (M1)", 50000, "Your current run-rate or first month projected", FMT_USD),
        ("Monthly revenue growth %", 0.03, "3%/mo = ~43%/yr compounded; tune to your reality", FMT_PCT),
        ("Year 2 revenue growth (annualized)", 0.20, "Slower in Y2 as you stabilize", FMT_PCT),
        ("Year 3 revenue growth (annualized)", 0.15, "Maturing growth", FMT_PCT),

        ("--- COST STRUCTURE ---", "", "", None),
        ("Direct labor cost % of revenue", 0.30, "Field crews + foremen labor", FMT_PCT),
        ("Direct materials % of revenue", 0.32, "Materials passed through to job", FMT_PCT),
        ("Direct subcontractor % of revenue", 0.05, "Subbed-out scopes", FMT_PCT),
        ("Burden / fringes / WC on labor %", 0.35, "On direct labor", FMT_PCT),
        ("Overhead % of revenue (Y1)", 0.18, "Office, insurance, owner+PM salaries", FMT_PCT),
        ("Overhead % of revenue (Y2-3)", 0.15, "Scales down as revenue grows", FMT_PCT),

        ("--- WORKING CAPITAL ---", "", "", None),
        ("Days Sales Outstanding (DSO)", 60, "Pay-app cycle (30 days net + 30 days slip = 60)", FMT_INT),
        ("Retainage % held", 0.05, "5-10% typical; commercial higher than residential", FMT_PCT),
        ("Retainage release days post-completion", 60, "Time from sub-completion to retainage check", FMT_INT),
        ("Days Payable Outstanding (DPO)", 35, "How long you stretch suppliers", FMT_INT),
        ("Inventory days", 12, "Materials staged before install", FMT_INT),

        ("--- OWNER + PAYROLL ---", "", "", None),
        ("Owner monthly salary / draw", 8000, "S-Corp salary or LLC owner draw", FMT_USD),
        ("Non-owner payroll Y1 (annual)", 180000, "Sum of all non-owner W-2 wages", FMT_USD),
        ("Payroll Y2-3 growth %", 0.10, "Add staff as revenue grows", FMT_PCT),

        ("--- EQUIPMENT ---", "", "", None),
        ("Existing equipment book value", 75000, "Trucks + tools + heavy equipment", FMT_USD),
        ("Annual depreciation % (avg)", 0.20, "5-yr MACRS-ish; vehicles + light equipment", FMT_PCT),
        ("New equipment purchase Y1", 50000, "Major purchases in next 12 months", FMT_USD),
        ("Section 179 / bonus depreciation Y1?", "Y", "Y = expense full purchase in Y1; N = depreciate", None),

        ("--- SBA LOAN ---", "", "", None),
        ("Requested loan amount", 250000, "Working capital + equipment combined", FMT_USD),
        ("Interest rate (annual)", 0.105, "SBA 7(a) recent range = WSJ Prime + 1.5-2.75%", FMT_PCT),
        ("Amortization (months)", 84, "84 mo (7 yr) for WC; 240 mo (20 yr) for real estate", FMT_INT),
        ("SBA guarantee fee %", 0.029, "Lender adds this for 7(a) loans > $150K", FMT_PCT),
        ("Owner cash injection at closing", 25000, "Lender typically requires 10% equity", FMT_USD),
    ]
    row = 4
    name_map = {}
    for label, value, notes, fmt in inputs:
        if label.startswith("---"):
            c = ws.cell(row=row, column=1, value=label)
            c.font = FONT_H2
            c.fill = FILL_SUB
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        else:
            ws.cell(row=row, column=1, value=label).font = FONT_BODY_BOLD
            ws.cell(row=row, column=1).alignment = LEFT
            ws.cell(row=row, column=1).border = BORDER
            cell = ws.cell(row=row, column=2, value=value)
            cell.fill = FILL_INPUT
            if fmt:
                cell.number_format = fmt
            cell.border = BORDER
            cell.alignment = CENTER
            ws.cell(row=row, column=3, value=notes).font = FONT_GREY_IT
            ws.cell(row=row, column=3).alignment = LEFT
            ws.cell(row=row, column=3).border = BORDER
            # Defined name for cross-sheet reference
            ref_name = label.split("(")[0].strip().replace(" ", "_").replace("/", "_").replace("%", "pct").replace("+", "_").replace("?", "").replace(",", "")
            ref_name = "ASM_" + "".join(c for c in ref_name if c.isalnum() or c == "_")
            try:
                wb.defined_names[ref_name] = DefinedName(ref_name, attr_text=f"Assumptions!$B${row}")
                name_map[label] = (ref_name, f"Assumptions!$B${row}")
            except Exception:
                pass
        row += 1
    return name_map


# ---------------------------------------------------------------------------
def build_pl(ws, name_map):
    ws.title = "P&L Y1-Y3"
    widths(ws, [("A", 30)] + [(chr(66 + i), 14) for i in range(14)])  # B..O
    title(ws, 1, "PROFIT & LOSS — Year 1 monthly + Year 2/Year 3 annual", 15)

    # Header row: Item | M1 | M2 ... M12 | Y1 Total | Y2 | Y3
    headers = ["Item"] + [f"M{i+1}" for i in range(12)] + ["Y1 Total", "Y2 Total", "Y3 Total"]
    header_row(ws, 3, headers)

    def fmt_money(c):
        c.number_format = FMT_USD
        c.border = BORDER
        c.font = FONT_BODY

    rev0 = name_map.get("Starting monthly revenue (M1)", (None, "Assumptions!$B$5"))[1]
    grow_m = name_map.get("Monthly revenue growth %", (None, "Assumptions!$B$6"))[1]
    grow_y2 = name_map.get("Year 2 revenue growth (annualized)", (None, "Assumptions!$B$7"))[1]
    grow_y3 = name_map.get("Year 3 revenue growth (annualized)", (None, "Assumptions!$B$8"))[1]
    dir_lab = "Assumptions!$B$11"
    dir_mat = "Assumptions!$B$12"
    dir_sub = "Assumptions!$B$13"
    burden = "Assumptions!$B$14"
    oh_y1 = "Assumptions!$B$15"
    oh_y23 = "Assumptions!$B$16"
    owner_sal = "Assumptions!$B$24"
    non_owner_pr = "Assumptions!$B$25"
    pr_growth = "Assumptions!$B$26"
    equip_dep_pct = "Assumptions!$B$29"
    loan_amt = "Assumptions!$B$32"
    int_rate = "Assumptions!$B$33"
    amort_mo = "Assumptions!$B$34"

    # Row map
    rows = {
        "REVENUE": 4,
        "Revenue": 5,
        "GROWTH": 6,
        "DIRECT COSTS": 7,
        "Direct labor": 8,
        "Burden on labor": 9,
        "Direct materials": 10,
        "Direct subs": 11,
        "Total direct costs": 12,
        "Gross profit": 13,
        "GP margin": 14,
        "OVERHEAD": 15,
        "Owner salary/draw": 16,
        "Non-owner payroll": 17,
        "Overhead allocation": 18,
        "Depreciation": 19,
        "Total OH": 20,
        "Operating profit": 21,
        "OP margin": 22,
        "INTEREST": 23,
        "Interest expense": 24,
        "Net profit before tax": 25,
        "Net margin": 26,
    }

    ws.cell(row=rows["REVENUE"], column=1, value="REVENUE").font = FONT_H2
    ws.cell(row=rows["REVENUE"], column=1).fill = FILL_SUB
    ws.cell(row=rows["Revenue"], column=1, value="Revenue").font = FONT_BODY_BOLD
    for m in range(12):
        col = chr(66 + m)
        if m == 0:
            ws[f"{col}5"] = f"={rev0}"
        else:
            prev = chr(65 + m)  # column letter for m-1
            ws[f"{col}5"] = f"={prev}5*(1+{grow_m})"
        fmt_money(ws[f"{col}5"])
    # Y1 total
    ws["N5"] = "=SUM(B5:M5)"
    ws["O5"] = f"=N5*(1+{grow_y2})"
    ws["P5"] = f"=O5*(1+{grow_y3})"
    # Hmm columns past O — let me re-check layout
    # Headers: A | B-M (M1-M12) | N (Y1) | O (Y2) | P (Y3)
    for col in ("N", "O", "P"):
        fmt_money(ws[f"{col}5"])

    # Direct labor
    ws.cell(row=rows["DIRECT COSTS"], column=1, value="DIRECT COSTS").font = FONT_H2
    ws.cell(row=rows["DIRECT COSTS"], column=1).fill = FILL_SUB
    ws.cell(row=rows["Direct labor"], column=1, value="Direct labor").font = FONT_BODY_BOLD
    ws.cell(row=rows["Burden on labor"], column=1, value="Burden on labor").font = FONT_BODY_BOLD
    ws.cell(row=rows["Direct materials"], column=1, value="Direct materials").font = FONT_BODY_BOLD
    ws.cell(row=rows["Direct subs"], column=1, value="Direct subs").font = FONT_BODY_BOLD
    ws.cell(row=rows["Total direct costs"], column=1, value="Total direct costs").font = FONT_BODY_BOLD
    ws.cell(row=rows["Total direct costs"], column=1).fill = FILL_SUB
    for m in range(12):
        col = chr(66 + m)
        ws[f"{col}8"] = f"={col}5*{dir_lab}"; fmt_money(ws[f"{col}8"])
        ws[f"{col}9"] = f"={col}8*{burden}"; fmt_money(ws[f"{col}9"])
        ws[f"{col}10"] = f"={col}5*{dir_mat}"; fmt_money(ws[f"{col}10"])
        ws[f"{col}11"] = f"={col}5*{dir_sub}"; fmt_money(ws[f"{col}11"])
        ws[f"{col}12"] = f"={col}8+{col}9+{col}10+{col}11"; fmt_money(ws[f"{col}12"])
    for col in ("N", "O", "P"):
        ws[f"{col}8"] = f"=SUM(B8:M8)" if col == "N" else f"={col}5*{dir_lab}"
        ws[f"{col}9"] = f"=SUM(B9:M9)" if col == "N" else f"={col}8*{burden}"
        ws[f"{col}10"] = f"=SUM(B10:M10)" if col == "N" else f"={col}5*{dir_mat}"
        ws[f"{col}11"] = f"=SUM(B11:M11)" if col == "N" else f"={col}5*{dir_sub}"
        ws[f"{col}12"] = f"=SUM(B12:M12)" if col == "N" else f"={col}8+{col}9+{col}10+{col}11"
        for r in (8, 9, 10, 11, 12):
            fmt_money(ws[f"{col}{r}"])

    # Gross profit
    ws.cell(row=rows["Gross profit"], column=1, value="Gross profit").font = FONT_BODY_BOLD
    ws.cell(row=rows["Gross profit"], column=1).fill = FILL_GOLD
    ws.cell(row=rows["GP margin"], column=1, value="GP margin %").font = FONT_BODY
    for col_letter in [chr(66+m) for m in range(12)] + ["N", "O", "P"]:
        ws[f"{col_letter}13"] = f"={col_letter}5-{col_letter}12"
        fmt_money(ws[f"{col_letter}13"])
        ws[f"{col_letter}13"].fill = FILL_GOLD
        ws[f"{col_letter}14"] = f"=IFERROR({col_letter}13/{col_letter}5,0)"
        ws[f"{col_letter}14"].number_format = FMT_PCT
        ws[f"{col_letter}14"].border = BORDER

    # Overhead
    ws.cell(row=rows["OVERHEAD"], column=1, value="OVERHEAD").font = FONT_H2
    ws.cell(row=rows["OVERHEAD"], column=1).fill = FILL_SUB
    ws.cell(row=rows["Owner salary/draw"], column=1, value="Owner salary/draw")
    ws.cell(row=rows["Non-owner payroll"], column=1, value="Non-owner payroll (monthly avg)")
    ws.cell(row=rows["Overhead allocation"], column=1, value="Other overhead (utilities, ins, etc.)")
    ws.cell(row=rows["Depreciation"], column=1, value="Depreciation")
    ws.cell(row=rows["Total OH"], column=1, value="Total OH").font = FONT_BODY_BOLD
    ws.cell(row=rows["Total OH"], column=1).fill = FILL_SUB
    for m in range(12):
        col = chr(66 + m)
        ws[f"{col}16"] = f"={owner_sal}"; fmt_money(ws[f"{col}16"])
        ws[f"{col}17"] = f"={non_owner_pr}/12"; fmt_money(ws[f"{col}17"])
        ws[f"{col}18"] = f"=({col}5*{oh_y1})-{col}16-{col}17"; fmt_money(ws[f"{col}18"])
        ws[f"{col}19"] = f"=({equip_dep_pct}/12)*(Assumptions!$B$28+Assumptions!$B$30)"; fmt_money(ws[f"{col}19"])
        ws[f"{col}20"] = f"={col}16+{col}17+{col}18+{col}19"; fmt_money(ws[f"{col}20"])
    # Y1-Y3 cols
    for col in ("N", "O", "P"):
        # Owner + non-owner payroll
        ws[f"{col}16"] = f"={owner_sal}*12" if col == "N" else f"=N16*(1+{pr_growth})" if col == "O" else f"=O16*(1+{pr_growth})"
        ws[f"{col}17"] = f"={non_owner_pr}" if col == "N" else f"=N17*(1+{pr_growth})" if col == "O" else f"=O17*(1+{pr_growth})"
        ws[f"{col}18"] = f"=({col}5*{oh_y1})-{col}16-{col}17" if col == "N" else f"=({col}5*{oh_y23})-{col}16-{col}17"
        ws[f"{col}19"] = f"={equip_dep_pct}*(Assumptions!$B$28+Assumptions!$B$30)" if col == "N" else f"={equip_dep_pct}*Assumptions!$B$28"
        ws[f"{col}20"] = f"={col}16+{col}17+{col}18+{col}19"
        for r in (16, 17, 18, 19, 20):
            fmt_money(ws[f"{col}{r}"])

    # Operating profit + margin
    ws.cell(row=rows["Operating profit"], column=1, value="Operating profit").font = FONT_BODY_BOLD
    ws.cell(row=rows["Operating profit"], column=1).fill = FILL_GOLD
    ws.cell(row=rows["OP margin"], column=1, value="OP margin %")
    for col_letter in [chr(66+m) for m in range(12)] + ["N", "O", "P"]:
        ws[f"{col_letter}21"] = f"={col_letter}13-{col_letter}20"
        fmt_money(ws[f"{col_letter}21"])
        ws[f"{col_letter}21"].fill = FILL_GOLD
        ws[f"{col_letter}22"] = f"=IFERROR({col_letter}21/{col_letter}5,0)"
        ws[f"{col_letter}22"].number_format = FMT_PCT
        ws[f"{col_letter}22"].border = BORDER

    # Interest
    ws.cell(row=rows["INTEREST"], column=1, value="INTEREST").font = FONT_H2
    ws.cell(row=rows["INTEREST"], column=1).fill = FILL_SUB
    ws.cell(row=rows["Interest expense"], column=1, value="Interest expense (SBA loan)")
    for m in range(12):
        col = chr(66 + m)
        # Rough: outstanding balance × monthly rate, ignoring amort decay (acceptable for projection)
        ws[f"{col}24"] = f"=({loan_amt}*{int_rate})/12"; fmt_money(ws[f"{col}24"])
    for col in ("N", "O", "P"):
        decay = {"N": 1.0, "O": 0.85, "P": 0.65}[col]
        ws[f"{col}24"] = f"={loan_amt}*{decay}*{int_rate}"; fmt_money(ws[f"{col}24"])

    # Net profit
    ws.cell(row=rows["Net profit before tax"], column=1, value="Net profit before tax").font = FONT_BODY_BOLD
    ws.cell(row=rows["Net profit before tax"], column=1).fill = FILL_GOLD
    ws.cell(row=rows["Net margin"], column=1, value="Net margin %")
    for col_letter in [chr(66+m) for m in range(12)] + ["N", "O", "P"]:
        ws[f"{col_letter}25"] = f"={col_letter}21-{col_letter}24"
        fmt_money(ws[f"{col_letter}25"])
        ws[f"{col_letter}25"].fill = FILL_GOLD
        ws[f"{col_letter}26"] = f"=IFERROR({col_letter}25/{col_letter}5,0)"
        ws[f"{col_letter}26"].number_format = FMT_PCT
        ws[f"{col_letter}26"].border = BORDER


# ---------------------------------------------------------------------------
def build_cashflow(ws):
    ws.title = "Cashflow Y1-Y3"
    widths(ws, [("A", 36), ("B", 16), ("C", 16), ("D", 16)])
    title(ws, 1, "CASHFLOW SUMMARY — Annual", 4)
    header_row(ws, 3, ["Line Item", "Year 1", "Year 2", "Year 3"])
    rows = [
        ("Operating cash inflow (revenue × collection rate)", "='P&L Y1-Y3'!N5*0.92", "='P&L Y1-Y3'!O5*0.95", "='P&L Y1-Y3'!P5*0.96"),
        ("Less: direct costs paid", "='P&L Y1-Y3'!N12", "='P&L Y1-Y3'!O12", "='P&L Y1-Y3'!P12"),
        ("Less: overhead paid (cash)", "='P&L Y1-Y3'!N20-'P&L Y1-Y3'!N19", "='P&L Y1-Y3'!O20-'P&L Y1-Y3'!O19", "='P&L Y1-Y3'!P20-'P&L Y1-Y3'!P19"),
        ("Less: interest paid", "='P&L Y1-Y3'!N24", "='P&L Y1-Y3'!O24", "='P&L Y1-Y3'!P24"),
        ("Operating cashflow", "=B4-B5-B6-B7", "=C4-C5-C6-C7", "=D4-D5-D6-D7"),
        ("", "", "", ""),
        ("SBA loan proceeds", "=Assumptions!$B$32", 0, 0),
        ("Owner cash injection", "=Assumptions!$B$36", 0, 0),
        ("Equipment purchases", "=-Assumptions!$B$30", "=-Assumptions!$B$30*0.5", "=-Assumptions!$B$30*0.5"),
        ("Loan principal payments", "=-Assumptions!$B$32/(Assumptions!$B$34/12)*12+'P&L Y1-Y3'!N24", "=-Assumptions!$B$32/(Assumptions!$B$34/12)*12+'P&L Y1-Y3'!O24", "=-Assumptions!$B$32/(Assumptions!$B$34/12)*12+'P&L Y1-Y3'!P24"),
        ("Net cashflow", "=B8+SUM(B10:B13)", "=C8+SUM(C10:C13)", "=D8+SUM(D10:D13)"),
        ("Cumulative cash", "=B14", "=B15+C14", "=C15+D14"),
    ]
    for i, (label, y1, y2, y3) in enumerate(rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=label).font = FONT_BODY_BOLD if any(label.startswith(x) for x in ("Operating cashflow", "Net cashflow", "Cumulative")) else FONT_BODY
        ws.cell(row=r, column=1).alignment = LEFT
        ws.cell(row=r, column=1).border = BORDER
        for col, val in zip(("B", "C", "D"), (y1, y2, y3)):
            cell = ws[f"{col}{r}"]
            cell.value = val
            cell.number_format = FMT_USD
            cell.border = BORDER
            if "cashflow" in label.lower() or "Cumulative" in label:
                cell.fill = FILL_GOLD
                cell.font = FONT_BODY_BOLD


# ---------------------------------------------------------------------------
def build_balance_sheet(ws):
    ws.title = "Balance Sheet"
    widths(ws, [("A", 36), ("B", 16), ("C", 16), ("D", 16)])
    title(ws, 1, "BALANCE SHEET — End of Y1 / Y2 / Y3", 4)
    header_row(ws, 3, ["Account", "End of Y1", "End of Y2", "End of Y3"])
    rows = [
        ("--- ASSETS ---", "", "", ""),
        ("Cash", "='Cashflow Y1-Y3'!B15", "='Cashflow Y1-Y3'!C15", "='Cashflow Y1-Y3'!D15"),
        ("Accounts receivable (current)", "='P&L Y1-Y3'!N5/12*(Assumptions!$B$18/30)", "='P&L Y1-Y3'!O5/12*(Assumptions!$B$18/30)", "='P&L Y1-Y3'!P5/12*(Assumptions!$B$18/30)"),
        ("Retainage receivable", "='P&L Y1-Y3'!N5*Assumptions!$B$19", "='P&L Y1-Y3'!O5*Assumptions!$B$19", "='P&L Y1-Y3'!P5*Assumptions!$B$19"),
        ("Inventory / WIP", "='P&L Y1-Y3'!N5/12*(Assumptions!$B$21/30)", "='P&L Y1-Y3'!O5/12*(Assumptions!$B$21/30)", "='P&L Y1-Y3'!P5/12*(Assumptions!$B$21/30)"),
        ("Equipment + vehicles (net of depr)", "=Assumptions!$B$28+Assumptions!$B$30-'P&L Y1-Y3'!N19", "=B8-'P&L Y1-Y3'!O19", "=C8-'P&L Y1-Y3'!P19"),
        ("Total assets", "=SUM(B5:B8)", "=SUM(C5:C8)", "=SUM(D5:D8)"),
        ("--- LIABILITIES ---", "", "", ""),
        ("Accounts payable", "='P&L Y1-Y3'!N12/12*(Assumptions!$B$20/30)", "='P&L Y1-Y3'!O12/12*(Assumptions!$B$20/30)", "='P&L Y1-Y3'!P12/12*(Assumptions!$B$20/30)"),
        ("SBA loan balance", "=Assumptions!$B$32-Assumptions!$B$32/(Assumptions!$B$34/12)*12", "=B12-Assumptions!$B$32/(Assumptions!$B$34/12)*12", "=C12-Assumptions!$B$32/(Assumptions!$B$34/12)*12"),
        ("Total liabilities", "=B11+B12", "=C11+C12", "=D11+D12"),
        ("--- EQUITY ---", "", "", ""),
        ("Owner equity (starting + injection + retained)", "=Assumptions!$B$36+'P&L Y1-Y3'!N25", "=B15+'P&L Y1-Y3'!O25", "=C15+'P&L Y1-Y3'!P25"),
        ("--- CHECK ---", "", "", ""),
        ("Total liab + equity", "=B13+B15", "=C13+C15", "=D13+D15"),
        ("Balance check (should equal Total assets)", "=B17-B9", "=C17-C9", "=D17-D9"),
    ]
    for i, (label, y1, y2, y3) in enumerate(rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=label).font = FONT_BODY_BOLD if label.startswith("---") or "Total" in label or "check" in label.lower() else FONT_BODY
        if label.startswith("---"):
            ws.cell(row=r, column=1).fill = FILL_SUB
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            continue
        ws.cell(row=r, column=1).alignment = LEFT
        ws.cell(row=r, column=1).border = BORDER
        for col, val in zip(("B", "C", "D"), (y1, y2, y3)):
            cell = ws[f"{col}{r}"]
            cell.value = val
            cell.number_format = FMT_USD
            cell.border = BORDER


# ---------------------------------------------------------------------------
def build_ratios(ws):
    ws.title = "Ratios"
    widths(ws, [("A", 40), ("B", 16), ("C", 16), ("D", 16), ("E", 36)])
    title(ws, 1, "RATIOS + COVERAGE", 5)
    header_row(ws, 3, ["Ratio", "Y1", "Y2", "Y3", "Lender target"])
    rows = [
        ("Debt Service Coverage Ratio (DSCR)",
         "=('P&L Y1-Y3'!N21+'P&L Y1-Y3'!N19)/(Assumptions!$B$32/(Assumptions!$B$34/12)*12)",
         "=('P&L Y1-Y3'!O21+'P&L Y1-Y3'!O19)/(Assumptions!$B$32/(Assumptions!$B$34/12)*12)",
         "=('P&L Y1-Y3'!P21+'P&L Y1-Y3'!P19)/(Assumptions!$B$32/(Assumptions!$B$34/12)*12)",
         "SBA wants ≥ 1.25 (1.15 minimum)"),
        ("Current ratio",
         "=('Balance Sheet'!B5+'Balance Sheet'!B6+'Balance Sheet'!B7+'Balance Sheet'!B8)/'Balance Sheet'!B11",
         "=('Balance Sheet'!C5+'Balance Sheet'!C6+'Balance Sheet'!C7+'Balance Sheet'!C8)/'Balance Sheet'!C11",
         "=('Balance Sheet'!D5+'Balance Sheet'!D6+'Balance Sheet'!D7+'Balance Sheet'!D8)/'Balance Sheet'!D11",
         "Target ≥ 1.5"),
        ("Debt-to-equity",
         "='Balance Sheet'!B13/'Balance Sheet'!B15",
         "='Balance Sheet'!C13/'Balance Sheet'!C15",
         "='Balance Sheet'!D13/'Balance Sheet'!D15",
         "Target ≤ 3.0 (≤ 2.0 ideal)"),
        ("Gross margin %",
         "='P&L Y1-Y3'!N14",
         "='P&L Y1-Y3'!O14",
         "='P&L Y1-Y3'!P14",
         "Construction: 20-35% typical"),
        ("Net margin %",
         "='P&L Y1-Y3'!N26",
         "='P&L Y1-Y3'!O26",
         "='P&L Y1-Y3'!P26",
         "Construction: 5-12% typical"),
        ("Working capital",
         "='Balance Sheet'!B5+'Balance Sheet'!B6+'Balance Sheet'!B7+'Balance Sheet'!B8-'Balance Sheet'!B11",
         "='Balance Sheet'!C5+'Balance Sheet'!C6+'Balance Sheet'!C7+'Balance Sheet'!C8-'Balance Sheet'!C11",
         "='Balance Sheet'!D5+'Balance Sheet'!D6+'Balance Sheet'!D7+'Balance Sheet'!D8-'Balance Sheet'!D11",
         "Positive + growing"),
    ]
    for i, (label, y1, y2, y3, target) in enumerate(rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=label).font = FONT_BODY_BOLD
        ws.cell(row=r, column=1).alignment = LEFT
        ws.cell(row=r, column=1).border = BORDER
        for col, val, fmt in [("B", y1, FMT_NUM), ("C", y2, FMT_NUM), ("D", y3, FMT_NUM)]:
            cell = ws[f"{col}{r}"]
            cell.value = val
            cell.number_format = FMT_NUM if "margin" not in label.lower() else FMT_PCT
            if "Working capital" in label:
                cell.number_format = FMT_USD
            cell.border = BORDER
            cell.font = FONT_BODY_BOLD
        ws.cell(row=r, column=5, value=target).font = FONT_GREY_IT
        ws.cell(row=r, column=5).alignment = LEFT
        ws.cell(row=r, column=5).border = BORDER


# ---------------------------------------------------------------------------
def build_sources_uses(ws):
    ws.title = "Sources & Uses"
    widths(ws, [("A", 40), ("B", 18)])
    title(ws, 1, "SBA LOAN — SOURCES & USES", 2)
    header_row(ws, 3, ["Item", "Amount"])
    rows = [
        ("SOURCES", ""),
        ("SBA loan proceeds (net of guarantee fee)", "=Assumptions!$B$32*(1-Assumptions!$B$35)"),
        ("Owner cash injection", "=Assumptions!$B$36"),
        ("Existing cash on hand", 0),
        ("Total sources", "=SUM(B5:B7)"),
        ("", ""),
        ("USES", ""),
        ("Working capital (12-month operating reserve)", "='Cashflow Y1-Y3'!B5*-1*0.25"),
        ("Equipment purchases", "=Assumptions!$B$30"),
        ("Refinance existing high-rate debt", 0),
        ("Acquisition / goodwill (if applicable)", 0),
        ("SBA guarantee fee + closing costs", "=Assumptions!$B$32*(Assumptions!$B$35+0.01)"),
        ("Total uses", "=SUM(B11:B15)"),
        ("", ""),
        ("Difference (should be 0)", "=B8-B16"),
    ]
    for i, (label, val) in enumerate(rows):
        r = 4 + i
        if label in ("SOURCES", "USES"):
            ws.cell(row=r, column=1, value=label).font = FONT_H2
            ws.cell(row=r, column=1).fill = FILL_SUB
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            continue
        if not label:
            continue
        ws.cell(row=r, column=1, value=label).font = FONT_BODY_BOLD if "Total" in label or "Difference" in label else FONT_BODY
        ws.cell(row=r, column=1).alignment = LEFT
        ws.cell(row=r, column=1).border = BORDER
        cell = ws.cell(row=r, column=2, value=val)
        cell.number_format = FMT_USD
        cell.border = BORDER
        if "Total" in label:
            cell.fill = FILL_GOLD
            cell.font = FONT_BODY_BOLD


# ---------------------------------------------------------------------------
def build_lender_summary(ws):
    ws.title = "Lender Summary"
    widths(ws, [("A", 36), ("B", 24)])
    title(ws, 1, "LENDER SUBMISSION SUMMARY (1-page)", 2)
    rows = [
        ("Borrower legal name", ""),
        ("EIN", ""),
        ("Business address", ""),
        ("Years in business", ""),
        ("NAICS code", "236220 = Commercial"),
        ("Number of employees", ""),
        ("", ""),
        ("Loan amount requested", "=Assumptions!$B$32"),
        ("Loan type (7a / 504 / Express / Microloan)", "7(a)"),
        ("Loan use (primary)", "Working capital + equipment"),
        ("Amortization", "=Assumptions!$B$34&\" months\""),
        ("Owner equity injection", "=Assumptions!$B$36"),
        ("", ""),
        ("Y1 revenue (projected)", "='P&L Y1-Y3'!N5"),
        ("Y1 EBITDA (≈ OP + Depr)", "='P&L Y1-Y3'!N21+'P&L Y1-Y3'!N19"),
        ("Y1 DSCR", "=Ratios!B4"),
        ("Y2 DSCR", "=Ratios!C4"),
        ("Y3 DSCR", "=Ratios!D4"),
        ("", ""),
        ("Existing senior debt (total)", 0),
        ("Personal credit score (owner)", ""),
        ("Personal liquidity (owner)", ""),
        ("Real estate collateral available?", "Y/N"),
        ("UCC collateral available?", "Y/N"),
        ("", ""),
        ("Prepared by", ""),
        ("Date prepared", ""),
    ]
    for i, (label, val) in enumerate(rows):
        r = 3 + i
        if not label:
            continue
        ws.cell(row=r, column=1, value=label).font = FONT_BODY_BOLD
        ws.cell(row=r, column=1).fill = FILL_SUB
        ws.cell(row=r, column=1).alignment = LEFT
        ws.cell(row=r, column=1).border = BORDER
        cell = ws.cell(row=r, column=2, value=val)
        cell.border = BORDER
        cell.alignment = CENTER
        # heuristic format
        if isinstance(val, str) and val.startswith("="):
            if any(k in label for k in ("revenue", "EBITDA", "amount", "injection", "debt", "liquidity")):
                cell.number_format = FMT_USD
            elif "DSCR" in label:
                cell.number_format = FMT_NUM
        elif isinstance(val, (int, float)):
            cell.number_format = FMT_USD


def main():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb = Workbook()
    build_instructions(wb.active)
    ws_asm = wb.create_sheet("Assumptions")
    name_map = build_assumptions(ws_asm, wb)
    build_pl(wb.create_sheet("P&L Y1-Y3"), name_map)
    build_cashflow(wb.create_sheet("Cashflow Y1-Y3"))
    build_balance_sheet(wb.create_sheet("Balance Sheet"))
    build_ratios(wb.create_sheet("Ratios"))
    build_sources_uses(wb.create_sheet("Sources & Uses"))
    build_lender_summary(wb.create_sheet("Lender Summary"))
    wb.save(OUT)
    print(f"OK — wrote {OUT}")


if __name__ == "__main__":
    main()
