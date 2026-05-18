#!/usr/bin/env python3
"""
Build ContrPro Sub Schedule of Values (XLSX) — Universal Sub Suite (Complete tier).

The sub-perspective companion to the GC's Application for Payment. Mirrors the
structure of AIA G-703 (Continuation Sheet) without reproducing the AIA form.
Used by a subcontractor to schedule its Subcontract Sum across line items and
to bill monthly progress against those line items.

Tabs:
  1. Instructions             (how to use, G-703 anatomy, retainage handling)
  2. Project Info             (anchor for project name, GC, period, percentages)
  3. SOV Setup                (1 row per line item — description, CSI code, scheduled value)
  4. Pay App — Current Period (G-703-style; this period, total to date, retainage)
  5. Pay App — History        (12-month roll-up; one column per billing period)
  6. Stored Materials Log     (Materials on site / off site not yet installed)
  7. CSI Reference            (Hidden — drives the CSI division dropdown)

Run:
    /Users/home/charles/.venv/bin/python3 \\
        /Users/home/charles/contrpro/scripts/build_sub_sov_xlsx.py

Output:
    /Users/home/charles/contrpro/files/packages/complete/sub/Sub_Schedule_of_Values.xlsx
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# ---------------------------------------------------------------------------
# Brand & styling (matches the rest of the ContrPro XLSX library)
# ---------------------------------------------------------------------------

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
ACCENT_GOLD = "C9A227"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
YELLOW_FILL = "FFEB9C"
YELLOW_FONT = "9C5700"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_SUMMARY_LABEL = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW_FILL)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H1 = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_BIG_NUMBER = Font(name="Calibri", size=18, bold=True, color=BRAND_BLUE)
FONT_GREY_ITALIC = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_PCT = "0.0%"
FMT_INT = "0"
FMT_DATE = "yyyy-mm-dd"

# ---------------------------------------------------------------------------
# CSI MasterFormat divisions — same source as the other ContrPro trackers
# ---------------------------------------------------------------------------

DIVISIONS: list[tuple[str, str]] = [
    ("01", "General Requirements"),
    ("02", "Existing Conditions"),
    ("03", "Concrete"),
    ("04", "Masonry"),
    ("05", "Metals"),
    ("06", "Wood, Plastics, and Composites"),
    ("07", "Thermal and Moisture Protection"),
    ("08", "Openings"),
    ("09", "Finishes"),
    ("10", "Specialties"),
    ("11", "Equipment"),
    ("12", "Furnishings"),
    ("13", "Special Construction"),
    ("14", "Conveying Equipment"),
    ("21", "Fire Suppression"),
    ("22", "Plumbing"),
    ("23", "HVAC"),
    ("25", "Integrated Automation"),
    ("26", "Electrical"),
    ("27", "Communications"),
    ("28", "Electronic Safety and Security"),
    ("31", "Earthwork"),
    ("32", "Exterior Improvements"),
    ("33", "Utilities"),
    ("34", "Transportation"),
    ("35", "Waterway and Marine Construction"),
    ("40", "Process Interconnections"),
    ("41", "Material Processing and Handling Equipment"),
    ("42", "Process Heating, Cooling, and Drying Equipment"),
    ("43", "Process Gas and Liquid Handling, Purification, and Storage Equipment"),
    ("44", "Pollution and Waste Control Equipment"),
    ("45", "Industry-Specific Manufacturing Equipment"),
    ("46", "Water and Wastewater Equipment"),
    ("48", "Electrical Power Generation"),
    ("49", "Reserved"),
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def set_col_widths(ws, widths: list[tuple[str, float]]):
    for col, w in widths:
        ws.column_dimensions[col].width = w


def style_header_row(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


def apply_body_style(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_BODY
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.border = BORDER


# ---------------------------------------------------------------------------
# Tab 1: Instructions
# ---------------------------------------------------------------------------

def build_instructions(ws):
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 100)])

    ws["A1"] = "SUB SCHEDULE OF VALUES — Instructions"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 36

    rows = [
        "",
        "WHAT THIS WORKBOOK IS",
        "This is the sub-perspective companion to AIA G-703 (Continuation Sheet for an Application for "
        "Payment). It lets you (1) schedule your Subcontract Sum across line items in SOV Setup, (2) bill "
        "monthly progress against those line items on the Pay App — Current Period tab, and (3) maintain a "
        "12-month rolling history of billings on the Pay App — History tab.",
        "",
        "WHY A SUB NEEDS ITS OWN SOV",
        "Even when the GC supplies a G-702 / G-703 to be returned with each pay app, the sub should maintain "
        "its own SOV ledger because: (a) the GC version is project-specific and will be filed with the GC's "
        "pay app, not yours; (b) retainage release tracking, lower-tier sub billings, and stored-materials "
        "values often diverge between what you submit to the GC and what your accounting actually tracks; "
        "and (c) if a dispute arises over partial payment or retainage, your own internal ledger is the "
        "evidentiary record.",
        "",
        "WORKFLOW — MONTHLY PAY APP",
        "1. Set up SOV Setup once at Subcontract execution. Each line item has a description, a CSI division/"
        "section, and a scheduled value. The scheduled values must sum to the Subcontract Sum on Project Info.",
        "2. At the end of each billing period (typically the 25th of the month for billing on the 1st of the "
        "next month, but confirm against the Subcontract), open Pay App — Current Period.",
        "3. For each line item, enter the % complete this period (column F). The workbook calculates this-"
        "period dollars, total-to-date dollars, retainage held, and net amount this period.",
        "4. Add stored materials values from the Stored Materials Log to the appropriate line items.",
        "5. Verify the totals against the GC's G-702 summary if the GC has issued one.",
        "6. After the GC has accepted and paid the pay app, transfer the column G (total to date) values into "
        "the next month's column on Pay App — History.",
        "",
        "RETAINAGE HANDLING",
        "Default retainage rate is set on Project Info (5% or 10% — confirm against Subcontract). Retainage is "
        "held on completed-work value and (depending on state law and Subcontract language) on stored "
        "materials. The workbook supports either treatment: set 'Retain on stored materials?' on Project Info "
        "to Yes or No. If your state caps retainage at a lower rate (e.g., CA at 5%), update the rate before "
        "first pay app. If your Subcontract allows retainage reduction at 50% Substantial Completion (an "
        "industry-standard provision), keep the rate at the full 10% in this workbook and reduce it on the "
        "specific pay app where reduction is requested.",
        "",
        "STORED MATERIALS",
        "Stored materials are billable when (a) the Subcontract permits it, (b) the materials are stored at a "
        "location protected from theft and weather, (c) materials are dedicated to the Project (segregated and "
        "labeled, with title transferring to Owner when billed), and (d) insurance / bonding is in place per "
        "the Subcontract. Use the Stored Materials Log to document each item, location, insurance, and lien "
        "waiver from the supplier.",
        "",
        "TIPS",
        "  - Never bill more than is physically complete. Front-loading the SOV (overweighting early line items "
        "like mobilization or layout) is a flag the GC's PM will catch and creates compounding problems on "
        "Subcontract close-out.",
        "  - Track every change order in a separate row on SOV Setup with a 'CO #' prefix; do not modify the "
        "scheduled value of base-bid line items.",
        "  - The 'Variance' column on SOV Setup compares the billed-to-date amount to the scheduled value. A "
        "consistently negative variance late in the project is the early warning of either scope creep that "
        "should have been a change order or front-loading that is now catching up.",
        "  - At Substantial Completion, the retainage held in column J should match the GC's accounting; "
        "reconcile before requesting retainage release.",
        "",
        "ON CSI MASTERFORMAT",
        "Each SOV line carries a CSI division and section. CSI MasterFormat is the industry standard for "
        "categorizing construction work — coding the SOV this way lets the GC's accounting system map your "
        "bills to its own job-cost ledger and lets you roll up totals by trade across projects.",
        "",
        "TROUBLESHOOTING",
        "  - 'Variance' column shows red: billed exceeds scheduled. Either bill less this period or add a "
        "change order on SOV Setup for the overage.",
        "  - SOV Setup totals don't equal Subcontract Sum on Project Info: cross-check each line, and confirm "
        "no rows are blank or missing values. The cell C5 of Project Info validates this.",
        "  - Retainage on Pay App doesn't match GC's: confirm Project Info retainage rate matches Subcontract "
        "and confirm 'Retain on stored materials?' setting matches the GC's practice.",
    ]

    for i, text in enumerate(rows, start=2):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if text and text == text.upper() and len(text) < 60:
            cell.font = FONT_H2
        else:
            cell.font = FONT_BODY


# ---------------------------------------------------------------------------
# Tab 2: Project Info
# ---------------------------------------------------------------------------

def build_project_info(ws):
    ws.title = "Project Info"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 36), ("C", 40)])

    ws["B1"] = "PROJECT INFORMATION"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:C1")
    ws.row_dimensions[1].height = 32

    info_rows = [
        ("Project Name", ""),
        ("Project No. (GC's)", ""),
        ("Subcontract No.", ""),
        ("Subcontractor (legal name)", ""),
        ("General Contractor (legal name)", ""),
        ("Owner", ""),
        ("Project Address", ""),
        ("Trade / Scope", ""),
        ("CSI Division (primary)", ""),
        ("Subcontract Date", ""),
        ("Original Subcontract Sum", 0),
        ("Approved Change Orders to date", 0),
        ("Current Subcontract Sum (Orig + COs)", "=C13+C14"),
        ("Retainage Rate (%) — confirm Subcontract", 0.10),
        ("Retain on stored materials?", "No"),
        ("Pay App Submission Deadline (day of month)", 25),
        ("Pay App Payment Due (days after GC acceptance)", 30),
        ("Project Manager (Sub)", ""),
        ("Project Manager (GC)", ""),
    ]

    for i, (label, val) in enumerate(info_rows, start=3):
        ws.cell(row=i, column=2, value=label).font = FONT_BODY_BOLD
        ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
        ws.cell(row=i, column=2).border = BORDER
        c = ws.cell(row=i, column=3, value=val)
        c.font = FONT_BODY
        c.border = BORDER
        if "Sum" in label or "Change Orders" in label or "Original" in label:
            c.number_format = FMT_USD
        if "Rate" in label:
            c.number_format = FMT_PCT
        if "Deadline" in label or "Due" in label or "day" in label:
            c.number_format = FMT_INT

    # Named ranges for cross-sheet references
    for name, ref in [
        ("ProjectName", "'Project Info'!$C$3"),
        ("SubContractSum", "'Project Info'!$C$15"),
        ("RetainageRate", "'Project Info'!$C$16"),
        ("RetainStored", "'Project Info'!$C$17"),
    ]:
        ws.parent.defined_names[name] = DefinedName(name=name, attr_text=ref)

    # Yes/No validation on retainage on stored
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    dv.add("C17")
    ws.add_data_validation(dv)

    # SOV-validation banner: shows whether SOV Setup totals match Subcontract Sum
    ws["B24"] = "SOV TOTAL VALIDATION"
    ws["B24"].font = FONT_H2
    ws["B25"] = "SOV Setup line-item total:"
    ws["B25"].font = FONT_BODY_BOLD
    ws["B25"].fill = FILL_SUMMARY_LABEL
    ws["C25"] = "=SUM('SOV Setup'!E:E)"
    ws["C25"].number_format = FMT_USD
    ws["C25"].font = FONT_BODY
    ws["B26"] = "Current Subcontract Sum:"
    ws["B26"].font = FONT_BODY_BOLD
    ws["B26"].fill = FILL_SUMMARY_LABEL
    ws["C26"] = "=C15"
    ws["C26"].number_format = FMT_USD
    ws["C26"].font = FONT_BODY
    ws["B27"] = "Variance (should be $0):"
    ws["B27"].font = FONT_BODY_BOLD
    ws["B27"].fill = FILL_SUMMARY_LABEL
    ws["C27"] = "=C25-C26"
    ws["C27"].number_format = FMT_USD
    ws["C27"].font = FONT_BIG_NUMBER

    # Conditional format on the variance cell — green if zero, red otherwise
    ws.conditional_formatting.add(
        "C27",
        FormulaRule(formula=["ABS(C27)<0.01"], fill=FILL_GREEN, font=Font(color=GREEN_FONT)),
    )
    ws.conditional_formatting.add(
        "C27",
        FormulaRule(formula=["ABS(C27)>=0.01"], fill=FILL_RED, font=Font(color=RED_FONT)),
    )


# ---------------------------------------------------------------------------
# Tab 3: SOV Setup
# ---------------------------------------------------------------------------

SOV_SETUP_COLS = [
    ("A", "Line #", 8),
    ("B", "Description", 38),
    ("C", "CSI Div", 9),
    ("D", "CSI Section", 14),
    ("E", "Scheduled Value ($)", 18),
    ("F", "% of Subcontract", 14),
    ("G", "Billed To Date ($)", 18),
    ("H", "Balance to Finish ($)", 18),
    ("I", "Variance", 14),
    ("J", "Type", 14),
    ("K", "Notes", 30),
]


def build_sov_setup(ws):
    ws.title = "SOV Setup"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in SOV_SETUP_COLS])

    ws["A1"] = "SOV SETUP — Line Items"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:K1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Set up one row per Subcontract line item. Scheduled Value must sum to Current Subcontract Sum "
        "on Project Info. Add change orders as separate rows prefixed 'CO #'."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:K2")
    ws.row_dimensions[2].height = 28

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(SOV_SETUP_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(SOV_SETUP_COLS))

    # Body — 50 rows starter
    DATA_ROWS = 50
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        # Line # auto-numbering
        ws.cell(row=r, column=1, value=f"=IF(B{r}=\"\",\"\",ROW()-{HEADER_ROW})")
        # % of Subcontract Sum
        ws.cell(row=r, column=6, value=f"=IF(E{r}=\"\",\"\",E{r}/SubContractSum)")
        ws.cell(row=r, column=6).number_format = FMT_PCT
        # Billed To Date — pulled from Pay App Current Period (this period + previous)
        ws.cell(
            row=r,
            column=7,
            value=f"=IFERROR(VLOOKUP(B{r},'Pay App - Current'!$B$5:$H$54,7,FALSE),0)",
        )
        ws.cell(row=r, column=7).number_format = FMT_USD
        # Balance to finish
        ws.cell(row=r, column=8, value=f"=IF(E{r}=\"\",\"\",E{r}-G{r})")
        ws.cell(row=r, column=8).number_format = FMT_USD
        # Variance — over-billed flag
        ws.cell(row=r, column=9, value=f"=IF(E{r}=\"\",\"\",G{r}-E{r})")
        ws.cell(row=r, column=9).number_format = FMT_USD

        ws.cell(row=r, column=5).number_format = FMT_USD
        apply_body_style(ws, r, len(SOV_SETUP_COLS))

    # Totals row
    TOTAL_ROW = HEADER_ROW + 1 + DATA_ROWS + 1
    ws.cell(row=TOTAL_ROW, column=2, value="TOTAL").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=2).fill = FILL_SUBHEADER
    for col_idx, col_letter in [(5, "E"), (7, "G"), (8, "H"), (9, "I")]:
        c = ws.cell(
            row=TOTAL_ROW,
            column=col_idx,
            value=f"=SUM({col_letter}{HEADER_ROW+1}:{col_letter}{HEADER_ROW+DATA_ROWS})",
        )
        c.font = FONT_BODY_BOLD
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_USD
        c.border = BORDER

    # Data validation: CSI Div dropdown
    div_codes = ",".join(code for code, _ in DIVISIONS)
    dv_div = DataValidation(
        type="list",
        formula1=f'"{div_codes}"',
        allow_blank=True,
    )
    dv_div.add(f"C{HEADER_ROW+1}:C{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_div)

    # Type dropdown
    dv_type = DataValidation(
        type="list",
        formula1='"Base Bid,Change Order,Allowance,Bond,Insurance,Other"',
        allow_blank=True,
    )
    dv_type.add(f"J{HEADER_ROW+1}:J{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_type)

    # Conditional formatting: Variance > 0 = over-billed (red)
    ws.conditional_formatting.add(
        f"I{HEADER_ROW+1}:I{HEADER_ROW+DATA_ROWS}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=FILL_RED, font=Font(color=RED_FONT)),
    )
    # Balance to Finish = 0 → green
    ws.conditional_formatting.add(
        f"H{HEADER_ROW+1}:H{HEADER_ROW+DATA_ROWS}",
        FormulaRule(
            formula=[f'AND(E{HEADER_ROW+1}<>"",H{HEADER_ROW+1}=0)'],
            fill=FILL_GREEN,
            font=Font(color=GREEN_FONT),
        ),
    )


# ---------------------------------------------------------------------------
# Tab 4: Pay App — Current Period (G-703 style)
# ---------------------------------------------------------------------------

PAY_APP_COLS = [
    ("A", "Line #", 8),
    ("B", "Description", 36),
    ("C", "CSI Div", 9),
    ("D", "Scheduled Value", 16),
    ("E", "Work Completed Previous ($)", 18),
    ("F", "% Complete This Period", 14),
    ("G", "Work Completed To Date ($)", 18),
    ("H", "Materials Stored ($)", 14),
    ("I", "Total To Date ($)", 16),
    ("J", "% Complete", 12),
    ("K", "Retainage ($)", 13),
    ("L", "Balance to Finish ($)", 16),
    ("M", "This Period Net ($)", 16),
]


def build_pay_app_current(ws):
    ws.title = "Pay App - Current"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in PAY_APP_COLS])

    ws["A1"] = "PAY APPLICATION — Current Period"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:M1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = "Project: =ProjectName"
    ws["A2"].font = FONT_BODY_BOLD
    ws.merge_cells("A2:F2")
    ws["G2"] = "Period: "
    ws["G2"].font = FONT_BODY_BOLD
    ws["H2"] = ""
    ws["I2"] = "Pay App #:"
    ws["I2"].font = FONT_BODY_BOLD
    ws["J2"] = 1
    ws["J2"].number_format = FMT_INT

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(PAY_APP_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(PAY_APP_COLS))

    DATA_ROWS = 50
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        sov_row = r - HEADER_ROW + 4  # match SOV Setup body row
        # Pull from SOV Setup
        ws.cell(row=r, column=1, value=f"=IF('SOV Setup'!A{sov_row}=\"\",\"\",'SOV Setup'!A{sov_row})")
        ws.cell(row=r, column=2, value=f"='SOV Setup'!B{sov_row}")
        ws.cell(row=r, column=3, value=f"='SOV Setup'!C{sov_row}")
        ws.cell(row=r, column=4, value=f"=IF('SOV Setup'!E{sov_row}=\"\",\"\",'SOV Setup'!E{sov_row})")
        ws.cell(row=r, column=4).number_format = FMT_USD
        # E = work completed previous (manual entry first period; carried-forward thereafter)
        ws.cell(row=r, column=5, value=0)
        ws.cell(row=r, column=5).number_format = FMT_USD
        # F = % complete this period (manual entry)
        ws.cell(row=r, column=6, value=0)
        ws.cell(row=r, column=6).number_format = FMT_PCT
        # G = work completed to date = previous + this period
        ws.cell(
            row=r,
            column=7,
            value=f"=IF(D{r}=\"\",\"\",E{r}+(D{r}*F{r}))",
        )
        ws.cell(row=r, column=7).number_format = FMT_USD
        # H = materials stored (manual; pulled from log)
        ws.cell(row=r, column=8, value=0)
        ws.cell(row=r, column=8).number_format = FMT_USD
        # I = total to date = G + H
        ws.cell(row=r, column=9, value=f"=IF(D{r}=\"\",\"\",G{r}+H{r})")
        ws.cell(row=r, column=9).number_format = FMT_USD
        # J = % complete
        ws.cell(row=r, column=10, value=f"=IF(OR(D{r}=\"\",D{r}=0),\"\",I{r}/D{r})")
        ws.cell(row=r, column=10).number_format = FMT_PCT
        # K = retainage = total-to-date * rate
        # If RetainStored == "No" then retain on G only; else on I
        ws.cell(
            row=r,
            column=11,
            value=f'=IF(D{r}="","",IF(RetainStored="Yes",I{r}*RetainageRate,G{r}*RetainageRate))',
        )
        ws.cell(row=r, column=11).number_format = FMT_USD
        # L = balance to finish = D - I
        ws.cell(row=r, column=12, value=f"=IF(D{r}=\"\",\"\",D{r}-I{r})")
        ws.cell(row=r, column=12).number_format = FMT_USD
        # M = this period net = (G - E + H) - retainage_delta — simplified to (I - E - prior_retainage)
        # For practical pay-app purposes, "this period net" = (current total - prior total) - (current retainage - prior retainage)
        # Prior retainage is approximated as E * RetainageRate (works for the steady-state case)
        ws.cell(
            row=r,
            column=13,
            value=(
                f'=IF(D{r}="","",'
                f'(I{r}-E{r})-(K{r}-E{r}*RetainageRate))'
            ),
        )
        ws.cell(row=r, column=13).number_format = FMT_USD

        apply_body_style(ws, r, len(PAY_APP_COLS))

    # Totals row
    TOTAL_ROW = HEADER_ROW + 1 + DATA_ROWS + 1
    ws.cell(row=TOTAL_ROW, column=2, value="GRAND TOTAL").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=2).fill = FILL_SUBHEADER
    for col_idx, col_letter in [(4, "D"), (5, "E"), (7, "G"), (8, "H"), (9, "I"), (11, "K"), (12, "L"), (13, "M")]:
        c = ws.cell(
            row=TOTAL_ROW,
            column=col_idx,
            value=f"=SUM({col_letter}{HEADER_ROW+1}:{col_letter}{HEADER_ROW+DATA_ROWS})",
        )
        c.font = FONT_BODY_BOLD
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_USD
        c.border = BORDER

    # Conditional format: % complete >= 100% → green
    ws.conditional_formatting.add(
        f"J{HEADER_ROW+1}:J{HEADER_ROW+DATA_ROWS}",
        CellIsRule(operator="greaterThanOrEqual", formula=["1"], fill=FILL_GREEN, font=Font(color=GREEN_FONT)),
    )
    # Balance to finish < 0 → red (over-billed)
    ws.conditional_formatting.add(
        f"L{HEADER_ROW+1}:L{HEADER_ROW+DATA_ROWS}",
        CellIsRule(operator="lessThan", formula=["0"], fill=FILL_RED, font=Font(color=RED_FONT)),
    )

    # Sub-application summary
    SUMMARY_ROW = TOTAL_ROW + 3
    ws.cell(row=SUMMARY_ROW, column=2, value="PAY APP SUMMARY").font = FONT_H2
    summary = [
        ("Total Subcontract Sum (incl. approved COs)", f"=SubContractSum"),
        ("Total Completed & Stored to Date", f"=I{TOTAL_ROW}"),
        ("Total Retainage Held", f"=K{TOTAL_ROW}"),
        ("Total Earned Less Retainage", f"=I{TOTAL_ROW}-K{TOTAL_ROW}"),
        ("Less Previous Certificates for Payment", f"=E{TOTAL_ROW}-E{TOTAL_ROW}*RetainageRate"),
        ("CURRENT PAYMENT DUE", f"=M{TOTAL_ROW}"),
        ("Balance to Finish (Including Retainage)", f"=SubContractSum-I{TOTAL_ROW}+K{TOTAL_ROW}"),
    ]
    for i, (label, val) in enumerate(summary, start=SUMMARY_ROW + 1):
        ws.cell(row=i, column=2, value=label).font = FONT_BODY_BOLD
        ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
        c = ws.cell(row=i, column=4, value=val)
        c.number_format = FMT_USD
        c.font = FONT_BIG_NUMBER if "CURRENT PAYMENT DUE" in label else FONT_BODY


# ---------------------------------------------------------------------------
# Tab 5: Pay App — History (12-month roll-up)
# ---------------------------------------------------------------------------

def build_pay_app_history(ws):
    ws.title = "Pay App - History"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "PAY APP HISTORY — 12-Month Roll-Up"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:O1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "After each pay app is paid by the GC, transfer the Total To Date values from Pay App — Current "
        "Period into the next column below."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:O2")

    HEADER_ROW = 4
    set_col_widths(ws, [
        ("A", 8), ("B", 36), ("C", 16),
        *[(get_column_letter(i), 14) for i in range(4, 16)],  # D..O = 12 months
        ("P", 16),
    ])

    headers = ["Line #", "Description", "Scheduled Value"]
    for m in range(1, 13):
        headers.append(f"Period {m}")
    headers.append("Final To Date")
    for i, h in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=h)
    style_header_row(ws, HEADER_ROW, len(headers))

    DATA_ROWS = 50
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        sov_row = r - HEADER_ROW + 4
        ws.cell(row=r, column=1, value=f"=IF('SOV Setup'!A{sov_row}=\"\",\"\",'SOV Setup'!A{sov_row})")
        ws.cell(row=r, column=2, value=f"='SOV Setup'!B{sov_row}")
        ws.cell(row=r, column=3, value=f"=IF('SOV Setup'!E{sov_row}=\"\",\"\",'SOV Setup'!E{sov_row})")
        ws.cell(row=r, column=3).number_format = FMT_USD
        for m in range(1, 13):
            col = 3 + m
            ws.cell(row=r, column=col, value="").number_format = FMT_USD
        # Final To Date = MAX of any populated period column (latest billed cumulative)
        ws.cell(
            row=r,
            column=16,
            value=f"=IF(C{r}=\"\",\"\",MAX(D{r}:O{r}))",
        )
        ws.cell(row=r, column=16).number_format = FMT_USD
        apply_body_style(ws, r, 16)

    # Totals row
    TOTAL_ROW = HEADER_ROW + 1 + DATA_ROWS + 1
    ws.cell(row=TOTAL_ROW, column=2, value="TOTAL").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=2).fill = FILL_SUBHEADER
    for col_idx in [3, *list(range(4, 17))]:
        col_letter = get_column_letter(col_idx)
        c = ws.cell(
            row=TOTAL_ROW,
            column=col_idx,
            value=f"=SUM({col_letter}{HEADER_ROW+1}:{col_letter}{HEADER_ROW+DATA_ROWS})",
        )
        c.font = FONT_BODY_BOLD
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_USD
        c.border = BORDER


# ---------------------------------------------------------------------------
# Tab 6: Stored Materials Log
# ---------------------------------------------------------------------------

STORED_COLS = [
    ("A", "Date Added", 12),
    ("B", "SOV Line Ref", 12),
    ("C", "Description", 30),
    ("D", "Supplier", 22),
    ("E", "PO #", 14),
    ("F", "Quantity", 10),
    ("G", "Unit", 8),
    ("H", "Unit Cost", 12),
    ("I", "Total Value ($)", 14),
    ("J", "Storage Location", 22),
    ("K", "Insured?", 10),
    ("L", "Lien Waiver from Supplier?", 14),
    ("M", "Date Installed", 12),
    ("N", "Status", 14),
]


def build_stored_materials(ws):
    ws.title = "Stored Materials"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in STORED_COLS])

    ws["A1"] = "STORED MATERIALS LOG"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:N1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Track materials billed-but-not-installed. The GC's pay-app reviewer will require evidence of off-site "
        "or on-site storage, dedicated segregation/labeling, insurance coverage, and a supplier lien waiver "
        "before approving stored-materials value."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:N2")
    ws.row_dimensions[2].height = 30

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(STORED_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(STORED_COLS))

    DATA_ROWS = 40
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        ws.cell(row=r, column=8).number_format = FMT_USD
        # Total value = qty * unit cost
        ws.cell(row=r, column=9, value=f"=IF(OR(F{r}=\"\",H{r}=\"\"),\"\",F{r}*H{r})")
        ws.cell(row=r, column=9).number_format = FMT_USD
        ws.cell(row=r, column=1).number_format = FMT_DATE
        ws.cell(row=r, column=13).number_format = FMT_DATE
        apply_body_style(ws, r, len(STORED_COLS))

    # Validations
    dv_yn = DataValidation(type="list", formula1='"Yes,No,Pending"', allow_blank=True)
    dv_yn.add(f"K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS}")
    dv_yn.add(f"L{HEADER_ROW+1}:L{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_yn)

    dv_status = DataValidation(
        type="list",
        formula1='"On Site,Off Site Bonded Warehouse,In Transit,Installed,Returned"',
        allow_blank=True,
    )
    dv_status.add(f"N{HEADER_ROW+1}:N{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_status)

    # Totals
    TOTAL_ROW = HEADER_ROW + 1 + DATA_ROWS + 1
    ws.cell(row=TOTAL_ROW, column=2, value="TOTAL").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=2).fill = FILL_SUBHEADER
    c = ws.cell(
        row=TOTAL_ROW,
        column=9,
        value=f"=SUM(I{HEADER_ROW+1}:I{HEADER_ROW+DATA_ROWS})",
    )
    c.font = FONT_BODY_BOLD
    c.fill = FILL_SUBHEADER
    c.number_format = FMT_USD
    c.border = BORDER


# ---------------------------------------------------------------------------
# Tab 7: CSI Reference (hidden)
# ---------------------------------------------------------------------------

def build_csi_reference(ws):
    ws.title = "CSI Reference"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 10), ("B", 60)])

    ws["A1"] = "CSI MasterFormat Divisions"
    ws["A1"].font = FONT_H2
    ws.merge_cells("A1:B1")

    ws["A3"] = "Div"
    ws["B3"] = "Name"
    for c in [ws["A3"], ws["B3"]]:
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    for i, (code, name) in enumerate(DIVISIONS, start=4):
        ws.cell(row=i, column=1, value=code).font = FONT_BODY
        ws.cell(row=i, column=2, value=name).font = FONT_BODY
        ws.cell(row=i, column=1).border = BORDER
        ws.cell(row=i, column=2).border = BORDER

    ws.sheet_state = "hidden"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    wb = Workbook()
    build_instructions(wb.active)
    build_project_info(wb.create_sheet())
    build_sov_setup(wb.create_sheet())
    build_pay_app_current(wb.create_sheet())
    build_pay_app_history(wb.create_sheet())
    build_stored_materials(wb.create_sheet())
    build_csi_reference(wb.create_sheet())

    out_dir = "/Users/home/charles/contrpro/files/packages/complete/sub"
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "Sub_Schedule_of_Values.xlsx")
    wb.save(out_path)
    sz = os.path.getsize(out_path)
    print(f"Wrote {out_path} ({sz//1024} KB)")


if __name__ == "__main__":
    main()
