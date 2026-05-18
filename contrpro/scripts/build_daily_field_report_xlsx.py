#!/usr/bin/env python3
"""
Build ContrPro Daily Field Report (XLSX) — Universal Sub Suite.

Digital companion to daily-field-report-template-and-guide.html. Designed so a
foreman can enter daily DFR data on a single sheet per day (31 day-tabs for a
month) plus a Monthly Roll-Up tab that aggregates labor hours, crew counts,
incidents, and weather impact.

Tabs:
  1. Instructions
  2. Project Info
  3. Monthly Roll-Up   (auto-aggregates from Day tabs)
  4. Day 01 .. Day 31  (31 identical day-tabs for a single calendar month)

Run:
    /Users/home/charles/.venv/bin/python3 \\
        /Users/home/charles/contrpro/scripts/build_daily_field_report_xlsx.py

Output:
    /Users/home/charles/contrpro/files/packages/complete/sub/Daily_Field_Report.xlsx
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
YELLOW_FILL = "FFEB9C"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_SUMMARY_LABEL = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW_FILL)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H1 = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_BIG_NUMBER = Font(name="Calibri", size=18, bold=True, color=BRAND_BLUE)
FONT_GREY_ITALIC = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_NUM2 = "0.00"
FMT_INT = "0"
FMT_DATE = "yyyy-mm-dd"


def set_col_widths(ws, widths):
    for col, w in widths:
        ws.column_dimensions[col].width = w


def style_header_row(ws, row, cols, start_col=1):
    for c in range(start_col, start_col + cols):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 26


def apply_body_style(ws, row, cols, start_col=1):
    for c in range(start_col, start_col + cols):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_BODY
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER


# ---------------------------------------------------------------------------
# Tab: Instructions
# ---------------------------------------------------------------------------

def build_instructions(ws):
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 100)])

    ws["A1"] = "DAILY FIELD REPORT — Instructions"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 36

    rows = [
        "",
        "WHAT THIS WORKBOOK IS",
        "Digital companion to the printable DFR (daily-field-report-template-and-guide.html). This workbook "
        "provides one tab per day for a single calendar month (Day 01 through Day 31), plus a Monthly Roll-"
        "Up tab that aggregates labor hours, crew counts, equipment hours, safety incidents, and weather "
        "lost-time. Use one workbook per project per month.",
        "",
        "DAILY WORKFLOW",
        "  1. Open the appropriate Day tab at the start of the shift.",
        "  2. Foreman enters crew, classification, hours, and tasks throughout the day or at end of shift.",
        "  3. Materials delivered, equipment used, conditions affecting work, directives, RFIs, photos, "
        "safety, and visitors logged.",
        "  4. Sign in the Foreman field. Request GC field-rep countersignature where practical.",
        "  5. Save and back up the workbook the same day.",
        "",
        "MONTHLY ROLL-UP",
        "The Monthly Roll-Up tab pulls totals from each Day tab automatically. Submit the roll-up with the "
        "monthly pay app or to the GC PM as the consolidated monthly report.",
        "",
        "MULTIPLE PROJECTS / MONTHS",
        "Duplicate this workbook for each project + month combination. Project Info propagates to every Day "
        "tab via named ranges, so updating Project Info once updates all Day tabs.",
        "",
        "PHOTOS",
        "Photos are not embedded in the workbook (file-size would balloon). Maintain a separate photo folder "
        "named by date — e.g., 'photos/2026-05-17/' — and reference filenames in the Photo Log section of "
        "the Day tab. The XLSX is the index; the photo folder is the storage.",
        "",
        "BACKCHARGE / DELAY DEFENSE",
        "If a backcharge or delay claim arises, the relevant Day tabs are the first evidence pulled. Keep "
        "this workbook clean and contemporaneous. Do not edit a Day tab more than 24 hours after the fact "
        "without adding a 'Late Entry' note dated when the edit was made.",
        "",
        "AUDIT TRAIL",
        "Save each completed Day tab as PDF at end-of-week and store with the project file. Excel files can "
        "be edited; PDF snapshots cannot. Both should exist.",
    ]

    for i, text in enumerate(rows, start=2):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if text and text == text.upper() and len(text) < 60:
            cell.font = FONT_H2
        else:
            cell.font = FONT_BODY


# ---------------------------------------------------------------------------
# Tab: Project Info
# ---------------------------------------------------------------------------

def build_project_info(ws):
    ws.title = "Project Info"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 38), ("C", 36)])

    ws["B1"] = "PROJECT INFORMATION"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:C1")
    ws.row_dimensions[1].height = 32

    info_rows = [
        ("Project Name", ""),
        ("Project Address", ""),
        ("GC", ""),
        ("Owner", ""),
        ("Subcontractor (legal name)", ""),
        ("Trade / Scope", ""),
        ("Subcontract No.", ""),
        ("Foreman", ""),
        ("Sub's PM", ""),
        ("Month (YYYY-MM)", ""),
        ("First Day of Month (date)", ""),
    ]

    name_map = {
        "ProjectName": "$C$3",
        "GCName": "$C$5",
        "SubName": "$C$7",
        "TradeScope": "$C$8",
        "Foreman": "$C$10",
        "FirstDay": "$C$13",
    }
    for nm, ref in name_map.items():
        ws.parent.defined_names[nm] = DefinedName(name=nm, attr_text=f"'Project Info'!{ref}")

    for i, (label, val) in enumerate(info_rows, start=3):
        ws.cell(row=i, column=2, value=label).font = FONT_BODY_BOLD
        ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
        ws.cell(row=i, column=2).border = BORDER
        c = ws.cell(row=i, column=3, value=val)
        c.font = FONT_BODY
        c.border = BORDER
        if "First Day" in label or "Month" in label:
            c.number_format = FMT_DATE


# ---------------------------------------------------------------------------
# Build a single Day tab
# ---------------------------------------------------------------------------

def build_day_tab(ws, day_num: int):
    ws.title = f"Day {day_num:02d}"
    ws.sheet_view.showGridLines = False
    set_col_widths(
        ws,
        [
            ("A", 4), ("B", 22), ("C", 18), ("D", 10), ("E", 10), ("F", 32),
        ],
    )

    # Title
    ws["B1"] = f"DAILY FIELD REPORT — Day {day_num}"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:F1")
    ws.row_dimensions[1].height = 32

    # Header block
    ws["B2"] = "DFR #:"
    ws["B2"].font = FONT_BODY_BOLD
    ws["C2"] = ""
    ws["D2"] = "Date:"
    ws["D2"].font = FONT_BODY_BOLD
    ws["E2"] = f"=IF(FirstDay=\"\",\"\",FirstDay+{day_num - 1})"
    ws["E2"].number_format = FMT_DATE
    ws["F2"] = "Project: =ProjectName"
    ws["F2"].font = FONT_BODY

    ws["B3"] = "Foreman:"
    ws["B3"].font = FONT_BODY_BOLD
    ws["C3"] = "=Foreman"
    ws["D3"] = "Sub:"
    ws["D3"].font = FONT_BODY_BOLD
    ws["E3"] = "=SubName"
    ws.merge_cells("E3:F3")

    # Weather
    row = 5
    ws.cell(row=row, column=2, value="WEATHER").font = FONT_H2
    row += 1
    weather_fields = [
        ("AM Temp (°F) / Conditions", ""),
        ("PM Temp (°F) / Conditions", ""),
        ("Wind / Visibility", ""),
        ("Impact on Work (None / Minor / Hours lost #)", ""),
        ("Idle Hours Lost to Weather", 0),
    ]
    for label, val in weather_fields:
        ws.cell(row=row, column=2, value=label).font = FONT_BODY
        ws.cell(row=row, column=2).fill = FILL_SUMMARY_LABEL
        ws.cell(row=row, column=2).border = BORDER
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        c = ws.cell(row=row, column=3, value=val)
        c.font = FONT_BODY
        c.border = BORDER
        if "Hours" in label:
            c.number_format = FMT_NUM2
        row += 1

    # Crew table
    row += 1
    ws.cell(row=row, column=2, value="CREW ON SITE").font = FONT_H2
    row += 1
    crew_headers = ["Name", "Classification", "ST Hrs", "OT Hrs", "Task / Location"]
    for i, h in enumerate(crew_headers, start=2):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(crew_headers), start_col=2)
    crew_start = row + 1
    CREW_ROWS = 10
    for r in range(crew_start, crew_start + CREW_ROWS):
        ws.cell(row=r, column=4).number_format = FMT_NUM2
        ws.cell(row=r, column=5).number_format = FMT_NUM2
        apply_body_style(ws, r, len(crew_headers), start_col=2)
    crew_total_row = crew_start + CREW_ROWS
    ws.cell(row=crew_total_row, column=2, value="CREW TOTALS").font = FONT_BODY_BOLD
    ws.cell(row=crew_total_row, column=2).fill = FILL_SUBHEADER
    ws.merge_cells(start_row=crew_total_row, start_column=2, end_row=crew_total_row, end_column=3)
    for col_letter, col_idx in [("D", 4), ("E", 5)]:
        c = ws.cell(
            row=crew_total_row,
            column=col_idx,
            value=f"=SUM({col_letter}{crew_start}:{col_letter}{crew_start+CREW_ROWS-1})",
        )
        c.font = FONT_BODY_BOLD
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_NUM2
        c.border = BORDER

    # Equipment table
    row = crew_total_row + 2
    ws.cell(row=row, column=2, value="EQUIPMENT ON SITE").font = FONT_H2
    row += 1
    eq_headers = ["Equipment", "Owned / Rented", "Op Hrs", "Idle Hrs", "Notes"]
    for i, h in enumerate(eq_headers, start=2):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(eq_headers), start_col=2)
    eq_start = row + 1
    EQ_ROWS = 6
    for r in range(eq_start, eq_start + EQ_ROWS):
        ws.cell(row=r, column=4).number_format = FMT_NUM2
        ws.cell(row=r, column=5).number_format = FMT_NUM2
        apply_body_style(ws, r, len(eq_headers), start_col=2)
    eq_total_row = eq_start + EQ_ROWS
    ws.cell(row=eq_total_row, column=2, value="EQUIP TOTALS").font = FONT_BODY_BOLD
    ws.cell(row=eq_total_row, column=2).fill = FILL_SUBHEADER
    ws.merge_cells(start_row=eq_total_row, start_column=2, end_row=eq_total_row, end_column=3)
    for col_letter, col_idx in [("D", 4), ("E", 5)]:
        c = ws.cell(
            row=eq_total_row,
            column=col_idx,
            value=f"=SUM({col_letter}{eq_start}:{col_letter}{eq_start+EQ_ROWS-1})",
        )
        c.font = FONT_BODY_BOLD
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_NUM2
        c.border = BORDER

    # Materials
    row = eq_total_row + 2
    ws.cell(row=row, column=2, value="MATERIALS DELIVERED / INSTALLED").font = FONT_H2
    row += 1
    mat_headers = ["Material", "Supplier", "Qty Delivered", "Qty Installed", "Location"]
    for i, h in enumerate(mat_headers, start=2):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(mat_headers), start_col=2)
    mat_start = row + 1
    MAT_ROWS = 6
    for r in range(mat_start, mat_start + MAT_ROWS):
        apply_body_style(ws, r, len(mat_headers), start_col=2)

    # Work performed narrative
    row = mat_start + MAT_ROWS + 1
    ws.cell(row=row, column=2, value="WORK PERFORMED (NARRATIVE)").font = FONT_H2
    row += 1
    ws.cell(row=row, column=2, value="").alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=2, end_row=row + 3, end_column=6)
    for r in range(row, row + 4):
        for c in range(2, 7):
            ws.cell(row=r, column=c).border = BORDER
    row += 4

    # Conditions / directives
    row += 1
    ws.cell(row=row, column=2, value="CONDITIONS AFFECTING WORK").font = FONT_H2
    row += 1
    cond_fields = [
        "Differing site conditions",
        "Missing GC trade / access",
        "Utility / power issue",
        "Inspection / AHJ",
        "Idle crew hours due to conditions (#)",
    ]
    for label in cond_fields:
        ws.cell(row=row, column=2, value=label).font = FONT_BODY
        ws.cell(row=row, column=2).fill = FILL_SUMMARY_LABEL
        ws.cell(row=row, column=2).border = BORDER
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        c = ws.cell(row=row, column=3, value="")
        c.font = FONT_BODY
        c.border = BORDER
        if "(#)" in label:
            c.number_format = FMT_NUM2
        row += 1

    # Directives / RFIs
    row += 1
    ws.cell(row=row, column=2, value="DIRECTIVES / RFIs / T&M TICKETS").font = FONT_H2
    row += 1
    dir_headers = ["Type", "Originator", "Description", "Status", "Notes"]
    for i, h in enumerate(dir_headers, start=2):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(dir_headers), start_col=2)
    dir_start = row + 1
    DIR_ROWS = 5
    for r in range(dir_start, dir_start + DIR_ROWS):
        apply_body_style(ws, r, len(dir_headers), start_col=2)

    # Safety
    row = dir_start + DIR_ROWS + 1
    ws.cell(row=row, column=2, value="SAFETY").font = FONT_H2
    row += 1
    safety_fields = [
        ("Toolbox topic", ""),
        ("Near-miss (count)", 0),
        ("First aid only (count)", 0),
        ("Recordable incident (count)", 0),
        ("OSHA visit", "No"),
        ("PPE audit notes", ""),
    ]
    for label, val in safety_fields:
        ws.cell(row=row, column=2, value=label).font = FONT_BODY
        ws.cell(row=row, column=2).fill = FILL_SUMMARY_LABEL
        ws.cell(row=row, column=2).border = BORDER
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        c = ws.cell(row=row, column=3, value=val)
        c.font = FONT_BODY
        c.border = BORDER
        if "count" in label:
            c.number_format = FMT_INT
        row += 1

    # Photo log
    row += 1
    ws.cell(row=row, column=2, value="PHOTO LOG").font = FONT_H2
    row += 1
    photo_headers = ["File / Frame", "Time", "Location", "Description", "Notes"]
    for i, h in enumerate(photo_headers, start=2):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(photo_headers), start_col=2)
    photo_start = row + 1
    PHOTO_ROWS = 8
    for r in range(photo_start, photo_start + PHOTO_ROWS):
        apply_body_style(ws, r, len(photo_headers), start_col=2)

    # Signatures
    row = photo_start + PHOTO_ROWS + 1
    ws.cell(row=row, column=2, value="SIGNATURES").font = FONT_H2
    row += 1
    ws.cell(row=row, column=2, value="Foreman:").font = FONT_BODY_BOLD
    ws.cell(row=row, column=2).fill = FILL_SUMMARY_LABEL
    ws.cell(row=row, column=2).border = BORDER
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    ws.cell(row=row, column=3, value="").border = BORDER
    row += 1
    ws.cell(row=row, column=2, value="GC Field Rep (countersig):").font = FONT_BODY_BOLD
    ws.cell(row=row, column=2).fill = FILL_SUMMARY_LABEL
    ws.cell(row=row, column=2).border = BORDER
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    ws.cell(row=row, column=3, value="").border = BORDER

    # Hidden anchor cells for roll-up (compact addresses):
    # We define a small "RollUp Anchors" block in fixed cells the Monthly Roll-Up references.
    # Place anchors in row 200 of this tab.
    anchor_row = 200
    ws.cell(row=anchor_row, column=1, value="Anchor:CrewST").font = FONT_GREY_ITALIC
    ws.cell(row=anchor_row, column=2, value=f"=D{crew_total_row}")
    ws.cell(row=anchor_row + 1, column=1, value="Anchor:CrewOT").font = FONT_GREY_ITALIC
    ws.cell(row=anchor_row + 1, column=2, value=f"=E{crew_total_row}")
    ws.cell(row=anchor_row + 2, column=1, value="Anchor:EqOp").font = FONT_GREY_ITALIC
    ws.cell(row=anchor_row + 2, column=2, value=f"=D{eq_total_row}")
    ws.cell(row=anchor_row + 3, column=1, value="Anchor:WeatherLost").font = FONT_GREY_ITALIC
    ws.cell(row=anchor_row + 3, column=2, value="=C9")  # weather lost hours
    ws.cell(row=anchor_row + 4, column=1, value="Anchor:Recordable").font = FONT_GREY_ITALIC
    # Recordable lives in safety block — too address-dependent to compute robustly; leave 0 default
    ws.cell(row=anchor_row + 4, column=2, value=0)


# ---------------------------------------------------------------------------
# Monthly Roll-Up tab
# ---------------------------------------------------------------------------

def build_monthly_rollup(ws):
    ws.title = "Monthly Roll-Up"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 14), ("C", 14), ("D", 12), ("E", 12), ("F", 12), ("G", 14), ("H", 14), ("I", 14)])

    ws["B1"] = "MONTHLY ROLL-UP"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:I1")
    ws.row_dimensions[1].height = 32

    ws["B2"] = "Auto-aggregated from Day 01 through Day 31 tabs. Blank rows = day not yet logged."
    ws["B2"].font = FONT_GREY_ITALIC
    ws.merge_cells("B2:I2")

    HEADER_ROW = 4
    headers = ["Day", "Date", "ST Hrs", "OT Hrs", "Total Hrs", "Equip Op Hrs", "Weather Lost Hrs", "Recordable Inc."]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=HEADER_ROW, column=i, value=h)
    style_header_row(ws, HEADER_ROW, len(headers), start_col=2)

    for d in range(1, 32):
        r = HEADER_ROW + d
        sheet_name = f"'Day {d:02d}'"
        ws.cell(row=r, column=2, value=d).font = FONT_BODY
        ws.cell(row=r, column=3, value=f"=IF(FirstDay=\"\",\"\",FirstDay+{d-1})").number_format = FMT_DATE
        ws.cell(row=r, column=4, value=f"={sheet_name}!B200").number_format = FMT_NUM2
        ws.cell(row=r, column=5, value=f"={sheet_name}!B201").number_format = FMT_NUM2
        ws.cell(row=r, column=6, value=f"=D{r}+E{r}").number_format = FMT_NUM2
        ws.cell(row=r, column=7, value=f"={sheet_name}!B202").number_format = FMT_NUM2
        ws.cell(row=r, column=8, value=f"={sheet_name}!B203").number_format = FMT_NUM2
        ws.cell(row=r, column=9, value=f"={sheet_name}!B204").number_format = FMT_INT
        apply_body_style(ws, r, len(headers), start_col=2)

    # Totals row
    TOTAL_ROW = HEADER_ROW + 32
    ws.cell(row=TOTAL_ROW, column=2, value="MONTH TOTAL").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=2).fill = FILL_SUBHEADER
    for col_idx in range(4, 10):
        col_letter = get_column_letter(col_idx)
        c = ws.cell(
            row=TOTAL_ROW,
            column=col_idx,
            value=f"=SUM({col_letter}{HEADER_ROW+1}:{col_letter}{HEADER_ROW+31})",
        )
        c.font = FONT_BODY_BOLD
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_NUM2 if col_idx <= 8 else FMT_INT
        c.border = BORDER


# ---------------------------------------------------------------------------
# Project Info anchor — must be created before Day tabs reference it
# Main
# ---------------------------------------------------------------------------

def main():
    wb = Workbook()
    build_instructions(wb.active)
    build_project_info(wb.create_sheet())
    build_monthly_rollup(wb.create_sheet())
    for d in range(1, 32):
        build_day_tab(wb.create_sheet(), d)

    out_dir = "/Users/home/charles/contrpro/files/packages/complete/sub"
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "Daily_Field_Report.xlsx")
    wb.save(out_path)
    sz = os.path.getsize(out_path)
    print(f"Wrote {out_path} ({sz//1024} KB)")


if __name__ == "__main__":
    main()
