#!/usr/bin/env python3
"""
Mechanical Bid Estimator (XLSX) — Mechanical trade pack.

Industry-standard mechanical bid estimator. CSI Division 23 primary with
cross-references to 22 (plumbing) + 26 (electrical) + 27 (BAS/controls).
Labor anchored to MCAA/SMACNA-style productivity units (per-CFM,
per-ton, per-LF by diameter, per-SF of sheet metal).

Tabs:
  1. Instructions
  2. Project Info
  3. Wage Rates (sheet metal / pipefitter / steamfitter / refrigeration)
  4. Material Takeoff (CSI 23 — equipment, ductwork, piping, controls)
  5. Labor Estimate
  6. Equipment Estimate
  7. Subcontract Estimate
  8. Premiums & Conditions
  9. Bid Summary
  10. CSI Reference

Output:
    /Users/home/charles/contrpro/files/packages/complete/mechanical/Mechanical_Bid_Estimator.xlsx
"""
from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BRAND_BLUE = "1E3A5F"; BRAND_BLUE_LIGHT = "D6E0EC"; ACCENT_GOLD = "C9A227"; GREY_TEXT = "808080"
FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GOLD = PatternFill("solid", fgColor=ACCENT_GOLD)
FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_GREY_IT = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)
THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_HRS = '_(* #,##0.0_);_(* (#,##0.0);_(* "-"_);_(@_)'
FMT_PCT = "0.0%"; FMT_INT = "0"; FMT_NUM2 = "0.00"
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

OUT = "/Users/home/charles/contrpro/files/packages/complete/mechanical/Mechanical_Bid_Estimator.xlsx"


def widths(ws, cols):
    for col, w in cols: ws.column_dimensions[col].width = w


def title(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text); c.font = FONT_TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def header(ws, row, cols):
    for i, h in enumerate(cols):
        c = ws.cell(row=row, column=1 + i, value=h)
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[row].height = 26


def build_instructions(ws):
    ws.title = "Instructions"
    widths(ws, [("A", 110)])
    ws["A1"] = "MECHANICAL BID ESTIMATOR — INSTRUCTIONS"
    ws["A1"].font = FONT_TITLE
    body = [
        "",
        "PURPOSE",
        "Production estimating tool for commercial mechanical (HVAC + hydronic + refrigeration)",
        "contractors. Combines per-equipment + per-ton + per-CFM + per-LF of pipe / per-SF of",
        "sheet-metal anchors. CSI Division 23 primary with cross-references to 22 (plumbing),",
        "26 (electrical), and 27 (BAS / controls).",
        "",
        "WORKFLOW",
        "  1. Project Info — fill every blue field.",
        "  2. Wage Rates — set base + fringe + burden by classification.",
        "  3. Material Takeoff — quantities by CSI line item.",
        "  4. Labor Estimate — productivity defaults assume average commercial conditions.",
        "  5. Equipment — daily/weekly rentals + tools.",
        "  6. Subcontract — TAB, refrigerant charging, controls programming, etc.",
        "  7. Premiums — height, healthcare, occupied building, prevailing wage.",
        "  8. Bid Summary — automatic roll-up + final bid.",
        "",
        "LABOR UNITS",
        "Anchors used: per-ton or per-CFM for equipment-set, per-LF by diameter for piping,",
        "per-SF for sheet-metal (gauge-adjusted), per-each for terminals + accessories.",
        "Calibrate against your historical job-cost. Underestimating labor is the most common",
        "cause of mechanical-bid losses.",
        "",
        "FIELD-ONLY SCOPE",
        "Calibrated for field installation. Sheet-metal fab-shop production work is typically",
        "in-house or subbed and not estimated here — material costs reflect delivered fabricated",
        "ductwork. Pre-fabricated piping rack assemblies similarly.",
        "",
        "DOCUMENT VERSION",
        "Mechanical_Bid_Estimator.xlsx v1.0 — 2026-05-19",
        "ContrPro Complete Tier · Mechanical Trade Pack",
    ]
    for i, line in enumerate(body, start=3):
        c = ws.cell(row=i, column=1, value=line)
        c.font = FONT_H2 if (line.isupper() and line.strip() and not line.startswith(" ")) else FONT_BODY
        c.alignment = LEFT


def build_project_info(ws):
    ws.title = "Project Info"
    widths(ws, [("A", 32), ("B", 60)])
    title(ws, 1, "PROJECT INFORMATION", 2)
    fields = [
        ("Project Name", ""),
        ("Project Address", ""),
        ("Owner", ""),
        ("General Contractor", ""),
        ("Architect / Engineer", ""),
        ("Bid Due Date", ""),
        ("Bid Reference / Number", ""),
        ("Project Type", "(New / Renovation / Tenant fit-out / Adaptive reuse)"),
        ("Building Use", "(Office / Retail / Healthcare / Education / Multi-family / Hospitality / Industrial)"),
        ("Approx. Square Footage", ""),
        ("Stories", ""),
        ("Total Tonnage (cooling)", ""),
        ("Total CFM (supply / return / OA)", ""),
        ("AHJ — Mechanical Inspector", ""),
        ("Applicable Mechanical Code", "(IMC 2024 / UMC 2024 / local amendments)"),
        ("Applicable Energy Code", "(IECC 2021 / ASHRAE 90.1-2019 / state amendments)"),
        ("Estimator Name", ""),
        ("Date Prepared", ""),
        ("Internal Bid Number", ""),
        ("Notes / Assumptions", ""),
    ]
    row = 3
    for label, default in fields:
        ws.cell(row=row, column=1, value=label).font = FONT_BODY_BOLD
        ws.cell(row=row, column=1).fill = FILL_SUBHEADER
        ws.cell(row=row, column=1).alignment = LEFT; ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=2, value=default).font = FONT_GREY_IT if default else FONT_BODY
        ws.cell(row=row, column=2).alignment = LEFT; ws.cell(row=row, column=2).border = BORDER
        row += 1


def build_wages(ws):
    ws.title = "Wage Rates"
    widths(ws, [("A", 28), ("B", 14), ("C", 14), ("D", 14), ("E", 16), ("F", 16)])
    title(ws, 1, "WAGE RATES (calibrate to your market)", 6)
    header(ws, 3, ["Classification", "Base Rate ($/hr)", "Fringe ($/hr)", "Burden % (WC/GL/FICA)", "OT Premium (1.5×)", "Loaded Rate ($/hr)"])
    rows = [
        ("Mechanical Foreman", 62.00, 19.50, 0.32),
        ("Sheet Metal Journeyman", 50.00, 18.50, 0.32),
        ("Pipefitter Journeyman", 54.00, 19.50, 0.32),
        ("Steamfitter Journeyman", 56.00, 19.50, 0.32),
        ("Refrigeration Tech (EPA 608)", 58.00, 19.50, 0.32),
        ("HVAC Tech", 50.00, 18.50, 0.32),
        ("Apprentice (Yr 4)", 38.00, 14.00, 0.32),
        ("Apprentice (Yr 2)", 28.00, 12.00, 0.32),
        ("Insulator", 42.00, 16.00, 0.32),
        ("Helper / Material Handler", 22.00, 9.50, 0.32),
        ("Controls Tech / BAS", 60.00, 19.00, 0.32),
        ("TAB Technician", 55.00, 18.00, 0.32),
    ]
    for i, (cls, base, fringe, burden) in enumerate(rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=cls).font = FONT_BODY_BOLD
        ws.cell(row=r, column=2, value=base).number_format = FMT_USD
        ws.cell(row=r, column=3, value=fringe).number_format = FMT_USD
        ws.cell(row=r, column=4, value=burden).number_format = FMT_PCT
        ws.cell(row=r, column=5, value=f"=B{r}*1.5").number_format = FMT_USD
        ws.cell(row=r, column=6, value=f"=(B{r}+C{r})*(1+D{r})").number_format = FMT_USD
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).alignment = CENTER if col > 1 else LEFT


def build_material(ws):
    ws.title = "Material Takeoff"
    widths(ws, [("A", 14), ("B", 50), ("C", 10), ("D", 10), ("E", 14), ("F", 16), ("G", 14)])
    title(ws, 1, "MATERIAL TAKEOFF (CSI Division 23)", 7)
    header(ws, 3, ["CSI Code", "Description", "QTY", "UOM", "Unit Cost ($)", "Material Subtotal ($)", "Notes"])
    items = [
        # 23 05 — Common work
        ("23 05 13", "Pipe hangers + supports — clevis 1\" - 4\"", 0, "EA", 9.50, ""),
        ("23 05 13", "Trapeze + unistrut assembly (multi-pipe)", 0, "EA", 38.00, ""),
        ("23 05 17", "Sleeves + escutcheons (per assembly)", 0, "EA", 6.50, ""),
        ("23 05 23", "Valve — ball, bronze, 1\"", 0, "EA", 28.00, ""),
        ("23 05 23", "Valve — ball, bronze, 2\"", 0, "EA", 65.00, ""),
        ("23 05 23", "Valve — globe, bronze, 1\"", 0, "EA", 65.00, ""),
        ("23 05 23", "Valve — butterfly, cast iron, 4\"", 0, "EA", 185.00, ""),
        ("23 05 23", "Valve — gate, cast iron, 6\"", 0, "EA", 425.00, ""),
        # 23 07 — Insulation
        ("23 07 13", "Duct insulation — fiberglass wrap, 1.5\" w/ FRK", 0, "SF", 1.85, ""),
        ("23 07 13", "Duct insulation — fiberglass duct liner, 1\"", 0, "SF", 2.65, ""),
        ("23 07 19", "Pipe insulation — fiberglass, 1\" pipe", 0, "LF", 2.85, ""),
        ("23 07 19", "Pipe insulation — fiberglass, 2\" pipe", 0, "LF", 4.20, ""),
        ("23 07 19", "Pipe insulation — fiberglass, 4\" pipe", 0, "LF", 7.85, ""),
        ("23 07 19", "Pipe insulation — Armaflex, 1\" pipe", 0, "LF", 3.85, ""),
        # 23 09 — Controls / BAS
        ("23 09 13", "BAS controller — VAV box", 0, "EA", 485.00, ""),
        ("23 09 13", "BAS controller — AHU controller", 0, "EA", 1450.00, ""),
        ("23 09 23", "Direct Digital Control (DDC) panel — 32-pt", 0, "EA", 3850.00, ""),
        ("23 09 33", "Thermostat — programmable", 0, "EA", 145.00, ""),
        ("23 09 33", "Thermostat — networked, BAS", 0, "EA", 385.00, ""),
        ("23 09 33", "Sensor — temperature, duct-mount", 0, "EA", 65.00, ""),
        ("23 09 33", "Sensor — pressure, differential", 0, "EA", 185.00, ""),
        ("23 09 33", "Sensor — CO2 / IAQ", 0, "EA", 285.00, ""),
        # 23 11 — Fuel
        ("23 11 13", "Fuel oil tank — 275 gal indoor", 0, "EA", 1850.00, ""),
        ("23 11 23", "Natural gas piping — black iron 1\"", 0, "LF", 8.50, ""),
        ("23 11 23", "Natural gas piping — black iron 2\"", 0, "LF", 16.50, ""),
        # 23 21 — Hydronic piping
        ("23 21 13", "Hydronic — copper Type L, 3/4\"", 0, "LF", 6.85, ""),
        ("23 21 13", "Hydronic — copper Type L, 1\"", 0, "LF", 9.40, ""),
        ("23 21 13", "Hydronic — copper Type L, 2\"", 0, "LF", 18.50, ""),
        ("23 21 13", "Hydronic — black steel, 2\"", 0, "LF", 8.50, ""),
        ("23 21 13", "Hydronic — black steel, 4\"", 0, "LF", 18.85, ""),
        ("23 21 13", "Hydronic — PEX-A w/ oxygen barrier, 1\"", 0, "LF", 1.85, ""),
        ("23 21 16", "Hydronic — pre-insulated underground", 0, "LF", 38.00, ""),
        ("23 21 23", "Circulator pump — 1/12 HP inline", 0, "EA", 285.00, ""),
        ("23 21 23", "Circulator pump — 1 HP base-mounted", 0, "EA", 1450.00, ""),
        ("23 21 23", "End-suction pump — 5 HP", 0, "EA", 3850.00, ""),
        ("23 21 29", "Expansion tank — 30 gal", 0, "EA", 485.00, ""),
        ("23 21 29", "Air separator — 4\"", 0, "EA", 685.00, ""),
        # 23 22 — Steam (if applicable)
        ("23 22 13", "Steam piping — black steel, 2\"", 0, "LF", 10.85, ""),
        ("23 22 16", "Steam trap — F&T 3/4\"", 0, "EA", 285.00, ""),
        # 23 25 — Water treatment
        ("23 25 13", "Glycol fill (per gallon)", 0, "GAL", 18.50, ""),
        ("23 25 13", "Chemical pot feeder", 0, "EA", 485.00, ""),
        # 23 31 — Ductwork
        ("23 31 13", "Galvanized rectangular duct — 24ga", 0, "SF", 8.50, ""),
        ("23 31 13", "Galvanized rectangular duct — 22ga", 0, "SF", 9.85, ""),
        ("23 31 13", "Galvanized rectangular duct — 20ga", 0, "SF", 11.50, ""),
        ("23 31 13", "Galvanized round duct — spiral, 24ga", 0, "SF", 7.50, ""),
        ("23 31 13", "Galvanized round duct — spiral, 22ga", 0, "SF", 8.85, ""),
        ("23 31 16", "Stainless steel duct — 22ga (kitchen exhaust)", 0, "SF", 28.50, ""),
        ("23 31 19", "Flex duct — insulated, 6\"", 0, "LF", 4.85, ""),
        ("23 31 19", "Flex duct — insulated, 12\"", 0, "LF", 9.85, ""),
        # 23 33 — Air-distribution accessories
        ("23 33 13", "Volume damper — manual, 12\"", 0, "EA", 85.00, ""),
        ("23 33 13", "Motorized damper — 24V actuator, 12\"", 0, "EA", 385.00, ""),
        ("23 33 33", "Fire damper — 1.5-hr UL listed, 12\"", 0, "EA", 285.00, ""),
        ("23 33 33", "Smoke damper — UL 555S, 12\"", 0, "EA", 485.00, ""),
        ("23 33 33", "Combination fire / smoke damper, 12\"", 0, "EA", 685.00, ""),
        ("23 33 46", "Flexible duct connector", 0, "LF", 14.50, ""),
        # 23 34 — Fans
        ("23 34 13", "Exhaust fan — inline 250 CFM", 0, "EA", 485.00, ""),
        ("23 34 13", "Exhaust fan — roof-mount 500 CFM", 0, "EA", 1485.00, ""),
        ("23 34 13", "Exhaust fan — kitchen hood 4000 CFM UL 762", 0, "EA", 4850.00, ""),
        ("23 34 16", "Supply fan — inline 800 CFM", 0, "EA", 685.00, ""),
        # 23 36 — Terminal units
        ("23 36 16", "VAV box — fan-powered 600 CFM", 0, "EA", 1685.00, ""),
        ("23 36 16", "VAV box — single-duct 800 CFM", 0, "EA", 1185.00, ""),
        ("23 36 16", "VAV box — single-duct 1200 CFM", 0, "EA", 1485.00, ""),
        # 23 37 — Air outlets/inlets
        ("23 37 13", "Supply diffuser — 4-way, 24x24", 0, "EA", 38.50, ""),
        ("23 37 13", "Return grille — 24x24", 0, "EA", 22.50, ""),
        ("23 37 13", "Exhaust grille — 12x12", 0, "EA", 14.50, ""),
        ("23 37 13", "Linear slot diffuser — 6 ft", 0, "EA", 285.00, ""),
        # 23 41 — Particulate air filtration
        ("23 41 13", "MERV 8 filter — 20x25x2", 0, "EA", 9.50, ""),
        ("23 41 13", "MERV 13 filter — 20x25x4", 0, "EA", 38.50, ""),
        ("23 41 13", "HEPA filter — 24x24x12 (hospital)", 0, "EA", 285.00, ""),
        # 23 51 — Breechings + Chimneys
        ("23 51 13", "B-vent — single wall 6\"", 0, "LF", 18.50, ""),
        ("23 51 13", "Stainless flue — 6\"", 0, "LF", 65.00, ""),
        # 23 52 — Heating boilers
        ("23 52 13", "Gas-fired boiler — 500 MBH atmospheric", 0, "EA", 5850.00, ""),
        ("23 52 16", "Gas-fired boiler — 1500 MBH condensing", 0, "EA", 18500.00, ""),
        ("23 52 39", "Electric boiler — 30 kW", 0, "EA", 6500.00, ""),
        # 23 54 — Furnaces
        ("23 54 13", "Furnace — gas, 80% AFUE 80 MBH", 0, "EA", 1850.00, ""),
        ("23 54 16", "Furnace — gas, 96% AFUE 80 MBH", 0, "EA", 2850.00, ""),
        # 23 57 — Heat exchangers
        ("23 57 13", "Plate-and-frame HX, 100 GPM", 0, "EA", 4850.00, ""),
        # 23 62 — Refrigerant
        ("23 62 13", "DX split — 3 ton outdoor unit", 0, "EA", 2850.00, ""),
        ("23 62 13", "DX split — 5 ton outdoor unit", 0, "EA", 4250.00, ""),
        ("23 62 13", "VRF outdoor — 10 ton", 0, "EA", 18500.00, ""),
        ("23 62 13", "Mini-split outdoor — 18 MBH", 0, "EA", 1685.00, ""),
        ("23 62 23", "Refrigerant piping — Type ACR copper 5/8\"", 0, "LF", 4.85, ""),
        ("23 62 23", "Refrigerant piping — Type ACR copper 7/8\"", 0, "LF", 7.85, ""),
        # 23 64 — Packaged chiller
        ("23 64 16", "Air-cooled chiller — 60 ton scroll", 0, "EA", 38500.00, ""),
        ("23 64 19", "Water-cooled chiller — 200 ton centrifugal", 0, "EA", 125000.00, ""),
        # 23 65 — Cooling tower
        ("23 65 13", "Cooling tower — 200 ton crossflow", 0, "EA", 38500.00, ""),
        # 23 72 — Air-to-air heat recovery
        ("23 72 13", "ERV — 1000 CFM packaged", 0, "EA", 6850.00, ""),
        # 23 73 — Indoor central-station air handler
        ("23 73 13", "Indoor AHU — 3000 CFM packaged", 0, "EA", 9500.00, ""),
        ("23 73 13", "Indoor AHU — 8000 CFM custom-built-up", 0, "EA", 24500.00, ""),
        ("23 73 23", "Rooftop unit — 5 ton packaged", 0, "EA", 5850.00, ""),
        ("23 73 23", "Rooftop unit — 15 ton packaged", 0, "EA", 18500.00, ""),
        ("23 73 23", "Rooftop unit — 25 ton packaged", 0, "EA", 32500.00, ""),
        # 23 81 — Unitary HVAC
        ("23 81 13", "Fan coil unit — vertical, 800 CFM", 0, "EA", 1485.00, ""),
        ("23 81 23", "Ductless mini-split — 18 MBH indoor head", 0, "EA", 485.00, ""),
        ("23 81 26", "Heat pump — 4 ton split outdoor unit", 0, "EA", 4850.00, ""),
        # 23 83 — Radiant
        ("23 83 13", "Radiant floor tubing — 1/2\" PEX", 0, "LF", 1.05, ""),
        ("23 83 13", "Radiant manifold — 6 loop", 0, "EA", 685.00, ""),
        # Permit / inspection
        ("23 00 10", "Permit — mechanical", 0, "LS", 0.00, "Allowance"),
    ]
    for i, (csi, desc, qty, uom, unit, notes) in enumerate(items):
        r = 4 + i
        ws.cell(row=r, column=1, value=csi).alignment = CENTER
        ws.cell(row=r, column=2, value=desc).alignment = LEFT
        ws.cell(row=r, column=3, value=qty).number_format = FMT_INT
        ws.cell(row=r, column=4, value=uom).alignment = CENTER
        ws.cell(row=r, column=5, value=unit).number_format = FMT_USD
        ws.cell(row=r, column=6, value=f"=C{r}*E{r}").number_format = FMT_USD
        ws.cell(row=r, column=7, value=notes).alignment = LEFT
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = FONT_BODY
    last = 4 + len(items) - 1
    sub_row = last + 2
    ws.cell(row=sub_row, column=5, value="Material Subtotal").font = FONT_BODY_BOLD
    ws.cell(row=sub_row, column=5).alignment = RIGHT
    ws.cell(row=sub_row, column=6, value=f"=SUM(F4:F{last})").number_format = FMT_USD
    ws.cell(row=sub_row, column=6).fill = FILL_SUBHEADER
    ws.cell(row=sub_row, column=6).font = FONT_BODY_BOLD
    tax_row = sub_row + 1
    ws.cell(row=tax_row, column=5, value="Sales tax %").alignment = RIGHT
    ws.cell(row=tax_row, column=5).font = FONT_BODY_BOLD
    ws.cell(row=tax_row, column=6, value=0.07).number_format = FMT_PCT
    tax_amt_row = tax_row + 1
    ws.cell(row=tax_amt_row, column=5, value="Sales tax $").alignment = RIGHT
    ws.cell(row=tax_amt_row, column=5).font = FONT_BODY_BOLD
    ws.cell(row=tax_amt_row, column=6, value=f"=F{sub_row}*F{tax_row}").number_format = FMT_USD
    tot_row = tax_amt_row + 1
    ws.cell(row=tot_row, column=5, value="Material TOTAL (incl tax)").font = FONT_BODY_BOLD
    ws.cell(row=tot_row, column=5).alignment = RIGHT
    ws.cell(row=tot_row, column=5).fill = FILL_GOLD
    ws.cell(row=tot_row, column=6, value=f"=F{sub_row}+F{tax_amt_row}").number_format = FMT_USD
    ws.cell(row=tot_row, column=6).fill = FILL_GOLD
    ws.cell(row=tot_row, column=6).font = FONT_BODY_BOLD
    return ("ME_MaterialTotal", f"'Material Takeoff'!$F${tot_row}")


def build_labor(ws):
    ws.title = "Labor Estimate"
    widths(ws, [("A", 14), ("B", 50), ("C", 10), ("D", 10), ("E", 16), ("F", 14), ("G", 16), ("H", 18)])
    title(ws, 1, "LABOR ESTIMATE (MCAA/SMACNA-style)", 8)
    header(ws, 3, ["CSI Code", "Activity", "QTY", "UOM", "Hours / UOM", "Subtotal Hours", "Crew Loaded Rate ($/hr)", "Subtotal Labor ($)"])
    items = [
        # Ductwork
        ("23 31 13", "Install galvanized rectangular duct (24ga)", 0, "SF", 0.18, 68),
        ("23 31 13", "Install galvanized rectangular duct (20ga)", 0, "SF", 0.25, 68),
        ("23 31 13", "Install round spiral duct", 0, "SF", 0.13, 68),
        ("23 31 16", "Install stainless kitchen exhaust", 0, "SF", 0.45, 68),
        ("23 31 19", "Install flex duct + connect", 0, "LF", 0.18, 68),
        ("23 33 13", "Install volume damper", 0, "EA", 0.40, 68),
        ("23 33 33", "Install fire damper (per UL system)", 0, "EA", 0.65, 68),
        ("23 33 33", "Install smoke/combination damper", 0, "EA", 0.85, 68),
        ("23 34 13", "Hang exhaust fan inline", 0, "EA", 1.50, 68),
        ("23 34 13", "Set roof-mount exhaust fan", 0, "EA", 4.50, 68),
        ("23 37 13", "Install supply diffuser", 0, "EA", 0.30, 68),
        ("23 37 13", "Install return grille", 0, "EA", 0.25, 68),
        ("23 37 13", "Install linear slot diffuser", 0, "EA", 0.65, 68),
        # Hydronic piping
        ("23 21 13", "Install copper hydronic 3/4\" - 1\"", 0, "LF", 0.18, 72),
        ("23 21 13", "Install copper hydronic 1-1/2\" - 2\"", 0, "LF", 0.32, 72),
        ("23 21 13", "Install black steel hydronic 2\"", 0, "LF", 0.35, 72),
        ("23 21 13", "Install black steel hydronic 4\"", 0, "LF", 0.65, 72),
        ("23 21 13", "Install PEX-A hydronic 1\"", 0, "LF", 0.08, 72),
        ("23 21 23", "Set inline circulator pump", 0, "EA", 1.20, 72),
        ("23 21 23", "Set base-mounted pump w/ flex connectors", 0, "EA", 4.50, 72),
        ("23 21 23", "Set end-suction pump on housekeeping pad", 0, "EA", 6.50, 72),
        ("23 21 29", "Install expansion tank", 0, "EA", 1.20, 72),
        ("23 21 29", "Install air separator", 0, "EA", 1.50, 72),
        ("23 05 23", "Install valve 1\" - 2\"", 0, "EA", 0.40, 72),
        ("23 05 23", "Install valve 4\"", 0, "EA", 0.85, 72),
        ("23 05 23", "Install valve 6\" + larger", 0, "EA", 1.50, 72),
        # Gas piping
        ("23 11 23", "Install gas piping 1\" black iron", 0, "LF", 0.35, 72),
        ("23 11 23", "Install gas piping 2\" black iron", 0, "LF", 0.52, 72),
        # Refrigerant
        ("23 62 23", "Install refrigerant pipe + insulate, 5/8\"", 0, "LF", 0.25, 74),
        ("23 62 23", "Install refrigerant pipe + insulate, 7/8\"", 0, "LF", 0.35, 74),
        ("23 62 23", "Refrigerant connection — brazed under N2", 0, "EA", 0.85, 74),
        # Insulation
        ("23 07 13", "Insulate duct — fiberglass wrap", 0, "SF", 0.08, 56),
        ("23 07 19", "Insulate pipe — 1\" - 2\"", 0, "LF", 0.12, 56),
        ("23 07 19", "Insulate pipe — 4\" + larger", 0, "LF", 0.20, 56),
        # Equipment set
        ("23 36 16", "Set + connect VAV box (single-duct)", 0, "EA", 3.50, 68),
        ("23 36 16", "Set + connect VAV box (fan-powered)", 0, "EA", 4.50, 68),
        ("23 52 13", "Set 500-MBH boiler + connect", 0, "EA", 18.00, 72),
        ("23 52 16", "Set 1500-MBH condensing boiler + connect", 0, "EA", 32.00, 72),
        ("23 54 13", "Set + connect gas furnace 80 MBH", 0, "EA", 4.50, 72),
        ("23 73 13", "Set indoor AHU + connect 3000 CFM", 0, "EA", 16.00, 68),
        ("23 73 13", "Set indoor AHU + connect 8000 CFM custom", 0, "EA", 28.00, 68),
        ("23 73 23", "Set rooftop unit 5 ton + connect", 0, "EA", 12.00, 68),
        ("23 73 23", "Set rooftop unit 15 ton + connect", 0, "EA", 18.00, 68),
        ("23 73 23", "Set rooftop unit 25 ton + connect", 0, "EA", 26.00, 68),
        ("23 81 13", "Set fan coil unit + connect", 0, "EA", 4.50, 72),
        ("23 81 23", "Set + connect mini-split indoor head", 0, "EA", 4.50, 74),
        ("23 81 26", "Set + connect heat pump outdoor 4 ton", 0, "EA", 9.50, 74),
        ("23 62 13", "Set DX split outdoor 3-5 ton", 0, "EA", 5.50, 74),
        ("23 62 13", "Set VRF outdoor 10 ton", 0, "EA", 18.00, 74),
        ("23 64 16", "Set chiller 60 ton air-cooled", 0, "EA", 28.00, 72),
        ("23 65 13", "Set cooling tower 200 ton crossflow", 0, "EA", 40.00, 72),
        ("23 72 13", "Set ERV 1000 CFM", 0, "EA", 12.00, 68),
        # Controls
        ("23 09 13", "Install + commission VAV controller", 0, "EA", 2.50, 78),
        ("23 09 13", "Install + commission AHU controller", 0, "EA", 8.00, 78),
        ("23 09 33", "Install thermostat + connect", 0, "EA", 0.85, 78),
        ("23 09 33", "Install duct sensor + connect", 0, "EA", 0.50, 78),
        # Commissioning
        ("23 05 93", "TAB — supply air balance per outlet", 0, "EA", 0.50, 70),
        ("23 05 93", "TAB — hydronic balance per circuit", 0, "EA", 1.20, 70),
        ("23 08 00", "Cx — equipment functional test", 0, "EA", 4.00, 78),
        ("23 08 00", "Cx — integration testing", 0, "LS", 16.00, 78),
    ]
    for i, (csi, act, qty, uom, hrs, rate) in enumerate(items):
        r = 4 + i
        ws.cell(row=r, column=1, value=csi).alignment = CENTER
        ws.cell(row=r, column=2, value=act).alignment = LEFT
        ws.cell(row=r, column=3, value=qty).number_format = FMT_INT
        ws.cell(row=r, column=4, value=uom).alignment = CENTER
        ws.cell(row=r, column=5, value=hrs).number_format = "0.00"
        ws.cell(row=r, column=6, value=f"=C{r}*E{r}").number_format = FMT_HRS
        ws.cell(row=r, column=7, value=rate).number_format = FMT_USD
        ws.cell(row=r, column=8, value=f"=F{r}*G{r}").number_format = FMT_USD
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = FONT_BODY
    last = 4 + len(items) - 1
    prod_row = last + 2
    ws.cell(row=prod_row, column=7, value="Productivity factor").font = FONT_BODY_BOLD
    ws.cell(row=prod_row, column=7).alignment = RIGHT
    ws.cell(row=prod_row, column=8, value=1.00).number_format = FMT_NUM2
    ws.cell(row=prod_row, column=8).fill = FILL_SUBHEADER
    sub_row = prod_row + 1
    ws.cell(row=sub_row, column=7, value="Labor Subtotal").font = FONT_BODY_BOLD
    ws.cell(row=sub_row, column=7).alignment = RIGHT
    ws.cell(row=sub_row, column=8, value=f"=SUM(H4:H{last})*H{prod_row}").number_format = FMT_USD
    ws.cell(row=sub_row, column=8).fill = FILL_GOLD
    ws.cell(row=sub_row, column=8).font = FONT_BODY_BOLD
    return ("ME_LaborTotal", f"'Labor Estimate'!$H${sub_row}")


def build_equipment(ws):
    ws.title = "Equipment"
    widths(ws, [("A", 42), ("B", 14), ("C", 14), ("D", 12), ("E", 16), ("F", 18)])
    title(ws, 1, "EQUIPMENT ESTIMATE", 6)
    header(ws, 3, ["Equipment", "Days", "Daily Rate ($)", "Weeks", "Weekly Rate ($)", "Subtotal ($)"])
    items = [
        ("Sheet-metal brake / shear (rental)", 0, 145.00, 0, 485.00),
        ("Threading machine 1\" - 2\"", 0, 145.00, 0, 525.00),
        ("Pipe-bending machine — hydronic", 0, 95.00, 0, 285.00),
        ("Refrigerant recovery machine + cylinders", 0, 95.00, 0, 285.00),
        ("Vacuum pump — 2-stage 8 CFM", 0, 65.00, 0, 195.00),
        ("Nitrogen tank + regulator", 0, 35.00, 0, 105.00),
        ("Brazing torch + tank set", 0, 45.00, 0, 145.00),
        ("Manifold gauge set + scale", 0, 35.00, 0, 105.00),
        ("Hydronic flush kit", 0, 95.00, 0, 285.00),
        ("Boom lift — 45'", 0, 425.00, 0, 1485.00),
        ("Scissor lift — 19'", 0, 145.00, 0, 485.00),
        ("Forklift — 5000 lb pneumatic", 0, 285.00, 0, 985.00),
        ("Generator — 5kW portable", 0, 65.00, 0, 195.00),
        ("Crane / rigging for heavy equipment lift", 0, 1850.00, 0, 0.00),
        ("Office trailer (allocation)", 0, 0.00, 0, 285.00),
        ("Truck / van allocation", 0, 0.00, 0, 425.00),
    ]
    for i, (name, days, drate, weeks, wrate) in enumerate(items):
        r = 4 + i
        ws.cell(row=r, column=1, value=name).alignment = LEFT
        ws.cell(row=r, column=2, value=days).number_format = FMT_INT
        ws.cell(row=r, column=3, value=drate).number_format = FMT_USD
        ws.cell(row=r, column=4, value=weeks).number_format = FMT_INT
        ws.cell(row=r, column=5, value=wrate).number_format = FMT_USD
        ws.cell(row=r, column=6, value=f"=(B{r}*C{r})+(D{r}*E{r})").number_format = FMT_USD
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = FONT_BODY
    last = 4 + len(items) - 1
    sub_row = last + 2
    ws.cell(row=sub_row, column=5, value="Equipment Subtotal").alignment = RIGHT
    ws.cell(row=sub_row, column=5).font = FONT_BODY_BOLD
    ws.cell(row=sub_row, column=6, value=f"=SUM(F4:F{last})").number_format = FMT_USD
    ws.cell(row=sub_row, column=6).fill = FILL_GOLD
    ws.cell(row=sub_row, column=6).font = FONT_BODY_BOLD
    return ("ME_EquipTotal", f"'Equipment'!$F${sub_row}")


def build_subcontract(ws):
    ws.title = "Subcontract"
    widths(ws, [("A", 42), ("B", 22), ("C", 18), ("D", 16), ("E", 18)])
    title(ws, 1, "SUBCONTRACT ESTIMATE", 5)
    header(ws, 3, ["Scope (if subbed)", "Vendor / Quote", "Quote $", "Markup %", "Subtotal ($)"])
    items = [
        ("Test, Adjust, Balance (TAB) — NEBB / AABC", "", 0.00, 0.10),
        ("Commissioning (Cx) — third-party CxA", "", 0.00, 0.10),
        ("BAS / DDC programming", "", 0.00, 0.10),
        ("Refrigerant charging + commissioning (EPA 608)", "", 0.00, 0.10),
        ("Boiler / chiller start-up — vendor", "", 0.00, 0.05),
        ("Insulation (if subbed)", "", 0.00, 0.10),
        ("Sheet-metal pre-fab (if subbed)", "", 0.00, 0.10),
        ("Crane / rigging — specialty heavy lift", "", 0.00, 0.10),
        ("Roof curbs + flashing", "", 0.00, 0.10),
        ("Hazmat / asbestos abatement (legacy MEP)", "", 0.00, 0.10),
        ("Concrete saw-cutting / core drilling", "", 0.00, 0.10),
        ("Hot-tap / live-cutover (gas / chilled water)", "", 0.00, 0.10),
    ]
    for i, (scope, vendor, quote, mk) in enumerate(items):
        r = 4 + i
        ws.cell(row=r, column=1, value=scope).alignment = LEFT
        ws.cell(row=r, column=2, value=vendor).alignment = LEFT
        ws.cell(row=r, column=3, value=quote).number_format = FMT_USD
        ws.cell(row=r, column=4, value=mk).number_format = FMT_PCT
        ws.cell(row=r, column=5, value=f"=C{r}*(1+D{r})").number_format = FMT_USD
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = FONT_BODY
    last = 4 + len(items) - 1
    sub_row = last + 2
    ws.cell(row=sub_row, column=4, value="Subcontract Subtotal").alignment = RIGHT
    ws.cell(row=sub_row, column=4).font = FONT_BODY_BOLD
    ws.cell(row=sub_row, column=5, value=f"=SUM(E4:E{last})").number_format = FMT_USD
    ws.cell(row=sub_row, column=5).fill = FILL_GOLD
    ws.cell(row=sub_row, column=5).font = FONT_BODY_BOLD
    return ("ME_SubTotal", f"'Subcontract'!$E${sub_row}")


def build_premiums(ws):
    ws.title = "Premiums"
    widths(ws, [("A", 50), ("B", 20), ("C", 18), ("D", 18)])
    title(ws, 1, "PREMIUMS & CONDITIONS", 4)
    header(ws, 3, ["Premium", "Applies? (Y/N)", "Adder %", "Adder $ (overrides %)"])
    items = [
        ("Prevailing wage", "N", 0.15, 0.00),
        ("Occupied building / live-cutover", "N", 0.10, 0.00),
        ("Hospital ICRA tier 3+", "N", 0.10, 0.00),
        ("Heights > 14 ft (lift / crane)", "N", 0.05, 0.00),
        ("Roof-top set requiring crane lift", "N", 0.06, 0.00),
        ("Tight mechanical room (extended labor)", "N", 0.05, 0.00),
        ("Hazardous-area (refrigerant + flammable)", "N", 0.06, 0.00),
        ("After-hours / weekend / night shift", "N", 0.20, 0.00),
        ("Travel / per diem", "N", 0.03, 0.00),
        ("Winter / cold-weather working", "N", 0.04, 0.00),
        ("LEED / Cx documentation", "N", 0.04, 0.00),
        ("Owner-direct equipment (no markup)", "N", -0.10, 0.00),
    ]
    for i, (label, applies, pct, dollar) in enumerate(items):
        r = 4 + i
        ws.cell(row=r, column=1, value=label).alignment = LEFT
        ws.cell(row=r, column=2, value=applies).alignment = CENTER
        ws.cell(row=r, column=3, value=pct).number_format = FMT_PCT
        ws.cell(row=r, column=4, value=dollar).number_format = FMT_USD
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = FONT_BODY
    last = 4 + len(items) - 1
    total_row = last + 2
    ws.cell(row=total_row, column=1, value="Premium $ total (applies-Y only)").alignment = RIGHT
    ws.cell(row=total_row, column=1).font = FONT_BODY_BOLD
    ws.cell(row=total_row, column=4, value=f"=SUMPRODUCT((B4:B{last}=\"Y\")*D4:D{last})").number_format = FMT_USD
    ws.cell(row=total_row, column=4).fill = FILL_GOLD
    ws.cell(row=total_row, column=4).font = FONT_BODY_BOLD
    pct_row = total_row + 1
    ws.cell(row=pct_row, column=1, value="Premium % total (applies-Y only)").alignment = RIGHT
    ws.cell(row=pct_row, column=1).font = FONT_BODY_BOLD
    ws.cell(row=pct_row, column=4, value=f"=SUMPRODUCT((B4:B{last}=\"Y\")*C4:C{last})").number_format = FMT_PCT
    ws.cell(row=pct_row, column=4).fill = FILL_GOLD
    ws.cell(row=pct_row, column=4).font = FONT_BODY_BOLD
    return ("ME_PremiumDollars", f"'Premiums'!$D${total_row}"), ("ME_PremiumPct", f"'Premiums'!$D${pct_row}")


def build_summary(ws, refs):
    ws.title = "Bid Summary"
    widths(ws, [("A", 4), ("B", 38), ("C", 20), ("D", 22)])
    title(ws, 1, "BID SUMMARY — Mechanical Trade Pack", 4)
    rows = [
        ("", "Material total (incl. tax)", f"={refs['ME_MaterialTotal']}"),
        ("", "Labor total", f"={refs['ME_LaborTotal']}"),
        ("", "Equipment total", f"={refs['ME_EquipTotal']}"),
        ("", "Subcontract total", f"={refs['ME_SubTotal']}"),
        ("", "Premiums $", f"={refs['ME_PremiumDollars']}"),
        ("", "Direct cost subtotal", "=SUM(C3:C7)"),
        ("", "Premium % applied", f"={refs['ME_PremiumPct']}"),
        ("", "Premium % adder $", "=C8*C9"),
        ("", "Direct + premium %", "=C8+C10"),
        ("", "Contingency %", 0.05),
        ("", "Contingency $", "=C11*C12"),
        ("", "Direct + contingency", "=C11+C13"),
        ("", "Overhead %", 0.10),
        ("", "Overhead $", "=C14*C15"),
        ("", "Profit %", 0.10),
        ("", "Profit $", "=(C14+C16)*C17"),
        ("", "Bond %", 0.012),
        ("", "Bond $", "=(C14+C16+C18)*C19"),
        ("", "Insurance / GL %", 0.018),
        ("", "Insurance $", "=(C14+C16+C18)*C21"),
        ("", "FINAL BID", "=C14+C16+C18+C20+C22"),
    ]
    for i, (a, label, val) in enumerate(rows):
        r = 3 + i
        ws.cell(row=r, column=2, value=label).font = FONT_BODY_BOLD
        ws.cell(row=r, column=2).alignment = RIGHT
        ws.cell(row=r, column=2).border = BORDER
        c = ws.cell(row=r, column=3, value=val)
        c.border = BORDER; c.font = FONT_BODY_BOLD
        c.number_format = FMT_PCT if label.endswith("%") else FMT_USD
        if label == "FINAL BID":
            c.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
            ws.cell(row=r, column=2).fill = FILL_GOLD
            ws.cell(row=r, column=2).font = Font(name="Calibri", size=14, bold=True, color=BRAND_BLUE)


def build_csi(ws):
    ws.title = "CSI Reference"
    widths(ws, [("A", 14), ("B", 80)])
    title(ws, 1, "CSI MASTERFORMAT — DIVISION 23 HEATING / VENTILATING / AIR-CONDITIONING", 2)
    rows = [
        ("23 00 00", "HEATING, VENTILATING, AND AIR-CONDITIONING (HVAC)"),
        ("23 05 00", "Common Work Results for HVAC"),
        ("23 05 13", "Common Motor Requirements for HVAC Equipment"),
        ("23 05 17", "Sleeves and Sleeve Seals for HVAC Piping"),
        ("23 05 19", "Meters and Gauges for HVAC Piping"),
        ("23 05 23", "General-Duty Valves for HVAC Piping"),
        ("23 05 29", "Hangers and Supports for HVAC Piping and Equipment"),
        ("23 05 53", "Identification for HVAC Piping and Equipment"),
        ("23 05 93", "Testing, Adjusting, and Balancing for HVAC"),
        ("23 07 00", "HVAC Insulation"),
        ("23 07 13", "Duct Insulation"),
        ("23 07 19", "HVAC Piping Insulation"),
        ("23 08 00", "Commissioning of HVAC"),
        ("23 09 00", "Instrumentation and Control for HVAC"),
        ("23 09 23", "Direct-Digital Control System for HVAC"),
        ("23 09 33", "Electric and Electronic Control System for HVAC"),
        ("23 11 23", "Facility Natural-Gas Piping"),
        ("23 21 13", "Hydronic Piping"),
        ("23 21 16", "Hydronic Piping Specialties"),
        ("23 21 23", "Hydronic Pumps"),
        ("23 22 13", "Steam and Condensate Heating Piping"),
        ("23 25 13", "Water Treatment for Closed-Loop Hydronic Systems"),
        ("23 31 13", "Metal Ducts"),
        ("23 31 16", "Nonmetal Ducts (fabric, FRP)"),
        ("23 31 19", "HVAC Casings"),
        ("23 33 13", "Dampers"),
        ("23 33 19", "Duct Silencers"),
        ("23 33 33", "Air-Duct Accessories — Fire Dampers, Smoke Dampers"),
        ("23 34 13", "Axial HVAC Fans"),
        ("23 34 16", "Centrifugal HVAC Fans"),
        ("23 36 16", "Variable Air Volume Boxes (VAV)"),
        ("23 37 13", "Diffusers, Registers, and Grilles"),
        ("23 41 00", "Particulate Air Filtration"),
        ("23 52 13", "Hot-Water Boilers"),
        ("23 52 39", "Electric Boilers"),
        ("23 54 13", "Gas Furnaces"),
        ("23 57 13", "Heat Exchangers for HVAC"),
        ("23 62 13", "Packaged Compressor and Condenser Units"),
        ("23 62 23", "Refrigerant Piping"),
        ("23 64 16", "Air-Cooled Chillers"),
        ("23 64 19", "Water-Cooled Chillers"),
        ("23 65 13", "Cooling Towers"),
        ("23 72 13", "Air-to-Air Energy Recovery Equipment (ERV / HRV)"),
        ("23 73 13", "Indoor Central-Station Air-Handling Units"),
        ("23 73 23", "Packaged Outdoor HVAC (Rooftop)"),
        ("23 81 13", "Fan-Coil Units"),
        ("23 81 23", "Computer-Room Air Conditioners (CRAC / CRAH)"),
        ("23 81 26", "Heat Pumps — Air-Source"),
        ("23 83 13", "Radiant Heating"),
        ("22 13 / 22 14", "Plumbing — sanitary / storm (cross-ref)"),
        ("26 27 / 26 28", "Electrical — power + disconnects to HVAC equipment (cross-ref)"),
        ("27 21 / 27 22", "BAS Communications (cross-ref)"),
    ]
    for i, (code, name) in enumerate(rows):
        r = 3 + i
        ws.cell(row=r, column=1, value=code).alignment = CENTER
        ws.cell(row=r, column=1).font = FONT_BODY_BOLD
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2, value=name).alignment = LEFT
        ws.cell(row=r, column=2).font = FONT_BODY
        ws.cell(row=r, column=2).border = BORDER


def main():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb = Workbook()
    build_instructions(wb.active)
    build_project_info(wb.create_sheet("Project Info"))
    build_wages(wb.create_sheet("Wage Rates"))
    mat_ref = build_material(wb.create_sheet("Material Takeoff"))
    lab_ref = build_labor(wb.create_sheet("Labor Estimate"))
    eq_ref = build_equipment(wb.create_sheet("Equipment"))
    sub_ref = build_subcontract(wb.create_sheet("Subcontract"))
    prem_d, prem_p = build_premiums(wb.create_sheet("Premiums"))
    refs = {mat_ref[0]: mat_ref[1], lab_ref[0]: lab_ref[1], eq_ref[0]: eq_ref[1],
            sub_ref[0]: sub_ref[1], prem_d[0]: prem_d[1], prem_p[0]: prem_p[1]}
    ws_sum = wb.create_sheet("Bid Summary")
    build_summary(ws_sum, refs)
    wb.move_sheet(ws_sum, offset=-7)
    build_csi(wb.create_sheet("CSI Reference"))
    wb.save(OUT)
    print(f"OK — wrote {OUT}")


if __name__ == "__main__":
    main()
