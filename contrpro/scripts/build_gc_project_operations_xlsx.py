#!/usr/bin/env python3
"""
Build ContrPro GC Project Operations (XLSX) — GC suite (Complete tier).

CONSOLIDATED workbook covering the GC's day-to-day project execution. Single
shared Project Info tab drives every other operational tab via named ranges,
so common fields (project name, GC, owner, etc.) are entered ONCE not 6x.

This is the answer to "don't make me fill in like information across 40
workbooks": RFI Log + Submittal Log + Schedule + DFR + Punch List + Meeting
Minutes all live in one workbook, sharing one Project Info header.

Tabs:
  1. Instructions
  2. Project Info               (shared header — pulled into every other tab)
  3. RFI Log
  4. Submittal Log
  5. Schedule / Lookahead
  6. Daily Field Report Log     (flat row-per-day log)
  7. Punch List
  8. Meeting Minutes
  9. CSI Reference              (hidden)

Run:
    /Users/home/charles/.venv/bin/python3 \\
        /Users/home/charles/contrpro/scripts/build_gc_project_operations_xlsx.py

Output:
    /Users/home/charles/contrpro/files/packages/complete/gc/GC_Project_Operations.xlsx
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
ACCENT_GOLD = "C9A227"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
YELLOW_FILL = "FFEB9C"
YELLOW_FONT = "9C5700"
ORANGE_FILL = "FFD8A8"
ORANGE_FONT = "8A4B00"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_SUMMARY_LABEL = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW_FILL)
FILL_ORANGE = PatternFill("solid", fgColor=ORANGE_FILL)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_BIG_NUMBER = Font(name="Calibri", size=18, bold=True, color=BRAND_BLUE)
FONT_GREY_ITALIC = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_PCT = "0.0%"
FMT_INT = "0"
FMT_NUM2 = "0.00"
FMT_DATE = "yyyy-mm-dd"

DIVISIONS = [
    ("01", "General Requirements"), ("02", "Existing Conditions"), ("03", "Concrete"),
    ("04", "Masonry"), ("05", "Metals"), ("06", "Wood, Plastics, and Composites"),
    ("07", "Thermal and Moisture Protection"), ("08", "Openings"), ("09", "Finishes"),
    ("10", "Specialties"), ("11", "Equipment"), ("12", "Furnishings"),
    ("13", "Special Construction"), ("14", "Conveying Equipment"), ("21", "Fire Suppression"),
    ("22", "Plumbing"), ("23", "HVAC"), ("25", "Integrated Automation"),
    ("26", "Electrical"), ("27", "Communications"), ("28", "Electronic Safety and Security"),
    ("31", "Earthwork"), ("32", "Exterior Improvements"), ("33", "Utilities"),
]


def set_col_widths(ws, widths):
    for col, w in widths:
        ws.column_dimensions[col].width = w


def style_header_row(ws, row, cols, start_col=1):
    for c in range(start_col, start_col + cols):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


def apply_body_style(ws, row, cols, start_col=1):
    for c in range(start_col, start_col + cols):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_BODY
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.border = BORDER


# ---------------------------------------------------------------------------
# Tab 1: Instructions
# ---------------------------------------------------------------------------

def build_instructions(ws):
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 110)])

    ws["A1"] = "GC PROJECT OPERATIONS — Instructions"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 36

    rows = [
        "",
        "WHAT THIS WORKBOOK IS",
        "Single consolidated workbook covering the GC's day-to-day project execution: RFI Log, Submittal "
        "Log, Schedule/Lookahead, Daily Field Report Log, Punch List, and Meeting Minutes. Project Info "
        "is entered ONCE on the Project Info tab and pulled into every other tab via named ranges — you "
        "do not re-type project name, owner, GC, etc., across six different workbooks.",
        "",
        "WHY CONSOLIDATED",
        "These are all tools the SAME user (the GC's Project Manager) uses in the SAME workflow (project "
        "execution between award and closeout). Keeping them in one workbook means: shared header data, "
        "shared CSI dropdowns, single file to back up, single file to email to a covering super.",
        "",
        "WHAT'S NOT IN HERE",
        "  - Bidding: use GC_Bid_Estimator.xlsx (different phase, different user — the estimator).",
        "  - Monthly billing to Owner: use GC_Application_for_Payment.xlsx (different cadence, different "
        "audience — Owner + Architect).",
        "  - Sub management: use Subcontractor_Tracker.xlsx (Business tier) for vendor master + COIs.",
        "  - Job cost accounting: use Job_Costing_Spreadsheet.xlsx (Business tier).",
        "  - Change Orders to subs: use Change_Order_Log.xlsx (Business tier).",
        "",
        "WORKFLOW",
        "1. Project Info: fill in once at award. Drives every other tab.",
        "2. RFI Log: log RFIs sent to architect/engineer/owner. Track response cadence.",
        "3. Submittal Log: track shop drawings, product data, samples, mockups through the approval cycle.",
        "4. Schedule / Lookahead: high-level activity tracking with 3-week lookahead. Not a CPM tool — use "
        "P6/MS Project/Smartsheet for true CPM scheduling.",
        "5. Daily Field Report Log: one ROW per day (not one tab per day). Crew counts, weather, key "
        "activities, issues. Use Daily_Field_Report.xlsx from Universal Sub Suite for per-day detail.",
        "6. Punch List: track punch-list items from substantial completion to final acceptance.",
        "7. Meeting Minutes: running log of all project meetings (Owner-Arch-GC, coordination, safety, "
        "subcontractor).",
        "",
        "DESIGN NOTE",
        "This workbook was deliberately consolidated AGAINST the 'one tool per file' pattern. The other "
        "ContrPro workbooks (Bid Estimator, Application for Payment, Sub SOV, T&M Tracker, etc.) are each "
        "their own files because they serve different users/phases. The operational tools all serve the GC "
        "PM during execution; consolidating them avoids the 'fill in project info six times' tax.",
    ]

    for i, text in enumerate(rows, start=2):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if text and text == text.upper() and len(text) < 70:
            cell.font = FONT_H2
        else:
            cell.font = FONT_BODY


# ---------------------------------------------------------------------------
# Tab 2: Project Info (shared)
# ---------------------------------------------------------------------------

def build_project_info(ws):
    ws.title = "Project Info"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 38), ("C", 40)])

    ws["B1"] = "PROJECT INFORMATION — Shared Across All Tabs"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:C1")
    ws.row_dimensions[1].height = 32

    ws["B2"] = "Enter once. Every other tab references this data via named ranges."
    ws["B2"].font = FONT_GREY_ITALIC
    ws.merge_cells("B2:C2")

    info_rows = [
        ("Project Name", ""),
        ("Project Address", ""),
        ("Project No. (GC's)", ""),
        ("Owner", ""),
        ("Architect", ""),
        ("Structural Engineer", ""),
        ("MEP Engineer", ""),
        ("Civil Engineer", ""),
        ("General Contractor", ""),
        ("Prime Contract No.", ""),
        ("Prime Contract Date", ""),
        ("Notice to Proceed Date", ""),
        ("Anticipated Substantial Completion", ""),
        ("Anticipated Final Completion", ""),
        ("", ""),
        ("KEY PERSONNEL", ""),
        ("Project Executive (GC)", ""),
        ("Project Manager (GC)", ""),
        ("Site Superintendent (GC)", ""),
        ("Safety Officer (GC)", ""),
        ("Owner's Rep", ""),
        ("Architect Project Manager", ""),
        ("", ""),
        ("RFI / SUBMITTAL SLA", ""),
        ("RFI Response SLA (days, per contract)", 10),
        ("Submittal Review SLA (days, per contract)", 14),
        ("Resubmittal SLA (days, per contract)", 10),
    ]

    name_map = {
        "ProjectName": "$C$3",
        "ProjectAddress": "$C$4",
        "OwnerName": "$C$6",
        "ArchitectName": "$C$7",
        "GCName": "$C$11",
        "PrimeContractNo": "$C$12",
        "GCPM": "$C$20",
        "GCSuper": "$C$21",
        "GCSafety": "$C$22",
        "RFI_SLA": "$C$27",
        "SubmittalSLA": "$C$28",
    }
    for nm, ref in name_map.items():
        ws.parent.defined_names[nm] = DefinedName(name=nm, attr_text=f"'Project Info'!{ref}")

    for i, (label, val) in enumerate(info_rows, start=3):
        is_section = label and label == label.upper() and not val and len(label) < 50
        if is_section:
            ws.cell(row=i, column=2, value=label).font = FONT_H2
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
            continue
        if label:
            ws.cell(row=i, column=2, value=label).font = FONT_BODY_BOLD
            ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
            ws.cell(row=i, column=2).border = BORDER
            c = ws.cell(row=i, column=3, value=val)
            c.font = FONT_BODY
            c.border = BORDER
            if "Date" in label or "Completion" in label or "Notice" in label:
                c.number_format = FMT_DATE
            if "SLA" in label or "days" in label.lower():
                c.number_format = FMT_INT


# ---------------------------------------------------------------------------
# Tab 3: RFI Log
# ---------------------------------------------------------------------------

RFI_COLS = [
    ("A", "RFI #", 8),
    ("B", "Date Sent", 12),
    ("C", "From", 22),
    ("D", "To", 22),
    ("E", "CSI Div", 9),
    ("F", "Subject", 30),
    ("G", "Description", 36),
    ("H", "Response Due (auto)", 14),
    ("I", "Response Date", 13),
    ("J", "Days Open", 11),
    ("K", "Status", 13),
    ("L", "Cost Impact ($)", 14),
    ("M", "Sched Impact (days)", 13),
    ("N", "Resolution", 28),
    ("O", "Notes", 22),
]


def build_rfi_log(ws):
    ws.title = "RFI Log"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in RFI_COLS])

    ws["A1"] = "RFI LOG"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:O1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Project: =ProjectName    GC: =GCName    Architect: =ArchitectName    RFI SLA: =RFI_SLA days"
    )
    ws["A2"].font = FONT_BODY_BOLD
    ws.merge_cells("A2:O2")

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(RFI_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(RFI_COLS))

    DATA_ROWS = 100
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        ws.cell(row=r, column=1, value=f"=IF(F{r}=\"\",\"\",ROW()-{HEADER_ROW})").number_format = FMT_INT
        ws.cell(row=r, column=2).number_format = FMT_DATE
        # Response due = sent + SLA
        ws.cell(row=r, column=8, value=f"=IF(B{r}=\"\",\"\",B{r}+RFI_SLA)").number_format = FMT_DATE
        ws.cell(row=r, column=9).number_format = FMT_DATE
        # Days open = response date or today - sent
        ws.cell(
            row=r,
            column=10,
            value=f"=IF(B{r}=\"\",\"\",IF(I{r}=\"\",TODAY()-B{r},I{r}-B{r}))",
        ).number_format = FMT_INT
        ws.cell(row=r, column=12).number_format = FMT_USD
        ws.cell(row=r, column=13).number_format = FMT_INT
        apply_body_style(ws, r, len(RFI_COLS))

    # CSI dropdown
    csi_codes = ",".join(c for c, _ in DIVISIONS)
    dv_csi = DataValidation(type="list", formula1=f'"{csi_codes}"', allow_blank=True)
    dv_csi.add(f"E{HEADER_ROW+1}:E{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_csi)

    # Status dropdown
    dv_status = DataValidation(
        type="list",
        formula1='"Open,Awaiting Response,Response Received,Resolved,Closed,Withdrawn"',
        allow_blank=True,
    )
    dv_status.add(f"K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_status)

    # Conditional format Status
    ws.conditional_formatting.add(
        f"K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'OR(K{HEADER_ROW+1}="Resolved",K{HEADER_ROW+1}="Closed")'], fill=FILL_GREEN, font=Font(color=GREEN_FONT)),
    )
    ws.conditional_formatting.add(
        f"K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'OR(K{HEADER_ROW+1}="Open",K{HEADER_ROW+1}="Awaiting Response")'], fill=FILL_YELLOW, font=Font(color=YELLOW_FONT)),
    )
    # Highlight days-open > SLA when still open
    ws.conditional_formatting.add(
        f"J{HEADER_ROW+1}:J{HEADER_ROW+DATA_ROWS}",
        FormulaRule(
            formula=[f'AND(I{HEADER_ROW+1}="",J{HEADER_ROW+1}>RFI_SLA)'],
            fill=FILL_RED,
            font=Font(bold=True, color=RED_FONT),
        ),
    )

    # Summary block above the data
    SUMMARY_START_ROW = HEADER_ROW + DATA_ROWS + 3
    ws.cell(row=SUMMARY_START_ROW, column=1, value="SUMMARY").font = FONT_H2
    summary_rows = [
        ("Total RFIs logged", f'=COUNTA(F{HEADER_ROW+1}:F{HEADER_ROW+DATA_ROWS})'),
        ("Open", f'=COUNTIF(K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS},"Open")+COUNTIF(K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS},"Awaiting Response")'),
        ("Resolved + Closed", f'=COUNTIF(K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS},"Resolved")+COUNTIF(K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS},"Closed")'),
        ("Past SLA (open + days > SLA)", f'=SUMPRODUCT((I{HEADER_ROW+1}:I{HEADER_ROW+DATA_ROWS}="")*(J{HEADER_ROW+1}:J{HEADER_ROW+DATA_ROWS}>RFI_SLA)*(B{HEADER_ROW+1}:B{HEADER_ROW+DATA_ROWS}<>""))'),
        ("Total cost impact ($)", f'=SUM(L{HEADER_ROW+1}:L{HEADER_ROW+DATA_ROWS})'),
        ("Total schedule impact (days)", f'=SUM(M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS})'),
    ]
    for i, (label, formula) in enumerate(summary_rows, start=SUMMARY_START_ROW + 1):
        ws.cell(row=i, column=1, value=label).font = FONT_BODY_BOLD
        ws.cell(row=i, column=1).fill = FILL_SUMMARY_LABEL
        ws.cell(row=i, column=1).border = BORDER
        c = ws.cell(row=i, column=2, value=formula)
        c.font = FONT_BIG_NUMBER
        c.number_format = FMT_USD if "$" in label else FMT_INT
        c.border = BORDER


# ---------------------------------------------------------------------------
# Tab 4: Submittal Log
# ---------------------------------------------------------------------------

SUB_COLS = [
    ("A", "Submittal #", 11),
    ("B", "Spec Section", 14),
    ("C", "Description", 32),
    ("D", "Type", 16),
    ("E", "Required by Date", 14),
    ("F", "Date Sent to Architect", 14),
    ("G", "Date Returned", 13),
    ("H", "Status", 22),
    ("I", "Resubmittal #", 11),
    ("J", "Final Approved Date", 14),
    ("K", "Days Out", 10),
    ("L", "Sub / Vendor", 22),
    ("M", "Notes", 22),
]


def build_submittal_log(ws):
    ws.title = "Submittal Log"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in SUB_COLS])

    ws["A1"] = "SUBMITTAL LOG"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:M1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Project: =ProjectName    GC: =GCName    Architect: =ArchitectName    Submittal SLA: =SubmittalSLA days"
    )
    ws["A2"].font = FONT_BODY_BOLD
    ws.merge_cells("A2:M2")

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(SUB_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(SUB_COLS))

    DATA_ROWS = 100
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        for c in (5, 6, 7, 10):
            ws.cell(row=r, column=c).number_format = FMT_DATE
        # Days out = returned - sent, or today - sent if not returned
        ws.cell(
            row=r,
            column=11,
            value=f"=IF(F{r}=\"\",\"\",IF(G{r}=\"\",TODAY()-F{r},G{r}-F{r}))",
        ).number_format = FMT_INT
        apply_body_style(ws, r, len(SUB_COLS))

    dv_type = DataValidation(
        type="list",
        formula1='"Shop Drawing,Product Data,Sample,Mockup,Schedule,Coordination Drawing,Other"',
        allow_blank=True,
    )
    dv_type.add(f"D{HEADER_ROW+1}:D{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_type)

    dv_status = DataValidation(
        type="list",
        formula1='"Pending,Submitted,Approved,Approved as Noted,Revise and Resubmit,Rejected,Closed"',
        allow_blank=True,
    )
    dv_status.add(f"H{HEADER_ROW+1}:H{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_status)

    # Conditional format
    ws.conditional_formatting.add(
        f"H{HEADER_ROW+1}:H{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'OR(H{HEADER_ROW+1}="Approved",H{HEADER_ROW+1}="Approved as Noted",H{HEADER_ROW+1}="Closed")'], fill=FILL_GREEN, font=Font(color=GREEN_FONT)),
    )
    ws.conditional_formatting.add(
        f"H{HEADER_ROW+1}:H{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'OR(H{HEADER_ROW+1}="Rejected",H{HEADER_ROW+1}="Revise and Resubmit")'], fill=FILL_RED, font=Font(color=RED_FONT)),
    )
    ws.conditional_formatting.add(
        f"K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS}",
        FormulaRule(
            formula=[f'AND(G{HEADER_ROW+1}="",K{HEADER_ROW+1}>SubmittalSLA)'],
            fill=FILL_RED,
            font=Font(bold=True, color=RED_FONT),
        ),
    )


# ---------------------------------------------------------------------------
# Tab 5: Schedule / Lookahead
# ---------------------------------------------------------------------------

SCHED_COLS = [
    ("A", "Activity ID", 10),
    ("B", "Activity / Milestone", 38),
    ("C", "Trade", 18),
    ("D", "CSI Div", 9),
    ("E", "Baseline Start", 14),
    ("F", "Baseline Finish", 14),
    ("G", "Actual / Forecast Start", 16),
    ("H", "Actual / Forecast Finish", 16),
    ("I", "Duration (days)", 12),
    ("J", "% Complete", 11),
    ("K", "Status", 14),
    ("L", "Predecessor IDs", 14),
    ("M", "Variance (days)", 12),
    ("N", "Notes", 22),
]


def build_schedule(ws):
    ws.title = "Schedule"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in SCHED_COLS])

    ws["A1"] = "SCHEDULE / 3-WEEK LOOKAHEAD"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:N1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Project: =ProjectName    GC: =GCName    "
        "Tracks high-level activities + milestones. NOT a CPM scheduling tool — use P6, MS Project, or "
        "Smartsheet for true CPM logic. This is for status reporting and weekly lookahead."
    )
    ws["A2"].font = FONT_BODY_BOLD
    ws.merge_cells("A2:N2")
    ws.row_dimensions[2].height = 36

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(SCHED_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(SCHED_COLS))

    DATA_ROWS = 80
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        for c in (5, 6, 7, 8):
            ws.cell(row=r, column=c).number_format = FMT_DATE
        # Duration = finish - start (use forecast if populated; else baseline)
        ws.cell(
            row=r,
            column=9,
            value=f"=IF(OR(E{r}=\"\",F{r}=\"\"),\"\",IF(AND(G{r}<>\"\",H{r}<>\"\"),H{r}-G{r},F{r}-E{r}))",
        ).number_format = FMT_INT
        ws.cell(row=r, column=10).number_format = FMT_PCT
        # Variance = (forecast finish - baseline finish) in days
        ws.cell(
            row=r,
            column=13,
            value=f"=IF(OR(F{r}=\"\",H{r}=\"\"),\"\",H{r}-F{r})",
        ).number_format = FMT_INT
        apply_body_style(ws, r, len(SCHED_COLS))

    dv_status = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Complete,Delayed,On Hold,Cancelled"',
        allow_blank=True,
    )
    dv_status.add(f"K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_status)

    csi_codes = ",".join(c for c, _ in DIVISIONS)
    dv_csi = DataValidation(type="list", formula1=f'"{csi_codes}"', allow_blank=True)
    dv_csi.add(f"D{HEADER_ROW+1}:D{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_csi)

    # Conditional formatting
    ws.conditional_formatting.add(
        f"K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'K{HEADER_ROW+1}="Complete"'], fill=FILL_GREEN, font=Font(color=GREEN_FONT)),
    )
    ws.conditional_formatting.add(
        f"K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'K{HEADER_ROW+1}="Delayed"'], fill=FILL_RED, font=Font(color=RED_FONT)),
    )
    ws.conditional_formatting.add(
        f"K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'K{HEADER_ROW+1}="In Progress"'], fill=FILL_YELLOW, font=Font(color=YELLOW_FONT)),
    )
    # Variance > 0 (late) red, < 0 (early) green
    ws.conditional_formatting.add(
        f"M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=FILL_RED, font=Font(color=RED_FONT)),
    )
    ws.conditional_formatting.add(
        f"M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS}",
        CellIsRule(operator="lessThan", formula=["0"], fill=FILL_GREEN, font=Font(color=GREEN_FONT)),
    )


# ---------------------------------------------------------------------------
# Tab 6: Daily Field Report Log (flat — row per day)
# ---------------------------------------------------------------------------

DFR_COLS = [
    ("A", "Date", 12),
    ("B", "Day", 8),
    ("C", "Weather AM", 14),
    ("D", "Weather PM", 14),
    ("E", "Temp °F (high/low)", 14),
    ("F", "Crew Count (Total)", 13),
    ("G", "Crews on Site (list)", 26),
    ("H", "Key Activities Today", 36),
    ("I", "Issues / Delays", 28),
    ("J", "Visitors", 22),
    ("K", "Safety Incidents", 16),
    ("L", "Inspections", 18),
    ("M", "Lost Hours (weather/delay)", 13),
    ("N", "Foreman / Super", 18),
    ("O", "Notes", 22),
]


def build_dfr_log(ws):
    ws.title = "DFR Log"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in DFR_COLS])

    ws["A1"] = "DAILY FIELD REPORT LOG"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:O1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Project: =ProjectName    GC: =GCName    Super: =GCSuper    "
        "One ROW per day. For per-day detail (crew names, hours per worker, materials), use the "
        "Daily_Field_Report.xlsx from Universal Sub Suite (33-tab monthly workbook)."
    )
    ws["A2"].font = FONT_BODY_BOLD
    ws.merge_cells("A2:O2")
    ws.row_dimensions[2].height = 36

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(DFR_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(DFR_COLS))

    DATA_ROWS = 200
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        ws.cell(row=r, column=1).number_format = FMT_DATE
        # Day-of-week derived from date
        ws.cell(row=r, column=2, value=f'=IF(A{r}="","",TEXT(A{r},"ddd"))')
        ws.cell(row=r, column=6).number_format = FMT_INT
        ws.cell(row=r, column=13).number_format = FMT_NUM2
        apply_body_style(ws, r, len(DFR_COLS))

    # Summary block
    SUMMARY_ROW = HEADER_ROW + DATA_ROWS + 3
    ws.cell(row=SUMMARY_ROW, column=1, value="DFR SUMMARY").font = FONT_H2
    summary = [
        ("Days logged",          f'=COUNTA(A{HEADER_ROW+1}:A{HEADER_ROW+DATA_ROWS})'),
        ("Total crew-days",      f'=SUM(F{HEADER_ROW+1}:F{HEADER_ROW+DATA_ROWS})'),
        ("Days with issues",     f'=COUNTIF(I{HEADER_ROW+1}:I{HEADER_ROW+DATA_ROWS},"<>")'),
        ("Days with safety inc.", f'=COUNTIF(K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS},"<>")'),
        ("Total lost hours",     f'=SUM(M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS})'),
    ]
    for i, (label, formula) in enumerate(summary, start=SUMMARY_ROW + 1):
        ws.cell(row=i, column=1, value=label).font = FONT_BODY_BOLD
        ws.cell(row=i, column=1).fill = FILL_SUMMARY_LABEL
        ws.cell(row=i, column=1).border = BORDER
        c = ws.cell(row=i, column=2, value=formula)
        c.font = FONT_BIG_NUMBER
        c.number_format = FMT_INT
        c.border = BORDER


# ---------------------------------------------------------------------------
# Tab 7: Punch List
# ---------------------------------------------------------------------------

PUNCH_COLS = [
    ("A", "Item #", 8),
    ("B", "Date Added", 12),
    ("C", "Source", 16),
    ("D", "Location", 22),
    ("E", "Trade", 16),
    ("F", "CSI Div", 9),
    ("G", "Description", 38),
    ("H", "Severity", 12),
    ("I", "Assigned To", 22),
    ("J", "Photo Ref", 14),
    ("K", "Date Completed", 13),
    ("L", "Verified By", 18),
    ("M", "Status", 14),
    ("N", "Notes", 22),
]


def build_punch_list(ws):
    ws.title = "Punch List"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in PUNCH_COLS])

    ws["A1"] = "PUNCH LIST"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:N1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = "Project: =ProjectName    GC: =GCName"
    ws["A2"].font = FONT_BODY_BOLD
    ws.merge_cells("A2:N2")

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(PUNCH_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(PUNCH_COLS))

    DATA_ROWS = 200
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        ws.cell(row=r, column=1, value=f"=IF(G{r}=\"\",\"\",ROW()-{HEADER_ROW})").number_format = FMT_INT
        ws.cell(row=r, column=2).number_format = FMT_DATE
        ws.cell(row=r, column=11).number_format = FMT_DATE
        apply_body_style(ws, r, len(PUNCH_COLS))

    dv_source = DataValidation(
        type="list",
        formula1='"GC Walk-Through,Owner Walk-Through,Architect Inspection,Substantial Completion,Pre-Final,Final,Warranty Item"',
        allow_blank=True,
    )
    dv_source.add(f"C{HEADER_ROW+1}:C{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_source)

    csi_codes = ",".join(c for c, _ in DIVISIONS)
    dv_csi = DataValidation(type="list", formula1=f'"{csi_codes}"', allow_blank=True)
    dv_csi.add(f"F{HEADER_ROW+1}:F{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_csi)

    dv_severity = DataValidation(
        type="list",
        formula1='"Minor,Standard,Major,Life Safety"',
        allow_blank=True,
    )
    dv_severity.add(f"H{HEADER_ROW+1}:H{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_severity)

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Awaiting Verification,Verified Complete,Closed,Deferred"',
        allow_blank=True,
    )
    dv_status.add(f"M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_status)

    # Conditional formatting
    ws.conditional_formatting.add(
        f"M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'OR(M{HEADER_ROW+1}="Verified Complete",M{HEADER_ROW+1}="Closed")'], fill=FILL_GREEN, font=Font(color=GREEN_FONT)),
    )
    ws.conditional_formatting.add(
        f"M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'OR(M{HEADER_ROW+1}="Open",M{HEADER_ROW+1}="In Progress")'], fill=FILL_YELLOW, font=Font(color=YELLOW_FONT)),
    )
    ws.conditional_formatting.add(
        f"H{HEADER_ROW+1}:H{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'H{HEADER_ROW+1}="Life Safety"'], fill=FILL_RED, font=Font(bold=True, color=RED_FONT)),
    )
    ws.conditional_formatting.add(
        f"H{HEADER_ROW+1}:H{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'H{HEADER_ROW+1}="Major"'], fill=FILL_ORANGE, font=Font(color=ORANGE_FONT)),
    )

    # Summary
    SUMMARY_ROW = HEADER_ROW + DATA_ROWS + 3
    ws.cell(row=SUMMARY_ROW, column=1, value="PUNCH LIST SUMMARY").font = FONT_H2
    summary = [
        ("Total items",        f'=COUNTA(G{HEADER_ROW+1}:G{HEADER_ROW+DATA_ROWS})'),
        ("Open",               f'=COUNTIF(M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS},"Open")+COUNTIF(M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS},"In Progress")'),
        ("Awaiting Verification", f'=COUNTIF(M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS},"Awaiting Verification")'),
        ("Closed",             f'=COUNTIF(M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS},"Verified Complete")+COUNTIF(M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS},"Closed")'),
        ("Life-Safety items",  f'=COUNTIF(H{HEADER_ROW+1}:H{HEADER_ROW+DATA_ROWS},"Life Safety")'),
    ]
    for i, (label, formula) in enumerate(summary, start=SUMMARY_ROW + 1):
        ws.cell(row=i, column=1, value=label).font = FONT_BODY_BOLD
        ws.cell(row=i, column=1).fill = FILL_SUMMARY_LABEL
        ws.cell(row=i, column=1).border = BORDER
        c = ws.cell(row=i, column=2, value=formula)
        c.font = FONT_BIG_NUMBER
        c.number_format = FMT_INT
        c.border = BORDER


# ---------------------------------------------------------------------------
# Tab 8: Meeting Minutes
# ---------------------------------------------------------------------------

MEETING_COLS = [
    ("A", "Meeting #", 10),
    ("B", "Date", 12),
    ("C", "Type", 22),
    ("D", "Attendees", 28),
    ("E", "Topics Discussed", 34),
    ("F", "Decisions", 30),
    ("G", "Action Items", 32),
    ("H", "Action Owners", 22),
    ("I", "Next Meeting", 12),
    ("J", "Notes", 20),
]


def build_meeting_minutes(ws):
    ws.title = "Meeting Minutes"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in MEETING_COLS])

    ws["A1"] = "MEETING MINUTES LOG"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = "Project: =ProjectName    GC: =GCName"
    ws["A2"].font = FONT_BODY_BOLD
    ws.merge_cells("A2:J2")

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(MEETING_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(MEETING_COLS))

    DATA_ROWS = 60
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        ws.cell(row=r, column=1, value=f"=IF(B{r}=\"\",\"\",ROW()-{HEADER_ROW})").number_format = FMT_INT
        ws.cell(row=r, column=2).number_format = FMT_DATE
        ws.cell(row=r, column=9).number_format = FMT_DATE
        apply_body_style(ws, r, len(MEETING_COLS))

    dv_type = DataValidation(
        type="list",
        formula1='"Owner-Architect-GC,Subcontractor Coordination,Safety,Pre-Construction,Pre-Pour,Pre-Erection,Site Walk,Closeout,Other"',
        allow_blank=True,
    )
    dv_type.add(f"C{HEADER_ROW+1}:C{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_type)


# ---------------------------------------------------------------------------
# Tab 9: CSI Reference (hidden)
# ---------------------------------------------------------------------------

def build_csi(ws):
    ws.title = "CSI Reference"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 10), ("B", 60)])
    ws["A1"] = "CSI MasterFormat Divisions"
    ws["A1"].font = FONT_H2
    for i, (code, name) in enumerate(DIVISIONS, start=3):
        ws.cell(row=i, column=1, value=code).font = FONT_BODY
        ws.cell(row=i, column=2, value=name).font = FONT_BODY
    ws.sheet_state = "hidden"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    wb = Workbook()
    build_instructions(wb.active)
    build_project_info(wb.create_sheet())
    build_rfi_log(wb.create_sheet())
    build_submittal_log(wb.create_sheet())
    build_schedule(wb.create_sheet())
    build_dfr_log(wb.create_sheet())
    build_punch_list(wb.create_sheet())
    build_meeting_minutes(wb.create_sheet())
    build_csi(wb.create_sheet())

    out_dir = "/Users/home/charles/contrpro/files/packages/complete/gc"
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "GC_Project_Operations.xlsx")
    wb.save(out_path)
    sz = os.path.getsize(out_path)
    print(f"Wrote {out_path} ({sz//1024} KB)")


if __name__ == "__main__":
    main()
