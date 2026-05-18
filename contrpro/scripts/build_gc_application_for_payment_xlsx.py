#!/usr/bin/env python3
"""
Build ContrPro GC Application for Payment (XLSX) — GC suite (Complete tier).

G-702/G-703 style billing workbook from the GC's side. GC bills the Owner
against the Schedule of Values (SOV) established at award. SOV typically
comes from the GC Bid Estimator's Trade Bids + Self-Perform + General
Conditions roll-up; copy line items into SOV Setup after award.

Mirrors the Universal Sub Suite's Sub_Schedule_of_Values.xlsx with parties
flipped: Contractor → GC, Owner is the payer. Adds G-702 Summary tab (Owner-
facing one-page summary) that the GC delivers with the G-703 Continuation
Sheet each month.

Tabs:
  1. Instructions
  2. Project Info               (project, owner, architect, contract sum)
  3. SOV Setup                  (line items — usually mirrors Trade Bids tab from Bid Estimator)
  4. G-703 Pay App — Current    (Continuation Sheet style)
  5. G-702 Summary — Current    (Application + Certificate for Payment, one page)
  6. Pay App History            (12-month rolling totals)
  7. Stored Materials Log
  8. CSI Reference              (hidden)

Run:
    /Users/home/charles/.venv/bin/python3 \\
        /Users/home/charles/contrpro/scripts/build_gc_application_for_payment_xlsx.py

Output:
    /Users/home/charles/contrpro/files/packages/complete/gc/GC_Application_for_Payment.xlsx
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
ACCENT_GOLD = "C9A227"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
YELLOW_FILL = "FFEB9C"
YELLOW_FONT = "9C5700"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_SUMMARY_LABEL = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW_FILL)
FILL_GOLD = PatternFill("solid", fgColor=ACCENT_GOLD)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_BIG_NUMBER = Font(name="Calibri", size=18, bold=True, color=BRAND_BLUE)
FONT_GREY_ITALIC = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_PCT = "0.0%"
FMT_INT = "0"
FMT_DATE = "yyyy-mm-dd"

DIVISIONS = [
    ("01", "General Requirements"), ("02", "Existing Conditions"), ("03", "Concrete"),
    ("04", "Masonry"), ("05", "Metals"), ("06", "Wood, Plastics, and Composites"),
    ("07", "Thermal and Moisture Protection"), ("08", "Openings"), ("09", "Finishes"),
    ("10", "Specialties"), ("11", "Equipment"), ("12", "Furnishings"),
    ("13", "Special Construction"), ("14", "Conveying Equipment"), ("21", "Fire Suppression"),
    ("22", "Plumbing"), ("23", "HVAC"), ("25", "Integrated Automation"),
    ("26", "Electrical"), ("27", "Communications"), ("28", "Electronic Safety and Security"),
    ("31", "Earthwork"), ("32", "Exterior Improvements"), ("33", "Utilities"),
]


def set_col_widths(ws, widths):
    for col, w in widths:
        ws.column_dimensions[col].width = w


def style_header_row(ws, row, cols, start_col=1):
    for c in range(start_col, start_col + cols):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


def apply_body_style(ws, row, cols, start_col=1):
    for c in range(start_col, start_col + cols):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_BODY
        cell.alignment = Alignment(vertical="center")
        cell.border = BORDER


# ---------------------------------------------------------------------------
# Tab 1: Instructions
# ---------------------------------------------------------------------------

def build_instructions(ws):
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 110)])

    ws["A1"] = "GC APPLICATION FOR PAYMENT — Instructions"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 36

    rows = [
        "",
        "WHAT THIS WORKBOOK IS",
        "Monthly billing workbook for a General Contractor billing the Owner against the prime contract. "
        "Mirrors AIA G-702 (Application + Certificate for Payment, the one-page summary the Owner / "
        "Architect signs) and G-703 (Continuation Sheet, the line-item detail). Use as the GC-side parallel "
        "of the sub's Sub_Schedule_of_Values.xlsx in the Universal Sub Suite.",
        "",
        "ONE-TIME SETUP AT AWARD",
        "1. Project Info: project name, owner, architect, contract date, Original Contract Sum, retainage "
        "rate, billing cycle.",
        "2. SOV Setup: line items. Usually mirror the Trade Bids tab from GC_Bid_Estimator.xlsx (copy/paste "
        "line items + scheduled values). Sum of scheduled values must equal Current Contract Sum on "
        "Project Info (the workbook validates this).",
        "",
        "MONTHLY WORKFLOW",
        "1. Open G-703 Pay App — Current at the end of each billing period.",
        "2. Enter % Complete This Period for each line item (column F). Workbook computes this-period $, "
        "total-to-date $, retainage, and net amount due.",
        "3. Update Stored Materials Log for any billed-but-not-installed materials.",
        "4. Open G-702 Summary — Current. Verify the auto-populated totals.",
        "5. Print BOTH tabs (G-703 + G-702) for transmittal to Architect + Owner.",
        "6. After Owner pays, copy column G (Total To Date) values into the next month's column on Pay App "
        "History.",
        "",
        "RETAINAGE",
        "Default retainage rate is set on Project Info (10% commercial standard; 5% on many state public "
        "works). The workbook supports either of two retention treatments — set 'Retain on stored "
        "materials?' on Project Info to Yes or No. If the contract allows retention reduction at 50% "
        "completion, keep the rate at 10% here and reduce it on the specific pay app where reduction is "
        "requested.",
        "",
        "RELATIONSHIP TO GC BID ESTIMATOR",
        "The SOV here is the line-item dollar values you ultimately accepted from the Bid Estimator's "
        "Trade Bids + Self-Perform + General Conditions roll-up. To set up:",
        "  - Open GC_Bid_Estimator.xlsx",
        "  - Trade Bids tab: pull each Line Total ($) into a row on SOV Setup here",
        "  - Self-Perform tab: pull each Line Total ($) into a row on SOV Setup",
        "  - General Conditions: usually combined into 1-3 SOV lines (Supervision, Field Office, etc.)",
        "  - Add OH + Profit + Bond + Sales Tax as separate SOV lines if Owner requires line-item detail",
        "",
        "CHANGE ORDERS",
        "Approved change orders are added as ADDITIONAL SOV lines (prefixed 'CO #'). Do NOT modify the "
        "Original Contract Sum line items. The Current Contract Sum field on Project Info captures the "
        "running total (Original + COs).",
        "",
        "TROUBLESHOOTING",
        "  - SOV Setup totals don't equal Contract Sum: cross-check each line, confirm no rows are blank "
        "or missing values. The validation banner at the bottom of Project Info will flag this.",
        "  - Retainage on Pay App doesn't match Architect's calc: confirm retainage rate matches the prime "
        "contract, and confirm 'Retain on stored materials?' setting matches.",
        "  - 'Variance' column on SOV Setup shows red (over-billed): you billed more than scheduled for a "
        "line. Reduce billing OR add a change order on SOV Setup for the overage.",
    ]

    for i, text in enumerate(rows, start=2):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if text and text == text.upper() and len(text) < 70:
            cell.font = FONT_H2
        else:
            cell.font = FONT_BODY


# ---------------------------------------------------------------------------
# Tab 2: Project Info
# ---------------------------------------------------------------------------

def build_project_info(ws):
    ws.title = "Project Info"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 38), ("C", 40)])

    ws["B1"] = "PROJECT INFORMATION — GC PAY APP"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:C1")
    ws.row_dimensions[1].height = 32

    info_rows = [
        ("Project Name", ""),
        ("Project No. (GC's)", ""),
        ("Owner", ""),
        ("Owner Address", ""),
        ("Architect", ""),
        ("General Contractor (this firm)", ""),
        ("GC Address", ""),
        ("Prime Contract No.", ""),
        ("Prime Contract Date", ""),
        ("Original Contract Sum", 0),
        ("Approved Change Orders to Date", 0),
        ("Current Contract Sum (Orig + COs)", "=C13+C14"),
        ("Retainage Rate (%)", 0.10),
        ("Retain on Stored Materials?", "No"),
        ("Pay App Submission Day of Month", 25),
        ("Owner Payment Terms (days after Architect cert)", 30),
        ("Project Manager (GC)", ""),
        ("Architect Project Manager", ""),
    ]

    for i, (label, val) in enumerate(info_rows, start=3):
        ws.cell(row=i, column=2, value=label).font = FONT_BODY_BOLD
        ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
        ws.cell(row=i, column=2).border = BORDER
        c = ws.cell(row=i, column=3, value=val)
        c.font = FONT_BODY
        c.border = BORDER
        if "Sum" in label or "Change Orders" in label or "Original" in label:
            c.number_format = FMT_USD
        if "Rate" in label:
            c.number_format = FMT_PCT
        if "Day" in label or "Days" in label or "after" in label:
            c.number_format = FMT_INT
        if "Date" in label:
            c.number_format = FMT_DATE

    # Named ranges
    for name, ref in [
        ("ProjectName", "'Project Info'!$C$3"),
        ("OwnerName", "'Project Info'!$C$5"),
        ("GCName", "'Project Info'!$C$8"),
        ("PrimeContractNo", "'Project Info'!$C$10"),
        ("ContractSum", "'Project Info'!$C$14"),
        ("RetainageRate", "'Project Info'!$C$15"),
        ("RetainStored", "'Project Info'!$C$16"),
    ]:
        ws.parent.defined_names[name] = DefinedName(name=name, attr_text=ref)

    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    dv.add("C16")
    ws.add_data_validation(dv)

    # SOV-validation banner
    ws["B22"] = "SOV TOTAL VALIDATION"
    ws["B22"].font = FONT_H2
    ws["B23"] = "SOV Setup line-item total:"
    ws["B23"].font = FONT_BODY_BOLD
    ws["B23"].fill = FILL_SUMMARY_LABEL
    ws["C23"] = "=SUM('SOV Setup'!E:E)"
    ws["C23"].number_format = FMT_USD
    ws["B24"] = "Current Contract Sum:"
    ws["B24"].font = FONT_BODY_BOLD
    ws["B24"].fill = FILL_SUMMARY_LABEL
    ws["C24"] = "=C14"
    ws["C24"].number_format = FMT_USD
    ws["B25"] = "Variance (should be $0):"
    ws["B25"].font = FONT_BODY_BOLD
    ws["B25"].fill = FILL_SUMMARY_LABEL
    ws["C25"] = "=C23-C24"
    ws["C25"].number_format = FMT_USD
    ws["C25"].font = FONT_BIG_NUMBER

    ws.conditional_formatting.add(
        "C25",
        FormulaRule(formula=["ABS(C25)<0.01"], fill=FILL_GREEN, font=Font(color=GREEN_FONT)),
    )
    ws.conditional_formatting.add(
        "C25",
        FormulaRule(formula=["ABS(C25)>=0.01"], fill=FILL_RED, font=Font(color=RED_FONT)),
    )


# ---------------------------------------------------------------------------
# Tab 3: SOV Setup
# ---------------------------------------------------------------------------

SOV_COLS = [
    ("A", "Line #", 8),
    ("B", "Description", 38),
    ("C", "CSI Div", 9),
    ("D", "CSI Section", 14),
    ("E", "Scheduled Value ($)", 18),
    ("F", "% of Contract", 13),
    ("G", "Billed To Date ($)", 18),
    ("H", "Balance to Finish ($)", 18),
    ("I", "Variance", 13),
    ("J", "Type", 14),
    ("K", "Notes", 26),
]


def build_sov_setup(ws):
    ws.title = "SOV Setup"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in SOV_COLS])

    ws["A1"] = "SOV SETUP — Schedule of Values Line Items"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:K1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "One row per Contract line item. Scheduled Value must sum to Current Contract Sum on Project Info. "
        "Typically copy from GC_Bid_Estimator.xlsx Trade Bids + Self-Perform + General Conditions. Add "
        "change orders as separate rows prefixed 'CO #'."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:K2")

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(SOV_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(SOV_COLS))

    DATA_ROWS = 60
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        ws.cell(row=r, column=1, value=f"=IF(B{r}=\"\",\"\",ROW()-{HEADER_ROW})")
        ws.cell(row=r, column=5).number_format = FMT_USD
        ws.cell(row=r, column=6, value=f"=IF(E{r}=\"\",\"\",E{r}/ContractSum)").number_format = FMT_PCT
        # Billed-to-date pulled from G-703 Pay App
        ws.cell(
            row=r,
            column=7,
            value=f"=IFERROR(VLOOKUP(B{r},'G-703 Pay App'!$B$5:$I$64,8,FALSE),0)",
        ).number_format = FMT_USD
        ws.cell(row=r, column=8, value=f"=IF(E{r}=\"\",\"\",E{r}-G{r})").number_format = FMT_USD
        ws.cell(row=r, column=9, value=f"=IF(E{r}=\"\",\"\",G{r}-E{r})").number_format = FMT_USD
        apply_body_style(ws, r, len(SOV_COLS))

    TOTAL_ROW = HEADER_ROW + DATA_ROWS + 1
    ws.cell(row=TOTAL_ROW, column=2, value="TOTAL").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=2).fill = FILL_SUBHEADER
    for col_idx, col_letter in [(5, "E"), (7, "G"), (8, "H"), (9, "I")]:
        c = ws.cell(
            row=TOTAL_ROW,
            column=col_idx,
            value=f"=SUM({col_letter}{HEADER_ROW+1}:{col_letter}{HEADER_ROW+DATA_ROWS})",
        )
        c.font = FONT_BODY_BOLD
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_USD
        c.border = BORDER

    # CSI dropdown
    csi_codes = ",".join(c for c, _ in DIVISIONS)
    dv_csi = DataValidation(type="list", formula1=f'"{csi_codes}"', allow_blank=True)
    dv_csi.add(f"C{HEADER_ROW+1}:C{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_csi)

    dv_type = DataValidation(
        type="list",
        formula1='"Base Contract,Change Order,Allowance,Bond,Insurance,Tax,General Conditions,Self-Perform,Other"',
        allow_blank=True,
    )
    dv_type.add(f"J{HEADER_ROW+1}:J{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_type)

    # Over-billed flag
    ws.conditional_formatting.add(
        f"I{HEADER_ROW+1}:I{HEADER_ROW+DATA_ROWS}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=FILL_RED, font=Font(color=RED_FONT)),
    )


# ---------------------------------------------------------------------------
# Tab 4: G-703 Pay App — Current Period
# ---------------------------------------------------------------------------

G703_COLS = [
    ("A", "Line #", 7),
    ("B", "Description", 32),
    ("C", "CSI Div", 8),
    ("D", "Scheduled Value", 15),
    ("E", "Previous Periods ($)", 16),
    ("F", "% Complete This Period", 14),
    ("G", "This Period ($)", 14),
    ("H", "Total Completed To Date ($)", 18),
    ("I", "Materials Stored ($)", 14),
    ("J", "Total + Stored ($)", 15),
    ("K", "% Complete", 11),
    ("L", "Retainage ($)", 13),
    ("M", "Balance to Finish ($)", 16),
    ("N", "This Period Net Due ($)", 16),
]


def build_g703(ws):
    ws.title = "G-703 Pay App"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in G703_COLS])

    ws["A1"] = "G-703 — Continuation Sheet (Pay App Detail)"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:N1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = "Project: =ProjectName"
    ws["A2"].font = FONT_BODY_BOLD
    ws.merge_cells("A2:F2")
    ws["G2"] = "Pay App #:"
    ws["G2"].font = FONT_BODY_BOLD
    ws["H2"] = 1
    ws["H2"].number_format = FMT_INT
    ws["I2"] = "Period Ending:"
    ws["I2"].font = FONT_BODY_BOLD
    ws["J2"] = ""
    ws["J2"].number_format = FMT_DATE
    ws.parent.defined_names["PayAppNum"] = DefinedName(name="PayAppNum", attr_text=f"'G-703 Pay App'!$H$2")
    ws.parent.defined_names["PeriodEnding"] = DefinedName(name="PeriodEnding", attr_text=f"'G-703 Pay App'!$J$2")

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(G703_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(G703_COLS))

    DATA_ROWS = 60
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        sov_row = r - HEADER_ROW + 4  # match SOV Setup row
        # Pull from SOV Setup
        ws.cell(row=r, column=1, value=f"=IF('SOV Setup'!A{sov_row}=\"\",\"\",'SOV Setup'!A{sov_row})")
        ws.cell(row=r, column=2, value=f"='SOV Setup'!B{sov_row}")
        ws.cell(row=r, column=3, value=f"='SOV Setup'!C{sov_row}")
        ws.cell(row=r, column=4, value=f"=IF('SOV Setup'!E{sov_row}=\"\",\"\",'SOV Setup'!E{sov_row})").number_format = FMT_USD
        # E = Previous period total (manual entry first period; carried forward)
        ws.cell(row=r, column=5, value=0).number_format = FMT_USD
        # F = % complete this period (manual)
        ws.cell(row=r, column=6, value=0).number_format = FMT_PCT
        # G = This period $ = Scheduled × F
        ws.cell(row=r, column=7, value=f"=IF(D{r}=\"\",\"\",D{r}*F{r})").number_format = FMT_USD
        # H = Total completed to date = E + G
        ws.cell(row=r, column=8, value=f"=IF(D{r}=\"\",\"\",E{r}+G{r})").number_format = FMT_USD
        # I = Materials stored (manual)
        ws.cell(row=r, column=9, value=0).number_format = FMT_USD
        # J = Total + stored
        ws.cell(row=r, column=10, value=f"=IF(D{r}=\"\",\"\",H{r}+I{r})").number_format = FMT_USD
        # K = % complete
        ws.cell(row=r, column=11, value=f"=IF(OR(D{r}=\"\",D{r}=0),\"\",J{r}/D{r})").number_format = FMT_PCT
        # L = Retainage (depends on RetainStored setting)
        ws.cell(
            row=r,
            column=12,
            value=(
                f'=IF(D{r}="","",'
                f'IF(RetainStored="Yes",J{r}*RetainageRate,H{r}*RetainageRate))'
            ),
        ).number_format = FMT_USD
        # M = Balance to finish
        ws.cell(row=r, column=13, value=f"=IF(D{r}=\"\",\"\",D{r}-J{r})").number_format = FMT_USD
        # N = This period net due = (J - E - retainage_delta)
        ws.cell(
            row=r,
            column=14,
            value=(
                f'=IF(D{r}="","",'
                f'(J{r}-E{r})-(L{r}-E{r}*RetainageRate))'
            ),
        ).number_format = FMT_USD
        apply_body_style(ws, r, len(G703_COLS))

    TOTAL_ROW = HEADER_ROW + DATA_ROWS + 1
    ws.cell(row=TOTAL_ROW, column=2, value="GRAND TOTAL").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=2).fill = FILL_SUBHEADER
    for col_idx, col_letter in [(4, "D"), (5, "E"), (7, "G"), (8, "H"), (9, "I"), (10, "J"), (12, "L"), (13, "M"), (14, "N")]:
        c = ws.cell(
            row=TOTAL_ROW,
            column=col_idx,
            value=f"=SUM({col_letter}{HEADER_ROW+1}:{col_letter}{HEADER_ROW+DATA_ROWS})",
        )
        c.font = FONT_BODY_BOLD
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_USD
        c.border = BORDER

    # Named ranges for G-702 Summary
    ws.parent.defined_names["G703Total_Scheduled"] = DefinedName(name="G703Total_Scheduled", attr_text=f"'G-703 Pay App'!$D${TOTAL_ROW}")
    ws.parent.defined_names["G703Total_TotalStored"] = DefinedName(name="G703Total_TotalStored", attr_text=f"'G-703 Pay App'!$J${TOTAL_ROW}")
    ws.parent.defined_names["G703Total_Retainage"] = DefinedName(name="G703Total_Retainage", attr_text=f"'G-703 Pay App'!$L${TOTAL_ROW}")
    ws.parent.defined_names["G703Total_Previous"] = DefinedName(name="G703Total_Previous", attr_text=f"'G-703 Pay App'!$E${TOTAL_ROW}")
    ws.parent.defined_names["G703Total_NetDue"] = DefinedName(name="G703Total_NetDue", attr_text=f"'G-703 Pay App'!$N${TOTAL_ROW}")

    # Conditional formatting
    ws.conditional_formatting.add(
        f"K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS}",
        CellIsRule(operator="greaterThanOrEqual", formula=["1"], fill=FILL_GREEN, font=Font(color=GREEN_FONT)),
    )
    ws.conditional_formatting.add(
        f"M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS}",
        CellIsRule(operator="lessThan", formula=["0"], fill=FILL_RED, font=Font(color=RED_FONT)),
    )


# ---------------------------------------------------------------------------
# Tab 5: G-702 Summary
# ---------------------------------------------------------------------------

def build_g702(ws):
    ws.title = "G-702 Summary"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 50), ("C", 22)])

    ws["B1"] = "G-702 — APPLICATION & CERTIFICATE FOR PAYMENT"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:C1")
    ws.row_dimensions[1].height = 38

    ws["B2"] = (
        "One-page Owner-facing summary. Prints with G-703 Pay App as the cover sheet. "
        "Architect signs the bottom to certify the amount due."
    )
    ws["B2"].font = FONT_GREY_ITALIC
    ws.merge_cells("B2:C2")

    rows = [
        ("PROJECT", ""),
        ("Project Name", "=ProjectName"),
        ("To Owner", "=OwnerName"),
        ("From Contractor", "=GCName"),
        ("Contract No.", "=PrimeContractNo"),
        ("Pay App No.", "=PayAppNum"),
        ("Period Ending", "=PeriodEnding"),
        ("", ""),
        ("CONTRACT SUMMARY", ""),
        ("1. Original Contract Sum", "='Project Info'!C12"),
        ("2. Net Change by Change Orders", "='Project Info'!C13"),
        ("3. Contract Sum to Date (1+2)", "=ContractSum"),
        ("4. Total Completed & Stored to Date", "=G703Total_TotalStored"),
        ("5. Retainage", "=G703Total_Retainage"),
        ("6. Total Earned Less Retainage (4-5)", "=G703Total_TotalStored-G703Total_Retainage"),
        ("7. Less Previous Certificates for Payment", "=G703Total_Previous-G703Total_Previous*RetainageRate"),
        ("8. CURRENT PAYMENT DUE (6-7)", "=G703Total_NetDue"),
        ("9. Balance to Finish, Plus Retainage", "=ContractSum-G703Total_TotalStored+G703Total_Retainage"),
        ("", ""),
        ("CONTRACTOR CERTIFICATION", ""),
        ("The undersigned Contractor certifies that to the best of the Contractor's knowledge, information, and belief, the Work covered by this Application has been completed in accordance with the Contract Documents, that all amounts have been paid by the Contractor for Work for which previous Certificates for Payment were issued and payments received from the Owner, and that current payment shown herein is now due.", ""),
        ("", ""),
        ("Contractor:", ""),
        ("By: ____________________________   Title: ____________________   Date: __________", ""),
        ("", ""),
        ("ARCHITECT'S CERTIFICATE", ""),
        ("In accordance with the Contract Documents, based on on-site observations and the data comprising this application, the Architect certifies to the Owner that the Work has progressed as indicated, the quality of the Work is in accordance with the Contract Documents, and the Contractor is entitled to payment of the AMOUNT CERTIFIED.", ""),
        ("", ""),
        ("AMOUNT CERTIFIED:  $ ____________________________", ""),
        ("Architect:", ""),
        ("By: ____________________________   Date: __________", ""),
    ]

    start_row = 4
    for i, (label, val) in enumerate(rows, start=start_row):
        is_section = label and label == label.upper() and not val and len(label) < 50
        if is_section:
            ws.cell(row=i, column=2, value=label).font = FONT_H2
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
            continue
        if label:
            cell = ws.cell(row=i, column=2, value=label)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = FONT_BODY_BOLD if label.startswith(("1.", "2.", "3.", "4.", "5.", "6.", "7.", "8.", "9.", "AMOUNT", "By:")) else FONT_BODY
            if val and not val.startswith("="):
                ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
            else:
                ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL if val else PatternFill()
            if val:
                c = ws.cell(row=i, column=3, value=val)
                c.font = FONT_BIG_NUMBER if "CURRENT PAYMENT" in label else FONT_BODY
                if "1." in label or "2." in label or "3." in label or "4." in label or "5." in label or "6." in label or "7." in label or "8." in label or "9." in label:
                    c.number_format = FMT_USD
                if "Period" in label:
                    c.number_format = FMT_DATE
                if "Pay App No" in label:
                    c.number_format = FMT_INT


# ---------------------------------------------------------------------------
# Tab 6: Pay App History (12-month roll-up)
# ---------------------------------------------------------------------------

def build_history(ws):
    ws.title = "Pay App History"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "PAY APP HISTORY — 12-Month Roll-Up"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:O1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = "After each pay app is paid, transfer the Total + Stored values from the G-703 tab into the next column below."
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:O2")

    HEADER_ROW = 4
    set_col_widths(ws, [
        ("A", 8), ("B", 36), ("C", 16),
        *[(get_column_letter(i), 14) for i in range(4, 16)],
        ("P", 16),
    ])

    headers = ["Line #", "Description", "Scheduled Value"] + [f"Period {m}" for m in range(1, 13)] + ["Final To Date"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=h)
    style_header_row(ws, HEADER_ROW, len(headers))

    DATA_ROWS = 60
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        sov_row = r - HEADER_ROW + 4
        ws.cell(row=r, column=1, value=f"=IF('SOV Setup'!A{sov_row}=\"\",\"\",'SOV Setup'!A{sov_row})")
        ws.cell(row=r, column=2, value=f"='SOV Setup'!B{sov_row}")
        ws.cell(row=r, column=3, value=f"=IF('SOV Setup'!E{sov_row}=\"\",\"\",'SOV Setup'!E{sov_row})").number_format = FMT_USD
        for m in range(1, 13):
            ws.cell(row=r, column=3 + m).number_format = FMT_USD
        ws.cell(row=r, column=16, value=f"=IF(C{r}=\"\",\"\",MAX(D{r}:O{r}))").number_format = FMT_USD
        apply_body_style(ws, r, 16)

    TOTAL_ROW = HEADER_ROW + DATA_ROWS + 1
    ws.cell(row=TOTAL_ROW, column=2, value="TOTAL").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=2).fill = FILL_SUBHEADER
    for col_idx in [3, *list(range(4, 17))]:
        col_letter = get_column_letter(col_idx)
        c = ws.cell(
            row=TOTAL_ROW,
            column=col_idx,
            value=f"=SUM({col_letter}{HEADER_ROW+1}:{col_letter}{HEADER_ROW+DATA_ROWS})",
        )
        c.font = FONT_BODY_BOLD
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_USD
        c.border = BORDER


# ---------------------------------------------------------------------------
# Tab 7: Stored Materials Log
# ---------------------------------------------------------------------------

STORED_COLS = [
    ("A", "Date Added", 12),
    ("B", "SOV Line Ref", 12),
    ("C", "Description", 30),
    ("D", "Supplier", 22),
    ("E", "PO #", 14),
    ("F", "Quantity", 10),
    ("G", "Unit", 8),
    ("H", "Unit Cost", 12),
    ("I", "Total Value ($)", 14),
    ("J", "Storage Location", 22),
    ("K", "Insured?", 10),
    ("L", "Supplier Lien Waiver?", 14),
    ("M", "Date Installed", 12),
    ("N", "Status", 14),
]


def build_stored(ws):
    ws.title = "Stored Materials"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in STORED_COLS])

    ws["A1"] = "STORED MATERIALS LOG"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:N1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = "Materials billed but not yet installed. Architect typically requires evidence of storage, insurance, segregation, and supplier lien waiver before approving stored materials in the pay app."
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:N2")

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(STORED_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(STORED_COLS))

    DATA_ROWS = 50
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        ws.cell(row=r, column=8).number_format = FMT_USD
        ws.cell(row=r, column=9, value=f"=IF(OR(F{r}=\"\",H{r}=\"\"),\"\",F{r}*H{r})").number_format = FMT_USD
        ws.cell(row=r, column=1).number_format = FMT_DATE
        ws.cell(row=r, column=13).number_format = FMT_DATE
        apply_body_style(ws, r, len(STORED_COLS))

    dv_yn = DataValidation(type="list", formula1='"Yes,No,Pending"', allow_blank=True)
    dv_yn.add(f"K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS}")
    dv_yn.add(f"L{HEADER_ROW+1}:L{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_yn)

    dv_status = DataValidation(
        type="list",
        formula1='"On Site,Off Site Bonded Warehouse,In Transit,Installed,Returned"',
        allow_blank=True,
    )
    dv_status.add(f"N{HEADER_ROW+1}:N{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_status)

    TOTAL_ROW = HEADER_ROW + DATA_ROWS + 1
    ws.cell(row=TOTAL_ROW, column=2, value="TOTAL").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=2).fill = FILL_SUBHEADER
    c = ws.cell(
        row=TOTAL_ROW,
        column=9,
        value=f"=SUM(I{HEADER_ROW+1}:I{HEADER_ROW+DATA_ROWS})",
    )
    c.font = FONT_BODY_BOLD
    c.fill = FILL_SUBHEADER
    c.number_format = FMT_USD
    c.border = BORDER


# ---------------------------------------------------------------------------
# Tab 8: CSI Reference (hidden)
# ---------------------------------------------------------------------------

def build_csi(ws):
    ws.title = "CSI Reference"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 10), ("B", 60)])
    ws["A1"] = "CSI MasterFormat Divisions"
    ws["A1"].font = FONT_H2
    for i, (code, name) in enumerate(DIVISIONS, start=3):
        ws.cell(row=i, column=1, value=code).font = FONT_BODY
        ws.cell(row=i, column=2, value=name).font = FONT_BODY
    ws.sheet_state = "hidden"


def main():
    wb = Workbook()
    build_instructions(wb.active)
    build_project_info(wb.create_sheet())
    build_sov_setup(wb.create_sheet())
    build_g703(wb.create_sheet())
    build_g702(wb.create_sheet())
    build_history(wb.create_sheet())
    build_stored(wb.create_sheet())
    build_csi(wb.create_sheet())

    out_dir = "/Users/home/charles/contrpro/files/packages/complete/gc"
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "GC_Application_for_Payment.xlsx")
    wb.save(out_path)
    sz = os.path.getsize(out_path)
    print(f"Wrote {out_path} ({sz//1024} KB)")


if __name__ == "__main__":
    main()
