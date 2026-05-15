#!/usr/bin/env python3
"""
Build ContrPro Change Order Log (XLSX).

Produces a production-grade workbook with:
  - Instructions tab (CO vs scope creep vs RFI, T&M, retainage, statute of limitations)
  - Project Info tab (contract anchor for Dashboard math)
  - Change Order Log tab (CSI MasterFormat coded, dropdowns, formulas, conditional formatting)
  - CO Summary by Division (SUMIFS roll-ups)
  - Dashboard (totals, contract impact %, status counts)
  - CSI Reference (hidden divisions list used by Named Range / dropdowns)

Run:
    /Users/home/charles/.venv/bin/python3 \
        /Users/home/charles/contrpro/scripts/build_change_order_log_xlsx.py

Output:
    /Users/home/charles/contrpro/files/packages/business/Change_Order_Log.xlsx
"""

from __future__ import annotations

import os
from typing import Tuple

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# ---------------------------------------------------------------------------
# Brand & styling (matches Job_Costing_Spreadsheet for visual consistency)
# ---------------------------------------------------------------------------

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
ACCENT_GOLD = "C9A227"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
YELLOW_FILL = "FFEB9C"
YELLOW_FONT = "9C5700"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_SUMMARY_LABEL = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW_FILL)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H1 = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_BIG_NUMBER = Font(name="Calibri", size=18, bold=True, color=BRAND_BLUE)
FONT_GREY_ITALIC = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_USD_BIG = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
FMT_PCT = "0.0%"
FMT_INT = "0"
FMT_DATE = "yyyy-mm-dd"


# ---------------------------------------------------------------------------
# CSI MasterFormat division reference (same source as Job Costing tracker)
# Full 01-49 (only canonical, in-use divisions; gaps reflect MasterFormat reality)
# ---------------------------------------------------------------------------

DIVISIONS: list[tuple[str, str]] = [
    ("01", "General Requirements"),
    ("02", "Existing Conditions"),
    ("03", "Concrete"),
    ("04", "Masonry"),
    ("05", "Metals"),
    ("06", "Wood, Plastics, and Composites"),
    ("07", "Thermal and Moisture Protection"),
    ("08", "Openings"),
    ("09", "Finishes"),
    ("10", "Specialties"),
    ("11", "Equipment"),
    ("12", "Furnishings"),
    ("13", "Special Construction"),
    ("14", "Conveying Equipment"),
    ("21", "Fire Suppression"),
    ("22", "Plumbing"),
    ("23", "HVAC"),
    ("25", "Integrated Automation"),
    ("26", "Electrical"),
    ("27", "Communications"),
    ("28", "Electronic Safety and Security"),
    ("31", "Earthwork"),
    ("32", "Exterior Improvements"),
    ("33", "Utilities"),
    ("34", "Transportation"),
    ("35", "Waterway and Marine Construction"),
    ("40", "Process Interconnections"),
    ("41", "Material Processing and Handling Equipment"),
    ("42", "Process Heating, Cooling, and Drying Equipment"),
    ("43", "Process Gas and Liquid Handling, Purification, and Storage Equipment"),
    ("44", "Pollution and Waste Control Equipment"),
    ("45", "Industry-Specific Manufacturing Equipment"),
    ("46", "Water and Wastewater Equipment"),
    ("48", "Electrical Power Generation"),
    ("49", "Reserved"),
]
DIV_LOOKUP: dict[str, str] = {d: n for d, n in DIVISIONS}
DIV_CODES_ORDERED = [d for d, _ in DIVISIONS]


# ---------------------------------------------------------------------------
# Pre-populated example change orders (varied types/statuses for showcase)
# (co_num, date_issued, date_approved, project, div, csi_code,
#  description, reason, co_type, amount, sched_days, status, approved_by, notes)
# ---------------------------------------------------------------------------

EXAMPLE_COS: list[tuple] = [
    ("CO-001", "2026-03-12", "2026-03-18",
     "Riverside Office Bldg",
     "03", "03 30 00",
     "Add 12 CY structural concrete at column footings F-4 to F-7 per RFI-024.",
     "Field Condition", "Lump Sum",
     8450.00, 2, "Approved",
     "K. Holmes (Owner Rep)",
     "Soft soil discovered at depth; required oversized footings."),
    ("CO-002", "2026-03-20", "2026-03-25",
     "Riverside Office Bldg",
     "23", "23 30 00",
     "Upgrade VAV boxes Level 2 from std to digital controls.",
     "Owner Request", "Lump Sum",
     14275.00, 0, "Approved",
     "K. Holmes (Owner Rep)",
     "Owner spec upgrade after value-engineering pass."),
    ("CO-003", "2026-04-02", "",
     "Riverside Office Bldg",
     "09", "09 60 00",
     "Substitute LVT flooring for spec'd carpet tile in open office.",
     "Owner Request", "Deductive",
     -6320.00, 0, "Pending",
     "",
     "Awaiting owner sign-off on LVT sample."),
    ("CO-004", "2026-04-10", "2026-04-15",
     "Riverside Office Bldg",
     "26", "26 50 00",
     "Add emergency lighting at stair towers per code review.",
     "Code Change", "T&M",
     3940.00, 1, "Approved",
     "K. Holmes (Owner Rep)",
     "Updated IBC interpretation from building dept."),
    ("CO-005", "2026-04-22", "",
     "Riverside Office Bldg",
     "07", "07 50 00",
     "Replace damaged membrane after wind event (40 sq).",
     "Force Majeure", "Unit Price",
     11200.00, 5, "Submitted",
     "",
     "Wind event 2026-04-19; insurance claim filed in parallel."),
]
EMPTY_CO_ROWS = 20
EMPTY_PROJECT_ROWS = 5

REASONS = [
    "Owner Request", "Design Change", "Field Condition",
    "Code Change", "Error/Omission", "Force Majeure", "Other",
]
CO_TYPES = ["Lump Sum", "T&M", "Unit Price", "Allowance", "Deductive"]
STATUSES = ["Submitted", "Approved", "Rejected", "Pending", "Disputed", "Voided"]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def set_col_widths(ws, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_header_row(ws, row: int, last_col: int) -> None:
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 28


def border_range(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER


# ---------------------------------------------------------------------------
# Hidden tab — CSI Reference (also feeds named ranges + dropdowns)
# ---------------------------------------------------------------------------

def build_csi_reference(wb: Workbook) -> str:
    ws = wb.create_sheet("CSI Reference")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "CSI MasterFormat Division Reference"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:B1")

    ws["A2"] = "Pulled from the same canonical list as ContrPro Job Costing. Do not edit."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)

    ws.cell(row=3, column=1, value="CSI Division")
    ws.cell(row=3, column=2, value="Division Name")
    style_header_row(ws, 3, 2)

    for i, (code, name) in enumerate(DIVISIONS, start=4):
        ws.cell(row=i, column=1, value=code).number_format = "@"
        ws.cell(row=i, column=2, value=name)
        for c in (1, 2):
            ws.cell(row=i, column=c).border = BORDER
            ws.cell(row=i, column=c).font = FONT_BODY

    set_col_widths(ws, {"A": 14, "B": 60})

    # Named range so VLOOKUPs and dropdowns can pull cleanly across sheets.
    first = 4
    last = 4 + len(DIVISIONS) - 1
    div_range = f"'CSI Reference'!$A${first}:$A${last}"
    full_range = f"'CSI Reference'!$A${first}:$B${last}"
    wb.defined_names["CSI_Divisions"] = DefinedName("CSI_Divisions", attr_text=div_range)
    wb.defined_names["CSI_Table"] = DefinedName("CSI_Table", attr_text=full_range)
    return ws.title


# ---------------------------------------------------------------------------
# Tab 1 — Instructions
# ---------------------------------------------------------------------------

def build_instructions(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {"A": 4, "B": 110})

    ws["B2"] = "ContrPro - Change Order Log"
    ws["B2"].font = FONT_TITLE
    ws.row_dimensions[2].height = 32

    ws["B3"] = "CSI MasterFormat-coded log of every change to a construction contract - the receipt drawer for the job."
    ws["B3"].font = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

    sections: list[tuple[str, list[str]]] = [
        (
            "What this workbook is",
            [
                "Six tabs work together: Project Info pins the contract, Change Order Log is where every CO lives, "
                "CO Summary by Division rolls them up by trade, and Dashboard shows contract impact at a glance. "
                "CSI Reference is a hidden-ish lookup list - leave it alone.",
                "Every row in Change Order Log is one paper trail: written description, dollar amount, schedule impact, "
                "and approval status. If it is not in this log, it did not happen.",
            ],
        ),
        (
            "Change Order vs Scope Creep vs RFI - know the difference",
            [
                "Change Order (CO) - A WRITTEN amendment to the contract that adjusts price, schedule, or scope. "
                "It is bilateral (owner and GC both sign) and binding. This log only tracks COs.",
                "Scope Creep - Owner or architect asks for 'just one more thing' verbally and you do it. "
                "DO NOT track scope creep here, track it as a CO the moment you sense it. Verbal is not a CO; "
                "verbal is unpaid work waiting to happen.",
                "RFI (Request for Information) - A question to the design team. Not a CO. But the answer often "
                "TRIGGERS a CO when the response changes scope or quantities. Log the resulting CO here and reference "
                "the RFI number in Notes.",
            ],
        ),
        (
            "When to issue a formal CO vs a T&M ticket",
            [
                "Formal CO (Lump Sum / Unit Price / Allowance / Deductive) - Use when scope is defined enough to "
                "price up-front. Get it signed BEFORE the work starts. This is the cleanest path and protects your "
                "margin.",
                "T&M (Time and Material) ticket - Use when scope is unclear or urgent (emergency repair, hidden "
                "condition). Track labor hours and material receipts daily, get signed T&M tickets each day, and roll "
                "them into a CO weekly. Type = T&M on the row.",
                "Rule of thumb: if you cannot say what the work is in one sentence, it is T&M until you can.",
            ],
        ),
        (
            "The CO Log and retainage discussions",
            [
                "Retainage (typically 5-10% withheld from pay apps) gets released at substantial completion. Owners "
                "will scrutinize every CO before cutting that final retainage check.",
                "A clean CO Log - with descriptions, reasons, signed approvals, and CSI coding - is what gets retainage "
                "released on time. A messy log is what funds 90-day retainage disputes.",
                "Before requesting retainage release, sort this tab by Status, confirm every Approved row has an "
                "approver name and date, and resolve every Pending/Disputed row.",
            ],
        ),
        (
            "Statute of Limitations awareness",
            [
                "Most U.S. states impose a 4-year statute of limitations on construction defect claims for breach of "
                "contract (some go to 6 or 10 - check your state). Some states also have a statute of repose (often "
                "10-12 years) that caps liability regardless of when a defect is discovered.",
                "What this means for the log: keep this file for AT LEAST the full statute of repose window after "
                "substantial completion. Cloud backup it. PDF the final version with all approvals and store it with "
                "the contract.",
                "This workbook is a record-keeping tool, not legal advice. Run anything that smells like a dispute "
                "past your construction attorney.",
            ],
        ),
        (
            "Tab-by-tab quick start",
            [
                "1. Project Info - fill in Project Name, Number, Client, Original Contract Amount, PM, and "
                "Architect/Engineer on row 4. The Dashboard reads Original Contract Amount from D4.",
                "2. Change Order Log - one row per CO. Fill in CO Number, dates, project, CSI Division (dropdown), "
                "CSI Code (6-digit text, e.g. 23 30 00), Division Name (auto-filled from VLOOKUP), description, "
                "reason, type, amount, schedule impact, status, approver.",
                "3. CO Summary by Division - DO NOT edit. It rolls up your CO Log automatically.",
                "4. Dashboard - DO NOT edit. Read-only summary of contract impact and CO counts by status.",
            ],
        ),
        (
            "CSI coding rules (must follow)",
            [
                "CSI Division: 2-digit text (01-49). Pick from the dropdown - this preserves leading zeros and matches "
                "the canonical list.",
                "CSI Code: 6-digit space-separated, e.g. 23 30 00 for HVAC Air Distribution. You can extend to "
                "Level 3 (e.g. 23 31 13) for finer detail.",
                "Division Name auto-fills from a VLOOKUP against CSI Reference. If it shows #N/A, your CSI Division "
                "value does not match the reference list - re-pick from the dropdown.",
            ],
        ),
        (
            "Color coding cheat sheet",
            [
                "Amount column: green = positive (you're getting paid more), red = negative (deductive CO).",
                "Status column: green = Approved, yellow = Pending or Submitted, red = Rejected or Disputed.",
                "Dashboard CO % of Contract: highlighted red when total approved COs exceed 15% of original "
                "contract - that's the threshold where most owners start asking hard questions and lenders may "
                "trigger a review.",
            ],
        ),
        (
            "Weekly cadence",
            [
                "Monday: review Pending and Submitted rows. Anything older than 14 days gets a follow-up email today.",
                "Friday: log every CO issued this week. No CO leaves the trailer unrecorded.",
                "End of month: reconcile CO totals against pay app G-703 line items. They must match - this is what "
                "your owner's accountant will check.",
            ],
        ),
        (
            "Tabs in this workbook",
            [
                "Instructions  -  this tab",
                "Project Info  -  contract anchor (Dashboard reads Original Contract Amount from here)",
                "Change Order Log  -  CSI-coded line items, the main worksheet",
                "CO Summary by Division  -  SUMIFS roll-up by division and status",
                "Dashboard  -  contract impact, CO counts, schedule impact",
                "CSI Reference  -  canonical division list (do not edit)",
            ],
        ),
    ]

    row = 5
    for heading, paragraphs in sections:
        ws.cell(row=row, column=2, value=heading)
        ws.cell(row=row, column=2).font = FONT_H2
        ws.cell(row=row, column=2).fill = FILL_SUBHEADER
        ws.cell(row=row, column=2).alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[row].height = 22
        row += 1
        for para in paragraphs:
            ws.cell(row=row, column=2, value=para)
            ws.cell(row=row, column=2).font = FONT_BODY
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = max(18, 16 * (1 + len(para) // 110))
            row += 1
        row += 1

    ws.cell(row=row + 1, column=2,
            value="ContrPro - built for builders. Questions: support@contrpro.com")
    ws.cell(row=row + 1, column=2).font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)


# ---------------------------------------------------------------------------
# Tab 2 — Project Info
# ---------------------------------------------------------------------------

PROJECT_HEADERS = [
    "Project Name", "Project Number", "Client",
    "Original Contract Amount", "Project Manager",
    "Architect / Engineer", "Notes",
]

PROJECT_EXAMPLE = [
    "EXAMPLE - Riverside Office Bldg",
    "2026-014",
    "Riverside Holdings LLC",
    1850000,
    "Sample PM",
    "Apex Architects + WSP Structural",
    "Example row - replace with your project.",
]


def build_project_info(wb: Workbook) -> None:
    ws = wb.create_sheet("Project Info")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Project Information"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:G1")

    for i, h in enumerate(PROJECT_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(PROJECT_HEADERS))

    # Example row (greyed/italic)
    for i, v in enumerate(PROJECT_EXAMPLE, start=1):
        cell = ws.cell(row=4, column=i, value=v)
        cell.font = FONT_GREY_ITALIC
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.cell(row=4, column=4).number_format = FMT_USD

    # Empty user rows
    for r in range(5, 5 + EMPTY_PROJECT_ROWS):
        for c in range(1, len(PROJECT_HEADERS) + 1):
            ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=4).number_format = FMT_USD
        ws.row_dimensions[r].height = 22

    set_col_widths(ws, {
        "A": 30, "B": 16, "C": 28, "D": 22,
        "E": 22, "F": 30, "G": 45,
    })

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Tab 3 — Change Order Log (main worksheet)
# ---------------------------------------------------------------------------

CO_HEADERS = [
    "CO Number", "Date Issued", "Date Approved",
    "Project / Job", "CSI Division", "CSI Code", "Division Name",
    "Description of Change", "Reason", "Type",
    "Amount", "Schedule Impact (days)", "Status",
    "Approved By", "Notes",
]


def build_change_order_log(wb: Workbook) -> Tuple[str, int, int]:
    ws = wb.create_sheet("Change Order Log")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Change Order Log (CSI MasterFormat)"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(CO_HEADERS))

    # Header row at row 3
    for i, h in enumerate(CO_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(CO_HEADERS))

    data_start = 4
    n_pre = len(EXAMPLE_COS)
    n_total = n_pre + EMPTY_CO_ROWS
    data_end = data_start + n_total - 1

    # Column letters reference
    COL = {h: get_column_letter(i + 1) for i, h in enumerate(CO_HEADERS)}
    col_div = COL["CSI Division"]          # E
    col_divname = COL["Division Name"]      # G

    # Pre-populated example rows
    for idx, row_data in enumerate(EXAMPLE_COS):
        r = data_start + idx
        (co_num, d_issued, d_approved, project, div, csi_code,
         desc, reason, co_type, amount, sched_days,
         status, approved_by, notes) = row_data

        ws.cell(row=r, column=1, value=co_num)
        ws.cell(row=r, column=2, value=d_issued)
        ws.cell(row=r, column=3, value=d_approved if d_approved else None)
        ws.cell(row=r, column=4, value=project)
        ws.cell(row=r, column=5, value=div)
        ws.cell(row=r, column=6, value=csi_code)
        # Division Name auto-filled via VLOOKUP against CSI_Table named range
        ws.cell(
            row=r, column=7,
            value=f'=IFERROR(VLOOKUP({col_div}{r},CSI_Table,2,FALSE),"")',
        )
        ws.cell(row=r, column=8, value=desc)
        ws.cell(row=r, column=9, value=reason)
        ws.cell(row=r, column=10, value=co_type)
        ws.cell(row=r, column=11, value=amount)
        ws.cell(row=r, column=12, value=sched_days)
        ws.cell(row=r, column=13, value=status)
        ws.cell(row=r, column=14, value=approved_by if approved_by else None)
        ws.cell(row=r, column=15, value=notes)

    # Empty user-fillable rows (VLOOKUP pre-seeded)
    for i in range(EMPTY_CO_ROWS):
        r = data_start + n_pre + i
        ws.cell(
            row=r, column=7,
            value=f'=IFERROR(VLOOKUP({col_div}{r},CSI_Table,2,FALSE),"")',
        )

    # Per-row formatting
    for r in range(data_start, data_end + 1):
        for c in range(1, len(CO_HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(c in (4, 7, 8, 15)),
            )
        # CSI Division and CSI Code as text so leading zeros stick
        ws.cell(row=r, column=5).number_format = "@"
        ws.cell(row=r, column=6).number_format = "@"
        # Dates
        ws.cell(row=r, column=2).number_format = FMT_DATE
        ws.cell(row=r, column=3).number_format = FMT_DATE
        # Amount as USD
        ws.cell(row=r, column=11).number_format = FMT_USD
        # Schedule impact as integer
        ws.cell(row=r, column=12).number_format = FMT_INT
        ws.row_dimensions[r].height = 30

    # --- Conditional formatting ---

    # Amount column (K): green if >0, red if <0
    amt_range = f"K{data_start}:K{data_end}"
    ws.conditional_formatting.add(
        amt_range,
        CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=False,
                   fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True)),
    )
    ws.conditional_formatting.add(
        amt_range,
        CellIsRule(operator="lessThan", formula=["0"], stopIfTrue=False,
                   fill=FILL_RED, font=Font(color=RED_FONT, bold=True)),
    )

    # Status column (M): green=Approved, yellow=Pending/Submitted, red=Rejected/Disputed
    status_range = f"M{data_start}:M{data_end}"
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(formula=[f'$M{data_start}="Approved"'], stopIfTrue=False,
                    fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True)),
    )
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(formula=[f'OR($M{data_start}="Pending",$M{data_start}="Submitted")'],
                    stopIfTrue=False,
                    fill=FILL_YELLOW, font=Font(color=YELLOW_FONT, bold=True)),
    )
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(formula=[f'OR($M{data_start}="Rejected",$M{data_start}="Disputed")'],
                    stopIfTrue=False,
                    fill=FILL_RED, font=Font(color=RED_FONT, bold=True)),
    )

    # --- Data validation ---

    # CSI Division dropdown (uses named range)
    dv_div = DataValidation(
        type="list", formula1="=CSI_Divisions", allow_blank=True,
    )
    dv_div.error = "Pick a CSI Division from the dropdown."
    dv_div.errorTitle = "Invalid CSI Division"
    ws.add_data_validation(dv_div)
    dv_div.add(f"E{data_start}:E{data_end}")

    # Reason dropdown
    dv_reason = DataValidation(
        type="list", formula1='"' + ",".join(REASONS) + '"', allow_blank=True,
    )
    dv_reason.error = "Pick a reason from the dropdown."
    dv_reason.errorTitle = "Invalid reason"
    ws.add_data_validation(dv_reason)
    dv_reason.add(f"I{data_start}:I{data_end}")

    # Type dropdown
    dv_type = DataValidation(
        type="list", formula1='"' + ",".join(CO_TYPES) + '"', allow_blank=True,
    )
    dv_type.error = "Pick a CO type from the dropdown."
    dv_type.errorTitle = "Invalid type"
    ws.add_data_validation(dv_type)
    dv_type.add(f"J{data_start}:J{data_end}")

    # Status dropdown
    dv_status = DataValidation(
        type="list", formula1='"' + ",".join(STATUSES) + '"', allow_blank=True,
    )
    dv_status.error = "Pick a status from the dropdown."
    dv_status.errorTitle = "Invalid status"
    ws.add_data_validation(dv_status)
    dv_status.add(f"M{data_start}:M{data_end}")

    # Totals row
    totals_row = data_end + 2
    ws.cell(row=totals_row, column=10, value="TOTALS").font = FONT_BODY_BOLD
    ws.cell(row=totals_row, column=10).alignment = Alignment(horizontal="right")
    ws.cell(row=totals_row, column=10).fill = FILL_SUBHEADER
    ws.cell(row=totals_row, column=10).border = BORDER
    # Sum amount
    ws.cell(row=totals_row, column=11,
            value=f"=SUM(K{data_start}:K{data_end})")
    ws.cell(row=totals_row, column=11).number_format = FMT_USD
    ws.cell(row=totals_row, column=11).font = FONT_BODY_BOLD
    ws.cell(row=totals_row, column=11).fill = FILL_SUBHEADER
    ws.cell(row=totals_row, column=11).border = BORDER
    # Sum schedule impact
    ws.cell(row=totals_row, column=12,
            value=f"=SUM(L{data_start}:L{data_end})")
    ws.cell(row=totals_row, column=12).number_format = FMT_INT
    ws.cell(row=totals_row, column=12).font = FONT_BODY_BOLD
    ws.cell(row=totals_row, column=12).fill = FILL_SUBHEADER
    ws.cell(row=totals_row, column=12).border = BORDER

    # Column widths
    set_col_widths(ws, {
        "A": 11, "B": 13, "C": 13, "D": 24,
        "E": 12, "F": 12, "G": 28,
        "H": 38, "I": 18, "J": 14,
        "K": 14, "L": 14, "M": 14, "N": 22, "O": 32,
    })

    ws.freeze_panes = "B4"

    return ws.title, data_start, data_end


# ---------------------------------------------------------------------------
# Tab 4 — CO Summary by Division
# ---------------------------------------------------------------------------

SUMMARY_HEADERS = [
    "CSI Division", "Division Name", "CO Count",
    "Approved Amount", "Pending Amount", "Rejected Amount",
    "Net Approved Amount", "Schedule Impact (Approved)",
]


def build_summary(wb: Workbook, co_sheet: str, co_start: int, co_end: int) -> None:
    ws = wb.create_sheet("CO Summary by Division")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "CO Summary by Division (auto roll-up)"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(SUMMARY_HEADERS))

    ws["A2"] = ("Reads from Change Order Log via SUMIFS - do not edit this tab directly. "
                "One row per CSI Division.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)

    for i, h in enumerate(SUMMARY_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(SUMMARY_HEADERS))

    co_div = f"'{co_sheet}'!$E${co_start}:$E${co_end}"
    co_status = f"'{co_sheet}'!$M${co_start}:$M${co_end}"
    co_amount = f"'{co_sheet}'!$K${co_start}:$K${co_end}"
    co_sched = f"'{co_sheet}'!$L${co_start}:$L${co_end}"

    start_row = 4
    for idx, (div, name) in enumerate(DIVISIONS):
        r = start_row + idx
        ws.cell(row=r, column=1, value=div).number_format = "@"
        ws.cell(row=r, column=2, value=name)
        # CO Count (any non-blank in this division)
        ws.cell(row=r, column=3, value=f'=COUNTIF({co_div},A{r})')
        # Approved Amount
        ws.cell(row=r, column=4,
                value=f'=SUMIFS({co_amount},{co_div},A{r},{co_status},"Approved")')
        # Pending Amount (Pending + Submitted)
        ws.cell(row=r, column=5,
                value=(f'=SUMIFS({co_amount},{co_div},A{r},{co_status},"Pending")'
                       f'+SUMIFS({co_amount},{co_div},A{r},{co_status},"Submitted")'))
        # Rejected Amount
        ws.cell(row=r, column=6,
                value=f'=SUMIFS({co_amount},{co_div},A{r},{co_status},"Rejected")')
        # Net Approved Amount (same as Approved; named for clarity in case future tier changes)
        ws.cell(row=r, column=7,
                value=f'=SUMIFS({co_amount},{co_div},A{r},{co_status},"Approved")')
        # Schedule Impact (Approved)
        ws.cell(row=r, column=8,
                value=f'=SUMIFS({co_sched},{co_div},A{r},{co_status},"Approved")')

        for c in (4, 5, 6, 7):
            ws.cell(row=r, column=c).number_format = FMT_USD
        ws.cell(row=r, column=3).number_format = FMT_INT
        ws.cell(row=r, column=8).number_format = FMT_INT

        for c in range(1, len(SUMMARY_HEADERS) + 1):
            ws.cell(row=r, column=c).border = BORDER
        ws.row_dimensions[r].height = 22

    end_row = start_row + len(DIVISIONS) - 1

    # Totals row
    tr = end_row + 2
    ws.cell(row=tr, column=2, value="TOTALS").font = FONT_BODY_BOLD
    ws.cell(row=tr, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=tr, column=2).fill = FILL_SUBHEADER
    ws.cell(row=tr, column=2).border = BORDER
    for col_letter in ("C", "D", "E", "F", "G", "H"):
        c_idx = ord(col_letter) - ord("A") + 1
        cell = ws.cell(row=tr, column=c_idx,
                       value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
        cell.font = FONT_BODY_BOLD
        cell.fill = FILL_SUBHEADER
        cell.border = BORDER
        if col_letter in ("C", "H"):
            cell.number_format = FMT_INT
        else:
            cell.number_format = FMT_USD

    # Conditional formatting on Approved (D), Pending (E), Rejected (F)
    ws.conditional_formatting.add(
        f"D{start_row}:D{end_row}",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True)),
    )
    ws.conditional_formatting.add(
        f"E{start_row}:E{end_row}",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=FILL_YELLOW, font=Font(color=YELLOW_FONT, bold=True)),
    )
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=FILL_RED, font=Font(color=RED_FONT, bold=True)),
    )

    set_col_widths(ws, {
        "A": 12, "B": 36, "C": 11,
        "D": 18, "E": 18, "F": 18, "G": 20, "H": 20,
    })

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Tab 5 — Dashboard
# ---------------------------------------------------------------------------

def build_dashboard(wb: Workbook, co_sheet: str, co_start: int, co_end: int) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Change Order Dashboard"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:F1")

    ws["A2"] = ("Auto-calculated from Project Info and Change Order Log. "
                "Update those tabs - this one stays in sync.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)
    ws.merge_cells("A2:F2")

    co_amount = f"'{co_sheet}'!$K${co_start}:$K${co_end}"
    co_status = f"'{co_sheet}'!$M${co_start}:$M${co_end}"
    co_sched = f"'{co_sheet}'!$L${co_start}:$L${co_end}"
    co_num = f"'{co_sheet}'!$A${co_start}:$A${co_end}"

    contract = "'Project Info'!D4"

    # ---------------- Top panel: contract impact ----------------
    panel_start = 4
    ws.cell(row=panel_start, column=1, value="Contract Impact").font = FONT_H1
    ws.cell(row=panel_start, column=1).fill = FILL_HEADER
    ws.cell(row=panel_start, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=panel_start, start_column=1,
                   end_row=panel_start, end_column=6)
    ws.row_dimensions[panel_start].height = 26

    # (label, formula, fmt, conditional_kind)
    #   conditional_kind:
    #     None       -> no CF
    #     "signed"   -> green if >=0, red if <0
    #     "co_pct"   -> red if > 0.15, yellow if > 0.05, green otherwise
    metrics = [
        ("Original Contract Amount", f"={contract}", FMT_USD, None),
        ("Total Approved CO Amount",
         f'=SUMIF({co_status},"Approved",{co_amount})', FMT_USD, None),
        ("Total Pending CO Amount",
         f'=SUMIF({co_status},"Pending",{co_amount})+SUMIF({co_status},"Submitted",{co_amount})',
         FMT_USD, None),
        ("Total Rejected CO Amount",
         f'=SUMIF({co_status},"Rejected",{co_amount})', FMT_USD, None),
        ("Adjusted Contract Amount (Original + Approved)",
         f'={contract}+SUMIF({co_status},"Approved",{co_amount})',
         FMT_USD, None),
        ("Approved CO % of Original Contract",
         f'=IFERROR(SUMIF({co_status},"Approved",{co_amount})/{contract},0)',
         FMT_PCT, "co_pct"),
        ("Total Schedule Impact - Approved (days)",
         f'=SUMIF({co_status},"Approved",{co_sched})',
         FMT_INT, "signed"),
    ]

    row = panel_start + 1
    co_pct_cells: list[str] = []
    signed_cells: list[str] = []
    for label, formula, fmt, kind in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = FONT_BODY_BOLD
        ws.cell(row=row, column=1).fill = FILL_SUMMARY_LABEL
        ws.cell(row=row, column=1).alignment = Alignment(vertical="center", indent=1)
        ws.cell(row=row, column=1).border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

        ws.cell(row=row, column=4, value=formula)
        ws.cell(row=row, column=4).font = FONT_BIG_NUMBER
        ws.cell(row=row, column=4).number_format = fmt
        ws.cell(row=row, column=4).alignment = Alignment(
            horizontal="right", vertical="center", indent=1,
        )
        ws.cell(row=row, column=4).border = BORDER
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)

        ref = f"D{row}"
        if kind == "co_pct":
            co_pct_cells.append(ref)
        elif kind == "signed":
            signed_cells.append(ref)
        ws.row_dimensions[row].height = 28
        row += 1

    # CO % of contract conditional formatting:
    #  - Red bg/font when > 15%
    #  - Yellow bg/font when > 5%
    #  - Green bg/font when <= 5%
    for ref in co_pct_cells:
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="greaterThan", formula=["0.15"], stopIfTrue=True,
                       fill=FILL_RED, font=Font(color=RED_FONT, bold=True, size=18)),
        )
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="greaterThan", formula=["0.05"], stopIfTrue=True,
                       fill=FILL_YELLOW, font=Font(color=YELLOW_FONT, bold=True, size=18)),
        )
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="lessThanOrEqual", formula=["0.05"], stopIfTrue=False,
                       fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True, size=18)),
        )

    # Signed (schedule) — green if >=0 (no slip), red if <0 (you bought back time / negative slip)
    # Convention: positive = added days, so green when 0 (no schedule slip), red when added days >0
    # Actually for schedule impact, more days = worse. So invert: red if >0, green if <=0.
    for ref in signed_cells:
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=False,
                       fill=FILL_RED, font=Font(color=RED_FONT, bold=True, size=18)),
        )
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="lessThanOrEqual", formula=["0"], stopIfTrue=False,
                       fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True, size=18)),
        )

    # ---------------- Bottom panel: CO counts by status ----------------
    section_row = row + 2
    ws.cell(row=section_row, column=1, value="CO Counts by Status").font = FONT_H1
    ws.cell(row=section_row, column=1).fill = FILL_HEADER
    ws.cell(row=section_row, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=section_row, start_column=1,
                   end_row=section_row, end_column=6)
    ws.row_dimensions[section_row].height = 26

    headers = ["Status", "Count", "Amount", "% of Total CO $"]
    hr = section_row + 1
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hr, column=i, value=h)
    style_header_row(ws, hr, len(headers))

    # Merge Amount to span two columns visually
    ws.merge_cells(start_row=hr, start_column=3, end_row=hr, end_column=3)
    ws.merge_cells(start_row=hr, start_column=4, end_row=hr, end_column=6)

    status_panel = [
        ("Approved", FILL_GREEN, GREEN_FONT),
        ("Pending", FILL_YELLOW, YELLOW_FONT),
        ("Submitted", FILL_YELLOW, YELLOW_FONT),
        ("Rejected", FILL_RED, RED_FONT),
        ("Disputed", FILL_RED, RED_FONT),
        ("Voided", FILL_SUBHEADER, BRAND_BLUE),
    ]

    total_amount_formula = f"SUMIF({co_status},\"<>\",{co_amount})"

    sr = hr + 1
    for idx, (status, fill, font_color) in enumerate(status_panel):
        r = sr + idx
        ws.cell(row=r, column=1, value=status)
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=12, bold=True, color=font_color)
        ws.cell(row=r, column=1).fill = fill
        ws.cell(row=r, column=1).alignment = Alignment(vertical="center", indent=1)
        ws.cell(row=r, column=1).border = BORDER
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

        # Count
        ws.cell(row=r, column=3,
                value=f'=COUNTIF({co_status},"{status}")')
        ws.cell(row=r, column=3).font = FONT_BODY_BOLD
        ws.cell(row=r, column=3).number_format = FMT_INT
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=3).border = BORDER

        # Amount
        ws.cell(row=r, column=4,
                value=f'=SUMIF({co_status},"{status}",{co_amount})')
        ws.cell(row=r, column=4).font = FONT_BODY_BOLD
        ws.cell(row=r, column=4).number_format = FMT_USD
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.cell(row=r, column=4).border = BORDER

        # % of Total CO $
        ws.cell(row=r, column=5,
                value=f'=IFERROR(SUMIF({co_status},"{status}",{co_amount})'
                      f'/{total_amount_formula},0)')
        ws.cell(row=r, column=5).font = FONT_BODY_BOLD
        ws.cell(row=r, column=5).number_format = FMT_PCT
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.cell(row=r, column=5).border = BORDER
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 24

    # Total row
    tr = sr + len(status_panel) + 1
    ws.cell(row=tr, column=1, value="ALL COs (any status)")
    ws.cell(row=tr, column=1).font = FONT_BODY_BOLD
    ws.cell(row=tr, column=1).fill = FILL_SUBHEADER
    ws.cell(row=tr, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.cell(row=tr, column=1).border = BORDER
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=2)

    ws.cell(row=tr, column=3, value=f'=COUNTA({co_num})')
    ws.cell(row=tr, column=3).font = FONT_BODY_BOLD
    ws.cell(row=tr, column=3).fill = FILL_SUBHEADER
    ws.cell(row=tr, column=3).number_format = FMT_INT
    ws.cell(row=tr, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=tr, column=3).border = BORDER

    ws.cell(row=tr, column=4, value=f'=SUM({co_amount})')
    ws.cell(row=tr, column=4).font = FONT_BODY_BOLD
    ws.cell(row=tr, column=4).fill = FILL_SUBHEADER
    ws.cell(row=tr, column=4).number_format = FMT_USD
    ws.cell(row=tr, column=4).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.cell(row=tr, column=4).border = BORDER

    ws.cell(row=tr, column=5, value="100.0%" if False else 1)
    ws.cell(row=tr, column=5).value = "=1"
    ws.cell(row=tr, column=5).font = FONT_BODY_BOLD
    ws.cell(row=tr, column=5).fill = FILL_SUBHEADER
    ws.cell(row=tr, column=5).number_format = FMT_PCT
    ws.cell(row=tr, column=5).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.cell(row=tr, column=5).border = BORDER
    ws.merge_cells(start_row=tr, start_column=5, end_row=tr, end_column=6)

    set_col_widths(ws, {
        "A": 22, "B": 12, "C": 14, "D": 18, "E": 14, "F": 10,
    })


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

OUTPUT_PATH = "/Users/home/charles/contrpro/files/packages/business/Change_Order_Log.xlsx"


def build() -> str:
    wb = Workbook()

    # Order: Instructions (active), Project Info, Change Order Log, Summary, Dashboard, CSI Reference
    # Build CSI Reference first so named ranges resolve, then move tab to the end.
    build_instructions(wb)            # Active sheet renamed to Instructions
    build_csi_reference(wb)           # Adds named ranges CSI_Divisions and CSI_Table
    build_project_info(wb)
    co_sheet, co_start, co_end = build_change_order_log(wb)
    build_summary(wb, co_sheet, co_start, co_end)
    build_dashboard(wb, co_sheet, co_start, co_end)

    # Re-order tabs: Instructions, Project Info, Change Order Log,
    # CO Summary by Division, Dashboard, CSI Reference (last so it feels like an appendix)
    desired_order = [
        "Instructions",
        "Project Info",
        "Change Order Log",
        "CO Summary by Division",
        "Dashboard",
        "CSI Reference",
    ]
    wb._sheets = [wb[name] for name in desired_order]

    # Hide CSI Reference (canonical lookup; user shouldn't see it by default)
    wb["CSI Reference"].sheet_state = "hidden"

    # Active tab = Dashboard on open
    wb.active = wb.sheetnames.index("Dashboard")

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
