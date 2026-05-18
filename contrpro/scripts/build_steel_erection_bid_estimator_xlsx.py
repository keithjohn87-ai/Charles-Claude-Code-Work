#!/usr/bin/env python3
"""
Build ContrPro Steel Erection Bid Estimator (XLSX) — Steel Erection trade pack
(Complete tier).

Tonnage-based bid estimator for a structural-steel erector. Mirrors the
estimating workflow a working steel-erection PM follows on a typical
commercial project: take off the tonnage by CSI section, apply productivity
rates by ironworker classification, layer equipment + subcontract + premiums,
and roll up to a final bid by CSI division.

This is the trade-specific delta on top of the Universal Sub Suite — use the
Sub Schedule of Values for monthly billing, T&M Tracker for force-account
work, etc. This workbook produces the original bid number.

Tabs:
  1. Instructions
  2. Project Info
  3. Wage Rates               (ironworker classifications + burden + markup)
  4. Equipment Rates          (crane + support equipment)
  5. Material Takeoff         (CSI-coded line items — tonnage, sqft, count)
  6. Labor Estimate           (hrs per ton × loaded rate)
  7. Equipment Estimate       (crane-days + support)
  8. Subcontract Estimate     (bolting / welding / fireproofing if subbed)
  9. Premiums & Conditions    (height, after-hours, distance, complexity adders)
 10. Bid Summary               (roll-up by CSI div → contingency, bond, OH, profit)
 11. CSI Reference             (hidden — drives dropdowns)

Run:
    /Users/home/charles/.venv/bin/python3 \\
        /Users/home/charles/contrpro/scripts/build_steel_erection_bid_estimator_xlsx.py

Output:
    /Users/home/charles/contrpro/files/packages/complete/steel-erection/Steel_Erection_Bid_Estimator.xlsx
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# ---------------------------------------------------------------------------
# Brand / styling
# ---------------------------------------------------------------------------

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
ACCENT_GOLD = "C9A227"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
YELLOW_FILL = "FFEB9C"
YELLOW_FONT = "9C5700"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_SUMMARY_LABEL = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW_FILL)
FILL_GOLD = PatternFill("solid", fgColor=ACCENT_GOLD)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H1 = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_BIG_NUMBER = Font(name="Calibri", size=18, bold=True, color=BRAND_BLUE)
FONT_GREY_ITALIC = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_USD0 = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
FMT_PCT = "0.0%"
FMT_INT = "0"
FMT_NUM2 = "0.00"
FMT_DATE = "yyyy-mm-dd"

# ---------------------------------------------------------------------------
# CSI MasterFormat — Division 05 (Metals) sections used by a steel erector
# Cross-references to other divisions where steel erector scope overlaps
# ---------------------------------------------------------------------------

STEEL_CSI_SECTIONS = [
    ("05 12 23", "Structural Steel Framing", "Primary structural steel members (W, HSS, C, L, MC shapes)"),
    ("05 12 33", "Architecturally Exposed Structural Steel (AESS)", "AESS tier finish steel — higher tolerance, more prep"),
    ("05 21 19", "Open-Web Steel Joists", "K-series, LH-series, DLH-series joists"),
    ("05 21 23", "Joist Girders", "Long-span joist girders supporting joists"),
    ("05 31 13", "Steel Floor Decking", "Composite or non-composite floor deck (1.5VL, 2VL, 3VL, etc.)"),
    ("05 31 23", "Steel Roof Decking", "Roof deck (1.5B, 2B, 3B, 1.5BR, etc.)"),
    ("05 31 33", "Steel Form Decking", "Permanent metal form decking"),
    ("05 41 00", "Structural Metal Stud Framing", "Load-bearing metal stud framing (rare for erector)"),
    ("05 50 00", "Metal Fabrications", "Miscellaneous metals — railings, ladders, embed plates"),
    ("05 51 13", "Metal Pan Stairs", "Metal pan stairs installed by erector"),
    ("05 51 33", "Metal Ladders", "Caged/uncaged metal ladders"),
    ("05 52 13", "Pipe and Tube Railings", "Stair, balcony, and platform railings"),
    ("05 71 13", "Fabricated Metal Spiral Stairs", "Spiral stairs (often shop-fab + field erect)"),
    ("05 73 16", "Wire Rope Decorative Metal Railings", "Tension wire-rope cable rail systems"),
    # Cross-references — steel erector scope occasionally extends into these
    ("03 30 00", "Cast-in-Place Concrete (embeds)", "Embed plates / anchor rods — supplied by erector, set by concrete contractor"),
    ("05 75 00", "Decorative Formed Metal", "Decorative cladding / canopies"),
    ("07 81 00", "Applied Fireproofing", "Spray-applied fireproofing on steel (typically separate sub)"),
    ("09 91 00", "Painting (Touch-Up)", "Shop primer touch-up on welds/abrasions"),
    ("13 34 19", "Pre-Engineered Metal Buildings", "PEMB — if erector takes on the metal-building scope"),
    ("33 16 00", "Water Utility Storage Tanks (Steel)", "Bolted/welded steel water tanks"),
]

# Default ironworker classifications & rates (placeholders the user edits)
# Based on industry-average non-union commercial rates in U.S. South ca. 2026.
# Union and prevailing-wage projects: replace with the controlling wage determination.
DEFAULT_IRONWORKER_CLASSIFICATIONS = [
    ("Foreman / Working Foreman",                58.00, 1.50),
    ("Journeyman Ironworker — Structural",       46.00, 1.50),
    ("Journeyman Ironworker — Connector",        48.00, 1.50),
    ("Journeyman Ironworker — Welder (Certified)", 52.00, 1.50),
    ("Apprentice Ironworker — Year 4",           36.00, 1.50),
    ("Apprentice Ironworker — Year 3",           32.00, 1.50),
    ("Apprentice Ironworker — Year 2",           28.00, 1.50),
    ("Apprentice Ironworker — Year 1",           24.00, 1.50),
    ("Crane Operator",                            54.00, 1.50),
    ("Signal Person / Rigger",                    44.00, 1.50),
    ("Decker / Bolter",                           42.00, 1.50),
    ("Detailer (field)",                          50.00, 1.00),  # rarely OT
    ("Field Engineer / Surveyor",                 56.00, 1.00),
]

# Default equipment — daily rate format. User updates per project.
DEFAULT_EQUIPMENT = [
    ("Crawler crane — 80T",                  "Rented",  185.00, 1850.00,  7400.00),
    ("Crawler crane — 110T",                 "Rented",  235.00, 2350.00,  9400.00),
    ("Crawler crane — 150T",                 "Rented",  310.00, 3100.00, 12400.00),
    ("Crawler crane — 200T",                 "Rented",  395.00, 3950.00, 15800.00),
    ("Truck-mounted crane — 50T (RT)",       "Rented",  165.00, 1650.00,  6600.00),
    ("Truck-mounted crane — 80T (RT)",       "Rented",  220.00, 2200.00,  8800.00),
    ("Boom truck — 18T (carrydeck)",         "Owned",    95.00,  950.00,  3800.00),
    ("Telehandler / forklift",               "Owned",    72.00,  720.00,  2880.00),
    ("Aerial lift — scissor 26'-32'",        "Rented",    48.00,  480.00,  1920.00),
    ("Aerial lift — boom 60'-85'",           "Rented",    98.00,  980.00,  3920.00),
    ("Engine-drive welder (lincoln/miller)", "Owned",    18.00,  180.00,   720.00),
    ("Air compressor — 185 cfm",             "Owned",    14.00,  140.00,   560.00),
    ("Generator — 25 kW",                    "Owned",    12.00,  120.00,   480.00),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def set_col_widths(ws, widths):
    for col, w in widths:
        ws.column_dimensions[col].width = w


def style_header_row(ws, row, cols, start_col=1):
    for c in range(start_col, start_col + cols):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


def apply_body_style(ws, row, cols, start_col=1):
    for c in range(start_col, start_col + cols):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_BODY
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.border = BORDER


# ---------------------------------------------------------------------------
# Tab 1: Instructions
# ---------------------------------------------------------------------------

def build_instructions(ws):
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 110)])

    ws["A1"] = "STEEL ERECTION BID ESTIMATOR — Instructions"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 36

    rows = [
        "",
        "WHAT THIS WORKBOOK IS",
        "A tonnage-based bid estimator for a structural-steel erection subcontractor. Output is a final bid "
        "number broken down by CSI division and section, suitable for submitting on a commercial bid form "
        "(use the Bid Package & Prequalification template from the Universal Sub Suite for the wrapper).",
        "",
        "WORKFLOW (TYPICAL BID DAY)",
        "1. Open the bid documents and identify the structural drawings (S-series sheets) + project manual "
        "(Division 05 specifications).",
        "2. Project Info: fill in the project, GC, delivery method, schedule, location, prevailing-wage "
        "flag, and contract-type assumptions. These drive every other tab.",
        "3. Wage Rates: review the defaults against the locality + project labor regime. On Davis-Bacon or "
        "state-prevailing-wage projects, REPLACE the defaults with the controlling wage determination "
        "(see Certified Payroll Tracker for parallel workflow).",
        "4. Equipment Rates: replace defaults with quotes from your crane vendor for the project tonnage + "
        "reach + pick weight. The 200T crawler default is a sample; bigger projects need bigger cranes.",
        "5. Material Takeoff: enter the tonnage / sqft / count by CSI section. The takeoff usually comes "
        "from a quantity surveyor's report, a fabricator's takeoff (if you bid as erector-only), or your "
        "own takeoff from the structural drawings.",
        "6. Labor Estimate: apply productivity rates (hrs/ton) by member class and crew composition. "
        "Industry-average productivity rates are pre-loaded; tighten or relax based on project access, "
        "complexity, weather window, etc.",
        "7. Equipment Estimate: estimate crane-days and support-equipment days by phase.",
        "8. Subcontract Estimate: if any scope is subbed (high-strength bolting, certified welding, "
        "fireproofing, painting), price as lump-sum or unit-price.",
        "9. Premiums & Conditions: apply percentage adders for working at heights, after-hours work, "
        "distance from your shop, complex geometry, weather contingency, and similar risk factors.",
        "10. Bid Summary: review the CSI-division roll-up. Apply contingency, bond cost, overhead, and "
        "profit to arrive at the final bid number.",
        "",
        "WHAT THIS WORKBOOK IS NOT",
        "  - This is not a substitute for a takeoff. Garbage in, garbage out — the bid is only as good as "
        "the tonnage entered.",
        "  - This is not a substitute for a contract review. The Subcontract Review Checklist (Universal "
        "Sub Suite) is for the deal terms; this estimator is for the price only.",
        "  - This is not a substitute for an erector's experience. The pre-loaded productivity rates are "
        "industry averages; your shop's actual productivity may be 20-40% different. Calibrate against "
        "your last 3 completed jobs before relying on the rates on a tight bid.",
        "",
        "CSI MASTERFORMAT INTEGRATION",
        "Every line item carries a CSI Division (05 primary for steel) and Section (e.g., 05 12 23 "
        "Structural Steel Framing, 05 31 13 Steel Floor Decking). The Bid Summary tab rolls up totals "
        "by CSI division so the bid integrates cleanly into the GC's job-cost accounting and into your "
        "monthly Sub Schedule of Values billing.",
        "",
        "STEEL ERECTOR'S SCOPE — TYPICAL DIVISIONS",
        "  - 05 12 23 Structural Steel Framing — primary structural members (W, HSS, C, L)",
        "  - 05 21 19 Open-Web Steel Joists — K, LH, DLH series",
        "  - 05 31 13/23/33 Steel Decking — floor / roof / form decking",
        "  - 05 50 00 Metal Fabrications — misc metals, embeds, railings (sometimes)",
        "  - 05 51 13 Metal Pan Stairs — stairs installed by erector",
        "  - 03 30 00 (cross-ref) — embed plates supplied by erector, set by concrete contractor",
        "  - 07 81 00 (cross-ref) — spray-applied fireproofing, typically a separate sub",
        "",
        "PRODUCTIVITY RATES (DEFAULTS — ADJUST BASED ON YOUR HISTORY)",
        "  - Structural steel: 8-15 tons/crew-day for normal commercial work; 4-8 for AESS or congested",
        "  - Joists: 25-45 joists/crew-day depending on span and bridging",
        "  - Decking: 7,000-12,000 sqft/crew-day on open floors, less on cut-up plans",
        "  - Bolting: 80-150 high-strength bolts/bolter-day (turn-of-nut method)",
        "  - Welding: depends entirely on weld size and position — estimate per joint, not per ton",
        "",
        "MARKUP DEFAULTS (REVIEW PER PROJECT)",
        "  - Overhead: 12-18% on commercial work",
        "  - Profit: 8-12%",
        "  - Contingency: 3-5% (more for design-build, less for fully-detailed bid)",
        "  - Bond: 0.5-1.5% of bid",
        "  - Total all-in markup typically lands at 25-35% above raw cost",
        "",
        "FIELD-ONLY SCOPE (CONTRPRO STEEL ERECTION TRADE PACK SCOPE)",
        "This estimator is calibrated for FIELD ERECTION only. Fabrication shop work is OUT OF SCOPE — "
        "if your firm fabricates AND erects, use a separate fabrication-shop estimator (different rate "
        "structure, different productivity, different material handling).",
    ]

    for i, text in enumerate(rows, start=2):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if text and text == text.upper() and len(text) < 70:
            cell.font = FONT_H2
        else:
            cell.font = FONT_BODY


# ---------------------------------------------------------------------------
# Tab 2: Project Info
# ---------------------------------------------------------------------------

def build_project_info(ws):
    ws.title = "Project Info"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 42), ("C", 36)])

    ws["B1"] = "PROJECT INFORMATION — Steel Erection Bid"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:C1")
    ws.row_dimensions[1].height = 32

    info_rows = [
        ("Project Name", ""),
        ("Project Address", ""),
        ("Owner", ""),
        ("General Contractor (bidding to)", ""),
        ("Bid Due Date", ""),
        ("Project Type", "Commercial — New Construction"),
        ("Delivery Method", "Design-Bid-Build"),
        ("Construction Schedule (months)", 0),
        ("Anticipated Steel Erection Start", ""),
        ("Anticipated Steel Erection Duration (weeks)", 0),
        ("", ""),
        ("LABOR REGIME", ""),
        ("Davis-Bacon / Prevailing Wage Project?", "No"),
        ("Wage Determination ID (if prevailing wage)", ""),
        ("Union Project?", "No"),
        ("Project State (controls workers' comp + EMR application)", ""),
        ("", ""),
        ("BID ECONOMICS", ""),
        ("Fringe Burden % (on bare wage)", 0.40),
        ("OT Premium Multiplier", 1.50),
        ("Labor Overhead %", 0.15),
        ("Labor Profit %", 0.10),
        ("Material Markup %", 0.10),
        ("Equipment Markup % (rented)", 0.08),
        ("Equipment Markup % (owned)", 0.10),
        ("Subcontract Markup %", 0.07),
        ("Project Contingency %", 0.04),
        ("Bond % of Bid", 0.012),
        ("Project-Level Overhead %", 0.13),
        ("Profit %", 0.10),
        ("", ""),
        ("KEY PROJECT ATTRIBUTES", ""),
        ("Maximum Working Height (ft)", 0),
        ("Heights >50ft Premium Trigger?", "No"),
        ("After-Hours / Weekend Work Required?", "No"),
        ("Distance from Shop (miles)", 0),
        ("AESS (Architecturally Exposed Steel)?", "No"),
        ("Complex Geometry (curved, slanted, offsets)?", "No"),
        ("Site Access Constraints (urban, occupied, phased)?", "No"),
        ("", ""),
        ("PRINCIPAL / BIDDER OF RECORD", ""),
        ("Estimator", ""),
        ("Project Executive (Sub)", ""),
        ("Authorized Signatory", ""),
    ]

    # Named ranges for cross-tab references
    name_map = {
        "ProjectName": "$C$3",
        "GCName": "$C$6",
        "FringeBurden": "$C$21",
        "OTPremium": "$C$22",
        "LaborOH": "$C$23",
        "LaborProfit": "$C$24",
        "MaterialMarkup": "$C$25",
        "EquipRentedMarkup": "$C$26",
        "EquipOwnedMarkup": "$C$27",
        "SubMarkup": "$C$28",
        "Contingency": "$C$29",
        "BondPct": "$C$30",
        "ProjectOH": "$C$31",
        "ProjectProfit": "$C$32",
        "MaxHeight": "$C$35",
        "HeightPremium": "$C$36",
        "AfterHours": "$C$37",
        "DistanceFromShop": "$C$38",
        "IsAESS": "$C$39",
        "IsComplex": "$C$40",
        "AccessConstrained": "$C$41",
        "IsPrevailingWage": "$C$15",
    }

    for nm, ref in name_map.items():
        ws.parent.defined_names[nm] = DefinedName(name=nm, attr_text=f"'Project Info'!{ref}")

    for i, (label, val) in enumerate(info_rows, start=3):
        is_section_header = label and label == label.upper() and label and not label.startswith(("Anticipated", "Authorized", "Estimator")) and len(label) < 32
        # Treat ALL-CAPS short labels as section headers
        if is_section_header:
            ws.cell(row=i, column=2, value=label).font = FONT_H2
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
        else:
            if label:
                ws.cell(row=i, column=2, value=label).font = FONT_BODY_BOLD
                ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
                ws.cell(row=i, column=2).border = BORDER
            c = ws.cell(row=i, column=3, value=val)
            c.font = FONT_BODY
            if label:
                c.border = BORDER
            if isinstance(val, float):
                c.number_format = FMT_PCT if val < 1 and val != 0 else FMT_NUM2
            elif isinstance(val, int) and val == 0:
                c.number_format = FMT_INT

    dv_yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    for cell in ("C15", "C17", "C36", "C37", "C39", "C40", "C41"):
        dv_yn.add(cell)
    ws.add_data_validation(dv_yn)

    dv_delivery = DataValidation(
        type="list",
        formula1='"Design-Bid-Build,Design-Build,GMP,CM at Risk,IPD,Other"',
        allow_blank=True,
    )
    dv_delivery.add("C8")
    ws.add_data_validation(dv_delivery)


# ---------------------------------------------------------------------------
# Tab 3: Wage Rates
# ---------------------------------------------------------------------------

WAGE_RATE_COLS = [
    ("A", "Classification", 38),
    ("B", "Bare Wage ($/hr)", 16),
    ("C", "OT Multiplier", 13),
    ("D", "Burdened ST ($/hr)", 16),
    ("E", "Burdened OT ($/hr)", 16),
    ("F", "Billed ST ($/hr)", 14),
    ("G", "Billed OT ($/hr)", 14),
    ("H", "Notes", 30),
]


def build_wage_rates(ws):
    ws.title = "Wage Rates"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in WAGE_RATE_COLS])

    ws["A1"] = "WAGE RATES — Ironworker Classifications"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Bare wage = pre-burden hourly rate. Burdened = bare × (1 + FringeBurden). Billed = burdened × "
        "(1 + LaborOH) × (1 + LaborProfit). On prevailing-wage / Davis-Bacon projects, replace bare wage "
        "with the wage determination base rate; fringe is paid in cash or to a bona-fide plan."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 32

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(WAGE_RATE_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(WAGE_RATE_COLS))

    for i, (cls, wage, ot) in enumerate(DEFAULT_IRONWORKER_CLASSIFICATIONS, start=HEADER_ROW + 1):
        ws.cell(row=i, column=1, value=cls)
        ws.cell(row=i, column=2, value=wage).number_format = FMT_USD
        ws.cell(row=i, column=3, value=ot).number_format = FMT_NUM2
        # Burdened ST
        ws.cell(row=i, column=4, value=f"=B{i}*(1+FringeBurden)").number_format = FMT_USD
        # Burdened OT (premium on bare only when prevailing wage; full when commercial)
        ws.cell(
            row=i,
            column=5,
            value=(
                f'=IF(IsPrevailingWage="Yes",'
                f'(B{i}*C{i})+(B{i}*FringeBurden),'
                f'B{i}*C{i}*(1+FringeBurden))'
            ),
        ).number_format = FMT_USD
        # Billed ST + OT
        ws.cell(row=i, column=6, value=f"=D{i}*(1+LaborOH)*(1+LaborProfit)").number_format = FMT_USD
        ws.cell(row=i, column=7, value=f"=E{i}*(1+LaborOH)*(1+LaborProfit)").number_format = FMT_USD
        apply_body_style(ws, i, len(WAGE_RATE_COLS))

    last_row = HEADER_ROW + len(DEFAULT_IRONWORKER_CLASSIFICATIONS)
    ws.parent.defined_names["WageRates"] = DefinedName(
        name="WageRates",
        attr_text=f"'Wage Rates'!$A${HEADER_ROW+1}:$G${last_row+5}",
    )


# ---------------------------------------------------------------------------
# Tab 4: Equipment Rates
# ---------------------------------------------------------------------------

EQUIPMENT_COLS = [
    ("A", "Equipment", 36),
    ("B", "Owned / Rented", 14),
    ("C", "Hourly Rate ($)", 14),
    ("D", "Daily Rate ($)", 14),
    ("E", "Weekly Rate ($)", 14),
    ("F", "Markup %", 12),
    ("G", "Billed Daily ($)", 14),
    ("H", "Billed Weekly ($)", 14),
    ("I", "Notes", 26),
]


def build_equipment_rates(ws):
    ws.title = "Equipment Rates"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in EQUIPMENT_COLS])

    ws["A1"] = "EQUIPMENT RATES"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Owned equipment: bare cost + internal markup. Rented: rental rate + small markup for handling. "
        "For project-specific bids, replace rental defaults with actual quotes from your crane vendor."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:I2")

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(EQUIPMENT_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(EQUIPMENT_COLS))

    for i, (name, owned, hr, day, wk) in enumerate(DEFAULT_EQUIPMENT, start=HEADER_ROW + 1):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=owned)
        ws.cell(row=i, column=3, value=hr).number_format = FMT_USD
        ws.cell(row=i, column=4, value=day).number_format = FMT_USD
        ws.cell(row=i, column=5, value=wk).number_format = FMT_USD
        # Markup based on owned vs rented
        ws.cell(
            row=i,
            column=6,
            value=f'=IF(B{i}="Owned",EquipOwnedMarkup,EquipRentedMarkup)',
        ).number_format = FMT_PCT
        # Billed daily + weekly = base * (1 + markup)
        ws.cell(row=i, column=7, value=f"=D{i}*(1+F{i})").number_format = FMT_USD
        ws.cell(row=i, column=8, value=f"=E{i}*(1+F{i})").number_format = FMT_USD
        apply_body_style(ws, i, len(EQUIPMENT_COLS))

    dv_owned = DataValidation(type="list", formula1='"Owned,Rented"', allow_blank=True)
    dv_owned.add(f"B{HEADER_ROW+1}:B{HEADER_ROW+len(DEFAULT_EQUIPMENT)+5}")
    ws.add_data_validation(dv_owned)

    last_row = HEADER_ROW + len(DEFAULT_EQUIPMENT)
    ws.parent.defined_names["EquipRates"] = DefinedName(
        name="EquipRates",
        attr_text=f"'Equipment Rates'!$A${HEADER_ROW+1}:$H${last_row+5}",
    )


# ---------------------------------------------------------------------------
# Tab 5: Material Takeoff (CSI-coded)
# ---------------------------------------------------------------------------

TAKEOFF_COLS = [
    ("A", "Line #", 7),
    ("B", "CSI Section", 12),
    ("C", "Description", 38),
    ("D", "UOM", 8),
    ("E", "Quantity", 12),
    ("F", "Unit Cost ($)", 14),
    ("G", "Material Subtotal ($)", 16),
    ("H", "Markup %", 11),
    ("I", "Material Total ($)", 16),
    ("J", "Notes", 22),
]


def build_takeoff(ws):
    ws.title = "Material Takeoff"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in TAKEOFF_COLS])

    ws["A1"] = "MATERIAL TAKEOFF — CSI-Coded"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Enter one row per CSI-coded material/system. UOM examples: TON (structural steel), EA (joists), "
        "SF (decking), LF (railings), LB (bolts), GAL (paint). Materials supplied by erector — if "
        "the fabricator supplies material direct, exclude from this tab and note in Assumptions."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:J2")
    ws.row_dimensions[2].height = 32

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(TAKEOFF_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(TAKEOFF_COLS))

    # Seed rows for the common steel-erector scopes
    seed = [
        ("05 12 23", "Structural Steel — Wide Flange (W shapes)",  "TON",  0, 0),
        ("05 12 23", "Structural Steel — HSS / Pipe",               "TON",  0, 0),
        ("05 12 23", "Structural Steel — Channels / Angles / MC",   "TON",  0, 0),
        ("05 12 23", "Structural Steel — Plate / Bent Plate",       "TON",  0, 0),
        ("05 12 33", "AESS — Architecturally Exposed Steel (premium)", "TON", 0, 0),
        ("05 12 23", "High-Strength Bolts — A325 (3/4\" ASTM F3125)", "EA",   0, 0),
        ("05 12 23", "High-Strength Bolts — A325 (7/8\" ASTM F3125)", "EA",   0, 0),
        ("05 12 23", "High-Strength Bolts — A490 (1\" ASTM F3125)",   "EA",   0, 0),
        ("05 12 23", "Tension-Control Bolts — F1852 / F2280",         "EA",   0, 0),
        ("05 12 23", "Anchor Rods — F1554 Grade 36",                  "EA",   0, 0),
        ("05 21 19", "Open-Web Steel Joists — K-Series",              "EA",   0, 0),
        ("05 21 19", "Open-Web Steel Joists — LH-Series",             "EA",   0, 0),
        ("05 21 23", "Joist Girders",                                  "EA",   0, 0),
        ("05 31 13", "Steel Floor Deck — 1.5VL Composite",            "SF",   0, 0),
        ("05 31 13", "Steel Floor Deck — 2VL Composite",              "SF",   0, 0),
        ("05 31 13", "Steel Floor Deck — 3VL Composite",              "SF",   0, 0),
        ("05 31 13", "Shear Studs — 3/4\" × 4-1/2\" Headed",          "EA",   0, 0),
        ("05 31 23", "Steel Roof Deck — 1.5B Wide-Rib",               "SF",   0, 0),
        ("05 31 23", "Steel Roof Deck — 1.5BR Wide-Rib Acoustic",     "SF",   0, 0),
        ("05 31 23", "Steel Roof Deck — 3B Long-Span",                "SF",   0, 0),
        ("05 50 00", "Misc Metals — Embed Plates",                    "EA",   0, 0),
        ("05 50 00", "Misc Metals — Lintels",                          "LF",  0, 0),
        ("05 51 13", "Metal Pan Stairs — Risers",                      "EA",  0, 0),
        ("05 51 13", "Stair Stringers / Landings",                     "EA",  0, 0),
        ("05 52 13", "Pipe Railings — Stair Rail",                     "LF",  0, 0),
        ("05 52 13", "Pipe Railings — Roof Edge / Platform",           "LF",  0, 0),
        ("09 91 00", "Touch-Up Primer (welds + abrasions)",            "GAL", 0, 0),
        ("", "", "", "", ""),
        ("", "", "", "", ""),
        ("", "", "", "", ""),
        ("", "", "", "", ""),
    ]

    for i, (csi, desc, uom, qty, ucost) in enumerate(seed, start=HEADER_ROW + 1):
        ws.cell(row=i, column=1, value=f"=IF(C{i}=\"\",\"\",ROW()-{HEADER_ROW})")
        ws.cell(row=i, column=2, value=csi)
        ws.cell(row=i, column=3, value=desc)
        ws.cell(row=i, column=4, value=uom)
        ws.cell(row=i, column=5, value=qty if qty else "").number_format = FMT_NUM2
        ws.cell(row=i, column=6, value=ucost if ucost else "").number_format = FMT_USD
        # Subtotal = qty × unit cost
        ws.cell(row=i, column=7, value=f'=IFERROR(E{i}*F{i},0)').number_format = FMT_USD
        # Markup
        ws.cell(row=i, column=8, value="=MaterialMarkup").number_format = FMT_PCT
        # Total = subtotal × (1 + markup)
        ws.cell(row=i, column=9, value=f'=IFERROR(G{i}*(1+H{i}),0)').number_format = FMT_USD
        apply_body_style(ws, i, len(TAKEOFF_COLS))

    # CSI dropdown
    csi_list = ",".join(sorted({s[0] for s in STEEL_CSI_SECTIONS}))
    dv_csi = DataValidation(type="list", formula1=f'"{csi_list}"', allow_blank=True)
    dv_csi.add(f"B{HEADER_ROW+1}:B{HEADER_ROW+len(seed)+10}")
    ws.add_data_validation(dv_csi)

    # Totals row
    TOTAL_ROW = HEADER_ROW + len(seed) + 1
    ws.cell(row=TOTAL_ROW, column=3, value="MATERIAL TOTAL").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=3).fill = FILL_SUBHEADER
    for col_idx, col_letter in [(7, "G"), (9, "I")]:
        c = ws.cell(
            row=TOTAL_ROW,
            column=col_idx,
            value=f"=SUMIFS({col_letter}{HEADER_ROW+1}:{col_letter}{HEADER_ROW+len(seed)},C{HEADER_ROW+1}:C{HEADER_ROW+len(seed)},\"<>\")",
        )
        c.font = FONT_BODY_BOLD
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_USD
        c.border = BORDER

    ws.parent.defined_names["TakeoffTable"] = DefinedName(
        name="TakeoffTable",
        attr_text=f"'Material Takeoff'!$A${HEADER_ROW+1}:$I${HEADER_ROW+len(seed)+10}",
    )


# ---------------------------------------------------------------------------
# Tab 6: Labor Estimate
# ---------------------------------------------------------------------------

LABOR_EST_COLS = [
    ("A", "Activity", 30),
    ("B", "CSI Section", 12),
    ("C", "Quantity", 12),
    ("D", "UOM", 8),
    ("E", "Productivity (hrs/UOM)", 16),
    ("F", "Crew Composition (#)", 14),
    ("G", "Total Labor Hours", 14),
    ("H", "Avg Crew Rate ($/hr)", 16),
    ("I", "Labor Total ($)", 16),
    ("J", "Notes", 22),
]

# Seeded activities with industry-average productivity rates
LABOR_ACTIVITIES = [
    ("Erect structural steel — typical",         "05 12 23",  0, "TON",  3.50, 5, 46.00),
    ("Erect structural steel — AESS (premium)",  "05 12 33",  0, "TON",  6.50, 5, 46.00),
    ("Erect open-web joists",                    "05 21 19",  0, "EA",   0.40, 4, 44.00),
    ("Erect joist girders",                      "05 21 23",  0, "EA",   1.50, 4, 46.00),
    ("Install floor decking",                    "05 31 13",  0, "SF",   0.020, 4, 42.00),
    ("Install roof decking",                     "05 31 23",  0, "SF",   0.018, 4, 42.00),
    ("Install shear studs (Nelson-stud welding)", "05 31 13", 0, "EA",  0.08, 2, 48.00),
    ("Set / shim embeds and base plates",        "05 12 23",  0, "EA",   0.50, 2, 44.00),
    ("Install misc metals (embeds, lintels)",    "05 50 00",  0, "EA",   1.00, 2, 44.00),
    ("Install metal pan stairs (per riser)",     "05 51 13",  0, "EA",   0.60, 3, 46.00),
    ("Install pipe railings",                    "05 52 13",  0, "LF",   0.30, 2, 42.00),
    ("Bolt-up (high-strength, turn-of-nut)",     "05 12 23",  0, "EA",   0.08, 2, 42.00),
    ("Field welding (per joint)",                "05 12 23",  0, "EA",   0.35, 2, 52.00),
    ("Touch-up primer (after erection)",         "09 91 00",  0, "GAL",  0.50, 1, 42.00),
    ("Plumb-and-true survey (per bay)",          "05 12 23",  0, "EA",   0.75, 2, 50.00),
    ("Final cleanup / demobilization",           "05 12 23",  0, "DAY",  8.00, 4, 42.00),
]


def build_labor_estimate(ws):
    ws.title = "Labor Estimate"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in LABOR_EST_COLS])

    ws["A1"] = "LABOR ESTIMATE — Productivity × Crew × Rate"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Productivity = crew-hours per unit of work (e.g., 3.50 crew-hrs/ton = a 5-person crew working "
        "an 8-hour day erects roughly 11.5 tons/day). Pull quantities from Material Takeoff. Avg Crew "
        "Rate is the loaded (burdened + marked-up) blended hourly rate across the crew composition."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:J2")
    ws.row_dimensions[2].height = 32

    # Productivity calibration banner. Pre-loaded defaults are industry-average
    # commercial-South non-union numbers; real productivity varies 20-40% by
    # crew, region, season, and project conditions. An estimator using the
    # defaults unmodified WILL underbid on union work, cold climates, AESS,
    # or congested sites. Reviewed against your last 3 job actuals before
    # any bid leaves the door.
    ws["A3"] = (
        "⚠ CALIBRATE AGAINST YOUR LAST 3 COMPLETED JOBS BEFORE TRUSTING THESE NUMBERS. "
        "Defaults are commercial-South non-union industry averages. Union work, "
        "cold climates, AESS, or congested sites typically run 20-40% slower."
    )
    ws["A3"].font = Font(name="Calibri", size=11, bold=True, color="9C0006")
    ws["A3"].fill = PatternFill("solid", fgColor="FFC7CE")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.merge_cells("A3:J3")
    ws.row_dimensions[3].height = 36

    HEADER_ROW = 5  # bumped down 1 row to make room for the calibration banner
    for i, (col, header, _) in enumerate(LABOR_EST_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(LABOR_EST_COLS))

    for i, (activity, csi, qty, uom, hrs_per, crew, avg_rate) in enumerate(LABOR_ACTIVITIES, start=HEADER_ROW + 1):
        ws.cell(row=i, column=1, value=activity)
        ws.cell(row=i, column=2, value=csi)
        ws.cell(row=i, column=3, value=qty if qty else "").number_format = FMT_NUM2
        ws.cell(row=i, column=4, value=uom)
        ws.cell(row=i, column=5, value=hrs_per).number_format = "0.000"
        ws.cell(row=i, column=6, value=crew).number_format = FMT_INT
        # Total labor hours = qty * hrs_per_UOM
        ws.cell(row=i, column=7, value=f'=IFERROR(C{i}*E{i},0)').number_format = FMT_NUM2
        # Avg crew rate — manually entered (defaults to journeyman-blended), can be tied to specific class
        ws.cell(row=i, column=8, value=avg_rate).number_format = FMT_USD
        # Labor total = total hours × avg crew rate × (1 + LaborOH) × (1 + LaborProfit) — but avg rate
        # is BARE wage, so we wrap it. Alternatively the user can enter the loaded rate. Default treats
        # column H as bare and applies burden + markup.
        ws.cell(
            row=i,
            column=9,
            value=f'=IFERROR(G{i}*H{i}*(1+FringeBurden)*(1+LaborOH)*(1+LaborProfit),0)',
        ).number_format = FMT_USD
        apply_body_style(ws, i, len(LABOR_EST_COLS))

    TOTAL_ROW = HEADER_ROW + len(LABOR_ACTIVITIES) + 1
    ws.cell(row=TOTAL_ROW, column=1, value="LABOR TOTAL").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=1).fill = FILL_SUBHEADER
    for col_idx, col_letter in [(7, "G"), (9, "I")]:
        c = ws.cell(
            row=TOTAL_ROW,
            column=col_idx,
            value=f"=SUM({col_letter}{HEADER_ROW+1}:{col_letter}{HEADER_ROW+len(LABOR_ACTIVITIES)})",
        )
        c.font = FONT_BODY_BOLD
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_USD if col_idx == 9 else FMT_NUM2
        c.border = BORDER


# ---------------------------------------------------------------------------
# Tab 7: Equipment Estimate
# ---------------------------------------------------------------------------

EQUIP_EST_COLS = [
    ("A", "Equipment", 36),
    ("B", "Phase / Activity", 24),
    ("C", "Days", 10),
    ("D", "Daily Rate ($)", 14),
    ("E", "Subtotal ($)", 14),
    ("F", "Markup %", 12),
    ("G", "Total ($)", 14),
    ("H", "Notes", 22),
]


def build_equipment_estimate(ws):
    ws.title = "Equipment Estimate"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in EQUIP_EST_COLS])

    ws["A1"] = "EQUIPMENT ESTIMATE"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Days = total billable days on the project. Crane phase typically = structural steel + joists + "
        "girders + decking days. Some equipment (welder, generator) runs the full erection duration."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:H2")

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(EQUIP_EST_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(EQUIP_EST_COLS))

    seed = [
        ("Primary crane (e.g., 150T crawler)",          "Steel + joists + decking", 0, 3100.00),
        ("Secondary crane (e.g., 80T)",                 "Misc lifts / overflow",     0, 1850.00),
        ("Boom truck / carrydeck",                       "Bolt-up + misc",            0,  950.00),
        ("Aerial lift — boom 60-85'",                    "Connection / welding",      0,  980.00),
        ("Aerial lift — scissor",                        "Decking / detailing",       0,  480.00),
        ("Engine-drive welder",                          "Field welding",             0,  180.00),
        ("Telehandler / forklift",                       "Material handling",         0,  720.00),
        ("Air compressor",                                "Bolting / impact",         0,  140.00),
        ("Generator",                                     "Power for tools",           0,  120.00),
        ("Mobilization / demobilization (allowance)",    "Each crane move",            0,  0.00),
        ("", "", 0, 0),
    ]

    for i, (eq, phase, days, rate) in enumerate(seed, start=HEADER_ROW + 1):
        ws.cell(row=i, column=1, value=eq)
        ws.cell(row=i, column=2, value=phase)
        ws.cell(row=i, column=3, value=days if days else "").number_format = FMT_INT
        ws.cell(row=i, column=4, value=rate).number_format = FMT_USD
        ws.cell(row=i, column=5, value=f'=IFERROR(C{i}*D{i},0)').number_format = FMT_USD
        # Default markup — rented (most cranes rented unless owner-operator)
        ws.cell(row=i, column=6, value="=EquipRentedMarkup").number_format = FMT_PCT
        ws.cell(row=i, column=7, value=f'=IFERROR(E{i}*(1+F{i}),0)').number_format = FMT_USD
        apply_body_style(ws, i, len(EQUIP_EST_COLS))

    TOTAL_ROW = HEADER_ROW + len(seed) + 1
    ws.cell(row=TOTAL_ROW, column=1, value="EQUIPMENT TOTAL").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=1).fill = FILL_SUBHEADER
    for col_idx, col_letter in [(5, "E"), (7, "G")]:
        c = ws.cell(
            row=TOTAL_ROW,
            column=col_idx,
            value=f"=SUM({col_letter}{HEADER_ROW+1}:{col_letter}{HEADER_ROW+len(seed)})",
        )
        c.font = FONT_BODY_BOLD
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_USD
        c.border = BORDER


# ---------------------------------------------------------------------------
# Tab 8: Subcontract Estimate
# ---------------------------------------------------------------------------

SUB_COLS = [
    ("A", "Scope", 36),
    ("B", "CSI Section", 12),
    ("C", "Sub Quote ($)", 14),
    ("D", "Markup %", 12),
    ("E", "Total ($)", 14),
    ("F", "Sub Name", 24),
    ("G", "Notes", 22),
]


def build_subcontract_estimate(ws):
    ws.title = "Subcontract Estimate"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in SUB_COLS])

    ws["A1"] = "SUBCONTRACT ESTIMATE — Lower-Tier Work"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Scopes commonly subbed by a steel erector: high-strength bolting, certified welding, spray-applied "
        "fireproofing, painting/galvanizing touch-up, decking install (if not self-performed). Get firm "
        "quotes from the lower-tier sub before bid day; lump-sum or unit-price."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 32

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(SUB_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(SUB_COLS))

    seed = [
        ("High-strength bolting (TC bolts, calibrated wrench)", "05 12 23", 0, "", ""),
        ("Certified welding (AWS D1.1 structural)",              "05 12 23", 0, "", ""),
        ("Spray-applied fireproofing (SFRM)",                    "07 81 00", 0, "", ""),
        ("Touch-up primer / paint",                              "09 91 00", 0, "", ""),
        ("Galvanizing repair (cold-galv compound)",              "05 12 23", 0, "", ""),
        ("Decking install (if not self-performed)",              "05 31 13", 0, "", ""),
        ("Bolt inspection (third-party)",                        "05 12 23", 0, "", ""),
        ("Welding inspection (third-party UT/MT/PT)",            "05 12 23", 0, "", ""),
        ("", "", 0, "", ""),
        ("", "", 0, "", ""),
    ]

    for i, (scope, csi, cost, sub, notes) in enumerate(seed, start=HEADER_ROW + 1):
        ws.cell(row=i, column=1, value=scope)
        ws.cell(row=i, column=2, value=csi)
        ws.cell(row=i, column=3, value=cost if cost else "").number_format = FMT_USD
        ws.cell(row=i, column=4, value="=SubMarkup").number_format = FMT_PCT
        ws.cell(row=i, column=5, value=f'=IFERROR(C{i}*(1+D{i}),0)').number_format = FMT_USD
        ws.cell(row=i, column=6, value=sub)
        ws.cell(row=i, column=7, value=notes)
        apply_body_style(ws, i, len(SUB_COLS))

    TOTAL_ROW = HEADER_ROW + len(seed) + 1
    ws.cell(row=TOTAL_ROW, column=1, value="SUBCONTRACT TOTAL").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=1).fill = FILL_SUBHEADER
    for col_idx, col_letter in [(3, "C"), (5, "E")]:
        c = ws.cell(
            row=TOTAL_ROW,
            column=col_idx,
            value=f"=SUM({col_letter}{HEADER_ROW+1}:{col_letter}{HEADER_ROW+len(seed)})",
        )
        c.font = FONT_BODY_BOLD
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_USD
        c.border = BORDER


# ---------------------------------------------------------------------------
# Tab 9: Premiums & Conditions
# ---------------------------------------------------------------------------

def build_premiums(ws):
    ws.title = "Premiums & Conditions"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 50), ("C", 16), ("D", 16), ("E", 36)])

    ws["B1"] = "PREMIUMS & CONDITIONS — Risk Adders"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:E1")
    ws.row_dimensions[1].height = 32

    ws["B2"] = (
        "Percentage adders applied to the base labor + equipment + subcontract subtotal. Use when a project "
        "attribute materially increases productivity loss, risk, or carrying cost."
    )
    ws["B2"].font = FONT_GREY_ITALIC
    ws.merge_cells("B2:E2")

    HEADER_ROW = 4
    headers = ["Condition", "Applies?", "Premium %", "Basis / Notes"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=HEADER_ROW, column=i, value=h)
    style_header_row(ws, HEADER_ROW, len(headers), start_col=2)

    seed = [
        ("Working at heights >50 ft (OSHA 1926.760 fall protection)", "=HeightPremium", 0.07, "Productivity loss, equipment for fall protection"),
        ("After-hours / weekend / night work (shift premium)", "=AfterHours", 0.15, "1.5× OT or shift differential"),
        ("Distance from shop >50 miles", '=IF(DistanceFromShop>50,"Yes","No")', 0.04, "Per-diem, travel time, equipment-move cost"),
        ("Architecturally Exposed Steel (AESS — Tier 2 or higher)", "=IsAESS", 0.12, "Slower productivity, higher tolerance, more handling"),
        ("Complex geometry (curved, slanted, twisted, offsets)", "=IsComplex", 0.10, "Slower placement, more shimming, more rework"),
        ("Constrained site (urban, occupied, phased, tight crane swing)", "=AccessConstrained", 0.08, "Slower productivity + more crane positions"),
        ("Winter weather (Nov-Mar in cold climates)", "No", 0.06, "Lost days, equipment heating, ice/snow removal"),
        ("Hurricane / wind contingency (coastal projects)", "No", 0.03, "Stoppage during named storms, anchor verification"),
        ("Davis-Bacon / state prevailing wage", "=IsPrevailingWage", 0.00, "Already captured in wage rates — informational"),
        ("Project-specific PPE (FR, arc-flash, special tie-off)", "No", 0.01, "PPE replacement frequency"),
        ("", "", 0, ""),
        ("", "", 0, ""),
    ]

    for i, (cond, applies, pct, notes) in enumerate(seed, start=HEADER_ROW + 1):
        ws.cell(row=i, column=2, value=cond).font = FONT_BODY
        ws.cell(row=i, column=2).border = BORDER
        c_applies = ws.cell(row=i, column=3, value=applies)
        c_applies.border = BORDER
        c_pct = ws.cell(row=i, column=4, value=pct)
        c_pct.number_format = FMT_PCT
        c_pct.border = BORDER
        ws.cell(row=i, column=5, value=notes).font = FONT_GREY_ITALIC
        ws.cell(row=i, column=5).border = BORDER

    # Total premium calculation — sum of (Premium % * 1) for each "Yes" / TRUE row
    TOTAL_ROW = HEADER_ROW + len(seed) + 1
    ws.cell(row=TOTAL_ROW, column=2, value="TOTAL PREMIUM % (applied to labor + equip + sub)").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=2).fill = FILL_SUBHEADER
    ws.cell(row=TOTAL_ROW, column=2).border = BORDER
    # SUMPRODUCT: (Applies = "Yes") × Premium %
    formula = (
        f"=SUMPRODUCT((C{HEADER_ROW+1}:C{HEADER_ROW+len(seed)}=\"Yes\")*"
        f"D{HEADER_ROW+1}:D{HEADER_ROW+len(seed)})"
    )
    c = ws.cell(row=TOTAL_ROW, column=4, value=formula)
    c.font = FONT_BIG_NUMBER
    c.number_format = FMT_PCT
    c.fill = FILL_SUBHEADER
    c.border = BORDER

    # Named range for Bid Summary
    ws.parent.defined_names["TotalPremiumPct"] = DefinedName(
        name="TotalPremiumPct",
        attr_text=f"'Premiums & Conditions'!$D${TOTAL_ROW}",
    )


# ---------------------------------------------------------------------------
# Tab 10: Bid Summary
# ---------------------------------------------------------------------------

def build_bid_summary(ws):
    ws.title = "Bid Summary"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 50), ("C", 22), ("D", 18), ("E", 28)])

    ws["B1"] = "BID SUMMARY — Roll-Up to Final Bid"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:E1")
    ws.row_dimensions[1].height = 32

    ws["B2"] = (
        "Project: =ProjectName    GC: =GCName"
    )
    ws["B2"].font = FONT_BODY_BOLD
    ws.merge_cells("B2:E2")

    # === Roll-up section ===
    ws["B4"] = "DIRECT COST ROLL-UP"
    ws["B4"].font = FONT_H2

    # Cross-tab pulls into the Bid Summary roll-up. Row math is HEADER_ROW + seed_count + 1.
    # Labor Estimate's HEADER_ROW is 5 (bumped from 4 in 2026-05-17 audit to add a productivity
    # calibration banner); all others still use HEADER_ROW=4.
    rows = [
        ("Material (from Material Takeoff)",   "='Material Takeoff'!I" + str(4 + 31 + 1)),       # Material Takeoff total at row 36
        ("Labor (from Labor Estimate)",        "='Labor Estimate'!I" + str(5 + 16 + 1)),         # Labor Estimate total at row 22
        ("Equipment (from Equipment Estimate)", "='Equipment Estimate'!G" + str(4 + 11 + 1)),    # Equipment Estimate total at row 16
        ("Subcontract (from Subcontract Estimate)", "='Subcontract Estimate'!E" + str(4 + 10 + 1)),  # Subcontract Estimate total at row 15
    ]

    start_row = 5
    for i, (label, formula) in enumerate(rows, start=start_row):
        ws.cell(row=i, column=2, value=label).font = FONT_BODY_BOLD
        ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
        ws.cell(row=i, column=2).border = BORDER
        c = ws.cell(row=i, column=3, value=formula)
        c.font = FONT_BODY
        c.number_format = FMT_USD
        c.border = BORDER

    subtotal_row = start_row + len(rows)
    ws.cell(row=subtotal_row, column=2, value="DIRECT COST SUBTOTAL").font = FONT_BODY_BOLD
    ws.cell(row=subtotal_row, column=2).fill = FILL_SUBHEADER
    ws.cell(row=subtotal_row, column=2).border = BORDER
    c = ws.cell(row=subtotal_row, column=3, value=f"=SUM(C{start_row}:C{subtotal_row-1})")
    c.font = FONT_BODY_BOLD
    c.fill = FILL_SUBHEADER
    c.number_format = FMT_USD
    c.border = BORDER

    # === Premium row ===
    premium_row = subtotal_row + 2
    ws.cell(row=premium_row, column=2, value="Premium (heights, after-hours, distance, AESS, etc.)").font = FONT_BODY_BOLD
    ws.cell(row=premium_row, column=2).fill = FILL_SUMMARY_LABEL
    ws.cell(row=premium_row, column=2).border = BORDER
    ws.cell(row=premium_row, column=3, value=f"=C{subtotal_row}*TotalPremiumPct").number_format = FMT_USD
    ws.cell(row=premium_row, column=3).border = BORDER
    ws.cell(row=premium_row, column=4, value="=TotalPremiumPct").number_format = FMT_PCT
    ws.cell(row=premium_row, column=4).border = BORDER

    # === Cost-of-work subtotal ===
    cost_row = premium_row + 1
    ws.cell(row=cost_row, column=2, value="COST OF WORK SUBTOTAL").font = FONT_BODY_BOLD
    ws.cell(row=cost_row, column=2).fill = FILL_SUBHEADER
    ws.cell(row=cost_row, column=2).border = BORDER
    c = ws.cell(row=cost_row, column=3, value=f"=C{subtotal_row}+C{premium_row}")
    c.font = FONT_BODY_BOLD
    c.fill = FILL_SUBHEADER
    c.number_format = FMT_USD
    c.border = BORDER

    # === Contingency + Overhead + Profit + Bond ===
    markup_rows = [
        ("Contingency", "Contingency"),
        ("Project Overhead (general conditions, supervision)", "ProjectOH"),
        ("Profit", "ProjectProfit"),
    ]
    cur_row = cost_row + 1
    for label, named_range in markup_rows:
        ws.cell(row=cur_row, column=2, value=label).font = FONT_BODY_BOLD
        ws.cell(row=cur_row, column=2).fill = FILL_SUMMARY_LABEL
        ws.cell(row=cur_row, column=2).border = BORDER
        ws.cell(row=cur_row, column=3, value=f"=C{cost_row}*{named_range}").number_format = FMT_USD
        ws.cell(row=cur_row, column=3).border = BORDER
        ws.cell(row=cur_row, column=4, value=f"={named_range}").number_format = FMT_PCT
        ws.cell(row=cur_row, column=4).border = BORDER
        cur_row += 1

    # Bid before bond
    pre_bond_row = cur_row
    ws.cell(row=pre_bond_row, column=2, value="BID BEFORE BOND").font = FONT_BODY_BOLD
    ws.cell(row=pre_bond_row, column=2).fill = FILL_SUBHEADER
    ws.cell(row=pre_bond_row, column=2).border = BORDER
    c = ws.cell(
        row=pre_bond_row,
        column=3,
        value=f"=C{cost_row}+C{cost_row+1}+C{cost_row+2}+C{cost_row+3}",
    )
    c.font = FONT_BODY_BOLD
    c.fill = FILL_SUBHEADER
    c.number_format = FMT_USD
    c.border = BORDER
    cur_row += 1

    # Bond
    ws.cell(row=cur_row, column=2, value="Performance + Payment Bond").font = FONT_BODY_BOLD
    ws.cell(row=cur_row, column=2).fill = FILL_SUMMARY_LABEL
    ws.cell(row=cur_row, column=2).border = BORDER
    ws.cell(row=cur_row, column=3, value=f"=C{pre_bond_row}*BondPct/(1-BondPct)").number_format = FMT_USD
    ws.cell(row=cur_row, column=3).border = BORDER
    ws.cell(row=cur_row, column=4, value="=BondPct").number_format = FMT_PCT
    ws.cell(row=cur_row, column=4).border = BORDER
    ws.cell(row=cur_row, column=5, value="Of final bid").font = FONT_GREY_ITALIC
    cur_row += 1

    # === FINAL BID ===
    final_row = cur_row + 1
    ws.cell(row=final_row, column=2, value="FINAL BID NUMBER").font = FONT_TITLE
    ws.cell(row=final_row, column=2).fill = FILL_GOLD
    ws.cell(row=final_row, column=2).border = BORDER
    c = ws.cell(row=final_row, column=3, value=f"=C{pre_bond_row}+C{cur_row-1}")
    c.font = FONT_TITLE
    c.fill = FILL_GOLD
    c.number_format = FMT_USD0
    c.border = BORDER
    ws.row_dimensions[final_row].height = 38

    # === Sanity checks ===
    sanity_row = final_row + 3
    ws.cell(row=sanity_row, column=2, value="SANITY CHECKS").font = FONT_H2

    checks = [
        ("$/Ton (final bid / total structural tons)", "=IFERROR(C{final}/(SUMIFS('Material Takeoff'!E:E,'Material Takeoff'!B:B,\"05 12 23\",'Material Takeoff'!D:D,\"TON\")+SUMIFS('Material Takeoff'!E:E,'Material Takeoff'!B:B,\"05 12 33\",'Material Takeoff'!D:D,\"TON\")),0)".format(final=final_row)),
        ("$/SF Decking (subcontract + decking labor)", "=IFERROR((SUMIFS('Labor Estimate'!I:I,'Labor Estimate'!B:B,\"05 31 13\")+SUMIFS('Labor Estimate'!I:I,'Labor Estimate'!B:B,\"05 31 23\"))/(SUMIFS('Material Takeoff'!E:E,'Material Takeoff'!B:B,\"05 31 13\",'Material Takeoff'!D:D,\"SF\")+SUMIFS('Material Takeoff'!E:E,'Material Takeoff'!B:B,\"05 31 23\",'Material Takeoff'!D:D,\"SF\")),0)"),
        # 2026-05-17 audit fix: Labor Estimate's HEADER_ROW bumped from 4 to 5 (calibration
        # banner). Total row at 5+16+1 = 22. Previous 4+17 = 21 was the last DATA row.
        ("Labor % of final bid (sanity: typically 25-45%)", f"=IFERROR('Labor Estimate'!I{5+16+1}/C{final_row},0)"),
        # 2026-05-17 audit fix: Equipment Estimate total at 4+11+1 = 16 (not 15).
        # Previous formula pointed to G15 which is an empty data row → sanity check always read 0%.
        ("Equipment % of final bid (sanity: typically 8-18%)", f"=IFERROR('Equipment Estimate'!G{4+11+1}/C{final_row},0)"),
        ("Markup % effective (above raw direct cost)", f"=IFERROR((C{final_row}-C{subtotal_row})/C{subtotal_row},0)"),
    ]
    for i, (label, formula) in enumerate(checks, start=sanity_row + 1):
        ws.cell(row=i, column=2, value=label).font = FONT_BODY
        ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
        ws.cell(row=i, column=2).border = BORDER
        c = ws.cell(row=i, column=3, value=formula)
        c.font = FONT_BODY
        c.border = BORDER
        if "%" in label:
            c.number_format = FMT_PCT
        else:
            c.number_format = FMT_USD

    # === CSI Division roll-up ===
    csi_row = sanity_row + len(checks) + 3
    ws.cell(row=csi_row, column=2, value="CSI DIVISION ROLL-UP (for Sub Schedule of Values)").font = FONT_H2

    headers = ["CSI Division", "Material ($)", "Labor ($)", "Total ($)"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=csi_row + 1, column=i, value=h)
    style_header_row(ws, csi_row + 1, len(headers), start_col=2)

    divisions = [
        ("03 — Cast-in-Place Concrete (embeds cross-ref)", "03"),
        ("05 — Metals", "05"),
        ("07 — Thermal & Moisture (fireproofing cross-ref)", "07"),
        ("09 — Finishes (paint cross-ref)", "09"),
    ]
    for i, (div_label, div_code) in enumerate(divisions, start=csi_row + 2):
        ws.cell(row=i, column=2, value=div_label).font = FONT_BODY
        ws.cell(row=i, column=2).border = BORDER
        # Material total for this division — match CSI section starting with the division code
        ws.cell(
            row=i,
            column=3,
            value=f'=SUMPRODUCT((LEFT(\'Material Takeoff\'!B5:B40,2)="{div_code}")*\'Material Takeoff\'!I5:I40)',
        ).number_format = FMT_USD
        ws.cell(row=i, column=3).border = BORDER
        # Labor total for this division
        ws.cell(
            row=i,
            column=4,
            value=f'=SUMPRODUCT((LEFT(\'Labor Estimate\'!B5:B25,2)="{div_code}")*\'Labor Estimate\'!I5:I25)',
        ).number_format = FMT_USD
        ws.cell(row=i, column=4).border = BORDER
        ws.cell(row=i, column=5, value=f"=C{i}+D{i}").number_format = FMT_USD
        ws.cell(row=i, column=5).border = BORDER


# ---------------------------------------------------------------------------
# Tab 11: CSI Reference (hidden)
# ---------------------------------------------------------------------------

def build_csi_reference(ws):
    ws.title = "CSI Reference"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 14), ("B", 50), ("C", 60)])

    ws["A1"] = "CSI MasterFormat — Steel Erector Scope"
    ws["A1"].font = FONT_H2

    ws["A3"] = "Section"
    ws["B3"] = "Title"
    ws["C3"] = "Scope Notes"
    for c in (ws["A3"], ws["B3"], ws["C3"]):
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    for i, (code, title, notes) in enumerate(STEEL_CSI_SECTIONS, start=4):
        ws.cell(row=i, column=1, value=code).font = FONT_BODY
        ws.cell(row=i, column=2, value=title).font = FONT_BODY
        ws.cell(row=i, column=3, value=notes).font = FONT_BODY
        for col in (1, 2, 3):
            ws.cell(row=i, column=col).border = BORDER

    ws.sheet_state = "hidden"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    wb = Workbook()
    build_instructions(wb.active)
    build_project_info(wb.create_sheet())
    build_wage_rates(wb.create_sheet())
    build_equipment_rates(wb.create_sheet())
    build_takeoff(wb.create_sheet())
    build_labor_estimate(wb.create_sheet())
    build_equipment_estimate(wb.create_sheet())
    build_subcontract_estimate(wb.create_sheet())
    build_premiums(wb.create_sheet())
    build_bid_summary(wb.create_sheet())
    build_csi_reference(wb.create_sheet())

    out_dir = "/Users/home/charles/contrpro/files/packages/complete/steel-erection"
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "Steel_Erection_Bid_Estimator.xlsx")
    wb.save(out_path)
    sz = os.path.getsize(out_path)
    print(f"Wrote {out_path} ({sz//1024} KB)")


if __name__ == "__main__":
    main()
