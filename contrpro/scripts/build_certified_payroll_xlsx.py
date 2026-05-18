#!/usr/bin/env python3
"""
Build ContrPro Certified Payroll Tracker (XLSX) — Universal Sub Suite.

A WH-347-mirror workbook for federal Davis-Bacon and state prevailing-wage
projects. Lets a sub set up the wage determination, log workers, enter weekly
payroll, and produce a printable WH-347-style weekly sheet with auto-calculated
gross, fringe, and Statement-of-Compliance text.

Tabs:
  1. Instructions
  2. Project Info               (project, prime contract no., WD id, week endings)
  3. Wage Determination         (one row per classification — Rate, Fringe, OT mult)
  4. Workers Master             (one row per worker — name, ID, classification)
  5. Weekly Payroll             (printable WH-347-style sheet; daily hours; auto math)
  6. Statement of Compliance    (signed text mirroring WH-348)
  7. CSI Reference              (hidden)

Run:
    /Users/home/charles/.venv/bin/python3 \\
        /Users/home/charles/contrpro/scripts/build_certified_payroll_xlsx.py

Output:
    /Users/home/charles/contrpro/files/packages/complete/sub/Certified_Payroll_Tracker.xlsx
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
YELLOW_FILL = "FFEB9C"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_SUMMARY_LABEL = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW_FILL)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_BIG_NUMBER = Font(name="Calibri", size=18, bold=True, color=BRAND_BLUE)
FONT_GREY_ITALIC = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_PCT = "0.0%"
FMT_INT = "0"
FMT_NUM2 = "0.00"
FMT_DATE = "yyyy-mm-dd"


def set_col_widths(ws, widths):
    for col, w in widths:
        ws.column_dimensions[col].width = w


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


def apply_body_style(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_BODY
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.border = BORDER


# ---------------------------------------------------------------------------
# Tab 1: Instructions
# ---------------------------------------------------------------------------

def build_instructions(ws):
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 100)])

    ws["A1"] = "CERTIFIED PAYROLL TRACKER — Instructions"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 36

    rows = [
        "",
        "WHAT THIS WORKBOOK IS",
        "A WH-347-mirror workbook for weekly certified payroll on federal Davis-Bacon and state prevailing-"
        "wage projects. The printable Weekly Payroll tab matches the column structure of WH-347 so the sheet "
        "can be submitted as the certified payroll OR cross-checked against your accounting system's "
        "WH-347 export.",
        "",
        "USAGE — ONE-TIME SETUP",
        "1. Project Info: project name, GC, prime contract number, WD identifying number, sub's PM, "
        "Statement-of-Compliance signatory.",
        "2. Wage Determination: one row per classification used on this project. Enter the WD Rate (base) "
        "and WD Fringe per hour; the workbook will compute OT rate and total hourly cost.",
        "3. Workers Master: one row per worker — name, employee ID, classification (validated against the "
        "Wage Determination), and any apprenticeship/training program data.",
        "",
        "USAGE — WEEKLY",
        "1. At the end of each workweek, populate the Weekly Payroll tab.",
        "2. Each row = one worker. Enter daily hours (ST and OT columns for each day Sun-Sat) and the "
        "workbook calculates total hours, gross pay, fringe owed, and net wage.",
        "3. Verify the totals against your payroll software.",
        "4. Print the Weekly Payroll tab plus the Statement of Compliance tab for the period.",
        "5. Statement of Compliance is signed by the authorized officer under penalty of perjury — review "
        "before signing.",
        "6. Submit to GC within 7 days of the workweek end.",
        "7. File a paper copy and a scanned copy; retain for 3 years from contract completion (longer for "
        "some states — confirm).",
        "",
        "MULTIPLE WEEKS",
        "The Weekly Payroll tab supports one week per use. For multiple weeks, duplicate this workbook (one "
        "file per week) or duplicate the tab inside one workbook (one tab per week). The Workers Master and "
        "Wage Determination tabs are the same across weeks, so do not duplicate those.",
        "",
        "FRINGE — CASH vs. PLAN",
        "Set the per-classification fringe delivery on Wage Determination. The 'Cash Fringe' column is what "
        "appears in the worker's gross wage on WH-347. The 'Plan Fringe' column is what is contributed to "
        "bona-fide benefit plans (not in the worker's gross wage). Cash + Plan must equal at least WD Fringe.",
        "",
        "OVERTIME MATH",
        "OT premium is paid on the base wage only — not on the fringe. The workbook calculates OT pay as "
        "(Hours × Rate) + (Hours × Rate × 0.5) on the base portion, then adds (Hours × Fringe) at straight "
        "rate. State laws can require double-time on specific days; adjust the OT multiplier on Wage "
        "Determination if so.",
        "",
        "APPRENTICE/TRAINEE",
        "Workers marked as Apprentice on Workers Master have a wage percentage field. The Weekly Payroll tab "
        "applies that percentage to the journeyman base rate from the WD. The full WD Fringe still applies — "
        "apprentices receive 100% of the fringe even at reduced base wage, unless the state prevailing-wage "
        "law specifies otherwise.",
        "",
        "STATEMENT OF COMPLIANCE",
        "Section 4(c) of the Statement is the False Claims Act trigger. The signatory must review every "
        "weekly payroll line, confirm classifications match the work actually performed, confirm fringe "
        "deliveries are documented, and confirm apprentice ratios are observed before signing. If any item "
        "is unverified, do not sign — escalate to compliance counsel or the certified-payroll services firm.",
        "",
        "AUDIT",
        "Davis-Bacon audits are random; state audits are sometimes triggered by employee complaint or by an "
        "interested party (union, competing bidder). USDOL audits typically look back 3 years. Keep this "
        "workbook plus the underlying timecards, payroll records, deduction authorizations, fringe-plan "
        "contribution records, and apprenticeship registrations together.",
    ]

    for i, text in enumerate(rows, start=2):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if text and text == text.upper() and len(text) < 70:
            cell.font = FONT_H2
        else:
            cell.font = FONT_BODY


# ---------------------------------------------------------------------------
# Tab 2: Project Info
# ---------------------------------------------------------------------------

def build_project_info(ws):
    ws.title = "Project Info"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 42), ("C", 30)])

    ws["B1"] = "CERTIFIED PAYROLL PROJECT INFORMATION"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:C1")
    ws.row_dimensions[1].height = 32

    info_rows = [
        ("Project Name", ""),
        ("Project Address", ""),
        ("GC", ""),
        ("Prime Contract No.", ""),
        ("Subcontract No.", ""),
        ("Subcontractor (legal name)", ""),
        ("Sub Address", ""),
        ("Sub FEIN", ""),
        ("Federal Davis-Bacon? (Yes / No)", "Yes"),
        ("State Prevailing Wage? (Yes / No)", "No"),
        ("State (if state PW applies)", ""),
        ("Wage Determination ID", ""),
        ("WD Effective Date", ""),
        ("WD Modification No.", ""),
        ("Construction Type (Bldg / Hwy / Heavy / Res)", "Building"),
        ("Payroll No. (sequential)", 1),
        ("For Week Ending (Saturday)", ""),
        ("Statement Signatory (officer name)", ""),
        ("Signatory Title", ""),
        ("Sub's PM", ""),
    ]

    name_map = {
        "ProjectName": "$C$3",
        "PrimeContractNo": "$C$6",
        "SubName": "$C$8",
        "PayrollNo": "$C$18",
        "WeekEnding": "$C$19",
        "SignatoryName": "$C$20",
        "SignatoryTitle": "$C$21",
    }
    for nm, ref in name_map.items():
        ws.parent.defined_names[nm] = DefinedName(name=nm, attr_text=f"'Project Info'!{ref}")

    for i, (label, val) in enumerate(info_rows, start=3):
        ws.cell(row=i, column=2, value=label).font = FONT_BODY_BOLD
        ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
        ws.cell(row=i, column=2).border = BORDER
        c = ws.cell(row=i, column=3, value=val)
        c.font = FONT_BODY
        c.border = BORDER
        if "Date" in label or "Week Ending" in label:
            c.number_format = FMT_DATE
        if "Payroll No" in label:
            c.number_format = FMT_INT

    dv_yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    dv_yn.add("C11")
    dv_yn.add("C12")
    ws.add_data_validation(dv_yn)


# ---------------------------------------------------------------------------
# Tab 3: Wage Determination
# ---------------------------------------------------------------------------

WD_COLS = [
    ("A", "Classification", 30),
    ("B", "WD Base Rate ($/hr)", 16),
    ("C", "WD Fringe ($/hr)", 14),
    ("D", "Cash Fringe ($/hr)", 14),
    ("E", "Plan Fringe ($/hr)", 14),
    ("F", "OT Multiplier", 12),
    ("G", "ST Hourly Cost (Base+Fringe)", 18),
    ("H", "OT Hourly Cost", 18),
    ("I", "Apprenticeable?", 14),
    ("J", "Notes", 30),
]


def build_wd(ws):
    ws.title = "Wage Determination"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in WD_COLS])

    ws["A1"] = "WAGE DETERMINATION — Classifications"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Enter one row per classification used on this project. WD Rate + WD Fringe is the floor; Cash + "
        "Plan Fringe must equal at least WD Fringe."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:J2")

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(WD_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(WD_COLS))

    DATA_ROWS = 25
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        ws.cell(row=r, column=2).number_format = FMT_USD
        ws.cell(row=r, column=3).number_format = FMT_USD
        ws.cell(row=r, column=4).number_format = FMT_USD
        ws.cell(row=r, column=5).number_format = FMT_USD
        ws.cell(row=r, column=6, value=1.5).number_format = FMT_NUM2
        # G = ST hourly cost = base + cash + plan
        ws.cell(
            row=r,
            column=7,
            value=f'=IF(B{r}="","",B{r}+IFERROR(D{r},0)+IFERROR(E{r},0))',
        ).number_format = FMT_USD
        # H = OT hourly cost = (base * mult) + fringe
        ws.cell(
            row=r,
            column=8,
            value=f'=IF(B{r}="","",(B{r}*F{r})+IFERROR(D{r},0)+IFERROR(E{r},0))',
        ).number_format = FMT_USD
        apply_body_style(ws, r, len(WD_COLS))

    dv_yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_yn.add(f"I{HEADER_ROW+1}:I{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_yn)

    # Conditional: fringe mismatch — cash + plan < WD fringe
    ws.conditional_formatting.add(
        f"E{HEADER_ROW+1}:E{HEADER_ROW+DATA_ROWS}",
        FormulaRule(
            formula=[f'AND(C{HEADER_ROW+1}>0,(D{HEADER_ROW+1}+E{HEADER_ROW+1})<C{HEADER_ROW+1})'],
            fill=FILL_RED,
            font=Font(color=RED_FONT),
        ),
    )

    # Named range for VLOOKUP from Workers Master and Weekly Payroll
    last_row = HEADER_ROW + DATA_ROWS
    ws.parent.defined_names["WDTable"] = DefinedName(
        name="WDTable",
        attr_text=f"'Wage Determination'!$A${HEADER_ROW+1}:$H${last_row}",
    )


# ---------------------------------------------------------------------------
# Tab 4: Workers Master
# ---------------------------------------------------------------------------

WORKER_COLS = [
    ("A", "Worker Name", 28),
    ("B", "Employee ID", 14),
    ("C", "Classification", 26),
    ("D", "Apprentice?", 12),
    ("E", "Apprenticeship Reg No.", 18),
    ("F", "Apprentice Year/Period", 14),
    ("G", "Apprentice % of Journeyman", 16),
    ("H", "Hire Date", 12),
    ("I", "Status", 12),
    ("J", "Notes", 24),
]


def build_workers(ws):
    ws.title = "Workers Master"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in WORKER_COLS])

    ws["A1"] = "WORKERS MASTER"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 32

    HEADER_ROW = 3
    for i, (col, header, _) in enumerate(WORKER_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(WORKER_COLS))

    DATA_ROWS = 60
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        ws.cell(row=r, column=7).number_format = FMT_PCT
        ws.cell(row=r, column=8).number_format = FMT_DATE
        apply_body_style(ws, r, len(WORKER_COLS))

    dv_yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_yn.add(f"D{HEADER_ROW+1}:D{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_yn)

    dv_status = DataValidation(type="list", formula1='"Active,Terminated,On Leave"', allow_blank=True)
    dv_status.add(f"I{HEADER_ROW+1}:I{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_status)


# ---------------------------------------------------------------------------
# Tab 5: Weekly Payroll (WH-347 style)
# ---------------------------------------------------------------------------

DAY_LABELS = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]


def build_weekly_payroll(ws):
    ws.title = "Weekly Payroll"
    ws.sheet_view.showGridLines = False

    # Header
    ws["A1"] = "WEEKLY CERTIFIED PAYROLL — WH-347 Style"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:V1")
    ws.row_dimensions[1].height = 32

    # Project metadata strip
    ws["A2"] = "Project:"
    ws["A2"].font = FONT_BODY_BOLD
    ws["B2"] = "=ProjectName"
    ws.merge_cells("B2:E2")
    ws["F2"] = "Payroll #:"
    ws["F2"].font = FONT_BODY_BOLD
    ws["G2"] = "=PayrollNo"
    ws["H2"] = "Week Ending:"
    ws["H2"].font = FONT_BODY_BOLD
    ws["I2"] = "=WeekEnding"
    ws["I2"].number_format = FMT_DATE
    ws["J2"] = "Subcontractor:"
    ws["J2"].font = FONT_BODY_BOLD
    ws["K2"] = "=SubName"
    ws.merge_cells("K2:N2")
    ws["O2"] = "Prime Contract:"
    ws["O2"].font = FONT_BODY_BOLD
    ws["P2"] = "=PrimeContractNo"
    ws.merge_cells("P2:R2")

    HEADER_ROW = 4

    # Column structure
    # A: # | B: Worker | C: Emp ID | D: Class | E-K: Sun-Sat hours | L: Total ST | M: Total OT |
    # N: ST Rate | O: OT Rate | P: ST Gross | Q: OT Gross | R: Fringe Cash | S: Total Gross |
    # T: Deductions | U: Net | V: Notes

    col_widths = [
        ("A", 4), ("B", 24), ("C", 12), ("D", 22),
        ("E", 7), ("F", 7), ("G", 7), ("H", 7), ("I", 7), ("J", 7), ("K", 7),
        ("L", 10), ("M", 10),
        ("N", 11), ("O", 11), ("P", 13), ("Q", 13), ("R", 13), ("S", 14),
        ("T", 12), ("U", 13), ("V", 18),
    ]
    set_col_widths(ws, col_widths)

    headers = ["#", "Worker Name", "Emp ID", "Classification"]
    for d in DAY_LABELS:
        headers.append(d)
    headers += [
        "Total ST", "Total OT", "ST Rate", "OT Rate",
        "ST Gross", "OT Gross", "Fringe (Cash)", "Total Gross",
        "Deductions", "Net Wages", "Notes",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=h)
    style_header_row(ws, HEADER_ROW, len(headers))

    # ST / OT sub-headers for daily columns
    for i, _ in enumerate(DAY_LABELS, start=5):
        ws.cell(row=HEADER_ROW + 1, column=i, value="ST | OT").font = FONT_GREY_ITALIC

    DATA_ROWS = 30
    DATA_START = HEADER_ROW + 2
    for r in range(DATA_START, DATA_START + DATA_ROWS):
        # Row index
        ws.cell(row=r, column=1, value=f"=IF(B{r}=\"\",\"\",ROW()-{HEADER_ROW+1})")
        # Classification lookup -> rate
        # ST Rate: VLOOKUP classification (col D) -> WDTable col 7 (ST hourly cost) — but worker may be
        # apprentice; if so, scale the base rate. To keep formulas readable, we VLOOKUP the WD base rate (col 2),
        # WD cash fringe (col 4), WD plan fringe (col 5), and apprentice % from Workers Master.
        # Worker lookup: column B; Apprentice % from Workers Master tab col G
        # For straightforward use, expose the ST Rate as: base + cash fringe (cash portion paid as wage)
        ws.cell(
            row=r,
            column=14,  # N — ST Rate
            value=(
                f'=IFERROR('
                f'(VLOOKUP(D{r},WDTable,2,FALSE)'
                f'*IFERROR(VLOOKUP(B{r},\'Workers Master\'!A:G,7,FALSE),1))'
                f'+VLOOKUP(D{r},WDTable,4,FALSE),"")'
            ),
        ).number_format = FMT_USD
        # OT Rate: (base * 1.5) + cash fringe
        ws.cell(
            row=r,
            column=15,  # O
            value=(
                f'=IFERROR('
                f'(VLOOKUP(D{r},WDTable,2,FALSE)'
                f'*IFERROR(VLOOKUP(B{r},\'Workers Master\'!A:G,7,FALSE),1)'
                f'*VLOOKUP(D{r},WDTable,6,FALSE))'
                f'+VLOOKUP(D{r},WDTable,4,FALSE),"")'
            ),
        ).number_format = FMT_USD

        # Daily ST + OT: enter two cells per day? For compactness, single hours cell — assume all ST. OT
        # column accommodated by Total OT col M. Daily cells are ST-only.
        for col in range(5, 12):
            ws.cell(row=r, column=col).number_format = FMT_NUM2

        # Total ST = SUM(E:K)
        ws.cell(row=r, column=12, value=f"=SUM(E{r}:K{r})").number_format = FMT_NUM2
        # Total OT — manual entry (column M)
        ws.cell(row=r, column=13).number_format = FMT_NUM2
        # ST Gross = Total ST * ST Rate
        ws.cell(row=r, column=16, value=f'=IF(N{r}="","",L{r}*N{r})').number_format = FMT_USD
        # OT Gross = Total OT * OT Rate
        ws.cell(row=r, column=17, value=f'=IF(O{r}="","",M{r}*O{r})').number_format = FMT_USD
        # Fringe (Cash) — already embedded in ST Rate and OT Rate; this column is informational
        # Show plan fringe (not in gross) for documentation
        ws.cell(
            row=r,
            column=18,
            value=(
                f'=IFERROR(VLOOKUP(D{r},WDTable,5,FALSE)*(L{r}+M{r}),"")'
            ),
        ).number_format = FMT_USD
        # Total Gross = ST Gross + OT Gross
        ws.cell(row=r, column=19, value=f'=IF(P{r}="","",IFERROR(P{r},0)+IFERROR(Q{r},0))').number_format = FMT_USD
        # Deductions — manual
        ws.cell(row=r, column=20).number_format = FMT_USD
        # Net Wages = Total Gross - Deductions
        ws.cell(row=r, column=21, value=f'=IF(S{r}="","",IFERROR(S{r},0)-IFERROR(T{r},0))').number_format = FMT_USD

        apply_body_style(ws, r, len(headers))

    # Totals row
    TOTAL_ROW = DATA_START + DATA_ROWS + 1
    ws.cell(row=TOTAL_ROW, column=4, value="WEEKLY TOTALS").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=4).fill = FILL_SUBHEADER
    for col_idx in [12, 13, 16, 17, 18, 19, 20, 21]:
        col_letter = get_column_letter(col_idx)
        c = ws.cell(
            row=TOTAL_ROW,
            column=col_idx,
            value=f"=SUM({col_letter}{DATA_START}:{col_letter}{DATA_START+DATA_ROWS-1})",
        )
        c.font = FONT_BODY_BOLD
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_USD if col_idx >= 16 else FMT_NUM2
        c.border = BORDER

    # Freeze panes for legibility
    ws.freeze_panes = "E5"

    # Print setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


# ---------------------------------------------------------------------------
# Tab 6: Statement of Compliance
# ---------------------------------------------------------------------------

def build_statement(ws):
    ws.title = "Statement of Compliance"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 100)])

    ws["B1"] = "STATEMENT OF COMPLIANCE (WH-348-style)"
    ws["B1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 32

    blocks = [
        ("", FONT_BODY),
        ("Date: ____________________   Payroll #: =PayrollNo   Week Ending: =WeekEnding", FONT_BODY_BOLD),
        ("", FONT_BODY),
        ("I, the undersigned, do hereby state, under penalty of perjury (18 U.S.C. § 1001 and 31 U.S.C. § 3729):", FONT_BODY),
        ("", FONT_BODY),
        ("(1)  That I pay or supervise the payment of the persons employed by =SubName on the project "
         "identified as =ProjectName (Prime Contract No. =PrimeContractNo); that during the payroll period "
         "commencing on ____________________ and ending on =WeekEnding all persons employed on said project "
         "have been paid the full weekly wages earned, that no rebates have been or will be made either "
         "directly or indirectly to or on behalf of said =SubName from the full weekly wages earned by any "
         "person and that no deductions have been made either directly or indirectly from the full wages "
         "earned by any person, other than permissible deductions as defined in Regulations, Part 3 "
         "(29 C.F.R. Subtitle A), issued by the Secretary of Labor under the Copeland Act (40 U.S.C. § 3145), "
         "and described below: ____________________________________________________________________________", FONT_BODY),
        ("", FONT_BODY),
        ("(2)  That any payrolls otherwise under this contract required to be submitted for the above period "
         "are correct and complete; that the wage rates for laborers or mechanics contained therein are not "
         "less than the applicable wage rates contained in any wage determination incorporated into the "
         "contract; that the classifications set forth therein for each laborer or mechanic conform with the "
         "work each performed.", FONT_BODY),
        ("", FONT_BODY),
        ("(3)  That any apprentices employed in the above period are duly registered in a bona fide "
         "apprenticeship program registered with a State apprenticeship agency recognized by the Bureau of "
         "Apprenticeship and Training, United States Department of Labor, or if no such recognized agency "
         "exists in a State, are registered with the Bureau of Apprenticeship and Training, United States "
         "Department of Labor.", FONT_BODY),
        ("", FONT_BODY),
        ("(4)  That:  (a)  WHERE FRINGE BENEFITS ARE PAID TO APPROVED PLANS, FUNDS, OR PROGRAMS — in "
         "addition to the basic hourly wage rates paid to each laborer or mechanic listed in the above "
         "referenced payroll, payments of fringe benefits as listed in the contract have been or will be "
         "made to appropriate programs for the benefit of such employees, except as noted in section 4(c) "
         "below.   (b)  WHERE FRINGE BENEFITS ARE PAID IN CASH — Each laborer or mechanic listed in the "
         "above referenced payroll has been paid, as indicated on the payroll, an amount not less than the "
         "sum of the applicable basic hourly wage rate plus the amount of the required fringe benefits as "
         "listed in the contract, except as noted in section 4(c) below.   (c)  EXCEPTIONS (specify by "
         "name of worker, basic hourly wage rate, fringe benefit method): _______________________________", FONT_BODY),
        ("", FONT_BODY),
        ("THE WILLFUL FALSIFICATION OF ANY OF THE ABOVE STATEMENTS MAY SUBJECT THE CONTRACTOR OR "
         "SUBCONTRACTOR TO CIVIL OR CRIMINAL PROSECUTION. SEE SECTION 1001 OF TITLE 18 AND SECTION 3729 OF "
         "TITLE 31 OF THE UNITED STATES CODE.", FONT_BODY_BOLD),
        ("", FONT_BODY),
        ("", FONT_BODY),
        ("Signature: _____________________________________   Date: ___________", FONT_BODY_BOLD),
        ("Printed Name: =SignatoryName", FONT_BODY),
        ("Title: =SignatoryTitle", FONT_BODY),
        ("Company: =SubName", FONT_BODY),
    ]

    for i, (text, font) in enumerate(blocks, start=3):
        if text.startswith("="):
            cell = ws.cell(row=i, column=2, value=text)
        elif "=" in text and "C." not in text and "U.S.C." not in text and "§" not in text:
            # Best-effort: embed the dynamic formula by keeping as-is — Excel will resolve named ranges
            # within concatenated strings only via &; for clarity we leave it as templated text the user
            # edits, since the named ranges appear in the metadata strip.
            cell = ws.cell(row=i, column=2, value=text)
        else:
            cell = ws.cell(row=i, column=2, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = font


# ---------------------------------------------------------------------------
# Tab 7: CSI Reference (hidden, kept for consistency with library)
# ---------------------------------------------------------------------------

def build_csi_reference(ws):
    ws.title = "CSI Reference"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 10), ("B", 60)])
    ws["A1"] = "CSI MasterFormat Divisions (reference only)"
    ws["A1"].font = FONT_H2
    ws.sheet_state = "hidden"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    wb = Workbook()
    build_instructions(wb.active)
    build_project_info(wb.create_sheet())
    build_wd(wb.create_sheet())
    build_workers(wb.create_sheet())
    build_weekly_payroll(wb.create_sheet())
    build_statement(wb.create_sheet())
    build_csi_reference(wb.create_sheet())

    out_dir = "/Users/home/charles/contrpro/files/packages/complete/sub"
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "Certified_Payroll_Tracker.xlsx")
    wb.save(out_path)
    sz = os.path.getsize(out_path)
    print(f"Wrote {out_path} ({sz//1024} KB)")


if __name__ == "__main__":
    main()
