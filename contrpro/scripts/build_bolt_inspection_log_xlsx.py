#!/usr/bin/env python3
"""
Build ContrPro Bolt Inspection Log (XLSX) — Steel Erection trade pack.

Companion to bolt-installation-and-inspection-guide.html. Captures bolt lot
metadata (MTR, grade, mfr, lot #), pre-installation Skidmore verification,
per-connection installation + inspection records, and a daily roll-up.

Tabs:
  1. Instructions
  2. Project Info
  3. Bolt Lot Master         (one row per lot — MTR, mfr, grade, Skidmore date)
  4. Pre-Installation Verification Log  (Skidmore tests, one row per test)
  5. Connection Log           (per connection — location, method, count, status, inspector)
  6. Daily Roll-Up            (totals by day + by inspector)
  7. CSI Reference            (hidden)

Run:
    /Users/home/charles/.venv/bin/python3 \\
        /Users/home/charles/contrpro/scripts/build_bolt_inspection_log_xlsx.py

Output:
    /Users/home/charles/contrpro/files/packages/complete/steel-erection/Bolt_Inspection_Log.xlsx
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
YELLOW_FILL = "FFEB9C"
YELLOW_FONT = "9C5700"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_SUMMARY_LABEL = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW_FILL)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_BIG_NUMBER = Font(name="Calibri", size=18, bold=True, color=BRAND_BLUE)
FONT_GREY_ITALIC = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_INT = "0"
FMT_NUM2 = "0.00"
FMT_DATE = "yyyy-mm-dd"
FMT_KIPS = '#,##0.0" kips"'

# RCSC Table 8.1 minimum pretension (kips) by bolt size + grade
RCSC_PRETENSION = [
    ("1/2",  "A325", 12.0),
    ("5/8",  "A325", 19.0),
    ("3/4",  "A325", 28.0),
    ("7/8",  "A325", 39.0),
    ("1",    "A325", 51.0),
    ("1-1/8","A325", 56.0),
    ("1-1/4","A325", 71.0),
    ("1-3/8","A325", 85.0),
    ("1-1/2","A325", 103.0),
    ("1/2",  "A490", 15.0),
    ("5/8",  "A490", 24.0),
    ("3/4",  "A490", 35.0),
    ("7/8",  "A490", 49.0),
    ("1",    "A490", 64.0),
    ("1-1/8","A490", 80.0),
    ("1-1/4","A490", 102.0),
    ("1-3/8","A490", 121.0),
    ("1-1/2","A490", 148.0),
]


def set_col_widths(ws, widths):
    for col, w in widths:
        ws.column_dimensions[col].width = w


def style_header_row(ws, row, cols, start_col=1):
    for c in range(start_col, start_col + cols):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 28


def apply_body_style(ws, row, cols, start_col=1):
    for c in range(start_col, start_col + cols):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_BODY
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.border = BORDER


# ---------------------------------------------------------------------------
# Tab 1: Instructions
# ---------------------------------------------------------------------------

def build_instructions(ws):
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 110)])

    ws["A1"] = "BOLT INSPECTION LOG — Instructions"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 36

    rows = [
        "",
        "WHAT THIS WORKBOOK IS",
        "Daily record-keeping companion to the Bolt Installation & Inspection Guide HTML. Captures the "
        "documentation the special inspector (and any post-completion forensic) will demand: bolt lot "
        "metadata, pre-installation Skidmore verification, per-connection install records, and a "
        "daily roll-up.",
        "",
        "WORKFLOW — ONE-TIME SETUP",
        "1. Project Info: project, GC, erector, special inspector, Skidmore serial.",
        "2. Bolt Lot Master: one row per bolt lot received from the fabricator/supplier. MTR # is the "
        "key — it must match the paperwork from the supplier and must be on file before any bolt "
        "from that lot is installed.",
        "",
        "WORKFLOW — DAILY",
        "1. First thing each shift, perform Skidmore pre-installation verification per RCSC §7.",
        "2. Log the verification on Pre-Installation Verification Log. Three bolt-nut-washer assemblies "
        "from each active lot, all reading ≥ 105% of RCSC Table 8.1 minimum pretension.",
        "3. As bolts are installed during the shift, log each connection on Connection Log (location, "
        "method used, bolt lot, count, installer, status).",
        "4. After installation, mark installed bolts (typically with a paint slash from bolt head to ply) "
        "so missed bolts are visually obvious to the special inspector.",
        "5. Inspector verifies installation — logs sign-off on Connection Log.",
        "6. End of shift: Daily Roll-Up tab auto-aggregates totals.",
        "",
        "RCSC TABLE 8.1 — MINIMUM PRETENSION (REFERENCE)",
        "  3/4\" A325: 28 kips     7/8\" A325: 39 kips     1\" A325: 51 kips",
        "  3/4\" A490: 35 kips     7/8\" A490: 49 kips     1\" A490: 64 kips",
        "  See the Bolt Lot Master tab for full table by size + grade.",
        "  Skidmore target = 105% of these minimums.",
        "",
        "INSPECTION SAMPLING (RCSC §9.2.2 / Table 9.2)",
        "  Slip-critical and fatigue-loaded: 100% routine observation during installation.",
        "  Standard pretensioned: reduced sampling per Special Inspector's program (typically 10-25%).",
        "  Snug-tight: visual only — no pretension verification required.",
        "",
        "MARKING CONVENTION",
        "  Standard color marking on installed-and-inspected bolts:",
        "  - Yellow paint slash = installed, awaiting inspection",
        "  - Green paint slash = installed + inspected, accepted",
        "  - Red paint slash = rejected — bolt to be replaced",
        "  Use the convention agreed with the special inspector at the pre-erection meeting.",
        "",
        "RETENTION",
        "Retain this workbook + all MTRs + Skidmore calibration certificate + Special Inspector reports "
        "for the project life PLUS the state's statute of repose for construction defect claims (typically "
        "6-10 years post-substantial completion).",
    ]

    for i, text in enumerate(rows, start=2):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if text and text == text.upper() and len(text) < 70:
            cell.font = FONT_H2
        else:
            cell.font = FONT_BODY


# ---------------------------------------------------------------------------
# Tab 2: Project Info
# ---------------------------------------------------------------------------

def build_project_info(ws):
    ws.title = "Project Info"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 42), ("C", 36)])

    ws["B1"] = "BOLT INSPECTION PROJECT INFO"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:C1")
    ws.row_dimensions[1].height = 32

    info_rows = [
        ("Project Name", ""),
        ("Project Address", ""),
        ("Owner", ""),
        ("Controlling Contractor (GC)", ""),
        ("Steel Erector", ""),
        ("Fabricator", ""),
        ("Special Inspection Agency (SAA)", ""),
        ("Lead Special Inspector (AWS CWI #)", ""),
        ("EOR", ""),
        ("Skidmore-Wilhelm serial #", ""),
        ("Skidmore calibration certificate date", ""),
        ("Skidmore calibration expiration", ""),
        ("Project Quality Manager (Sub)", ""),
        ("Erection start date", ""),
        ("Anticipated bolt installation completion", ""),
    ]

    name_map = {
        "ProjectName": "$C$3",
        "Erector": "$C$7",
        "Inspector": "$C$10",
        "SkidmoreSerial": "$C$12",
    }
    for nm, ref in name_map.items():
        ws.parent.defined_names[nm] = DefinedName(name=nm, attr_text=f"'Project Info'!{ref}")

    for i, (label, val) in enumerate(info_rows, start=3):
        ws.cell(row=i, column=2, value=label).font = FONT_BODY_BOLD
        ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
        ws.cell(row=i, column=2).border = BORDER
        c = ws.cell(row=i, column=3, value=val)
        c.font = FONT_BODY
        c.border = BORDER
        if "date" in label.lower() or "expiration" in label.lower():
            c.number_format = FMT_DATE


# ---------------------------------------------------------------------------
# Tab 3: Bolt Lot Master
# ---------------------------------------------------------------------------

LOT_COLS = [
    ("A", "Lot #", 12),
    ("B", "Bolt Size", 10),
    ("C", "Grade", 10),
    ("D", "Length (in)", 11),
    ("E", "Finish", 16),
    ("F", "Manufacturer", 22),
    ("G", "MTR / Cert #", 16),
    ("H", "Date Received", 13),
    ("I", "RCSC Min Pretension (kips)", 18),
    ("J", "Skidmore Target (105% kips)", 18),
    ("K", "Pre-Install Verified?", 16),
    ("L", "Last Verification Date", 14),
    ("M", "Quantity Received", 14),
    ("N", "Quantity Installed", 14),
    ("O", "Notes", 22),
]


def build_lot_master(ws):
    ws.title = "Bolt Lot Master"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in LOT_COLS])

    ws["A1"] = "BOLT LOT MASTER"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:O1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "One row per bolt lot received. RCSC minimum pretension auto-fills based on size + grade — "
        "Skidmore target is 105% of that minimum."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:O2")

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(LOT_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(LOT_COLS))

    DATA_ROWS = 30
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        ws.cell(row=r, column=4).number_format = FMT_NUM2
        ws.cell(row=r, column=8).number_format = FMT_DATE
        # RCSC min pretension — SUMIFS lookup against PretensionSize/Grade/Kips
        # named ranges. 2026-05-17 audit fix: previous INDEX/MATCH with concatenated
        # column refs was an array-formula construct that worked in Excel 365 but
        # failed silently in older Excel / LibreOffice / Numbers / Sheets. SUMIFS
        # works as a regular formula in all spreadsheet versions from 2007 onward.
        ws.cell(
            row=r,
            column=9,
            value=(
                f'=IF(OR(B{r}="",C{r}=""),"",'
                f'IFERROR(SUMIFS(PretensionKips,PretensionSize,B{r},PretensionGrade,C{r}),""))'
            ),
        ).number_format = FMT_KIPS
        # Skidmore target = 105% of min
        ws.cell(row=r, column=10, value=f'=IFERROR(I{r}*1.05,"")').number_format = FMT_KIPS
        ws.cell(row=r, column=12).number_format = FMT_DATE
        ws.cell(row=r, column=13).number_format = FMT_INT
        ws.cell(row=r, column=14).number_format = FMT_INT
        apply_body_style(ws, r, len(LOT_COLS))

    # Validations
    sizes = ['1/2','5/8','3/4','7/8','1','1-1/8','1-1/4','1-3/8','1-1/2']
    dv_size = DataValidation(type="list", formula1=f'"{",".join(sizes)}"', allow_blank=True)
    dv_size.add(f"B{HEADER_ROW+1}:B{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_size)

    dv_grade = DataValidation(type="list", formula1='"A325,A490,F1852,F2280,F3148"', allow_blank=True)
    dv_grade.add(f"C{HEADER_ROW+1}:C{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_grade)

    dv_finish = DataValidation(type="list", formula1='"Plain,Mechanically Galvanized,Hot-Dip Galvanized,Weathering,Zinc/Aluminum Coated"', allow_blank=True)
    dv_finish.add(f"E{HEADER_ROW+1}:E{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_finish)

    dv_verified = DataValidation(type="list", formula1='"Yes,No,Expired"', allow_blank=True)
    dv_verified.add(f"K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_verified)

    # Conditional format on Verified column
    ws.conditional_formatting.add(
        f"K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'K{HEADER_ROW+1}="Yes"'], fill=FILL_GREEN, font=Font(color=GREEN_FONT)),
    )
    ws.conditional_formatting.add(
        f"K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'K{HEADER_ROW+1}="No"'], fill=FILL_RED, font=Font(color=RED_FONT)),
    )
    ws.conditional_formatting.add(
        f"K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'K{HEADER_ROW+1}="Expired"'], fill=FILL_YELLOW, font=Font(color=YELLOW_FONT)),
    )

    # Named range for VLOOKUP from Connection Log
    last_row = HEADER_ROW + DATA_ROWS
    ws.parent.defined_names["LotTable"] = DefinedName(
        name="LotTable",
        attr_text=f"'Bolt Lot Master'!$A${HEADER_ROW+1}:$O${last_row}",
    )


# ---------------------------------------------------------------------------
# Tab 4: Pre-Installation Verification Log
# ---------------------------------------------------------------------------

VERIF_COLS = [
    ("A", "Date", 12),
    ("B", "Time", 10),
    ("C", "Lot #", 12),
    ("D", "Bolt Size", 10),
    ("E", "Grade", 10),
    ("F", "Method", 18),
    ("G", "Test #1 Pretension (kips)", 16),
    ("H", "Test #2 Pretension (kips)", 16),
    ("I", "Test #3 Pretension (kips)", 16),
    ("J", "Target (105% min)", 14),
    ("K", "Pass / Fail", 12),
    ("L", "Ironworker", 18),
    ("M", "Inspector", 18),
    ("N", "Notes", 20),
]


def build_verification_log(ws):
    ws.title = "Pre-Install Verification"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in VERIF_COLS])

    ws["A1"] = "PRE-INSTALLATION VERIFICATION LOG"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:N1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "RCSC §7. Test 3 bolt-nut-washer assemblies from each active lot at start of shift OR when "
        "anything changes. All 3 readings must meet or exceed the 105% target."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:N2")

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(VERIF_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(VERIF_COLS))

    DATA_ROWS = 100
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        ws.cell(row=r, column=1).number_format = FMT_DATE
        for c in (7, 8, 9):
            ws.cell(row=r, column=c).number_format = FMT_KIPS
        # Target = VLOOKUP lot, get target column (J = col 10) from Bolt Lot Master
        ws.cell(
            row=r,
            column=10,
            value=f'=IFERROR(VLOOKUP(C{r},LotTable,10,FALSE),"")',
        ).number_format = FMT_KIPS
        # Pass/Fail: all 3 tests >= target
        ws.cell(
            row=r,
            column=11,
            value=(
                f'=IF(OR(G{r}="",H{r}="",I{r}="",J{r}=""),"",'
                f'IF(AND(G{r}>=J{r},H{r}>=J{r},I{r}>=J{r}),"Pass","Fail"))'
            ),
        )
        ws.cell(row=r, column=11).font = FONT_BODY_BOLD
        apply_body_style(ws, r, len(VERIF_COLS))

    dv_method = DataValidation(
        type="list",
        formula1='"Turn-of-Nut,Calibrated Wrench,Twist-Off (TC),DTI Washer"',
        allow_blank=True,
    )
    dv_method.add(f"F{HEADER_ROW+1}:F{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_method)

    # Conditional format Pass/Fail
    ws.conditional_formatting.add(
        f"K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'K{HEADER_ROW+1}="Pass"'], fill=FILL_GREEN, font=Font(bold=True, color=GREEN_FONT)),
    )
    ws.conditional_formatting.add(
        f"K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'K{HEADER_ROW+1}="Fail"'], fill=FILL_RED, font=Font(bold=True, color=RED_FONT)),
    )


# ---------------------------------------------------------------------------
# Tab 5: Connection Log
# ---------------------------------------------------------------------------

CONN_COLS = [
    ("A", "Date", 12),
    ("B", "Connection ID", 14),
    ("C", "Location (Gridline/Floor)", 22),
    ("D", "Connection Type", 18),
    ("E", "Method", 18),
    ("F", "Lot #", 12),
    ("G", "# Bolts", 9),
    ("H", "Installer / Crew", 18),
    ("I", "Snug-Tight Complete", 16),
    ("J", "Final Tighten Complete", 16),
    ("K", "Marked (color)", 13),
    ("L", "Inspector Sign-Off", 16),
    ("M", "Status", 14),
    ("N", "Slip-Critical Faying OK?", 16),
    ("O", "Notes", 22),
]


def build_connection_log(ws):
    ws.title = "Connection Log"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in CONN_COLS])

    ws["A1"] = "CONNECTION INSTALLATION + INSPECTION LOG"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:O1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "One row per connection. Status auto-rolls: Snug → Tightened → Inspected → Accepted / Rejected. "
        "Slip-critical column N is only relevant for SC connections — verify faying surface class."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:O2")

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(CONN_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(CONN_COLS))

    DATA_ROWS = 300
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        ws.cell(row=r, column=1).number_format = FMT_DATE
        ws.cell(row=r, column=7).number_format = FMT_INT
        # Status — derive from check-marks
        ws.cell(
            row=r,
            column=13,
            value=(
                f'=IF(B{r}="","",'
                f'IF(L{r}="Accepted","Accepted",'
                f'IF(L{r}="Rejected","Rejected",'
                f'IF(J{r}="Yes","Awaiting Inspection",'
                f'IF(I{r}="Yes","Snug-Tight","Pending")))))'
            ),
        )
        ws.cell(row=r, column=13).font = FONT_BODY_BOLD
        apply_body_style(ws, r, len(CONN_COLS))

    dv_ctype = DataValidation(
        type="list",
        formula1='"Snug-Tight,Pretensioned,Slip-Critical"',
        allow_blank=True,
    )
    dv_ctype.add(f"D{HEADER_ROW+1}:D{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_ctype)

    dv_method = DataValidation(
        type="list",
        formula1='"N/A — Snug,Turn-of-Nut,Calibrated Wrench,Twist-Off (TC),DTI Washer"',
        allow_blank=True,
    )
    dv_method.add(f"E{HEADER_ROW+1}:E{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_method)

    dv_yn = DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=True)
    dv_yn.add(f"I{HEADER_ROW+1}:I{HEADER_ROW+DATA_ROWS}")
    dv_yn.add(f"J{HEADER_ROW+1}:J{HEADER_ROW+DATA_ROWS}")
    dv_yn.add(f"N{HEADER_ROW+1}:N{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_yn)

    dv_color = DataValidation(type="list", formula1='"Yellow,Green,Red,Not Marked"', allow_blank=True)
    dv_color.add(f"K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_color)

    dv_inspect = DataValidation(type="list", formula1='"Pending,Accepted,Rejected,Re-inspect"', allow_blank=True)
    dv_inspect.add(f"L{HEADER_ROW+1}:L{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_inspect)

    # Conditional format on Status
    ws.conditional_formatting.add(
        f"M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'M{HEADER_ROW+1}="Accepted"'], fill=FILL_GREEN, font=Font(bold=True, color=GREEN_FONT)),
    )
    ws.conditional_formatting.add(
        f"M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'M{HEADER_ROW+1}="Rejected"'], fill=FILL_RED, font=Font(bold=True, color=RED_FONT)),
    )
    ws.conditional_formatting.add(
        f"M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'M{HEADER_ROW+1}="Awaiting Inspection"'], fill=FILL_YELLOW, font=Font(color=YELLOW_FONT)),
    )

    # Totals row
    TOTAL_ROW = HEADER_ROW + DATA_ROWS + 1
    ws.cell(row=TOTAL_ROW, column=2, value="TOTALS").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=2).fill = FILL_SUBHEADER
    ws.cell(row=TOTAL_ROW, column=7, value=f"=SUM(G{HEADER_ROW+1}:G{HEADER_ROW+DATA_ROWS})").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=7).number_format = FMT_INT
    ws.cell(row=TOTAL_ROW, column=7).fill = FILL_SUBHEADER
    ws.cell(row=TOTAL_ROW, column=7).border = BORDER


# ---------------------------------------------------------------------------
# Tab 6: Daily Roll-Up
# ---------------------------------------------------------------------------

def build_daily_rollup(ws):
    ws.title = "Daily Roll-Up"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 38), ("C", 18), ("D", 28)])

    ws["B1"] = "DAILY ROLL-UP — Bolt Installation Status"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:D1")
    ws.row_dimensions[1].height = 32

    ws["B2"] = "Auto-aggregated from Connection Log. Use for daily progress reporting + monthly billing back-up."
    ws["B2"].font = FONT_GREY_ITALIC
    ws.merge_cells("B2:D2")

    rows = [
        ("Total connections logged",        '=COUNTA(\'Connection Log\'!B5:B304)'),
        ("Total bolts installed",            "=SUM('Connection Log'!G5:G304)"),
        ("Connections — Pending",            '=COUNTIF(\'Connection Log\'!M5:M304,"Pending")'),
        ("Connections — Snug-Tight only",    '=COUNTIF(\'Connection Log\'!M5:M304,"Snug-Tight")'),
        ("Connections — Awaiting Inspection", '=COUNTIF(\'Connection Log\'!M5:M304,"Awaiting Inspection")'),
        ("Connections — Accepted",           '=COUNTIF(\'Connection Log\'!M5:M304,"Accepted")'),
        ("Connections — Rejected",           '=COUNTIF(\'Connection Log\'!M5:M304,"Rejected")'),
        ("Connections — Re-inspect",         '=COUNTIF(\'Connection Log\'!L5:L304,"Re-inspect")'),
        ("", ""),
        ("By Connection Type", ""),
        ("Snug-Tight type connections",      '=COUNTIF(\'Connection Log\'!D5:D304,"Snug-Tight")'),
        ("Pretensioned type connections",    '=COUNTIF(\'Connection Log\'!D5:D304,"Pretensioned")'),
        ("Slip-Critical type connections",   '=COUNTIF(\'Connection Log\'!D5:D304,"Slip-Critical")'),
        ("", ""),
        ("By Method", ""),
        ("Turn-of-Nut",                       '=COUNTIF(\'Connection Log\'!E5:E304,"Turn-of-Nut")'),
        ("Calibrated Wrench",                 '=COUNTIF(\'Connection Log\'!E5:E304,"Calibrated Wrench")'),
        ("Twist-Off (TC)",                    '=COUNTIF(\'Connection Log\'!E5:E304,"Twist-Off (TC)")'),
        ("DTI Washer",                         '=COUNTIF(\'Connection Log\'!E5:E304,"DTI Washer")'),
        ("", ""),
        ("Skidmore verifications logged",    '=COUNTA(\'Pre-Install Verification\'!C5:C104)'),
        ("Skidmore verifications PASSED",    '=COUNTIF(\'Pre-Install Verification\'!K5:K104,"Pass")'),
        ("Skidmore verifications FAILED",    '=COUNTIF(\'Pre-Install Verification\'!K5:K104,"Fail")'),
    ]

    start_row = 4
    for i, (label, formula) in enumerate(rows, start=start_row):
        is_subhead = label and not formula
        if is_subhead:
            ws.cell(row=i, column=2, value=label).font = FONT_H2
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
            continue
        if label:
            ws.cell(row=i, column=2, value=label).font = FONT_BODY_BOLD
            ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
            ws.cell(row=i, column=2).border = BORDER
            c = ws.cell(row=i, column=3, value=formula)
            c.font = FONT_BIG_NUMBER if "Total" in label else FONT_BODY_BOLD
            c.number_format = FMT_INT
            c.border = BORDER


# ---------------------------------------------------------------------------
# Tab 7: CSI Reference / RCSC Pretension Table (hidden but used as VLOOKUP source)
# ---------------------------------------------------------------------------

def build_csi_reference(ws):
    ws.title = "Reference"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 10), ("B", 10), ("C", 14)])

    ws["A1"] = "RCSC Table 8.1 — Min Pretension (kips)"
    ws["A1"].font = FONT_H2
    ws.merge_cells("A1:C1")

    headers = ["Size", "Grade", "Min Pretension (kips)"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=h)
        ws.cell(row=3, column=i).font = FONT_HEADER
        ws.cell(row=3, column=i).fill = FILL_HEADER
        ws.cell(row=3, column=i).alignment = Alignment(horizontal="center")
        ws.cell(row=3, column=i).border = BORDER

    for i, (size, grade, kips) in enumerate(RCSC_PRETENSION, start=4):
        ws.cell(row=i, column=1, value=size).font = FONT_BODY
        ws.cell(row=i, column=2, value=grade).font = FONT_BODY
        ws.cell(row=i, column=3, value=kips).font = FONT_BODY
        ws.cell(row=i, column=3).number_format = FMT_NUM2
        for col in (1, 2, 3):
            ws.cell(row=i, column=col).border = BORDER

    # Named ranges. PretensionTable is the legacy 3-column block (kept for any
    # downstream INDEX/MATCH consumers); the three single-column ranges drive
    # the portable SUMIFS lookup used in Bolt Lot Master (audit fix 2026-05-17).
    last_row = 3 + len(RCSC_PRETENSION)
    ws.parent.defined_names["PretensionTable"] = DefinedName(
        name="PretensionTable",
        attr_text=f"'Reference'!$A$4:$C${last_row}",
    )
    ws.parent.defined_names["PretensionSize"] = DefinedName(
        name="PretensionSize",
        attr_text=f"'Reference'!$A$4:$A${last_row}",
    )
    ws.parent.defined_names["PretensionGrade"] = DefinedName(
        name="PretensionGrade",
        attr_text=f"'Reference'!$B$4:$B${last_row}",
    )
    ws.parent.defined_names["PretensionKips"] = DefinedName(
        name="PretensionKips",
        attr_text=f"'Reference'!$C$4:$C${last_row}",
    )

    ws.sheet_state = "hidden"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    wb = Workbook()
    build_instructions(wb.active)
    build_project_info(wb.create_sheet())
    build_csi_reference(wb.create_sheet())  # build reference first so named ranges exist for VLOOKUP
    build_lot_master(wb.create_sheet())
    build_verification_log(wb.create_sheet())
    build_connection_log(wb.create_sheet())
    build_daily_rollup(wb.create_sheet())

    out_dir = "/Users/home/charles/contrpro/files/packages/complete/steel-erection"
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "Bolt_Inspection_Log.xlsx")
    wb.save(out_path)
    sz = os.path.getsize(out_path)
    print(f"Wrote {out_path} ({sz//1024} KB)")


if __name__ == "__main__":
    main()
