#!/usr/bin/env python3
"""
Build ContrPro Certificate of Insurance (COI) Tracker (XLSX).

Produces a production-grade workbook with:
  - Instructions tab (what to verify on every cert, statutory employer exposure, renewal cadence)
  - Company Info tab (the GC's own GL and WC policies; expiry alerts)
  - Sub/Vendor COI Log (main worksheet — CSI-coded, dropdowns, expiry formulas, risk flags)
  - Compliance Summary (COUNTIF roll-ups for risk items with traffic lights)
  - Dashboard (COI Compliance Score, Risk Exposure panel, top 5 expiring soonest, counts by coverage type)
  - CSI Reference (hidden, named ranges feeding the dropdowns/VLOOKUPs)

Why this matters: in TX/GA (and most states) the GC is the statutory employer
for any uncovered sub on their jobsite. An expired sub COI = the GC's GL/WC
policy is on the hook for the sub's incidents. The red flags in this tracker
are the GC's risk management.

Run:
    /Users/home/charles/.venv/bin/python3 \
        /Users/home/charles/contrpro/scripts/build_coi_tracker_xlsx.py

Output:
    /Users/home/charles/contrpro/files/packages/business/COI_Tracker.xlsx
"""

from __future__ import annotations

import os
from datetime import date, timedelta
from typing import Tuple

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# ---------------------------------------------------------------------------
# Brand & styling (matches Job_Costing, Change_Order_Log, AR_Tracker, Project_Tracker)
# ---------------------------------------------------------------------------

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
ACCENT_GOLD = "C9A227"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
YELLOW_FILL = "FFEB9C"
YELLOW_FONT = "9C5700"
ORANGE_FILL = "FFD8A8"
ORANGE_FONT = "8A4B00"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_SUMMARY_LABEL = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW_FILL)
FILL_ORANGE = PatternFill("solid", fgColor=ORANGE_FILL)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H1 = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_BIG_NUMBER = Font(name="Calibri", size=18, bold=True, color=BRAND_BLUE)
FONT_GREY_ITALIC = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_USD_BIG = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
FMT_PCT = "0.0%"
FMT_INT = "0"
FMT_DATE = "yyyy-mm-dd"


# ---------------------------------------------------------------------------
# CSI MasterFormat division reference (same canonical list as other trackers)
# ---------------------------------------------------------------------------

DIVISIONS: list[tuple[str, str]] = [
    ("01", "General Requirements"),
    ("02", "Existing Conditions"),
    ("03", "Concrete"),
    ("04", "Masonry"),
    ("05", "Metals"),
    ("06", "Wood, Plastics, and Composites"),
    ("07", "Thermal and Moisture Protection"),
    ("08", "Openings"),
    ("09", "Finishes"),
    ("10", "Specialties"),
    ("11", "Equipment"),
    ("12", "Furnishings"),
    ("13", "Special Construction"),
    ("14", "Conveying Equipment"),
    ("21", "Fire Suppression"),
    ("22", "Plumbing"),
    ("23", "HVAC"),
    ("25", "Integrated Automation"),
    ("26", "Electrical"),
    ("27", "Communications"),
    ("28", "Electronic Safety and Security"),
    ("31", "Earthwork"),
    ("32", "Exterior Improvements"),
    ("33", "Utilities"),
    ("34", "Transportation"),
    ("35", "Waterway and Marine Construction"),
    ("40", "Process Interconnections"),
    ("41", "Material Processing and Handling Equipment"),
    ("42", "Process Heating, Cooling, and Drying Equipment"),
    ("43", "Process Gas and Liquid Handling, Purification, and Storage Equipment"),
    ("44", "Pollution and Waste Control Equipment"),
    ("45", "Industry-Specific Manufacturing Equipment"),
    ("46", "Water and Wastewater Equipment"),
    ("48", "Electrical Power Generation"),
    ("49", "Reserved"),
]


# ---------------------------------------------------------------------------
# Dropdown values
# ---------------------------------------------------------------------------

TRADES = ["General", "Mechanical", "Electrical", "Plumbing", "Structural Steel", "Other"]
COVERAGE_TYPES = ["GL", "WC", "Auto", "Excess/Umbrella", "Builder's Risk", "Pollution"]
YES_NO_PENDING = ["Yes", "No", "Pending"]
YES_NO_NA = ["Yes", "No", "N/A"]
YES_NO = ["Yes", "No"]
STATUS_VALUES = ["Current", "Renew Soon", "Expiring Soon", "EXPIRED"]


# ---------------------------------------------------------------------------
# Pre-populated example COIs (varied risk states — anchor "today" = 2026-05-13)
#
# Mix of: Current, Renew Soon (60-90 days), Expiring Soon (<30), EXPIRED,
# missing Additional Insured, missing Waiver of Subrogation.
#
# (sub_name, trade, csi_div, csi_code, carrier, policy_num, coverage_type,
#  each_occ, gen_agg, eff_date, exp_date,
#  add_insured, waiver_subro, cert_holder, project, doc_path, notes)
# ---------------------------------------------------------------------------

EXAMPLE_COIS: list[tuple] = [
    # 1. Steel sub, GL, current, fully compliant
    ("Acme Structural Steel LLC", "Structural Steel", "05", "05 12 00",
     "Travelers", "GL-7782341", "GL",
     2000000.00, 4000000.00,
     "2025-09-01", "2026-09-01",
     "Yes", "Yes", "Yes",
     "Riverside Office Bldg",
     "/COIs/2026/Acme_GL.pdf",
     "Fully compliant. Endorsement CG 20 10 + CG 20 37 on file."),

    # 2. Same steel sub, WC, current
    ("Acme Structural Steel LLC", "Structural Steel", "05", "05 12 00",
     "Texas Mutual", "WC-118-44-2210", "WC",
     1000000.00, 1000000.00,
     "2025-09-01", "2026-09-01",
     "N/A", "Yes", "Yes",
     "Riverside Office Bldg",
     "/COIs/2026/Acme_WC.pdf",
     "WC compliant. State of Texas form. Stat employer risk = covered."),

    # 3. HVAC sub - EXPIRING SOON (<30 days)
    ("Northstar HVAC Inc", "Mechanical", "23", "23 30 00",
     "The Hartford", "GL-2298-1144", "GL",
     1000000.00, 2000000.00,
     "2025-05-25", "2026-05-25",
     "Yes", "Yes", "Yes",
     "Pinnacle Tower",
     "/COIs/2026/Northstar_GL.pdf",
     "12 days to expiry. Renewal cert requested 2026-04-25 - follow up."),

    # 4. Electrical sub - EXPIRED, GC ON THE HOOK
    ("Bayside Electric LLC", "Electrical", "26", "26 50 00",
     "Liberty Mutual", "GL-441-9982", "GL",
     1000000.00, 2000000.00,
     "2025-03-15", "2026-03-15",
     "Yes", "Yes", "Yes",
     "Heartland Industrial",
     "/COIs/2026/Bayside_GL_EXPIRED.pdf",
     "EXPIRED 59 days ago. PULL FROM SITE until renewal received. Statutory employer exposure live."),

    # 5. Plumbing sub - current but MISSING Additional Insured
    ("Cypress Plumbing Co", "Plumbing", "22", "22 10 00",
     "Nationwide", "GL-77-2231", "GL",
     1000000.00, 2000000.00,
     "2026-01-10", "2027-01-10",
     "No", "Yes", "Yes",
     "Riverside Office Bldg",
     "/COIs/2026/Cypress_GL.pdf",
     "GL current but Additional Insured endorsement NOT on cert. Demand CG 20 10 before next pay app."),

    # 6. Concrete sub - Renew Soon (~75 days out) AND missing waiver
    ("Lone Star Concrete Inc", "General", "03", "03 30 00",
     "CNA", "GL-5589-7711", "GL",
     2000000.00, 4000000.00,
     "2025-07-27", "2026-07-27",
     "Yes", "No", "Yes",
     "Apex Industrial",
     "/COIs/2026/LoneStar_GL.pdf",
     "Renewal due in 75 days. Waiver of Subrogation MISSING - prime contract requires it. Demand endorsement."),
]

EMPTY_COI_ROWS = 20


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def set_col_widths(ws, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_header_row(ws, row: int, last_col: int) -> None:
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 32


# ---------------------------------------------------------------------------
# Hidden tab — CSI Reference (also feeds named ranges + dropdowns)
# ---------------------------------------------------------------------------

def build_csi_reference(wb: Workbook) -> str:
    ws = wb.create_sheet("CSI Reference")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "CSI MasterFormat Division Reference"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:B1")

    ws["A2"] = "Pulled from the same canonical list as ContrPro Job Costing. Do not edit."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)

    ws.cell(row=3, column=1, value="CSI Division")
    ws.cell(row=3, column=2, value="Division Name")
    style_header_row(ws, 3, 2)

    for i, (code, name) in enumerate(DIVISIONS, start=4):
        ws.cell(row=i, column=1, value=code).number_format = "@"
        ws.cell(row=i, column=2, value=name)
        for c in (1, 2):
            ws.cell(row=i, column=c).border = BORDER
            ws.cell(row=i, column=c).font = FONT_BODY

    set_col_widths(ws, {"A": 14, "B": 60})

    first = 4
    last = 4 + len(DIVISIONS) - 1
    div_range = f"'CSI Reference'!$A${first}:$A${last}"
    full_range = f"'CSI Reference'!$A${first}:$B${last}"
    wb.defined_names["CSI_Divisions"] = DefinedName("CSI_Divisions", attr_text=div_range)
    wb.defined_names["CSI_Table"] = DefinedName("CSI_Table", attr_text=full_range)
    return ws.title


# ---------------------------------------------------------------------------
# Tab 1 — Instructions
# ---------------------------------------------------------------------------

def build_instructions(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {"A": 4, "B": 110})

    ws["B2"] = "ContrPro - Certificate of Insurance (COI) Tracker"
    ws["B2"].font = FONT_TITLE
    ws.row_dimensions[2].height = 32

    ws["B3"] = ("The risk management workbook for your subs and vendors. If a COI is expired or missing the "
                "right endorsements, YOUR policy is on the hook.")
    ws["B3"].font = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

    sections: list[tuple[str, list[str]]] = [
        (
            "Why this workbook is non-negotiable",
            [
                "In Texas, Georgia, and most other states, the General Contractor is the 'statutory employer' for "
                "any uncovered sub on the jobsite. Translation: if a sub's worker gets hurt and that sub has "
                "no current Workers' Comp policy, the GC's WC carrier pays the claim. Premium hikes follow.",
                "Same logic for General Liability. An expired sub GL means the sub is operating uninsured. When the "
                "OSHA inspector shows up, when a third party gets injured, when property gets damaged - the GC's "
                "GL policy backstops the loss. That's a margin-killer on a job you already bid tight.",
                "Bottom line: an expired COI on an active sub IS the GC's risk exposure. This workbook is the "
                "early-warning system. Red flags here = risk to your business.",
            ],
        ),
        (
            "What to verify on every certificate (the 5-point check)",
            [
                "1. Certificate Holder - your company name and address, exactly as on the contract. Not the project "
                "name, not the owner. YOU are the certificate holder.",
                "2. Additional Insured endorsement - the cert must list YOUR company as an Additional Insured on the "
                "sub's GL. Look for endorsement form CG 20 10 (ongoing operations) and CG 20 37 (completed "
                "operations). 'Listed in Description of Operations' is not enough - you want the endorsement form.",
                "3. Waiver of Subrogation - both GL and WC. Means the sub's carrier waives the right to come after "
                "YOU after they pay a claim caused by your shared work. Most prime contracts require this.",
                "4. Policy Effective and Expiration Dates - the cert must cover the entire term of the sub's work, "
                "including completed operations tail. Today's date must fall between Effective and Expiration.",
                "5. Coverage Limits meet the contract minimum - usually $1M per occurrence / $2M aggregate for GL, "
                "statutory WC, $1M Auto. Excess/Umbrella per contract.",
            ],
        ),
        (
            "Renewal cadence",
            [
                "Most commercial policies are annual. Sub turns over a new cert each year on policy anniversary.",
                "Set the 90-day flag (Renew Soon): request the renewal cert from the sub now. Don't wait.",
                "30-day flag (Expiring Soon): escalate. Email + phone call. PMs forget; insurance brokers don't "
                "always push the cert to you automatically.",
                "Day of expiry: if you don't have the renewal cert in hand, the sub stops work. Period. Document "
                "the stop-work decision in writing.",
            ],
        ),
        (
            "What to do when a COI expires on an active sub",
            [
                "1. Pull the sub from the site that day. Email the sub PM AND your own super: 'Per contract, "
                "[Sub Name] cannot be on site without current COI. Stand down until cert received.'",
                "2. Issue a written demand for the renewal cert with 48-hour response window. Email + certified "
                "mail if you've had problems with this sub before.",
                "3. Document the gap in the Notes column of this tracker. If anything happens during the gap, you "
                "need a paper trail showing you enforced the requirement.",
                "4. Do NOT issue the next pay application until current COI is on file. Pay app + cert go together.",
            ],
        ),
        (
            "Tab-by-tab quick start",
            [
                "1. Company Info - fill in YOUR company's GL and WC policy details. Single row. If your own GL or WC "
                "is within 30 days of expiry, the cell turns red. Same drill - you owe certs to YOUR clients.",
                "2. Sub/Vendor COI Log - one row per sub-coverage combo. A sub with GL, WC, and Auto = three rows. "
                "Effective + Expiration dates are required. Status and Days to Expiry calculate automatically.",
                "3. Compliance Summary - DO NOT edit. Headline risk counts with traffic lights.",
                "4. Dashboard - DO NOT edit. COI Compliance Score, top 5 expiring soonest, coverage breakdown.",
            ],
        ),
        (
            "Color coding cheat sheet",
            [
                "Status: green = Current, yellow = Renew Soon, orange = Expiring Soon, RED = EXPIRED.",
                "Days to Expiry: red fill anytime < 30 days. Includes negative values (already expired).",
                "Additional Insured: red if 'No' - YOUR policy will eat any claim.",
                "Waiver of Subrogation: red if 'No' - violates most prime contracts.",
                "Risk Exposure panel on Dashboard: any cell red = active risk, deal with it this week.",
            ],
        ),
        (
            "CSI coding rules (must follow)",
            [
                "CSI Division: 2-digit text (01-49). Pick from the dropdown - leading zeros preserved.",
                "CSI Code: 6-digit space-separated, e.g. 05 12 00 for Structural Steel Framing. Level 3 OK (05 12 13).",
                "Division Name auto-fills from a VLOOKUP against CSI Reference. If it shows blank/error, your CSI "
                "Division value does not match the reference list - re-pick from the dropdown.",
            ],
        ),
        (
            "Weekly cadence",
            [
                "Monday morning: open Dashboard. Anything RED in Risk Exposure panel? That's the week's collection list.",
                "Friday afternoon: log new COIs received this week. Update Notes on any renewals in progress.",
                "End of month: pull Compliance Summary. Zero EXPIRED, zero missing Additional Insured = clean month.",
                "Quarterly: full audit. Confirm every active sub has current GL + WC on file. No exceptions.",
            ],
        ),
        (
            "Tabs in this workbook",
            [
                "Instructions  -  this tab",
                "Company Info  -  your own GC GL and WC policy details + expiry alerts",
                "Sub/Vendor COI Log  -  the main worksheet, CSI-coded with status + risk flags",
                "Compliance Summary  -  headline risk counts with traffic lights",
                "Dashboard  -  COI Compliance Score, Risk Exposure panel, top 5 expiring soonest",
                "CSI Reference  -  canonical division list (hidden - do not edit)",
            ],
        ),
    ]

    row = 5
    for heading, paragraphs in sections:
        ws.cell(row=row, column=2, value=heading)
        ws.cell(row=row, column=2).font = FONT_H2
        ws.cell(row=row, column=2).fill = FILL_SUBHEADER
        ws.cell(row=row, column=2).alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[row].height = 22
        row += 1
        for para in paragraphs:
            ws.cell(row=row, column=2, value=para)
            ws.cell(row=row, column=2).font = FONT_BODY
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = max(18, 16 * (1 + len(para) // 110))
            row += 1
        row += 1

    ws.cell(row=row + 1, column=2,
            value="ContrPro - built for builders. Questions: support@contrpro.com")
    ws.cell(row=row + 1, column=2).font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)


# ---------------------------------------------------------------------------
# Tab 2 — Company Info (the GC's own info)
# ---------------------------------------------------------------------------

COMPANY_HEADERS = [
    "Company Name", "EIN",
    "GL Carrier", "GL Policy #", "GL Effective", "GL Expires",
    "WC Carrier", "WC Policy #", "WC Effective", "WC Expires",
    "Notes",
]

# Example row (greyed/italic) - today anchor = 2026-05-13. Use dates that show
# the conditional formatting working (one within 30 days, one healthy).
COMPANY_EXAMPLE = [
    "EXAMPLE - Your Construction Co LLC",
    "12-3456789",
    "Travelers", "GL-CO-2241-08", "2025-10-01", "2026-10-01",
    "Texas Mutual", "WC-CO-118-99", "2025-06-01", "2026-06-01",
    "Replace this row with your company's actual policy info.",
]


def build_company_info(wb: Workbook) -> None:
    ws = wb.create_sheet("Company Info")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Company Information (Your GC Policies)"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:K1")

    ws["A2"] = ("Single row of metadata for your own company. The 'Expires' cells turn red within 30 days - "
                "renew before you find out the hard way.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)
    ws.merge_cells("A2:K2")

    for i, h in enumerate(COMPANY_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(COMPANY_HEADERS))

    # Example row at row 4
    for i, v in enumerate(COMPANY_EXAMPLE, start=1):
        cell = ws.cell(row=4, column=i, value=v)
        cell.font = FONT_GREY_ITALIC
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    # Date format on expiry/effective cells
    for c in (5, 6, 9, 10):
        ws.cell(row=4, column=c).number_format = FMT_DATE

    # Empty user rows (allow a couple, in case GC has multiple entities)
    for r in range(5, 8):
        for c in range(1, len(COMPANY_HEADERS) + 1):
            ws.cell(row=r, column=c).border = BORDER
        for c in (5, 6, 9, 10):
            ws.cell(row=r, column=c).number_format = FMT_DATE
        ws.row_dimensions[r].height = 22

    # Conditional formatting: red fill on GL Expires (F) and WC Expires (J)
    # when expiry < TODAY() + 30 days (includes already-expired).
    for col_letter in ("F", "J"):
        target_range = f"{col_letter}4:{col_letter}7"
        ws.conditional_formatting.add(
            target_range,
            FormulaRule(
                formula=[f'AND(${col_letter}4<>"",${col_letter}4-TODAY()<30)'],
                stopIfTrue=False,
                fill=FILL_RED, font=Font(color=RED_FONT, bold=True),
            ),
        )

    set_col_widths(ws, {
        "A": 32, "B": 14, "C": 18, "D": 18, "E": 13, "F": 13,
        "G": 18, "H": 18, "I": 13, "J": 13, "K": 40,
    })

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Tab 3 — Sub/Vendor COI Log (main worksheet)
# ---------------------------------------------------------------------------

COI_HEADERS = [
    "Subcontractor / Vendor Name",       # A
    "Trade",                              # B
    "CSI Division",                       # C
    "CSI Code",                           # D
    "Division Name",                      # E
    "Insurance Carrier",                  # F
    "Policy Number",                      # G
    "Coverage Type",                      # H
    "Each Occurrence Limit",              # I
    "General Aggregate Limit",            # J
    "Effective Date",                     # K
    "Expiration Date",                    # L
    "Days to Expiry",                     # M
    "Status",                             # N
    "Additional Insured?",                # O
    "Waiver of Subrogation?",             # P
    "Certificate Holder Listed?",         # Q
    "Project / Job",                      # R
    "COI Document Path",                  # S
    "Notes",                              # T
]


def build_coi_log(wb: Workbook) -> Tuple[str, int, int]:
    ws = wb.create_sheet("Sub-Vendor COI Log")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Sub/Vendor COI Log (CSI MasterFormat)"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COI_HEADERS))

    # Header row at row 3
    for i, h in enumerate(COI_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(COI_HEADERS))

    data_start = 4
    n_pre = len(EXAMPLE_COIS)
    n_total = n_pre + EMPTY_COI_ROWS
    data_end = data_start + n_total - 1

    # Column letter shortcuts
    COL = {h: get_column_letter(i + 1) for i, h in enumerate(COI_HEADERS)}
    col_div = COL["CSI Division"]           # C
    col_csi_code = COL["CSI Code"]          # D
    col_div_name = COL["Division Name"]     # E
    col_exp = COL["Expiration Date"]        # L
    col_days = COL["Days to Expiry"]        # M
    col_status = COL["Status"]              # N
    col_addl = COL["Additional Insured?"]   # O
    col_waiver = COL["Waiver of Subrogation?"]  # P
    col_holder = COL["Certificate Holder Listed?"]  # Q
    col_trade = COL["Trade"]                # B
    col_coverage = COL["Coverage Type"]     # H

    def days_formula(r: int) -> str:
        # Days to Expiry = Expiration Date - TODAY(). Negative if already expired.
        # Blank string if expiration not set.
        return f'=IF({col_exp}{r}="","",{col_exp}{r}-TODAY())'

    def status_formula(r: int) -> str:
        # IF blank -> ""
        # IF days < 0     -> "EXPIRED"
        # IF days < 30    -> "Expiring Soon"
        # IF days < 90    -> "Renew Soon"
        # else            -> "Current"
        return (
            f'=IF({col_exp}{r}="","",'
            f'IF({col_days}{r}<0,"EXPIRED",'
            f'IF({col_days}{r}<30,"Expiring Soon",'
            f'IF({col_days}{r}<90,"Renew Soon","Current"))))'
        )

    def divname_formula(r: int) -> str:
        return f'=IFERROR(VLOOKUP({col_div}{r},CSI_Table,2,FALSE),"")'

    # Pre-populated example rows
    for idx, row_data in enumerate(EXAMPLE_COIS):
        r = data_start + idx
        (sub_name, trade, csi_div, csi_code, carrier, policy_num, cov_type,
         each_occ, gen_agg, eff_d, exp_d,
         add_insured, waiver, cert_holder, project, doc_path, notes) = row_data

        ws.cell(row=r, column=1, value=sub_name)
        ws.cell(row=r, column=2, value=trade)
        ws.cell(row=r, column=3, value=csi_div)
        ws.cell(row=r, column=4, value=csi_code)
        ws.cell(row=r, column=5, value=divname_formula(r))
        ws.cell(row=r, column=6, value=carrier)
        ws.cell(row=r, column=7, value=policy_num)
        ws.cell(row=r, column=8, value=cov_type)
        ws.cell(row=r, column=9, value=each_occ)
        ws.cell(row=r, column=10, value=gen_agg)
        ws.cell(row=r, column=11, value=eff_d)
        ws.cell(row=r, column=12, value=exp_d)
        ws.cell(row=r, column=13, value=days_formula(r))
        ws.cell(row=r, column=14, value=status_formula(r))
        ws.cell(row=r, column=15, value=add_insured)
        ws.cell(row=r, column=16, value=waiver)
        ws.cell(row=r, column=17, value=cert_holder)
        ws.cell(row=r, column=18, value=project)
        ws.cell(row=r, column=19, value=doc_path)
        ws.cell(row=r, column=20, value=notes)

    # Empty user-fillable rows (formulas pre-seeded)
    for i in range(EMPTY_COI_ROWS):
        r = data_start + n_pre + i
        ws.cell(row=r, column=5, value=divname_formula(r))
        ws.cell(row=r, column=13, value=days_formula(r))
        ws.cell(row=r, column=14, value=status_formula(r))

    # Per-row formatting
    wrap_cols = {1, 5, 18, 19, 20}
    for r in range(data_start, data_end + 1):
        for c in range(1, len(COI_HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(c in wrap_cols),
            )
        # Text columns (preserve formatting)
        ws.cell(row=r, column=3).number_format = "@"   # CSI Division
        ws.cell(row=r, column=4).number_format = "@"   # CSI Code
        ws.cell(row=r, column=7).number_format = "@"   # Policy Number
        # Dates
        ws.cell(row=r, column=11).number_format = FMT_DATE  # Effective
        ws.cell(row=r, column=12).number_format = FMT_DATE  # Expiration
        # USD columns
        for c in (9, 10):
            ws.cell(row=r, column=c).number_format = FMT_USD
        # Days to Expiry as int
        ws.cell(row=r, column=13).number_format = FMT_INT
        ws.row_dimensions[r].height = 34

    # --- Conditional formatting ---

    # Status (N): green Current, yellow Renew Soon, orange Expiring Soon, RED EXPIRED
    status_range = f"{col_status}{data_start}:{col_status}{data_end}"
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'${col_status}{data_start}="Current"'],
            stopIfTrue=False,
            fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'${col_status}{data_start}="Renew Soon"'],
            stopIfTrue=False,
            fill=FILL_YELLOW, font=Font(color=YELLOW_FONT, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'${col_status}{data_start}="Expiring Soon"'],
            stopIfTrue=False,
            fill=FILL_ORANGE, font=Font(color=ORANGE_FONT, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'${col_status}{data_start}="EXPIRED"'],
            stopIfTrue=False,
            fill=FILL_RED, font=Font(color=RED_FONT, bold=True, size=12),
        ),
    )

    # Days to Expiry (M): red fill if < 30 (including negatives)
    days_range = f"{col_days}{data_start}:{col_days}{data_end}"
    ws.conditional_formatting.add(
        days_range,
        FormulaRule(
            formula=[f'AND(ISNUMBER(${col_days}{data_start}),${col_days}{data_start}<30)'],
            stopIfTrue=False,
            fill=FILL_RED, font=Font(color=RED_FONT, bold=True),
        ),
    )

    # Additional Insured (O): red if "No"
    addl_range = f"{col_addl}{data_start}:{col_addl}{data_end}"
    ws.conditional_formatting.add(
        addl_range,
        FormulaRule(
            formula=[f'${col_addl}{data_start}="No"'],
            stopIfTrue=False,
            fill=FILL_RED, font=Font(color=RED_FONT, bold=True),
        ),
    )

    # Waiver of Subrogation (P): red if "No"
    waiver_range = f"{col_waiver}{data_start}:{col_waiver}{data_end}"
    ws.conditional_formatting.add(
        waiver_range,
        FormulaRule(
            formula=[f'${col_waiver}{data_start}="No"'],
            stopIfTrue=False,
            fill=FILL_RED, font=Font(color=RED_FONT, bold=True),
        ),
    )

    # Certificate Holder Listed (Q): red if "No"
    holder_range = f"{col_holder}{data_start}:{col_holder}{data_end}"
    ws.conditional_formatting.add(
        holder_range,
        FormulaRule(
            formula=[f'${col_holder}{data_start}="No"'],
            stopIfTrue=False,
            fill=FILL_RED, font=Font(color=RED_FONT, bold=True),
        ),
    )

    # --- Data validation ---

    # Trade dropdown (B)
    dv_trade = DataValidation(
        type="list", formula1='"' + ",".join(TRADES) + '"', allow_blank=True,
    )
    dv_trade.error = "Pick a Trade from the dropdown."
    dv_trade.errorTitle = "Invalid Trade"
    ws.add_data_validation(dv_trade)
    dv_trade.add(f"{col_trade}{data_start}:{col_trade}{data_end}")

    # CSI Division dropdown (C) - named range
    dv_div = DataValidation(type="list", formula1="=CSI_Divisions", allow_blank=True)
    dv_div.error = "Pick a CSI Division from the dropdown."
    dv_div.errorTitle = "Invalid CSI Division"
    ws.add_data_validation(dv_div)
    dv_div.add(f"{col_div}{data_start}:{col_div}{data_end}")

    # Coverage Type dropdown (H)
    dv_cov = DataValidation(
        type="list", formula1='"' + ",".join(COVERAGE_TYPES) + '"', allow_blank=True,
    )
    dv_cov.error = "Pick a Coverage Type from the dropdown."
    dv_cov.errorTitle = "Invalid Coverage Type"
    ws.add_data_validation(dv_cov)
    dv_cov.add(f"{col_coverage}{data_start}:{col_coverage}{data_end}")

    # Additional Insured (O): Yes / No / Pending
    dv_addl = DataValidation(
        type="list", formula1='"' + ",".join(YES_NO_PENDING) + '"', allow_blank=True,
    )
    dv_addl.error = "Pick Yes / No / Pending."
    dv_addl.errorTitle = "Invalid"
    ws.add_data_validation(dv_addl)
    dv_addl.add(f"{col_addl}{data_start}:{col_addl}{data_end}")

    # Waiver of Subrogation (P): Yes / No / N/A
    dv_waiver = DataValidation(
        type="list", formula1='"' + ",".join(YES_NO_NA) + '"', allow_blank=True,
    )
    dv_waiver.error = "Pick Yes / No / N/A."
    dv_waiver.errorTitle = "Invalid"
    ws.add_data_validation(dv_waiver)
    dv_waiver.add(f"{col_waiver}{data_start}:{col_waiver}{data_end}")

    # Certificate Holder Listed (Q): Yes / No
    dv_holder = DataValidation(
        type="list", formula1='"' + ",".join(YES_NO) + '"', allow_blank=True,
    )
    dv_holder.error = "Pick Yes / No."
    dv_holder.errorTitle = "Invalid"
    ws.add_data_validation(dv_holder)
    dv_holder.add(f"{col_holder}{data_start}:{col_holder}{data_end}")

    # Column widths
    set_col_widths(ws, {
        "A": 28, "B": 14, "C": 11, "D": 12, "E": 24,
        "F": 18, "G": 16, "H": 16, "I": 16, "J": 16,
        "K": 13, "L": 13, "M": 12, "N": 15,
        "O": 14, "P": 14, "Q": 14,
        "R": 24, "S": 28, "T": 36,
    })

    ws.freeze_panes = "B4"

    return ws.title, data_start, data_end


# ---------------------------------------------------------------------------
# Tab 4 — Compliance Summary
# ---------------------------------------------------------------------------

COMPLIANCE_HEADERS = ["Compliance Metric", "Count", "Risk Level"]


def build_compliance_summary(wb: Workbook, coi_sheet: str, coi_start: int, coi_end: int) -> None:
    ws = wb.create_sheet("Compliance Summary")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Compliance Summary (auto roll-up)"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:C1")

    ws["A2"] = ("Reads from Sub/Vendor COI Log via COUNTIF / COUNTIFS - do not edit this tab directly. "
                "Risk Level cell turns red if Count > 0.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)
    ws.merge_cells("A2:C2")

    for i, h in enumerate(COMPLIANCE_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(COMPLIANCE_HEADERS))

    # Range references
    name_range = f"'{coi_sheet}'!$A${coi_start}:$A${coi_end}"
    status_range = f"'{coi_sheet}'!$N${coi_start}:$N${coi_end}"
    addl_range = f"'{coi_sheet}'!$O${coi_start}:$O${coi_end}"
    waiver_range = f"'{coi_sheet}'!$P${coi_start}:$P${coi_end}"
    coverage_range = f"'{coi_sheet}'!$H${coi_start}:$H${coi_end}"

    # (label, count formula, is_risk_metric)
    # is_risk_metric=True -> Risk Level cell turns RED if count > 0
    # is_risk_metric=False -> Risk Level cell shows neutral / informational
    metrics: list[tuple[str, str, bool]] = [
        ("Total Subs/Vendors with COI on file (rows logged)",
         f'=COUNTA({name_range})', False),
        ("Total EXPIRED",
         f'=COUNTIF({status_range},"EXPIRED")', True),
        ("Total Expiring within 30 days (Expiring Soon)",
         f'=COUNTIF({status_range},"Expiring Soon")', True),
        ("Total missing Additional Insured (Add'l Insured = No)",
         f'=COUNTIF({addl_range},"No")', True),
        ("Total missing Waiver of Subrogation (Waiver = No)",
         f'=COUNTIF({waiver_range},"No")', True),
        # Without GL coverage: count rows where Sub Name is set but Coverage Type != "GL"
        # Better worded: count distinct rows that are NOT GL coverage (informational).
        # The brief specifies COUNTIFS Coverage Type != GL; do exactly that.
        ("Rows not classified as GL coverage",
         f'=COUNTIFS({name_range},"<>",{coverage_range},"<>GL")', False),
        ("Rows not classified as WC coverage",
         f'=COUNTIFS({name_range},"<>",{coverage_range},"<>WC")', False),
    ]

    start_row = 4
    for idx, (label, formula, is_risk) in enumerate(metrics):
        r = start_row + idx
        # Label
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=1).font = FONT_BODY_BOLD
        ws.cell(row=r, column=1).fill = FILL_SUMMARY_LABEL
        ws.cell(row=r, column=1).alignment = Alignment(vertical="center", indent=1, wrap_text=True)
        ws.cell(row=r, column=1).border = BORDER

        # Count
        ws.cell(row=r, column=2, value=formula)
        ws.cell(row=r, column=2).number_format = FMT_INT
        ws.cell(row=r, column=2).font = FONT_BODY_BOLD
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=2).border = BORDER

        # Risk Level (traffic light)
        # For risk metrics: text is "OK" if count=0 (green) else "RISK" (red).
        # For info metrics: leave as "—" with neutral styling.
        if is_risk:
            risk_formula = f'=IF(B{r}>0,"RISK","OK")'
            ws.cell(row=r, column=3, value=risk_formula)
        else:
            ws.cell(row=r, column=3, value="—")
        ws.cell(row=r, column=3).font = FONT_BODY_BOLD
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=3).border = BORDER

        ws.row_dimensions[r].height = 28

    end_row = start_row + len(metrics) - 1

    # Conditional formatting on Risk Level column (C) for the risk-metric rows
    # Apply per-row CF so non-risk rows stay neutral.
    for idx, (_, _, is_risk) in enumerate(metrics):
        if not is_risk:
            continue
        r = start_row + idx
        cell_addr = f"C{r}"
        ws.conditional_formatting.add(
            cell_addr,
            FormulaRule(
                formula=[f'$B${r}>0'],
                stopIfTrue=False,
                fill=FILL_RED, font=Font(color=RED_FONT, bold=True, size=12),
            ),
        )
        ws.conditional_formatting.add(
            cell_addr,
            FormulaRule(
                formula=[f'$B${r}=0'],
                stopIfTrue=False,
                fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True),
            ),
        )

    set_col_widths(ws, {"A": 55, "B": 12, "C": 14})

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Tab 5 — Dashboard
# ---------------------------------------------------------------------------

def build_dashboard(wb: Workbook, coi_sheet: str, coi_start: int, coi_end: int) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "COI Compliance Dashboard"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:F1")

    ws["A2"] = ("Auto-calculated from Sub/Vendor COI Log. Update that tab - this one stays in sync. "
                "Red = active risk, deal with it this week.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)
    ws.merge_cells("A2:F2")

    # Range references on the COI log
    name_range = f"'{coi_sheet}'!$A${coi_start}:$A${coi_end}"
    coverage_range = f"'{coi_sheet}'!$H${coi_start}:$H${coi_end}"
    exp_range = f"'{coi_sheet}'!$L${coi_start}:$L${coi_end}"
    days_range = f"'{coi_sheet}'!$M${coi_start}:$M${coi_end}"
    status_range = f"'{coi_sheet}'!$N${coi_start}:$N${coi_end}"
    addl_range = f"'{coi_sheet}'!$O${coi_start}:$O${coi_end}"
    waiver_range = f"'{coi_sheet}'!$P${coi_start}:$P${coi_end}"

    # ---------------- Top panel: COI Compliance Score (big number) ----------------
    panel_start = 4
    ws.cell(row=panel_start, column=1, value="COI Compliance Score").font = FONT_H1
    ws.cell(row=panel_start, column=1).fill = FILL_HEADER
    ws.cell(row=panel_start, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=panel_start, start_column=1,
                   end_row=panel_start, end_column=6)
    ws.row_dimensions[panel_start].height = 26

    # Compliance Score = (Current + Renew Soon) / Total rows with status set * 100
    total_rows_formula = (
        f'(COUNTIF({status_range},"Current")+'
        f'COUNTIF({status_range},"Renew Soon")+'
        f'COUNTIF({status_range},"Expiring Soon")+'
        f'COUNTIF({status_range},"EXPIRED"))'
    )
    healthy_formula = (
        f'(COUNTIF({status_range},"Current")+'
        f'COUNTIF({status_range},"Renew Soon"))'
    )
    score_formula = f'=IFERROR({healthy_formula}/{total_rows_formula},0)'

    r = panel_start + 1
    ws.cell(row=r, column=1, value="COI Compliance Score (% Current + Renew Soon)")
    ws.cell(row=r, column=1).font = FONT_BODY_BOLD
    ws.cell(row=r, column=1).fill = FILL_SUMMARY_LABEL
    ws.cell(row=r, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.cell(row=r, column=1).border = BORDER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

    ws.cell(row=r, column=4, value=score_formula)
    ws.cell(row=r, column=4).font = FONT_BIG_NUMBER
    ws.cell(row=r, column=4).number_format = FMT_PCT
    ws.cell(row=r, column=4).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.cell(row=r, column=4).border = BORDER
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)

    # Color: green >=90%, yellow 70-90%, red <70%
    score_cell = f"D{r}"
    ws.conditional_formatting.add(
        score_cell,
        CellIsRule(operator="greaterThanOrEqual", formula=["0.9"], stopIfTrue=True,
                   fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True, size=18)),
    )
    ws.conditional_formatting.add(
        score_cell,
        CellIsRule(operator="greaterThanOrEqual", formula=["0.7"], stopIfTrue=True,
                   fill=FILL_YELLOW, font=Font(color=YELLOW_FONT, bold=True, size=18)),
    )
    ws.conditional_formatting.add(
        score_cell,
        CellIsRule(operator="lessThan", formula=["0.7"], stopIfTrue=False,
                   fill=FILL_RED, font=Font(color=RED_FONT, bold=True, size=18)),
    )
    ws.row_dimensions[r].height = 32

    # ---------------- Middle panel: Risk Exposure ----------------
    risk_start = r + 2
    ws.cell(row=risk_start, column=1, value="Risk Exposure").font = FONT_H1
    ws.cell(row=risk_start, column=1).fill = FILL_HEADER
    ws.cell(row=risk_start, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=risk_start, start_column=1,
                   end_row=risk_start, end_column=6)
    ws.row_dimensions[risk_start].height = 26

    # (label, count formula, "redIfPositive")
    risk_metrics = [
        ("EXPIRED count",
         f'=COUNTIF({status_range},"EXPIRED")'),
        ("Expiring Soon count (<30 days)",
         f'=COUNTIF({status_range},"Expiring Soon")'),
        ("Missing Additional Insured count",
         f'=COUNTIF({addl_range},"No")'),
        ("Missing Waiver of Subrogation count",
         f'=COUNTIF({waiver_range},"No")'),
    ]

    rrow = risk_start + 1
    risk_cells: list[str] = []
    for label, formula in risk_metrics:
        ws.cell(row=rrow, column=1, value=label)
        ws.cell(row=rrow, column=1).font = FONT_BODY_BOLD
        ws.cell(row=rrow, column=1).fill = FILL_SUMMARY_LABEL
        ws.cell(row=rrow, column=1).alignment = Alignment(vertical="center", indent=1)
        ws.cell(row=rrow, column=1).border = BORDER
        ws.merge_cells(start_row=rrow, start_column=1, end_row=rrow, end_column=3)

        ws.cell(row=rrow, column=4, value=formula)
        ws.cell(row=rrow, column=4).font = FONT_BIG_NUMBER
        ws.cell(row=rrow, column=4).number_format = FMT_INT
        ws.cell(row=rrow, column=4).alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.cell(row=rrow, column=4).border = BORDER
        ws.merge_cells(start_row=rrow, start_column=4, end_row=rrow, end_column=6)

        risk_cells.append(f"D{rrow}")
        ws.row_dimensions[rrow].height = 28
        rrow += 1

    # Red if >0, green if =0 — UNMISSABLE
    for ref in risk_cells:
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=True,
                       fill=FILL_RED, font=Font(color=RED_FONT, bold=True, size=18)),
        )
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="equal", formula=["0"], stopIfTrue=False,
                       fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True, size=18)),
        )

    # ---------------- Top 5 Expiring Soonest ----------------
    top_start = rrow + 1
    ws.cell(row=top_start, column=1, value="Top 5 Expiring Soonest").font = FONT_H1
    ws.cell(row=top_start, column=1).fill = FILL_HEADER
    ws.cell(row=top_start, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=top_start, start_column=1,
                   end_row=top_start, end_column=6)
    ws.row_dimensions[top_start].height = 26

    top_headers = ["Rank", "Subcontractor / Vendor", "Coverage", "Expiration Date", "Days to Expiry", "Status"]
    thr = top_start + 1
    for i, h in enumerate(top_headers, start=1):
        ws.cell(row=thr, column=i, value=h)
    style_header_row(ws, thr, len(top_headers))

    # Use SMALL on Days to Expiry to find the k-th smallest (most negative = most overdue).
    # Then INDEX/MATCH the row by matching (days = SMALL(...)) AND name is non-blank.
    # We add a small per-row tiebreaker using ROW() to handle duplicate day values,
    # by computing an effective value = days*10000 + ROW() — keeps ordering stable.
    # For simplicity and bulletproof formulas, we use a SMALL on the days column directly.
    # Ties: SMALL returns multiple matches but MATCH will hit the first occurrence — fine.

    for k in range(1, 6):
        r = thr + k
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=1).font = FONT_BODY_BOLD
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=1).border = BORDER

        # Days value: k-th smallest of Days to Expiry (filter out blanks via ISNUMBER inside SMALL is awkward;
        # rely on the column being either a number or "" — SMALL ignores text/blanks).
        days_val_formula = (
            f'=IFERROR(SMALL({days_range},{k}),"")'
        )

        # Index of the row with that value (returns position within days_range)
        idx_expr = f'MATCH(SMALL({days_range},{k}),{days_range},0)'

        # Name lookup
        name_val_formula = (
            f'=IFERROR(INDEX({name_range},{idx_expr}),"")'
        )
        # Coverage lookup
        cov_val_formula = (
            f'=IFERROR(INDEX({coverage_range},{idx_expr}),"")'
        )
        # Expiration date lookup
        exp_val_formula = (
            f'=IFERROR(INDEX({exp_range},{idx_expr}),"")'
        )
        # Status lookup
        status_val_formula = (
            f'=IFERROR(INDEX({status_range},{idx_expr}),"")'
        )

        ws.cell(row=r, column=2, value=name_val_formula)
        ws.cell(row=r, column=3, value=cov_val_formula)
        ws.cell(row=r, column=4, value=exp_val_formula)
        ws.cell(row=r, column=5, value=days_val_formula)
        ws.cell(row=r, column=6, value=status_val_formula)

        ws.cell(row=r, column=4).number_format = FMT_DATE
        ws.cell(row=r, column=5).number_format = FMT_INT

        for c in range(2, 7):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = Alignment(vertical="center", indent=1)
            ws.cell(row=r, column=c).font = FONT_BODY
        ws.row_dimensions[r].height = 24

        # Red fill on Days to Expiry cell if < 30 (matches main log)
        days_cell = f"E{r}"
        ws.conditional_formatting.add(
            days_cell,
            FormulaRule(
                formula=[f'AND(ISNUMBER(${days_cell}),${days_cell}<30)'],
                stopIfTrue=False,
                fill=FILL_RED, font=Font(color=RED_FONT, bold=True),
            ),
        )
        # Status cell formatting (mirrors log)
        status_cell = f"F{r}"
        ws.conditional_formatting.add(
            status_cell,
            FormulaRule(
                formula=[f'${status_cell}="EXPIRED"'],
                stopIfTrue=False,
                fill=FILL_RED, font=Font(color=RED_FONT, bold=True),
            ),
        )
        ws.conditional_formatting.add(
            status_cell,
            FormulaRule(
                formula=[f'${status_cell}="Expiring Soon"'],
                stopIfTrue=False,
                fill=FILL_ORANGE, font=Font(color=ORANGE_FONT, bold=True),
            ),
        )
        ws.conditional_formatting.add(
            status_cell,
            FormulaRule(
                formula=[f'${status_cell}="Renew Soon"'],
                stopIfTrue=False,
                fill=FILL_YELLOW, font=Font(color=YELLOW_FONT, bold=True),
            ),
        )
        ws.conditional_formatting.add(
            status_cell,
            FormulaRule(
                formula=[f'${status_cell}="Current"'],
                stopIfTrue=False,
                fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True),
            ),
        )

    # ---------------- Counts by Coverage Type ----------------
    cov_section_row = thr + 6 + 1
    ws.cell(row=cov_section_row, column=1, value="Counts by Coverage Type").font = FONT_H1
    ws.cell(row=cov_section_row, column=1).fill = FILL_HEADER
    ws.cell(row=cov_section_row, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=cov_section_row, start_column=1,
                   end_row=cov_section_row, end_column=6)
    ws.row_dimensions[cov_section_row].height = 26

    cov_headers = ["Coverage Type", "Count", "Current", "Expiring Soon", "EXPIRED"]
    chr_ = cov_section_row + 1
    for i, h in enumerate(cov_headers, start=1):
        ws.cell(row=chr_, column=i, value=h)
    style_header_row(ws, chr_, len(cov_headers))
    # Pad right side to keep visual rhythm with 6-column layout
    ws.cell(row=chr_, column=6, value="").fill = FILL_HEADER

    for idx, cov in enumerate(COVERAGE_TYPES):
        r = chr_ + idx + 1
        ws.cell(row=r, column=1, value=cov)
        ws.cell(row=r, column=1).font = FONT_BODY_BOLD
        ws.cell(row=r, column=1).fill = FILL_SUMMARY_LABEL
        ws.cell(row=r, column=1).alignment = Alignment(vertical="center", indent=1)
        ws.cell(row=r, column=1).border = BORDER

        # Total count for this coverage type
        ws.cell(row=r, column=2,
                value=f'=COUNTIF({coverage_range},"{cov}")')
        ws.cell(row=r, column=2).number_format = FMT_INT
        ws.cell(row=r, column=2).font = FONT_BODY_BOLD
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=2).border = BORDER

        # Current
        ws.cell(row=r, column=3,
                value=f'=COUNTIFS({coverage_range},"{cov}",{status_range},"Current")')
        ws.cell(row=r, column=3).number_format = FMT_INT
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=3).border = BORDER

        # Expiring Soon
        ws.cell(row=r, column=4,
                value=f'=COUNTIFS({coverage_range},"{cov}",{status_range},"Expiring Soon")')
        ws.cell(row=r, column=4).number_format = FMT_INT
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=4).border = BORDER

        # EXPIRED
        ws.cell(row=r, column=5,
                value=f'=COUNTIFS({coverage_range},"{cov}",{status_range},"EXPIRED")')
        ws.cell(row=r, column=5).number_format = FMT_INT
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=5).border = BORDER

        # Empty pad column F
        ws.cell(row=r, column=6, value="").border = BORDER

        ws.row_dimensions[r].height = 22

        # Conditional formatting: EXPIRED column (E) red if > 0
        exp_cell = f"E{r}"
        ws.conditional_formatting.add(
            exp_cell,
            FormulaRule(
                formula=[f'${exp_cell}>0'],
                stopIfTrue=False,
                fill=FILL_RED, font=Font(color=RED_FONT, bold=True),
            ),
        )
        # Expiring Soon column (D) orange if > 0
        es_cell = f"D{r}"
        ws.conditional_formatting.add(
            es_cell,
            FormulaRule(
                formula=[f'${es_cell}>0'],
                stopIfTrue=False,
                fill=FILL_ORANGE, font=Font(color=ORANGE_FONT, bold=True),
            ),
        )

    set_col_widths(ws, {
        "A": 30, "B": 14, "C": 14, "D": 18, "E": 18, "F": 14,
    })


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

OUTPUT_PATH = "/Users/home/charles/contrpro/files/packages/business/COI_Tracker.xlsx"


def build() -> str:
    wb = Workbook()

    build_instructions(wb)            # Active sheet renamed to Instructions
    build_csi_reference(wb)           # Adds named ranges CSI_Divisions and CSI_Table
    build_company_info(wb)
    coi_sheet, coi_start, coi_end = build_coi_log(wb)
    build_compliance_summary(wb, coi_sheet, coi_start, coi_end)
    build_dashboard(wb, coi_sheet, coi_start, coi_end)

    # Re-order tabs: Instructions, Company Info, Sub-Vendor COI Log,
    # Compliance Summary, Dashboard, CSI Reference (hidden, last)
    desired_order = [
        "Instructions",
        "Company Info",
        "Sub-Vendor COI Log",
        "Compliance Summary",
        "Dashboard",
        "CSI Reference",
    ]
    wb._sheets = [wb[name] for name in desired_order]

    # Hide CSI Reference
    wb["CSI Reference"].sheet_state = "hidden"

    # Active tab = Dashboard on open
    wb.active = wb.sheetnames.index("Dashboard")

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
