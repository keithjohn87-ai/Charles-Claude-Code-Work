#!/usr/bin/env python3
"""
Build ContrPro Backflow Test Log (XLSX) — Plumbing trade pack.

Companion to cross-connection-control-and-backflow-guide.html. Captures the
per-device test records the water purveyor + AHJ require.

Tabs:
  1. Instructions
  2. Device Inventory (master list of every backflow device on the project)
  3. Initial Test Log (per ASSE 1013 RPZ / 1015 DCV)
  4. Annual Test Log
  5. Repair Log
  6. Reporting Summary

Output:
    /Users/home/charles/contrpro/files/packages/complete/plumbing/Backflow_Test_Log.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
ACCENT_GOLD = "C9A227"
GREEN = "C6EFCE"
RED = "FFC7CE"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GOLD = PatternFill("solid", fgColor=ACCENT_GOLD)

FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=BRAND_BLUE)
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_GREY_IT = Font(name="Calibri", size=10.5, italic=True, color="808080")

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

OUT = "/Users/home/charles/contrpro/files/packages/complete/plumbing/Backflow_Test_Log.xlsx"


def set_col_widths(ws, widths):
    for col, w in widths:
        ws.column_dimensions[col].width = w


def title(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def header(ws, row, cols, start_col=1):
    for i, h in enumerate(cols):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 28


def add_blank_rows(ws, start_row, n_rows, n_cols, formats=None):
    """Add blank pre-formatted entry rows."""
    formats = formats or {}
    for r in range(start_row, start_row + n_rows):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c, value="")
            cell.border = BORDER
            cell.font = FONT_BODY
            cell.alignment = CENTER if c > 1 else LEFT
            if c in formats:
                cell.number_format = formats[c]


# ---------------------------------------------------------------------------
def build_instructions(ws):
    ws.title = "Instructions"
    set_col_widths(ws, [("A", 110)])
    ws["A1"] = "BACKFLOW TEST LOG — INSTRUCTIONS"
    ws["A1"].font = FONT_TITLE
    body = [
        "",
        "PURPOSE",
        "Per-device test record for every backflow-prevention device on a project. Captures",
        "initial test (after installation, before service), annual tests (12-month intervals),",
        "and repair history. Submission-ready format for water purveyor + local AHJ.",
        "",
        "WORKFLOW",
        "  1. Device Inventory — list every backflow device on the project at install time.",
        "     Assign an owner-defined Device ID for the column. Include make / model / size /",
        "     serial / location.",
        "  2. Initial Test Log — one row per device, completed by state-licensed tester after",
        "     install + before placing system in service. All three test readings recorded.",
        "  3. Annual Test Log — recurring annual record. Same tester licensure required.",
        "  4. Repair Log — any failure leading to a repair or replacement. Drives traceability.",
        "  5. Reporting Summary — pulls test status by device for the building owner + AHJ.",
        "",
        "TEST READINGS — WHAT EACH NUMBER MEANS",
        "  Check #1 differential — pressure across the upstream check (RPZ + DCV). Pass:",
        "    ≥1.0 psid (RPZ); ≥1.0 psid (DCV).",
        "  Relief Opening — RPZ only. Pressure at which the relief valve opens during the",
        "    upstream-isolation test. Pass: ≥2.0 psid.",
        "  Check #2 held — DCV + RPZ. Verifies #2 holds against zero downstream pressure.",
        "    Pass: visible hold for the test duration.",
        "",
        "REQUIRED CALIBRATION",
        "Test gauges must be calibrated within the prior 12 months. Record the gauge serial",
        "number + calibration date on every test row. AHJs reject tests with out-of-cal gauges.",
        "",
        "REPORT SUBMISSION",
        "Most jurisdictions require submission within 10 days of testing. Send to: (1) building",
        "owner / facility manager, (2) water purveyor, (3) local AHJ plumbing inspector.",
        "Confirm specific submission requirements with each entity at the initial install.",
        "",
        "OUT-OF-SERVICE PROTOCOL",
        "A failed device must be tagged ('OUT OF SERVICE — DO NOT USE'), the owner notified in",
        "writing within 24 hours, and repaired or replaced within 30 days. Document the tagging",
        "+ notification in the Repair Log.",
        "",
        "DOCUMENT VERSION",
        "Backflow_Test_Log.xlsx v1.0 — 2026-05-19",
        "ContrPro Complete Tier · Plumbing Trade Pack",
    ]
    for i, line in enumerate(body, start=3):
        c = ws.cell(row=i, column=1, value=line)
        if line.isupper() and line.strip() and not line.startswith(" "):
            c.font = FONT_H2
        else:
            c.font = FONT_BODY
        c.alignment = LEFT


def build_device_inventory(ws):
    ws.title = "Device Inventory"
    set_col_widths(ws, [("A", 14), ("B", 22), ("C", 22), ("D", 12), ("E", 14), ("F", 16), ("G", 26), ("H", 26)])
    title(ws, 1, "DEVICE INVENTORY", 8)
    header(ws, 3, ["Device ID", "Location (Bldg/Room)", "Application", "Type (RPZ/DCV/PVB/AG)", "Size", "Make / Model", "Serial #", "Install Date"])
    add_blank_rows(ws, 4, 30, 8)


def build_initial_test(ws):
    ws.title = "Initial Test Log"
    set_col_widths(ws, [("A", 14), ("B", 14), ("C", 12), ("D", 14), ("E", 14), ("F", 12),
                        ("G", 14), ("H", 16), ("I", 18), ("J", 14), ("K", 22)])
    title(ws, 1, "INITIAL TEST LOG (after install, before service)", 11)
    header(ws, 3, ["Device ID", "Test Date", "Tester Name", "Tester License #", "Gauge Serial",
                   "Gauge Cal Date", "Check #1 (psid)", "Relief Open (psid, RPZ)", "Check #2 Held",
                   "Pass / Fail", "Notes"])
    formats = {2: "yyyy-mm-dd", 6: "yyyy-mm-dd", 7: "0.0", 8: "0.0"}
    add_blank_rows(ws, 4, 30, 11, formats)


def build_annual_test(ws):
    ws.title = "Annual Test Log"
    set_col_widths(ws, [("A", 14), ("B", 14), ("C", 12), ("D", 14), ("E", 14), ("F", 12),
                        ("G", 14), ("H", 16), ("I", 18), ("J", 14), ("K", 22)])
    title(ws, 1, "ANNUAL TEST LOG", 11)
    header(ws, 3, ["Device ID", "Test Date", "Tester Name", "Tester License #", "Gauge Serial",
                   "Gauge Cal Date", "Check #1 (psid)", "Relief Open (psid, RPZ)", "Check #2 Held",
                   "Pass / Fail", "Notes"])
    formats = {2: "yyyy-mm-dd", 6: "yyyy-mm-dd", 7: "0.0", 8: "0.0"}
    add_blank_rows(ws, 4, 60, 11, formats)


def build_repair_log(ws):
    ws.title = "Repair Log"
    set_col_widths(ws, [("A", 14), ("B", 14), ("C", 24), ("D", 24), ("E", 18), ("F", 14),
                        ("G", 14), ("H", 16), ("I", 22)])
    title(ws, 1, "REPAIR LOG", 9)
    header(ws, 3, ["Device ID", "Fail Date", "Symptom", "Cause Identified",
                   "Repair Performed", "Repaired By", "Re-test Date",
                   "Re-test Pass / Fail", "Owner Notified (Date)"])
    formats = {2: "yyyy-mm-dd", 7: "yyyy-mm-dd", 9: "yyyy-mm-dd"}
    add_blank_rows(ws, 4, 30, 9, formats)


def build_reporting(ws):
    ws.title = "Reporting Summary"
    set_col_widths(ws, [("A", 14), ("B", 24), ("C", 18), ("D", 18), ("E", 18), ("F", 20)])
    title(ws, 1, "REPORTING SUMMARY (informational)", 6)
    header(ws, 3, ["Device ID", "Application", "Last Test Date", "Next Test Due", "Status",
                   "Sent To (Purveyor / AHJ / Owner)"])
    # 30 rows that look up the Device ID + apply formulas for last/next test
    for r in range(4, 4 + 30):
        ws.cell(row=r, column=1, value=f"=IF('Device Inventory'!A{r}<>\"\",'Device Inventory'!A{r},\"\")")
        ws.cell(row=r, column=2, value=f"=IF('Device Inventory'!A{r}<>\"\",'Device Inventory'!C{r},\"\")")
        # Last test date — MAX of initial + annual log
        ws.cell(row=r, column=3, value=(
            f"=IFERROR(MAX("
            f"IF('Initial Test Log'!A:A=A{r},'Initial Test Log'!B:B,0),"
            f"IF('Annual Test Log'!A:A=A{r},'Annual Test Log'!B:B,0)"
            f"),\"\")"
        ))
        ws.cell(row=r, column=3).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=4, value=f"=IF(C{r}=\"\",\"\",DATE(YEAR(C{r})+1,MONTH(C{r}),DAY(C{r})))")
        ws.cell(row=r, column=4).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=5, value=(
            f"=IF(D{r}=\"\",\"\",IF(D{r}<TODAY(),\"OVERDUE\","
            f"IF(D{r}<TODAY()+30,\"DUE SOON\",\"CURRENT\")))"
        ))
        ws.cell(row=r, column=6, value="")
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).font = FONT_BODY
            ws.cell(row=r, column=c).alignment = CENTER if c > 1 else LEFT


def main():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb = Workbook()
    build_instructions(wb.active)
    build_device_inventory(wb.create_sheet("Device Inventory"))
    build_initial_test(wb.create_sheet("Initial Test Log"))
    build_annual_test(wb.create_sheet("Annual Test Log"))
    build_repair_log(wb.create_sheet("Repair Log"))
    build_reporting(wb.create_sheet("Reporting Summary"))
    wb.save(OUT)
    print(f"OK — wrote {OUT}")


if __name__ == "__main__":
    main()
