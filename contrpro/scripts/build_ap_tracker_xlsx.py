#!/usr/bin/env python3
"""
Build ContrPro Accounts Payable (AP) Tracker (XLSX).

Mirror image of the AR Tracker for what the contractor OWES. Tracks vendor,
supplier, and subcontractor invoices with the two construction-CFO pain points
front and center: lien-waiver-risk (paid a sub without protection) and
COI-expiry-risk (sub on site with no current cert of insurance).

Produces a production-grade workbook with:
  - Instructions tab (AP purpose, lien waivers, trust fund law, aging)
  - Company Info tab
  - Vendor Invoice Log (CSI-coded, dropdowns, formulas, conditional formatting)
  - Aging Summary by Vendor Type (SUMIFS cross-tab)
  - Dashboard (outstanding AP, retainage held, risk flags, DPO, aging)
  - CSI Reference (hidden, named ranges for VLOOKUP / dropdowns)

Run:
    /Users/home/charles/.venv/bin/python3 \
        /Users/home/charles/contrpro/scripts/build_ap_tracker_xlsx.py

Output:
    /Users/home/charles/contrpro/files/packages/business/AP_Tracker.xlsx
"""

from __future__ import annotations

import os
from typing import Tuple

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# ---------------------------------------------------------------------------
# Brand & styling (matches AR Tracker / Job Costing / Change Order Log)
# ---------------------------------------------------------------------------

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
ACCENT_GOLD = "C9A227"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
YELLOW_FILL = "FFEB9C"
YELLOW_FONT = "9C5700"
ORANGE_FILL = "FFD8A8"
ORANGE_FONT = "8A4B00"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_SUMMARY_LABEL = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW_FILL)
FILL_ORANGE = PatternFill("solid", fgColor=ORANGE_FILL)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H1 = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_BIG_NUMBER = Font(name="Calibri", size=18, bold=True, color=BRAND_BLUE)
FONT_GREY_ITALIC = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_USD_BIG = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
FMT_PCT = "0.0%"
FMT_INT = "0"
FMT_DATE = "yyyy-mm-dd"


# ---------------------------------------------------------------------------
# CSI MasterFormat division reference (same canonical source)
# ---------------------------------------------------------------------------

DIVISIONS: list[tuple[str, str]] = [
    ("01", "General Requirements"),
    ("02", "Existing Conditions"),
    ("03", "Concrete"),
    ("04", "Masonry"),
    ("05", "Metals"),
    ("06", "Wood, Plastics, and Composites"),
    ("07", "Thermal and Moisture Protection"),
    ("08", "Openings"),
    ("09", "Finishes"),
    ("10", "Specialties"),
    ("11", "Equipment"),
    ("12", "Furnishings"),
    ("13", "Special Construction"),
    ("14", "Conveying Equipment"),
    ("21", "Fire Suppression"),
    ("22", "Plumbing"),
    ("23", "HVAC"),
    ("25", "Integrated Automation"),
    ("26", "Electrical"),
    ("27", "Communications"),
    ("28", "Electronic Safety and Security"),
    ("31", "Earthwork"),
    ("32", "Exterior Improvements"),
    ("33", "Utilities"),
    ("34", "Transportation"),
    ("35", "Waterway and Marine Construction"),
    ("40", "Process Interconnections"),
    ("41", "Material Processing and Handling Equipment"),
    ("42", "Process Heating, Cooling, and Drying Equipment"),
    ("43", "Process Gas and Liquid Handling, Purification, and Storage Equipment"),
    ("44", "Pollution and Waste Control Equipment"),
    ("45", "Industry-Specific Manufacturing Equipment"),
    ("46", "Water and Wastewater Equipment"),
    ("48", "Electrical Power Generation"),
    ("49", "Reserved"),
]


# ---------------------------------------------------------------------------
# Dropdown vocabularies
# ---------------------------------------------------------------------------

VENDOR_TYPES = ["Subcontractor", "Material Supplier", "Equipment Rental", "Service", "Other"]
LIEN_WAIVER_STATES = ["None", "Conditional", "Unconditional Partial", "Unconditional Final", "N/A"]
COI_STATES = ["Yes", "No", "Expired", "N/A"]
STATUSES = ["Received", "Approved for Payment", "Paid", "On Hold", "Disputed"]
AGING_BUCKETS = ["Current", "1-30", "31-60", "61-90", "91-120", "120+", "Paid"]
AGING_BUCKETS_REPORT = ["Current", "1-30", "31-60", "61-90", "91-120", "120+"]


# ---------------------------------------------------------------------------
# Pre-populated example bills (varied aging / risk states for showcase)
#
# Anchor today = 2026-05-13 (today's date in the build environment).
#   B-2001: Paid sub, unconditional final waiver in hand, COI current (clean)
#   B-2002: Sub bill due 2026-05-30, current, conditional waiver only
#   B-2003: Material supplier 31-60 bucket, COI N/A (no on-site work)
#   B-2004: Sub 61-90, NO COI ON FILE - red flag for risk dashboard
#   B-2005: PAID sub but NO LIEN WAIVER received - red flag (paid without protection)
#
# Fields:
#  (inv_num, date_received, vendor, vtype, div, csi_code, project, desc,
#   amount, retainage_withheld, date_due, date_paid, amount_paid,
#   lien_waiver, coi, status, notes)
# ---------------------------------------------------------------------------

EXAMPLE_BILLS: list[tuple] = [
    ("B-2001", "2026-03-15",
     "Pacific Steel Erectors", "Subcontractor", "05", "05 12 00",
     "Riverside Office Bldg",
     "Pay App #2 - Structural steel erection, Level 1-2.",
     58400.00, 2920.00,
     "2026-04-14", "2026-04-12", 55480.00,
     "Unconditional Partial", "Yes", "Paid",
     "Paid via ACH. Conditional issued with invoice, unconditional received "
     "after funds cleared. COI on file thru 2026-09-30."),

    ("B-2002", "2026-04-30",
     "Apex Plumbing & Mechanical", "Subcontractor", "22", "22 10 00",
     "Riverside Office Bldg",
     "Pay App #3 - Plumbing rough-in, Level 2.",
     34200.00, 1710.00,
     "2026-05-30", "", 0.00,
     "Conditional", "Yes", "Approved for Payment",
     "Pay app reviewed and approved. Conditional waiver on file. Releasing "
     "check this week on receipt of unconditional partial."),

    ("B-2003", "2026-03-02",
     "Heartland Lumber Supply", "Material Supplier", "06", "06 10 00",
     "Riverside Office Bldg",
     "Framing lumber package - Level 2 partitions.",
     11850.00, 0.00,
     "2026-04-01", "", 0.00,
     "N/A", "N/A", "Received",
     "Material supplier - no retainage, no COI required (off-site delivery). "
     "Disputing 12 sticks short on tally. Hold check pending credit memo."),

    ("B-2004", "2026-02-12",
     "Bayside Electric LLC", "Subcontractor", "26", "26 50 00",
     "Pinnacle Tower Renovation",
     "Pay App #1 - Lighting rough-in & fixture install Level 3.",
     22600.00, 1130.00,
     "2026-03-14", "", 10000.00,
     "Conditional", "Expired", "On Hold",
     "RISK: Sub COI expired 2026-04-30. Required new cert before next "
     "payment. Partial payment was released before expiry. Follow-up sent "
     "2026-05-05; no response. Stop further on-site work."),

    ("B-2005", "2026-01-22",
     "Granite Concrete Cutters", "Subcontractor", "03", "03 30 00",
     "Pinnacle Tower Renovation",
     "Saw-cut & remove existing SOG for new MEP chase.",
     8650.00, 0.00,
     "2026-02-21", "2026-02-19", 8650.00,
     "None", "Yes", "Paid",
     "RISK: Paid in full without any lien waiver received. Followed up "
     "twice for unconditional final - radio silence. Lien exposure window "
     "still open per state statute. Escalate to attorney if not resolved by 5/20."),
]

EMPTY_BILL_ROWS = 25
EMPTY_COMPANY_ROWS = 2


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def set_col_widths(ws, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_header_row(ws, row: int, last_col: int) -> None:
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 32


# ---------------------------------------------------------------------------
# Hidden tab — CSI Reference (also feeds named ranges + dropdowns)
# ---------------------------------------------------------------------------

def build_csi_reference(wb: Workbook) -> str:
    ws = wb.create_sheet("CSI Reference")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "CSI MasterFormat Division Reference"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:B1")

    ws["A2"] = "Pulled from the same canonical list as ContrPro Job Costing. Do not edit."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)

    ws.cell(row=3, column=1, value="CSI Division")
    ws.cell(row=3, column=2, value="Division Name")
    style_header_row(ws, 3, 2)

    for i, (code, name) in enumerate(DIVISIONS, start=4):
        ws.cell(row=i, column=1, value=code).number_format = "@"
        ws.cell(row=i, column=2, value=name)
        for c in (1, 2):
            ws.cell(row=i, column=c).border = BORDER
            ws.cell(row=i, column=c).font = FONT_BODY

    set_col_widths(ws, {"A": 14, "B": 60})

    first = 4
    last = 4 + len(DIVISIONS) - 1
    div_range = f"'CSI Reference'!$A${first}:$A${last}"
    full_range = f"'CSI Reference'!$A${first}:$B${last}"
    wb.defined_names["CSI_Divisions"] = DefinedName("CSI_Divisions", attr_text=div_range)
    wb.defined_names["CSI_Table"] = DefinedName("CSI_Table", attr_text=full_range)
    return ws.title


# ---------------------------------------------------------------------------
# Tab 1 — Instructions
# ---------------------------------------------------------------------------

def build_instructions(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {"A": 4, "B": 110})

    ws["B2"] = "ContrPro - Accounts Payable (AP) Tracker"
    ws["B2"].font = FONT_TITLE
    ws.row_dimensions[2].height = 32

    ws["B3"] = ("Mirror image of the AR Tracker - what YOU owe vendors, subs, and suppliers. "
                "Cash flow and lien-waiver discipline live here.")
    ws["B3"].font = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

    sections: list[tuple[str, list[str]]] = [
        (
            "What this workbook is",
            [
                "Five visible tabs plus a hidden CSI Reference. Company Info anchors your entity, Vendor Invoice "
                "Log is the main worksheet, Aging Summary buckets exposure by vendor type, and Dashboard surfaces the "
                "two metrics that wreck contractors: Subs Paid Without Lien Waiver and Subs On Site With Expired COIs.",
                "Every bill you receive goes in the Vendor Invoice Log. The aging math is automatic - you fill in dates "
                "and amounts, the workbook tells you what's due this week, what's late, and which payments carry risk.",
            ],
        ),
        (
            "Vendor vs sub vs material supplier - tag every row",
            [
                "Subcontractor - performs labor on site. Owes you a lien waiver before AND after payment. Must carry "
                "current COI naming you as additional insured. Highest risk per dollar - mechanic's-lien exposure.",
                "Material Supplier - delivers materials, no on-site labor. May still have lien rights (depends on state). "
                "Joint check agreements common when supplier sells through your sub.",
                "Equipment Rental - cranes, lifts, generators. Time-and-materials billing. Watch for damage waivers and "
                "fuel/run-time true-ups.",
                "Service - utilities, dumpsters, sanitation, professional services (engineer, surveyor, attorney).",
                "Other - catch-all. Use sparingly; re-tag when you can.",
            ],
        ),
        (
            "Lien waiver workflow - DO NOT pay without one",
            [
                "Conditional waiver - get from sub BEFORE you cut the check. Says 'lien rights waived CONDITIONAL on "
                "receipt of payment.' Protects you the moment the funds clear. Required at every progress payment.",
                "Unconditional Partial - get AFTER the check clears, for the amount actually paid. Says 'lien rights "
                "waived up to $X paid through date Y.' Required to close out each pay cycle.",
                "Unconditional Final - get on the FINAL payment after the check clears. Says 'all lien rights waived, "
                "paid in full.' Required to close the job.",
                "If the Status is Paid and Lien Waiver Received is None - that row turns red. Track it down NOW. Every "
                "day you let it sit, the sub has a longer window to lien your project.",
            ],
        ),
        (
            "Trust fund implications - this is statutory law in some states",
            [
                "Texas Construction Trust Fund Act (Tex. Prop. Code Ch. 162): money you receive from an owner is held "
                "in TRUST for downstream subs and suppliers. Spending it on overhead before paying subs is a felony "
                "in Texas. Pay your subs from project money first - period.",
                "New York Lien Law Article 3-A: similar statutory trust. Funds received are trust funds; diversion is "
                "larceny. Personal liability for officers/managers who divert.",
                "Maryland, Michigan, Minnesota, New Jersey, Washington, Wisconsin, Illinois (residential): comparable "
                "trust fund or prompt-payment statutes. Confirm your state with counsel before assuming you're free "
                "to commingle project receipts with operating cash.",
                "The Total Retainage Held figure on the Dashboard is the dollar amount you have BACK-pocketed from "
                "subs. In trust-fund states, that money may be statutorily owed - get the lien waiver, release on the "
                "contract milestone, do not sit on it past the agreed release date.",
            ],
        ),
        (
            "COI (Certificate of Insurance) discipline",
            [
                "Every sub on site MUST have current general liability + workers comp coverage naming you as "
                "additional insured. Expired = stop work, no exceptions.",
                "COI Current column: Yes / No / Expired / N/A. If Expired or No appears on a sub with an open balance, "
                "that row turns red and the Dashboard increments the COI risk counter.",
                "Material suppliers and most services are N/A. Equipment rentals usually require a different "
                "certificate type (rental liability). Use your judgment - the dropdown is a tool, not a doctrine.",
            ],
        ),
        (
            "Aging buckets and prompt-payment statutes",
            [
                "Current   - not yet past due. Pay on the negotiated terms.",
                "1-30      - 1 to 30 days past due. Sub will start calling. Most states' prompt-payment statutes "
                "trigger interest in this window (commonly 1-2% per month).",
                "31-60     - 31 to 60 days. Sub may stop work. State licensing boards take complaints seriously.",
                "61-90     - 61 to 90 days. Most state prompt-pay statutes have ALREADY granted statutory interest. "
                "Sub may file a stop-work notice or mechanic's lien.",
                "91-120    - 91 to 120 days. Lien is likely. Bond claim if you bonded the job.",
                "120+      - over 120 days. Litigation or bond claim is probably already in motion.",
                "Note: many states' prompt-payment statutes require GC to pay subs within 7-30 days of receiving "
                "owner payment for that sub's work. Falling outside that window can void contract defenses.",
            ],
        ),
        (
            "Color coding cheat sheet",
            [
                "Aging Bucket: green = Current/Paid, yellow = 1-30/31-60, orange = 61-90, red = 91-120/120+.",
                "Lien Waiver: RED if Status is Paid and Waiver is None - you paid a sub without protection.",
                "COI Current: RED if No or Expired - sub is on site without coverage.",
                "Status: green = Paid, yellow = Received/Approved for Payment, red = On Hold/Disputed.",
            ],
        ),
        (
            "Tab-by-tab quick start",
            [
                "1. Company Info - your entity name, EIN, primary state, and any AP-specific notes.",
                "2. Vendor Invoice Log - one row per bill. Pick Vendor Type and CSI Division from dropdowns. Fill in "
                "Amount, Retainage Withheld, Date Due, Date Paid, Amount Paid. Outstanding, Aging, and Days "
                "Outstanding all calculate. Set Lien Waiver and COI status honestly - the risk panel depends on it.",
                "3. Aging Summary by Vendor Type - DO NOT edit. Cross-tabs aging x vendor type.",
                "4. Dashboard - DO NOT edit. Top of stack: outstanding AP, retainage held from subs, subs with bad "
                "COIs, subs paid without waivers, DPO, aging panel.",
            ],
        ),
        (
            "Weekly cadence",
            [
                "Monday morning: open Dashboard. Anything red gets handled before lunch. COI-expired subs - call them "
                "off site until they email a current cert.",
                "Wednesday: pay run. Cut checks ONLY against rows with a current conditional waiver on file. If you "
                "have to pay without one, document why in Notes and chase the unconditional inside 5 business days.",
                "Friday afternoon: log every bill received and payment made this week. Reconcile against the bank.",
                "End of month: tie Total Outstanding AP to your accounting system. They must match. Reconcile "
                "retainage held to your sub contracts.",
            ],
        ),
        (
            "Tabs in this workbook",
            [
                "Instructions             -  this tab",
                "Company Info             -  your entity anchor",
                "Vendor Invoice Log       -  CSI-coded bill line items, the main worksheet",
                "Aging Summary by Type    -  aging x vendor type cross-tab",
                "Dashboard                -  AP headline, risk flags, DPO, aging panel",
                "CSI Reference            -  canonical division list (hidden - do not edit)",
            ],
        ),
    ]

    row = 5
    for heading, paragraphs in sections:
        ws.cell(row=row, column=2, value=heading)
        ws.cell(row=row, column=2).font = FONT_H2
        ws.cell(row=row, column=2).fill = FILL_SUBHEADER
        ws.cell(row=row, column=2).alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[row].height = 22
        row += 1
        for para in paragraphs:
            ws.cell(row=row, column=2, value=para)
            ws.cell(row=row, column=2).font = FONT_BODY
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = max(18, 16 * (1 + len(para) // 110))
            row += 1
        row += 1

    ws.cell(row=row + 1, column=2,
            value="ContrPro - built for builders. Questions: support@contrpro.com")
    ws.cell(row=row + 1, column=2).font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)


# ---------------------------------------------------------------------------
# Tab 2 — Company Info
# ---------------------------------------------------------------------------

COMPANY_HEADERS = [
    "Company Name", "EIN", "Primary State", "Notes",
]

COMPANY_EXAMPLE = [
    "EXAMPLE - Acme Construction LLC",
    "12-3456789",
    "TX",
    "Texas Construction Trust Fund Act applies - pay subs from project receipts first.",
]


def build_company_info(wb: Workbook) -> None:
    ws = wb.create_sheet("Company Info")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Company Information"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:D1")

    for i, h in enumerate(COMPANY_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(COMPANY_HEADERS))

    # Example row (italic grey)
    for i, v in enumerate(COMPANY_EXAMPLE, start=1):
        cell = ws.cell(row=4, column=i, value=v)
        cell.font = FONT_GREY_ITALIC
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER

    # Empty user rows
    for r in range(5, 5 + EMPTY_COMPANY_ROWS):
        for c in range(1, len(COMPANY_HEADERS) + 1):
            ws.cell(row=r, column=c).border = BORDER
        ws.row_dimensions[r].height = 22

    set_col_widths(ws, {
        "A": 36, "B": 18, "C": 16, "D": 70,
    })

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Tab 3 — Vendor Invoice Log (main worksheet)
# ---------------------------------------------------------------------------

BILL_HEADERS = [
    "Invoice / Bill Number",        # A
    "Date Received",                # B
    "Vendor / Supplier / Sub",      # C
    "Vendor Type",                  # D
    "CSI Division",                 # E
    "CSI Code",                     # F
    "Division Name",                # G  (VLOOKUP)
    "Project / Job",                # H
    "Description",                  # I
    "Amount",                       # J
    "Retainage Withheld",           # K
    "Net Payable",                  # L  (FORMULA)
    "Date Due",                     # M
    "Date Paid",                    # N
    "Amount Paid",                  # O
    "Outstanding",                  # P  (FORMULA)
    "Days Outstanding",             # Q  (FORMULA)
    "Aging Bucket",                 # R  (FORMULA)
    "Lien Waiver Received?",        # S
    "COI Current?",                 # T
    "Status",                       # U
    "Notes",                        # V
]


def build_invoice_log(wb: Workbook) -> Tuple[str, int, int]:
    ws = wb.create_sheet("Vendor Invoice Log")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Vendor Invoice Log (CSI MasterFormat)"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(BILL_HEADERS))

    # Header row at row 3
    for i, h in enumerate(BILL_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(BILL_HEADERS))

    data_start = 4
    n_pre = len(EXAMPLE_BILLS)
    n_total = n_pre + EMPTY_BILL_ROWS
    data_end = data_start + n_total - 1

    COL = {h: get_column_letter(i + 1) for i, h in enumerate(BILL_HEADERS)}
    col_vtype = COL["Vendor Type"]              # D
    col_div = COL["CSI Division"]               # E
    col_amount = COL["Amount"]                  # J
    col_retain = COL["Retainage Withheld"]      # K
    col_net = COL["Net Payable"]                # L
    col_due = COL["Date Due"]                   # M
    col_paid_date = COL["Date Paid"]            # N
    col_amt_paid = COL["Amount Paid"]           # O
    col_outstanding = COL["Outstanding"]        # P
    col_days = COL["Days Outstanding"]          # Q
    col_aging = COL["Aging Bucket"]             # R
    col_lien = COL["Lien Waiver Received?"]    # S
    col_coi = COL["COI Current?"]              # T
    col_status = COL["Status"]                  # U

    def net_payable_formula(r: int) -> str:
        # Net Payable = Amount - Retainage Withheld
        return f"=IFERROR({col_amount}{r},0)-IFERROR({col_retain}{r},0)"

    def outstanding_formula(r: int) -> str:
        return f"=IFERROR({col_net}{r},0)-IFERROR({col_amt_paid}{r},0)"

    def days_outstanding_formula(r: int) -> str:
        # Days = TODAY() - DateDue when outstanding > 0; else blank
        return (
            f'=IF(OR({col_due}{r}="",IFERROR({col_outstanding}{r},0)<=0),"",'
            f'MAX(0,TODAY()-{col_due}{r}))'
        )

    def aging_formula(r: int) -> str:
        # If fully paid (Amount Paid >= Net Payable and Date Paid present) -> Paid
        # Else compute days = TODAY()-DateDue.
        return (
            f'=IF(AND({col_paid_date}{r}<>"",IFERROR({col_amt_paid}{r},0)>=IFERROR({col_net}{r},0),'
            f'IFERROR({col_net}{r},0)>0),"Paid",'
            f'IF({col_due}{r}="","",'
            f'IF(TODAY()-{col_due}{r}<=0,"Current",'
            f'IF(TODAY()-{col_due}{r}<=30,"1-30",'
            f'IF(TODAY()-{col_due}{r}<=60,"31-60",'
            f'IF(TODAY()-{col_due}{r}<=90,"61-90",'
            f'IF(TODAY()-{col_due}{r}<=120,"91-120","120+")))))))'
        )

    def divname_formula(r: int) -> str:
        return f'=IFERROR(VLOOKUP({col_div}{r},CSI_Table,2,FALSE),"")'

    # Pre-populated example rows
    for idx, row_data in enumerate(EXAMPLE_BILLS):
        r = data_start + idx
        (inv_num, d_received, vendor, vtype, div, csi_code, project, desc,
         amount, retain, d_due, d_paid, amt_paid,
         lien, coi, status, notes) = row_data

        ws.cell(row=r, column=1, value=inv_num)
        ws.cell(row=r, column=2, value=d_received)
        ws.cell(row=r, column=3, value=vendor)
        ws.cell(row=r, column=4, value=vtype)
        ws.cell(row=r, column=5, value=div)
        ws.cell(row=r, column=6, value=csi_code)
        ws.cell(row=r, column=7, value=divname_formula(r))
        ws.cell(row=r, column=8, value=project)
        ws.cell(row=r, column=9, value=desc)
        ws.cell(row=r, column=10, value=amount)
        ws.cell(row=r, column=11, value=retain)
        ws.cell(row=r, column=12, value=net_payable_formula(r))
        ws.cell(row=r, column=13, value=d_due)
        ws.cell(row=r, column=14, value=d_paid if d_paid else None)
        ws.cell(row=r, column=15, value=amt_paid if amt_paid else 0)
        ws.cell(row=r, column=16, value=outstanding_formula(r))
        ws.cell(row=r, column=17, value=days_outstanding_formula(r))
        ws.cell(row=r, column=18, value=aging_formula(r))
        ws.cell(row=r, column=19, value=lien)
        ws.cell(row=r, column=20, value=coi)
        ws.cell(row=r, column=21, value=status)
        ws.cell(row=r, column=22, value=notes)

    # Empty user-fillable rows (formulas pre-seeded)
    for i in range(EMPTY_BILL_ROWS):
        r = data_start + n_pre + i
        ws.cell(row=r, column=7, value=divname_formula(r))
        ws.cell(row=r, column=12, value=net_payable_formula(r))
        ws.cell(row=r, column=16, value=outstanding_formula(r))
        ws.cell(row=r, column=17, value=days_outstanding_formula(r))
        ws.cell(row=r, column=18, value=aging_formula(r))

    # Per-row formatting
    wrap_cols = {3, 7, 8, 9, 22}
    for r in range(data_start, data_end + 1):
        for c in range(1, len(BILL_HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(c in wrap_cols),
            )
        # Text columns
        ws.cell(row=r, column=1).number_format = "@"
        ws.cell(row=r, column=5).number_format = "@"
        ws.cell(row=r, column=6).number_format = "@"
        # Dates
        ws.cell(row=r, column=2).number_format = FMT_DATE
        ws.cell(row=r, column=13).number_format = FMT_DATE
        ws.cell(row=r, column=14).number_format = FMT_DATE
        # USD columns: Amount, Retainage, Net Payable, Amount Paid, Outstanding
        for c in (10, 11, 12, 15, 16):
            ws.cell(row=r, column=c).number_format = FMT_USD
        # Days outstanding as int
        ws.cell(row=r, column=17).number_format = FMT_INT
        ws.row_dimensions[r].height = 36

    # --- Conditional formatting ---

    # Aging Bucket (R): green Current/Paid, yellow 1-30/31-60, orange 61-90, red 91+
    aging_range = f"{col_aging}{data_start}:{col_aging}{data_end}"
    ws.conditional_formatting.add(
        aging_range,
        FormulaRule(
            formula=[f'OR(${col_aging}{data_start}="Current",${col_aging}{data_start}="Paid")'],
            stopIfTrue=False,
            fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        aging_range,
        FormulaRule(
            formula=[f'OR(${col_aging}{data_start}="1-30",${col_aging}{data_start}="31-60")'],
            stopIfTrue=False,
            fill=FILL_YELLOW, font=Font(color=YELLOW_FONT, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        aging_range,
        FormulaRule(
            formula=[f'${col_aging}{data_start}="61-90"'],
            stopIfTrue=False,
            fill=FILL_ORANGE, font=Font(color=ORANGE_FONT, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        aging_range,
        FormulaRule(
            formula=[f'OR(${col_aging}{data_start}="91-120",${col_aging}{data_start}="120+")'],
            stopIfTrue=False,
            fill=FILL_RED, font=Font(color=RED_FONT, bold=True),
        ),
    )

    # Lien Waiver (S): RED if Status="Paid" AND Lien Waiver="None"
    # (i.e. we paid the sub with NO protection - the worst kind of mistake)
    lien_range = f"{col_lien}{data_start}:{col_lien}{data_end}"
    ws.conditional_formatting.add(
        lien_range,
        FormulaRule(
            formula=[
                f'AND(${col_status}{data_start}="Paid",'
                f'${col_lien}{data_start}="None")'
            ],
            stopIfTrue=True,
            fill=FILL_RED, font=Font(color=RED_FONT, bold=True),
        ),
    )
    # Green when we have a full unconditional waiver
    ws.conditional_formatting.add(
        lien_range,
        FormulaRule(
            formula=[
                f'OR(${col_lien}{data_start}="Unconditional Partial",'
                f'${col_lien}{data_start}="Unconditional Final")'
            ],
            stopIfTrue=False,
            fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True),
        ),
    )
    # Yellow for Conditional or None when not yet paid
    ws.conditional_formatting.add(
        lien_range,
        FormulaRule(
            formula=[f'${col_lien}{data_start}="Conditional"'],
            stopIfTrue=False,
            fill=FILL_YELLOW, font=Font(color=YELLOW_FONT, bold=True),
        ),
    )

    # COI Current? (T): red if No or Expired
    coi_range = f"{col_coi}{data_start}:{col_coi}{data_end}"
    ws.conditional_formatting.add(
        coi_range,
        FormulaRule(
            formula=[
                f'OR(${col_coi}{data_start}="No",'
                f'${col_coi}{data_start}="Expired")'
            ],
            stopIfTrue=True,
            fill=FILL_RED, font=Font(color=RED_FONT, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        coi_range,
        FormulaRule(
            formula=[f'${col_coi}{data_start}="Yes"'],
            stopIfTrue=False,
            fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True),
        ),
    )

    # Status (U): green Paid, yellow Received/Approved, red On Hold/Disputed
    status_range = f"{col_status}{data_start}:{col_status}{data_end}"
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'${col_status}{data_start}="Paid"'],
            stopIfTrue=False,
            fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[
                f'OR(${col_status}{data_start}="Received",'
                f'${col_status}{data_start}="Approved for Payment")'
            ],
            stopIfTrue=False,
            fill=FILL_YELLOW, font=Font(color=YELLOW_FONT, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[
                f'OR(${col_status}{data_start}="On Hold",'
                f'${col_status}{data_start}="Disputed")'
            ],
            stopIfTrue=False,
            fill=FILL_RED, font=Font(color=RED_FONT, bold=True),
        ),
    )

    # --- Data validation ---

    # Vendor Type dropdown
    dv_vtype = DataValidation(
        type="list", formula1='"' + ",".join(VENDOR_TYPES) + '"', allow_blank=True,
    )
    dv_vtype.error = "Pick a vendor type from the dropdown."
    dv_vtype.errorTitle = "Invalid vendor type"
    ws.add_data_validation(dv_vtype)
    dv_vtype.add(f"{col_vtype}{data_start}:{col_vtype}{data_end}")

    # CSI Division dropdown (named range)
    dv_div = DataValidation(type="list", formula1="=CSI_Divisions", allow_blank=True)
    dv_div.error = "Pick a CSI Division from the dropdown."
    dv_div.errorTitle = "Invalid CSI Division"
    ws.add_data_validation(dv_div)
    dv_div.add(f"{col_div}{data_start}:{col_div}{data_end}")

    # Lien Waiver dropdown
    dv_lien = DataValidation(
        type="list", formula1='"' + ",".join(LIEN_WAIVER_STATES) + '"', allow_blank=True,
    )
    dv_lien.error = "Pick a waiver state from the dropdown."
    dv_lien.errorTitle = "Invalid lien waiver"
    ws.add_data_validation(dv_lien)
    dv_lien.add(f"{col_lien}{data_start}:{col_lien}{data_end}")

    # COI dropdown
    dv_coi = DataValidation(
        type="list", formula1='"' + ",".join(COI_STATES) + '"', allow_blank=True,
    )
    dv_coi.error = "Pick a COI state from the dropdown."
    dv_coi.errorTitle = "Invalid COI state"
    ws.add_data_validation(dv_coi)
    dv_coi.add(f"{col_coi}{data_start}:{col_coi}{data_end}")

    # Status dropdown
    dv_status = DataValidation(
        type="list", formula1='"' + ",".join(STATUSES) + '"', allow_blank=True,
    )
    dv_status.error = "Pick a status from the dropdown."
    dv_status.errorTitle = "Invalid status"
    ws.add_data_validation(dv_status)
    dv_status.add(f"{col_status}{data_start}:{col_status}{data_end}")

    # Totals row
    totals_row = data_end + 2
    ws.cell(row=totals_row, column=9, value="TOTALS").font = FONT_BODY_BOLD
    ws.cell(row=totals_row, column=9).alignment = Alignment(horizontal="right")
    ws.cell(row=totals_row, column=9).fill = FILL_SUBHEADER
    ws.cell(row=totals_row, column=9).border = BORDER
    # Sum Amount(J), Retainage(K), Net Payable(L), Amount Paid(O), Outstanding(P)
    for c in (10, 11, 12, 15, 16):
        col_letter = get_column_letter(c)
        cell = ws.cell(
            row=totals_row, column=c,
            value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})",
        )
        cell.number_format = FMT_USD
        cell.font = FONT_BODY_BOLD
        cell.fill = FILL_SUBHEADER
        cell.border = BORDER

    # Column widths
    set_col_widths(ws, {
        "A": 14, "B": 12, "C": 24, "D": 14, "E": 11, "F": 12, "G": 24,
        "H": 22, "I": 36, "J": 14, "K": 14, "L": 14, "M": 12, "N": 12,
        "O": 14, "P": 14, "Q": 11, "R": 12, "S": 18, "T": 11, "U": 18,
        "V": 36,
    })

    ws.freeze_panes = "C4"

    return ws.title, data_start, data_end


# ---------------------------------------------------------------------------
# Tab 4 — Aging Summary by Vendor Type
# ---------------------------------------------------------------------------

def build_aging_summary(wb: Workbook, inv_sheet: str, inv_start: int, inv_end: int) -> None:
    ws = wb.create_sheet("Aging Summary by Type")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "AP Aging Summary by Vendor Type"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:G1")

    ws["A2"] = ("Reads from Vendor Invoice Log via COUNTIFS / SUMIFS - do not edit. "
                "Cross-tabs aging bucket by vendor type. Material Supplier + Equipment Rental + "
                "Service + Other are rolled into the 'Other' columns.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)
    ws.merge_cells("A2:G2")

    headers = [
        "Aging Bucket",
        "Subcontractor Count", "Sub $",
        "Supplier Count", "Supplier $",
        "Other Count", "Other $",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(headers))

    inv_vtype = f"'{inv_sheet}'!$D${inv_start}:$D${inv_end}"
    inv_aging = f"'{inv_sheet}'!$R${inv_start}:$R${inv_end}"
    inv_outstanding = f"'{inv_sheet}'!$P${inv_start}:$P${inv_end}"

    start_row = 4
    bucket_fills: dict[str, tuple[PatternFill, str]] = {
        "Current": (FILL_GREEN, GREEN_FONT),
        "1-30": (FILL_YELLOW, YELLOW_FONT),
        "31-60": (FILL_YELLOW, YELLOW_FONT),
        "61-90": (FILL_ORANGE, ORANGE_FONT),
        "91-120": (FILL_RED, RED_FONT),
        "120+": (FILL_RED, RED_FONT),
    }

    for idx, bucket in enumerate(AGING_BUCKETS_REPORT):
        r = start_row + idx
        fill, font_color = bucket_fills[bucket]
        ws.cell(row=r, column=1, value=bucket)
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=11, bold=True, color=font_color)
        ws.cell(row=r, column=1).fill = fill
        ws.cell(row=r, column=1).alignment = Alignment(vertical="center", indent=1)

        # Subcontractor count & $
        ws.cell(row=r, column=2,
                value=f'=COUNTIFS({inv_aging},A{r},{inv_vtype},"Subcontractor")')
        ws.cell(row=r, column=2).number_format = FMT_INT
        ws.cell(row=r, column=3,
                value=f'=SUMIFS({inv_outstanding},{inv_aging},A{r},{inv_vtype},"Subcontractor")')
        ws.cell(row=r, column=3).number_format = FMT_USD

        # Material Supplier
        ws.cell(row=r, column=4,
                value=f'=COUNTIFS({inv_aging},A{r},{inv_vtype},"Material Supplier")')
        ws.cell(row=r, column=4).number_format = FMT_INT
        ws.cell(row=r, column=5,
                value=f'=SUMIFS({inv_outstanding},{inv_aging},A{r},{inv_vtype},"Material Supplier")')
        ws.cell(row=r, column=5).number_format = FMT_USD

        # Other (everything else: Equipment Rental + Service + Other + blanks)
        # Computed as total-for-bucket minus sub & supplier.
        ws.cell(
            row=r, column=6,
            value=(
                f'=COUNTIF({inv_aging},A{r})'
                f'-COUNTIFS({inv_aging},A{r},{inv_vtype},"Subcontractor")'
                f'-COUNTIFS({inv_aging},A{r},{inv_vtype},"Material Supplier")'
            ),
        )
        ws.cell(row=r, column=6).number_format = FMT_INT
        ws.cell(
            row=r, column=7,
            value=(
                f'=SUMIFS({inv_outstanding},{inv_aging},A{r})'
                f'-SUMIFS({inv_outstanding},{inv_aging},A{r},{inv_vtype},"Subcontractor")'
                f'-SUMIFS({inv_outstanding},{inv_aging},A{r},{inv_vtype},"Material Supplier")'
            ),
        )
        ws.cell(row=r, column=7).number_format = FMT_USD

    end_row = start_row + len(AGING_BUCKETS_REPORT) - 1

    # Totals row
    tr = end_row + 1
    ws.cell(row=tr, column=1, value="TOTAL OUTSTANDING")
    ws.cell(row=tr, column=1).font = FONT_BODY_BOLD
    ws.cell(row=tr, column=1).fill = FILL_SUBHEADER
    ws.cell(row=tr, column=1).alignment = Alignment(vertical="center", indent=1)

    for c in (2, 4, 6):
        col_letter = get_column_letter(c)
        ws.cell(row=tr, column=c, value=f'=SUM({col_letter}{start_row}:{col_letter}{end_row})')
        ws.cell(row=tr, column=c).number_format = FMT_INT
        ws.cell(row=tr, column=c).font = FONT_BODY_BOLD
        ws.cell(row=tr, column=c).fill = FILL_SUBHEADER
    for c in (3, 5, 7):
        col_letter = get_column_letter(c)
        ws.cell(row=tr, column=c, value=f'=SUM({col_letter}{start_row}:{col_letter}{end_row})')
        ws.cell(row=tr, column=c).number_format = FMT_USD
        ws.cell(row=tr, column=c).font = FONT_BODY_BOLD
        ws.cell(row=tr, column=c).fill = FILL_SUBHEADER

    # Borders on the whole table
    for r in range(3, tr + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = BORDER
        ws.row_dimensions[r].height = 24

    set_col_widths(ws, {
        "A": 18, "B": 16, "C": 18, "D": 16, "E": 18, "F": 16, "G": 18,
    })

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Tab 5 — Dashboard
# ---------------------------------------------------------------------------

def build_dashboard(wb: Workbook, inv_sheet: str, inv_start: int, inv_end: int) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Accounts Payable Dashboard"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:F1")

    ws["A2"] = ("Auto-calculated from Company Info and Vendor Invoice Log. Update those tabs - this one stays in sync. "
                "Pay extra attention to the two RISK rows: paying subs without lien waivers and subs on site with expired COIs.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)
    ws.merge_cells("A2:F2")

    inv_vtype = f"'{inv_sheet}'!$D${inv_start}:$D${inv_end}"
    inv_amount = f"'{inv_sheet}'!$J${inv_start}:$J${inv_end}"
    inv_retain = f"'{inv_sheet}'!$K${inv_start}:$K${inv_end}"
    inv_net = f"'{inv_sheet}'!$L${inv_start}:$L${inv_end}"
    inv_paid_date = f"'{inv_sheet}'!$N${inv_start}:$N${inv_end}"
    inv_amt_paid = f"'{inv_sheet}'!$O${inv_start}:$O${inv_end}"
    inv_outstanding = f"'{inv_sheet}'!$P${inv_start}:$P${inv_end}"
    inv_days = f"'{inv_sheet}'!$Q${inv_start}:$Q${inv_end}"
    inv_aging = f"'{inv_sheet}'!$R${inv_start}:$R${inv_end}"
    inv_lien = f"'{inv_sheet}'!$S${inv_start}:$S${inv_end}"
    inv_coi = f"'{inv_sheet}'!$T${inv_start}:$T${inv_end}"
    inv_status = f"'{inv_sheet}'!$U${inv_start}:$U${inv_end}"
    inv_num = f"'{inv_sheet}'!$A${inv_start}:$A${inv_end}"

    # ---------------- Top panel: AP headline ----------------
    panel_start = 4
    ws.cell(row=panel_start, column=1, value="AP Headline").font = FONT_H1
    ws.cell(row=panel_start, column=1).fill = FILL_HEADER
    ws.cell(row=panel_start, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=panel_start, start_column=1,
                   end_row=panel_start, end_column=6)
    ws.row_dimensions[panel_start].height = 26

    # YTD Paid: sum of Amount Paid where Date Paid is in current calendar year.
    # Using SUMPRODUCT to support the YEAR() guard cleanly.
    ytd_paid = (
        f'=SUMPRODUCT(({inv_paid_date}<>"")*(IFERROR(YEAR({inv_paid_date}),0)=YEAR(TODAY()))*'
        f'IFERROR({inv_amt_paid}+0,0))'
    )

    # Avg Days to Pay = average of (Date Paid - Date Received) for paid bills.
    # Pulls Date Received (col B) and Date Paid (col N).
    inv_received = f"'{inv_sheet}'!$B${inv_start}:$B${inv_end}"
    avg_days_to_pay = (
        f'=IFERROR(SUMPRODUCT(({inv_paid_date}<>"")*({inv_received}<>"")*'
        f'IFERROR({inv_paid_date}-{inv_received},0))/'
        f'MAX(1,SUMPRODUCT(({inv_paid_date}<>"")*({inv_received}<>"")*1)),0)'
    )

    # (label, formula, fmt, conditional_kind)
    metrics = [
        ("Total Outstanding AP", f"=SUM({inv_outstanding})", FMT_USD, "outstanding_ap"),
        ("Total Retainage Held (from subs)",
         f'=SUMIFS({inv_retain},{inv_vtype},"Subcontractor")', FMT_USD, None),
        ("Total Paid YTD", ytd_paid, FMT_USD, None),
        ("Avg Days to Pay (DPO equivalent)", avg_days_to_pay, FMT_INT, "dpo"),
    ]

    row = panel_start + 1
    outstanding_cells: list[str] = []
    dpo_cells: list[str] = []
    for label, formula, fmt, kind in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = FONT_BODY_BOLD
        ws.cell(row=row, column=1).fill = FILL_SUMMARY_LABEL
        ws.cell(row=row, column=1).alignment = Alignment(vertical="center", indent=1)
        ws.cell(row=row, column=1).border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

        ws.cell(row=row, column=4, value=formula)
        ws.cell(row=row, column=4).font = FONT_BIG_NUMBER
        ws.cell(row=row, column=4).number_format = fmt
        ws.cell(row=row, column=4).alignment = Alignment(
            horizontal="right", vertical="center", indent=1,
        )
        ws.cell(row=row, column=4).border = BORDER
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)

        ref = f"D{row}"
        if kind == "outstanding_ap":
            outstanding_cells.append(ref)
        elif kind == "dpo":
            dpo_cells.append(ref)
        ws.row_dimensions[row].height = 28
        row += 1

    # Outstanding AP: red if >0 (you owe money), green if 0.
    # NOTE: outstanding AP being high is not inherently bad (it's normal float),
    # but we still surface it via fill so the eye finds it.
    for ref in outstanding_cells:
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=False,
                       fill=FILL_YELLOW, font=Font(color=YELLOW_FONT, bold=True, size=18)),
        )
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="lessThanOrEqual", formula=["0"], stopIfTrue=False,
                       fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True, size=18)),
        )

    # DPO: too fast may not be optimal cash management, but too slow burns subs.
    # Green 14-45 days, yellow 0-14 / 45-60, red >60 or <0.
    for ref in dpo_cells:
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="greaterThan", formula=["60"], stopIfTrue=True,
                       fill=FILL_RED, font=Font(color=RED_FONT, bold=True, size=18)),
        )
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="between", formula=["14", "45"], stopIfTrue=True,
                       fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True, size=18)),
        )
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=False,
                       fill=FILL_YELLOW, font=Font(color=YELLOW_FONT, bold=True, size=18)),
        )

    # ---------------- Risk panel ----------------
    risk_row = row + 1
    ws.cell(row=risk_row, column=1, value="Risk Flags - These Are Where CFOs Lose Sleep").font = FONT_H1
    ws.cell(row=risk_row, column=1).fill = FILL_HEADER
    ws.cell(row=risk_row, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=risk_row, start_column=1,
                   end_row=risk_row, end_column=6)
    ws.row_dimensions[risk_row].height = 26

    # Subs without current COI: count of rows where Vendor Type = Subcontractor
    # AND COI is No or Expired AND Outstanding > 0 (open exposure).
    # We can't easily do a multi-condition SUMPRODUCT with OR on COI, so do it
    # as the sum of two COUNTIFS.
    coi_risk_count = (
        f'=COUNTIFS({inv_vtype},"Subcontractor",{inv_coi},"No",{inv_outstanding},">0")'
        f'+COUNTIFS({inv_vtype},"Subcontractor",{inv_coi},"Expired",{inv_outstanding},">0")'
    )
    # Subs paid without any lien waiver
    lien_risk_count = (
        f'=COUNTIFS({inv_vtype},"Subcontractor",{inv_status},"Paid",{inv_lien},"None")'
    )
    # Dollar exposure on subs paid without waiver
    lien_risk_amount = (
        f'=SUMIFS({inv_amt_paid},{inv_vtype},"Subcontractor",{inv_status},"Paid",{inv_lien},"None")'
    )

    risk_metrics = [
        ("Subs without current COI (open exposure)", coi_risk_count, FMT_INT, "risk_count"),
        ("Subs paid without lien waiver - count", lien_risk_count, FMT_INT, "risk_count"),
        ("Subs paid without lien waiver - $ exposure", lien_risk_amount, FMT_USD, "risk_amount"),
    ]

    row = risk_row + 1
    risk_cells: list[str] = []
    for label, formula, fmt, kind in risk_metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = FONT_BODY_BOLD
        ws.cell(row=row, column=1).fill = FILL_SUMMARY_LABEL
        ws.cell(row=row, column=1).alignment = Alignment(vertical="center", indent=1)
        ws.cell(row=row, column=1).border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

        ws.cell(row=row, column=4, value=formula)
        ws.cell(row=row, column=4).font = FONT_BIG_NUMBER
        ws.cell(row=row, column=4).number_format = fmt
        ws.cell(row=row, column=4).alignment = Alignment(
            horizontal="right", vertical="center", indent=1,
        )
        ws.cell(row=row, column=4).border = BORDER
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)

        risk_cells.append(f"D{row}")
        ws.row_dimensions[row].height = 28
        row += 1

    # Risk cells: red when >0, green when 0
    for ref in risk_cells:
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=False,
                       fill=FILL_RED, font=Font(color=RED_FONT, bold=True, size=18)),
        )
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="lessThanOrEqual", formula=["0"], stopIfTrue=False,
                       fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True, size=18)),
        )

    # ---------------- Aging panel ----------------
    section_row = row + 1
    ws.cell(row=section_row, column=1, value="Aging Breakdown").font = FONT_H1
    ws.cell(row=section_row, column=1).fill = FILL_HEADER
    ws.cell(row=section_row, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=section_row, start_column=1,
                   end_row=section_row, end_column=6)
    ws.row_dimensions[section_row].height = 26

    aging_headers = ["Aging Bucket", "Count", "Outstanding", "% of Outstanding"]
    hr = section_row + 1
    for i, h in enumerate(aging_headers, start=1):
        ws.cell(row=hr, column=i, value=h)
    style_header_row(ws, hr, len(aging_headers))
    ws.merge_cells(start_row=hr, start_column=3, end_row=hr, end_column=4)
    ws.merge_cells(start_row=hr, start_column=5, end_row=hr, end_column=6)

    bucket_panel = [
        ("Current", FILL_GREEN, GREEN_FONT),
        ("1-30", FILL_YELLOW, YELLOW_FONT),
        ("31-60", FILL_YELLOW, YELLOW_FONT),
        ("61-90", FILL_ORANGE, ORANGE_FONT),
        ("91-120", FILL_RED, RED_FONT),
        ("120+", FILL_RED, RED_FONT),
    ]

    total_outstanding_formula = f"SUM({inv_outstanding})"

    br = hr + 1
    for idx, (bucket, fill, font_color) in enumerate(bucket_panel):
        r = br + idx
        ws.cell(row=r, column=1, value=bucket)
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=12, bold=True, color=font_color)
        ws.cell(row=r, column=1).fill = fill
        ws.cell(row=r, column=1).alignment = Alignment(vertical="center", indent=1)
        ws.cell(row=r, column=1).border = BORDER
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

        ws.cell(row=r, column=3, value=f'=COUNTIF({inv_aging},"{bucket}")')
        ws.cell(row=r, column=3).font = FONT_BODY_BOLD
        ws.cell(row=r, column=3).number_format = FMT_INT
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=3).border = BORDER

        ws.cell(row=r, column=4,
                value=f'=SUMIFS({inv_outstanding},{inv_aging},"{bucket}")')
        ws.cell(row=r, column=4).font = FONT_BODY_BOLD
        ws.cell(row=r, column=4).number_format = FMT_USD
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.cell(row=r, column=4).border = BORDER

        ws.cell(row=r, column=5,
                value=f'=IFERROR(SUMIFS({inv_outstanding},{inv_aging},"{bucket}")/{total_outstanding_formula},0)')
        ws.cell(row=r, column=5).font = FONT_BODY_BOLD
        ws.cell(row=r, column=5).number_format = FMT_PCT
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.cell(row=r, column=5).border = BORDER
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 22

    # Aging total row
    btr = br + len(bucket_panel)
    ws.cell(row=btr, column=1, value="TOTAL OUTSTANDING")
    ws.cell(row=btr, column=1).font = FONT_BODY_BOLD
    ws.cell(row=btr, column=1).fill = FILL_SUBHEADER
    ws.cell(row=btr, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.cell(row=btr, column=1).border = BORDER
    ws.merge_cells(start_row=btr, start_column=1, end_row=btr, end_column=2)

    ws.cell(row=btr, column=3,
            value=f'=COUNTIF({inv_aging},"Current")+COUNTIF({inv_aging},"1-30")'
                  f'+COUNTIF({inv_aging},"31-60")+COUNTIF({inv_aging},"61-90")'
                  f'+COUNTIF({inv_aging},"91-120")+COUNTIF({inv_aging},"120+")')
    ws.cell(row=btr, column=3).font = FONT_BODY_BOLD
    ws.cell(row=btr, column=3).fill = FILL_SUBHEADER
    ws.cell(row=btr, column=3).number_format = FMT_INT
    ws.cell(row=btr, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=btr, column=3).border = BORDER

    ws.cell(row=btr, column=4, value=f'=SUM({inv_outstanding})')
    ws.cell(row=btr, column=4).font = FONT_BODY_BOLD
    ws.cell(row=btr, column=4).fill = FILL_SUBHEADER
    ws.cell(row=btr, column=4).number_format = FMT_USD
    ws.cell(row=btr, column=4).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.cell(row=btr, column=4).border = BORDER

    ws.cell(row=btr, column=5, value="=1")
    ws.cell(row=btr, column=5).font = FONT_BODY_BOLD
    ws.cell(row=btr, column=5).fill = FILL_SUBHEADER
    ws.cell(row=btr, column=5).number_format = FMT_PCT
    ws.cell(row=btr, column=5).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.cell(row=btr, column=5).border = BORDER
    ws.merge_cells(start_row=btr, start_column=5, end_row=btr, end_column=6)

    set_col_widths(ws, {
        "A": 26, "B": 14, "C": 14, "D": 18, "E": 14, "F": 10,
    })


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

OUTPUT_PATH = "/Users/home/charles/contrpro/files/packages/business/AP_Tracker.xlsx"


def build() -> str:
    wb = Workbook()

    build_instructions(wb)            # Active sheet renamed to Instructions
    build_csi_reference(wb)           # Adds named ranges CSI_Divisions and CSI_Table
    build_company_info(wb)
    inv_sheet, inv_start, inv_end = build_invoice_log(wb)
    build_aging_summary(wb, inv_sheet, inv_start, inv_end)
    build_dashboard(wb, inv_sheet, inv_start, inv_end)

    # Re-order tabs for proper UX: Instructions, Company Info, Vendor Invoice Log,
    # Aging Summary by Type, Dashboard, CSI Reference (hidden, last)
    desired_order = [
        "Instructions",
        "Company Info",
        "Vendor Invoice Log",
        "Aging Summary by Type",
        "Dashboard",
        "CSI Reference",
    ]
    wb._sheets = [wb[name] for name in desired_order]

    # Hide CSI Reference (canonical lookup; user shouldn't see it by default)
    wb["CSI Reference"].sheet_state = "hidden"

    # Active tab = Dashboard on open
    wb.active = wb.sheetnames.index("Dashboard")

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
