#!/usr/bin/env python3
"""Build all three Mechanical companion XLSX logs in one pass.

Outputs:
  - Refrigerant_and_HotWork_Log.xlsx
  - Mechanical_Pressure_and_Leakage_Test_Log.xlsx
  - TAB_and_Cx_FPT_Log.xlsx
"""
from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BRAND_BLUE = "1E3A5F"
FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=BRAND_BLUE)
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

OUT_DIR = "/Users/home/charles/contrpro/files/packages/complete/mechanical"


def widths(ws, cols):
    for col, w in cols:
        ws.column_dimensions[col].width = w


def title(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text); c.font = FONT_TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def header(ws, row, cols):
    for i, h in enumerate(cols):
        c = ws.cell(row=row, column=1 + i, value=h)
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[row].height = 28


def blanks(ws, start, n_rows, n_cols, fmts=None):
    fmts = fmts or {}
    for r in range(start, start + n_rows):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c, value="")
            cell.border = BORDER; cell.font = FONT_BODY
            cell.alignment = CENTER if c > 1 else LEFT
            if c in fmts: cell.number_format = fmts[c]


def instr_block(ws, ws_title, body_lines):
    ws.title = "Instructions"
    widths(ws, [("A", 110)])
    ws["A1"] = ws_title; ws["A1"].font = FONT_TITLE
    for i, line in enumerate(body_lines, start=3):
        c = ws.cell(row=i, column=1, value=line)
        c.font = FONT_H2 if (line.isupper() and line.strip() and not line.startswith(" ")) else FONT_BODY
        c.alignment = LEFT


# ===========================================================================
# Refrigerant + Hot Work Log
# ===========================================================================
def build_refrigerant_log():
    wb = Workbook()
    instr_block(wb.active, "REFRIGERANT + HOT WORK LOG — INSTRUCTIONS", [
        "",
        "PURPOSE",
        "Captures EPA 608 recovery + charge events, leak-detection + repair, nitrogen pressure",
        "tests + vacuum decay, hot-work permits, and per-tech certification tracking.",
        "Submission-ready for EPA / OSHA / AHJ inquiries.",
        "",
        "WORKFLOW",
        "  • Tech Certs — list every refrigeration tech + EPA 608 type + cert # + expiry.",
        "  • Recovery / Charge Log — one row per event; lbs in / lbs out, cylinder ID.",
        "  • Leak Detection + Repair Log — per EPA Subpart F triggers.",
        "  • Nitrogen Pressure Test + Vacuum Decay — one row per system / circuit.",
        "  • Hot Work Permit Log — daily permit issuance + fire watch.",
        "",
        "EPA SECTION 608 REPORTING",
        "Annual leak rate triggers for commercial refrigeration (≥30%) and comfort cooling",
        "(≥10%) require leak repair + verification. Document each event.",
        "",
        "DOCUMENT VERSION",
        "Refrigerant_and_HotWork_Log.xlsx v1.0 — 2026-05-19",
        "ContrPro Complete Tier · Mechanical Trade Pack",
    ])

    ws = wb.create_sheet("Tech Certs")
    widths(ws, [("A", 8), ("B", 22), ("C", 14), ("D", 18), ("E", 14), ("F", 20)])
    title(ws, 1, "TECH CERTS (EPA 608)", 6)
    header(ws, 3, ["#", "Tech Name", "Cert Type (I/II/III/Universal)",
                   "Cert # / Org", "Expiry Date", "Other Certs (ASSE 6010, etc.)"])
    blanks(ws, 4, 20, 6, {5: "yyyy-mm-dd"})

    ws = wb.create_sheet("Recovery+Charge")
    widths(ws, [("A", 8), ("B", 14), ("C", 22), ("D", 14), ("E", 14), ("F", 14), ("G", 14), ("H", 18), ("I", 22)])
    title(ws, 1, "RECOVERY + CHARGE LOG", 9)
    header(ws, 3, ["#", "Date", "Equipment", "Refrigerant Type", "Lbs Recovered",
                   "Lbs Charged", "Cylinder ID", "Tech (cert #)", "Notes"])
    blanks(ws, 4, 60, 9, {2: "yyyy-mm-dd", 5: "0.0", 6: "0.0"})

    ws = wb.create_sheet("Leak+Repair")
    widths(ws, [("A", 8), ("B", 14), ("C", 22), ("D", 14), ("E", 24), ("F", 14), ("G", 14), ("H", 14)])
    title(ws, 1, "LEAK DETECTION + REPAIR LOG", 8)
    header(ws, 3, ["#", "Date", "Equipment", "Annual Leak Rate %",
                   "Repair Performed", "Tech", "Initial Verify Date", "Follow-up Verify Date"])
    blanks(ws, 4, 30, 8, {2: "yyyy-mm-dd", 4: "0.0%", 7: "yyyy-mm-dd", 8: "yyyy-mm-dd"})

    ws = wb.create_sheet("N2 Pressure+Vacuum")
    widths(ws, [("A", 8), ("B", 14), ("C", 22), ("D", 14), ("E", 14), ("F", 14), ("G", 14), ("H", 14), ("I", 14), ("J", 22)])
    title(ws, 1, "N2 PRESSURE TEST + VACUUM DECAY", 10)
    header(ws, 3, ["#", "Date", "System / Circuit", "Test Pressure (psig)",
                   "Hold Duration", "Pressure Result", "Vacuum (microns)",
                   "Decay Test Result", "Tech", "Notes"])
    blanks(ws, 4, 40, 10, {2: "yyyy-mm-dd", 4: "0", 7: "0"})

    ws = wb.create_sheet("Hot Work Permits")
    widths(ws, [("A", 8), ("B", 14), ("C", 22), ("D", 22), ("E", 18), ("F", 18), ("G", 22)])
    title(ws, 1, "HOT WORK PERMIT LOG", 7)
    header(ws, 3, ["#", "Date", "Location", "Work Type (Braze/Weld/Cut)",
                   "Worker(s)", "Fire Watch", "Issued By / Notes"])
    blanks(ws, 4, 60, 7, {2: "yyyy-mm-dd"})

    out = os.path.join(OUT_DIR, "Refrigerant_and_HotWork_Log.xlsx")
    wb.save(out); print(f"OK — wrote {out}")


# ===========================================================================
# Mechanical Pressure + Leakage Test Log
# ===========================================================================
def build_pressure_log():
    wb = Workbook()
    instr_block(wb.active, "MECHANICAL PRESSURE + LEAKAGE TEST LOG — INSTRUCTIONS", [
        "",
        "PURPOSE",
        "Captures duct leakage tests (SMACNA DCS Chapter 5), hydronic pressure tests, gas pressure",
        "tests (IFGC 406), steam pressure tests, and chemistry / flush logs.",
        "",
        "WORKFLOW",
        "  • Duct Leakage Log — one row per zone tested.",
        "  • Hydronic Pressure Log — pre-fill pressure test + post-fill operating pressure verification.",
        "  • Gas Pressure Log — per IFGC 406 / NFPA 54.",
        "  • Steam Pressure Log (if applicable).",
        "  • Flush + Fill Log — hydronic flush + chemical fill / glycol charge.",
        "",
        "DOCUMENT VERSION",
        "Mechanical_Pressure_and_Leakage_Test_Log.xlsx v1.0 — 2026-05-19",
        "ContrPro Complete Tier · Mechanical Trade Pack",
    ])

    ws = wb.create_sheet("Duct Leakage")
    widths(ws, [("A", 8), ("B", 14), ("C", 22), ("D", 14), ("E", 14), ("F", 16), ("G", 16), ("H", 16), ("I", 14), ("J", 22)])
    title(ws, 1, "DUCT LEAKAGE TEST LOG (SMACNA DCS)", 10)
    header(ws, 3, ["#", "Date", "Zone / System", "Pressure Class",
                   "Test Pressure (in wg)", "Surface Area (sf)", "Measured Leakage (cfm)",
                   "Calculated CL", "Pass/Fail", "Tester"])
    blanks(ws, 4, 30, 10, {2: "yyyy-mm-dd", 5: "0.0", 6: "0", 7: "0.0", 8: "0.0"})

    ws = wb.create_sheet("Hydronic Pressure")
    widths(ws, [("A", 8), ("B", 14), ("C", 22), ("D", 14), ("E", 14), ("F", 14), ("G", 14), ("H", 14), ("I", 22)])
    title(ws, 1, "HYDRONIC PRESSURE TEST LOG", 9)
    header(ws, 3, ["#", "Date", "System / Loop", "Test Pressure (psig)",
                   "Working Pressure", "Hold (hr)", "Start (psig)", "Finish (psig)",
                   "Pass/Fail + Tester"])
    blanks(ws, 4, 30, 9, {2: "yyyy-mm-dd", 4: "0", 5: "0", 6: "0.0", 7: "0", 8: "0"})

    ws = wb.create_sheet("Gas Pressure")
    widths(ws, [("A", 8), ("B", 14), ("C", 22), ("D", 14), ("E", 14), ("F", 14), ("G", 14), ("H", 22)])
    title(ws, 1, "GAS PRESSURE TEST LOG (IFGC 406)", 8)
    header(ws, 3, ["#", "Date", "System", "Test Pressure (psig)",
                   "Hold (min)", "Start / Finish (psig)", "Medium (Air/N2)",
                   "Tester / Witness / Notes"])
    blanks(ws, 4, 25, 8, {2: "yyyy-mm-dd", 4: "0.0", 5: "0"})

    ws = wb.create_sheet("Steam Pressure")
    widths(ws, [("A", 8), ("B", 14), ("C", 22), ("D", 14), ("E", 14), ("F", 14), ("G", 22)])
    title(ws, 1, "STEAM PRESSURE TEST LOG", 7)
    header(ws, 3, ["#", "Date", "System", "Test Pressure (psig)",
                   "Hold (hr)", "Pass/Fail", "Tester / Notes"])
    blanks(ws, 4, 15, 7, {2: "yyyy-mm-dd", 4: "0"})

    ws = wb.create_sheet("Flush + Fill")
    widths(ws, [("A", 8), ("B", 14), ("C", 22), ("D", 22), ("E", 18), ("F", 22)])
    title(ws, 1, "FLUSH + FILL LOG", 6)
    header(ws, 3, ["#", "Date", "Loop / System", "Flush Method + Result",
                   "Chemical / Glycol Added", "Tested By / Lab Sample Date"])
    blanks(ws, 4, 20, 6, {2: "yyyy-mm-dd"})

    out = os.path.join(OUT_DIR, "Mechanical_Pressure_and_Leakage_Test_Log.xlsx")
    wb.save(out); print(f"OK — wrote {out}")


# ===========================================================================
# TAB + Cx FPT Log
# ===========================================================================
def build_tab_cx_log():
    wb = Workbook()
    instr_block(wb.active, "TAB + CX FPT LOG — INSTRUCTIONS", [
        "",
        "PURPOSE",
        "Captures TAB air-side + water-side measurements, Pre-Functional Checklists (PFC),",
        "Functional Performance Tests (FPT), Integrated Systems Tests (IST), and deficiency",
        "tracking through resolution.",
        "",
        "WORKFLOW",
        "  • PFC Log — one row per piece of equipment; installer signs.",
        "  • TAB Air-Side — per terminal + per equipment. Design vs. measured + delta.",
        "  • TAB Water-Side — per loop + per pump.",
        "  • FPT Log — per FPT test step + pass/fail.",
        "  • IST Log — integrated multi-system tests.",
        "  • Deficiency Log — track each deficiency from FPT/IST through retest.",
        "",
        "ACCEPTANCE TOLERANCE",
        "  • Airflow: typically ±10% of design.",
        "  • Water flow: typically ±5% of design.",
        "  • Verify project spec for tighter tolerances on critical applications.",
        "",
        "DOCUMENT VERSION",
        "TAB_and_Cx_FPT_Log.xlsx v1.0 — 2026-05-19",
        "ContrPro Complete Tier · Mechanical Trade Pack",
    ])

    ws = wb.create_sheet("PFC Log")
    widths(ws, [("A", 8), ("B", 14), ("C", 22), ("D", 22), ("E", 18), ("F", 18)])
    title(ws, 1, "PRE-FUNCTIONAL CHECKLIST LOG", 6)
    header(ws, 3, ["#", "Date", "Equipment", "Installer Verification",
                   "CxA Verification", "Status (Pass/Fail/N-A)"])
    blanks(ws, 4, 30, 6, {2: "yyyy-mm-dd"})

    ws = wb.create_sheet("TAB Air-Side")
    widths(ws, [("A", 8), ("B", 14), ("C", 22), ("D", 14), ("E", 14), ("F", 14), ("G", 14), ("H", 14), ("I", 18)])
    title(ws, 1, "TAB AIR-SIDE LOG", 9)
    header(ws, 3, ["#", "Date", "Terminal / Equipment", "Type (Diff/Grille/VAV/etc)",
                   "Design CFM", "Measured CFM", "Delta %",
                   "Damper Position", "Tech / Notes"])
    blanks(ws, 4, 80, 9, {2: "yyyy-mm-dd", 5: "0", 6: "0", 7: "0.0%"})

    ws = wb.create_sheet("TAB Water-Side")
    widths(ws, [("A", 8), ("B", 14), ("C", 22), ("D", 14), ("E", 14), ("F", 14), ("G", 14), ("H", 14), ("I", 18)])
    title(ws, 1, "TAB WATER-SIDE LOG", 9)
    header(ws, 3, ["#", "Date", "Loop / Coil / Pump", "Type",
                   "Design GPM / TDH", "Measured GPM / TDH", "Delta %",
                   "Balance Valve Position", "Tech / Notes"])
    blanks(ws, 4, 40, 9, {2: "yyyy-mm-dd", 7: "0.0%"})

    ws = wb.create_sheet("FPT Log")
    widths(ws, [("A", 8), ("B", 14), ("C", 22), ("D", 32), ("E", 18), ("F", 18), ("G", 22)])
    title(ws, 1, "FUNCTIONAL PERFORMANCE TEST LOG", 7)
    header(ws, 3, ["#", "Date", "Equipment / System", "Test Step",
                   "Expected Result", "Actual Result", "Pass / Fail / Notes"])
    blanks(ws, 4, 50, 7, {2: "yyyy-mm-dd"})

    ws = wb.create_sheet("IST Log")
    widths(ws, [("A", 8), ("B", 14), ("C", 22), ("D", 32), ("E", 18), ("F", 22)])
    title(ws, 1, "INTEGRATED SYSTEMS TEST LOG", 6)
    header(ws, 3, ["#", "Date", "Test Name", "Scenario / Steps",
                   "Pass / Fail", "Witness / Notes"])
    blanks(ws, 4, 20, 6, {2: "yyyy-mm-dd"})

    ws = wb.create_sheet("Deficiency Log")
    widths(ws, [("A", 8), ("B", 14), ("C", 22), ("D", 32), ("E", 14), ("F", 14), ("G", 14), ("H", 22)])
    title(ws, 1, "DEFICIENCY LOG", 8)
    header(ws, 3, ["#", "Date Opened", "Equipment / System", "Description",
                   "Severity (P/F/C)", "Owner", "Closed Date", "Resolution Notes"])
    blanks(ws, 4, 40, 8, {2: "yyyy-mm-dd", 7: "yyyy-mm-dd"})

    out = os.path.join(OUT_DIR, "TAB_and_Cx_FPT_Log.xlsx")
    wb.save(out); print(f"OK — wrote {out}")


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    build_refrigerant_log()
    build_pressure_log()
    build_tab_cx_log()


if __name__ == "__main__":
    main()
