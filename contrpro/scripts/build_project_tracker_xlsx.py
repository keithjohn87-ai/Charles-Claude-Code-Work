#!/usr/bin/env python3
"""
Build ContrPro Project Tracker (XLSX).

Portfolio-level project tracker - the GC's "all my open jobs" master log.
Different from Job_Costing_Spreadsheet (single job, CSI-coded). This rolls
multiple concurrent projects into a portfolio with backlog, margin, and
completion KPIs.

Produces:
  - Instructions tab
  - Company Info tab (shared company metadata)
  - Project Portfolio tab (main worksheet, one row per job)
  - Portfolio Summary by Status tab (auto roll-up)
  - Dashboard tab (top KPIs, top-5 backlogs, aged projects)
  - CSI Reference (hidden, feeds named ranges + dropdowns)

Run:
    /Users/home/charles/.venv/bin/python3 \
        /Users/home/charles/contrpro/scripts/build_project_tracker_xlsx.py

Output:
    /Users/home/charles/contrpro/files/packages/business/Project_Tracker.xlsx
"""

from __future__ import annotations

import os
from typing import Tuple

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule

# ---------------------------------------------------------------------------
# Brand & styling (matches Job_Costing_Spreadsheet + Change_Order_Log)
# ---------------------------------------------------------------------------

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
ACCENT_GOLD = "C9A227"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
YELLOW_FILL = "FFEB9C"
YELLOW_FONT = "9C5700"
BLUE_FILL = "BDD7EE"
BLUE_FONT = "1F4E78"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_SUMMARY_LABEL = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW_FILL)
FILL_BLUE = PatternFill("solid", fgColor=BLUE_FILL)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H1 = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_BIG_NUMBER = Font(name="Calibri", size=18, bold=True, color=BRAND_BLUE)
FONT_GREY_ITALIC = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_USD_BIG = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
FMT_PCT = "0.0%"
FMT_PCT_INT = "0%"
FMT_INT = "0"
FMT_DATE = "yyyy-mm-dd"


# ---------------------------------------------------------------------------
# Reference data
# ---------------------------------------------------------------------------

DIVISIONS: list[tuple[str, str]] = [
    ("01", "General Requirements"),
    ("02", "Existing Conditions"),
    ("03", "Concrete"),
    ("04", "Masonry"),
    ("05", "Metals"),
    ("06", "Wood, Plastics, and Composites"),
    ("07", "Thermal and Moisture Protection"),
    ("08", "Openings"),
    ("09", "Finishes"),
    ("10", "Specialties"),
    ("11", "Equipment"),
    ("12", "Furnishings"),
    ("13", "Special Construction"),
    ("14", "Conveying Equipment"),
    ("21", "Fire Suppression"),
    ("22", "Plumbing"),
    ("23", "HVAC"),
    ("25", "Integrated Automation"),
    ("26", "Electrical"),
    ("27", "Communications"),
    ("28", "Electronic Safety and Security"),
    ("31", "Earthwork"),
    ("32", "Exterior Improvements"),
    ("33", "Utilities"),
    ("34", "Transportation"),
    ("35", "Waterway and Marine Construction"),
    ("40", "Process Interconnections"),
    ("41", "Material Processing and Handling Equipment"),
    ("42", "Process Heating, Cooling, and Drying Equipment"),
    ("43", "Process Gas and Liquid Handling, Purification, and Storage Equipment"),
    ("44", "Pollution and Waste Control Equipment"),
    ("45", "Industry-Specific Manufacturing Equipment"),
    ("46", "Water and Wastewater Equipment"),
    ("48", "Electrical Power Generation"),
    ("49", "Reserved"),
]

PROJECT_TYPES = [
    "New Construction", "Remodel", "Tenant Improvement",
    "Repair", "Service", "Specialty",
]

PROJECT_STATUSES = [
    "Bidding", "Awarded", "Mobilizing", "Active",
    "Substantial Completion", "Final", "Closed", "Lost",
]

STATES_50_DC = [
    "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "DC", "FL",
    "GA", "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME",
    "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH",
    "NJ", "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI",
    "SC", "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI", "WY",
]


# ---------------------------------------------------------------------------
# Pre-populated example projects (diverse statuses, geographies, sizes)
# (proj_num, name, client, ptype, csi_div, state, city,
#  start, target, actual, status, contract, co_net, cost_to_date, eac,
#  pct_complete, pm, notes)
# ---------------------------------------------------------------------------

EXAMPLE_PROJECTS: list[tuple] = [
    ("2026-014", "Riverside Office Bldg", "Riverside Holdings LLC",
     "New Construction", "03", "TN", "Knoxville",
     "2026-03-01", "2026-11-30", "",
     "Active", 1850000.00, 22725.00, 540000.00, 1720000.00, 30,
     "M. Reyes", "On schedule; concrete topping out next month."),

    ("2026-021", "Maple Plaza TI Buildout", "Maple Capital Partners",
     "Tenant Improvement", "09", "TN", "Nashville",
     "2026-04-15", "2026-09-15", "",
     "Mobilizing", 425000.00, 0.00, 18000.00, 390000.00, 5,
     "M. Reyes", "Permits in hand; demo starts next week."),

    ("2026-009", "Hillcrest Medical Remodel", "Hillcrest Health System",
     "Remodel", "23", "GA", "Atlanta",
     "2026-01-10", "2026-04-30", "2026-05-08",
     "Substantial Completion", 1240000.00, 41200.00, 1198000.00, 1215000.00, 98,
     "S. Patel", "Punchlist in progress; final inspection scheduled."),

    ("2026-027", "Eastside Warehouse Roof Repair", "Eastside Logistics Inc",
     "Repair", "07", "AL", "Birmingham",
     "2026-05-05", "2026-05-20", "",
     "Awarded", 87500.00, 0.00, 0.00, 76000.00, 0,
     "J. Cole", "Insurance claim cleared; mob next Monday."),

    ("2026-031", "Civic Center Expansion (BID)", "City of Greenville",
     "New Construction", "01", "SC", "Greenville",
     "", "", "",
     "Bidding", 0.00, 0.00, 0.00, 0.00, 0,
     "M. Reyes", "Bid due 2026-06-02; preconstruction cost estimate $4.2M."),
]
EMPTY_PROJECT_ROWS = 20


# Company Info row (one row of company metadata)
COMPANY_INFO_HEADERS = [
    "Company Name", "EIN", "Primary State",
    "Workers' Comp Carrier", "Bond Carrier", "GL Carrier", "Notes",
]
COMPANY_EXAMPLE = [
    "EXAMPLE - Carolina Builders LLC",
    "00-0000000",
    "TN",
    "Travelers (WC)",
    "Liberty Mutual Surety",
    "The Hartford (GL)",
    "Example row - replace with your company info.",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def set_col_widths(ws, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_header_row(ws, row: int, last_col: int) -> None:
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 28


def border_range(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER


# ---------------------------------------------------------------------------
# Hidden tab — CSI Reference + named ranges (including States)
# ---------------------------------------------------------------------------

def build_csi_reference(wb: Workbook) -> str:
    ws = wb.create_sheet("CSI Reference")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "CSI MasterFormat + Reference Lists"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:B1")

    ws["A2"] = ("Pulled from the same canonical lists as ContrPro Job Costing "
                "and Change Order Log. Do not edit.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)

    # CSI Division table
    ws.cell(row=3, column=1, value="CSI Division")
    ws.cell(row=3, column=2, value="Division Name")
    style_header_row(ws, 3, 2)

    for i, (code, name) in enumerate(DIVISIONS, start=4):
        ws.cell(row=i, column=1, value=code).number_format = "@"
        ws.cell(row=i, column=2, value=name)
        for c in (1, 2):
            ws.cell(row=i, column=c).border = BORDER
            ws.cell(row=i, column=c).font = FONT_BODY

    csi_first = 4
    csi_last = 4 + len(DIVISIONS) - 1

    # States column (D) - kept in a separate column so range is contiguous
    ws.cell(row=3, column=4, value="State")
    for c in (4,):
        cell = ws.cell(row=3, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for i, st in enumerate(STATES_50_DC, start=4):
        ws.cell(row=i, column=4, value=st)
        ws.cell(row=i, column=4).number_format = "@"
        ws.cell(row=i, column=4).border = BORDER
        ws.cell(row=i, column=4).font = FONT_BODY

    state_first = 4
    state_last = 4 + len(STATES_50_DC) - 1

    set_col_widths(ws, {"A": 14, "B": 60, "C": 4, "D": 10})

    # Named ranges
    div_range = f"'CSI Reference'!$A${csi_first}:$A${csi_last}"
    full_range = f"'CSI Reference'!$A${csi_first}:$B${csi_last}"
    state_range = f"'CSI Reference'!$D${state_first}:$D${state_last}"
    wb.defined_names["CSI_Divisions"] = DefinedName("CSI_Divisions", attr_text=div_range)
    wb.defined_names["CSI_Table"] = DefinedName("CSI_Table", attr_text=full_range)
    wb.defined_names["States_List"] = DefinedName("States_List", attr_text=state_range)

    return ws.title


# ---------------------------------------------------------------------------
# Tab 1 — Instructions
# ---------------------------------------------------------------------------

def build_instructions(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {"A": 4, "B": 110})

    ws["B2"] = "ContrPro - Project Tracker (Portfolio View)"
    ws["B2"].font = FONT_TITLE
    ws.row_dimensions[2].height = 32

    ws["B3"] = ("Portfolio-level tracker for every open job at once. The view a GC owner needs Monday morning.")
    ws["B3"].font = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

    sections: list[tuple[str, list[str]]] = [
        (
            "What this workbook is (and what it is NOT)",
            [
                "This is a PORTFOLIO tracker - one row per project, all your active jobs in one view. "
                "Use Job_Costing_Spreadsheet.xlsx (per-job, CSI-coded) for the deep cost detail on any single job. "
                "Use Change_Order_Log.xlsx for the CO paper trail. This file is the bird's-eye view.",
                "Six tabs work together: Company Info pins your company metadata, Project Portfolio holds the rows, "
                "Portfolio Summary by Status auto-rolls, and Dashboard surfaces the KPIs that matter. "
                "CSI Reference is hidden - it feeds the dropdowns.",
            ],
        ),
        (
            "KPIs explained (read this once - it changes how you use this file)",
            [
                "Backlog = Adjusted Contract minus Earned Value. The work you have under contract but haven't "
                "billed yet. This is what banks and bonding companies care about most. Healthy backlog = 6-12 "
                "months of revenue at your current burn rate.",
                "Burn Rate = Cost to Date divided by elapsed days. Tells you how fast a job is consuming cash. "
                "Compare against EAC / total duration to spot jobs that are eating themselves.",
                "Projected Margin % = (Adjusted Contract - Estimated Cost at Completion) / Adjusted Contract. "
                "Green > 10%, yellow 0-10%, RED < 0. A red row means you are losing money on that job - call "
                "your PM today.",
                "% Complete = self-reported field assessment, dollar-weighted on the Dashboard. Use the field "
                "PM's number, not a calculated one - the gap between field % and earned % is a leading indicator "
                "of trouble.",
                "Earned Value = Adjusted Contract * %Complete. This is what you SHOULD have billed by now if "
                "your schedule of values is honest.",
            ],
        ),
        (
            "How to use the Project Portfolio tab",
            [
                "One row per project. Pre-populated rows show diverse statuses (Active, Mobilizing, Substantial "
                "Completion, Awarded, Bidding) - replace them or add new rows below.",
                "Dropdowns: Project Type, Primary CSI Division, State, Status. Pick from the list - typos break "
                "the Summary roll-up.",
                "FORMULA columns (do NOT type into these - they calculate themselves): Adjusted Contract, "
                "Projected Gross Margin, Projected Margin %, Earned Value, Backlog.",
                "Status colors: green for Active / Substantial Completion / Final, yellow for Mobilizing / Bidding, "
                "blue for Awarded, red for Lost. Closed shows as default.",
                "% Complete column has a green data bar so you can scan progress visually down the column.",
            ],
        ),
        (
            "Company Info tab",
            [
                "One row with your company's metadata - Company Name, EIN, primary state of operation, and your "
                "insurance / bond carriers. The Dashboard does not pull from this tab, but downstream files "
                "(COI Tracker, Subcontractor Tracker) will reference the same fields - keep it current so a "
                "future ContrPro release can auto-fill those workbooks.",
            ],
        ),
        (
            "Portfolio Summary by Status (auto roll-up)",
            [
                "Reads Project Portfolio via COUNTIF/SUMIFS. Do not type into this tab.",
                "Shows count, total contract, total earned, total cost-to-date, total backlog, and average margin "
                "% for every status bucket. If your active backlog drops while your closed-job count climbs, you "
                "have a sales pipeline problem - not a delivery problem.",
            ],
        ),
        (
            "Dashboard",
            [
                "Top panel = portfolio KPIs: Total Active Backlog, Total Cost to Date, Weighted Average Margin %, "
                "and Active Project Count.",
                "Top 5 Largest Backlogs panel uses LARGE() so it always shows your biggest exposures, regardless "
                "of row order in the Portfolio tab.",
                "Aged Projects panel flags any project where Target Completion is BEFORE today AND status is not "
                "yet Final / Closed. These are the jobs that need a hard conversation this week.",
            ],
        ),
        (
            "Weekly cadence (Monday review meeting)",
            [
                "Monday AM: open this file. Scan Dashboard KPIs first. Then scroll the Portfolio tab from top to "
                "bottom - any red margin gets a phone call before lunch.",
                "Friday PM: PMs update % Complete, Cost to Date, and Estimated Cost at Completion on their rows. "
                "No row goes a week without an update; a stale row is a lying row.",
                "End of month: archive a copy with a date stamp (Project_Tracker_2026-05-31.xlsx) for board "
                "reviews and lender packages.",
            ],
        ),
        (
            "Tabs in this workbook",
            [
                "Instructions  -  this tab",
                "Company Info  -  one-row company metadata",
                "Project Portfolio  -  main worksheet, one row per job",
                "Portfolio Summary by Status  -  auto roll-up by status",
                "Dashboard  -  portfolio KPIs, top-5 backlogs, aged projects",
                "CSI Reference  -  canonical division + state lists (hidden, do not edit)",
            ],
        ),
    ]

    row = 5
    for heading, paragraphs in sections:
        ws.cell(row=row, column=2, value=heading)
        ws.cell(row=row, column=2).font = FONT_H2
        ws.cell(row=row, column=2).fill = FILL_SUBHEADER
        ws.cell(row=row, column=2).alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[row].height = 22
        row += 1
        for para in paragraphs:
            ws.cell(row=row, column=2, value=para)
            ws.cell(row=row, column=2).font = FONT_BODY
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = max(18, 16 * (1 + len(para) // 110))
            row += 1
        row += 1

    ws.cell(row=row + 1, column=2,
            value="ContrPro - built for builders. Questions: support@contrpro.com")
    ws.cell(row=row + 1, column=2).font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)


# ---------------------------------------------------------------------------
# Tab 2 — Company Info
# ---------------------------------------------------------------------------

def build_company_info(wb: Workbook) -> None:
    ws = wb.create_sheet("Company Info")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Company Information"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:G1")

    ws["A2"] = "Shared metadata referenced by future ContrPro workbooks (COI, Subcontractor, etc.)."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)
    ws.merge_cells("A2:G2")

    for i, h in enumerate(COMPANY_INFO_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(COMPANY_INFO_HEADERS))

    # Example row (greyed/italic)
    for i, v in enumerate(COMPANY_EXAMPLE, start=1):
        cell = ws.cell(row=4, column=i, value=v)
        cell.font = FONT_GREY_ITALIC
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[4].height = 24

    # State dropdown on Primary State (col C)
    dv_state = DataValidation(
        type="list", formula1="=States_List", allow_blank=True,
    )
    dv_state.error = "Pick a state from the dropdown."
    dv_state.errorTitle = "Invalid state"
    ws.add_data_validation(dv_state)
    dv_state.add("C4")

    set_col_widths(ws, {
        "A": 32, "B": 14, "C": 14,
        "D": 26, "E": 26, "F": 26, "G": 45,
    })

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Tab 3 — Project Portfolio (main worksheet)
# ---------------------------------------------------------------------------

PORTFOLIO_HEADERS = [
    "Project Number",                      # A
    "Project Name",                         # B
    "Client",                               # C
    "Project Type",                         # D (dropdown)
    "Primary CSI Division",                 # E (dropdown)
    "State",                                # F (dropdown)
    "City",                                 # G
    "Start Date",                           # H
    "Target Completion",                    # I
    "Actual Completion",                    # J
    "Status",                               # K (dropdown)
    "Contract Amount",                      # L
    "Approved CO Net",                      # M
    "Adjusted Contract",                    # N (FORMULA = L + M)
    "Cost to Date",                         # O
    "Estimated Cost at Completion",         # P
    "Projected Gross Margin",               # Q (FORMULA = N - P)
    "Projected Margin %",                   # R (FORMULA = Q / N)
    "% Complete",                           # S
    "Earned Value",                         # T (FORMULA = N * S/100)
    "Backlog",                              # U (FORMULA = N - T)
    "Project Manager",                      # V
    "Notes",                                # W
]


def build_portfolio(wb: Workbook) -> Tuple[str, int, int]:
    ws = wb.create_sheet("Project Portfolio")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Project Portfolio (one row per job)"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(PORTFOLIO_HEADERS))

    # Header row at row 3
    for i, h in enumerate(PORTFOLIO_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(PORTFOLIO_HEADERS))

    data_start = 4
    n_pre = len(EXAMPLE_PROJECTS)
    n_total = n_pre + EMPTY_PROJECT_ROWS
    data_end = data_start + n_total - 1

    # Pre-populated example rows
    for idx, row_data in enumerate(EXAMPLE_PROJECTS):
        r = data_start + idx
        (proj_num, name, client, ptype, csi_div, state, city,
         start, target, actual, status, contract, co_net,
         cost_to_date, eac, pct_complete, pm, notes) = row_data

        ws.cell(row=r, column=1, value=proj_num)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=client)
        ws.cell(row=r, column=4, value=ptype)
        ws.cell(row=r, column=5, value=csi_div)
        ws.cell(row=r, column=6, value=state)
        ws.cell(row=r, column=7, value=city)
        ws.cell(row=r, column=8, value=start if start else None)
        ws.cell(row=r, column=9, value=target if target else None)
        ws.cell(row=r, column=10, value=actual if actual else None)
        ws.cell(row=r, column=11, value=status)
        ws.cell(row=r, column=12, value=contract)
        ws.cell(row=r, column=13, value=co_net)
        # N = Adjusted Contract = L + M
        ws.cell(row=r, column=14, value=f"=IFERROR(L{r}+M{r},0)")
        ws.cell(row=r, column=15, value=cost_to_date)
        ws.cell(row=r, column=16, value=eac)
        # Q = Projected Gross Margin = N - P
        ws.cell(row=r, column=17, value=f"=IFERROR(N{r}-P{r},0)")
        # R = Projected Margin % = Q / N
        ws.cell(row=r, column=18, value=f'=IFERROR(Q{r}/N{r},0)')
        ws.cell(row=r, column=19, value=pct_complete)
        # T = Earned Value = N * S / 100
        ws.cell(row=r, column=20, value=f"=IFERROR(N{r}*S{r}/100,0)")
        # U = Backlog = N - T
        ws.cell(row=r, column=21, value=f"=IFERROR(N{r}-T{r},0)")
        ws.cell(row=r, column=22, value=pm)
        ws.cell(row=r, column=23, value=notes)

    # Empty user-fillable rows (formulas pre-seeded)
    for i in range(EMPTY_PROJECT_ROWS):
        r = data_start + n_pre + i
        ws.cell(row=r, column=14, value=f"=IFERROR(L{r}+M{r},0)")
        ws.cell(row=r, column=17, value=f"=IFERROR(N{r}-P{r},0)")
        ws.cell(row=r, column=18, value=f'=IFERROR(Q{r}/N{r},0)')
        ws.cell(row=r, column=20, value=f"=IFERROR(N{r}*S{r}/100,0)")
        ws.cell(row=r, column=21, value=f"=IFERROR(N{r}-T{r},0)")

    # Per-row formatting
    for r in range(data_start, data_end + 1):
        for c in range(1, len(PORTFOLIO_HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(c in (2, 3, 23)),
            )
        # CSI Division as text (preserve leading zeros)
        ws.cell(row=r, column=5).number_format = "@"
        ws.cell(row=r, column=6).number_format = "@"
        # Dates
        for c in (8, 9, 10):
            ws.cell(row=r, column=c).number_format = FMT_DATE
        # Currency cols
        for c in (12, 13, 14, 15, 16, 17, 20, 21):
            ws.cell(row=r, column=c).number_format = FMT_USD
        # Margin %
        ws.cell(row=r, column=18).number_format = FMT_PCT
        # % Complete (integer 0-100, displayed as "0%")
        ws.cell(row=r, column=19).number_format = "0\"%\""
        ws.row_dimensions[r].height = 30

    # --- Data validation ---

    # Project Type (col D)
    dv_ptype = DataValidation(
        type="list", formula1='"' + ",".join(PROJECT_TYPES) + '"', allow_blank=True,
    )
    dv_ptype.error = "Pick a project type from the dropdown."
    dv_ptype.errorTitle = "Invalid project type"
    ws.add_data_validation(dv_ptype)
    dv_ptype.add(f"D{data_start}:D{data_end}")

    # CSI Division (col E)
    dv_csi = DataValidation(
        type="list", formula1="=CSI_Divisions", allow_blank=True,
    )
    dv_csi.error = "Pick a CSI Division from the dropdown."
    dv_csi.errorTitle = "Invalid CSI Division"
    ws.add_data_validation(dv_csi)
    dv_csi.add(f"E{data_start}:E{data_end}")

    # State (col F)
    dv_state = DataValidation(
        type="list", formula1="=States_List", allow_blank=True,
    )
    dv_state.error = "Pick a state from the dropdown."
    dv_state.errorTitle = "Invalid state"
    ws.add_data_validation(dv_state)
    dv_state.add(f"F{data_start}:F{data_end}")

    # Status (col K)
    dv_status = DataValidation(
        type="list", formula1='"' + ",".join(PROJECT_STATUSES) + '"', allow_blank=True,
    )
    dv_status.error = "Pick a status from the dropdown."
    dv_status.errorTitle = "Invalid status"
    ws.add_data_validation(dv_status)
    dv_status.add(f"K{data_start}:K{data_end}")

    # % Complete must be 0..100
    dv_pct = DataValidation(
        type="decimal", operator="between",
        formula1=0, formula2=100, allow_blank=True,
    )
    dv_pct.error = "Enter a value between 0 and 100."
    dv_pct.errorTitle = "Out of range"
    ws.add_data_validation(dv_pct)
    dv_pct.add(f"S{data_start}:S{data_end}")

    # --- Conditional formatting ---

    # Status (col K) — colored by value
    status_range = f"K{data_start}:K{data_end}"
    green_statuses = ["Active", "Substantial Completion", "Final"]
    yellow_statuses = ["Mobilizing", "Bidding"]
    blue_statuses = ["Awarded"]
    red_statuses = ["Lost"]

    for st in green_statuses:
        ws.conditional_formatting.add(
            status_range,
            FormulaRule(formula=[f'$K{data_start}="{st}"'], stopIfTrue=False,
                        fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True)),
        )
    for st in yellow_statuses:
        ws.conditional_formatting.add(
            status_range,
            FormulaRule(formula=[f'$K{data_start}="{st}"'], stopIfTrue=False,
                        fill=FILL_YELLOW, font=Font(color=YELLOW_FONT, bold=True)),
        )
    for st in blue_statuses:
        ws.conditional_formatting.add(
            status_range,
            FormulaRule(formula=[f'$K{data_start}="{st}"'], stopIfTrue=False,
                        fill=FILL_BLUE, font=Font(color=BLUE_FONT, bold=True)),
        )
    for st in red_statuses:
        ws.conditional_formatting.add(
            status_range,
            FormulaRule(formula=[f'$K{data_start}="{st}"'], stopIfTrue=False,
                        fill=FILL_RED, font=Font(color=RED_FONT, bold=True)),
        )

    # Projected Margin % (col R)
    margin_range = f"R{data_start}:R{data_end}"
    ws.conditional_formatting.add(
        margin_range,
        CellIsRule(operator="lessThan", formula=["0"], stopIfTrue=True,
                   fill=FILL_RED, font=Font(color=RED_FONT, bold=True)),
    )
    ws.conditional_formatting.add(
        margin_range,
        CellIsRule(operator="between", formula=["0", "0.10"], stopIfTrue=True,
                   fill=FILL_YELLOW, font=Font(color=YELLOW_FONT, bold=True)),
    )
    ws.conditional_formatting.add(
        margin_range,
        CellIsRule(operator="greaterThan", formula=["0.10"], stopIfTrue=False,
                   fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True)),
    )

    # % Complete (col S) — green data bar
    pct_range = f"S{data_start}:S{data_end}"
    ws.conditional_formatting.add(
        pct_range,
        DataBarRule(
            start_type="num", start_value=0,
            end_type="num", end_value=100,
            color=GREEN_FONT,
            showValue=True,
        ),
    )

    # --- Totals row ---
    totals_row = data_end + 2
    ws.cell(row=totals_row, column=11, value="TOTALS").font = FONT_BODY_BOLD
    ws.cell(row=totals_row, column=11).alignment = Alignment(horizontal="right")
    ws.cell(row=totals_row, column=11).fill = FILL_SUBHEADER
    ws.cell(row=totals_row, column=11).border = BORDER

    for col_letter in ("L", "M", "N", "O", "P", "Q", "T", "U"):
        c_idx = ord(col_letter) - ord("A") + 1
        cell = ws.cell(row=totals_row, column=c_idx,
                       value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})")
        cell.number_format = FMT_USD
        cell.font = FONT_BODY_BOLD
        cell.fill = FILL_SUBHEADER
        cell.border = BORDER

    # Weighted margin % at totals row
    ws.cell(row=totals_row, column=18,
            value=(f"=IFERROR(SUM(Q{data_start}:Q{data_end})"
                   f"/SUM(N{data_start}:N{data_end}),0)"))
    ws.cell(row=totals_row, column=18).number_format = FMT_PCT
    ws.cell(row=totals_row, column=18).font = FONT_BODY_BOLD
    ws.cell(row=totals_row, column=18).fill = FILL_SUBHEADER
    ws.cell(row=totals_row, column=18).border = BORDER

    # Weighted % complete at totals row (dollar-weighted by Adjusted Contract)
    ws.cell(row=totals_row, column=19,
            value=(f"=IFERROR(SUMPRODUCT(N{data_start}:N{data_end},"
                   f"S{data_start}:S{data_end})"
                   f"/SUM(N{data_start}:N{data_end}),0)"))
    ws.cell(row=totals_row, column=19).number_format = "0\"%\""
    ws.cell(row=totals_row, column=19).font = FONT_BODY_BOLD
    ws.cell(row=totals_row, column=19).fill = FILL_SUBHEADER
    ws.cell(row=totals_row, column=19).border = BORDER

    # --- Column widths ---
    set_col_widths(ws, {
        "A": 12, "B": 28, "C": 24, "D": 16, "E": 12, "F": 8, "G": 14,
        "H": 12, "I": 13, "J": 13, "K": 16,
        "L": 14, "M": 14, "N": 16, "O": 14, "P": 16,
        "Q": 16, "R": 12, "S": 11, "T": 14, "U": 14,
        "V": 16, "W": 32,
    })

    ws.freeze_panes = "C4"

    return ws.title, data_start, data_end


# ---------------------------------------------------------------------------
# Tab 4 — Portfolio Summary by Status
# ---------------------------------------------------------------------------

STATUS_SUM_HEADERS = [
    "Status", "Count", "Total Contract", "Total Earned Value",
    "Total Cost to Date", "Total Backlog", "Avg Margin %",
]


def build_status_summary(wb: Workbook, pt_sheet: str, pt_start: int, pt_end: int) -> None:
    ws = wb.create_sheet("Portfolio Summary by Status")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Portfolio Summary by Status (auto roll-up)"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(STATUS_SUM_HEADERS))

    ws["A2"] = ("Reads from Project Portfolio via COUNTIF/SUMIFS - do not edit. "
                "One row per status bucket.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)

    for i, h in enumerate(STATUS_SUM_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(STATUS_SUM_HEADERS))

    # Portfolio ranges
    pt_status = f"'{pt_sheet}'!$K${pt_start}:$K${pt_end}"
    pt_adjcontract = f"'{pt_sheet}'!$N${pt_start}:$N${pt_end}"
    pt_costtodate = f"'{pt_sheet}'!$O${pt_start}:$O${pt_end}"
    pt_margin = f"'{pt_sheet}'!$R${pt_start}:$R${pt_end}"
    pt_earned = f"'{pt_sheet}'!$T${pt_start}:$T${pt_end}"
    pt_backlog = f"'{pt_sheet}'!$U${pt_start}:$U${pt_end}"

    start_row = 4
    for idx, status in enumerate(PROJECT_STATUSES):
        r = start_row + idx
        ws.cell(row=r, column=1, value=status)
        # Count
        ws.cell(row=r, column=2,
                value=f'=COUNTIF({pt_status},A{r})')
        # Total Contract (uses Adjusted Contract for accuracy)
        ws.cell(row=r, column=3,
                value=f'=SUMIFS({pt_adjcontract},{pt_status},A{r})')
        # Total Earned Value
        ws.cell(row=r, column=4,
                value=f'=SUMIFS({pt_earned},{pt_status},A{r})')
        # Total Cost to Date
        ws.cell(row=r, column=5,
                value=f'=SUMIFS({pt_costtodate},{pt_status},A{r})')
        # Total Backlog
        ws.cell(row=r, column=6,
                value=f'=SUMIFS({pt_backlog},{pt_status},A{r})')
        # Avg Margin % (simple average across rows where status matches AND has a contract)
        ws.cell(row=r, column=7,
                value=(f'=IFERROR(AVERAGEIFS({pt_margin},'
                       f'{pt_status},A{r},{pt_adjcontract},">0"),0)'))

        ws.cell(row=r, column=2).number_format = FMT_INT
        for c in (3, 4, 5, 6):
            ws.cell(row=r, column=c).number_format = FMT_USD
        ws.cell(row=r, column=7).number_format = FMT_PCT

        for c in range(1, len(STATUS_SUM_HEADERS) + 1):
            ws.cell(row=r, column=c).border = BORDER
        ws.row_dimensions[r].height = 22

        # Color-code status label cell
        if status in ("Active", "Substantial Completion", "Final"):
            ws.cell(row=r, column=1).fill = FILL_GREEN
            ws.cell(row=r, column=1).font = Font(color=GREEN_FONT, bold=True)
        elif status in ("Mobilizing", "Bidding"):
            ws.cell(row=r, column=1).fill = FILL_YELLOW
            ws.cell(row=r, column=1).font = Font(color=YELLOW_FONT, bold=True)
        elif status == "Awarded":
            ws.cell(row=r, column=1).fill = FILL_BLUE
            ws.cell(row=r, column=1).font = Font(color=BLUE_FONT, bold=True)
        elif status == "Lost":
            ws.cell(row=r, column=1).fill = FILL_RED
            ws.cell(row=r, column=1).font = Font(color=RED_FONT, bold=True)
        else:
            ws.cell(row=r, column=1).fill = FILL_SUBHEADER
            ws.cell(row=r, column=1).font = FONT_BODY_BOLD

    end_row = start_row + len(PROJECT_STATUSES) - 1

    # Totals row
    tr = end_row + 2
    ws.cell(row=tr, column=1, value="TOTALS").font = FONT_BODY_BOLD
    ws.cell(row=tr, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=tr, column=1).fill = FILL_SUBHEADER
    ws.cell(row=tr, column=1).border = BORDER

    ws.cell(row=tr, column=2, value=f"=SUM(B{start_row}:B{end_row})")
    ws.cell(row=tr, column=2).number_format = FMT_INT
    ws.cell(row=tr, column=2).font = FONT_BODY_BOLD
    ws.cell(row=tr, column=2).fill = FILL_SUBHEADER
    ws.cell(row=tr, column=2).border = BORDER

    for col_letter in ("C", "D", "E", "F"):
        c_idx = ord(col_letter) - ord("A") + 1
        cell = ws.cell(row=tr, column=c_idx,
                       value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
        cell.number_format = FMT_USD
        cell.font = FONT_BODY_BOLD
        cell.fill = FILL_SUBHEADER
        cell.border = BORDER

    # Avg Margin % weighted by contract
    ws.cell(row=tr, column=7,
            value=(f"=IFERROR(SUMPRODUCT(C{start_row}:C{end_row},"
                   f"G{start_row}:G{end_row})"
                   f"/SUM(C{start_row}:C{end_row}),0)"))
    ws.cell(row=tr, column=7).number_format = FMT_PCT
    ws.cell(row=tr, column=7).font = FONT_BODY_BOLD
    ws.cell(row=tr, column=7).fill = FILL_SUBHEADER
    ws.cell(row=tr, column=7).border = BORDER

    set_col_widths(ws, {
        "A": 24, "B": 10, "C": 20, "D": 20, "E": 20, "F": 20, "G": 14,
    })

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Tab 5 — Dashboard
# ---------------------------------------------------------------------------

def build_dashboard(wb: Workbook, pt_sheet: str, pt_start: int, pt_end: int) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Portfolio Dashboard"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:F1")

    ws["A2"] = ("Auto-calculated from Project Portfolio. Update that tab - this one stays in sync.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)
    ws.merge_cells("A2:F2")

    # Portfolio ranges
    pt_name = f"'{pt_sheet}'!$B${pt_start}:$B${pt_end}"
    pt_status = f"'{pt_sheet}'!$K${pt_start}:$K${pt_end}"
    pt_adjcontract = f"'{pt_sheet}'!$N${pt_start}:$N${pt_end}"
    pt_costtodate = f"'{pt_sheet}'!$O${pt_start}:$O${pt_end}"
    pt_target = f"'{pt_sheet}'!$I${pt_start}:$I${pt_end}"
    pt_backlog = f"'{pt_sheet}'!$U${pt_start}:$U${pt_end}"
    pt_margin_q = f"'{pt_sheet}'!$Q${pt_start}:$Q${pt_end}"

    # ---------------- Top panel: portfolio KPIs ----------------
    panel_start = 4
    ws.cell(row=panel_start, column=1, value="Portfolio KPIs").font = FONT_H1
    ws.cell(row=panel_start, column=1).fill = FILL_HEADER
    ws.cell(row=panel_start, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=panel_start, start_column=1,
                   end_row=panel_start, end_column=6)
    ws.row_dimensions[panel_start].height = 26

    # "Active" backlog = backlog where status is in the active-revenue bucket
    # (Active, Mobilizing, Substantial Completion, Awarded). Lost / Closed / Bidding excluded.
    active_states = ("Active", "Mobilizing", "Substantial Completion", "Awarded")
    active_sumifs_backlog = " + ".join(
        f'SUMIFS({pt_backlog},{pt_status},"{s}")' for s in active_states
    )
    active_sumifs_cost = " + ".join(
        f'SUMIFS({pt_costtodate},{pt_status},"{s}")' for s in active_states
    )
    active_sumifs_contract = " + ".join(
        f'SUMIFS({pt_adjcontract},{pt_status},"{s}")' for s in active_states
    )
    active_sumifs_margin = " + ".join(
        f'SUMIFS({pt_margin_q},{pt_status},"{s}")' for s in active_states
    )
    active_count_formula = " + ".join(
        f'COUNTIF({pt_status},"{s}")' for s in active_states
    )

    metrics = [
        ("Total Active Backlog",
         f"={active_sumifs_backlog}", FMT_USD, None),
        ("Total Cost to Date (Active)",
         f"={active_sumifs_cost}", FMT_USD, None),
        ("Weighted Avg Margin % (Active)",
         f"=IFERROR(({active_sumifs_margin})/({active_sumifs_contract}),0)",
         FMT_PCT, "margin"),
        ("Active Project Count",
         f"={active_count_formula}", FMT_INT, None),
        ("All Statuses - Total Contract",
         f"=SUM({pt_adjcontract})", FMT_USD, None),
        ("All Statuses - Total Backlog",
         f"=SUM({pt_backlog})", FMT_USD, None),
    ]

    row = panel_start + 1
    margin_cells: list[str] = []
    for label, formula, fmt, kind in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = FONT_BODY_BOLD
        ws.cell(row=row, column=1).fill = FILL_SUMMARY_LABEL
        ws.cell(row=row, column=1).alignment = Alignment(vertical="center", indent=1)
        ws.cell(row=row, column=1).border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

        ws.cell(row=row, column=4, value=formula)
        ws.cell(row=row, column=4).font = FONT_BIG_NUMBER
        ws.cell(row=row, column=4).number_format = fmt
        ws.cell(row=row, column=4).alignment = Alignment(
            horizontal="right", vertical="center", indent=1,
        )
        ws.cell(row=row, column=4).border = BORDER
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)

        if kind == "margin":
            margin_cells.append(f"D{row}")
        ws.row_dimensions[row].height = 28
        row += 1

    # CF on margin cells: red < 0, yellow 0-10%, green > 10%
    for ref in margin_cells:
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="lessThan", formula=["0"], stopIfTrue=True,
                       fill=FILL_RED, font=Font(color=RED_FONT, bold=True, size=18)),
        )
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="between", formula=["0", "0.10"], stopIfTrue=True,
                       fill=FILL_YELLOW, font=Font(color=YELLOW_FONT, bold=True, size=18)),
        )
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="greaterThan", formula=["0.10"], stopIfTrue=False,
                       fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True, size=18)),
        )

    # ---------------- Middle panel: Top 5 largest backlogs ----------------
    section_row = row + 2
    ws.cell(row=section_row, column=1, value="Top 5 Largest Backlogs").font = FONT_H1
    ws.cell(row=section_row, column=1).fill = FILL_HEADER
    ws.cell(row=section_row, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=section_row, start_column=1,
                   end_row=section_row, end_column=6)
    ws.row_dimensions[section_row].height = 26

    headers = ["Rank", "Project Name", "Status", "Backlog"]
    hr = section_row + 1
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hr, column=i, value=h)
    # Merge Project Name across cols 2-3, Status col 4, Backlog cols 5-6
    ws.merge_cells(start_row=hr, start_column=2, end_row=hr, end_column=3)
    ws.merge_cells(start_row=hr, start_column=5, end_row=hr, end_column=6)
    style_header_row(ws, hr, 6)

    top5_start = hr + 1
    for i in range(5):
        r = top5_start + i
        rank = i + 1
        ws.cell(row=r, column=1, value=rank)
        ws.cell(row=r, column=1).font = FONT_BODY_BOLD
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=1).fill = FILL_SUMMARY_LABEL
        ws.cell(row=r, column=1).border = BORDER

        # Backlog rank using LARGE
        backlog_formula = f"=IFERROR(LARGE({pt_backlog},{rank}),0)"
        # Use INDEX/MATCH to find the project name/status for that backlog value
        # (handles ties imperfectly but works for typical small portfolios)
        name_formula = (
            f'=IFERROR(INDEX({pt_name},MATCH(LARGE({pt_backlog},{rank}),'
            f'{pt_backlog},0)),"")'
        )
        status_formula = (
            f'=IFERROR(INDEX({pt_status},MATCH(LARGE({pt_backlog},{rank}),'
            f'{pt_backlog},0)),"")'
        )

        ws.cell(row=r, column=2, value=name_formula)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.cell(row=r, column=2).font = FONT_BODY
        ws.cell(row=r, column=2).alignment = Alignment(vertical="center", indent=1, wrap_text=True)
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=3).border = BORDER

        ws.cell(row=r, column=4, value=status_formula)
        ws.cell(row=r, column=4).font = FONT_BODY_BOLD
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=4).border = BORDER

        ws.cell(row=r, column=5, value=backlog_formula)
        ws.cell(row=r, column=5).font = FONT_BODY_BOLD
        ws.cell(row=r, column=5).number_format = FMT_USD
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.cell(row=r, column=5).border = BORDER
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        ws.cell(row=r, column=6).border = BORDER
        ws.row_dimensions[r].height = 24

    # ---------------- Bottom panel: Aged projects ----------------
    section_row = top5_start + 5 + 2
    ws.cell(row=section_row, column=1,
            value="Aged Projects (past Target Completion, not yet Final/Closed)").font = FONT_H1
    ws.cell(row=section_row, column=1).fill = FILL_HEADER
    ws.cell(row=section_row, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=section_row, start_column=1,
                   end_row=section_row, end_column=6)
    ws.row_dimensions[section_row].height = 26

    # Show a count + an instruction to filter the Portfolio tab to find them.
    # Excel SUMPRODUCT trick: count rows where target < today() AND status not in (Final, Closed, Lost, Bidding)
    # and contract > 0.
    aged_count_formula = (
        f'=SUMPRODUCT(({pt_target}<TODAY())*({pt_target}<>"")'
        f'*({pt_status}<>"Final")*({pt_status}<>"Closed")'
        f'*({pt_status}<>"Lost")*({pt_status}<>"Bidding"))'
    )
    aged_backlog_formula = (
        f'=SUMPRODUCT(({pt_target}<TODAY())*({pt_target}<>"")'
        f'*({pt_status}<>"Final")*({pt_status}<>"Closed")'
        f'*({pt_status}<>"Lost")*({pt_status}<>"Bidding")*{pt_backlog})'
    )

    aged_row = section_row + 1
    ws.cell(row=aged_row, column=1, value="Count of aged projects")
    ws.cell(row=aged_row, column=1).font = FONT_BODY_BOLD
    ws.cell(row=aged_row, column=1).fill = FILL_SUMMARY_LABEL
    ws.cell(row=aged_row, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.cell(row=aged_row, column=1).border = BORDER
    ws.merge_cells(start_row=aged_row, start_column=1, end_row=aged_row, end_column=3)

    ws.cell(row=aged_row, column=4, value=aged_count_formula)
    ws.cell(row=aged_row, column=4).font = FONT_BIG_NUMBER
    ws.cell(row=aged_row, column=4).number_format = FMT_INT
    ws.cell(row=aged_row, column=4).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.cell(row=aged_row, column=4).border = BORDER
    ws.merge_cells(start_row=aged_row, start_column=4, end_row=aged_row, end_column=6)
    ws.row_dimensions[aged_row].height = 28

    # Red highlight if any aged projects
    ws.conditional_formatting.add(
        f"D{aged_row}",
        CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=False,
                   fill=FILL_RED, font=Font(color=RED_FONT, bold=True, size=18)),
    )

    aged2_row = aged_row + 1
    ws.cell(row=aged2_row, column=1, value="Backlog tied up in aged projects")
    ws.cell(row=aged2_row, column=1).font = FONT_BODY_BOLD
    ws.cell(row=aged2_row, column=1).fill = FILL_SUMMARY_LABEL
    ws.cell(row=aged2_row, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.cell(row=aged2_row, column=1).border = BORDER
    ws.merge_cells(start_row=aged2_row, start_column=1, end_row=aged2_row, end_column=3)

    ws.cell(row=aged2_row, column=4, value=aged_backlog_formula)
    ws.cell(row=aged2_row, column=4).font = FONT_BIG_NUMBER
    ws.cell(row=aged2_row, column=4).number_format = FMT_USD
    ws.cell(row=aged2_row, column=4).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.cell(row=aged2_row, column=4).border = BORDER
    ws.merge_cells(start_row=aged2_row, start_column=4, end_row=aged2_row, end_column=6)
    ws.row_dimensions[aged2_row].height = 28

    ws.conditional_formatting.add(
        f"D{aged2_row}",
        CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=False,
                   fill=FILL_RED, font=Font(color=RED_FONT, bold=True, size=18)),
    )

    # Hint row
    hint_row = aged2_row + 2
    ws.cell(row=hint_row, column=1,
            value="To see which projects are aged: open Project Portfolio, "
                  "filter Target Completion < today and Status not Final/Closed/Lost/Bidding.")
    ws.cell(row=hint_row, column=1).font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)
    ws.merge_cells(start_row=hint_row, start_column=1, end_row=hint_row, end_column=6)
    ws.row_dimensions[hint_row].height = 20

    set_col_widths(ws, {
        "A": 14, "B": 18, "C": 16, "D": 20, "E": 16, "F": 10,
    })


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

OUTPUT_PATH = "/Users/home/charles/contrpro/files/packages/business/Project_Tracker.xlsx"


def build() -> str:
    wb = Workbook()

    # Order: Instructions (active) -> CSI Reference (for named ranges)
    #     -> Company Info -> Project Portfolio -> Summary -> Dashboard.
    # Build CSI Reference first so named ranges resolve before being referenced.
    build_instructions(wb)            # Active sheet renamed to Instructions
    build_csi_reference(wb)           # Defines CSI_Divisions, CSI_Table, States_List
    build_company_info(wb)
    pt_sheet, pt_start, pt_end = build_portfolio(wb)
    build_status_summary(wb, pt_sheet, pt_start, pt_end)
    build_dashboard(wb, pt_sheet, pt_start, pt_end)

    # Re-order tabs
    desired_order = [
        "Instructions",
        "Company Info",
        "Project Portfolio",
        "Portfolio Summary by Status",
        "Dashboard",
        "CSI Reference",
    ]
    wb._sheets = [wb[name] for name in desired_order]

    # Hide CSI Reference
    wb["CSI Reference"].sheet_state = "hidden"

    # Active tab = Dashboard on open
    wb.active = wb.sheetnames.index("Dashboard")

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
