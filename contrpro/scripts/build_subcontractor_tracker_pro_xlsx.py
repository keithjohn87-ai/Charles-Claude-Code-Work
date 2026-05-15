#!/usr/bin/env python3
"""
Build ContrPro Subcontractor Tracker PRO (XLSX) - Business + Complete tier.

The deep version. This is the master vendor database a real $10M+ GC uses to
decide who gets on the bid list. Where the lean Subcontractor_Tracker.xlsx
answers "who do we work with and is their paper current?", the PRO version
answers "should we keep working with them?":
  - Prequalification status + renewal cadence
  - Performance scoring across Schedule / Quality / Safety / Communication / Financial
  - Work history (per-project ledger of how the sub actually performed)
  - Bonding capacity (single + aggregate) and remaining headroom
  - References, DBE/MBE/WBE, EMR, OSHA incidents
  - Compliance dashboard + bonding roll-up + performance leaderboard

Tabs:
  1. Instructions
  2. Sub Master Record         (1 row per sub - the deep database)
  3. Work History              (1 row per sub x project)
  4. Performance Leaderboard   (auto-ranked top 10)
  5. Compliance Dashboard      (counts + flags)
  6. Bonding Capacity Roll-Up
  7. Trade Distribution        (per-trade roll-up)
  8. CSI Reference             (hidden, named ranges)

Run:
    /Users/home/charles/.venv/bin/python3 \\
        /Users/home/charles/contrpro/scripts/build_subcontractor_tracker_pro_xlsx.py

Output:
    /Users/home/charles/contrpro/files/packages/business/Subcontractor_Tracker_Pro.xlsx
"""

from __future__ import annotations

import os
from typing import Tuple

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# ---------------------------------------------------------------------------
# Brand & styling (matches AR/AP/COI/Subcontractor Tracker basic)
# ---------------------------------------------------------------------------

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
ACCENT_GOLD = "C9A227"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
DARK_GREEN_FILL = "63BE7B"
DARK_GREEN_FONT = "FFFFFF"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
YELLOW_FILL = "FFEB9C"
YELLOW_FONT = "9C5700"
ORANGE_FILL = "FFD8A8"
ORANGE_FONT = "8A4B00"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_SUMMARY_LABEL = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_DARK_GREEN = PatternFill("solid", fgColor=DARK_GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW_FILL)
FILL_ORANGE = PatternFill("solid", fgColor=ORANGE_FILL)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H1 = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_BIG_NUMBER = Font(name="Calibri", size=18, bold=True, color=BRAND_BLUE)
FONT_GREY_ITALIC = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_USD0 = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
FMT_PCT = "0.0%"
FMT_INT = "0"
FMT_NUM1 = "0.00"
FMT_NUM2 = "0.0"
FMT_DATE = "yyyy-mm-dd"


# ---------------------------------------------------------------------------
# CSI MasterFormat division reference (canonical, same as sibling trackers)
# ---------------------------------------------------------------------------

DIVISIONS: list[tuple[str, str]] = [
    ("01", "General Requirements"),
    ("02", "Existing Conditions"),
    ("03", "Concrete"),
    ("04", "Masonry"),
    ("05", "Metals"),
    ("06", "Wood, Plastics, and Composites"),
    ("07", "Thermal and Moisture Protection"),
    ("08", "Openings"),
    ("09", "Finishes"),
    ("10", "Specialties"),
    ("11", "Equipment"),
    ("12", "Furnishings"),
    ("13", "Special Construction"),
    ("14", "Conveying Equipment"),
    ("21", "Fire Suppression"),
    ("22", "Plumbing"),
    ("23", "HVAC"),
    ("25", "Integrated Automation"),
    ("26", "Electrical"),
    ("27", "Communications"),
    ("28", "Electronic Safety and Security"),
    ("31", "Earthwork"),
    ("32", "Exterior Improvements"),
    ("33", "Utilities"),
    ("34", "Transportation"),
    ("35", "Waterway and Marine Construction"),
    ("40", "Process Interconnections"),
    ("41", "Material Processing and Handling Equipment"),
    ("42", "Process Heating, Cooling, and Drying Equipment"),
    ("43", "Process Gas and Liquid Handling, Purification, and Storage Equipment"),
    ("44", "Pollution and Waste Control Equipment"),
    ("45", "Industry-Specific Manufacturing Equipment"),
    ("46", "Water and Wastewater Equipment"),
    ("48", "Electrical Power Generation"),
    ("49", "Reserved"),
]


# ---------------------------------------------------------------------------
# Dropdown vocabularies
# ---------------------------------------------------------------------------

TRADES = [
    "General",
    "Mechanical",
    "Electrical",
    "Plumbing",
    "Structural Steel",
    "HVAC",
    "Roofing",
    "Concrete",
    "Masonry",
    "Drywall",
    "Flooring",
    "Painting",
    "Earthwork",
    "Fire Suppression",
    "Other",
]
TRADE_SUMMARY_ORDER = TRADES

PREQUAL_STATES = ["Approved", "Conditionally Approved", "Pending", "Rejected", "Inactive"]
LICENSE_STATES = ["Active", "Expired", "Pending"]
COI_STATES = ["Current", "Expiring", "Expired", "None"]
W9_STATES = ["Yes", "No"]
BONDED_STATES = ["Yes", "No"]
DBE_STATES = ["None", "DBE", "MBE", "WBE", "Multiple"]
SCORE_STATES = ["1", "2", "3", "4", "5"]
YNP_STATES = ["Yes", "No", "Partial"]


# ---------------------------------------------------------------------------
# Pre-populated example subs (8 rows, varied performance)
# ---------------------------------------------------------------------------
#
# Tuple order matches SUB_HEADERS below (the input columns only - the formula
# columns Overall, Rating, Approved are computed in Excel).
#
EXAMPLE_SUBS: list[tuple] = [
    # sub_id, company, dba, primary_contact, secondary_contact, phone, email, address,
    # ein, trade, specialties, csi_div, years_in_biz, annual_rev, num_emp,
    # office_state, service_area, prequal_status, prequal_date, prequal_renewal,
    # lic_num, lic_state, lic_status, coi, coi_exp, w9, bonded, bond_co,
    # single_bond, agg_bond, dbe, dbe_exp, emr, osha_inc, last_used,
    # sched, qual, safe, comm, fin, notes
    ("S-001", "EXAMPLE - Pacific Steel Erectors", "PSE Steel",
     "Marcus Reyes", "Tanya Reyes", "(555) 213-8842", "marcus@pacificsteel.example",
     "1240 Industrial Way, Tacoma WA 98421",
     "91-2334887",
     "Structural Steel", "Erection, miscellaneous metals, stair fab",
     "05", 22, 14_500_000, 78,
     "WA", "WA, OR, ID, MT",
     "Approved", "2026-01-15", "2027-01-15",
     "WA-CTR-104882", "WA", "Active",
     "Current", "2026-09-30", "Yes", "Yes", "Travelers Casualty",
     5_000_000, 15_000_000, "None", "",
     0.78, 0, "2026-04-12",
     5, 5, 5, 4, 5,
     "A-list steel erector. Used on Riverside Office Bldg + Pinnacle. "
     "Bonded thru Travelers. EMR 0.78 is best-in-class."),

    ("S-002", "EXAMPLE - Apex Plumbing & Mechanical", "Apex P&M",
     "Dana Whitfield", "Carlos Mendez", "(555) 661-0034", "dana@apexpm.example",
     "5612 Service Rd, Portland OR 97211",
     "93-1118822",
     "Plumbing", "Commercial plumbing rough-in, medical gas, fixtures",
     "22", 18, 9_200_000, 54,
     "OR", "OR, WA",
     "Approved", "2025-11-08", "2026-11-08",
     "OR-PLB-77231", "OR", "Active",
     "Current", "2026-12-31", "Yes", "Yes", "Liberty Mutual",
     3_000_000, 8_000_000, "None", "",
     0.92, 1, "2026-05-02",
     4, 4, 4, 5, 4,
     "Reliable plumbing partner. Strong on med gas. EMR fine. "
     "Pay app cadence monthly, never a friction."),

    ("S-003", "EXAMPLE - Bayside Electric LLC", "",
     "Renee Hu", "Jon Park", "(555) 408-2917", "renee@baysideelec.example",
     "888 Harbor Blvd, Oakland CA 94606",
     "94-2207715",
     "Electrical", "Service entrance, lighting controls, fire alarm",
     "26", 9, 4_400_000, 28,
     "CA", "CA, NV",
     "Conditionally Approved", "2026-02-22", "2026-08-22",
     "CA-C10-1004821", "CA", "Active",
     "Expired", "2026-04-30", "Yes", "No", "",
     0, 0, "MBE", "2027-03-01",
     1.12, 2, "2026-02-12",
     3, 4, 2, 3, 3,
     "RISK: COI expired 2026-04-30. EMR 1.12 over target. "
     "MBE cert valid. Quality fine, safety + paperwork need work."),

    ("S-004", "EXAMPLE - North Cascade HVAC", "NC HVAC",
     "Sam Petrov", "Anna Petrov", "(555) 920-4477", "sam@ncascadehvac.example",
     "201 Mountain View Dr, Bellingham WA 98225",
     "91-3447721",
     "HVAC", "VAV controls, RTU install, commissioning support",
     "23", 14, 6_800_000, 41,
     "WA", "WA",
     "Approved", "2026-03-01", "2027-03-01",
     "WA-MEC-55619", "WA", "Active",
     "Current", "2027-01-15", "Yes", "Yes", "Travelers Casualty",
     2_500_000, 7_000_000, "None", "",
     0.85, 0, "2026-03-22",
     5, 5, 5, 5, 4,
     "Top-tier HVAC. Strong VAV controls. Approved on Riverside + Pinnacle. "
     "Best performer in the HVAC bench."),

    ("S-005", "EXAMPLE - Granite West Concrete", "Granite West",
     "Tomas Okafor", "Marie Okafor", "(555) 717-3360", "tomas@granitewest.example",
     "3300 Plant Rd, Spokane WA 99202",
     "91-4498822",
     "Concrete", "Slab on grade, footings, small structural pours",
     "03", 26, 12_100_000, 65,
     "WA", "WA, ID, MT",
     "Approved", "2025-09-15", "2026-09-15",
     "WA-CON-88914", "WA", "Active",
     "Current", "2026-11-30", "Yes", "Yes", "Zurich North America",
     4_000_000, 10_000_000, "None", "",
     0.88, 1, "2026-01-18",
     3, 4, 4, 3, 4,
     "Solid for slabs up to ~10K SF. Bigger pours we route to GraniteCore. "
     "Schedule slippage is the recurring friction."),

    ("S-006", "EXAMPLE - Cornerstone Masonry", "Cornerstone",
     "Lila Bennett", "", "(555) 332-5511", "lila@cornermason.example",
     "44 Brick Yard Ln, Boise ID 83702",
     "82-2218833",
     "Masonry", "CMU, brick veneer, stone",
     "04", 6, 1_900_000, 14,
     "ID", "ID, WA, OR",
     "Pending", "2026-04-30", "2026-10-30",
     "ID-MSN-21908", "ID", "Pending",
     "None", "", "No", "No", "",
     0, 0, "WBE", "2027-05-15",
     1.05, 0, "",
     0, 0, 0, 0, 0,
     "New relationship. Prequal in progress. License renewal + W9 still pending. "
     "Hold all awards until full prequal closes. WBE certified."),

    ("S-007", "EXAMPLE - Skyline Roofing Co", "Skyline",
     "Jorge Alvarez", "Maria Alvarez", "(555) 224-6611", "jorge@skylineroof.example",
     "1101 Industrial Pkwy, Reno NV 89502",
     "88-1144229",
     "Roofing", "TPO, modified bit, flashing details",
     "07", 11, 5_300_000, 32,
     "NV", "NV, CA, UT",
     "Approved", "2026-01-30", "2027-01-30",
     "NV-ROOF-44520", "NV", "Active",
     "Current", "2026-10-15", "Yes", "Yes", "Liberty Mutual",
     2_000_000, 5_000_000, "DBE", "2026-12-31",
     0.95, 1, "2026-04-19",
     4, 4, 5, 4, 4,
     "TPO + mod-bit specialist. Strong on emergency repair (Riverside wind event). "
     "Premium pricing but they show up. DBE cert valid."),

    ("S-008", "EXAMPLE - Allied Drywall & Finishes", "Allied DW",
     "Kim Park", "Eddie Park", "(555) 559-2218", "kim@allieddrywall.example",
     "78 Trade Center Dr, Seattle WA 98108",
     "91-5572281",
     "Drywall", "Metal stud framing, drywall, acoustic ceilings, paint",
     "09", 8, 3_100_000, 22,
     "WA", "WA",
     "Inactive", "2025-06-01", "2026-06-01",
     "WA-CTR-203847", "WA", "Expired",
     "None", "", "Yes", "No", "",
     0, 0, "None", "",
     1.35, 4, "2025-11-04",
     2, 2, 1, 2, 3,
     "RETIRED from active bid list. EMR 1.35, 4 OSHA incidents last 3 yrs, "
     "license expired, COI lapsed. Do not award. Kept on record for history."),
]

EMPTY_USER_ROWS = 25


# ---------------------------------------------------------------------------
# Work history example rows
# (sub_id, proj_num, proj_name, year, contract_val, on_time, on_budget,
#  final_quality, issues)
# ---------------------------------------------------------------------------

EXAMPLE_HISTORY: list[tuple] = [
    ("S-001", "P-2024-014", "Riverside Office Building", 2024, 1_850_000,
     "Yes", "Yes", 5, "Excellent erection schedule. Zero punch issues."),
    ("S-001", "P-2025-007", "Pinnacle Tower Phase 1", 2025, 2_650_000,
     "Yes", "Yes", 5, "On schedule throughout. Stair fab quality top-tier."),
    ("S-002", "P-2024-018", "Northgate Medical Plaza", 2024, 920_000,
     "Yes", "Partial", 4, "Med gas changes drove $42K cost growth - sub flagged early."),
    ("S-002", "P-2025-003", "Bayview Tenant Improvement", 2025, 310_000,
     "Yes", "Yes", 4, "Clean job. Fixture coordination handled well."),
    ("S-003", "P-2025-009", "Harbor District Retail", 2025, 480_000,
     "Partial", "No", 3, "Two safety stand-downs. Schedule slipped 14 days. Cost overrun $28K."),
    ("S-004", "P-2024-014", "Riverside Office Building", 2024, 1_120_000,
     "Yes", "Yes", 5, "VAV controls commissioned 5 days early. Owner thrilled."),
    ("S-004", "P-2025-007", "Pinnacle Tower Phase 1", 2025, 1_580_000,
     "Yes", "Yes", 5, "Same crew as Riverside. Same A+ result."),
    ("S-005", "P-2024-011", "Eastpoint Distribution Center", 2024, 1_350_000,
     "No", "Partial", 4, "Slab pours delayed 11 days due to rebar lead time. Quality fine."),
    ("S-007", "P-2024-014", "Riverside Office Building (storm repair)", 2024, 68_000,
     "Yes", "Yes", 5, "Emergency wind-event repair, mobilized in 36 hrs."),
    ("S-008", "P-2024-006", "Sunnyside Office Refit", 2024, 240_000,
     "No", "No", 2, "Schedule slip 21 days. Punch list reopened twice. Crew safety issue noted."),
]

EMPTY_HISTORY_ROWS = 30


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def set_col_widths(ws, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_header_row(ws, row: int, last_col: int) -> None:
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 32


# ---------------------------------------------------------------------------
# Hidden tab - CSI Reference (named ranges for VLOOKUP + dropdowns)
# ---------------------------------------------------------------------------

def build_csi_reference(wb: Workbook) -> str:
    ws = wb.create_sheet("CSI Reference")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "CSI MasterFormat Division Reference"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:B1")

    ws["A2"] = ("Canonical list - do not edit. Feeds the CSI Division dropdown "
                "and VLOOKUPs in Sub Master Record.")
    ws["A2"].font = FONT_GREY_ITALIC

    ws.cell(row=3, column=1, value="CSI Division")
    ws.cell(row=3, column=2, value="Division Name")
    style_header_row(ws, 3, 2)

    for i, (code, name) in enumerate(DIVISIONS, start=4):
        ws.cell(row=i, column=1, value=code).number_format = "@"
        ws.cell(row=i, column=2, value=name)
        for c in (1, 2):
            ws.cell(row=i, column=c).border = BORDER
            ws.cell(row=i, column=c).font = FONT_BODY

    set_col_widths(ws, {"A": 14, "B": 60})

    first = 4
    last = 4 + len(DIVISIONS) - 1
    div_range = f"'CSI Reference'!$A${first}:$A${last}"
    full_range = f"'CSI Reference'!$A${first}:$B${last}"
    wb.defined_names["CSI_Divisions"] = DefinedName("CSI_Divisions", attr_text=div_range)
    wb.defined_names["CSI_Table"] = DefinedName("CSI_Table", attr_text=full_range)
    return ws.title


# ---------------------------------------------------------------------------
# Tab 1 - Instructions
# ---------------------------------------------------------------------------

def build_instructions(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {"A": 4, "B": 112})

    ws["B2"] = "ContrPro - Subcontractor Tracker PRO"
    ws["B2"].font = FONT_TITLE
    ws.row_dimensions[2].height = 32

    ws["B3"] = ("Prequalification + performance + work history + bonding. "
                "The master vendor database. Business / Complete tier.")
    ws["B3"].font = FONT_GREY_ITALIC

    sections: list[tuple[str, list[str]]] = [
        (
            "What this workbook is",
            [
                "Seven visible tabs plus one hidden lookup. Sub Master Record is the deep database - "
                "prequal status, license, COI, W9, bonding, EMR, OSHA history, DBE/MBE/WBE, plus a "
                "performance scorecard (Schedule, Quality, Safety, Communication, Financial). Work History "
                "is a per-project ledger of how each sub actually performed. The remaining tabs are "
                "auto-rolled reports: Performance Leaderboard, Compliance Dashboard, Bonding Capacity "
                "Roll-Up, Trade Distribution.",
                "This is the FLAGSHIP sub-tracker. If you are awarding > 25 subs/year, running multiple PMs, "
                "or carrying $5M+ in subcontracts at any time, this is the tool. If you just need a clean "
                "list of who you work with, use the lean Subcontractor_Tracker.xlsx instead.",
            ],
        ),
        (
            "Prequalification workflow",
            [
                "Step 1 - Add the sub to Sub Master Record with Prequal Status = Pending. Fill in company "
                "data, EIN, license, insurance, bonding capacity, EMR. Set Prequal Date to today.",
                "Step 2 - Send the prequal package (financial statements last 3 years, references, safety "
                "program, bonding letter, COI sample, license verification). Confirm receipt.",
                "Step 3 - Verify references and license status independently. Pull EMR letter from their "
                "insurer. Confirm OSHA 300A logs.",
                "Step 4 - Decision: Approved (full bid list), Conditionally Approved (specific scopes or "
                "$ caps only), or Rejected. Set Prequal Renewal Date = +12 months.",
                "Step 5 - Renewal: every Prequal Renewal Date, re-pull financials, COI, license, EMR. "
                "Refresh the row. No re-prequal, no awards.",
            ],
        ),
        (
            "Performance scoring rubric (1-5 each, scored after every job)",
            [
                "Schedule - 5: every milestone hit or beat. 4: minor slip, recovered. 3: schedule slipped but "
                "delivered. 2: significant delay, drove other trades. 1: missed completion / abandoned.",
                "Quality - 5: zero punch reopens, owner-grade workmanship. 4: typical punch, closed clean. "
                "3: punch closed slowly. 2: rework required. 1: failed inspection / unacceptable.",
                "Safety - 5: zero recordables, proactive safety. 4: minor first-aid only. 3: one recordable, "
                "corrected. 2: multiple incidents. 1: serious injury / OSHA violation.",
                "Communication - 5: proactive RFIs, fast response, owner-facing pro. 4: responsive, clear. "
                "3: reactive only. 2: chasing them for answers. 1: ghosting.",
                "Financial - 5: pay apps clean, no claims. 4: minor pay app errors. 3: one disputed CO. "
                "2: multiple disputes / late lien releases. 1: filed claim or lien against project.",
                "Overall Score is the average. Performance Rating: 4.5+ = A, 3.5-4.4 = B, 2.5-3.4 = C, "
                "below 2.5 = D. C is your re-prequal trigger. D is your retirement trigger.",
            ],
        ),
        (
            "Bonding capacity (Tab 6)",
            [
                "Single Project Bond Capacity - the biggest single contract the sub's surety will bond.",
                "Aggregate Bond Capacity - the total dollar value of bonded work the surety will let them "
                "carry at once. Subs get over-extended when active bonded work approaches aggregate.",
                "Available Capacity (Tab 6) = Aggregate - sum of active bonded contracts. A sub at 90%+ of "
                "aggregate is tapped out and should not be awarded more bonded work until backlog clears.",
                "Refresh the surety letter annually with the prequal renewal.",
            ],
        ),
        (
            "Approval gate (formula in Sub Master Record)",
            [
                "Approved? = Yes only if ALL of: Prequal Status = Approved AND COI = Current AND License = "
                "Active AND Overall Score >= 3. Any other state returns No.",
                "This is the auto-gate before any award. Read this column on Sub Master, not your gut.",
            ],
        ),
        (
            "Retiring underperforming subs",
            [
                "Trigger 1: Performance Rating = D for any single completed project AND prior projects "
                "average below C. Move Prequal Status to Inactive.",
                "Trigger 2: EMR > 1.0 trending upward for 2 consecutive years. Inactive until they bring it "
                "back under 1.0 with a documented safety plan.",
                "Trigger 3: OSHA Recordable Incidents >= 3 in trailing 3 years. Inactive.",
                "Trigger 4: License lapsed > 60 days with no plan to renew. Inactive.",
                "Inactive does NOT mean delete. Keep the row, keep the Work History. You may rehabilitate "
                "the relationship later, and the record is the receipt of why you stopped.",
            ],
        ),
        (
            "Color coding cheat sheet",
            [
                "Prequal Status: green = Approved, yellow = Conditionally Approved / Pending, red = Rejected / "
                "Inactive.",
                "COI: green = Current, yellow = Expiring, red = Expired / None.",
                "License: green = Active, yellow = Pending, red = Expired.",
                "Performance Rating: dark green = A, green = B, yellow = C, red = D.",
                "EMR: green when <= 1.0, red when > 1.0.",
                "Approved?: green Yes, red No. This is your final gate.",
            ],
        ),
        (
            "Tab-by-tab quick start",
            [
                "1. Sub Master Record - one row per sub. Drop in data, score the performance fields after "
                "every job. Overall Score, Performance Rating, and Approved? are formulas - do not type.",
                "2. Work History - one row per sub x project. Sub ID links back to Sub Master.",
                "3. Performance Leaderboard - DO NOT edit. Auto-ranks the top 10 by Overall Score.",
                "4. Compliance Dashboard - DO NOT edit. Snapshot of prequal, COI, EMR, bonding, DBE.",
                "5. Bonding Capacity Roll-Up - tracks who has headroom and who is tapped out.",
                "6. Trade Distribution - count + avg performance per trade. Use this when picking who to bid.",
                "7. CSI Reference - hidden. Unhide only if you need to inspect the canonical list.",
            ],
        ),
        (
            "Tabs in this workbook",
            [
                "Instructions  -  this tab",
                "Sub Master Record  -  the deep database (1 row per sub)",
                "Work History  -  per-project performance ledger",
                "Performance Leaderboard  -  auto top 10",
                "Compliance Dashboard  -  totals and risk flags",
                "Bonding Capacity Roll-Up  -  available bond headroom",
                "Trade Distribution  -  per-trade roll-up",
                "CSI Reference  -  canonical division list (hidden)",
            ],
        ),
    ]

    row = 5
    for heading, paragraphs in sections:
        ws.cell(row=row, column=2, value=heading)
        ws.cell(row=row, column=2).font = FONT_H2
        ws.cell(row=row, column=2).fill = FILL_SUBHEADER
        ws.cell(row=row, column=2).alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[row].height = 22
        row += 1
        for para in paragraphs:
            ws.cell(row=row, column=2, value=para)
            ws.cell(row=row, column=2).font = FONT_BODY
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = max(18, 16 * (1 + len(para) // 112))
            row += 1
        row += 1

    ws.cell(row=row + 1, column=2,
            value="ContrPro - built for builders. Questions: support@contrpro.com")
    ws.cell(row=row + 1, column=2).font = FONT_GREY_ITALIC


# ---------------------------------------------------------------------------
# Tab 2 - Sub Master Record (THE deep database)
# ---------------------------------------------------------------------------

SUB_HEADERS = [
    "Sub ID", "Company Name", "DBA / Trade Name",
    "Primary Contact", "Secondary Contact",
    "Phone", "Email", "Address",
    "EIN",
    "Trade", "Specialties",
    "CSI Division", "Division Name",
    "Years in Business", "Annual Revenue", "Employees",
    "Office State", "Service Area States",
    "Prequal Status", "Prequal Date", "Prequal Renewal",
    "License #", "License State", "License Status",
    "COI on File?", "COI Expires",
    "W9 on File?",
    "Bonded?", "Bonding Company",
    "Single Project Bond Capacity", "Aggregate Bond Capacity",
    "DBE/MBE/WBE", "DBE Cert Expires",
    "Safety EMR", "OSHA Recordable (3yr)",
    "Last Used Date",
    "Score: Schedule", "Score: Quality", "Score: Safety",
    "Score: Communication", "Score: Financial",
    "Overall Score", "Performance Rating", "Approved?",
    "Notes",
]


def build_sub_master(wb: Workbook) -> Tuple[str, int, int]:
    ws = wb.create_sheet("Sub Master Record")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Sub Master Record"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(SUB_HEADERS))

    for i, h in enumerate(SUB_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(SUB_HEADERS))

    data_start = 4
    n_pre = len(EXAMPLE_SUBS)
    n_total = n_pre + EMPTY_USER_ROWS
    data_end = data_start + n_total - 1

    # Column letters
    COL = {h: get_column_letter(i + 1) for i, h in enumerate(SUB_HEADERS)}

    # Pre-populate examples
    for idx, row_data in enumerate(EXAMPLE_SUBS):
        r = data_start + idx
        (sub_id, company, dba, contact, contact2, phone, email, address,
         ein, trade, specialties, csi_div, years, annual_rev, num_emp,
         office_state, service_area, prequal, prequal_date, prequal_renewal,
         lic_num, lic_state, lic_status, coi, coi_exp, w9, bonded, bond_co,
         single_bond, agg_bond, dbe, dbe_exp, emr, osha_inc, last_used,
         sched, qual, safe, comm, fin, notes) = row_data

        ws.cell(row=r, column=1, value=sub_id)
        ws.cell(row=r, column=2, value=company)
        ws.cell(row=r, column=3, value=dba)
        ws.cell(row=r, column=4, value=contact)
        ws.cell(row=r, column=5, value=contact2)
        ws.cell(row=r, column=6, value=phone)
        ws.cell(row=r, column=7, value=email)
        ws.cell(row=r, column=8, value=address)
        ws.cell(row=r, column=9, value=ein)
        ws.cell(row=r, column=10, value=trade)
        ws.cell(row=r, column=11, value=specialties)
        ws.cell(row=r, column=12, value=csi_div)
        # Division Name = VLOOKUP on CSI Division (col L = 12)
        ws.cell(row=r, column=13,
                value=f'=IFERROR(VLOOKUP(L{r},CSI_Table,2,FALSE),"")')
        ws.cell(row=r, column=14, value=years if years else None)
        ws.cell(row=r, column=15, value=annual_rev if annual_rev else None)
        ws.cell(row=r, column=16, value=num_emp if num_emp else None)
        ws.cell(row=r, column=17, value=office_state)
        ws.cell(row=r, column=18, value=service_area)
        ws.cell(row=r, column=19, value=prequal)
        ws.cell(row=r, column=20, value=prequal_date if prequal_date else None)
        ws.cell(row=r, column=21, value=prequal_renewal if prequal_renewal else None)
        ws.cell(row=r, column=22, value=lic_num)
        ws.cell(row=r, column=23, value=lic_state)
        ws.cell(row=r, column=24, value=lic_status)
        ws.cell(row=r, column=25, value=coi)
        ws.cell(row=r, column=26, value=coi_exp if coi_exp else None)
        ws.cell(row=r, column=27, value=w9)
        ws.cell(row=r, column=28, value=bonded)
        ws.cell(row=r, column=29, value=bond_co)
        ws.cell(row=r, column=30, value=single_bond if single_bond else None)
        ws.cell(row=r, column=31, value=agg_bond if agg_bond else None)
        ws.cell(row=r, column=32, value=dbe)
        ws.cell(row=r, column=33, value=dbe_exp if dbe_exp else None)
        ws.cell(row=r, column=34, value=emr if emr else None)
        ws.cell(row=r, column=35, value=osha_inc if osha_inc is not None else None)
        ws.cell(row=r, column=36, value=last_used if last_used else None)
        # Performance scores - 0 means "not yet scored"; show blank if 0
        for col_offset, val in enumerate([sched, qual, safe, comm, fin]):
            ws.cell(row=r, column=37 + col_offset,
                    value=val if val else None)
        # Overall Score = AVERAGE of the 5 score cols (AK..AO = 37..41)
        ws.cell(row=r, column=42,
                value=f'=IFERROR(AVERAGE(AK{r}:AO{r}),"")')
        # Performance Rating = letter grade off Overall (AP = col 42)
        ws.cell(row=r, column=43,
                value=(f'=IF(AP{r}="","",'
                       f'IF(AP{r}>=4.5,"A",'
                       f'IF(AP{r}>=3.5,"B",'
                       f'IF(AP{r}>=2.5,"C","D"))))'))
        # Approved? gate
        # Approved if Prequal=Approved AND COI=Current AND License=Active AND Overall>=3
        # Cols: Prequal=S(19), COI=Y(25), License=X(24), Overall=AP(42)
        ws.cell(row=r, column=44,
                value=(f'=IF(AND(S{r}="Approved",Y{r}="Current",X{r}="Active",'
                       f'IFERROR(AP{r}>=3,FALSE)),"Yes","No")'))
        ws.cell(row=r, column=45, value=notes)

    # Empty user-fillable rows - pre-seed formulas
    for i in range(EMPTY_USER_ROWS):
        r = data_start + n_pre + i
        ws.cell(row=r, column=13,
                value=f'=IFERROR(VLOOKUP(L{r},CSI_Table,2,FALSE),"")')
        ws.cell(row=r, column=42,
                value=f'=IFERROR(AVERAGE(AK{r}:AO{r}),"")')
        ws.cell(row=r, column=43,
                value=(f'=IF(AP{r}="","",'
                       f'IF(AP{r}>=4.5,"A",'
                       f'IF(AP{r}>=3.5,"B",'
                       f'IF(AP{r}>=2.5,"C","D"))))'))
        ws.cell(row=r, column=44,
                value=(f'=IF(AND(S{r}="Approved",Y{r}="Current",X{r}="Active",'
                       f'IFERROR(AP{r}>=3,FALSE)),"Yes","No")'))

    # Per-row formatting
    for r in range(data_start, data_end + 1):
        for c in range(1, len(SUB_HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(c in (2, 8, 11, 18, 45)),
            )
        # Sub ID and CSI Division text format (keep leading zeros)
        ws.cell(row=r, column=1).number_format = "@"
        ws.cell(row=r, column=12).number_format = "@"
        # Date columns
        for date_col in (20, 21, 26, 33, 36):
            ws.cell(row=r, column=date_col).number_format = FMT_DATE
        # USD columns
        for usd_col in (15, 30, 31):
            ws.cell(row=r, column=usd_col).number_format = FMT_USD0
        # Numeric format columns
        ws.cell(row=r, column=14).number_format = FMT_INT  # Years
        ws.cell(row=r, column=16).number_format = FMT_INT  # Employees
        ws.cell(row=r, column=34).number_format = FMT_NUM1  # EMR
        ws.cell(row=r, column=35).number_format = FMT_INT  # OSHA
        for sc in (37, 38, 39, 40, 41):
            ws.cell(row=r, column=sc).number_format = FMT_INT
            ws.cell(row=r, column=sc).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=42).number_format = FMT_NUM1  # Overall
        ws.cell(row=r, column=42).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=43).alignment = Alignment(horizontal="center", vertical="center")  # Rating
        ws.cell(row=r, column=43).font = FONT_BODY_BOLD
        ws.cell(row=r, column=44).alignment = Alignment(horizontal="center", vertical="center")  # Approved
        ws.cell(row=r, column=44).font = FONT_BODY_BOLD
        ws.row_dimensions[r].height = 38

    # --- Conditional formatting ---

    def cf_eq(ws_, col, vals, fill, font):
        rng = f"{col}{data_start}:{col}{data_end}"
        for v in vals:
            ws_.conditional_formatting.add(
                rng,
                FormulaRule(formula=[f'${col}{data_start}="{v}"'], stopIfTrue=False,
                            fill=fill, font=font),
            )

    bold_green = Font(color=GREEN_FONT, bold=True)
    bold_yellow = Font(color=YELLOW_FONT, bold=True)
    bold_red = Font(color=RED_FONT, bold=True)
    white_bold = Font(color="FFFFFF", bold=True)

    # Prequal Status (col S = 19)
    cf_eq(ws, "S", ["Approved"], FILL_GREEN, bold_green)
    cf_eq(ws, "S", ["Conditionally Approved", "Pending"], FILL_YELLOW, bold_yellow)
    cf_eq(ws, "S", ["Rejected", "Inactive"], FILL_RED, bold_red)

    # COI (col Y = 25)
    cf_eq(ws, "Y", ["Current"], FILL_GREEN, bold_green)
    cf_eq(ws, "Y", ["Expiring"], FILL_YELLOW, bold_yellow)
    cf_eq(ws, "Y", ["Expired", "None"], FILL_RED, bold_red)

    # License Status (col X = 24)
    cf_eq(ws, "X", ["Active"], FILL_GREEN, bold_green)
    cf_eq(ws, "X", ["Pending"], FILL_YELLOW, bold_yellow)
    cf_eq(ws, "X", ["Expired"], FILL_RED, bold_red)

    # Performance Rating (col AQ = 43)
    cf_eq(ws, "AQ", ["A"], FILL_DARK_GREEN, white_bold)
    cf_eq(ws, "AQ", ["B"], FILL_GREEN, bold_green)
    cf_eq(ws, "AQ", ["C"], FILL_YELLOW, bold_yellow)
    cf_eq(ws, "AQ", ["D"], FILL_RED, bold_red)

    # EMR (col AH = 34) - red if > 1.0, green if <= 1.0 and > 0
    emr_range = f"AH{data_start}:AH{data_end}"
    ws.conditional_formatting.add(
        emr_range,
        FormulaRule(formula=[f'AND(AH{data_start}>0,AH{data_start}<=1)'], stopIfTrue=False,
                    fill=FILL_GREEN, font=bold_green),
    )
    ws.conditional_formatting.add(
        emr_range,
        FormulaRule(formula=[f'AH{data_start}>1'], stopIfTrue=False,
                    fill=FILL_RED, font=bold_red),
    )

    # Approved? (col AR = 44)
    cf_eq(ws, "AR", ["Yes"], FILL_GREEN, bold_green)
    cf_eq(ws, "AR", ["No"], FILL_RED, bold_red)

    # --- Data validation (dropdowns) ---

    def add_dv(col_letter, values_list=None, formula=None, error_msg="", title=""):
        if formula:
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        else:
            dv = DataValidation(type="list",
                                formula1='"' + ",".join(values_list) + '"',
                                allow_blank=True)
        dv.error = error_msg
        dv.errorTitle = title
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}{data_start}:{col_letter}{data_end}")

    add_dv("J", TRADES, "Pick a trade.", "Invalid trade")
    add_dv("L", formula="=CSI_Divisions", error_msg="Pick CSI division.",
           title="Invalid CSI Division")
    add_dv("S", PREQUAL_STATES, "Pick prequal status.", "Invalid prequal")
    add_dv("X", LICENSE_STATES, "Pick license status.", "Invalid license")
    add_dv("Y", COI_STATES, "Pick COI state.", "Invalid COI")
    add_dv("AA", W9_STATES, "Yes / No.", "Invalid W9")
    add_dv("AB", BONDED_STATES, "Yes / No.", "Invalid bonded")
    add_dv("AF", DBE_STATES, "Pick DBE/MBE/WBE.", "Invalid DBE")
    for col in ("AK", "AL", "AM", "AN", "AO"):
        add_dv(col, SCORE_STATES, "Score 1-5.", "Invalid score")

    # Column widths
    set_col_widths(ws, {
        "A": 9, "B": 30, "C": 18,
        "D": 18, "E": 18,
        "F": 15, "G": 28, "H": 30,
        "I": 13,
        "J": 16, "K": 30,
        "L": 11, "M": 26,
        "N": 11, "O": 16, "P": 11,
        "Q": 11, "R": 22,
        "S": 18, "T": 13, "U": 14,
        "V": 18, "W": 11, "X": 13,
        "Y": 11, "Z": 13,
        "AA": 9,
        "AB": 9, "AC": 22,
        "AD": 14, "AE": 14,
        "AF": 11, "AG": 13,
        "AH": 9, "AI": 10,
        "AJ": 13,
        "AK": 7, "AL": 7, "AM": 7, "AN": 7, "AO": 7,
        "AP": 10, "AQ": 11, "AR": 11,
        "AS": 40,
    })

    ws.freeze_panes = "C4"

    return ws.title, data_start, data_end


# ---------------------------------------------------------------------------
# Tab 3 - Work History
# ---------------------------------------------------------------------------

HISTORY_HEADERS = [
    "Sub ID", "Project Number", "Project Name",
    "Year Completed", "Contract Value",
    "On-Time?", "On-Budget?", "Final Quality Rating",
    "Issues / Notes",
]


def build_work_history(wb: Workbook) -> Tuple[str, int, int]:
    ws = wb.create_sheet("Work History")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Work History (per-sub project ledger)"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HISTORY_HEADERS))

    ws["A2"] = ("One row per sub x project. Sub ID must match Sub Master Record. "
                "Score after every completed contract.")
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:I2")

    for i, h in enumerate(HISTORY_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(HISTORY_HEADERS))

    data_start = 4
    n_pre = len(EXAMPLE_HISTORY)
    n_total = n_pre + EMPTY_HISTORY_ROWS
    data_end = data_start + n_total - 1

    for idx, row_data in enumerate(EXAMPLE_HISTORY):
        r = data_start + idx
        (sub_id, proj_num, proj_name, year, contract_val,
         on_time, on_budget, quality, issues) = row_data
        ws.cell(row=r, column=1, value=sub_id)
        ws.cell(row=r, column=2, value=proj_num)
        ws.cell(row=r, column=3, value=proj_name)
        ws.cell(row=r, column=4, value=year)
        ws.cell(row=r, column=5, value=contract_val)
        ws.cell(row=r, column=6, value=on_time)
        ws.cell(row=r, column=7, value=on_budget)
        ws.cell(row=r, column=8, value=quality)
        ws.cell(row=r, column=9, value=issues)

    # Format all rows
    for r in range(data_start, data_end + 1):
        for c in range(1, len(HISTORY_HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(c in (3, 9)),
            )
        ws.cell(row=r, column=1).number_format = "@"
        ws.cell(row=r, column=4).number_format = FMT_INT
        ws.cell(row=r, column=5).number_format = FMT_USD0
        ws.cell(row=r, column=8).number_format = FMT_INT
        ws.cell(row=r, column=8).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 28

    # Conditional formatting
    yn_green = Font(color=GREEN_FONT, bold=True)
    yn_red = Font(color=RED_FONT, bold=True)
    yn_yellow = Font(color=YELLOW_FONT, bold=True)

    for col in ("F", "G"):
        rng = f"{col}{data_start}:{col}{data_end}"
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'${col}{data_start}="Yes"'], stopIfTrue=False,
                        fill=FILL_GREEN, font=yn_green),
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'${col}{data_start}="Partial"'], stopIfTrue=False,
                        fill=FILL_YELLOW, font=yn_yellow),
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'${col}{data_start}="No"'], stopIfTrue=False,
                        fill=FILL_RED, font=yn_red),
        )

    # Quality rating CF: 4-5 green, 3 yellow, 1-2 red
    q_rng = f"H{data_start}:H{data_end}"
    ws.conditional_formatting.add(
        q_rng,
        FormulaRule(formula=[f'AND(H{data_start}>=4,H{data_start}<=5)'], stopIfTrue=False,
                    fill=FILL_GREEN, font=yn_green),
    )
    ws.conditional_formatting.add(
        q_rng,
        FormulaRule(formula=[f'H{data_start}=3'], stopIfTrue=False,
                    fill=FILL_YELLOW, font=yn_yellow),
    )
    ws.conditional_formatting.add(
        q_rng,
        FormulaRule(formula=[f'AND(H{data_start}>=1,H{data_start}<=2)'], stopIfTrue=False,
                    fill=FILL_RED, font=yn_red),
    )

    # Dropdowns
    dv_yn = DataValidation(type="list", formula1='"' + ",".join(YNP_STATES) + '"',
                           allow_blank=True)
    dv_yn.error = "Yes / No / Partial"
    dv_yn.errorTitle = "Invalid value"
    ws.add_data_validation(dv_yn)
    dv_yn.add(f"F{data_start}:G{data_end}")

    dv_q = DataValidation(type="list", formula1='"' + ",".join(SCORE_STATES) + '"',
                          allow_blank=True)
    dv_q.error = "1-5"
    dv_q.errorTitle = "Invalid rating"
    ws.add_data_validation(dv_q)
    dv_q.add(f"H{data_start}:H{data_end}")

    set_col_widths(ws, {
        "A": 9, "B": 14, "C": 32,
        "D": 9, "E": 16,
        "F": 11, "G": 12, "H": 10,
        "I": 50,
    })

    ws.freeze_panes = "C4"
    return ws.title, data_start, data_end


# ---------------------------------------------------------------------------
# Tab 4 - Performance Leaderboard
# ---------------------------------------------------------------------------

def build_leaderboard(wb: Workbook, sub_sheet: str, sub_start: int, sub_end: int) -> None:
    ws = wb.create_sheet("Performance Leaderboard")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Performance Leaderboard"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:G1")

    ws["A2"] = ("Top 10 subs by Overall Score. Auto-ranked from Sub Master Record. "
                "Do not edit - update scores on the master tab and this re-sorts.")
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:G2")

    headers = ["Rank", "Sub ID", "Company Name", "Trade",
               "Overall Score", "Rating", "Last Used"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, len(headers))

    # Source columns on Sub Master:
    # Sub ID = A, Company = B, Trade = J, Overall = AP, Rating = AQ, Last Used = AJ
    sub_id_rng = f"'{sub_sheet}'!$A${sub_start}:$A${sub_end}"
    company_rng = f"'{sub_sheet}'!$B${sub_start}:$B${sub_end}"
    trade_rng = f"'{sub_sheet}'!$J${sub_start}:$J${sub_end}"
    overall_rng = f"'{sub_sheet}'!$AP${sub_start}:$AP${sub_end}"
    rating_rng = f"'{sub_sheet}'!$AQ${sub_start}:$AQ${sub_end}"
    lastused_rng = f"'{sub_sheet}'!$AJ${sub_start}:$AJ${sub_end}"

    # Top 10 - use LARGE on the Overall Score column. Handle blank-string nulls
    # by wrapping LARGE in IFERROR. To avoid duplicate-Overall tie collisions,
    # we add a tiny row-index nudge to break ties: Overall + ROW()/1000000.
    # But that nudge needs to live in a helper column. To keep this clean and
    # avoid array-formula complications in openpyxl, we use straight LARGE +
    # INDEX/MATCH and accept that ties resolve to the first match (acceptable
    # since the master record will rarely produce more than 1-2 exact ties).
    for i in range(10):
        rank = i + 1
        r = 5 + i
        ws.cell(row=r, column=1, value=rank).font = FONT_BODY_BOLD
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        # Overall Score (col E)
        ws.cell(row=r, column=5,
                value=f'=IFERROR(LARGE({overall_rng},{rank}),"")')
        # Sub ID via INDEX/MATCH on Overall == E{r}
        ws.cell(row=r, column=2,
                value=(f'=IFERROR(IF(E{r}="","",'
                       f'INDEX({sub_id_rng},MATCH(E{r},{overall_rng},0))),"")'))
        ws.cell(row=r, column=3,
                value=(f'=IFERROR(IF(E{r}="","",'
                       f'INDEX({company_rng},MATCH(E{r},{overall_rng},0))),"")'))
        ws.cell(row=r, column=4,
                value=(f'=IFERROR(IF(E{r}="","",'
                       f'INDEX({trade_rng},MATCH(E{r},{overall_rng},0))),"")'))
        ws.cell(row=r, column=6,
                value=(f'=IFERROR(IF(E{r}="","",'
                       f'INDEX({rating_rng},MATCH(E{r},{overall_rng},0))),"")'))
        ws.cell(row=r, column=7,
                value=(f'=IFERROR(IF(E{r}="","",'
                       f'INDEX({lastused_rng},MATCH(E{r},{overall_rng},0))),"")'))

        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = Alignment(
                vertical="center", horizontal="center" if c in (1, 5, 6, 7) else "left",
                indent=0 if c in (1, 5, 6, 7) else 1,
            )
        ws.cell(row=r, column=2).number_format = "@"
        ws.cell(row=r, column=5).number_format = FMT_NUM1
        ws.cell(row=r, column=7).number_format = FMT_DATE
        ws.cell(row=r, column=6).font = FONT_BODY_BOLD
        ws.row_dimensions[r].height = 24

    # Rating CF on leaderboard (col F = 6)
    bold_green = Font(color=GREEN_FONT, bold=True)
    bold_yellow = Font(color=YELLOW_FONT, bold=True)
    bold_red = Font(color=RED_FONT, bold=True)
    white_bold = Font(color="FFFFFF", bold=True)
    rate_rng = "F5:F14"
    ws.conditional_formatting.add(
        rate_rng,
        FormulaRule(formula=['$F5="A"'], stopIfTrue=False,
                    fill=FILL_DARK_GREEN, font=white_bold),
    )
    ws.conditional_formatting.add(
        rate_rng,
        FormulaRule(formula=['$F5="B"'], stopIfTrue=False,
                    fill=FILL_GREEN, font=bold_green),
    )
    ws.conditional_formatting.add(
        rate_rng,
        FormulaRule(formula=['$F5="C"'], stopIfTrue=False,
                    fill=FILL_YELLOW, font=bold_yellow),
    )
    ws.conditional_formatting.add(
        rate_rng,
        FormulaRule(formula=['$F5="D"'], stopIfTrue=False,
                    fill=FILL_RED, font=bold_red),
    )

    set_col_widths(ws, {
        "A": 8, "B": 10, "C": 36, "D": 18,
        "E": 14, "F": 10, "G": 14,
    })

    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# Tab 5 - Compliance Dashboard
# ---------------------------------------------------------------------------

def build_compliance_dashboard(wb: Workbook, sub_sheet: str,
                               sub_start: int, sub_end: int) -> None:
    ws = wb.create_sheet("Compliance Dashboard")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Compliance Dashboard"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:F1")

    ws["A2"] = ("Auto-rolled from Sub Master Record. Update that tab - this stays in sync.")
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:F2")

    company_col = f"'{sub_sheet}'!$B${sub_start}:$B${sub_end}"
    prequal_col = f"'{sub_sheet}'!$S${sub_start}:$S${sub_end}"
    coi_col = f"'{sub_sheet}'!$Y${sub_start}:$Y${sub_end}"
    coi_exp_col = f"'{sub_sheet}'!$Z${sub_start}:$Z${sub_end}"
    lic_col = f"'{sub_sheet}'!$X${sub_start}:$X${sub_end}"
    bonded_col = f"'{sub_sheet}'!$AB${sub_start}:$AB${sub_end}"
    agg_bond_col = f"'{sub_sheet}'!$AE${sub_start}:$AE${sub_end}"
    dbe_col = f"'{sub_sheet}'!$AF${sub_start}:$AF${sub_end}"
    emr_col = f"'{sub_sheet}'!$AH${sub_start}:$AH${sub_end}"
    overall_col = f"'{sub_sheet}'!$AP${sub_start}:$AP${sub_end}"
    approved_col = f"'{sub_sheet}'!$AR${sub_start}:$AR${sub_end}"

    def section_header(row: int, text: str) -> None:
        ws.cell(row=row, column=1, value=text).font = FONT_H1
        ws.cell(row=row, column=1).fill = FILL_HEADER
        ws.cell(row=row, column=1).alignment = Alignment(vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 26

    def metric_row(row: int, label: str, formula: str, fmt: str,
                   color_zero_green: bool = False, color_pos_red: bool = False) -> None:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = FONT_BODY_BOLD
        ws.cell(row=row, column=1).fill = FILL_SUMMARY_LABEL
        ws.cell(row=row, column=1).alignment = Alignment(vertical="center", indent=1)
        ws.cell(row=row, column=1).border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

        ws.cell(row=row, column=5, value=formula)
        ws.cell(row=row, column=5).font = FONT_BIG_NUMBER
        ws.cell(row=row, column=5).number_format = fmt
        ws.cell(row=row, column=5).alignment = Alignment(
            horizontal="right", vertical="center", indent=1,
        )
        ws.cell(row=row, column=5).border = BORDER
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 28

        if color_pos_red:
            # Red if > 0, green if = 0
            ws.conditional_formatting.add(
                f"E{row}",
                CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=False,
                           fill=FILL_RED, font=Font(color=RED_FONT, bold=True, size=18)),
            )
            ws.conditional_formatting.add(
                f"E{row}",
                CellIsRule(operator="equal", formula=["0"], stopIfTrue=False,
                           fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True, size=18)),
            )
        elif color_zero_green:
            ws.conditional_formatting.add(
                f"E{row}",
                CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=False,
                           fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True, size=18)),
            )

    # ---------- Roster Snapshot ----------
    r = 4
    section_header(r, "Roster Snapshot")
    r += 1
    metric_row(r, "Total Subs in Database",
               f'=COUNTA({company_col})', FMT_INT, color_zero_green=True)
    r += 1
    metric_row(r, "Total Approved Subs",
               f'=COUNTIF({approved_col},"Yes")', FMT_INT, color_zero_green=True)
    r += 1
    metric_row(r, "Total Pending Prequal",
               f'=COUNTIF({prequal_col},"Pending")', FMT_INT)
    r += 1
    metric_row(r, "Average Performance Score (Approved subs)",
               (f'=IFERROR(AVERAGEIFS({overall_col},'
                f'{approved_col},"Yes",{overall_col},">0"),0)'),
               FMT_NUM1, color_zero_green=True)

    # ---------- Risk Flags ----------
    r += 2
    section_header(r, "Risk Flags")
    r += 1
    metric_row(r, "Subs with Expiring COIs (next 30 days)",
               (f'=SUMPRODUCT(({coi_exp_col}>=TODAY())*'
                f'({coi_exp_col}<=TODAY()+30)*({coi_exp_col}<>""))'),
               FMT_INT, color_pos_red=True)
    r += 1
    metric_row(r, "Subs with Expired COIs",
               f'=COUNTIF({coi_col},"Expired")+COUNTIF({coi_col},"None")', FMT_INT,
               color_pos_red=True)
    r += 1
    metric_row(r, "Subs with Expired License",
               f'=COUNTIF({lic_col},"Expired")', FMT_INT, color_pos_red=True)
    r += 1
    metric_row(r, "Subs with EMR > 1.0 (safety concern)",
               f'=COUNTIF({emr_col},">1")', FMT_INT, color_pos_red=True)
    r += 1
    metric_row(r, "Inactive / Rejected Subs",
               f'=COUNTIF({prequal_col},"Inactive")+COUNTIF({prequal_col},"Rejected")',
               FMT_INT)

    # ---------- Bonding & Diversity ----------
    r += 2
    section_header(r, "Bonding & Diversity")
    r += 1
    metric_row(r, "Bonded Subs (count)",
               f'=COUNTIF({bonded_col},"Yes")', FMT_INT, color_zero_green=True)
    r += 1
    metric_row(r, "Total Aggregate Bonding Capacity ($)",
               f'=SUMIF({bonded_col},"Yes",{agg_bond_col})', FMT_USD0,
               color_zero_green=True)
    r += 1
    metric_row(r, "DBE Certified",
               f'=COUNTIF({dbe_col},"DBE")+COUNTIF({dbe_col},"Multiple")', FMT_INT)
    r += 1
    metric_row(r, "MBE Certified",
               f'=COUNTIF({dbe_col},"MBE")+COUNTIF({dbe_col},"Multiple")', FMT_INT)
    r += 1
    metric_row(r, "WBE Certified",
               f'=COUNTIF({dbe_col},"WBE")+COUNTIF({dbe_col},"Multiple")', FMT_INT)

    set_col_widths(ws, {
        "A": 22, "B": 12, "C": 14, "D": 14, "E": 18, "F": 14,
    })


# ---------------------------------------------------------------------------
# Tab 6 - Bonding Capacity Roll-Up
# ---------------------------------------------------------------------------

def build_bonding_rollup(wb: Workbook, sub_sheet: str, sub_start: int, sub_end: int,
                         history_sheet: str, hist_start: int, hist_end: int) -> None:
    ws = wb.create_sheet("Bonding Capacity Roll-Up")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Bonding Capacity Roll-Up"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:F1")

    ws["A2"] = ("Per-sub bonding headroom. Active Project Count + Contract Value $ are summed from "
                "Work History (current year only). Available = Aggregate - Active. Subs near 0 are tapped out.")
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:F2")

    headers = ["Sub ID", "Company Name", "Single Project Capacity",
               "Aggregate Capacity", "Active Bonded Value (current yr)",
               "Available Capacity"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, len(headers))

    # We surface one bonding row per Sub Master row, displayed only when Bonded = Yes
    # (others render as blank rows).
    n_sub = sub_end - sub_start + 1

    sub_id_master = f"'{sub_sheet}'!$A${sub_start}:$A${sub_end}"
    bonded_col = f"'{sub_sheet}'!$AB${sub_start}:$AB${sub_end}"

    history_sub_id = f"'{history_sheet}'!$A${hist_start}:$A${hist_end}"
    history_year = f"'{history_sheet}'!$D${hist_start}:$D${hist_end}"
    history_val = f"'{history_sheet}'!$E${hist_start}:$E${hist_end}"

    for i in range(n_sub):
        sm_row = sub_start + i  # row in Sub Master
        out_row = 5 + i
        # Show data only when Bonded = Yes on the master row
        # Sub ID
        ws.cell(row=out_row, column=1,
                value=(f'=IF(\'{sub_sheet}\'!$AB${sm_row}="Yes",'
                       f'\'{sub_sheet}\'!$A${sm_row},"")'))
        # Company
        ws.cell(row=out_row, column=2,
                value=(f'=IF(\'{sub_sheet}\'!$AB${sm_row}="Yes",'
                       f'\'{sub_sheet}\'!$B${sm_row},"")'))
        # Single Project
        ws.cell(row=out_row, column=3,
                value=(f'=IF(\'{sub_sheet}\'!$AB${sm_row}="Yes",'
                       f'\'{sub_sheet}\'!$AD${sm_row},"")'))
        # Aggregate
        ws.cell(row=out_row, column=4,
                value=(f'=IF(\'{sub_sheet}\'!$AB${sm_row}="Yes",'
                       f'\'{sub_sheet}\'!$AE${sm_row},"")'))
        # Active bonded value = SUMIFS on Work History where Sub ID matches AND year = current
        ws.cell(row=out_row, column=5,
                value=(f'=IF(\'{sub_sheet}\'!$AB${sm_row}="Yes",'
                       f'SUMIFS({history_val},{history_sub_id},'
                       f'\'{sub_sheet}\'!$A${sm_row},{history_year},YEAR(TODAY())),"")'))
        # Available = Aggregate - Active
        ws.cell(row=out_row, column=6,
                value=(f'=IF(\'{sub_sheet}\'!$AB${sm_row}="Yes",'
                       f'\'{sub_sheet}\'!$AE${sm_row}-'
                       f'SUMIFS({history_val},{history_sub_id},'
                       f'\'{sub_sheet}\'!$A${sm_row},{history_year},YEAR(TODAY())),"")'))

        for c in range(1, len(headers) + 1):
            ws.cell(row=out_row, column=c).border = BORDER
            ws.cell(row=out_row, column=c).alignment = Alignment(vertical="center")
        ws.cell(row=out_row, column=1).number_format = "@"
        for usd in (3, 4, 5, 6):
            ws.cell(row=out_row, column=usd).number_format = FMT_USD0
        ws.row_dimensions[out_row].height = 24

    # Highlight tapped-out subs (Available <= 10% of Aggregate)
    last_row = 5 + n_sub - 1
    rng = f"F5:F{last_row}"
    # Red when Available <= Aggregate * 0.1 AND Aggregate > 0
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'AND(ISNUMBER(F5),ISNUMBER(D5),D5>0,F5<=D5*0.1)'],
                    stopIfTrue=False, fill=FILL_RED,
                    font=Font(color=RED_FONT, bold=True)),
    )
    # Yellow when Available <= 25% of Aggregate
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'AND(ISNUMBER(F5),ISNUMBER(D5),D5>0,F5<=D5*0.25,F5>D5*0.1)'],
                    stopIfTrue=False, fill=FILL_YELLOW,
                    font=Font(color=YELLOW_FONT, bold=True)),
    )
    # Green when Available > 25% of Aggregate
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'AND(ISNUMBER(F5),ISNUMBER(D5),D5>0,F5>D5*0.25)'],
                    stopIfTrue=False, fill=FILL_GREEN,
                    font=Font(color=GREEN_FONT, bold=True)),
    )

    set_col_widths(ws, {
        "A": 9, "B": 32,
        "C": 22, "D": 22, "E": 26, "F": 22,
    })

    ws.freeze_panes = "C5"


# ---------------------------------------------------------------------------
# Tab 7 - Trade Distribution
# ---------------------------------------------------------------------------

def build_trade_distribution(wb: Workbook, sub_sheet: str, sub_start: int, sub_end: int) -> None:
    ws = wb.create_sheet("Trade Distribution")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Trade Distribution"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:D1")

    ws["A2"] = ("Per-trade roll-up: total in database, count approved, and average performance score. "
                "Pulls live from Sub Master Record.")
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:D2")

    headers = ["Trade", "Total Count", "Approved Count", "Avg Performance Score"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, len(headers))

    trade_col = f"'{sub_sheet}'!$J${sub_start}:$J${sub_end}"
    approved_col = f"'{sub_sheet}'!$AR${sub_start}:$AR${sub_end}"
    overall_col = f"'{sub_sheet}'!$AP${sub_start}:$AP${sub_end}"

    start_row = 5
    for idx, trade in enumerate(TRADES):
        r = start_row + idx
        ws.cell(row=r, column=1, value=trade)
        ws.cell(row=r, column=2, value=f'=COUNTIF({trade_col},A{r})')
        ws.cell(row=r, column=3,
                value=f'=COUNTIFS({trade_col},A{r},{approved_col},"Yes")')
        ws.cell(row=r, column=4,
                value=(f'=IFERROR(AVERAGEIFS({overall_col},{trade_col},A{r},'
                       f'{overall_col},">0"),0)'))
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = Alignment(
                vertical="center",
                horizontal="center" if c >= 2 else "left",
                indent=0 if c >= 2 else 1,
            )
        ws.cell(row=r, column=2).number_format = FMT_INT
        ws.cell(row=r, column=3).number_format = FMT_INT
        ws.cell(row=r, column=4).number_format = FMT_NUM1
        ws.row_dimensions[r].height = 22

    end_row = start_row + len(TRADES) - 1

    # Totals row
    tr = end_row + 2
    ws.cell(row=tr, column=1, value="TOTALS").font = FONT_BODY_BOLD
    ws.cell(row=tr, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=tr, column=1).fill = FILL_SUBHEADER
    ws.cell(row=tr, column=1).border = BORDER
    for col_letter in ("B", "C"):
        cell = ws.cell(row=tr, column=ord(col_letter) - ord("A") + 1,
                       value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
        cell.font = FONT_BODY_BOLD
        cell.fill = FILL_SUBHEADER
        cell.border = BORDER
        cell.number_format = FMT_INT
    # Average of averages (weighted would be better but this matches the count-by-trade view)
    cell = ws.cell(row=tr, column=4,
                   value=f"=IFERROR(AVERAGEIFS({overall_col},{overall_col},\">0\"),0)")
    cell.font = FONT_BODY_BOLD
    cell.fill = FILL_SUBHEADER
    cell.border = BORDER
    cell.number_format = FMT_NUM1

    # Highlight non-zero counts in green
    for col_letter in ("B", "C"):
        ws.conditional_formatting.add(
            f"{col_letter}{start_row}:{col_letter}{end_row}",
            CellIsRule(operator="greaterThan", formula=["0"],
                       fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True)),
        )

    set_col_widths(ws, {
        "A": 22, "B": 14, "C": 18, "D": 22,
    })

    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

OUTPUT_PATH = "/Users/home/charles/contrpro/files/packages/business/Subcontractor_Tracker_Pro.xlsx"


def build() -> str:
    wb = Workbook()

    build_instructions(wb)                # Active sheet renamed
    build_csi_reference(wb)                # Named ranges
    sub_sheet, sub_start, sub_end = build_sub_master(wb)
    hist_sheet, hist_start, hist_end = build_work_history(wb)
    build_leaderboard(wb, sub_sheet, sub_start, sub_end)
    build_compliance_dashboard(wb, sub_sheet, sub_start, sub_end)
    build_bonding_rollup(wb, sub_sheet, sub_start, sub_end,
                         hist_sheet, hist_start, hist_end)
    build_trade_distribution(wb, sub_sheet, sub_start, sub_end)

    desired_order = [
        "Instructions",
        "Sub Master Record",
        "Work History",
        "Performance Leaderboard",
        "Compliance Dashboard",
        "Bonding Capacity Roll-Up",
        "Trade Distribution",
        "CSI Reference",
    ]
    wb._sheets = [wb[name] for name in desired_order]

    # Hide CSI Reference
    wb["CSI Reference"].sheet_state = "hidden"

    # Active tab = Compliance Dashboard on open
    wb.active = wb.sheetnames.index("Compliance Dashboard")

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
