#!/usr/bin/env python3
"""Ground Resistance Test Log (XLSX) — Electrical trade pack."""
from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BRAND_BLUE = "1E3A5F"
FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=BRAND_BLUE)
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

OUT = "/Users/home/charles/contrpro/files/packages/complete/electrical/Ground_Resistance_Test_Log.xlsx"


def widths(ws, cols):
    for col, w in cols: ws.column_dimensions[col].width = w


def title(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text); c.font = FONT_TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def header(ws, row, cols):
    for i, h in enumerate(cols):
        c = ws.cell(row=row, column=1 + i, value=h)
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[row].height = 28


def blanks(ws, start, n_rows, n_cols, fmts=None):
    fmts = fmts or {}
    for r in range(start, start + n_rows):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c, value="")
            cell.border = BORDER; cell.font = FONT_BODY
            cell.alignment = CENTER if c > 1 else LEFT
            if c in fmts: cell.number_format = fmts[c]


def build_instructions(ws):
    ws.title = "Instructions"
    widths(ws, [("A", 110)])
    ws["A1"] = "GROUND RESISTANCE TEST LOG — INSTRUCTIONS"
    ws["A1"].font = FONT_TITLE
    body = [
        "",
        "PURPOSE",
        "Captures every ground resistance measurement — 3-point fall-of-potential (acceptance),",
        "clamp-on (verification), Wenner 4-point soil resistivity (design). Submission-ready for",
        "AHJ + commissioning.",
        "",
        "TYPICAL THRESHOLDS",
        "  • NEC 250.53(A)(2) — single rod electrode: ≤25 Ω or add second electrode.",
        "  • Hospital + data center spec: ≤5 Ω, sometimes ≤3 Ω.",
        "  • Telecom + sensitive electronics: ≤1 Ω.",
        "",
        "TEST METHODS",
        "  3-point fall-of-potential: drive aux rods; measure resistance at 62% spacing.",
        "  Clamp-on: non-invasive; valid where multiple ground paths form a loop.",
        "  4-point Wenner: soil resistivity for design.",
        "",
        "DOCUMENT VERSION",
        "Ground_Resistance_Test_Log.xlsx v1.0 — 2026-05-19",
        "ContrPro Complete Tier · Electrical Trade Pack",
    ]
    for i, line in enumerate(body, start=3):
        c = ws.cell(row=i, column=1, value=line)
        c.font = FONT_H2 if line.isupper() and line.strip() and not line.startswith(" ") else FONT_BODY
        c.alignment = LEFT


def build_electrodes(ws):
    ws.title = "Electrode Inventory"
    widths(ws, [("A", 8), ("B", 22), ("C", 14), ("D", 14), ("E", 16), ("F", 22)])
    title(ws, 1, "GROUNDING-ELECTRODE INVENTORY", 6)
    header(ws, 3, ["#", "Type (Rod/Ufer/Ring/Plate/Water)", "Location", "Size/Length",
                   "Install Date", "Notes"])
    blanks(ws, 4, 30, 6, {5: "yyyy-mm-dd"})


def build_3point(ws):
    ws.title = "3-Point Test Log"
    widths(ws, [("A", 8), ("B", 14), ("C", 18), ("D", 14), ("E", 14), ("F", 14), ("G", 14), ("H", 18), ("I", 22)])
    title(ws, 1, "3-POINT FALL-OF-POTENTIAL TEST LOG", 9)
    header(ws, 3, ["#", "Date", "Electrode Tested", "Aux #1 Dist (ft)", "Aux #2 Dist (ft)",
                   "Resistance (Ω)", "Pass / Fail", "Equipment Serial / Cal Date", "Tester"])
    blanks(ws, 4, 40, 9, {2: "yyyy-mm-dd", 4: "0", 5: "0", 6: "0.0"})


def build_clamp(ws):
    ws.title = "Clamp-On Test Log"
    widths(ws, [("A", 8), ("B", 14), ("C", 22), ("D", 14), ("E", 18), ("F", 22)])
    title(ws, 1, "CLAMP-ON TEST LOG", 6)
    header(ws, 3, ["#", "Date", "Conductor / Path", "Resistance (Ω)", "Equipment Serial", "Tester"])
    blanks(ws, 4, 40, 6, {2: "yyyy-mm-dd", 4: "0.0"})


def build_soil(ws):
    ws.title = "Soil Resistivity"
    widths(ws, [("A", 8), ("B", 14), ("C", 18), ("D", 14), ("E", 14), ("F", 14), ("G", 14), ("H", 22)])
    title(ws, 1, "SOIL RESISTIVITY (4-POINT WENNER)", 8)
    header(ws, 3, ["#", "Date", "Site Location", "Probe Spacing (ft)",
                   "Resistance (Ω)", "Calc'd ρ (Ω·m)", "Soil Type", "Tester"])
    blanks(ws, 4, 30, 8, {2: "yyyy-mm-dd", 4: "0.0", 5: "0.0", 6: "0.0"})


def build_summary(ws):
    ws.title = "Reporting Summary"
    widths(ws, [("A", 30), ("B", 18), ("C", 18), ("D", 22)])
    title(ws, 1, "REPORTING SUMMARY", 4)
    header(ws, 3, ["Category", "Total Tests", "Passed", "Notes"])
    rows = [
        "3-Point Fall-of-Potential",
        "Clamp-On Tests",
        "Soil Resistivity Surveys",
        "Electrodes Inventoried",
    ]
    for i, cat in enumerate(rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=cat).alignment = LEFT
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).font = FONT_BODY


def main():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb = Workbook()
    build_instructions(wb.active)
    build_electrodes(wb.create_sheet("Electrode Inventory"))
    build_3point(wb.create_sheet("3-Point Test Log"))
    build_clamp(wb.create_sheet("Clamp-On Test Log"))
    build_soil(wb.create_sheet("Soil Resistivity"))
    build_summary(wb.create_sheet("Reporting Summary"))
    wb.save(OUT)
    print(f"OK — wrote {OUT}")


if __name__ == "__main__":
    main()
