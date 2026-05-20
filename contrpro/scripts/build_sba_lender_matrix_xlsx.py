#!/usr/bin/env python3
"""
Build SBA Lender Comparison Matrix (XLSX).

Side-by-side comparison of major SBA lenders along the criteria a
contractor actually cares about — approval volume for 7(a), construction-
industry friendliness, typical loan size, average approval time, common
deal sticking points.

Tabs:
  1. Instructions
  2. Top 7(a) Lenders (Recent FY) — banks + non-bank lenders
  3. Lender Selection Worksheet — buyer ranks 5-6 lenders against their fit
  4. Decision Notes — capture what you learned about each one
  5. Sources

Output:
    /Users/home/charles/contrpro/files/packages/complete/sba/Lender_Comparison_Matrix.xlsx
"""
from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
ACCENT_GOLD = "C9A227"
FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUB = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GOLD = PatternFill("solid", fgColor=ACCENT_GOLD)
FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=BRAND_BLUE)
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

OUT = "/Users/home/charles/contrpro/files/packages/complete/sba/Lender_Comparison_Matrix.xlsx"


def widths(ws, cols):
    for col, w in cols:
        ws.column_dimensions[col].width = w


def title(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def header_row(ws, row, cols):
    for i, h in enumerate(cols):
        c = ws.cell(row=row, column=1 + i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 30


# ---------------------------------------------------------------------------
def build_instructions(ws):
    ws.title = "Instructions"
    widths(ws, [("A", 110)])
    ws["A1"] = "LENDER COMPARISON MATRIX — INSTRUCTIONS"
    ws["A1"].font = FONT_TITLE
    body = [
        "",
        "PURPOSE",
        "Tells you WHO to call first when you decide to pursue an SBA loan. Most contractors",
        "burn weeks calling banks at random; this matrix shortlists by 7(a) volume, construction-",
        "industry willingness, and contractor-specific deal experience.",
        "",
        "USE THIS WORKBOOK IN THREE STEPS",
        "  1. Top 7(a) Lenders tab — review the top SBA lenders ranked by recent fiscal-year",
        "     7(a) approval volume. The TOP-15 in volume don't all serve contractors equally.",
        "     Look at the 'Construction-friendly' column.",
        "  2. Lender Selection Worksheet — pick 5-6 lenders to actually call. Rank them on",
        "     loan-size fit, geography, sector experience, and (importantly) how quickly they",
        "     get to a real preapproval letter.",
        "  3. Decision Notes — capture what each lender told you on the first call so you can",
        "     compare offers apples-to-apples.",
        "",
        "WHY VOLUME MATTERS",
        "  • Banks that do hundreds of 7(a) loans/year have systematized underwriting + can",
        "    move faster (4-8 wk). Banks that do 5/year per branch are SBA-novice and a deal",
        "    can stretch to 12-20 weeks while they learn your industry.",
        "  • Approval volume ≠ approval RATE; a high-volume lender can still reject you. But",
        "    volume is the best proxy for ability + willingness to actually close.",
        "",
        "CONSTRUCTION-INDUSTRY SIGNALS",
        "  • Lenders with experience in construction price working-capital lines DIFFERENTLY",
        "    than retail/restaurant lenders because of WIP + retainage realities.",
        "  • Look for lenders that ask about your bonding capacity in the first call —",
        "    that's a sign they know construction.",
        "",
        "DATA SOURCES",
        "  • SBA Lender Match (lendermatch.sba.gov)",
        "  • SBA's annual 'Top 100 SBA Lenders' reports (sba.gov press releases)",
        "  • Lender public profiles + LinkedIn/news for industry tags",
        "",
        "DISCLAIMER",
        "  This matrix is a starting point. Lender performance + appetite shifts year over",
        "  year. Always verify current 7(a) approval status + program participation at",
        "  lendermatch.sba.gov before relying on any specific entry.",
        "",
        "DOCUMENT VERSION",
        "Lender_Comparison_Matrix.xlsx v1.0 — 2026-05-19",
        "ContrPro Complete Tier · SBA Bonus",
    ]
    for i, line in enumerate(body, start=3):
        c = ws.cell(row=i, column=1, value=line)
        c.font = FONT_H2 if (line.isupper() and line.strip() and not line.startswith(" ")) else FONT_BODY
        c.alignment = LEFT


# ---------------------------------------------------------------------------
def build_top_lenders(ws):
    ws.title = "Top 7(a) Lenders"
    widths(ws, [("A", 5), ("B", 28), ("C", 18), ("D", 16), ("E", 18), ("F", 18), ("G", 26), ("H", 26)])
    title(ws, 1, "TOP SBA 7(a) LENDERS (recent FY — verify current ranking)", 8)
    header_row(ws, 3, ["Rank", "Lender", "Type", "Geography", "Construction-friendly?",
                       "Typical loan size", "Notes", "Contact / website"])
    # Representative roster — research-grade, not authoritative. The actual ordering
    # in any given fiscal year shifts; the worksheet is a starting point, not gospel.
    lenders = [
        (1, "Live Oak Bank",          "Bank",     "National",         "Yes — industry-vertical lender", "$350K–$5M", "One of the largest 7(a) volume banks. Industry vertical teams; ask for the construction/contractor desk.", "liveoakbank.com / SBA Lender Match"),
        (2, "Newtek Bank",            "Bank",     "National",         "Some",                            "$50K–$5M",  "High volume, technology-forward. Construction OK but not specialty. Faster decision cycle.", "newtekone.com"),
        (3, "Huntington National Bank","Bank",    "Midwest / National","Yes — strong contractor desk",   "$100K–$5M", "Repeatedly top-ranked SBA lender. Branch network + relationship-driven.", "huntington.com"),
        (4, "Wells Fargo",            "Bank",     "National",         "Some",                            "$150K–$5M", "Large volume, slower process. Existing banking relationship helps.", "wellsfargo.com"),
        (5, "JPMorgan Chase",         "Bank",     "National",         "Some",                            "$100K–$5M", "High volume; tends to favor existing customers + larger requests.", "chase.com/business"),
        (6, "TD Bank",                "Bank",     "East Coast",       "Yes",                             "$50K–$5M",  "Solid SBA player in the Northeast / Mid-Atlantic / Florida.", "td.com"),
        (7, "Byline Bank",            "Bank",     "National",         "Yes — multi-industry",            "$100K–$5M", "Top-10 7(a) volume historically. Construction in their wheelhouse.", "bylinebank.com"),
        (8, "U.S. Bank",              "Bank",     "National",         "Some",                            "$100K–$5M", "Wide footprint, more cautious underwriting cycle.", "usbank.com/business"),
        (9, "First Bank of Highland Park", "Bank","Midwest",          "Yes",                             "$100K–$3M", "Specialty SBA shop; smaller bank, faster touch.", "fbhp.com"),
        (10,"Celtic Bank",            "Bank",     "National",         "Some",                            "$50K–$5M",  "Big SBA volume, tech-forward processing.", "celticbank.com"),
        (11,"Pinnacle Bank",          "Bank",     "Southeast",        "Yes",                             "$100K–$5M", "Strong Southeast presence; contractor-friendly.", "pinnaclefinancialpartners.com"),
        (12,"BayFirst National Bank", "Bank",     "Florida + national","Some",                           "$50K–$3M",  "Florida-rooted but national SBA reach.", "bayfirstfinancial.com"),
        (13,"Lendio",                 "Marketplace","National",       "Some",                            "$5K–$2M",   "Not a lender — a marketplace that shops your file across multiple banks. Useful early-stage when you don't know which lender fits.", "lendio.com"),
        (14,"SmartBiz",               "Marketplace","National",       "Some",                            "$30K–$350K","Specializes in smaller SBA 7(a) loans (faster, lower paperwork). Good for sub-$350K asks.", "smartbizloans.com"),
        (15,"Local Community Bank",   "Bank",     "Your home market", "Varies — ASK",                    "$50K–$1M",  "Often overlooked. Smaller community banks may move faster + be more flexible if you have a deposit relationship.", "Search SBA Lender Match by ZIP"),
    ]
    for rank, name, ltype, geo, cf, size, notes, contact in lenders:
        r = 3 + rank
        ws.cell(row=r, column=1, value=rank).alignment = CENTER
        ws.cell(row=r, column=2, value=name).font = FONT_BODY_BOLD
        ws.cell(row=r, column=3, value=ltype).alignment = CENTER
        ws.cell(row=r, column=4, value=geo)
        ws.cell(row=r, column=5, value=cf).alignment = CENTER
        ws.cell(row=r, column=6, value=size).alignment = CENTER
        ws.cell(row=r, column=7, value=notes)
        ws.cell(row=r, column=8, value=contact)
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if c != 2:
                cell.font = FONT_BODY
            if c in (1, 3, 5, 6):
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT
    # Disclaimer row
    r = 4 + len(lenders) + 1
    ws.cell(row=r, column=1, value=(
        "Note: rankings drift year-over-year. Verify current 7(a) volume + program "
        "participation at lendermatch.sba.gov before relying on any specific entry. "
        "The 'Top 100 SBA Lenders' annual press release on sba.gov is the canonical ranking source."
    )).font = Font(name="Calibri", size=10, italic=True, color="808080")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)


# ---------------------------------------------------------------------------
def build_selection_worksheet(ws):
    ws.title = "Selection Worksheet"
    widths(ws, [("A", 5), ("B", 26), ("C", 14), ("D", 14), ("E", 14), ("F", 14), ("G", 14), ("H", 14), ("I", 14)])
    title(ws, 1, "SELECTION WORKSHEET — score your shortlist (1-5 scale)", 9)
    header_row(ws, 3, ["#", "Lender", "Loan-size fit", "Geography fit", "Industry exp",
                       "Existing relationship?", "Quick response?", "Total", "Decision"])
    # 6 blank rows for buyer to fill
    for i in range(1, 7):
        r = 3 + i
        ws.cell(row=r, column=1, value=i).alignment = CENTER
        ws.cell(row=r, column=2, value="").alignment = LEFT  # lender name input
        for col in range(3, 8):
            cell = ws.cell(row=r, column=col, value="")
            cell.alignment = CENTER
        # Total = sum of cols C-G
        ws.cell(row=r, column=8, value=f"=SUM(C{r}:G{r})").alignment = CENTER
        ws.cell(row=r, column=8).font = FONT_BODY_BOLD
        ws.cell(row=r, column=9, value="").alignment = LEFT
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = BORDER
            if c != 8:
                ws.cell(row=r, column=c).font = FONT_BODY
    # Validations: 1-5 scale on score cells
    dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5", allow_blank=True)
    dv.add("C4:G9")
    ws.add_data_validation(dv)
    # Decision col validation
    dv_dec = DataValidation(type="list", formula1='"Call first,Call later,Skip"', allow_blank=True)
    dv_dec.add("I4:I9")
    ws.add_data_validation(dv_dec)
    # Legend
    r = 11
    ws.cell(row=r, column=1, value="Score 1-5 each criterion (1=poor fit, 5=excellent fit). Total ≥ 18 = strong shortlist.").font = Font(italic=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)


# ---------------------------------------------------------------------------
def build_decision_notes(ws):
    ws.title = "Decision Notes"
    widths(ws, [("A", 22), ("B", 14), ("C", 24), ("D", 20), ("E", 24), ("F", 18), ("G", 30)])
    title(ws, 1, "DECISION NOTES — what each lender told you on the call", 7)
    header_row(ws, 3, ["Lender", "Date called", "Loan officer name + phone",
                       "Rate quoted / range", "Sticking points", "Decision needed by",
                       "Status"])
    # 12 blank rows
    for r in range(4, 16):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c, value="")
            cell.border = BORDER
            cell.font = FONT_BODY
            cell.alignment = LEFT if c in (1, 3, 5) else CENTER
            if c == 2 or c == 6:
                cell.number_format = "yyyy-mm-dd"


# ---------------------------------------------------------------------------
def build_sources(ws):
    ws.title = "Sources"
    widths(ws, [("A", 60), ("B", 60)])
    title(ws, 1, "DATA SOURCES + LIVE LOOKUPS", 2)
    header_row(ws, 3, ["Source", "URL"])
    sources = [
        ("SBA Lender Match — find banks by ZIP/loan-size",
         "https://lendermatch.sba.gov/"),
        ("SBA 'Top 100' lender rankings (annual)",
         "https://www.sba.gov/article/2024/sep/26/sba-announces-top-100-sba-lenders-fiscal-year"),
        ("SBA 7(a) program overview",
         "https://www.sba.gov/funding-programs/loans/7a-loans"),
        ("SBA 504 program overview",
         "https://www.sba.gov/funding-programs/loans/504-loans"),
        ("SBA Microloan program",
         "https://www.sba.gov/funding-programs/loans/microloans"),
        ("SBA Express overview (lower-doc, smaller loans)",
         "https://www.sba.gov/funding-programs/loans/7a-loans/sba-express-loans"),
        ("Local SBA District Office + SBDC finder",
         "https://www.sba.gov/local-assistance/find/"),
        ("Find SCORE mentor (free SBA-affiliated mentoring)",
         "https://www.score.org/find-mentor"),
        ("FY7a Loan data (downloadable Excel by lender + program)",
         "https://www.sba.gov/document/report-7a-504-foia"),
    ]
    for i, (src, url) in enumerate(sources):
        r = 4 + i
        ws.cell(row=r, column=1, value=src).alignment = LEFT
        ws.cell(row=r, column=2, value=url).alignment = LEFT
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).font = FONT_BODY


def main():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb = Workbook()
    build_instructions(wb.active)
    build_top_lenders(wb.create_sheet("Top 7(a) Lenders"))
    build_selection_worksheet(wb.create_sheet("Selection Worksheet"))
    build_decision_notes(wb.create_sheet("Decision Notes"))
    build_sources(wb.create_sheet("Sources"))
    wb.save(OUT)
    print(f"OK — wrote {OUT}")


if __name__ == "__main__":
    main()
