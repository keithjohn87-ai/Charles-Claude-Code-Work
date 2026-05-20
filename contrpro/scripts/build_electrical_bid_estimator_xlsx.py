#!/usr/bin/env python3
"""
Build ContrPro Electrical Bid Estimator (XLSX) — Electrical trade pack.

Industry-standard NECA-labor-unit-based bid estimator. CSI Division 26
primary with cross-references to 27 (communications) and 28 (electronic
safety + security).

Tabs:
  1. Instructions
  2. Project Info
  3. Wage Rates
  4. Material Takeoff (CSI 26-coded line items — conduit, wire, devices, gear)
  5. Labor Estimate (NECA Labor Units × crew × rate)
  6. Equipment Estimate
  7. Subcontract Estimate
  8. Premiums & Conditions
  9. Bid Summary
  10. CSI Reference

Output:
    /Users/home/charles/contrpro/files/packages/complete/electrical/Electrical_Bid_Estimator.xlsx
"""
from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
ACCENT_GOLD = "C9A227"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GOLD = PatternFill("solid", fgColor=ACCENT_GOLD)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_GREY_IT = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_HRS = '_(* #,##0.0_);_(* (#,##0.0);_(* "-"_);_(@_)'
FMT_PCT = "0.0%"
FMT_INT = "0"
FMT_NUM2 = "0.00"

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

OUT = "/Users/home/charles/contrpro/files/packages/complete/electrical/Electrical_Bid_Estimator.xlsx"


def widths(ws, cols):
    for col, w in cols:
        ws.column_dimensions[col].width = w


def title(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def header(ws, row, cols):
    for i, h in enumerate(cols):
        c = ws.cell(row=row, column=1 + i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 26


def build_instructions(ws):
    ws.title = "Instructions"
    widths(ws, [("A", 110)])
    ws["A1"] = "ELECTRICAL BID ESTIMATOR — INSTRUCTIONS"
    ws["A1"].font = FONT_TITLE
    body = [
        "",
        "PURPOSE",
        "Production estimating tool for commercial electrical contractors. Uses NECA Labor",
        "Units (per device + per LF of conduit + per termination). CSI Division 26 primary",
        "with cross-references to 27 (communications) and 28 (electronic safety + security).",
        "Output rolls up by CSI division so the bid integrates cleanly into the GC's job-cost",
        "accounting + into your monthly billing on the Universal Sub Suite's",
        "Sub_Schedule_of_Values.xlsx.",
        "",
        "WORKFLOW",
        "  1. Project Info — fill every blue field.",
        "  2. Wage Rates — set base + fringe + OT premium by classification.",
        "  3. Material Takeoff — quantities by CSI line item (conduit type+size, wire AWG,",
        "     device class, gear). Unit costs flow into Bid Summary.",
        "  4. Labor Estimate — NECA Labor Units assume average commercial conditions.",
        "     Adjust productivity factor for difficult work.",
        "  5. Equipment — daily/weekly rentals (lifts, generator, bender, knockout sets).",
        "  6. Subcontract — fire alarm certification, low-voltage, gear assembly + factory test.",
        "  7. Premiums — height, hazardous, prevailing wage, occupied-building disruption.",
        "  8. Bid Summary — automatic roll-up + final bid number.",
        "",
        "NECA LABOR UNITS",
        "Hours-per-unit values approximate published NECA Labor Units for standard commercial",
        "work. Confirm against your jurisdiction + the latest NECA edition. Calibrate against",
        "your historical productivity. Underestimating labor is the #1 cause of electrical-",
        "bid losses; if your historical productivity differs from these defaults by more than",
        "15%, change the defaults — don't ignore the gap.",
        "",
        "FIELD-ONLY SCOPE",
        "Calibrated for field installation. Switchgear assembly + factory acceptance testing,",
        "specialty fabricated bus assemblies, and modular substation work are typically",
        "subbed and entered on the Subcontract tab.",
        "",
        "DOCUMENT VERSION",
        "Electrical_Bid_Estimator.xlsx v1.0 — 2026-05-19",
        "ContrPro Complete Tier · Electrical Trade Pack",
    ]
    for i, line in enumerate(body, start=3):
        c = ws.cell(row=i, column=1, value=line)
        if line.isupper() and line.strip() and not line.startswith(" "):
            c.font = FONT_H2
        else:
            c.font = FONT_BODY
        c.alignment = LEFT


def build_project_info(ws):
    ws.title = "Project Info"
    widths(ws, [("A", 32), ("B", 60)])
    title(ws, 1, "PROJECT INFORMATION", 2)
    fields = [
        ("Project Name", ""),
        ("Project Address", ""),
        ("Owner", ""),
        ("General Contractor", ""),
        ("Architect / Engineer", ""),
        ("Bid Due Date", ""),
        ("Bid Reference / Number", ""),
        ("Project Type", "(New construction / Renovation / Tenant fit-out / Adaptive reuse)"),
        ("Building Use", "(Office / Retail / Healthcare / Education / Multi-family / Hospitality / Industrial / Mixed)"),
        ("Approx. Square Footage", ""),
        ("Stories", ""),
        ("Service Size (Amps / Volts / Phase)", ""),
        ("AHJ — Electrical Inspector", ""),
        ("Applicable NEC Edition", "(NEC 2023 + local amendments)"),
        ("Estimator Name", ""),
        ("Estimator Phone", ""),
        ("Estimator Email", ""),
        ("Date Prepared", ""),
        ("Internal Bid Number", ""),
        ("Notes / Assumptions", ""),
    ]
    row = 3
    for label, default in fields:
        ws.cell(row=row, column=1, value=label).font = FONT_BODY_BOLD
        ws.cell(row=row, column=1).fill = FILL_SUBHEADER
        ws.cell(row=row, column=1).alignment = LEFT
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=2, value=default).font = FONT_GREY_IT if default else FONT_BODY
        ws.cell(row=row, column=2).alignment = LEFT
        ws.cell(row=row, column=2).border = BORDER
        row += 1


def build_wages(ws, wb):
    ws.title = "Wage Rates"
    widths(ws, [("A", 28), ("B", 14), ("C", 14), ("D", 14), ("E", 16), ("F", 16)])
    title(ws, 1, "WAGE RATES (calibrate to your market)", 6)
    header(ws, 3, ["Classification", "Base Rate ($/hr)", "Fringe ($/hr)", "Burden % (WC/GL/FICA)", "OT Premium (1.5×)", "Loaded Rate ($/hr)"])
    rows = [
        ("Electrical Foreman", 62.00, 19.50, 0.32),
        ("Journeyman Electrician", 52.00, 19.50, 0.32),
        ("Apprentice (Yr 4)", 40.00, 14.50, 0.32),
        ("Apprentice (Yr 2)", 28.00, 12.00, 0.32),
        ("Material Handler / Helper", 22.00, 9.50, 0.32),
        ("Low-Voltage Tech", 48.00, 17.00, 0.32),
        ("Fire Alarm Certified Tech", 58.00, 19.00, 0.32),
    ]
    for i, (cls, base, fringe, burden) in enumerate(rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=cls).font = FONT_BODY_BOLD
        ws.cell(row=r, column=2, value=base).number_format = FMT_USD
        ws.cell(row=r, column=3, value=fringe).number_format = FMT_USD
        ws.cell(row=r, column=4, value=burden).number_format = FMT_PCT
        ws.cell(row=r, column=5, value=f"=B{r}*1.5").number_format = FMT_USD
        ws.cell(row=r, column=6, value=f"=(B{r}+C{r})*(1+D{r})").number_format = FMT_USD
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).alignment = CENTER if col > 1 else LEFT
    last = 4 + len(rows) - 1
    wb.defined_names["EC_LoadedRates"] = DefinedName("EC_LoadedRates", attr_text=f"'Wage Rates'!$F$4:$F${last}")


def build_material(ws):
    ws.title = "Material Takeoff"
    widths(ws, [("A", 14), ("B", 46), ("C", 10), ("D", 10), ("E", 14), ("F", 16), ("G", 14)])
    title(ws, 1, "MATERIAL TAKEOFF (CSI Division 26)", 7)
    header(ws, 3, ["CSI Code", "Description", "QTY", "UOM", "Unit Cost ($)", "Material Subtotal ($)", "Notes"])
    items = [
        # 26 05 — Common work results
        ("26 05 19", "Conductor — THHN/THWN-2, #14 AWG copper", 0, "LF", 0.18, ""),
        ("26 05 19", "Conductor — THHN/THWN-2, #12 AWG copper", 0, "LF", 0.28, ""),
        ("26 05 19", "Conductor — THHN/THWN-2, #10 AWG copper", 0, "LF", 0.48, ""),
        ("26 05 19", "Conductor — THHN/THWN-2, #8 AWG copper", 0, "LF", 0.85, ""),
        ("26 05 19", "Conductor — THHN/THWN-2, #6 AWG copper", 0, "LF", 1.45, ""),
        ("26 05 19", "Conductor — THHN/THWN-2, #4 AWG copper", 0, "LF", 2.40, ""),
        ("26 05 19", "Conductor — THHN/THWN-2, #2 AWG copper", 0, "LF", 3.85, ""),
        ("26 05 19", "Conductor — THHN/THWN-2, 1/0 AWG copper", 0, "LF", 6.50, ""),
        ("26 05 19", "Conductor — THHN/THWN-2, 2/0 AWG copper", 0, "LF", 8.40, ""),
        ("26 05 19", "Conductor — THHN/THWN-2, 4/0 AWG copper", 0, "LF", 13.25, ""),
        ("26 05 19", "Conductor — THHN/THWN-2, 250 MCM copper", 0, "LF", 15.75, ""),
        ("26 05 19", "Conductor — THHN/THWN-2, 500 MCM copper", 0, "LF", 28.00, ""),
        ("26 05 19", "Conductor — XHHW-2 aluminum, #6 AWG", 0, "LF", 0.85, ""),
        ("26 05 19", "Conductor — XHHW-2 aluminum, 250 MCM", 0, "LF", 5.95, ""),
        ("26 05 19", "Conductor — XHHW-2 aluminum, 500 MCM", 0, "LF", 11.50, ""),
        # 26 05 26 — Grounding + Bonding
        ("26 05 26", "Ground rod — 5/8\" × 8' copper", 0, "EA", 14.50, ""),
        ("26 05 26", "Ground rod — 5/8\" × 10' copper", 0, "EA", 21.50, ""),
        ("26 05 26", "Ground clamp — acorn", 0, "EA", 3.85, ""),
        ("26 05 26", "Bare copper bond — #6 AWG", 0, "LF", 1.20, ""),
        ("26 05 26", "Bare copper bond — #4 AWG", 0, "LF", 1.85, ""),
        ("26 05 26", "Bare copper bond — 2/0 AWG", 0, "LF", 7.95, ""),
        # 26 05 29 — Hangers + Supports
        ("26 05 29", "Unistrut B-line 12 ga, 10' length", 0, "EA", 18.50, ""),
        ("26 05 29", "Conduit clamp + spring nut", 0, "EA", 1.85, ""),
        ("26 05 29", "Threaded rod 3/8\" × 6'", 0, "EA", 6.50, ""),
        # 26 05 33 — Raceway
        ("26 05 33", "EMT conduit — 1/2\"", 0, "LF", 0.65, ""),
        ("26 05 33", "EMT conduit — 3/4\"", 0, "LF", 0.95, ""),
        ("26 05 33", "EMT conduit — 1\"", 0, "LF", 1.45, ""),
        ("26 05 33", "EMT conduit — 1-1/2\"", 0, "LF", 2.65, ""),
        ("26 05 33", "EMT conduit — 2\"", 0, "LF", 3.85, ""),
        ("26 05 33", "EMT conduit — 3\"", 0, "LF", 7.85, ""),
        ("26 05 33", "EMT conduit — 4\"", 0, "LF", 11.50, ""),
        ("26 05 33", "Rigid steel conduit — 1\"", 0, "LF", 5.85, ""),
        ("26 05 33", "Rigid steel conduit — 2\"", 0, "LF", 12.50, ""),
        ("26 05 33", "PVC schedule 40 — 3/4\"", 0, "LF", 0.55, ""),
        ("26 05 33", "PVC schedule 40 — 1\"", 0, "LF", 0.85, ""),
        ("26 05 33", "PVC schedule 40 — 2\"", 0, "LF", 2.20, ""),
        ("26 05 33", "PVC schedule 40 — 4\"", 0, "LF", 6.50, ""),
        ("26 05 33", "MC cable — 12/2 with ground", 0, "LF", 0.85, ""),
        ("26 05 33", "MC cable — 12/3 with ground", 0, "LF", 1.15, ""),
        ("26 05 33", "Flex conduit — 1/2\" Greenfield", 0, "LF", 1.20, ""),
        ("26 05 33", "Liquidtight flex — 3/4\"", 0, "LF", 4.85, ""),
        # Boxes + fittings
        ("26 05 33", "4-S box — galvanized", 0, "EA", 4.85, ""),
        ("26 05 33", "Single gang box — handy box", 0, "EA", 2.85, ""),
        ("26 05 33", "Plaster ring — single gang", 0, "EA", 1.45, ""),
        ("26 05 33", "Plaster ring — 2-gang", 0, "EA", 2.20, ""),
        ("26 05 33", "PVC adapter — male, 3/4\"", 0, "EA", 0.85, ""),
        ("26 05 33", "EMT set-screw connector — 1/2\"", 0, "EA", 0.55, ""),
        # 26 24 — Switchboards + panels
        ("26 24 16", "Panelboard — 200A, 42-circuit, MLO", 0, "EA", 1685.00, ""),
        ("26 24 16", "Panelboard — 400A, 84-circuit MCB", 0, "EA", 3850.00, ""),
        ("26 24 16", "Panelboard — 800A, MCB + sub-feed", 0, "EA", 7250.00, ""),
        ("26 24 13", "Switchboard — 1200A service entrance", 0, "EA", 18500.00, ""),
        ("26 24 13", "Switchboard — 2500A main", 0, "EA", 38500.00, ""),
        ("26 24 19", "MCC bucket — combination starter", 0, "EA", 1850.00, ""),
        # 26 27 — Service entrance
        ("26 27 13", "Service entrance — 200A residential", 0, "EA", 1450.00, ""),
        ("26 27 13", "CT cabinet — 800A", 0, "EA", 3850.00, ""),
        ("26 27 13", "Meter base — 200A self-contained", 0, "EA", 285.00, ""),
        # 26 28 — Disconnects + OCP
        ("26 28 16", "Safety switch — 60A fused 600V", 0, "EA", 145.00, ""),
        ("26 28 16", "Safety switch — 200A non-fused 600V", 0, "EA", 485.00, ""),
        ("26 28 16", "Safety switch — 400A fused", 0, "EA", 985.00, ""),
        ("26 28 18", "Circuit breaker — 20A 1P", 0, "EA", 12.50, ""),
        ("26 28 18", "Circuit breaker — 50A 2P", 0, "EA", 38.50, ""),
        ("26 28 18", "Circuit breaker — 100A 3P", 0, "EA", 145.00, ""),
        ("26 28 18", "Circuit breaker — 400A 3P", 0, "EA", 1485.00, ""),
        # 26 29 — Motor controllers
        ("26 29 13", "Motor starter — size 0", 0, "EA", 285.00, ""),
        ("26 29 13", "Motor starter — size 2", 0, "EA", 485.00, ""),
        ("26 29 23", "Variable frequency drive — 5 HP", 0, "EA", 685.00, ""),
        ("26 29 23", "Variable frequency drive — 25 HP", 0, "EA", 2450.00, ""),
        # 26 32 — Generators (where in scope)
        ("26 32 13", "Standby generator — 50 kW diesel", 0, "EA", 28500.00, "Confirm scope split"),
        ("26 36 23", "Automatic transfer switch — 200A", 0, "EA", 4850.00, ""),
        # 26 41 — Lightning protection
        ("26 41 13", "Air terminal + base", 0, "EA", 38.50, ""),
        ("26 41 16", "Down conductor — copper braid", 0, "LF", 4.85, ""),
        # 26 51 — Interior lighting
        ("26 51 13", "LED 2x4 troffer — 40W", 0, "EA", 145.00, ""),
        ("26 51 13", "LED 2x2 troffer — 25W", 0, "EA", 115.00, ""),
        ("26 51 13", "LED high-bay — 150W warehouse", 0, "EA", 185.00, ""),
        ("26 51 13", "LED downlight — 6\" can", 0, "EA", 65.00, ""),
        ("26 51 13", "Linear LED strip — 4'", 0, "EA", 95.00, ""),
        ("26 51 13", "Wall sconce — LED", 0, "EA", 145.00, ""),
        ("26 51 13", "Occupancy sensor — wall switch", 0, "EA", 45.00, ""),
        ("26 51 13", "Occupancy sensor — ceiling-mount PIR", 0, "EA", 78.00, ""),
        ("26 51 13", "Daylight sensor", 0, "EA", 145.00, ""),
        ("26 51 13", "Emergency battery pack — 90 min", 0, "EA", 285.00, ""),
        ("26 51 13", "Exit sign — LED w/ battery", 0, "EA", 95.00, ""),
        # 26 52 — Exterior lighting
        ("26 52 13", "Wall pack — LED 40W full cutoff", 0, "EA", 145.00, ""),
        ("26 52 13", "Pole-mount LED — 150W shoebox", 0, "EA", 685.00, ""),
        ("26 52 13", "Bollard light — LED 15W", 0, "EA", 285.00, ""),
        # Devices
        ("26 27 26", "Receptacle — 20A duplex, decora", 0, "EA", 4.85, ""),
        ("26 27 26", "Receptacle — 20A GFCI", 0, "EA", 18.50, ""),
        ("26 27 26", "Receptacle — 30A 250V", 0, "EA", 32.00, ""),
        ("26 27 26", "Receptacle — 50A 250V dryer/range", 0, "EA", 38.00, ""),
        ("26 27 26", "Switch — 20A single-pole", 0, "EA", 3.85, ""),
        ("26 27 26", "Switch — 20A 3-way", 0, "EA", 6.50, ""),
        ("26 27 26", "Switch — 20A 4-way", 0, "EA", 18.50, ""),
        ("26 27 26", "Dimmer — 0-10V LED-rated", 0, "EA", 38.00, ""),
        ("26 27 26", "Wallplate — single gang", 0, "EA", 1.45, ""),
        # Specialty
        ("26 27 73", "Floor box — flush 6-gang", 0, "EA", 285.00, ""),
        ("26 27 73", "Poke-thru floor box", 0, "EA", 485.00, ""),
        # Cabling — low voltage
        ("27 11 13", "Cat 6A cable", 0, "LF", 0.85, "If in EC scope"),
        ("27 11 13", "Cat 6A RJ-45 jack + faceplate", 0, "EA", 14.50, ""),
        ("27 13 23", "Fiber multimode 12-strand OM4", 0, "LF", 4.85, ""),
        # Permit / inspection
        ("26 00 10", "Permit — electrical", 0, "LS", 0.00, "Allowance — AHJ amount"),
        ("26 00 10", "Inspection re-trip allowance", 0, "LS", 0.00, ""),
    ]
    for i, (csi, desc, qty, uom, unit, notes) in enumerate(items):
        r = 4 + i
        ws.cell(row=r, column=1, value=csi).alignment = CENTER
        ws.cell(row=r, column=2, value=desc).alignment = LEFT
        ws.cell(row=r, column=3, value=qty).number_format = FMT_INT
        ws.cell(row=r, column=4, value=uom).alignment = CENTER
        ws.cell(row=r, column=5, value=unit).number_format = FMT_USD
        ws.cell(row=r, column=6, value=f"=C{r}*E{r}").number_format = FMT_USD
        ws.cell(row=r, column=7, value=notes).alignment = LEFT
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = FONT_BODY
    last = 4 + len(items) - 1
    sub_row = last + 2
    ws.cell(row=sub_row, column=5, value="Material Subtotal").font = FONT_BODY_BOLD
    ws.cell(row=sub_row, column=5).alignment = RIGHT
    ws.cell(row=sub_row, column=6, value=f"=SUM(F4:F{last})").number_format = FMT_USD
    ws.cell(row=sub_row, column=6).font = FONT_BODY_BOLD
    ws.cell(row=sub_row, column=6).fill = FILL_SUBHEADER
    tax_row = sub_row + 1
    ws.cell(row=tax_row, column=5, value="Sales tax %").alignment = RIGHT
    ws.cell(row=tax_row, column=5).font = FONT_BODY_BOLD
    ws.cell(row=tax_row, column=6, value=0.07).number_format = FMT_PCT
    tax_amt_row = tax_row + 1
    ws.cell(row=tax_amt_row, column=5, value="Sales tax $").alignment = RIGHT
    ws.cell(row=tax_amt_row, column=5).font = FONT_BODY_BOLD
    ws.cell(row=tax_amt_row, column=6, value=f"=F{sub_row}*F{tax_row}").number_format = FMT_USD
    tot_row = tax_amt_row + 1
    ws.cell(row=tot_row, column=5, value="Material TOTAL (incl tax)").font = FONT_BODY_BOLD
    ws.cell(row=tot_row, column=5).alignment = RIGHT
    ws.cell(row=tot_row, column=5).fill = FILL_GOLD
    ws.cell(row=tot_row, column=6, value=f"=F{sub_row}+F{tax_amt_row}").number_format = FMT_USD
    ws.cell(row=tot_row, column=6).fill = FILL_GOLD
    ws.cell(row=tot_row, column=6).font = FONT_BODY_BOLD
    return ("EC_MaterialTotal", f"'Material Takeoff'!$F${tot_row}")


def build_labor(ws):
    ws.title = "Labor Estimate"
    widths(ws, [("A", 14), ("B", 44), ("C", 10), ("D", 10), ("E", 16), ("F", 14), ("G", 16), ("H", 18)])
    title(ws, 1, "LABOR ESTIMATE (NECA Labor Units)", 8)
    header(ws, 3, ["CSI Code", "Activity", "QTY", "UOM", "Hours / UOM", "Subtotal Hours", "Crew Loaded Rate ($/hr)", "Subtotal Labor ($)"])
    # NECA-style productivity placeholders — calibrate to local
    items = [
        # Conduit install (per LF)
        ("26 05 33", "Install EMT 1/2\" - 3/4\" exposed", 0, "LF", 0.045, 70),
        ("26 05 33", "Install EMT 1\" - 1-1/4\" exposed", 0, "LF", 0.060, 70),
        ("26 05 33", "Install EMT 2\" - 3\" exposed", 0, "LF", 0.090, 70),
        ("26 05 33", "Install EMT 4\" exposed", 0, "LF", 0.130, 70),
        ("26 05 33", "Install EMT 1/2\" - 1\" concealed", 0, "LF", 0.055, 70),
        ("26 05 33", "Install rigid steel 1\" - 2\"", 0, "LF", 0.110, 70),
        ("26 05 33", "Install rigid steel 3\" - 4\"", 0, "LF", 0.180, 70),
        ("26 05 33", "Install PVC schedule 40 1\" - 2\"", 0, "LF", 0.040, 70),
        ("26 05 33", "Install PVC schedule 40 3\" - 4\"", 0, "LF", 0.075, 70),
        ("26 05 33", "Install MC cable per LF", 0, "LF", 0.025, 70),
        ("26 05 33", "Install liquidtight flex", 0, "LF", 0.075, 70),
        # Wire pulls (per LF)
        ("26 05 19", "Pull #14 - #10 AWG, 3-conductor", 0, "LF", 0.012, 70),
        ("26 05 19", "Pull #8 - #4 AWG, 3-conductor", 0, "LF", 0.025, 70),
        ("26 05 19", "Pull #2 AWG - 1/0, 3-conductor", 0, "LF", 0.045, 70),
        ("26 05 19", "Pull 2/0 - 250 MCM, 3-conductor", 0, "LF", 0.075, 70),
        ("26 05 19", "Pull 500 MCM, 3-conductor", 0, "LF", 0.150, 70),
        # Terminations (per termination)
        ("26 05 19", "Terminate #14 - #10 AWG", 0, "EA", 0.10, 70),
        ("26 05 19", "Terminate #8 - #4 AWG (lug)", 0, "EA", 0.20, 70),
        ("26 05 19", "Terminate #2 - 1/0 AWG (lug)", 0, "EA", 0.30, 70),
        ("26 05 19", "Terminate 2/0 - 250 MCM (compression)", 0, "EA", 0.50, 70),
        ("26 05 19", "Terminate 500 MCM (compression)", 0, "EA", 0.75, 70),
        # Boxes + plates
        ("26 05 33", "Install 4-S box exposed", 0, "EA", 0.20, 70),
        ("26 05 33", "Install handy box + plate exposed", 0, "EA", 0.15, 70),
        ("26 05 33", "Install device box concealed (rough-in)", 0, "EA", 0.18, 70),
        # Devices (set + connect)
        ("26 27 26", "Set + connect receptacle 20A", 0, "EA", 0.20, 70),
        ("26 27 26", "Set + connect GFCI receptacle", 0, "EA", 0.30, 70),
        ("26 27 26", "Set + connect single-pole switch", 0, "EA", 0.18, 70),
        ("26 27 26", "Set + connect 3-way switch", 0, "EA", 0.30, 70),
        ("26 27 26", "Set + connect dimmer (0-10V)", 0, "EA", 0.35, 70),
        ("26 27 26", "Set + connect 30A receptacle", 0, "EA", 0.45, 70),
        # Lighting
        ("26 51 13", "Hang LED 2x4 troffer + connect", 0, "EA", 0.85, 70),
        ("26 51 13", "Hang LED 2x2 troffer + connect", 0, "EA", 0.70, 70),
        ("26 51 13", "Hang LED high-bay (lift)", 0, "EA", 1.50, 70),
        ("26 51 13", "Install LED downlight 6\"", 0, "EA", 0.60, 70),
        ("26 51 13", "Install wall sconce", 0, "EA", 0.55, 70),
        ("26 51 13", "Install occupancy sensor wall", 0, "EA", 0.35, 70),
        ("26 51 13", "Install occupancy sensor ceiling", 0, "EA", 0.60, 70),
        ("26 51 13", "Install emergency battery pack", 0, "EA", 0.60, 70),
        ("26 51 13", "Install exit sign", 0, "EA", 0.50, 70),
        ("26 52 13", "Install wall-pack exterior", 0, "EA", 0.75, 70),
        ("26 52 13", "Install pole-mount fixture + arm", 0, "EA", 3.50, 70),
        # Panels + gear
        ("26 24 16", "Set + connect 200A panelboard", 0, "EA", 4.50, 70),
        ("26 24 16", "Set + connect 400A panelboard", 0, "EA", 6.50, 70),
        ("26 24 16", "Set + connect 800A panelboard", 0, "EA", 10.50, 70),
        ("26 24 13", "Set + connect 1200A switchboard (rig + connect)", 0, "EA", 24.00, 70),
        ("26 27 13", "Set + connect 200A service entrance", 0, "EA", 6.00, 70),
        ("26 28 16", "Set + connect 60A-200A safety switch", 0, "EA", 1.20, 70),
        ("26 28 16", "Set + connect 400A safety switch", 0, "EA", 2.50, 70),
        # Grounding
        ("26 05 26", "Drive ground rod + bond", 0, "EA", 0.45, 70),
        ("26 05 26", "Install ground bar in panel", 0, "EA", 0.65, 70),
        ("26 05 26", "Run bare copper bond + terminate", 0, "LF", 0.025, 70),
        ("26 05 26", "Cad-weld bond connection", 0, "EA", 0.85, 70),
        # Low voltage (if EC scope)
        ("27 11 13", "Pull Cat 6A cable + terminate jack", 0, "EA", 0.40, 65),
        ("27 13 23", "Pull fiber 12-strand multimode + terminate", 0, "EA", 6.00, 75),
        # Testing
        ("26 08 00", "Megger test feeder (1 hour minimum)", 0, "LS", 1.50, 70),
        ("26 08 00", "Ground resistance test (3-point fall-of-potential)", 0, "LS", 2.00, 70),
        ("26 08 00", "Acceptance test — panel/breaker exercise", 0, "EA", 0.50, 70),
    ]
    for i, (csi, act, qty, uom, hrs, rate) in enumerate(items):
        r = 4 + i
        ws.cell(row=r, column=1, value=csi).alignment = CENTER
        ws.cell(row=r, column=2, value=act).alignment = LEFT
        ws.cell(row=r, column=3, value=qty).number_format = FMT_INT
        ws.cell(row=r, column=4, value=uom).alignment = CENTER
        ws.cell(row=r, column=5, value=hrs).number_format = "0.000"
        ws.cell(row=r, column=6, value=f"=C{r}*E{r}").number_format = FMT_HRS
        ws.cell(row=r, column=7, value=rate).number_format = FMT_USD
        ws.cell(row=r, column=8, value=f"=F{r}*G{r}").number_format = FMT_USD
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = FONT_BODY
    last = 4 + len(items) - 1
    prod_row = last + 2
    ws.cell(row=prod_row, column=7, value="Productivity factor").font = FONT_BODY_BOLD
    ws.cell(row=prod_row, column=7).alignment = RIGHT
    ws.cell(row=prod_row, column=8, value=1.00).number_format = FMT_NUM2
    ws.cell(row=prod_row, column=8).fill = FILL_SUBHEADER
    sub_row = prod_row + 1
    ws.cell(row=sub_row, column=7, value="Labor Subtotal").font = FONT_BODY_BOLD
    ws.cell(row=sub_row, column=7).alignment = RIGHT
    ws.cell(row=sub_row, column=8, value=f"=SUM(H4:H{last})*H{prod_row}").number_format = FMT_USD
    ws.cell(row=sub_row, column=8).fill = FILL_GOLD
    ws.cell(row=sub_row, column=8).font = FONT_BODY_BOLD
    return ("EC_LaborTotal", f"'Labor Estimate'!$H${sub_row}")


def build_equipment(ws):
    ws.title = "Equipment"
    widths(ws, [("A", 42), ("B", 14), ("C", 14), ("D", 12), ("E", 16), ("F", 18)])
    title(ws, 1, "EQUIPMENT ESTIMATE", 6)
    header(ws, 3, ["Equipment", "Days", "Daily Rate ($)", "Weeks", "Weekly Rate ($)", "Subtotal ($)"])
    items = [
        ("Scissor lift — 19'", 0, 145.00, 0, 485.00),
        ("Boom lift — 45'", 0, 425.00, 0, 1485.00),
        ("Conduit bender — mechanical 1\"-2\"", 0, 65.00, 0, 195.00),
        ("Conduit bender — hydraulic 2-1/2\"-4\"", 0, 185.00, 0, 685.00),
        ("Knockout set (1/2\"-2\")", 0, 45.00, 0, 135.00),
        ("Knockout set — hydraulic 2-1/2\"-4\"", 0, 145.00, 0, 485.00),
        ("Wire-pulling equipment (tugger)", 0, 165.00, 0, 545.00),
        ("Megger / insulation tester", 0, 55.00, 0, 165.00),
        ("Ground resistance tester (3-point)", 0, 95.00, 0, 285.00),
        ("Fluke 87V / 1587 multimeter", 0, 35.00, 0, 105.00),
        ("Hot-stick set", 0, 95.00, 0, 285.00),
        ("Arc-flash PPE (Cat 2-4) per worker / week", 0, 0.00, 0, 145.00),
        ("Voltage detector / proximity meter", 0, 25.00, 0, 75.00),
        ("Trencher — walk-behind", 0, 285.00, 0, 1185.00),
        ("Concrete saw + core drill", 0, 145.00, 0, 525.00),
        ("Fiber fusion splicer", 0, 285.00, 0, 1185.00),
        ("Cat 6 cable tester (Fluke DSX)", 0, 145.00, 0, 485.00),
        ("Generator — 5kW portable", 0, 65.00, 0, 195.00),
        ("Office trailer (allocation)", 0, 0.00, 0, 285.00),
        ("Truck / van allocation", 0, 0.00, 0, 425.00),
    ]
    for i, (name, days, drate, weeks, wrate) in enumerate(items):
        r = 4 + i
        ws.cell(row=r, column=1, value=name).alignment = LEFT
        ws.cell(row=r, column=2, value=days).number_format = FMT_INT
        ws.cell(row=r, column=3, value=drate).number_format = FMT_USD
        ws.cell(row=r, column=4, value=weeks).number_format = FMT_INT
        ws.cell(row=r, column=5, value=wrate).number_format = FMT_USD
        ws.cell(row=r, column=6, value=f"=(B{r}*C{r})+(D{r}*E{r})").number_format = FMT_USD
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = FONT_BODY
    last = 4 + len(items) - 1
    sub_row = last + 2
    ws.cell(row=sub_row, column=5, value="Equipment Subtotal").alignment = RIGHT
    ws.cell(row=sub_row, column=5).font = FONT_BODY_BOLD
    ws.cell(row=sub_row, column=6, value=f"=SUM(F4:F{last})").number_format = FMT_USD
    ws.cell(row=sub_row, column=6).fill = FILL_GOLD
    ws.cell(row=sub_row, column=6).font = FONT_BODY_BOLD
    return ("EC_EquipTotal", f"'Equipment'!$F${sub_row}")


def build_subcontract(ws):
    ws.title = "Subcontract"
    widths(ws, [("A", 42), ("B", 22), ("C", 18), ("D", 16), ("E", 18)])
    title(ws, 1, "SUBCONTRACT ESTIMATE", 5)
    header(ws, 3, ["Scope (if subbed)", "Vendor / Quote", "Quote $", "Markup %", "Subtotal ($)"])
    items = [
        ("Fire-alarm cert + commissioning", "", 0.00, 0.10),
        ("Switchgear assembly + factory test", "", 0.00, 0.10),
        ("Generator startup + commissioning", "", 0.00, 0.10),
        ("ATS commissioning", "", 0.00, 0.10),
        ("Low-voltage cabling (if subbed)", "", 0.00, 0.10),
        ("Security system installation", "", 0.00, 0.10),
        ("Access-control install + commissioning", "", 0.00, 0.10),
        ("Audio-visual install", "", 0.00, 0.10),
        ("Excavation / boring (deep raceway)", "", 0.00, 0.10),
        ("Specialty grounding (counterpoise/MGB)", "", 0.00, 0.10),
        ("Arc-flash study (engineered)", "", 0.00, 0.10),
        ("Concrete saw-cutting / core drilling", "", 0.00, 0.10),
        ("Acceptance testing (NETA cert)", "", 0.00, 0.10),
    ]
    for i, (scope, vendor, quote, mk) in enumerate(items):
        r = 4 + i
        ws.cell(row=r, column=1, value=scope).alignment = LEFT
        ws.cell(row=r, column=2, value=vendor).alignment = LEFT
        ws.cell(row=r, column=3, value=quote).number_format = FMT_USD
        ws.cell(row=r, column=4, value=mk).number_format = FMT_PCT
        ws.cell(row=r, column=5, value=f"=C{r}*(1+D{r})").number_format = FMT_USD
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = FONT_BODY
    last = 4 + len(items) - 1
    sub_row = last + 2
    ws.cell(row=sub_row, column=4, value="Subcontract Subtotal").alignment = RIGHT
    ws.cell(row=sub_row, column=4).font = FONT_BODY_BOLD
    ws.cell(row=sub_row, column=5, value=f"=SUM(E4:E{last})").number_format = FMT_USD
    ws.cell(row=sub_row, column=5).fill = FILL_GOLD
    ws.cell(row=sub_row, column=5).font = FONT_BODY_BOLD
    return ("EC_SubTotal", f"'Subcontract'!$E${sub_row}")


def build_premiums(ws):
    ws.title = "Premiums"
    widths(ws, [("A", 50), ("B", 20), ("C", 18), ("D", 18)])
    title(ws, 1, "PREMIUMS & CONDITIONS", 4)
    header(ws, 3, ["Premium", "Applies? (Y/N)", "Adder %", "Adder $ (overrides %)"])
    items = [
        ("Prevailing wage (Davis-Bacon or state)", "N", 0.15, 0.00),
        ("Occupied building / live-rewire", "N", 0.12, 0.00),
        ("Hazardous (Class I Div 1) area", "N", 0.18, 0.00),
        ("Energized work / hot work permits", "N", 0.08, 0.00),
        ("Hospital ICRA tier 3+", "N", 0.08, 0.00),
        ("Heights > 14 ft (lift premium)", "N", 0.05, 0.00),
        ("Underground duct-bank concrete encasement", "N", 0.06, 0.00),
        ("Tight ceiling / plenum-rated cable upgrade", "N", 0.04, 0.00),
        ("After-hours / weekend / night shift", "N", 0.20, 0.00),
        ("Travel / per diem", "N", 0.03, 0.00),
        ("Cold-weather working conditions", "N", 0.04, 0.00),
        ("LEED documentation / Cx", "N", 0.03, 0.00),
        ("Owner-direct equipment (no markup)", "N", -0.10, 0.00),
    ]
    for i, (label, applies, pct, dollar) in enumerate(items):
        r = 4 + i
        ws.cell(row=r, column=1, value=label).alignment = LEFT
        ws.cell(row=r, column=2, value=applies).alignment = CENTER
        ws.cell(row=r, column=3, value=pct).number_format = FMT_PCT
        ws.cell(row=r, column=4, value=dollar).number_format = FMT_USD
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = FONT_BODY
    last = 4 + len(items) - 1
    total_row = last + 2
    ws.cell(row=total_row, column=1, value="Premium $ total (applies-Y only)").alignment = RIGHT
    ws.cell(row=total_row, column=1).font = FONT_BODY_BOLD
    ws.cell(row=total_row, column=4, value=f"=SUMPRODUCT((B4:B{last}=\"Y\")*D4:D{last})").number_format = FMT_USD
    ws.cell(row=total_row, column=4).fill = FILL_GOLD
    ws.cell(row=total_row, column=4).font = FONT_BODY_BOLD
    pct_row = total_row + 1
    ws.cell(row=pct_row, column=1, value="Premium % total (applies-Y only)").alignment = RIGHT
    ws.cell(row=pct_row, column=1).font = FONT_BODY_BOLD
    ws.cell(row=pct_row, column=4, value=f"=SUMPRODUCT((B4:B{last}=\"Y\")*C4:C{last})").number_format = FMT_PCT
    ws.cell(row=pct_row, column=4).fill = FILL_GOLD
    ws.cell(row=pct_row, column=4).font = FONT_BODY_BOLD
    return ("EC_PremiumDollars", f"'Premiums'!$D${total_row}"), ("EC_PremiumPct", f"'Premiums'!$D${pct_row}")


def build_summary(ws, refs):
    ws.title = "Bid Summary"
    widths(ws, [("A", 4), ("B", 38), ("C", 20), ("D", 22)])
    title(ws, 1, "BID SUMMARY — Electrical Trade Pack", 4)
    rows = [
        ("", "Material total (incl. tax)", f"={refs['EC_MaterialTotal']}"),
        ("", "Labor total", f"={refs['EC_LaborTotal']}"),
        ("", "Equipment total", f"={refs['EC_EquipTotal']}"),
        ("", "Subcontract total", f"={refs['EC_SubTotal']}"),
        ("", "Premiums $", f"={refs['EC_PremiumDollars']}"),
        ("", "Direct cost subtotal", "=SUM(C3:C7)"),
        ("", "Premium % applied", f"={refs['EC_PremiumPct']}"),
        ("", "Premium % adder $", "=C8*C9"),
        ("", "Direct + premium %", "=C8+C10"),
        ("", "Contingency %", 0.05),
        ("", "Contingency $", "=C11*C12"),
        ("", "Direct + contingency", "=C11+C13"),
        ("", "Overhead %", 0.10),
        ("", "Overhead $", "=C14*C15"),
        ("", "Profit %", 0.10),
        ("", "Profit $", "=(C14+C16)*C17"),
        ("", "Bond %", 0.012),
        ("", "Bond $", "=(C14+C16+C18)*C19"),
        ("", "Insurance / GL %", 0.018),
        ("", "Insurance $", "=(C14+C16+C18)*C21"),
        ("", "FINAL BID", "=C14+C16+C18+C20+C22"),
    ]
    for i, (a, label, val) in enumerate(rows):
        r = 3 + i
        ws.cell(row=r, column=2, value=label).font = FONT_BODY_BOLD
        ws.cell(row=r, column=2).alignment = RIGHT
        ws.cell(row=r, column=2).border = BORDER
        c = ws.cell(row=r, column=3, value=val)
        c.border = BORDER
        c.font = FONT_BODY_BOLD
        c.number_format = FMT_PCT if label.endswith("%") else FMT_USD
        if label == "FINAL BID":
            c.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
            ws.cell(row=r, column=2).fill = FILL_GOLD
            ws.cell(row=r, column=2).font = Font(name="Calibri", size=14, bold=True, color=BRAND_BLUE)


def build_csi(ws):
    ws.title = "CSI Reference"
    widths(ws, [("A", 14), ("B", 80)])
    title(ws, 1, "CSI MASTERFORMAT — DIVISION 26 ELECTRICAL", 2)
    rows = [
        ("26 00 00", "ELECTRICAL (overall division)"),
        ("26 05 00", "Common Work Results for Electrical"),
        ("26 05 19", "Low-Voltage Electrical Power Conductors and Cables"),
        ("26 05 26", "Grounding and Bonding for Electrical Systems"),
        ("26 05 29", "Hangers and Supports for Electrical Systems"),
        ("26 05 33", "Raceways and Boxes for Electrical Systems"),
        ("26 05 36", "Cable Trays for Electrical Systems"),
        ("26 05 43", "Underground Ducts and Raceways for Electrical Systems"),
        ("26 05 53", "Identification for Electrical Systems"),
        ("26 08 00", "Commissioning of Electrical Systems"),
        ("26 09 23", "Lighting Control Devices"),
        ("26 09 26", "Lighting Control Panelboards"),
        ("26 09 43", "Network Lighting Controls"),
        ("26 22 00", "Low-Voltage Transformers"),
        ("26 24 13", "Switchboards"),
        ("26 24 16", "Panelboards"),
        ("26 24 19", "Motor-Control Centers"),
        ("26 25 00", "Enclosed Bus Assemblies"),
        ("26 27 13", "Electricity Metering"),
        ("26 27 16", "Electrical Cabinets and Enclosures"),
        ("26 27 26", "Wiring Devices"),
        ("26 27 73", "Door Chimes / Specialty Devices"),
        ("26 28 13", "Fuses"),
        ("26 28 16", "Enclosed Switches and Circuit Breakers"),
        ("26 28 18", "Enclosed Circuit Breakers"),
        ("26 29 13", "Enclosed Controllers"),
        ("26 29 23", "Variable-Frequency Motor Controllers"),
        ("26 32 13", "Engine Generators"),
        ("26 36 23", "Automatic Transfer Switches"),
        ("26 41 13", "Lightning Protection for Structures"),
        ("26 43 13", "Surge Protective Devices"),
        ("26 51 13", "Interior Lighting Fixtures, Lamps, and Ballasts"),
        ("26 52 13", "Exterior Lighting Fixtures"),
        ("26 53 00", "Exit Signs"),
        ("26 56 00", "Sports / Stadium / Site Lighting"),
        ("27 11 13", "Communications Copper Cable (cross-ref)"),
        ("27 13 23", "Communications Optical Fiber (cross-ref)"),
        ("28 13 00", "Access Control (cross-ref)"),
        ("28 23 00", "Video Surveillance (cross-ref)"),
        ("28 31 00", "Fire Detection and Alarm (cross-ref)"),
    ]
    for i, (code, name) in enumerate(rows):
        r = 3 + i
        ws.cell(row=r, column=1, value=code).alignment = CENTER
        ws.cell(row=r, column=1).font = FONT_BODY_BOLD
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2, value=name).alignment = LEFT
        ws.cell(row=r, column=2).font = FONT_BODY
        ws.cell(row=r, column=2).border = BORDER


def main():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb = Workbook()
    build_instructions(wb.active)
    build_project_info(wb.create_sheet("Project Info"))
    build_wages(wb.create_sheet("Wage Rates"), wb)
    mat_ref = build_material(wb.create_sheet("Material Takeoff"))
    lab_ref = build_labor(wb.create_sheet("Labor Estimate"))
    eq_ref = build_equipment(wb.create_sheet("Equipment"))
    sub_ref = build_subcontract(wb.create_sheet("Subcontract"))
    prem_d, prem_p = build_premiums(wb.create_sheet("Premiums"))

    refs = {mat_ref[0]: mat_ref[1], lab_ref[0]: lab_ref[1], eq_ref[0]: eq_ref[1],
            sub_ref[0]: sub_ref[1], prem_d[0]: prem_d[1], prem_p[0]: prem_p[1]}
    ws_sum = wb.create_sheet("Bid Summary")
    build_summary(ws_sum, refs)
    wb.move_sheet(ws_sum, offset=-7)

    build_csi(wb.create_sheet("CSI Reference"))
    wb.save(OUT)
    print(f"OK — wrote {OUT}")


if __name__ == "__main__":
    main()
