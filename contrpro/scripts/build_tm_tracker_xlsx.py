#!/usr/bin/env python3
"""
Build ContrPro T&M (Time & Material) Tracker (XLSX) — Universal Sub Suite.

When changed work is authorized on a T&M basis (no agreed lump-sum price), the
sub must capture daily labor, equipment, materials, and lower-tier subcontract
costs with sufficient evidence to bill the GC and survive an audit. This
workbook is the per-job ledger that produces invoiceable T&M line items.

Tabs:
  1. Instructions
  2. Project Info        (markup rates, OT premium, fringe burden, prevailing wage flag)
  3. Labor Rate Schedule (named rates by classification — drives Labor Tickets via lookup)
  4. Labor Tickets       (1 row per worker per day per task)
  5. Equipment Tickets   (1 row per equipment-day)
  6. Material Tickets    (1 row per material delivery/consumption)
  7. Subcontract Tickets (lower-tier work performed under T&M)
  8. Invoice Roll-Up     (groups by ticket # and approval status, produces invoice lines)
  9. CSI Reference       (hidden)

Run:
    /Users/home/charles/.venv/bin/python3 \\
        /Users/home/charles/contrpro/scripts/build_tm_tracker_xlsx.py

Output:
    /Users/home/charles/contrpro/files/packages/complete/sub/TM_Tracker.xlsx
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# ---------------------------------------------------------------------------
# Brand / styling
# ---------------------------------------------------------------------------

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
YELLOW_FILL = "FFEB9C"
YELLOW_FONT = "9C5700"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_SUMMARY_LABEL = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW_FILL)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_BIG_NUMBER = Font(name="Calibri", size=18, bold=True, color=BRAND_BLUE)
FONT_GREY_ITALIC = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_PCT = "0.0%"
FMT_INT = "0"
FMT_NUM2 = "0.00"
FMT_DATE = "yyyy-mm-dd"

DIVISIONS = [
    ("01", "General Requirements"), ("02", "Existing Conditions"), ("03", "Concrete"),
    ("04", "Masonry"), ("05", "Metals"), ("06", "Wood, Plastics, and Composites"),
    ("07", "Thermal and Moisture Protection"), ("08", "Openings"), ("09", "Finishes"),
    ("10", "Specialties"), ("11", "Equipment"), ("12", "Furnishings"),
    ("13", "Special Construction"), ("14", "Conveying Equipment"), ("21", "Fire Suppression"),
    ("22", "Plumbing"), ("23", "HVAC"), ("25", "Integrated Automation"),
    ("26", "Electrical"), ("27", "Communications"), ("28", "Electronic Safety and Security"),
    ("31", "Earthwork"), ("32", "Exterior Improvements"), ("33", "Utilities"),
    ("34", "Transportation"), ("35", "Waterway and Marine Construction"),
]

# Default labor classifications & rates (placeholders the user edits)
DEFAULT_CLASSIFICATIONS = [
    ("Foreman", 52.00, 1.50),
    ("Journeyman", 42.00, 1.50),
    ("Apprentice — Year 4", 33.00, 1.50),
    ("Apprentice — Year 3", 28.00, 1.50),
    ("Apprentice — Year 2", 24.00, 1.50),
    ("Apprentice — Year 1", 20.00, 1.50),
    ("Helper / Laborer", 22.00, 1.50),
    ("Operating Engineer", 48.00, 1.50),
    ("Project Manager (T&M when authorized)", 65.00, 1.00),
    ("Superintendent (T&M when authorized)", 58.00, 1.00),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def set_col_widths(ws, widths):
    for col, w in widths:
        ws.column_dimensions[col].width = w


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


def apply_body_style(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_BODY
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.border = BORDER


# ---------------------------------------------------------------------------
# Tab 1: Instructions
# ---------------------------------------------------------------------------

def build_instructions(ws):
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 100)])

    ws["A1"] = "T&M TRACKER — Instructions"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 36

    rows = [
        "",
        "WHEN T&M APPLIES",
        "Time & Material billing is the default pricing method when changed Work is authorized without an "
        "agreed lump-sum price (a 'force-account' or 'directive' instruction). The GC's right to direct T&M "
        "work and the markup rates the sub may charge are governed by the Changes Article of the Subcontract "
        "and (on flow-down) the Prime Contract. Confirm both before invoicing.",
        "",
        "DAILY DISCIPLINE — THE FIVE INPUTS",
        "Each T&M day on the project must capture: (1) labor — every worker, classification, hours straight "
        "time and overtime, task; (2) equipment — every piece of equipment used and for how long; (3) "
        "materials — every material consumed or installed during the directed Work, with receipts; (4) "
        "subcontract — any lower-tier sub work performed under T&M; (5) GC field signature — a GC "
        "representative's same-day signature confirming the labor, equipment, and materials actually "
        "performed on directed Work. Without same-day signatures, T&M tickets are routinely disputed on "
        "billing.",
        "",
        "MARKUPS",
        "Industry-standard commercial markups, subject to Subcontract:",
        "  - Labor: cost + fringe burden + 10-15% overhead + 10% profit",
        "  - Material: cost + 10-15% markup (or as Subcontract specifies)",
        "  - Equipment (owned): published rental rate (e.g., RS Means or AED) with no further markup, OR "
        "internal rate + 10% (Subcontract controls)",
        "  - Equipment (rented): pass-through + 5-10% markup",
        "  - Subcontract: cost + 5-7% markup (lower than self-performed because no direct supervision risk)",
        "  - Total: rarely exceeds 25-30% all-in markup on costs",
        "",
        "FRINGE BURDEN",
        "Fringe burden = the cost of employing a worker beyond bare wages: Workers' Compensation insurance, "
        "FICA, FUTA, SUTA, health/welfare, pension, vacation, holiday, and (on prevailing wage projects) the "
        "fringe portion of the prevailing wage. Industry-standard fringe burden on commercial non-union "
        "work is 30-45% of bare wage; on union or prevailing-wage work it can be 60-100%. Set the rate on "
        "Project Info and the Labor Tickets tab applies it automatically.",
        "",
        "OVERTIME",
        "OT premium = the multiplier above straight time (typically 1.5× for daily OT, 2.0× for double time / "
        "Sunday / holiday in some jurisdictions). Confirm against the controlling collective-bargaining "
        "agreement or state wage law. Note: on Davis-Bacon / prevailing-wage work, the OT premium is "
        "calculated on the bare wage only; the fringe portion does not multiply.",
        "",
        "PREVAILING WAGE FLAG",
        "If the Project is subject to Davis-Bacon or state prevailing wage, the Labor Rate Schedule must "
        "match (at minimum) the prevailing wage determination for each classification. The 'Prevailing "
        "Wage?' flag on Project Info exists as a reminder; the actual wage determination is project-specific "
        "and must be loaded into the Labor Rate Schedule manually before T&M work begins.",
        "",
        "WORKFLOW — DAILY",
        "  1. Foreman fills paper T&M tickets in the field, GC field rep signs same-day.",
        "  2. End of shift: foreman enters tickets into the Labor / Equipment / Material / Subcontract tabs.",
        "  3. PM reviews against directives and Subcontract.",
        "  4. Tickets approved by GC roll up on the Invoice Roll-Up tab.",
        "  5. Monthly: produce invoice from Roll-Up for inclusion in the regular Pay App or as a "
        "separate Change-Order invoice (per Subcontract).",
        "",
        "DISPUTED TICKETS",
        "If the GC field rep refuses to sign a ticket, do not throw the ticket away. Mark it 'Unsigned' in "
        "the Status column. Send the ticket with photos and timestamps to the PM same-day. Preserve the "
        "claim per the Changes Article notice provisions. An unsigned ticket is still a claim — but only if "
        "you preserve it on time.",
        "",
        "AUDIT TRAIL",
        "Keep paper tickets bundled by week, scanned into the project file. The XLSX is the working ledger; "
        "the paper is the evidence. In a dispute, the GC's PM will demand the original signed paper.",
    ]

    for i, text in enumerate(rows, start=2):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if text and text == text.upper() and len(text) < 60:
            cell.font = FONT_H2
        else:
            cell.font = FONT_BODY


# ---------------------------------------------------------------------------
# Tab 2: Project Info
# ---------------------------------------------------------------------------

def build_project_info(ws):
    ws.title = "Project Info"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 42), ("C", 24)])

    ws["B1"] = "T&M PROJECT INFORMATION"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:C1")
    ws.row_dimensions[1].height = 32

    info_rows = [
        ("Project Name", ""),
        ("Subcontract No.", ""),
        ("GC", ""),
        ("Trade / Scope", ""),
        ("Subcontract Date", ""),
        ("", ""),
        ("MARKUP RATES — confirm against Subcontract", ""),
        ("Labor Overhead %", 0.15),
        ("Labor Profit %", 0.10),
        ("Material Markup %", 0.15),
        ("Equipment (Owned) Markup %", 0.10),
        ("Equipment (Rented) Markup %", 0.10),
        ("Subcontract Markup %", 0.07),
        ("", ""),
        ("LABOR — burdens and OT", ""),
        ("Fringe Burden % (on bare wage)", 0.35),
        ("OT Premium Multiplier (daily)", 1.50),
        ("DT Premium Multiplier (Sun / Holiday)", 2.00),
        ("Prevailing Wage Project?", "No"),
        ("", ""),
        ("APPROVAL", ""),
        ("Project Manager (Sub)", ""),
        ("Project Manager (GC)", ""),
    ]

    named = {}
    for i, (label, val) in enumerate(info_rows, start=3):
        ws.cell(row=i, column=2, value=label).font = FONT_BODY_BOLD if label and not label.startswith(("MARKUP", "LABOR", "APPROVAL")) else FONT_H2
        ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL if label and not label.startswith(("MARKUP", "LABOR", "APPROVAL")) else PatternFill()
        ws.cell(row=i, column=2).border = BORDER if label and not label.startswith(("MARKUP", "LABOR", "APPROVAL")) else Border()
        c = ws.cell(row=i, column=3, value=val)
        c.font = FONT_BODY
        c.border = BORDER if val != "" or (label and not label.startswith(("MARKUP", "LABOR", "APPROVAL"))) else Border()
        if isinstance(val, float):
            c.number_format = FMT_PCT if val < 1 else FMT_NUM2

    # Named ranges
    name_map = {
        "LaborOH": "$C$10",
        "LaborProfit": "$C$11",
        "MaterialMarkup": "$C$12",
        "EquipOwnedMarkup": "$C$13",
        "EquipRentedMarkup": "$C$14",
        "SubMarkup": "$C$15",
        "FringeBurden": "$C$18",
        "OTPremium": "$C$19",
        "DTPremium": "$C$20",
        "PrevailingWage": "$C$21",
    }
    for nm, ref in name_map.items():
        ws.parent.defined_names[nm] = DefinedName(name=nm, attr_text=f"'Project Info'!{ref}")

    dv_yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    dv_yn.add("C21")
    ws.add_data_validation(dv_yn)


# ---------------------------------------------------------------------------
# Tab 3: Labor Rate Schedule
# ---------------------------------------------------------------------------

LABOR_RATE_COLS = [
    ("A", "Classification", 38),
    ("B", "Bare Wage ($/hr)", 16),
    ("C", "OT Multiplier", 14),
    ("D", "Burdened Wage ($/hr)", 18),
    ("E", "Burdened OT ($/hr)", 18),
    ("F", "Billed ST Rate ($/hr)", 18),
    ("G", "Billed OT Rate ($/hr)", 18),
    ("H", "Notes", 30),
]


def build_labor_rate_schedule(ws):
    ws.title = "Labor Rates"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in LABOR_RATE_COLS])

    ws["A1"] = "LABOR RATE SCHEDULE"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Bare Wage = pre-burden hourly wage. Burdened = wage * (1 + fringe burden). Billed = burdened * "
        "(1 + LaborOH) * (1 + LaborProfit). On prevailing-wage projects, replace the default rates with "
        "the project-specific wage determination."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 32

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(LABOR_RATE_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(LABOR_RATE_COLS))

    for i, (cls, wage, ot) in enumerate(DEFAULT_CLASSIFICATIONS, start=HEADER_ROW + 1):
        ws.cell(row=i, column=1, value=cls)
        ws.cell(row=i, column=2, value=wage).number_format = FMT_USD
        ws.cell(row=i, column=3, value=ot).number_format = FMT_NUM2
        # Burdened ST
        ws.cell(row=i, column=4, value=f"=B{i}*(1+FringeBurden)").number_format = FMT_USD
        # Burdened OT — premium applies to bare wage only on prevailing wage; full on commercial
        ws.cell(
            row=i,
            column=5,
            value=(
                f'=IF(PrevailingWage="Yes",'
                f'(B{i}*C{i})+(B{i}*FringeBurden),'
                f'B{i}*C{i}*(1+FringeBurden))'
            ),
        ).number_format = FMT_USD
        # Billed ST = burdened * (1 + OH) * (1 + Profit)
        ws.cell(
            row=i,
            column=6,
            value=f"=D{i}*(1+LaborOH)*(1+LaborProfit)",
        ).number_format = FMT_USD
        # Billed OT
        ws.cell(
            row=i,
            column=7,
            value=f"=E{i}*(1+LaborOH)*(1+LaborProfit)",
        ).number_format = FMT_USD
        apply_body_style(ws, i, len(LABOR_RATE_COLS))

    # Named range for VLOOKUPs from Labor Tickets
    last_row = HEADER_ROW + len(DEFAULT_CLASSIFICATIONS)
    ws.parent.defined_names["LaborRates"] = DefinedName(
        name="LaborRates",
        attr_text=f"'Labor Rates'!$A${HEADER_ROW+1}:$G${last_row+10}",
    )


# ---------------------------------------------------------------------------
# Tab 4: Labor Tickets
# ---------------------------------------------------------------------------

LABOR_TICKET_COLS = [
    ("A", "Date", 12),
    ("B", "Ticket #", 10),
    ("C", "Worker Name", 22),
    ("D", "Classification", 26),
    ("E", "Task / CO Ref", 30),
    ("F", "ST Hours", 10),
    ("G", "OT Hours", 10),
    ("H", "Billed ST Rate", 14),
    ("I", "Billed OT Rate", 14),
    ("J", "Labor $ (ST)", 14),
    ("K", "Labor $ (OT)", 14),
    ("L", "Total Labor $", 14),
    ("M", "Status", 14),
    ("N", "GC Signature", 16),
    ("O", "Notes", 24),
]


def build_labor_tickets(ws):
    ws.title = "Labor Tickets"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in LABOR_TICKET_COLS])

    ws["A1"] = "LABOR TICKETS — Daily Force Account"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:O1")
    ws.row_dimensions[1].height = 32

    HEADER_ROW = 3
    for i, (col, header, _) in enumerate(LABOR_TICKET_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(LABOR_TICKET_COLS))

    DATA_ROWS = 200
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        ws.cell(row=r, column=1).number_format = FMT_DATE
        ws.cell(row=r, column=6).number_format = FMT_NUM2
        ws.cell(row=r, column=7).number_format = FMT_NUM2
        # Billed ST rate — VLOOKUP from Labor Rates (col F = billed ST)
        ws.cell(
            row=r,
            column=8,
            value=f'=IFERROR(VLOOKUP(D{r},LaborRates,6,FALSE),"")',
        ).number_format = FMT_USD
        # Billed OT rate — VLOOKUP col G
        ws.cell(
            row=r,
            column=9,
            value=f'=IFERROR(VLOOKUP(D{r},LaborRates,7,FALSE),"")',
        ).number_format = FMT_USD
        # Labor $ ST
        ws.cell(
            row=r,
            column=10,
            value=f'=IF(OR(F{r}="",H{r}=""),"",F{r}*H{r})',
        ).number_format = FMT_USD
        # Labor $ OT
        ws.cell(
            row=r,
            column=11,
            value=f'=IF(OR(G{r}="",I{r}=""),"",G{r}*I{r})',
        ).number_format = FMT_USD
        # Total labor
        ws.cell(
            row=r,
            column=12,
            value=f'=IF(J{r}="",K{r},IF(K{r}="",J{r},J{r}+K{r}))',
        ).number_format = FMT_USD
        apply_body_style(ws, r, len(LABOR_TICKET_COLS))

    # Validations
    dv_status = DataValidation(
        type="list",
        formula1='"Pending,Signed,Disputed,Approved,Invoiced,Paid,Unsigned"',
        allow_blank=True,
    )
    dv_status.add(f"M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_status)

    # Conditional format on Status
    ws.conditional_formatting.add(
        f"M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'M{HEADER_ROW+1}="Disputed"'], fill=FILL_RED, font=Font(color=RED_FONT)),
    )
    ws.conditional_formatting.add(
        f"M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'M{HEADER_ROW+1}="Unsigned"'], fill=FILL_YELLOW, font=Font(color=YELLOW_FONT)),
    )
    ws.conditional_formatting.add(
        f"M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'M{HEADER_ROW+1}="Approved"'], fill=FILL_GREEN, font=Font(color=GREEN_FONT)),
    )
    ws.conditional_formatting.add(
        f"M{HEADER_ROW+1}:M{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'M{HEADER_ROW+1}="Paid"'], fill=FILL_GREEN, font=Font(color=GREEN_FONT)),
    )

    # Totals row
    TOTAL_ROW = HEADER_ROW + 1 + DATA_ROWS + 1
    ws.cell(row=TOTAL_ROW, column=5, value="TOTALS").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=5).fill = FILL_SUBHEADER
    for col_idx, col_letter in [(6, "F"), (7, "G"), (10, "J"), (11, "K"), (12, "L")]:
        c = ws.cell(
            row=TOTAL_ROW,
            column=col_idx,
            value=f"=SUM({col_letter}{HEADER_ROW+1}:{col_letter}{HEADER_ROW+DATA_ROWS})",
        )
        c.font = FONT_BODY_BOLD
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_USD if col_idx >= 10 else FMT_NUM2
        c.border = BORDER


# ---------------------------------------------------------------------------
# Tab 5: Equipment Tickets
# ---------------------------------------------------------------------------

EQUIPMENT_COLS = [
    ("A", "Date", 12),
    ("B", "Ticket #", 10),
    ("C", "Equipment", 28),
    ("D", "Owned / Rented", 14),
    ("E", "Task / CO Ref", 28),
    ("F", "Hours", 10),
    ("G", "Hourly Rate ($)", 14),
    ("H", "Subtotal ($)", 14),
    ("I", "Markup", 12),
    ("J", "Billed Total ($)", 14),
    ("K", "Status", 14),
    ("L", "GC Signature", 16),
    ("M", "Notes", 24),
]


def build_equipment_tickets(ws):
    ws.title = "Equipment Tickets"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in EQUIPMENT_COLS])

    ws["A1"] = "EQUIPMENT TICKETS"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:M1")
    ws.row_dimensions[1].height = 32

    HEADER_ROW = 3
    for i, (col, header, _) in enumerate(EQUIPMENT_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(EQUIPMENT_COLS))

    DATA_ROWS = 120
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        ws.cell(row=r, column=1).number_format = FMT_DATE
        ws.cell(row=r, column=6).number_format = FMT_NUM2
        ws.cell(row=r, column=7).number_format = FMT_USD
        ws.cell(row=r, column=8, value=f'=IF(OR(F{r}="",G{r}=""),"",F{r}*G{r})').number_format = FMT_USD
        # Markup = owned or rented
        ws.cell(
            row=r,
            column=9,
            value=f'=IF(D{r}="Owned",EquipOwnedMarkup,IF(D{r}="Rented",EquipRentedMarkup,0))',
        ).number_format = FMT_PCT
        # Billed total = subtotal * (1 + markup)
        ws.cell(
            row=r,
            column=10,
            value=f'=IF(H{r}="","",H{r}*(1+I{r}))',
        ).number_format = FMT_USD
        apply_body_style(ws, r, len(EQUIPMENT_COLS))

    dv_owned = DataValidation(
        type="list",
        formula1='"Owned,Rented"',
        allow_blank=True,
    )
    dv_owned.add(f"D{HEADER_ROW+1}:D{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_owned)

    dv_status = DataValidation(
        type="list",
        formula1='"Pending,Signed,Disputed,Approved,Invoiced,Paid,Unsigned"',
        allow_blank=True,
    )
    dv_status.add(f"K{HEADER_ROW+1}:K{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_status)

    # Totals row
    TOTAL_ROW = HEADER_ROW + 1 + DATA_ROWS + 1
    ws.cell(row=TOTAL_ROW, column=5, value="TOTALS").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=5).fill = FILL_SUBHEADER
    for col_idx, col_letter in [(6, "F"), (8, "H"), (10, "J")]:
        c = ws.cell(
            row=TOTAL_ROW,
            column=col_idx,
            value=f"=SUM({col_letter}{HEADER_ROW+1}:{col_letter}{HEADER_ROW+DATA_ROWS})",
        )
        c.font = FONT_BODY_BOLD
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_USD if col_idx >= 8 else FMT_NUM2
        c.border = BORDER


# ---------------------------------------------------------------------------
# Tab 6: Material Tickets
# ---------------------------------------------------------------------------

MATERIAL_COLS = [
    ("A", "Date", 12),
    ("B", "Ticket #", 10),
    ("C", "Material / Description", 28),
    ("D", "Supplier", 20),
    ("E", "Invoice / PO #", 14),
    ("F", "CSI Div", 9),
    ("G", "Task / CO Ref", 22),
    ("H", "Quantity", 10),
    ("I", "Unit", 8),
    ("J", "Unit Cost ($)", 12),
    ("K", "Subtotal ($)", 12),
    ("L", "Markup", 10),
    ("M", "Billed ($)", 14),
    ("N", "Status", 14),
    ("O", "GC Signature", 16),
]


def build_material_tickets(ws):
    ws.title = "Material Tickets"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in MATERIAL_COLS])

    ws["A1"] = "MATERIAL TICKETS"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:O1")
    ws.row_dimensions[1].height = 32

    HEADER_ROW = 3
    for i, (col, header, _) in enumerate(MATERIAL_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(MATERIAL_COLS))

    DATA_ROWS = 150
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        ws.cell(row=r, column=1).number_format = FMT_DATE
        ws.cell(row=r, column=10).number_format = FMT_USD
        ws.cell(row=r, column=11, value=f'=IF(OR(H{r}="",J{r}=""),"",H{r}*J{r})').number_format = FMT_USD
        ws.cell(row=r, column=12, value="=MaterialMarkup").number_format = FMT_PCT
        ws.cell(row=r, column=13, value=f'=IF(K{r}="","",K{r}*(1+L{r}))').number_format = FMT_USD
        apply_body_style(ws, r, len(MATERIAL_COLS))

    # CSI dropdown
    div_codes = ",".join(code for code, _ in DIVISIONS)
    dv_div = DataValidation(type="list", formula1=f'"{div_codes}"', allow_blank=True)
    dv_div.add(f"F{HEADER_ROW+1}:F{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_div)

    dv_status = DataValidation(
        type="list",
        formula1='"Pending,Signed,Disputed,Approved,Invoiced,Paid,Unsigned"',
        allow_blank=True,
    )
    dv_status.add(f"N{HEADER_ROW+1}:N{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_status)

    TOTAL_ROW = HEADER_ROW + 1 + DATA_ROWS + 1
    ws.cell(row=TOTAL_ROW, column=7, value="TOTALS").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=7).fill = FILL_SUBHEADER
    for col_idx, col_letter in [(11, "K"), (13, "M")]:
        c = ws.cell(
            row=TOTAL_ROW,
            column=col_idx,
            value=f"=SUM({col_letter}{HEADER_ROW+1}:{col_letter}{HEADER_ROW+DATA_ROWS})",
        )
        c.font = FONT_BODY_BOLD
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_USD
        c.border = BORDER


# ---------------------------------------------------------------------------
# Tab 7: Subcontract Tickets
# ---------------------------------------------------------------------------

SUBCONTRACT_COLS = [
    ("A", "Date", 12),
    ("B", "Ticket #", 10),
    ("C", "Lower-Tier Sub", 26),
    ("D", "Description of Work", 32),
    ("E", "Task / CO Ref", 24),
    ("F", "Sub's Invoice $", 14),
    ("G", "Markup", 10),
    ("H", "Billed ($)", 14),
    ("I", "Status", 14),
    ("J", "GC Signature", 16),
    ("K", "Notes", 24),
]


def build_subcontract_tickets(ws):
    ws.title = "Subcontract Tickets"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in SUBCONTRACT_COLS])

    ws["A1"] = "LOWER-TIER SUBCONTRACT TICKETS"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:K1")
    ws.row_dimensions[1].height = 32

    HEADER_ROW = 3
    for i, (col, header, _) in enumerate(SUBCONTRACT_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(SUBCONTRACT_COLS))

    DATA_ROWS = 80
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        ws.cell(row=r, column=1).number_format = FMT_DATE
        ws.cell(row=r, column=6).number_format = FMT_USD
        ws.cell(row=r, column=7, value="=SubMarkup").number_format = FMT_PCT
        ws.cell(row=r, column=8, value=f'=IF(F{r}="","",F{r}*(1+G{r}))').number_format = FMT_USD
        apply_body_style(ws, r, len(SUBCONTRACT_COLS))

    dv_status = DataValidation(
        type="list",
        formula1='"Pending,Signed,Disputed,Approved,Invoiced,Paid,Unsigned"',
        allow_blank=True,
    )
    dv_status.add(f"I{HEADER_ROW+1}:I{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_status)

    TOTAL_ROW = HEADER_ROW + 1 + DATA_ROWS + 1
    ws.cell(row=TOTAL_ROW, column=5, value="TOTALS").font = FONT_BODY_BOLD
    ws.cell(row=TOTAL_ROW, column=5).fill = FILL_SUBHEADER
    for col_idx, col_letter in [(6, "F"), (8, "H")]:
        c = ws.cell(
            row=TOTAL_ROW,
            column=col_idx,
            value=f"=SUM({col_letter}{HEADER_ROW+1}:{col_letter}{HEADER_ROW+DATA_ROWS})",
        )
        c.font = FONT_BODY_BOLD
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_USD
        c.border = BORDER


# ---------------------------------------------------------------------------
# Tab 8: Invoice Roll-Up
# ---------------------------------------------------------------------------

def build_invoice_rollup(ws):
    ws.title = "Invoice Roll-Up"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 36), ("C", 22), ("D", 22), ("E", 22)])

    ws["B1"] = "T&M INVOICE ROLL-UP"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:E1")
    ws.row_dimensions[1].height = 32

    ws["B2"] = "Roll-up by status. Approved totals feed the invoice; Disputed / Unsigned tracked separately."
    ws["B2"].font = FONT_GREY_ITALIC
    ws.merge_cells("B2:E2")

    HEADER_ROW = 4
    headers = ["Category", "Approved", "Pending / Signed", "Disputed / Unsigned"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=HEADER_ROW, column=i, value=h)
    style_header_row(ws, HEADER_ROW, 4 + 1)  # B..E

    # Labor row
    ws.cell(row=5, column=2, value="Labor").font = FONT_BODY_BOLD
    ws.cell(row=5, column=3, value='=SUMIF(\'Labor Tickets\'!M:M,"Approved",\'Labor Tickets\'!L:L)+SUMIF(\'Labor Tickets\'!M:M,"Paid",\'Labor Tickets\'!L:L)+SUMIF(\'Labor Tickets\'!M:M,"Invoiced",\'Labor Tickets\'!L:L)').number_format = FMT_USD
    ws.cell(row=5, column=4, value='=SUMIF(\'Labor Tickets\'!M:M,"Pending",\'Labor Tickets\'!L:L)+SUMIF(\'Labor Tickets\'!M:M,"Signed",\'Labor Tickets\'!L:L)').number_format = FMT_USD
    ws.cell(row=5, column=5, value='=SUMIF(\'Labor Tickets\'!M:M,"Disputed",\'Labor Tickets\'!L:L)+SUMIF(\'Labor Tickets\'!M:M,"Unsigned",\'Labor Tickets\'!L:L)').number_format = FMT_USD

    # Equipment row
    ws.cell(row=6, column=2, value="Equipment").font = FONT_BODY_BOLD
    ws.cell(row=6, column=3, value='=SUMIF(\'Equipment Tickets\'!K:K,"Approved",\'Equipment Tickets\'!J:J)+SUMIF(\'Equipment Tickets\'!K:K,"Paid",\'Equipment Tickets\'!J:J)+SUMIF(\'Equipment Tickets\'!K:K,"Invoiced",\'Equipment Tickets\'!J:J)').number_format = FMT_USD
    ws.cell(row=6, column=4, value='=SUMIF(\'Equipment Tickets\'!K:K,"Pending",\'Equipment Tickets\'!J:J)+SUMIF(\'Equipment Tickets\'!K:K,"Signed",\'Equipment Tickets\'!J:J)').number_format = FMT_USD
    ws.cell(row=6, column=5, value='=SUMIF(\'Equipment Tickets\'!K:K,"Disputed",\'Equipment Tickets\'!J:J)+SUMIF(\'Equipment Tickets\'!K:K,"Unsigned",\'Equipment Tickets\'!J:J)').number_format = FMT_USD

    # Materials row
    ws.cell(row=7, column=2, value="Materials").font = FONT_BODY_BOLD
    ws.cell(row=7, column=3, value='=SUMIF(\'Material Tickets\'!N:N,"Approved",\'Material Tickets\'!M:M)+SUMIF(\'Material Tickets\'!N:N,"Paid",\'Material Tickets\'!M:M)+SUMIF(\'Material Tickets\'!N:N,"Invoiced",\'Material Tickets\'!M:M)').number_format = FMT_USD
    ws.cell(row=7, column=4, value='=SUMIF(\'Material Tickets\'!N:N,"Pending",\'Material Tickets\'!M:M)+SUMIF(\'Material Tickets\'!N:N,"Signed",\'Material Tickets\'!M:M)').number_format = FMT_USD
    ws.cell(row=7, column=5, value='=SUMIF(\'Material Tickets\'!N:N,"Disputed",\'Material Tickets\'!M:M)+SUMIF(\'Material Tickets\'!N:N,"Unsigned",\'Material Tickets\'!M:M)').number_format = FMT_USD

    # Subcontract row
    ws.cell(row=8, column=2, value="Subcontract").font = FONT_BODY_BOLD
    ws.cell(row=8, column=3, value='=SUMIF(\'Subcontract Tickets\'!I:I,"Approved",\'Subcontract Tickets\'!H:H)+SUMIF(\'Subcontract Tickets\'!I:I,"Paid",\'Subcontract Tickets\'!H:H)+SUMIF(\'Subcontract Tickets\'!I:I,"Invoiced",\'Subcontract Tickets\'!H:H)').number_format = FMT_USD
    ws.cell(row=8, column=4, value='=SUMIF(\'Subcontract Tickets\'!I:I,"Pending",\'Subcontract Tickets\'!H:H)+SUMIF(\'Subcontract Tickets\'!I:I,"Signed",\'Subcontract Tickets\'!H:H)').number_format = FMT_USD
    ws.cell(row=8, column=5, value='=SUMIF(\'Subcontract Tickets\'!I:I,"Disputed",\'Subcontract Tickets\'!H:H)+SUMIF(\'Subcontract Tickets\'!I:I,"Unsigned",\'Subcontract Tickets\'!H:H)').number_format = FMT_USD

    # Total row
    ws.cell(row=9, column=2, value="TOTAL").font = FONT_BODY_BOLD
    ws.cell(row=9, column=2).fill = FILL_SUBHEADER
    for col in [3, 4, 5]:
        col_letter = get_column_letter(col)
        c = ws.cell(row=9, column=col, value=f"=SUM({col_letter}5:{col_letter}8)")
        c.font = FONT_BIG_NUMBER
        c.fill = FILL_SUBHEADER
        c.number_format = FMT_USD
        c.border = BORDER

    for r in range(5, 10):
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = BORDER

    # Invoice next steps
    ws["B12"] = "INVOICE NEXT STEPS"
    ws["B12"].font = FONT_H2
    notes = [
        "1. Confirm Approved column totals match the signed-and-approved tickets.",
        "2. Disputed / Unsigned column is the working dispute file; track preservation-of-claim notice dates.",
        "3. Build the T&M invoice from Approved totals only. Pending / Signed remains on the next month's invoice.",
        "4. After invoice is sent, change ticket statuses from Approved → Invoiced.",
        "5. After payment is received, change ticket statuses from Invoiced → Paid.",
    ]
    for i, t in enumerate(notes, start=13):
        ws.cell(row=i, column=2, value=t).font = FONT_BODY
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)


# ---------------------------------------------------------------------------
# Tab 9: CSI Reference (hidden)
# ---------------------------------------------------------------------------

def build_csi_reference(ws):
    ws.title = "CSI Reference"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 10), ("B", 60)])
    ws["A1"] = "CSI MasterFormat Divisions"
    ws["A1"].font = FONT_H2
    ws["A3"] = "Div"
    ws["B3"] = "Name"
    for c in [ws["A3"], ws["B3"]]:
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
    for i, (code, name) in enumerate(DIVISIONS, start=4):
        ws.cell(row=i, column=1, value=code).font = FONT_BODY
        ws.cell(row=i, column=2, value=name).font = FONT_BODY
    ws.sheet_state = "hidden"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    wb = Workbook()
    build_instructions(wb.active)
    build_project_info(wb.create_sheet())
    build_labor_rate_schedule(wb.create_sheet())
    build_labor_tickets(wb.create_sheet())
    build_equipment_tickets(wb.create_sheet())
    build_material_tickets(wb.create_sheet())
    build_subcontract_tickets(wb.create_sheet())
    build_invoice_rollup(wb.create_sheet())
    build_csi_reference(wb.create_sheet())

    out_dir = "/Users/home/charles/contrpro/files/packages/complete/sub"
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "TM_Tracker.xlsx")
    wb.save(out_path)
    sz = os.path.getsize(out_path)
    print(f"Wrote {out_path} ({sz//1024} KB)")


if __name__ == "__main__":
    main()
