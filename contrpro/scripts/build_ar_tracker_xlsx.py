#!/usr/bin/env python3
"""
Build ContrPro Accounts Receivable (AR) Tracker (XLSX).

Produces a production-grade workbook with:
  - Instructions tab (aging buckets, retainage, lien waiver workflow, escalation)
  - Project Info tab (contract anchor)
  - Invoice Log tab (CSI-coded, dropdowns, aging formulas, conditional formatting)
  - Aging Summary tab (SUMIFS roll-ups by aging bucket)
  - Dashboard tab (totals, DSO, retainage, aging panel, status counts)
  - CSI Reference tab (hidden, named ranges used by VLOOKUP / dropdowns)

Run:
    /Users/home/charles/.venv/bin/python3 \
        /Users/home/charles/contrpro/scripts/build_ar_tracker_xlsx.py

Output:
    /Users/home/charles/contrpro/files/packages/business/AR_Tracker.xlsx
"""

from __future__ import annotations

import os
from typing import Tuple

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# ---------------------------------------------------------------------------
# Brand & styling (matches Job_Costing_Spreadsheet & Change_Order_Log)
# ---------------------------------------------------------------------------

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
ACCENT_GOLD = "C9A227"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
YELLOW_FILL = "FFEB9C"
YELLOW_FONT = "9C5700"
ORANGE_FILL = "FFD8A8"
ORANGE_FONT = "8A4B00"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_SUMMARY_LABEL = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW_FILL)
FILL_ORANGE = PatternFill("solid", fgColor=ORANGE_FILL)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H1 = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_BIG_NUMBER = Font(name="Calibri", size=18, bold=True, color=BRAND_BLUE)
FONT_GREY_ITALIC = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_USD_BIG = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
FMT_PCT = "0.0%"
FMT_INT = "0"
FMT_DATE = "yyyy-mm-dd"


# ---------------------------------------------------------------------------
# CSI MasterFormat division reference (same source as other trackers)
# ---------------------------------------------------------------------------

DIVISIONS: list[tuple[str, str]] = [
    ("01", "General Requirements"),
    ("02", "Existing Conditions"),
    ("03", "Concrete"),
    ("04", "Masonry"),
    ("05", "Metals"),
    ("06", "Wood, Plastics, and Composites"),
    ("07", "Thermal and Moisture Protection"),
    ("08", "Openings"),
    ("09", "Finishes"),
    ("10", "Specialties"),
    ("11", "Equipment"),
    ("12", "Furnishings"),
    ("13", "Special Construction"),
    ("14", "Conveying Equipment"),
    ("21", "Fire Suppression"),
    ("22", "Plumbing"),
    ("23", "HVAC"),
    ("25", "Integrated Automation"),
    ("26", "Electrical"),
    ("27", "Communications"),
    ("28", "Electronic Safety and Security"),
    ("31", "Earthwork"),
    ("32", "Exterior Improvements"),
    ("33", "Utilities"),
    ("34", "Transportation"),
    ("35", "Waterway and Marine Construction"),
    ("40", "Process Interconnections"),
    ("41", "Material Processing and Handling Equipment"),
    ("42", "Process Heating, Cooling, and Drying Equipment"),
    ("43", "Process Gas and Liquid Handling, Purification, and Storage Equipment"),
    ("44", "Pollution and Waste Control Equipment"),
    ("45", "Industry-Specific Manufacturing Equipment"),
    ("46", "Water and Wastewater Equipment"),
    ("48", "Electrical Power Generation"),
    ("49", "Reserved"),
]
DIV_LOOKUP: dict[str, str] = {d: n for d, n in DIVISIONS}


# ---------------------------------------------------------------------------
# Pre-populated example invoices (varied aging states for showcase)
#
# Anchor today = 2026-05-13 (today's date in the build environment).
#  - INV-1001: paid in full (Paid bucket)
#  - INV-1002: due 2026-05-25, not yet due (Current bucket)
#  - INV-1003: due 2026-04-05, 38 days late (31-60 bucket)
#  - INV-1004: due 2026-03-01, 73 days late (61-90 bucket)
#  - INV-1005: due 2026-01-10, 123 days late (120+ bucket)
#
# (inv_num, date_issued, customer, div, csi_code, description,
#  base_amount, retainage_withheld, date_due, date_paid, amount_paid,
#  lien_waiver, status, notes)
# ---------------------------------------------------------------------------

EXAMPLE_INVOICES: list[tuple] = [
    ("INV-1001", "2026-03-20",
     "Riverside Holdings LLC", "03", "03 30 00",
     "Pay App #2 - Cast-in-place concrete, footings & SOG.",
     48500.00, 2425.00,
     "2026-04-19", "2026-04-17", 48500.00,
     "Yes", "Paid",
     "Paid in full. Unconditional lien waiver on file."),
    ("INV-1002", "2026-04-25",
     "Riverside Holdings LLC", "05", "05 12 00",
     "Pay App #3 - Structural steel erection, Level 1-2.",
     62300.00, 3115.00,
     "2026-05-25", "", 0.00,
     "Conditional", "Sent",
     "Pay app submitted. Conditional waiver issued; final on receipt."),
    ("INV-1003", "2026-03-06",
     "Pinnacle Dev Corp", "23", "23 30 00",
     "Pay App #1 - HVAC rough-in, Level 2 air distribution.",
     28750.00, 1437.50,
     "2026-04-05", "", 0.00,
     "No", "Partially Paid",
     "Owner remitted $15k against this invoice; $13,750 outstanding. Follow-up email sent 2026-04-22."),
    ("INV-1004", "2026-01-30",
     "Heartland Builders", "26", "26 50 00",
     "Pay App #4 - Lighting fixtures & emergency egress install.",
     19400.00, 970.00,
     "2026-03-01", "", 0.00,
     "No", "Disputed",
     "Owner disputes 6 fixture count. Demand letter draft prepared; escalate to attorney week of 5/19."),
    ("INV-1005", "2025-12-11",
     "Apex Industrial", "07", "07 50 00",
     "Pay App #2 - Membrane roofing, 32 sq install.",
     14200.00, 0.00,
     "2026-01-10", "", 0.00,
     "No", "Disputed",
     "120+ days. Notice of intent to lien filed 2026-04-15. Lien deadline tracking - confirm state statute window."),
]

EMPTY_INVOICE_ROWS = 25
EMPTY_PROJECT_ROWS = 5

STATUSES = ["Sent", "Partially Paid", "Paid", "Disputed", "Written Off"]
LIEN_WAIVERS = ["Yes", "No", "Conditional", "N/A"]
AGING_BUCKETS = ["Current", "1-30", "31-60", "61-90", "91-120", "120+", "Paid"]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def set_col_widths(ws, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_header_row(ws, row: int, last_col: int) -> None:
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 32


def border_range(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER


# ---------------------------------------------------------------------------
# Hidden tab — CSI Reference (also feeds named ranges + dropdowns)
# ---------------------------------------------------------------------------

def build_csi_reference(wb: Workbook) -> str:
    ws = wb.create_sheet("CSI Reference")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "CSI MasterFormat Division Reference"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:B1")

    ws["A2"] = "Pulled from the same canonical list as ContrPro Job Costing. Do not edit."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)

    ws.cell(row=3, column=1, value="CSI Division")
    ws.cell(row=3, column=2, value="Division Name")
    style_header_row(ws, 3, 2)

    for i, (code, name) in enumerate(DIVISIONS, start=4):
        ws.cell(row=i, column=1, value=code).number_format = "@"
        ws.cell(row=i, column=2, value=name)
        for c in (1, 2):
            ws.cell(row=i, column=c).border = BORDER
            ws.cell(row=i, column=c).font = FONT_BODY

    set_col_widths(ws, {"A": 14, "B": 60})

    first = 4
    last = 4 + len(DIVISIONS) - 1
    div_range = f"'CSI Reference'!$A${first}:$A${last}"
    full_range = f"'CSI Reference'!$A${first}:$B${last}"
    wb.defined_names["CSI_Divisions"] = DefinedName("CSI_Divisions", attr_text=div_range)
    wb.defined_names["CSI_Table"] = DefinedName("CSI_Table", attr_text=full_range)
    return ws.title


# ---------------------------------------------------------------------------
# Tab 1 — Instructions
# ---------------------------------------------------------------------------

def build_instructions(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {"A": 4, "B": 110})

    ws["B2"] = "ContrPro - Accounts Receivable (AR) Tracker"
    ws["B2"].font = FONT_TITLE
    ws.row_dimensions[2].height = 32

    ws["B3"] = ("CSI MasterFormat-coded invoice tracker - who owes you what, "
                "how old it is, and when to start pulling on the line.")
    ws["B3"].font = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

    sections: list[tuple[str, list[str]]] = [
        (
            "What this workbook is",
            [
                "Five visible tabs plus a hidden CSI Reference. Project Info pins the contract, Invoice Log is the "
                "main worksheet, Aging Summary buckets every dollar by how late it is, and Dashboard shows the "
                "headline numbers (Total Outstanding, DSO, retainage held).",
                "Every invoice you cut goes in the Invoice Log. The aging math is automatic - you fill in dates and "
                "amounts, the workbook tells you who is current, who is 30 days late, and who is past 90 and burning.",
            ],
        ),
        (
            "Aging buckets - the cadence",
            [
                "Current   - not yet past due. Healthy. Do nothing.",
                "1-30      - 1 to 30 days past due. Standard reminder call/email. Most owners pay in this window.",
                "31-60     - 31 to 60 days past due. Escalate to the PM and copy the owner's accountant.",
                "61-90     - 61 to 90 days past due. CFO-level conversation. Stop further work if contract allows.",
                "91-120    - 91 to 120 days past due. Formal demand letter. Pre-lien notices in many states.",
                "120+      - over 120 days. Mechanic's lien window or attorney referral. Statute clock is ticking.",
            ],
        ),
        (
            "Retainage vs base receivable - keep them separate",
            [
                "Base Amount is the dollar value of work performed this pay period. That is the number the owner owes "
                "you against this invoice on the due date.",
                "Retainage Withheld is the percentage (usually 5-10%) the owner holds back per the contract. It is "
                "NOT past due until substantial completion - tracked separately so it does not pollute your aging.",
                "Total Invoice = Base Amount. Retainage is logged for visibility only; it gets released at the end of "
                "the job via a separate retainage invoice.",
            ],
        ),
        (
            "Lien waiver workflow",
            [
                "Conditional lien waiver - issue WITH the invoice. Says 'I waive my lien rights conditional on receipt "
                "of payment.' This is what owners require to release the check.",
                "Unconditional lien waiver - issue ONLY after the check has cleared. Says 'paid in full, lien rights "
                "waived.' Never issue unconditional before the funds settle - this is how contractors lose lien rights.",
                "Lien Waiver Received? column tracks what waivers YOUR subs have given YOU. Yes = unconditional in hand. "
                "Conditional = conditional on file. No = open exposure. N/A = self-performed.",
            ],
        ),
        (
            "When to escalate",
            [
                "60+ days past due: stop and call. PM-to-PM conversation. Document the call in Notes with date and who "
                "you talked to. If the answer is 'next week,' put a follow-up on your calendar.",
                "90+ days past due: written demand letter. Reference the contract pay terms, the invoice number, and "
                "the amount. Copy the owner principal, not just the PM.",
                "120+ days past due: pre-lien notice if your state requires one (most do; window is typically 60-120 "
                "days from last work). Then mechanic's lien before the statute window closes (commonly 90 days from "
                "last work, but varies by state). This is attorney territory - do not freelance it.",
            ],
        ),
        (
            "Days Sales Outstanding (DSO)",
            [
                "DSO is the average number of days it takes you to collect after issuing an invoice. The Dashboard "
                "computes it as a dollar-weighted average across all outstanding invoices.",
                "Construction industry benchmark: 60-75 days is normal for commercial GCs, 30-45 days is normal for "
                "residential. DSO above 90 means you're financing the project for the owner - margin killer.",
                "Bring DSO down by: invoicing on day-1 of each cycle, issuing conditional waivers at submission, and "
                "calling on the 1st late day rather than the 30th.",
            ],
        ),
        (
            "Tab-by-tab quick start",
            [
                "1. Project Info - fill in Project Name, Number, Client, Contract Amount, PM, and notes on row 4.",
                "2. Invoice Log - one row per invoice. Fill in Invoice Number, dates, customer, CSI Division "
                "(dropdown), CSI Code, description, Base Amount, Retainage Withheld, Date Due, Date Paid, Amount "
                "Paid. Outstanding, Days Outstanding, and Aging Bucket calculate automatically.",
                "3. Aging Summary - DO NOT edit. Rolls up Invoice Log by bucket.",
                "4. Dashboard - DO NOT edit. Read-only headline numbers and aging panel.",
            ],
        ),
        (
            "CSI coding rules (must follow)",
            [
                "CSI Division: 2-digit text (01-49). Pick from the dropdown - leading zeros preserved.",
                "CSI Code: 6-digit space-separated, e.g. 23 30 00 for HVAC Air Distribution. Level 3 OK (23 31 13).",
                "Division Name auto-fills from a VLOOKUP against CSI Reference. If it shows blank/error, your CSI "
                "Division value does not match the reference list - re-pick from the dropdown.",
            ],
        ),
        (
            "Color coding cheat sheet",
            [
                "Aging Bucket: green = Current/Paid, yellow = 1-30/31-60, orange = 61-90, red = 91-120 and 120+.",
                "Outstanding: red fill when amount > 0 AND aging is past 60 days - your collection priority queue.",
                "Status: green = Paid, yellow = Sent/Partially Paid, red = Disputed/Written Off.",
            ],
        ),
        (
            "Weekly cadence",
            [
                "Monday morning: open Dashboard, sort Invoice Log by Days Outstanding desc, call/email anything 14+ days late.",
                "Friday afternoon: log every invoice issued this week and every payment received. Reconcile against the bank.",
                "End of month: tie the Total Outstanding figure to your accounting system AR aging report. They must match.",
            ],
        ),
        (
            "Tabs in this workbook",
            [
                "Instructions  -  this tab",
                "Project Info  -  contract anchor",
                "Invoice Log  -  CSI-coded invoice line items, the main worksheet",
                "Aging Summary  -  SUMIFS roll-up by aging bucket",
                "Dashboard  -  headline AR numbers, DSO, retainage, status counts",
                "CSI Reference  -  canonical division list (hidden - do not edit)",
            ],
        ),
    ]

    row = 5
    for heading, paragraphs in sections:
        ws.cell(row=row, column=2, value=heading)
        ws.cell(row=row, column=2).font = FONT_H2
        ws.cell(row=row, column=2).fill = FILL_SUBHEADER
        ws.cell(row=row, column=2).alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[row].height = 22
        row += 1
        for para in paragraphs:
            ws.cell(row=row, column=2, value=para)
            ws.cell(row=row, column=2).font = FONT_BODY
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = max(18, 16 * (1 + len(para) // 110))
            row += 1
        row += 1

    ws.cell(row=row + 1, column=2,
            value="ContrPro - built for builders. Questions: support@contrpro.com")
    ws.cell(row=row + 1, column=2).font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)


# ---------------------------------------------------------------------------
# Tab 2 — Project Info
# ---------------------------------------------------------------------------

PROJECT_HEADERS = [
    "Project Name", "Project Number", "Client",
    "Contract Amount", "Project Manager", "Notes",
]

PROJECT_EXAMPLE = [
    "EXAMPLE - Riverside Office Bldg",
    "2026-014",
    "Riverside Holdings LLC",
    1850000,
    "Sample PM",
    "Example row - replace with your project.",
]


def build_project_info(wb: Workbook) -> None:
    ws = wb.create_sheet("Project Info")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Project Information"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:F1")

    for i, h in enumerate(PROJECT_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(PROJECT_HEADERS))

    # Example row (greyed/italic)
    for i, v in enumerate(PROJECT_EXAMPLE, start=1):
        cell = ws.cell(row=4, column=i, value=v)
        cell.font = FONT_GREY_ITALIC
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.cell(row=4, column=4).number_format = FMT_USD

    # Empty user rows
    for r in range(5, 5 + EMPTY_PROJECT_ROWS):
        for c in range(1, len(PROJECT_HEADERS) + 1):
            ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=4).number_format = FMT_USD
        ws.row_dimensions[r].height = 22

    set_col_widths(ws, {
        "A": 30, "B": 16, "C": 28, "D": 18, "E": 22, "F": 45,
    })

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Tab 3 — Invoice Log (main worksheet)
# ---------------------------------------------------------------------------

INVOICE_HEADERS = [
    "Invoice Number", "Date Issued", "Customer / Job",
    "CSI Division", "CSI Code", "Division Name",
    "Description",
    "Base Amount", "Retainage Withheld", "Total Invoice",
    "Date Due", "Date Paid", "Amount Paid",
    "Outstanding", "Aging Bucket", "Days Outstanding",
    "Lien Waiver Received?", "Status", "Notes",
]


def build_invoice_log(wb: Workbook) -> Tuple[str, int, int]:
    ws = wb.create_sheet("Invoice Log")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Invoice Log (CSI MasterFormat)"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(INVOICE_HEADERS))

    # Header row at row 3
    for i, h in enumerate(INVOICE_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(INVOICE_HEADERS))

    data_start = 4
    n_pre = len(EXAMPLE_INVOICES)
    n_total = n_pre + EMPTY_INVOICE_ROWS
    data_end = data_start + n_total - 1

    # Column letter map
    COL = {h: get_column_letter(i + 1) for i, h in enumerate(INVOICE_HEADERS)}
    col_div = COL["CSI Division"]            # D
    col_base = COL["Base Amount"]            # H
    col_retain = COL["Retainage Withheld"]   # I
    col_total = COL["Total Invoice"]         # J
    col_due = COL["Date Due"]                # K
    col_paid_date = COL["Date Paid"]         # L
    col_amt_paid = COL["Amount Paid"]        # M
    col_outstanding = COL["Outstanding"]      # N
    col_aging = COL["Aging Bucket"]           # O
    col_days = COL["Days Outstanding"]        # P
    col_status = COL["Status"]                # R

    def total_formula(r: int) -> str:
        # Total Invoice = Base Amount (retainage is held aside, logged for visibility)
        return f"=IFERROR({col_base}{r},0)"

    def outstanding_formula(r: int) -> str:
        return f"=IFERROR({col_total}{r}-IFERROR({col_amt_paid}{r},0),0)"

    def aging_formula(r: int) -> str:
        # If fully paid (Amount Paid >= Total Invoice and Date Paid present) -> Paid
        # Else compute days = TODAY()-DateDue.
        #   days <= 0     -> Current
        #   1..30         -> 1-30
        #   31..60        -> 31-60
        #   61..90        -> 61-90
        #   91..120       -> 91-120
        #   >120          -> 120+
        # If DateDue blank -> ""
        return (
            f'=IF(AND({col_paid_date}{r}<>"",IFERROR({col_amt_paid}{r},0)>=IFERROR({col_total}{r},0),'
            f'IFERROR({col_total}{r},0)>0),"Paid",'
            f'IF({col_due}{r}="","",'
            f'IF(TODAY()-{col_due}{r}<=0,"Current",'
            f'IF(TODAY()-{col_due}{r}<=30,"1-30",'
            f'IF(TODAY()-{col_due}{r}<=60,"31-60",'
            f'IF(TODAY()-{col_due}{r}<=90,"61-90",'
            f'IF(TODAY()-{col_due}{r}<=120,"91-120","120+")))))))'
        )

    def days_outstanding_formula(r: int) -> str:
        # Days = TODAY() - DateDue when outstanding > 0; else blank
        return (
            f'=IF(OR({col_due}{r}="",IFERROR({col_outstanding}{r},0)<=0),"",'
            f'MAX(0,TODAY()-{col_due}{r}))'
        )

    def divname_formula(r: int) -> str:
        return f'=IFERROR(VLOOKUP({col_div}{r},CSI_Table,2,FALSE),"")'

    # Pre-populated example rows
    for idx, row_data in enumerate(EXAMPLE_INVOICES):
        r = data_start + idx
        (inv_num, d_issued, customer, div, csi_code, desc,
         base, retain, d_due, d_paid, amt_paid,
         lien, status, notes) = row_data

        ws.cell(row=r, column=1, value=inv_num)
        ws.cell(row=r, column=2, value=d_issued)
        ws.cell(row=r, column=3, value=customer)
        ws.cell(row=r, column=4, value=div)
        ws.cell(row=r, column=5, value=csi_code)
        ws.cell(row=r, column=6, value=divname_formula(r))
        ws.cell(row=r, column=7, value=desc)
        ws.cell(row=r, column=8, value=base)
        ws.cell(row=r, column=9, value=retain)
        ws.cell(row=r, column=10, value=total_formula(r))
        ws.cell(row=r, column=11, value=d_due)
        ws.cell(row=r, column=12, value=d_paid if d_paid else None)
        ws.cell(row=r, column=13, value=amt_paid if amt_paid else 0)
        ws.cell(row=r, column=14, value=outstanding_formula(r))
        ws.cell(row=r, column=15, value=aging_formula(r))
        ws.cell(row=r, column=16, value=days_outstanding_formula(r))
        ws.cell(row=r, column=17, value=lien)
        ws.cell(row=r, column=18, value=status)
        ws.cell(row=r, column=19, value=notes)

    # Empty user-fillable rows (formulas pre-seeded)
    for i in range(EMPTY_INVOICE_ROWS):
        r = data_start + n_pre + i
        ws.cell(row=r, column=6, value=divname_formula(r))
        ws.cell(row=r, column=10, value=total_formula(r))
        ws.cell(row=r, column=14, value=outstanding_formula(r))
        ws.cell(row=r, column=15, value=aging_formula(r))
        ws.cell(row=r, column=16, value=days_outstanding_formula(r))

    # Per-row formatting
    wrap_cols = {3, 6, 7, 19}
    for r in range(data_start, data_end + 1):
        for c in range(1, len(INVOICE_HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(c in wrap_cols),
            )
        # Text columns (preserve leading zeros, invoice IDs)
        ws.cell(row=r, column=1).number_format = "@"
        ws.cell(row=r, column=4).number_format = "@"
        ws.cell(row=r, column=5).number_format = "@"
        # Dates
        ws.cell(row=r, column=2).number_format = FMT_DATE
        ws.cell(row=r, column=11).number_format = FMT_DATE
        ws.cell(row=r, column=12).number_format = FMT_DATE
        # USD columns
        for c in (8, 9, 10, 13, 14):
            ws.cell(row=r, column=c).number_format = FMT_USD
        # Days outstanding as int
        ws.cell(row=r, column=16).number_format = FMT_INT
        ws.row_dimensions[r].height = 32

    # --- Conditional formatting ---

    # Aging Bucket (O): green=Current/Paid, yellow=1-30/31-60, orange=61-90, red=91-120/120+
    aging_range = f"{col_aging}{data_start}:{col_aging}{data_end}"
    ws.conditional_formatting.add(
        aging_range,
        FormulaRule(
            formula=[f'OR(${col_aging}{data_start}="Current",${col_aging}{data_start}="Paid")'],
            stopIfTrue=False,
            fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        aging_range,
        FormulaRule(
            formula=[f'OR(${col_aging}{data_start}="1-30",${col_aging}{data_start}="31-60")'],
            stopIfTrue=False,
            fill=FILL_YELLOW, font=Font(color=YELLOW_FONT, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        aging_range,
        FormulaRule(
            formula=[f'${col_aging}{data_start}="61-90"'],
            stopIfTrue=False,
            fill=FILL_ORANGE, font=Font(color=ORANGE_FONT, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        aging_range,
        FormulaRule(
            formula=[f'OR(${col_aging}{data_start}="91-120",${col_aging}{data_start}="120+")'],
            stopIfTrue=False,
            fill=FILL_RED, font=Font(color=RED_FONT, bold=True),
        ),
    )

    # Outstanding (N): red fill when >0 AND aging >60 days (i.e. 61-90, 91-120, 120+)
    out_range = f"{col_outstanding}{data_start}:{col_outstanding}{data_end}"
    ws.conditional_formatting.add(
        out_range,
        FormulaRule(
            formula=[
                f'AND(${col_outstanding}{data_start}>0,'
                f'OR(${col_aging}{data_start}="61-90",'
                f'${col_aging}{data_start}="91-120",'
                f'${col_aging}{data_start}="120+"))'
            ],
            stopIfTrue=False,
            fill=FILL_RED, font=Font(color=RED_FONT, bold=True),
        ),
    )

    # Status (R): green=Paid, yellow=Sent/Partially Paid, red=Disputed/Written Off
    status_range = f"{col_status}{data_start}:{col_status}{data_end}"
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'${col_status}{data_start}="Paid"'],
            stopIfTrue=False,
            fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'OR(${col_status}{data_start}="Sent",${col_status}{data_start}="Partially Paid")'],
            stopIfTrue=False,
            fill=FILL_YELLOW, font=Font(color=YELLOW_FONT, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'OR(${col_status}{data_start}="Disputed",${col_status}{data_start}="Written Off")'],
            stopIfTrue=False,
            fill=FILL_RED, font=Font(color=RED_FONT, bold=True),
        ),
    )

    # --- Data validation ---

    # CSI Division dropdown (uses named range)
    dv_div = DataValidation(type="list", formula1="=CSI_Divisions", allow_blank=True)
    dv_div.error = "Pick a CSI Division from the dropdown."
    dv_div.errorTitle = "Invalid CSI Division"
    ws.add_data_validation(dv_div)
    dv_div.add(f"{col_div}{data_start}:{col_div}{data_end}")

    # Status dropdown
    dv_status = DataValidation(
        type="list", formula1='"' + ",".join(STATUSES) + '"', allow_blank=True,
    )
    dv_status.error = "Pick a status from the dropdown."
    dv_status.errorTitle = "Invalid status"
    ws.add_data_validation(dv_status)
    dv_status.add(f"{col_status}{data_start}:{col_status}{data_end}")

    # Lien Waiver dropdown
    dv_lien = DataValidation(
        type="list", formula1='"' + ",".join(LIEN_WAIVERS) + '"', allow_blank=True,
    )
    dv_lien.error = "Pick from the dropdown."
    dv_lien.errorTitle = "Invalid lien waiver"
    ws.add_data_validation(dv_lien)
    dv_lien.add(f"Q{data_start}:Q{data_end}")

    # Totals row
    totals_row = data_end + 2
    ws.cell(row=totals_row, column=7, value="TOTALS").font = FONT_BODY_BOLD
    ws.cell(row=totals_row, column=7).alignment = Alignment(horizontal="right")
    ws.cell(row=totals_row, column=7).fill = FILL_SUBHEADER
    ws.cell(row=totals_row, column=7).border = BORDER
    for c in (8, 9, 10, 13, 14):
        col_letter = get_column_letter(c)
        cell = ws.cell(
            row=totals_row, column=c,
            value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})",
        )
        cell.number_format = FMT_USD
        cell.font = FONT_BODY_BOLD
        cell.fill = FILL_SUBHEADER
        cell.border = BORDER

    # Column widths
    set_col_widths(ws, {
        "A": 12, "B": 12, "C": 22, "D": 11, "E": 12, "F": 24,
        "G": 36, "H": 14, "I": 14, "J": 14, "K": 12, "L": 12,
        "M": 14, "N": 14, "O": 12, "P": 11, "Q": 16, "R": 16, "S": 32,
    })

    ws.freeze_panes = "B4"

    return ws.title, data_start, data_end


# ---------------------------------------------------------------------------
# Tab 4 — Aging Summary
# ---------------------------------------------------------------------------

AGING_HEADERS = [
    "Aging Bucket", "Count", "Total Outstanding", "% of Total",
]

# Buckets shown in the summary (Paid intentionally excluded — this is an
# aging report of outstanding receivables)
AGING_BUCKETS_REPORT = ["Current", "1-30", "31-60", "61-90", "91-120", "120+"]


def build_aging_summary(wb: Workbook, inv_sheet: str, inv_start: int, inv_end: int) -> None:
    ws = wb.create_sheet("Aging Summary")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "AR Aging Summary (auto roll-up)"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(AGING_HEADERS))

    ws["A2"] = ("Reads from Invoice Log via COUNTIF / SUMIFS - do not edit this tab directly. "
                "One row per aging bucket.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)

    for i, h in enumerate(AGING_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(AGING_HEADERS))

    inv_aging = f"'{inv_sheet}'!$O${inv_start}:$O${inv_end}"
    inv_outstanding = f"'{inv_sheet}'!$N${inv_start}:$N${inv_end}"

    start_row = 4
    bucket_fills: dict[str, tuple[PatternFill, str]] = {
        "Current": (FILL_GREEN, GREEN_FONT),
        "1-30": (FILL_YELLOW, YELLOW_FONT),
        "31-60": (FILL_YELLOW, YELLOW_FONT),
        "61-90": (FILL_ORANGE, ORANGE_FONT),
        "91-120": (FILL_RED, RED_FONT),
        "120+": (FILL_RED, RED_FONT),
    }

    for idx, bucket in enumerate(AGING_BUCKETS_REPORT):
        r = start_row + idx
        fill, font_color = bucket_fills[bucket]
        ws.cell(row=r, column=1, value=bucket)
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=11, bold=True, color=font_color)
        ws.cell(row=r, column=1).fill = fill
        ws.cell(row=r, column=1).alignment = Alignment(vertical="center", indent=1)
        # Count
        ws.cell(row=r, column=2, value=f'=COUNTIF({inv_aging},A{r})')
        ws.cell(row=r, column=2).number_format = FMT_INT
        # Total Outstanding for this bucket
        ws.cell(row=r, column=3, value=f'=SUMIFS({inv_outstanding},{inv_aging},A{r})')
        ws.cell(row=r, column=3).number_format = FMT_USD

    end_row = start_row + len(AGING_BUCKETS_REPORT) - 1

    # Totals row (sums of outstanding across all aging buckets shown)
    tr = end_row + 1
    ws.cell(row=tr, column=1, value="TOTAL OUTSTANDING")
    ws.cell(row=tr, column=1).font = FONT_BODY_BOLD
    ws.cell(row=tr, column=1).fill = FILL_SUBHEADER
    ws.cell(row=tr, column=1).alignment = Alignment(vertical="center", indent=1)

    ws.cell(row=tr, column=2, value=f'=SUM(B{start_row}:B{end_row})')
    ws.cell(row=tr, column=2).number_format = FMT_INT
    ws.cell(row=tr, column=2).font = FONT_BODY_BOLD
    ws.cell(row=tr, column=2).fill = FILL_SUBHEADER

    ws.cell(row=tr, column=3, value=f'=SUM(C{start_row}:C{end_row})')
    ws.cell(row=tr, column=3).number_format = FMT_USD
    ws.cell(row=tr, column=3).font = FONT_BODY_BOLD
    ws.cell(row=tr, column=3).fill = FILL_SUBHEADER

    ws.cell(row=tr, column=4, value="=1")
    ws.cell(row=tr, column=4).number_format = FMT_PCT
    ws.cell(row=tr, column=4).font = FONT_BODY_BOLD
    ws.cell(row=tr, column=4).fill = FILL_SUBHEADER

    # Now fill % of Total for each row (referencing the total at C{tr})
    for idx, bucket in enumerate(AGING_BUCKETS_REPORT):
        r = start_row + idx
        ws.cell(row=r, column=4,
                value=f'=IFERROR(C{r}/$C${tr},0)')
        ws.cell(row=r, column=4).number_format = FMT_PCT

    # Borders across the whole table
    for r in range(3, tr + 1):
        for c in range(1, len(AGING_HEADERS) + 1):
            ws.cell(row=r, column=c).border = BORDER
        ws.row_dimensions[r].height = 24

    set_col_widths(ws, {"A": 18, "B": 12, "C": 22, "D": 14})

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Tab 5 — Dashboard
# ---------------------------------------------------------------------------

def build_dashboard(wb: Workbook, inv_sheet: str, inv_start: int, inv_end: int) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Accounts Receivable Dashboard"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:F1")

    ws["A2"] = ("Auto-calculated from Project Info and Invoice Log. Update those tabs - this one stays in sync.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)
    ws.merge_cells("A2:F2")

    inv_total = f"'{inv_sheet}'!$J${inv_start}:$J${inv_end}"
    inv_paid = f"'{inv_sheet}'!$M${inv_start}:$M${inv_end}"
    inv_outstanding = f"'{inv_sheet}'!$N${inv_start}:$N${inv_end}"
    inv_retain = f"'{inv_sheet}'!$I${inv_start}:$I${inv_end}"
    inv_aging = f"'{inv_sheet}'!$O${inv_start}:$O${inv_end}"
    inv_days = f"'{inv_sheet}'!$P${inv_start}:$P${inv_end}"
    inv_status = f"'{inv_sheet}'!$R${inv_start}:$R${inv_end}"
    inv_num = f"'{inv_sheet}'!$A${inv_start}:$A${inv_end}"

    # ---------------- Top panel: AR headline ----------------
    panel_start = 4
    ws.cell(row=panel_start, column=1, value="AR Headline").font = FONT_H1
    ws.cell(row=panel_start, column=1).fill = FILL_HEADER
    ws.cell(row=panel_start, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=panel_start, start_column=1,
                   end_row=panel_start, end_column=6)
    ws.row_dimensions[panel_start].height = 26

    # (label, formula, fmt, conditional_kind)
    #   conditional_kind:
    #     None       -> no CF
    #     "outstanding" -> red if > 0 (you have money on the street), green if 0
    #     "dso"      -> red if >90, yellow if >60, green if <=60
    metrics = [
        ("Total Invoiced", f"=SUM({inv_total})", FMT_USD, None),
        ("Total Paid", f"=SUM({inv_paid})", FMT_USD, None),
        ("Total Outstanding", f"=SUM({inv_outstanding})", FMT_USD, "outstanding"),
        ("Total Retainage Held", f"=SUM({inv_retain})", FMT_USD, None),
        # DSO = SUMPRODUCT(Outstanding * Days) / SUM(Outstanding)
        # i.e. dollar-weighted average days-late for the outstanding pool.
        ("Days Sales Outstanding (DSO)",
         f'=IFERROR(SUMPRODUCT({inv_outstanding},IFERROR({inv_days}+0,0))/SUM({inv_outstanding}),0)',
         FMT_INT, "dso"),
    ]

    row = panel_start + 1
    outstanding_cells: list[str] = []
    dso_cells: list[str] = []
    for label, formula, fmt, kind in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = FONT_BODY_BOLD
        ws.cell(row=row, column=1).fill = FILL_SUMMARY_LABEL
        ws.cell(row=row, column=1).alignment = Alignment(vertical="center", indent=1)
        ws.cell(row=row, column=1).border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

        ws.cell(row=row, column=4, value=formula)
        ws.cell(row=row, column=4).font = FONT_BIG_NUMBER
        ws.cell(row=row, column=4).number_format = fmt
        ws.cell(row=row, column=4).alignment = Alignment(
            horizontal="right", vertical="center", indent=1,
        )
        ws.cell(row=row, column=4).border = BORDER
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)

        ref = f"D{row}"
        if kind == "outstanding":
            outstanding_cells.append(ref)
        elif kind == "dso":
            dso_cells.append(ref)
        ws.row_dimensions[row].height = 28
        row += 1

    # Outstanding: red if >0
    for ref in outstanding_cells:
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=False,
                       fill=FILL_RED, font=Font(color=RED_FONT, bold=True, size=18)),
        )
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="lessThanOrEqual", formula=["0"], stopIfTrue=False,
                       fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True, size=18)),
        )

    # DSO: green <=60, yellow 60-90, red >90
    for ref in dso_cells:
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="greaterThan", formula=["90"], stopIfTrue=True,
                       fill=FILL_RED, font=Font(color=RED_FONT, bold=True, size=18)),
        )
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="greaterThan", formula=["60"], stopIfTrue=True,
                       fill=FILL_YELLOW, font=Font(color=YELLOW_FONT, bold=True, size=18)),
        )
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="lessThanOrEqual", formula=["60"], stopIfTrue=False,
                       fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True, size=18)),
        )

    # ---------------- Middle panel: Aging breakdown reference ----------------
    section_row = row + 2
    ws.cell(row=section_row, column=1, value="Aging Breakdown").font = FONT_H1
    ws.cell(row=section_row, column=1).fill = FILL_HEADER
    ws.cell(row=section_row, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=section_row, start_column=1,
                   end_row=section_row, end_column=6)
    ws.row_dimensions[section_row].height = 26

    aging_headers = ["Aging Bucket", "Count", "Outstanding", "% of Outstanding"]
    hr = section_row + 1
    for i, h in enumerate(aging_headers, start=1):
        ws.cell(row=hr, column=i, value=h)
    style_header_row(ws, hr, len(aging_headers))
    # Merge Outstanding and % to span two cols visually
    ws.merge_cells(start_row=hr, start_column=3, end_row=hr, end_column=4)
    ws.merge_cells(start_row=hr, start_column=5, end_row=hr, end_column=6)

    bucket_panel = [
        ("Current", FILL_GREEN, GREEN_FONT),
        ("1-30", FILL_YELLOW, YELLOW_FONT),
        ("31-60", FILL_YELLOW, YELLOW_FONT),
        ("61-90", FILL_ORANGE, ORANGE_FONT),
        ("91-120", FILL_RED, RED_FONT),
        ("120+", FILL_RED, RED_FONT),
    ]

    total_outstanding_formula = f"SUM({inv_outstanding})"

    br = hr + 1
    for idx, (bucket, fill, font_color) in enumerate(bucket_panel):
        r = br + idx
        ws.cell(row=r, column=1, value=bucket)
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=12, bold=True, color=font_color)
        ws.cell(row=r, column=1).fill = fill
        ws.cell(row=r, column=1).alignment = Alignment(vertical="center", indent=1)
        ws.cell(row=r, column=1).border = BORDER
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

        # Count
        ws.cell(row=r, column=3, value=f'=COUNTIF({inv_aging},"{bucket}")')
        ws.cell(row=r, column=3).font = FONT_BODY_BOLD
        ws.cell(row=r, column=3).number_format = FMT_INT
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=3).border = BORDER

        # Outstanding amount in this bucket
        ws.cell(row=r, column=4,
                value=f'=SUMIFS({inv_outstanding},{inv_aging},"{bucket}")')
        ws.cell(row=r, column=4).font = FONT_BODY_BOLD
        ws.cell(row=r, column=4).number_format = FMT_USD
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.cell(row=r, column=4).border = BORDER

        # % of total outstanding
        ws.cell(row=r, column=5,
                value=f'=IFERROR(SUMIFS({inv_outstanding},{inv_aging},"{bucket}")/{total_outstanding_formula},0)')
        ws.cell(row=r, column=5).font = FONT_BODY_BOLD
        ws.cell(row=r, column=5).number_format = FMT_PCT
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.cell(row=r, column=5).border = BORDER
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 22

    # Aging total row
    btr = br + len(bucket_panel)
    ws.cell(row=btr, column=1, value="TOTAL OUTSTANDING")
    ws.cell(row=btr, column=1).font = FONT_BODY_BOLD
    ws.cell(row=btr, column=1).fill = FILL_SUBHEADER
    ws.cell(row=btr, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.cell(row=btr, column=1).border = BORDER
    ws.merge_cells(start_row=btr, start_column=1, end_row=btr, end_column=2)

    ws.cell(row=btr, column=3,
            value=f'=COUNTIF({inv_aging},"Current")+COUNTIF({inv_aging},"1-30")'
                  f'+COUNTIF({inv_aging},"31-60")+COUNTIF({inv_aging},"61-90")'
                  f'+COUNTIF({inv_aging},"91-120")+COUNTIF({inv_aging},"120+")')
    ws.cell(row=btr, column=3).font = FONT_BODY_BOLD
    ws.cell(row=btr, column=3).fill = FILL_SUBHEADER
    ws.cell(row=btr, column=3).number_format = FMT_INT
    ws.cell(row=btr, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=btr, column=3).border = BORDER

    ws.cell(row=btr, column=4, value=f'=SUM({inv_outstanding})')
    ws.cell(row=btr, column=4).font = FONT_BODY_BOLD
    ws.cell(row=btr, column=4).fill = FILL_SUBHEADER
    ws.cell(row=btr, column=4).number_format = FMT_USD
    ws.cell(row=btr, column=4).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.cell(row=btr, column=4).border = BORDER

    ws.cell(row=btr, column=5, value="=1")
    ws.cell(row=btr, column=5).font = FONT_BODY_BOLD
    ws.cell(row=btr, column=5).fill = FILL_SUBHEADER
    ws.cell(row=btr, column=5).number_format = FMT_PCT
    ws.cell(row=btr, column=5).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.cell(row=btr, column=5).border = BORDER
    ws.merge_cells(start_row=btr, start_column=5, end_row=btr, end_column=6)

    # ---------------- Bottom panel: Count by Status ----------------
    status_section_row = btr + 2
    ws.cell(row=status_section_row, column=1, value="Count by Status").font = FONT_H1
    ws.cell(row=status_section_row, column=1).fill = FILL_HEADER
    ws.cell(row=status_section_row, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=status_section_row, start_column=1,
                   end_row=status_section_row, end_column=6)
    ws.row_dimensions[status_section_row].height = 26

    status_headers = ["Status", "Count", "Amount Outstanding", "% of Outstanding"]
    shr = status_section_row + 1
    for i, h in enumerate(status_headers, start=1):
        ws.cell(row=shr, column=i, value=h)
    style_header_row(ws, shr, len(status_headers))
    ws.merge_cells(start_row=shr, start_column=3, end_row=shr, end_column=4)
    ws.merge_cells(start_row=shr, start_column=5, end_row=shr, end_column=6)

    status_panel = [
        ("Sent", FILL_YELLOW, YELLOW_FONT),
        ("Partially Paid", FILL_YELLOW, YELLOW_FONT),
        ("Paid", FILL_GREEN, GREEN_FONT),
        ("Disputed", FILL_RED, RED_FONT),
        ("Written Off", FILL_RED, RED_FONT),
    ]

    sr = shr + 1
    for idx, (status, fill, font_color) in enumerate(status_panel):
        r = sr + idx
        ws.cell(row=r, column=1, value=status)
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=12, bold=True, color=font_color)
        ws.cell(row=r, column=1).fill = fill
        ws.cell(row=r, column=1).alignment = Alignment(vertical="center", indent=1)
        ws.cell(row=r, column=1).border = BORDER
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

        # Count
        ws.cell(row=r, column=3, value=f'=COUNTIF({inv_status},"{status}")')
        ws.cell(row=r, column=3).font = FONT_BODY_BOLD
        ws.cell(row=r, column=3).number_format = FMT_INT
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=3).border = BORDER

        # Outstanding amount for this status
        ws.cell(row=r, column=4,
                value=f'=SUMIFS({inv_outstanding},{inv_status},"{status}")')
        ws.cell(row=r, column=4).font = FONT_BODY_BOLD
        ws.cell(row=r, column=4).number_format = FMT_USD
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.cell(row=r, column=4).border = BORDER

        # % of outstanding
        ws.cell(row=r, column=5,
                value=f'=IFERROR(SUMIFS({inv_outstanding},{inv_status},"{status}")/{total_outstanding_formula},0)')
        ws.cell(row=r, column=5).font = FONT_BODY_BOLD
        ws.cell(row=r, column=5).number_format = FMT_PCT
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.cell(row=r, column=5).border = BORDER
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 22

    # Status totals row
    str_total = sr + len(status_panel)
    ws.cell(row=str_total, column=1, value="ALL INVOICES (any status)")
    ws.cell(row=str_total, column=1).font = FONT_BODY_BOLD
    ws.cell(row=str_total, column=1).fill = FILL_SUBHEADER
    ws.cell(row=str_total, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.cell(row=str_total, column=1).border = BORDER
    ws.merge_cells(start_row=str_total, start_column=1, end_row=str_total, end_column=2)

    ws.cell(row=str_total, column=3, value=f'=COUNTA({inv_num})')
    ws.cell(row=str_total, column=3).font = FONT_BODY_BOLD
    ws.cell(row=str_total, column=3).fill = FILL_SUBHEADER
    ws.cell(row=str_total, column=3).number_format = FMT_INT
    ws.cell(row=str_total, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=str_total, column=3).border = BORDER

    ws.cell(row=str_total, column=4, value=f'=SUM({inv_outstanding})')
    ws.cell(row=str_total, column=4).font = FONT_BODY_BOLD
    ws.cell(row=str_total, column=4).fill = FILL_SUBHEADER
    ws.cell(row=str_total, column=4).number_format = FMT_USD
    ws.cell(row=str_total, column=4).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.cell(row=str_total, column=4).border = BORDER

    ws.cell(row=str_total, column=5, value="=1")
    ws.cell(row=str_total, column=5).font = FONT_BODY_BOLD
    ws.cell(row=str_total, column=5).fill = FILL_SUBHEADER
    ws.cell(row=str_total, column=5).number_format = FMT_PCT
    ws.cell(row=str_total, column=5).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.cell(row=str_total, column=5).border = BORDER
    ws.merge_cells(start_row=str_total, start_column=5, end_row=str_total, end_column=6)

    set_col_widths(ws, {
        "A": 22, "B": 12, "C": 14, "D": 18, "E": 14, "F": 10,
    })


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

OUTPUT_PATH = "/Users/home/charles/contrpro/files/packages/business/AR_Tracker.xlsx"


def build() -> str:
    wb = Workbook()

    build_instructions(wb)            # Active sheet renamed to Instructions
    build_csi_reference(wb)           # Adds named ranges CSI_Divisions and CSI_Table
    build_project_info(wb)
    inv_sheet, inv_start, inv_end = build_invoice_log(wb)
    build_aging_summary(wb, inv_sheet, inv_start, inv_end)
    build_dashboard(wb, inv_sheet, inv_start, inv_end)

    # Re-order tabs for proper UX: Instructions, Project Info, Invoice Log,
    # Aging Summary, Dashboard, CSI Reference (hidden, last)
    desired_order = [
        "Instructions",
        "Project Info",
        "Invoice Log",
        "Aging Summary",
        "Dashboard",
        "CSI Reference",
    ]
    wb._sheets = [wb[name] for name in desired_order]

    # Hide CSI Reference (canonical lookup; user shouldn't see it by default)
    wb["CSI Reference"].sheet_state = "hidden"

    # Active tab = Dashboard on open
    wb.active = wb.sheetnames.index("Dashboard")

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
