#!/usr/bin/env python3
"""
Build ContrPro Job Costing Spreadsheet (XLSX).

Produces a production-grade workbook with:
  - Instructions tab
  - Job Info tab (data validation, currency formats)
  - Cost Tracking tab (CSI MasterFormat coded, formulas, conditional formatting)
  - Division Summary tab (SUMIF roll-ups)
  - Dashboard tab (totals, profit projection, weighted % complete)

Run:
    /Users/home/charles/.venv/bin/python3 \
        /Users/home/charles/contrpro/scripts/build_job_costing_xlsx.py

Output:
    /Users/home/charles/contrpro/files/packages/business/Job_Costing_Spreadsheet.xlsx
"""

from __future__ import annotations

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ---------------------------------------------------------------------------
# Brand & styling
# ---------------------------------------------------------------------------

BRAND_BLUE = "1E3A5F"       # ContrPro brand blue
BRAND_BLUE_LIGHT = "D6E0EC"  # tint for banded rows
ACCENT_GOLD = "C9A227"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_SUMMARY_LABEL = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H1 = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_BIG_NUMBER = Font(name="Calibri", size=18, bold=True, color=BRAND_BLUE)
FONT_GREY_ITALIC = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)
FONT_HYPERLINK = Font(name="Calibri", size=11, color="0563C1", underline="single")

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_USD_BIG = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
FMT_PCT = "0.0%"
FMT_PCT_INT = "0%"


# ---------------------------------------------------------------------------
# CSI MasterFormat reference data (pre-populated rows)
# ---------------------------------------------------------------------------

# (division, code, division_name, scope_description, default_trade)
COST_ROWS: list[tuple[str, str, str, str, str]] = [
    ("01", "01 10 00", "General Requirements", "Summary", "General"),
    ("01", "01 50 00", "General Requirements", "Temporary Facilities and Controls", "General"),
    ("01", "01 70 00", "General Requirements", "Execution and Closeout Requirements", "General"),

    ("02", "02 30 00", "Existing Conditions", "Assessment", "General"),
    ("02", "02 41 00", "Existing Conditions", "Demolition", "General"),

    ("03", "03 30 00", "Concrete", "Cast-in-Place Concrete", "Structural Steel"),
    ("03", "03 40 00", "Concrete", "Precast Concrete", "Structural Steel"),

    ("04", "04 20 00", "Masonry", "Unit Masonry", "General"),

    ("05", "05 12 00", "Metals", "Structural Steel Framing", "Structural Steel"),
    ("05", "05 31 00", "Metals", "Steel Decking", "Structural Steel"),
    ("05", "05 50 00", "Metals", "Metal Fabrications", "Structural Steel"),

    ("06", "06 10 00", "Wood, Plastics, and Composites", "Rough Carpentry", "General"),
    ("06", "06 20 00", "Wood, Plastics, and Composites", "Finish Carpentry", "General"),

    ("07", "07 20 00", "Thermal and Moisture Protection", "Insulation", "General"),
    ("07", "07 50 00", "Thermal and Moisture Protection", "Membrane Roofing", "General"),
    ("07", "07 90 00", "Thermal and Moisture Protection", "Joint Protection", "General"),

    ("08", "08 10 00", "Openings", "Doors and Frames", "General"),
    ("08", "08 50 00", "Openings", "Windows", "General"),

    ("09", "09 20 00", "Finishes", "Plaster and Gypsum Board", "General"),
    ("09", "09 60 00", "Finishes", "Flooring", "General"),
    ("09", "09 90 00", "Finishes", "Painting and Coating", "General"),

    ("10", "10 50 00", "Specialties", "Storage Specialties", "General"),

    ("21", "21 10 00", "Fire Suppression", "Water-Based Fire-Suppression Systems", "Mechanical"),

    ("22", "22 10 00", "Plumbing", "Plumbing Piping", "Plumbing"),
    ("22", "22 30 00", "Plumbing", "Plumbing Equipment", "Plumbing"),
    ("22", "22 40 00", "Plumbing", "Plumbing Fixtures", "Plumbing"),

    ("23", "23 20 00", "HVAC", "HVAC Piping and Pumps", "Mechanical"),
    ("23", "23 30 00", "HVAC", "HVAC Air Distribution", "Mechanical"),
    ("23", "23 70 00", "HVAC", "Central HVAC Equipment", "Mechanical"),
    ("23", "23 80 00", "HVAC", "Decentralized HVAC Equipment", "Mechanical"),

    ("26", "26 20 00", "Electrical", "Low-Voltage Electrical Distribution", "Electrical"),
    ("26", "26 50 00", "Electrical", "Lighting", "Electrical"),

    ("31", "31 20 00", "Earthwork", "Earth Moving", "General"),
    ("31", "31 60 00", "Earthwork", "Special Foundations", "General"),

    ("32", "32 12 00", "Exterior Improvements", "Flexible Paving", "General"),
    ("32", "32 31 00", "Exterior Improvements", "Fences and Gates", "General"),

    ("33", "33 10 00", "Utilities", "Water Utilities", "General"),
]

# Division Summary rows — one per division actually used above.
DIVISION_NAMES: dict[str, str] = {}
for div, _code, name, _desc, _trade in COST_ROWS:
    DIVISION_NAMES.setdefault(div, name)
DIVISIONS_ORDERED = sorted(DIVISION_NAMES.keys(), key=lambda d: int(d))

EMPTY_ROWS_AT_END = 10  # user-fillable rows on Cost Tracking
EMPTY_JOB_ROWS = 5      # user-fillable rows on Job Info

TRADES = ["General", "Mechanical", "Electrical", "Plumbing", "Structural Steel", "Other"]
STATUSES = ["Not Started", "In Progress", "Substantial Completion", "Final", "On Hold"]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def set_col_widths(ws, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_header_row(ws, row: int, last_col: int) -> None:
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 28


def border_range(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER


# ---------------------------------------------------------------------------
# Tab 1 — Instructions
# ---------------------------------------------------------------------------

def build_instructions(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {"A": 4, "B": 110})

    ws["B2"] = "ContrPro - Job Costing Spreadsheet"
    ws["B2"].font = FONT_TITLE
    ws.row_dimensions[2].height = 32

    ws["B3"] = "CSI MasterFormat-coded job cost tracker for general contractors and trade subs."
    ws["B3"].font = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

    sections: list[tuple[str, list[str]]] = [
        (
            "How to use this workbook",
            [
                "This workbook tracks budget vs. actual cost on a single construction job, organized by CSI MasterFormat division and section. Five tabs work together: Job Info captures the contract, Cost Tracking is where line-item costs go, Division Summary rolls them up, and Dashboard shows the bottom line.",
                "Open the tabs at the bottom of the window in this order: Job Info first, then Cost Tracking, then check Division Summary and Dashboard.",
            ],
        ),
        (
            "1. Job Info",
            [
                "Fill in Job Name, Job Number, Client, Contract Amount, Start Date, Target Completion, Project Manager, and Estimator on row 4 (the first blank row).",
                "Status is a dropdown: Not Started, In Progress, Substantial Completion, Final, On Hold.",
                "Contract Amount on the FIRST job row (row 4) is the value the Dashboard reads for profit projections. If you track multiple jobs, build a separate file per job.",
            ],
        ),
        (
            "2. Cost Tracking",
            [
                "Each row is one scope item, coded by CSI Division (2-digit) and CSI Code (6-digit, e.g. 05 12 00).",
                "Pre-populated rows cover the divisions most jobs touch. Add more at the bottom in the empty rows.",
                "Enter Budget, Committed (PO/subcontract value), Actual to Date, and % Complete. Variance and Earned Value are calculated automatically.",
                "Trade column is a dropdown - use it to filter Subcontractor View (just sort or filter by Trade).",
                "Variance turns GREEN when you're under or on budget and RED when you're over.",
            ],
        ),
        (
            "3. Division Summary",
            [
                "Auto-rolls Cost Tracking up to the division level using SUMIF. Do not type into this tab - it reads from Cost Tracking.",
                "Weighted % Complete is dollar-weighted by Budget so a 100% complete trim line does not skew a 10% complete steel package.",
            ],
        ),
        (
            "4. Dashboard",
            [
                "Top panel shows Contract Amount, Total Budget, Committed, Actual, Variance, Weighted % Complete, Projected Profit, and Profit Margin.",
                "Projected Profit assumes cost-to-complete tracks linearly with % complete; refine your Actuals weekly for accuracy.",
                "Below the panel is the per-division breakdown - same numbers as Division Summary, formatted for at-a-glance review.",
            ],
        ),
        (
            "5. CSI MasterFormat coding rules",
            [
                "CSI Division: 2-digit number, leading zero preserved as text (01, 02 ... 33). Do NOT enter as a number or Excel will drop the leading zero.",
                "CSI Code: 6-digit space-separated format - Level 2 (e.g. 05 12 00 for Structural Steel Framing). You can extend to Level 3 (05 12 23) if you need finer detail.",
                "Division Name should match the standard MasterFormat name so subs and architects recognize it on pay apps and change orders.",
            ],
        ),
        (
            "6. GC vs Subcontractor view",
            [
                "GC View: leave Cost Tracking unfiltered - all divisions visible, Division Summary shows every trade.",
                "Subcontractor View: use Excel/Numbers/Sheets filter on the Trade column (or CSI Division) to show only your scope. The formulas still work because SUMIF references the full range.",
            ],
        ),
        (
            "7. Weekly update cadence",
            [
                "Friday afternoon: walk the job, update % Complete and Actual to Date on every active line.",
                "Monday morning: review Dashboard - any RED Variance gets a phone call before the week starts.",
                "End of month: reconcile Committed against signed POs and subcontracts; this is your CFO-grade snapshot.",
            ],
        ),
        (
            "Tabs in this workbook",
            [
                "Instructions  -  this tab",
                "Job Info  -  contract and project info (Dashboard reads Contract Amount from here)",
                "Cost Tracking  -  CSI-coded line items, the main worksheet",
                "Division Summary  -  SUMIF roll-up by division",
                "Dashboard  -  totals, projected profit, % complete by division",
            ],
        ),
    ]

    row = 5
    for heading, paragraphs in sections:
        ws.cell(row=row, column=2, value=heading)
        ws.cell(row=row, column=2).font = FONT_H2
        ws.cell(row=row, column=2).fill = FILL_SUBHEADER
        ws.cell(row=row, column=2).alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[row].height = 22
        row += 1
        for para in paragraphs:
            ws.cell(row=row, column=2, value=para)
            ws.cell(row=row, column=2).font = FONT_BODY
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            # Approximate height: 18 per ~110 chars
            ws.row_dimensions[row].height = max(18, 16 * (1 + len(para) // 110))
            row += 1
        row += 1  # gap between sections

    # Footer note
    ws.cell(row=row + 1, column=2, value="ContrPro - built for builders. Questions: support@contrpro.com")
    ws.cell(row=row + 1, column=2).font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)


# ---------------------------------------------------------------------------
# Tab 2 — Job Info
# ---------------------------------------------------------------------------

JOB_INFO_HEADERS = [
    "Job Name", "Job Number", "Client Name", "Contract Amount",
    "Start Date", "Target Completion", "Actual Completion",
    "Status", "Project Manager", "Estimator", "Notes",
]

# Pre-filled example row (greyed/italic)
JOB_INFO_EXAMPLE = [
    "EXAMPLE - Riverside Office Bldg",
    "2026-014",
    "Riverside Holdings LLC",
    1850000,
    "2026-03-01",
    "2026-11-30",
    "",
    "In Progress",
    "Sample PM",
    "Sample Estimator",
    "Example row - replace with your job.",
]


def build_job_info(wb: Workbook) -> None:
    ws = wb.create_sheet("Job Info")
    ws.sheet_view.showGridLines = False

    # Title
    ws["A1"] = "Job Information"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:K1")

    # Header row at row 3
    for i, h in enumerate(JOB_INFO_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(JOB_INFO_HEADERS))

    # Example row at row 4 (greyed/italic)
    for i, v in enumerate(JOB_INFO_EXAMPLE, start=1):
        cell = ws.cell(row=4, column=i, value=v)
        cell.font = FONT_GREY_ITALIC
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.cell(row=4, column=4).number_format = FMT_USD

    # Empty rows
    for r in range(5, 5 + EMPTY_JOB_ROWS):
        for c in range(1, len(JOB_INFO_HEADERS) + 1):
            ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=4).number_format = FMT_USD
        ws.cell(row=r, column=5).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=6).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=7).number_format = "yyyy-mm-dd"
        ws.row_dimensions[r].height = 22

    # Status dropdown on H4:H9
    dv_status = DataValidation(
        type="list",
        formula1='"' + ",".join(STATUSES) + '"',
        allow_blank=True,
        showDropDown=False,  # show arrow
    )
    dv_status.error = "Pick a status from the dropdown."
    dv_status.errorTitle = "Invalid status"
    ws.add_data_validation(dv_status)
    dv_status.add(f"H4:H{4 + EMPTY_JOB_ROWS}")

    set_col_widths(ws, {
        "A": 28, "B": 14, "C": 28, "D": 16, "E": 14, "F": 16,
        "G": 16, "H": 22, "I": 18, "J": 18, "K": 40,
    })

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Tab 3 — Cost Tracking
# ---------------------------------------------------------------------------

COST_HEADERS = [
    "CSI Division", "CSI Code", "Division Name", "Description / Scope Item",
    "Trade", "Budget", "Committed", "Actual to Date",
    "Variance", "% Complete", "Earned Value",
    "Schedule of Values Item", "Notes",
]


def build_cost_tracking(wb: Workbook) -> Tuple[str, int, int]:
    ws = wb.create_sheet("Cost Tracking")
    ws.sheet_view.showGridLines = False

    # Title
    ws["A1"] = "Cost Tracking (CSI MasterFormat)"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COST_HEADERS))

    # Header row at row 3
    for i, h in enumerate(COST_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(COST_HEADERS))

    # Data starts at row 4
    data_start = 4
    n_pre = len(COST_ROWS)
    n_total = n_pre + EMPTY_ROWS_AT_END
    data_end = data_start + n_total - 1

    for idx, (div, code, name, desc, trade) in enumerate(COST_ROWS):
        r = data_start + idx
        ws.cell(row=r, column=1, value=div)
        ws.cell(row=r, column=2, value=code)
        ws.cell(row=r, column=3, value=name)
        ws.cell(row=r, column=4, value=desc)
        ws.cell(row=r, column=5, value=trade)
        # Budget / Committed / Actual blank for user input
        ws.cell(row=r, column=6, value=None)
        ws.cell(row=r, column=7, value=None)
        ws.cell(row=r, column=8, value=None)
        # Variance = Budget - Actual
        ws.cell(row=r, column=9, value=f"=IFERROR(F{r}-H{r},0)")
        # % Complete blank
        ws.cell(row=r, column=10, value=None)
        # Earned Value = Budget * %Complete
        ws.cell(row=r, column=11, value=f"=IFERROR(F{r}*J{r},0)")
        ws.cell(row=r, column=12, value=None)
        ws.cell(row=r, column=13, value=None)

    # Empty user-fillable rows
    for i in range(EMPTY_ROWS_AT_END):
        r = data_start + n_pre + i
        ws.cell(row=r, column=9, value=f"=IFERROR(F{r}-H{r},0)")
        ws.cell(row=r, column=11, value=f"=IFERROR(F{r}*J{r},0)")

    # Formatting all data rows
    for r in range(data_start, data_end + 1):
        for c in range(1, len(COST_HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(c in (3, 4, 12, 13)))
        # CSI Division and Code as text so leading zeros stick
        ws.cell(row=r, column=1).number_format = "@"
        ws.cell(row=r, column=2).number_format = "@"
        # Currency columns
        for c in (6, 7, 8, 9, 11):
            ws.cell(row=r, column=c).number_format = FMT_USD
        # % Complete
        ws.cell(row=r, column=10).number_format = FMT_PCT
        ws.row_dimensions[r].height = 22

    # Conditional formatting on Variance column (I)
    var_range = f"I{data_start}:I{data_end}"
    ws.conditional_formatting.add(
        var_range,
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["0"],
            stopIfTrue=False,
            fill=FILL_GREEN,
            font=Font(color=GREEN_FONT, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        var_range,
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            stopIfTrue=False,
            fill=FILL_RED,
            font=Font(color=RED_FONT, bold=True),
        ),
    )

    # Data validation: Trade column (E)
    dv_trade = DataValidation(
        type="list",
        formula1='"' + ",".join(TRADES) + '"',
        allow_blank=True,
    )
    dv_trade.error = "Pick a trade from the dropdown."
    dv_trade.errorTitle = "Invalid trade"
    ws.add_data_validation(dv_trade)
    dv_trade.add(f"E{data_start}:E{data_end}")

    # Data validation: % Complete must be 0..1
    dv_pct = DataValidation(
        type="decimal",
        operator="between",
        formula1=0,
        formula2=1,
        allow_blank=True,
    )
    dv_pct.error = "Enter a value between 0 and 100% (e.g. 0.25 for 25%)."
    dv_pct.errorTitle = "Out of range"
    ws.add_data_validation(dv_pct)
    dv_pct.add(f"J{data_start}:J{data_end}")

    # Totals row at the bottom
    totals_row = data_end + 2
    ws.cell(row=totals_row, column=5, value="TOTALS").font = FONT_BODY_BOLD
    ws.cell(row=totals_row, column=5).alignment = Alignment(horizontal="right")
    for col_letter in ("F", "G", "H", "I", "K"):
        c_idx = ord(col_letter) - ord("A") + 1
        cell = ws.cell(
            row=totals_row,
            column=c_idx,
            value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})",
        )
        cell.number_format = FMT_USD
        cell.font = FONT_BODY_BOLD
        cell.fill = FILL_SUBHEADER
        cell.border = BORDER
    # Weighted % complete
    ws.cell(row=totals_row, column=10,
            value=f"=IFERROR(SUMPRODUCT(F{data_start}:F{data_end},J{data_start}:J{data_end})"
                  f"/SUM(F{data_start}:F{data_end}),0)")
    ws.cell(row=totals_row, column=10).number_format = FMT_PCT
    ws.cell(row=totals_row, column=10).font = FONT_BODY_BOLD
    ws.cell(row=totals_row, column=10).fill = FILL_SUBHEADER
    ws.cell(row=totals_row, column=10).border = BORDER
    ws.cell(row=totals_row, column=5).fill = FILL_SUBHEADER
    ws.cell(row=totals_row, column=5).border = BORDER

    # Column widths
    set_col_widths(ws, {
        "A": 11, "B": 12, "C": 26, "D": 38, "E": 16,
        "F": 14, "G": 14, "H": 14, "I": 14, "J": 13,
        "K": 14, "L": 26, "M": 30,
    })

    ws.freeze_panes = "C4"

    return ws.title, data_start, data_end


# ---------------------------------------------------------------------------
# Tab 4 — Division Summary
# ---------------------------------------------------------------------------

DIV_SUM_HEADERS = [
    "CSI Division", "Division Name",
    "Total Budget", "Total Committed", "Total Actual", "Total Variance",
    "Weighted % Complete", "Total Earned Value",
]


def build_division_summary(wb: Workbook, ct_sheet: str, ct_start: int, ct_end: int) -> None:
    ws = wb.create_sheet("Division Summary")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Division Summary (auto roll-up)"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(DIV_SUM_HEADERS))

    ws["A2"] = "Reads from Cost Tracking via SUMIF - do not edit this tab directly."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)

    for i, h in enumerate(DIV_SUM_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(DIV_SUM_HEADERS))

    # Absolute references for the Cost Tracking ranges
    ct_div = f"'{ct_sheet}'!$A${ct_start}:$A${ct_end}"
    ct_budget = f"'{ct_sheet}'!$F${ct_start}:$F${ct_end}"
    ct_committed = f"'{ct_sheet}'!$G${ct_start}:$G${ct_end}"
    ct_actual = f"'{ct_sheet}'!$H${ct_start}:$H${ct_end}"
    ct_variance = f"'{ct_sheet}'!$I${ct_start}:$I${ct_end}"
    ct_pct = f"'{ct_sheet}'!$J${ct_start}:$J${ct_end}"
    ct_ev = f"'{ct_sheet}'!$K${ct_start}:$K${ct_end}"

    start_row = 4
    for idx, div in enumerate(DIVISIONS_ORDERED):
        r = start_row + idx
        ws.cell(row=r, column=1, value=div).number_format = "@"
        ws.cell(row=r, column=2, value=DIVISION_NAMES[div])
        # SUMIF by division
        ws.cell(row=r, column=3, value=f"=SUMIF({ct_div},A{r},{ct_budget})")
        ws.cell(row=r, column=4, value=f"=SUMIF({ct_div},A{r},{ct_committed})")
        ws.cell(row=r, column=5, value=f"=SUMIF({ct_div},A{r},{ct_actual})")
        ws.cell(row=r, column=6, value=f"=SUMIF({ct_div},A{r},{ct_variance})")
        # Weighted % complete = SUMPRODUCT(budget * pct where div matches) / SUMIF(div, budget)
        ws.cell(
            row=r, column=7,
            value=(
                f"=IFERROR(SUMPRODUCT(({ct_div}=A{r})*{ct_budget}*{ct_pct})"
                f"/SUMIF({ct_div},A{r},{ct_budget}),0)"
            ),
        )
        ws.cell(row=r, column=8, value=f"=SUMIF({ct_div},A{r},{ct_ev})")

        for c in (3, 4, 5, 6, 8):
            ws.cell(row=r, column=c).number_format = FMT_USD
        ws.cell(row=r, column=7).number_format = FMT_PCT

        for c in range(1, len(DIV_SUM_HEADERS) + 1):
            ws.cell(row=r, column=c).border = BORDER
        ws.row_dimensions[r].height = 22

    end_row = start_row + len(DIVISIONS_ORDERED) - 1

    # Totals row
    tr = end_row + 2
    ws.cell(row=tr, column=2, value="TOTALS").font = FONT_BODY_BOLD
    ws.cell(row=tr, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=tr, column=2).fill = FILL_SUBHEADER
    ws.cell(row=tr, column=2).border = BORDER
    for col_letter in ("C", "D", "E", "F", "H"):
        c_idx = ord(col_letter) - ord("A") + 1
        cell = ws.cell(row=tr, column=c_idx,
                       value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
        cell.number_format = FMT_USD
        cell.font = FONT_BODY_BOLD
        cell.fill = FILL_SUBHEADER
        cell.border = BORDER
    ws.cell(row=tr, column=7,
            value=f"=IFERROR(SUMPRODUCT(C{start_row}:C{end_row},G{start_row}:G{end_row})"
                  f"/SUM(C{start_row}:C{end_row}),0)")
    ws.cell(row=tr, column=7).number_format = FMT_PCT
    ws.cell(row=tr, column=7).font = FONT_BODY_BOLD
    ws.cell(row=tr, column=7).fill = FILL_SUBHEADER
    ws.cell(row=tr, column=7).border = BORDER

    # Conditional formatting on Total Variance (column F)
    var_range = f"F{start_row}:F{end_row}"
    ws.conditional_formatting.add(
        var_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["0"],
                   fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True)),
    )
    ws.conditional_formatting.add(
        var_range,
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=FILL_RED, font=Font(color=RED_FONT, bold=True)),
    )

    set_col_widths(ws, {
        "A": 12, "B": 32, "C": 16, "D": 16, "E": 16, "F": 16, "G": 18, "H": 18,
    })

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Tab 5 — Dashboard
# ---------------------------------------------------------------------------

def build_dashboard(wb: Workbook, ct_sheet: str, ct_start: int, ct_end: int) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Job Dashboard"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:F1")

    ws["A2"] = "Auto-calculated from Job Info and Cost Tracking. Update those tabs - this one stays in sync."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY_TEXT)
    ws.merge_cells("A2:F2")

    # ---- Summary panel ----
    panel_start = 4
    # Header band
    ws.cell(row=panel_start, column=1, value="Job Summary").font = FONT_H1
    ws.cell(row=panel_start, column=1).fill = FILL_HEADER
    ws.cell(row=panel_start, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=panel_start, start_column=1,
                   end_row=panel_start, end_column=6)
    ws.row_dimensions[panel_start].height = 26

    ct_budget = f"'{ct_sheet}'!F{ct_start}:F{ct_end}"
    ct_committed = f"'{ct_sheet}'!G{ct_start}:G{ct_end}"
    ct_actual = f"'{ct_sheet}'!H{ct_start}:H{ct_end}"
    ct_pct = f"'{ct_sheet}'!J{ct_start}:J{ct_end}"

    # (label, formula, number_format, is_variance_for_cf)
    metrics = [
        ("Contract Amount", "='Job Info'!D4", FMT_USD, False),
        ("Total Budget", f"=SUM({ct_budget})", FMT_USD, False),
        ("Total Committed", f"=SUM({ct_committed})", FMT_USD, False),
        ("Total Actual to Date", f"=SUM({ct_actual})", FMT_USD, False),
        ("Total Variance (Budget - Actual)",
         f"=SUM({ct_budget})-SUM({ct_actual})", FMT_USD, True),
        ("Weighted % Complete",
         f"=IFERROR(SUMPRODUCT({ct_budget},{ct_pct})/SUM({ct_budget}),0)",
         FMT_PCT, False),
        # Projected cost-at-completion = Actual / %Complete (linear extrapolation),
        # falls back to Total Budget when % complete is 0 to avoid div/0.
        ("Projected Cost at Completion",
         f"=IFERROR(SUM({ct_actual})/IFERROR(SUMPRODUCT({ct_budget},{ct_pct})"
         f"/SUM({ct_budget}),0),SUM({ct_budget}))",
         FMT_USD, False),
        ("Projected Profit",
         f"='Job Info'!D4-IFERROR(SUM({ct_actual})/IFERROR(SUMPRODUCT({ct_budget},{ct_pct})"
         f"/SUM({ct_budget}),0),SUM({ct_budget}))",
         FMT_USD, True),
        ("Profit Margin %",
         f"=IFERROR(('Job Info'!D4-IFERROR(SUM({ct_actual})/IFERROR(SUMPRODUCT({ct_budget},{ct_pct})"
         f"/SUM({ct_budget}),0),SUM({ct_budget})))/'Job Info'!D4,0)",
         FMT_PCT, True),
    ]

    row = panel_start + 1
    variance_cells: list[str] = []
    for label, formula, fmt, is_variance in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = FONT_BODY_BOLD
        ws.cell(row=row, column=1).fill = FILL_SUMMARY_LABEL
        ws.cell(row=row, column=1).alignment = Alignment(vertical="center", indent=1)
        ws.cell(row=row, column=1).border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

        ws.cell(row=row, column=4, value=formula)
        ws.cell(row=row, column=4).font = FONT_BIG_NUMBER
        ws.cell(row=row, column=4).number_format = fmt
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.cell(row=row, column=4).border = BORDER
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)

        if is_variance:
            variance_cells.append(f"D{row}")
        ws.row_dimensions[row].height = 26
        row += 1

    # Conditional formatting on variance/profit cells
    for cell_ref in variance_cells:
        ws.conditional_formatting.add(
            cell_ref,
            CellIsRule(operator="greaterThanOrEqual", formula=["0"],
                       fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True, size=18)),
        )
        ws.conditional_formatting.add(
            cell_ref,
            CellIsRule(operator="lessThan", formula=["0"],
                       fill=FILL_RED, font=Font(color=RED_FONT, bold=True, size=18)),
        )

    # ---- Per-division breakdown ----
    section_row = row + 2
    ws.cell(row=section_row, column=1, value="Per-Division Breakdown").font = FONT_H1
    ws.cell(row=section_row, column=1).fill = FILL_HEADER
    ws.cell(row=section_row, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=section_row, start_column=1,
                   end_row=section_row, end_column=6)
    ws.row_dimensions[section_row].height = 26

    headers = ["Div", "Division Name", "Budget", "Actual", "Variance", "% Complete"]
    hr = section_row + 1
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hr, column=i, value=h)
    style_header_row(ws, hr, len(headers))

    ct_div = f"'{ct_sheet}'!$A${ct_start}:$A${ct_end}"
    ct_budget_abs = f"'{ct_sheet}'!$F${ct_start}:$F${ct_end}"
    ct_actual_abs = f"'{ct_sheet}'!$H${ct_start}:$H${ct_end}"
    ct_variance_abs = f"'{ct_sheet}'!$I${ct_start}:$I${ct_end}"
    ct_pct_abs = f"'{ct_sheet}'!$J${ct_start}:$J${ct_end}"

    div_start = hr + 1
    for idx, div in enumerate(DIVISIONS_ORDERED):
        r = div_start + idx
        ws.cell(row=r, column=1, value=div).number_format = "@"
        ws.cell(row=r, column=2, value=DIVISION_NAMES[div])
        ws.cell(row=r, column=3, value=f"=SUMIF({ct_div},A{r},{ct_budget_abs})")
        ws.cell(row=r, column=4, value=f"=SUMIF({ct_div},A{r},{ct_actual_abs})")
        ws.cell(row=r, column=5, value=f"=SUMIF({ct_div},A{r},{ct_variance_abs})")
        ws.cell(row=r, column=6,
                value=(
                    f"=IFERROR(SUMPRODUCT(({ct_div}=A{r})*{ct_budget_abs}*{ct_pct_abs})"
                    f"/SUMIF({ct_div},A{r},{ct_budget_abs}),0)"
                ))
        for c in (3, 4, 5):
            ws.cell(row=r, column=c).number_format = FMT_USD
        ws.cell(row=r, column=6).number_format = FMT_PCT
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = BORDER
        ws.row_dimensions[r].height = 20

    div_end = div_start + len(DIVISIONS_ORDERED) - 1

    # Conditional fill on per-division variance (col E)
    ws.conditional_formatting.add(
        f"E{div_start}:E{div_end}",
        CellIsRule(operator="greaterThanOrEqual", formula=["0"],
                   fill=FILL_GREEN, font=Font(color=GREEN_FONT, bold=True)),
    )
    ws.conditional_formatting.add(
        f"E{div_start}:E{div_end}",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=FILL_RED, font=Font(color=RED_FONT, bold=True)),
    )

    # Column widths
    set_col_widths(ws, {
        "A": 8, "B": 34, "C": 18, "D": 18, "E": 18, "F": 16,
    })


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

# `Tuple` is only used as a type hint in build_cost_tracking — import here
# instead of at the top so the module still works under older Python.
from typing import Tuple  # noqa: E402


OUTPUT_PATH = "/Users/home/charles/contrpro/files/packages/business/Job_Costing_Spreadsheet.xlsx"


def build() -> str:
    wb = Workbook()

    build_instructions(wb)
    build_job_info(wb)
    ct_sheet, ct_start, ct_end = build_cost_tracking(wb)
    build_division_summary(wb, ct_sheet, ct_start, ct_end)
    build_dashboard(wb, ct_sheet, ct_start, ct_end)

    # Active tab = Dashboard on open
    wb.active = wb.sheetnames.index("Dashboard")

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
