#!/usr/bin/env python3
"""
Build ContrPro Punch List + As-Built Log (XLSX) — Plumbing trade pack.

Companion to closeout-and-inspection-compliance-guide.html.

Tabs:
  1. Instructions
  2. Punch List (open + closed in one running log)
  3. Closed Punch Archive (auto-filter view via status)
  4. As-Built Documentation Checklist
  5. O&M Manual Index
  6. Final Submissions Tracker

Output:
    /Users/home/charles/contrpro/files/packages/complete/plumbing/Punch_List_and_AsBuilt_Log.xlsx
"""
from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=BRAND_BLUE)
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

OUT = "/Users/home/charles/contrpro/files/packages/complete/plumbing/Punch_List_and_AsBuilt_Log.xlsx"


def widths(ws, cols):
    for col, w in cols:
        ws.column_dimensions[col].width = w


def title(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def header(ws, row, cols):
    for i, h in enumerate(cols):
        c = ws.cell(row=row, column=1 + i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 28


def blanks(ws, start, n_rows, n_cols, fmts=None):
    fmts = fmts or {}
    for r in range(start, start + n_rows):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c, value="")
            cell.border = BORDER
            cell.font = FONT_BODY
            cell.alignment = CENTER if c > 1 else LEFT
            if c in fmts:
                cell.number_format = fmts[c]


def build_instructions(ws):
    ws.title = "Instructions"
    widths(ws, [("A", 110)])
    ws["A1"] = "PUNCH LIST + AS-BUILT LOG — INSTRUCTIONS"
    ws["A1"].font = FONT_TITLE
    body = [
        "",
        "PURPOSE",
        "Single running log of every closeout item: punch items, as-built deliverables, O&M",
        "manual sections, and final submissions. Tabs are independent — fill what applies.",
        "",
        "PUNCH LIST WORKFLOW",
        "  1. Self-punch before GC walk. Walk every fixture + system; capture deficiencies.",
        "  2. Add each item to the Punch List tab with: location, fixture / system, description,",
        "     priority, assigned to, target date, photo reference.",
        "  3. As items clear, set Status='Closed' + record closed-date + closed-by.",
        "  4. Closed Punch Archive tab pulls a filter-view of closed items for retention payback.",
        "",
        "AS-BUILT WORKFLOW",
        "  1. During construction, foreman maintains daily redlines on the field set.",
        "  2. At closeout, transfer all redlines to a clean as-built set.",
        "  3. As-Built Documentation Checklist tracks every drawing sheet + the redline transfer",
        "     status. Don't sign off until every sheet is checked.",
        "",
        "O&M MANUAL",
        "  Use the index tab to confirm every required section is populated before binding.",
        "",
        "FINAL SUBMISSIONS",
        "  Use the Final Submissions Tracker to confirm every required document went to every",
        "  required recipient (Owner / GC / AHJ / Water Purveyor). Cleared retention depends",
        "  on this tab.",
        "",
        "DOCUMENT VERSION",
        "Punch_List_and_AsBuilt_Log.xlsx v1.0 — 2026-05-19",
        "ContrPro Complete Tier · Plumbing Trade Pack",
    ]
    for i, line in enumerate(body, start=3):
        c = ws.cell(row=i, column=1, value=line)
        if line.isupper() and line.strip() and not line.startswith(" "):
            c.font = FONT_H2
        else:
            c.font = FONT_BODY
        c.alignment = LEFT


def build_punch(ws):
    ws.title = "Punch List"
    widths(ws, [("A", 8), ("B", 18), ("C", 22), ("D", 32), ("E", 12), ("F", 18), ("G", 14), ("H", 14), ("I", 14), ("J", 20)])
    title(ws, 1, "PUNCH LIST — open + closed running log", 10)
    header(ws, 3, ["#", "Location (Bldg/Floor/Room)", "System / Fixture",
                   "Description", "Priority (H/M/L)", "Assigned To", "Target Date",
                   "Status (Open/Closed)", "Closed Date", "Photo Ref"])
    blanks(ws, 4, 80, 10, {7: "yyyy-mm-dd", 9: "yyyy-mm-dd"})

    # Status validation
    dv = DataValidation(type="list", formula1='"Open,Closed"', allow_blank=True)
    dv.add(f"H4:H{4+80-1}")
    ws.add_data_validation(dv)
    pri = DataValidation(type="list", formula1='"H,M,L"', allow_blank=True)
    pri.add(f"E4:E{4+80-1}")
    ws.add_data_validation(pri)


def build_closed_archive(ws):
    ws.title = "Closed Archive"
    widths(ws, [("A", 8), ("B", 18), ("C", 22), ("D", 32), ("E", 12), ("F", 18), ("G", 14), ("H", 14)])
    title(ws, 1, "CLOSED PUNCH ARCHIVE (filter view — paste closed rows for retention)", 8)
    header(ws, 3, ["#", "Location", "System / Fixture", "Description",
                   "Priority", "Closed By", "Closed Date", "Photo Ref"])
    blanks(ws, 4, 80, 8, {7: "yyyy-mm-dd"})


def build_asbuilt(ws):
    ws.title = "As-Built Checklist"
    widths(ws, [("A", 8), ("B", 18), ("C", 36), ("D", 14), ("E", 22), ("F", 22)])
    title(ws, 1, "AS-BUILT DOCUMENTATION CHECKLIST", 6)
    header(ws, 3, ["Sheet #", "Drawing Title", "Redlines (description)",
                   "Transferred Y/N", "Verified By", "Date"])
    standard = [
        ("P-0.0", "Plumbing Cover / Index", ""),
        ("P-1.0", "Underground Sanitary Plan", ""),
        ("P-1.1", "Underground Storm Plan", ""),
        ("P-1.2", "Underground Water + Gas Plan", ""),
        ("P-2.0", "First-Floor Plumbing", ""),
        ("P-2.1", "Second-Floor Plumbing", ""),
        ("P-2.2", "Roof Plumbing Plan", ""),
        ("P-3.0", "Sanitary Riser Diagrams", ""),
        ("P-3.1", "Water Riser Diagrams", ""),
        ("P-3.2", "Gas Riser + Sizing", ""),
        ("P-4.0", "Plumbing Details", ""),
        ("P-5.0", "Plumbing Schedules + Fixture Cuts", ""),
        ("P-6.0", "Specifications Volume", ""),
    ]
    for i, (sht, title_txt, redline) in enumerate(standard):
        r = 4 + i
        ws.cell(row=r, column=1, value=sht).alignment = CENTER
        ws.cell(row=r, column=2, value=title_txt).alignment = LEFT
        ws.cell(row=r, column=3, value=redline).alignment = LEFT
        ws.cell(row=r, column=4, value="").alignment = CENTER
        ws.cell(row=r, column=5, value="").alignment = LEFT
        ws.cell(row=r, column=6, value="").number_format = "yyyy-mm-dd"
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).font = FONT_BODY
    # Append blanks for project-specific sheets
    blanks(ws, 4 + len(standard), 20, 6, {6: "yyyy-mm-dd"})


def build_om(ws):
    ws.title = "O&M Index"
    widths(ws, [("A", 8), ("B", 32), ("C", 36), ("D", 14), ("E", 22)])
    title(ws, 1, "O&M MANUAL — REQUIRED SECTION INDEX", 5)
    header(ws, 3, ["Tab", "Section", "Contents", "Populated Y/N", "Verified By"])
    sections = [
        (1, "Cover Sheet", "Project name / address / contractor / version"),
        (2, "Table of Contents", "Sectional TOC + page numbers"),
        (3, "System Overview", "Schematic + narrative description"),
        (4, "Equipment Data", "Cut sheets — water heaters / boosters / pumps / interceptors / BFPs"),
        (5, "Approved Submittals", "Every approved submittal in CSI order"),
        (6, "Mfr O&M Booklets", "Original manufacturer booklets"),
        (7, "Test Reports", "Pressure / DWV / gas test logs + disinfection cert + BFP initial tests"),
        (8, "As-Builts", "Reduced + full-size in rear pocket"),
        (9, "Warranties", "Contractor warranty letter + mfr warranty cards"),
        (10, "Maintenance Schedule", "Recommended PM intervals + tasks"),
        (11, "Vendor Contacts", "Service techs / suppliers / water purveyor"),
        (12, "Emergency Procedures", "Shutoffs / contacts / backup vendor"),
    ]
    for i, (tab, sec, contents) in enumerate(sections):
        r = 4 + i
        ws.cell(row=r, column=1, value=tab).alignment = CENTER
        ws.cell(row=r, column=2, value=sec).alignment = LEFT
        ws.cell(row=r, column=3, value=contents).alignment = LEFT
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).font = FONT_BODY


def build_submissions(ws):
    ws.title = "Final Submissions"
    widths(ws, [("A", 32), ("B", 22), ("C", 16), ("D", 16), ("E", 22)])
    title(ws, 1, "FINAL SUBMISSIONS TRACKER", 5)
    header(ws, 3, ["Document", "Recipient", "Sent Date", "Confirmed", "Notes"])
    items = [
        "Pressure test log",
        "DWV / stack test results",
        "Gas pressure test report",
        "Backflow initial test reports",
        "Disinfection certificate",
        "Final as-built drawings (paper)",
        "Final as-built drawings (PDF)",
        "Final as-built drawings (CAD)",
        "O&M Manual (binder)",
        "O&M Manual (PDF/USB)",
        "Warranty letter",
        "Equipment registrations",
        "Conditional final lien waiver",
        "Unconditional final lien waiver",
        "Punch list closeout sign-off",
    ]
    for i, doc in enumerate(items):
        r = 4 + i
        ws.cell(row=r, column=1, value=doc).alignment = LEFT
        ws.cell(row=r, column=2, value="").alignment = LEFT
        ws.cell(row=r, column=3, value="").number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=4, value="").alignment = CENTER
        ws.cell(row=r, column=5, value="").alignment = LEFT
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).font = FONT_BODY
    # Blank rows for project-specific extras
    blanks(ws, 4 + len(items), 10, 5, {3: "yyyy-mm-dd"})


def main():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb = Workbook()
    build_instructions(wb.active)
    build_punch(wb.create_sheet("Punch List"))
    build_closed_archive(wb.create_sheet("Closed Archive"))
    build_asbuilt(wb.create_sheet("As-Built Checklist"))
    build_om(wb.create_sheet("O&M Index"))
    build_submissions(wb.create_sheet("Final Submissions"))
    wb.save(OUT)
    print(f"OK — wrote {OUT}")


if __name__ == "__main__":
    main()
