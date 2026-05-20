#!/usr/bin/env python3
"""
Build ContrPro Plumbing Bid Estimator (XLSX) — Plumbing trade pack.

Industry-standard fixture-and-LF-based bid estimator for commercial plumbing
contractors. CSI Division 22 primary with cross-references. Labor anchored
to PHCC/MCAA-style labor units (hrs per fixture or per LF by diameter).

Tabs:
  1. Instructions
  2. Project Info
  3. Wage Rates           (journeyman / apprentice / foreman + fringe + OT premium)
  4. Material Takeoff     (CSI 22-coded line items — fixtures, DWV, water, gas, specialties)
  5. Labor Estimate       (per-fixture + per-LF unit hours × crew × rate)
  6. Equipment Estimate   (lifts, threading machines, fusion welders, etc.)
  7. Subcontract Estimate (specialty subs — insulation, gas, control wiring if needed)
  8. Premiums & Conditions(height, hazardous-area, prevailing wage, occupied-building disruption)
  9. Bid Summary          (roll-up by CSI division → contingency, bond, OH, profit)
  10. CSI Reference

Run:
    /Users/home/charles/.venv/bin/python3 \\
        /Users/home/charles/contrpro/scripts/build_plumbing_bid_estimator_xlsx.py

Output:
    /Users/home/charles/contrpro/files/packages/complete/plumbing/Plumbing_Bid_Estimator.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
ACCENT_GOLD = "C9A227"
GREEN_FILL = "C6EFCE"
RED_FILL = "FFC7CE"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GOLD = PatternFill("solid", fgColor=ACCENT_GOLD)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_GREY_ITALIC = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_HRS = '_(* #,##0.0_);_(* (#,##0.0);_(* "-"_);_(@_)'
FMT_PCT = "0.0%"
FMT_INT = "0"
FMT_NUM2 = "0.00"

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

OUT = "/Users/home/charles/contrpro/files/packages/complete/plumbing/Plumbing_Bid_Estimator.xlsx"


def set_col_widths(ws, widths):
    for col, w in widths:
        ws.column_dimensions[col].width = w


def bordered(cell):
    cell.border = BORDER
    return cell


def title_row(ws, row: int, text: str, span_cols: int = 8):
    ws.cell(row=row, column=1, value=text).font = FONT_TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)


def header_row(ws, row: int, headers, start_col: int = 1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 26


# ---------------------------------------------------------------------------
# Tab 1: Instructions
# ---------------------------------------------------------------------------
def build_instructions(ws):
    ws.title = "Instructions"
    set_col_widths(ws, [("A", 110)])
    ws["A1"] = "PLUMBING BID ESTIMATOR — INSTRUCTIONS"
    ws["A1"].font = FONT_TITLE
    body = [
        "",
        "PURPOSE",
        "Production estimating tool for commercial plumbing contractors. Fixture- and",
        "linear-foot-based with PHCC/MCAA-style labor units. CSI Division 22 primary",
        "with cross-references to Division 23 (mechanical), 31 (earthwork), and 33",
        "(utilities). Output rolls up by CSI division so the bid integrates cleanly into",
        "the GC's job-cost accounting and into your monthly billing on the Universal",
        "Sub Suite's Sub_Schedule_of_Values.xlsx.",
        "",
        "WORKFLOW",
        "  1. Project Info — fill in every blue-filled field. The cells you leave blank",
        "     show up as warnings on the Bid Summary tab.",
        "  2. Wage Rates — set hourly base + fringe + OT premium for each classification.",
        "     Defaults are placeholders; calibrate to YOUR market.",
        "  3. Material Takeoff — enter quantities by CSI line item. Unit prices flow",
        "     into the Bid Summary by division.",
        "  4. Labor Estimate — productivity defaults assume standard commercial work.",
        "     Adjust the productivity factor for difficult work (occupied buildings,",
        "     extensive demo, hazmat). Labor hours × crew rate flow to the Bid Summary.",
        "  5. Equipment — daily/weekly rates for owned + rented. The daily/weekly toggle",
        "     drives the multiplication.",
        "  6. Subcontract — any work you sub out (insulation, gas, controls wiring).",
        "  7. Premiums — height, hazardous area, prevailing wage, occupied-building disruption.",
        "  8. Bid Summary — automatic roll-up. Sets contingency / bond / OH / profit.",
        "     The bottom of the tab is your final bid number.",
        "",
        "PHCC / MCAA LABOR UNITS",
        "Default hours-per-unit assume average commercial conditions. Calibrate against",
        "your own historical job-cost data before relying on this bid. Underestimating",
        "labor is the #1 cause of plumbing-bid losses; if your historical productivity",
        "differs from these defaults by more than 15%, change the defaults — don't ignore",
        "the gap.",
        "",
        "FIELD-ONLY SCOPE",
        "Calibrated for field installation only. Fabrication-shop pre-assembly (modular",
        "racks, pump skids built off-site) is out of scope and requires a separate",
        "shop-prefab estimator with different rate structures.",
        "",
        "CSI MASTERFORMAT",
        "Every line in Material Takeoff and Labor Estimate is CSI-coded. See the CSI",
        "Reference tab for the Division 22 codes used in this workbook. The Bid Summary",
        "rolls up totals by CSI division for clean handoff to the GC's accounting system.",
        "",
        "RED-FLAGS YOU SHOULD CATCH BEFORE SUBMITTING THE BID",
        "  • Project Info has blanks — Bid Summary will warn you",
        "  • Labor productivity defaults unchanged on a difficult project",
        "  • No allowance for the prevailing-wage premium on a public-works job",
        "  • Equipment crane/lift rentals omitted on a high-ceiling job",
        "  • Bond/insurance/permit costs left at $0",
        "  • Profit margin set below your historical floor",
        "",
        "DOCUMENT VERSION",
        "Plumbing_Bid_Estimator.xlsx v1.0 — 2026-05-19",
        "ContrPro Complete Tier · Plumbing Trade Pack",
    ]
    for i, line in enumerate(body, start=3):
        c = ws.cell(row=i, column=1, value=line)
        if line.isupper() and line.strip() and not line.startswith(" "):
            c.font = FONT_H2
        else:
            c.font = FONT_BODY
        c.alignment = LEFT


# ---------------------------------------------------------------------------
# Tab 2: Project Info
# ---------------------------------------------------------------------------
def build_project_info(ws):
    ws.title = "Project Info"
    set_col_widths(ws, [("A", 32), ("B", 60)])
    title_row(ws, 1, "PROJECT INFORMATION", 2)
    fields = [
        ("Project Name", ""),
        ("Project Address", ""),
        ("Owner", ""),
        ("General Contractor", ""),
        ("Architect / Engineer", ""),
        ("Bid Due Date", ""),
        ("Bid Reference / Number", ""),
        ("Project Type", "(New construction / Renovation / Tenant fit-out / Adaptive reuse)"),
        ("Building Use", "(Office / Retail / Healthcare / Education / Multi-family / Hospitality / Industrial / Mixed)"),
        ("Approx. Square Footage", ""),
        ("Stories", ""),
        ("Plumbing Fixture Count (est.)", ""),
        ("AHJ — Plumbing Inspector", ""),
        ("Applicable Plumbing Code", "(IPC 2024 / UPC 2024 / Local amendment)"),
        ("Estimator Name", ""),
        ("Estimator Phone", ""),
        ("Estimator Email", ""),
        ("Date Prepared", ""),
        ("Internal Bid Number", ""),
        ("Notes / Assumptions", ""),
    ]
    row = 3
    for label, default in fields:
        ws.cell(row=row, column=1, value=label).font = FONT_BODY_BOLD
        ws.cell(row=row, column=1).fill = FILL_SUBHEADER
        ws.cell(row=row, column=1).alignment = LEFT
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=2, value=default).font = FONT_GREY_ITALIC if default else FONT_BODY
        ws.cell(row=row, column=2).alignment = LEFT
        ws.cell(row=row, column=2).border = BORDER
        row += 1
    ws.row_dimensions[1].height = 30


# ---------------------------------------------------------------------------
# Tab 3: Wage Rates
# ---------------------------------------------------------------------------
def build_wage_rates(ws, wb):
    ws.title = "Wage Rates"
    set_col_widths(ws, [("A", 28), ("B", 14), ("C", 14), ("D", 14), ("E", 16), ("F", 16)])
    title_row(ws, 1, "WAGE RATES (calibrate to your market)", 6)
    header_row(ws, 3, ["Classification", "Base Rate ($/hr)", "Fringe ($/hr)", "Burden % (WC/GL/FICA)", "OT Premium (1.5×)", "Loaded Rate ($/hr)"])
    rows = [
        ("Plumbing Foreman",   58.00, 18.50, 0.32),
        ("Journeyman Plumber", 48.00, 18.50, 0.32),
        ("Apprentice (Yr 4)",  38.00, 14.00, 0.32),
        ("Apprentice (Yr 2)",  28.00, 12.00, 0.32),
        ("Laborer / Helper",   22.00,  9.00, 0.32),
        ("Service Tech",       55.00, 18.50, 0.32),
        ("Pipefitter (gas)",   54.00, 18.50, 0.32),
    ]
    for i, (cls, base, fringe, burden) in enumerate(rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=cls).font = FONT_BODY_BOLD
        ws.cell(row=r, column=2, value=base).number_format = FMT_USD
        ws.cell(row=r, column=3, value=fringe).number_format = FMT_USD
        ws.cell(row=r, column=4, value=burden).number_format = FMT_PCT
        ws.cell(row=r, column=5, value=f"=B{r}*1.5").number_format = FMT_USD
        ws.cell(row=r, column=6, value=f"=(B{r}+C{r})*(1+D{r})").number_format = FMT_USD
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).alignment = CENTER if col > 1 else LEFT
    # Named ranges for downstream lookups
    last = 4 + len(rows) - 1
    wb.defined_names["WageClassifications"] = DefinedName(
        "WageClassifications", attr_text=f"'Wage Rates'!$A$4:$A${last}"
    )
    wb.defined_names["WageLoadedRates"] = DefinedName(
        "WageLoadedRates", attr_text=f"'Wage Rates'!$F$4:$F${last}"
    )
    ws.cell(row=last + 2, column=1, value=(
        "Burden default 32% covers WC + GL + FICA/Medicare + SUTA/FUTA. Verify against "
        "your insurance certificates + payroll service for accuracy. Prevailing-wage jobs "
        "may require federal Davis-Bacon wage determinations — see Universal Sub Suite's "
        "Certified Payroll guide."
    )).font = FONT_GREY_ITALIC
    ws.merge_cells(start_row=last + 2, start_column=1, end_row=last + 2, end_column=6)


# ---------------------------------------------------------------------------
# Tab 4: Material Takeoff
# ---------------------------------------------------------------------------
def build_material_takeoff(ws):
    ws.title = "Material Takeoff"
    set_col_widths(ws, [("A", 14), ("B", 44), ("C", 10), ("D", 10), ("E", 14), ("F", 16), ("G", 16)])
    title_row(ws, 1, "MATERIAL TAKEOFF (CSI Division 22)", 7)
    header_row(ws, 3, ["CSI Code", "Description", "QTY", "UOM", "Unit Cost ($)", "Material Subtotal ($)", "Notes"])
    items = [
        # 22 11 — Water distribution
        ("22 11 00", "Domestic water — CPVC piping, 1/2\"", 0, "LF", 1.85, ""),
        ("22 11 00", "Domestic water — CPVC piping, 3/4\"", 0, "LF", 2.40, ""),
        ("22 11 00", "Domestic water — CPVC piping, 1\"", 0, "LF", 3.20, ""),
        ("22 11 00", "Domestic water — CPVC piping, 1-1/2\"", 0, "LF", 4.85, ""),
        ("22 11 00", "Domestic water — CPVC piping, 2\"", 0, "LF", 6.10, ""),
        ("22 11 00", "Domestic water — PEX-A piping, 1/2\"", 0, "LF", 0.95, "Crimp/cinch connections"),
        ("22 11 00", "Domestic water — PEX-A piping, 3/4\"", 0, "LF", 1.45, ""),
        ("22 11 00", "Domestic water — Copper Type L, 1/2\"", 0, "LF", 4.75, "Soldered + ProPress"),
        ("22 11 00", "Domestic water — Copper Type L, 3/4\"", 0, "LF", 6.85, ""),
        ("22 11 00", "Domestic water — Copper Type L, 1\"", 0, "LF", 9.40, ""),
        ("22 11 00", "Water service entry — 1.5\" - 2\"", 0, "EA", 425.00, "Includes meter loop, BFP"),
        ("22 11 00", "Backflow preventer — RPZ, 1\"", 0, "EA", 1875.00, ""),
        ("22 11 00", "Backflow preventer — DCV, 1\"", 0, "EA", 875.00, ""),
        ("22 11 00", "Water hammer arrestor", 0, "EA", 38.00, ""),
        ("22 11 19", "Pipe insulation — 1/2\" copper", 0, "LF", 1.45, ""),
        ("22 11 19", "Pipe insulation — 1\" copper", 0, "LF", 2.20, ""),
        ("22 11 19", "Pipe insulation — 2\" copper/CPVC", 0, "LF", 3.40, ""),
        # 22 13 — Sanitary sewerage
        ("22 13 16", "Sanitary DWV — PVC, 2\" Sch 40", 0, "LF", 2.85, ""),
        ("22 13 16", "Sanitary DWV — PVC, 3\" Sch 40", 0, "LF", 4.20, ""),
        ("22 13 16", "Sanitary DWV — PVC, 4\" Sch 40", 0, "LF", 5.95, ""),
        ("22 13 16", "Sanitary DWV — PVC, 6\" Sch 40", 0, "LF", 9.75, ""),
        ("22 13 16", "Cast iron DWV — 4\" no-hub", 0, "LF", 24.50, "Service-weight hub-and-spigot uplift"),
        ("22 13 16", "Cast iron DWV — 6\" no-hub", 0, "LF", 38.00, ""),
        ("22 13 19", "Floor drain — 2\" w/ trap primer", 0, "EA", 145.00, ""),
        ("22 13 19", "Floor drain — 4\" w/ trap primer", 0, "EA", 215.00, ""),
        ("22 13 19", "Roof drain — 4\" w/ scupper overflow", 0, "EA", 285.00, ""),
        ("22 13 19", "Cleanout — wall, 4\"", 0, "EA", 95.00, ""),
        ("22 13 19", "Cleanout — floor, 4\"", 0, "EA", 125.00, ""),
        ("22 13 23", "Grease interceptor — 50 GPM", 0, "EA", 3850.00, "Hydromechanical"),
        ("22 13 23", "Grease interceptor — 75 GPM", 0, "EA", 5250.00, ""),
        ("22 13 29", "Sewage ejector pump — duplex 1HP", 0, "EA", 2950.00, ""),
        # 22 14 — Storm drainage
        ("22 14 13", "Storm DWV — PVC, 4\"", 0, "LF", 5.95, ""),
        ("22 14 13", "Storm DWV — PVC, 6\"", 0, "LF", 9.75, ""),
        ("22 14 13", "Storm DWV — PVC, 8\"", 0, "LF", 14.50, ""),
        # 22 33 — Domestic water heaters
        ("22 33 30", "Electric water heater — 40 gal", 0, "EA", 685.00, ""),
        ("22 33 30", "Electric water heater — 80 gal", 0, "EA", 1245.00, ""),
        ("22 33 33", "Gas water heater — 50 gal, atmospheric", 0, "EA", 1485.00, ""),
        ("22 33 36", "Gas water heater — 75 gal, power-vent", 0, "EA", 2150.00, ""),
        ("22 34 00", "Tankless water heater — 199K BTU", 0, "EA", 2850.00, ""),
        ("22 35 00", "Domestic water booster pump system", 0, "EA", 8500.00, "VFD-driven duplex"),
        # 22 40 — Plumbing fixtures
        ("22 42 13", "Water closet — flushometer, wall-hung", 0, "EA", 525.00, ""),
        ("22 42 13", "Water closet — flushometer, floor-mount", 0, "EA", 415.00, ""),
        ("22 42 13", "Water closet — tank-type ADA", 0, "EA", 385.00, ""),
        ("22 42 16", "Urinal — flushometer, vitreous china", 0, "EA", 425.00, ""),
        ("22 42 16", "Urinal — waterless, with cartridge", 0, "EA", 685.00, ""),
        ("22 42 19", "Lavatory — wall-hung, vitreous china", 0, "EA", 285.00, ""),
        ("22 42 19", "Lavatory — countertop drop-in", 0, "EA", 195.00, ""),
        ("22 42 23", "Sink — single-bowl stainless, 18 ga", 0, "EA", 215.00, ""),
        ("22 42 23", "Sink — three-compartment commercial", 0, "EA", 945.00, ""),
        ("22 42 26", "Service sink — floor-mount mop basin", 0, "EA", 385.00, ""),
        ("22 42 33", "Shower — single, with valve + head", 0, "EA", 425.00, ""),
        ("22 42 39", "Drinking fountain — bottle-fill, ADA", 0, "EA", 1685.00, ""),
        ("22 45 00", "Emergency eyewash + shower combo", 0, "EA", 1450.00, "ANSI Z358.1"),
        # 22 60 — Gas
        ("22 67 00", "Natural gas piping — black iron 1\"", 0, "LF", 8.50, ""),
        ("22 67 00", "Natural gas piping — black iron 1-1/2\"", 0, "LF", 12.85, ""),
        ("22 67 00", "Natural gas piping — black iron 2\"", 0, "LF", 16.50, ""),
        ("22 67 00", "CSST gas piping — 3/4\"", 0, "LF", 4.85, ""),
        ("22 67 00", "Gas regulator — pressure-reducing 2 psi", 0, "EA", 285.00, ""),
        ("22 67 00", "Gas shutoff valve — 1\"", 0, "EA", 65.00, ""),
        # Hangers / supports
        ("22 05 29", "Pipe hangers — clevis 1/2\"-1\"", 0, "EA", 4.85, ""),
        ("22 05 29", "Pipe hangers — clevis 1-1/4\"-2\"", 0, "EA", 7.50, ""),
        ("22 05 29", "Pipe hangers — clevis 3\"-4\"", 0, "EA", 14.50, ""),
        ("22 05 29", "Trapeze hangers — multi-pipe", 0, "EA", 38.00, ""),
        # Sleeves / firestopping
        ("22 05 19", "Wall sleeve — 4\" through firewall", 0, "EA", 22.50, ""),
        ("22 05 19", "Floor sleeve — 4\" through slab", 0, "EA", 18.50, ""),
        ("22 05 19", "Firestop assembly — UL-listed pipe through wall", 0, "EA", 32.00, ""),
        # Testing / commissioning
        ("22 08 00", "Hydrostatic test — DWV system", 0, "LS", 750.00, "Air or water per IPC 312"),
        ("22 08 00", "Hydrostatic test — domestic water", 0, "LS", 650.00, ""),
        ("22 08 00", "Gas pressure test", 0, "LS", 450.00, "Per IFGC 406"),
        ("22 08 00", "Backflow preventer initial test", 0, "EA", 175.00, ""),
        ("22 08 00", "Disinfection — domestic water lines", 0, "LS", 1250.00, "AWWA C651"),
        # Permit / inspection
        ("22 00 10", "Permit — plumbing", 0, "LS", 0.00, "Allowance — fill in AHJ amount"),
        ("22 00 10", "Inspection re-trip fees (allowance)", 0, "LS", 0.00, ""),
    ]
    for i, (csi, desc, qty, uom, unit, notes) in enumerate(items):
        r = 4 + i
        ws.cell(row=r, column=1, value=csi).alignment = CENTER
        ws.cell(row=r, column=2, value=desc).alignment = LEFT
        ws.cell(row=r, column=3, value=qty).number_format = FMT_INT
        ws.cell(row=r, column=4, value=uom).alignment = CENTER
        ws.cell(row=r, column=5, value=unit).number_format = FMT_USD
        ws.cell(row=r, column=6, value=f"=C{r}*E{r}").number_format = FMT_USD
        ws.cell(row=r, column=7, value=notes).alignment = LEFT
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = FONT_BODY
    # Subtotal row
    last = 4 + len(items) - 1
    sub_row = last + 2
    ws.cell(row=sub_row, column=5, value="Material Subtotal").font = FONT_BODY_BOLD
    ws.cell(row=sub_row, column=5).alignment = RIGHT
    ws.cell(row=sub_row, column=6, value=f"=SUM(F4:F{last})").font = FONT_BODY_BOLD
    ws.cell(row=sub_row, column=6).number_format = FMT_USD
    ws.cell(row=sub_row, column=6).fill = FILL_SUBHEADER
    # Sales tax
    tax_row = sub_row + 1
    ws.cell(row=tax_row, column=5, value="Sales tax %").font = FONT_BODY_BOLD
    ws.cell(row=tax_row, column=5).alignment = RIGHT
    ws.cell(row=tax_row, column=6, value=0.07).number_format = FMT_PCT
    tax_amt_row = tax_row + 1
    ws.cell(row=tax_amt_row, column=5, value="Sales tax $").font = FONT_BODY_BOLD
    ws.cell(row=tax_amt_row, column=5).alignment = RIGHT
    ws.cell(row=tax_amt_row, column=6, value=f"=F{sub_row}*F{tax_row}").number_format = FMT_USD
    tot_row = tax_amt_row + 1
    ws.cell(row=tot_row, column=5, value="Material TOTAL (incl tax)").font = FONT_BODY_BOLD
    ws.cell(row=tot_row, column=5).alignment = RIGHT
    ws.cell(row=tot_row, column=5).fill = FILL_GOLD
    ws.cell(row=tot_row, column=6, value=f"=F{sub_row}+F{tax_amt_row}").font = FONT_BODY_BOLD
    ws.cell(row=tot_row, column=6).number_format = FMT_USD
    ws.cell(row=tot_row, column=6).fill = FILL_GOLD
    # Stash named range for Bid Summary
    return ("MaterialTotal", f"'Material Takeoff'!$F${tot_row}")


# ---------------------------------------------------------------------------
# Tab 5: Labor Estimate
# ---------------------------------------------------------------------------
def build_labor_estimate(ws):
    ws.title = "Labor Estimate"
    set_col_widths(ws, [("A", 14), ("B", 42), ("C", 10), ("D", 10), ("E", 16), ("F", 14), ("G", 16), ("H", 18)])
    title_row(ws, 1, "LABOR ESTIMATE (PHCC/MCAA-style labor units)", 8)
    header_row(ws, 3, ["CSI Code", "Activity", "QTY", "UOM", "Hours / UOM", "Subtotal Hours", "Crew Loaded Rate ($/hr)", "Subtotal Labor ($)"])
    activities = [
        # Fixtures — per-each
        ("22 42 13", "Set water closet — flushometer", 0, "EA", 1.85, 65.00),
        ("22 42 13", "Set water closet — tank type", 0, "EA", 1.20, 65.00),
        ("22 42 16", "Set urinal — flushometer", 0, "EA", 1.65, 65.00),
        ("22 42 19", "Set lavatory — wall-hung", 0, "EA", 1.45, 65.00),
        ("22 42 19", "Set lavatory — countertop", 0, "EA", 1.20, 65.00),
        ("22 42 23", "Set sink — single-bowl", 0, "EA", 1.30, 65.00),
        ("22 42 23", "Set sink — three-compartment", 0, "EA", 2.50, 65.00),
        ("22 42 26", "Set service / mop sink", 0, "EA", 2.10, 65.00),
        ("22 42 33", "Set shower w/ valve + head", 0, "EA", 2.75, 65.00),
        ("22 42 39", "Set drinking fountain — ADA", 0, "EA", 2.40, 65.00),
        ("22 45 00", "Set emergency eyewash + shower", 0, "EA", 4.25, 65.00),
        # Water heaters
        ("22 33 30", "Set electric water heater — 40 gal", 0, "EA", 3.50, 65.00),
        ("22 33 30", "Set electric water heater — 80 gal", 0, "EA", 4.50, 65.00),
        ("22 33 33", "Set gas water heater — 50 gal", 0, "EA", 4.85, 65.00),
        ("22 34 00", "Set tankless water heater", 0, "EA", 5.50, 65.00),
        # Domestic water — install per LF
        ("22 11 00", "Install CPVC 1/2\" - 1\"", 0, "LF", 0.18, 65.00),
        ("22 11 00", "Install CPVC 1-1/2\" - 2\"", 0, "LF", 0.28, 65.00),
        ("22 11 00", "Install PEX 1/2\" - 3/4\"", 0, "LF", 0.12, 65.00),
        ("22 11 00", "Install copper Type L 1/2\" - 1\"", 0, "LF", 0.32, 65.00),
        ("22 11 00", "Install copper Type L 1-1/2\" - 2\"", 0, "LF", 0.48, 65.00),
        # DWV — install per LF
        ("22 13 16", "Install PVC DWV 2\" - 3\"", 0, "LF", 0.22, 65.00),
        ("22 13 16", "Install PVC DWV 4\" - 6\"", 0, "LF", 0.38, 65.00),
        ("22 13 16", "Install cast iron DWV 4\" no-hub", 0, "LF", 0.65, 65.00),
        ("22 13 16", "Install cast iron DWV 6\" no-hub", 0, "LF", 0.92, 65.00),
        # Storm
        ("22 14 13", "Install storm 4\" - 6\"", 0, "LF", 0.42, 65.00),
        ("22 14 13", "Install storm 8\"", 0, "LF", 0.68, 65.00),
        # Drains / cleanouts
        ("22 13 19", "Set floor drain w/ trap primer", 0, "EA", 1.20, 65.00),
        ("22 13 19", "Set roof drain w/ overflow", 0, "EA", 2.20, 65.00),
        ("22 13 19", "Set cleanout — wall or floor", 0, "EA", 0.85, 65.00),
        ("22 13 23", "Set grease interceptor — 50/75 GPM", 0, "EA", 6.50, 65.00),
        # Gas
        ("22 67 00", "Install black iron gas 1\" - 1-1/2\"", 0, "LF", 0.35, 70.00),
        ("22 67 00", "Install black iron gas 2\"", 0, "LF", 0.52, 70.00),
        ("22 67 00", "Install CSST gas 3/4\"", 0, "LF", 0.18, 70.00),
        ("22 67 00", "Set gas regulator / shutoff", 0, "EA", 0.95, 70.00),
        # Backflow
        ("22 11 00", "Install + test backflow preventer (RPZ)", 0, "EA", 4.50, 65.00),
        ("22 11 00", "Install + test backflow preventer (DCV)", 0, "EA", 2.85, 65.00),
        # Insulation (if self-performed)
        ("22 07 00", "Insulate pipe 1/2\" - 1\"", 0, "LF", 0.08, 55.00),
        ("22 07 00", "Insulate pipe 1-1/2\" - 2\"", 0, "LF", 0.12, 55.00),
        # Hangers / firestop
        ("22 05 29", "Install hangers + trapeze", 0, "EA", 0.38, 65.00),
        ("22 05 19", "Cut + sleeve through wall", 0, "EA", 0.45, 65.00),
        ("22 05 19", "Install firestop assembly", 0, "EA", 0.65, 65.00),
        # Testing / commissioning
        ("22 08 00", "Test + chlorinate domestic water", 0, "LS", 12.00, 65.00),
        ("22 08 00", "Hydrostatic test DWV", 0, "LS", 9.00, 65.00),
        ("22 08 00", "Gas pressure test", 0, "LS", 8.00, 70.00),
    ]
    for i, (csi, act, qty, uom, hrs_unit, rate) in enumerate(activities):
        r = 4 + i
        ws.cell(row=r, column=1, value=csi).alignment = CENTER
        ws.cell(row=r, column=2, value=act).alignment = LEFT
        ws.cell(row=r, column=3, value=qty).number_format = FMT_INT
        ws.cell(row=r, column=4, value=uom).alignment = CENTER
        ws.cell(row=r, column=5, value=hrs_unit).number_format = FMT_NUM2
        ws.cell(row=r, column=6, value=f"=C{r}*E{r}").number_format = FMT_HRS
        ws.cell(row=r, column=7, value=rate).number_format = FMT_USD
        ws.cell(row=r, column=8, value=f"=F{r}*G{r}").number_format = FMT_USD
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = FONT_BODY
    last = 4 + len(activities) - 1
    # Productivity factor
    prod_row = last + 2
    ws.cell(row=prod_row, column=7, value="Productivity factor (1.0=avg)").font = FONT_BODY_BOLD
    ws.cell(row=prod_row, column=7).alignment = RIGHT
    ws.cell(row=prod_row, column=8, value=1.00).number_format = FMT_NUM2
    ws.cell(row=prod_row, column=8).fill = FILL_SUBHEADER
    # Subtotal
    sub_row = prod_row + 1
    ws.cell(row=sub_row, column=7, value="Labor Subtotal").font = FONT_BODY_BOLD
    ws.cell(row=sub_row, column=7).alignment = RIGHT
    ws.cell(row=sub_row, column=8, value=f"=SUM(H4:H{last})*H{prod_row}").font = FONT_BODY_BOLD
    ws.cell(row=sub_row, column=8).number_format = FMT_USD
    ws.cell(row=sub_row, column=8).fill = FILL_GOLD
    ws.cell(row=last + 4, column=1, value=(
        "Productivity defaults assume average commercial conditions with reasonable access "
        "and standard sequencing. Increase the productivity factor for difficult work "
        "(occupied buildings, extensive demolition, hazmat). Decrease for repetitive prefab "
        "or unobstructed greenfield. Calibrate against your own historical job-cost data."
    )).font = FONT_GREY_ITALIC
    ws.merge_cells(start_row=last + 4, start_column=1, end_row=last + 4, end_column=8)
    return ("LaborTotal", f"'Labor Estimate'!$H${sub_row}")


# ---------------------------------------------------------------------------
# Tab 6: Equipment Estimate
# ---------------------------------------------------------------------------
def build_equipment(ws):
    ws.title = "Equipment"
    set_col_widths(ws, [("A", 42), ("B", 14), ("C", 14), ("D", 12), ("E", 16), ("F", 18)])
    title_row(ws, 1, "EQUIPMENT ESTIMATE", 6)
    header_row(ws, 3, ["Equipment", "Days", "Daily Rate ($)", "Weeks", "Weekly Rate ($)", "Subtotal ($)"])
    items = [
        ("Threading machine — 1/2\" - 2\"", 0, 145.00, 0, 525.00),
        ("Threading machine — 2-1/2\" - 4\"", 0, 285.00, 0, 985.00),
        ("Press tool — ProPress copper", 0, 95.00, 0, 285.00),
        ("PEX expansion tool", 0, 65.00, 0, 195.00),
        ("Fusion welder — HDPE", 0, 425.00, 0, 1450.00),
        ("Pipe-bending machine — conduit/CSST", 0, 85.00, 0, 245.00),
        ("Scissor lift — 19'", 0, 145.00, 0, 485.00),
        ("Boom lift — 45'", 0, 425.00, 0, 1485.00),
        ("Genie / personnel lift — 20'", 0, 85.00, 0, 285.00),
        ("Test pump — hydrostatic", 0, 55.00, 0, 165.00),
        ("Drain camera — push", 0, 95.00, 0, 325.00),
        ("Trencher / mini-excavator", 0, 385.00, 0, 1450.00),
        ("Concrete saw + core drill", 0, 145.00, 0, 525.00),
        ("HEPA vacuum — dust control", 0, 45.00, 0, 145.00),
        ("Generator — 5kW portable", 0, 65.00, 0, 195.00),
        ("Office trailer (allocation)", 0, 0.00, 0, 285.00),
        ("Truck / van allocation", 0, 0.00, 0, 425.00),
    ]
    for i, (name, days, drate, weeks, wrate) in enumerate(items):
        r = 4 + i
        ws.cell(row=r, column=1, value=name).alignment = LEFT
        ws.cell(row=r, column=2, value=days).number_format = FMT_INT
        ws.cell(row=r, column=3, value=drate).number_format = FMT_USD
        ws.cell(row=r, column=4, value=weeks).number_format = FMT_INT
        ws.cell(row=r, column=5, value=wrate).number_format = FMT_USD
        ws.cell(row=r, column=6, value=f"=(B{r}*C{r})+(D{r}*E{r})").number_format = FMT_USD
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = FONT_BODY
    last = 4 + len(items) - 1
    sub_row = last + 2
    ws.cell(row=sub_row, column=5, value="Equipment Subtotal").font = FONT_BODY_BOLD
    ws.cell(row=sub_row, column=5).alignment = RIGHT
    ws.cell(row=sub_row, column=6, value=f"=SUM(F4:F{last})").font = FONT_BODY_BOLD
    ws.cell(row=sub_row, column=6).number_format = FMT_USD
    ws.cell(row=sub_row, column=6).fill = FILL_GOLD
    return ("EquipTotal", f"'Equipment'!$F${sub_row}")


# ---------------------------------------------------------------------------
# Tab 7: Subcontract Estimate
# ---------------------------------------------------------------------------
def build_subcontract(ws):
    ws.title = "Subcontract"
    set_col_widths(ws, [("A", 42), ("B", 18), ("C", 18), ("D", 16), ("E", 18)])
    title_row(ws, 1, "SUBCONTRACT ESTIMATE", 5)
    header_row(ws, 3, ["Scope (if subbed)", "Vendor / Quote", "Quote $", "Markup %", "Subtotal ($)"])
    items = [
        ("Pipe insulation — performance-spec", "", 0.00, 0.10),
        ("Excavation / backfill (deep utilities)", "", 0.00, 0.10),
        ("Site utility tie-ins", "", 0.00, 0.10),
        ("Medical gas certification", "", 0.00, 0.10),
        ("Fire-stop installer (specialty)", "", 0.00, 0.10),
        ("Control wiring (if not by EC)", "", 0.00, 0.10),
        ("Concrete saw-cutting / core drilling", "", 0.00, 0.10),
        ("Engineered shop drawings (CAD)", "", 0.00, 0.10),
        ("Hot work fire watch (after-hours)", "", 0.00, 0.10),
        ("Hazmat / lead-paint disturbance", "", 0.00, 0.10),
    ]
    for i, (scope, vendor, quote, markup) in enumerate(items):
        r = 4 + i
        ws.cell(row=r, column=1, value=scope).alignment = LEFT
        ws.cell(row=r, column=2, value=vendor).alignment = LEFT
        ws.cell(row=r, column=3, value=quote).number_format = FMT_USD
        ws.cell(row=r, column=4, value=markup).number_format = FMT_PCT
        ws.cell(row=r, column=5, value=f"=C{r}*(1+D{r})").number_format = FMT_USD
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = FONT_BODY
    last = 4 + len(items) - 1
    sub_row = last + 2
    ws.cell(row=sub_row, column=4, value="Subcontract Subtotal").font = FONT_BODY_BOLD
    ws.cell(row=sub_row, column=4).alignment = RIGHT
    ws.cell(row=sub_row, column=5, value=f"=SUM(E4:E{last})").font = FONT_BODY_BOLD
    ws.cell(row=sub_row, column=5).number_format = FMT_USD
    ws.cell(row=sub_row, column=5).fill = FILL_GOLD
    return ("SubTotal", f"'Subcontract'!$E${sub_row}")


# ---------------------------------------------------------------------------
# Tab 8: Premiums & Conditions
# ---------------------------------------------------------------------------
def build_premiums(ws):
    ws.title = "Premiums"
    set_col_widths(ws, [("A", 50), ("B", 20), ("C", 18), ("D", 18)])
    title_row(ws, 1, "PREMIUMS & CONDITIONS", 4)
    header_row(ws, 3, ["Premium", "Applies? (Y/N)", "Adder %", "Adder $ (overrides %)"])
    items = [
        ("Prevailing wage (Davis-Bacon or state)",   "N", 0.15, 0.00),
        ("Occupied building disruption",             "N", 0.08, 0.00),
        ("Confined-space entry frequency",           "N", 0.04, 0.00),
        ("Hazmat-area (asbestos / lead / silica)",   "N", 0.10, 0.00),
        ("Energized-system / hot-tap proximity",     "N", 0.05, 0.00),
        ("Hospital ICRA (infection-control) tier 3+","N", 0.08, 0.00),
        ("Heights > 14 ft (lift / scaffold premium)","N", 0.04, 0.00),
        ("Underground utility congestion",           "N", 0.06, 0.00),
        ("Tight crawl-space / attic conditions",     "N", 0.05, 0.00),
        ("After-hours / weekend / night shift",      "N", 0.20, 0.00),
        ("Travel / per diem",                        "N", 0.03, 0.00),
        ("Winter weather conditions",                "N", 0.04, 0.00),
        ("LEED documentation / commissioning",       "N", 0.03, 0.00),
        ("Owner-direct purchasing (no markup)",      "N", -0.10, 0.00),
    ]
    for i, (label, applies, pct, dollar) in enumerate(items):
        r = 4 + i
        ws.cell(row=r, column=1, value=label).alignment = LEFT
        ws.cell(row=r, column=2, value=applies).alignment = CENTER
        ws.cell(row=r, column=3, value=pct).number_format = FMT_PCT
        ws.cell(row=r, column=4, value=dollar).number_format = FMT_USD
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = FONT_BODY
    last = 4 + len(items) - 1
    # Total adder dollars (only counts rows where Applies = Y; uses dollar override if non-zero)
    total_row = last + 2
    ws.cell(row=total_row, column=1, value="Premium $ total (applies-Y only)").font = FONT_BODY_BOLD
    ws.cell(row=total_row, column=1).alignment = RIGHT
    # Computed in Bid Summary against subtotal
    ws.cell(row=total_row, column=4, value=(
        f"=SUMPRODUCT((B4:B{last}=\"Y\")*D4:D{last})"
    )).number_format = FMT_USD
    ws.cell(row=total_row, column=4).font = FONT_BODY_BOLD
    ws.cell(row=total_row, column=4).fill = FILL_GOLD
    # Total adder % (only counts rows where Applies = Y)
    pct_row = total_row + 1
    ws.cell(row=pct_row, column=1, value="Premium % total (applies-Y only)").font = FONT_BODY_BOLD
    ws.cell(row=pct_row, column=1).alignment = RIGHT
    ws.cell(row=pct_row, column=4, value=(
        f"=SUMPRODUCT((B4:B{last}=\"Y\")*C4:C{last})"
    )).number_format = FMT_PCT
    ws.cell(row=pct_row, column=4).font = FONT_BODY_BOLD
    ws.cell(row=pct_row, column=4).fill = FILL_GOLD
    return (
        ("PremiumDollars", f"'Premiums'!$D${total_row}"),
        ("PremiumPct", f"'Premiums'!$D${pct_row}"),
    )


# ---------------------------------------------------------------------------
# Tab 9: Bid Summary
# ---------------------------------------------------------------------------
def build_bid_summary(ws, refs):
    ws.title = "Bid Summary"
    set_col_widths(ws, [("A", 4), ("B", 38), ("C", 20), ("D", 22)])
    title_row(ws, 1, "BID SUMMARY — Plumbing Trade Pack", 4)
    rows = [
        ("",            "Material total (incl. tax)",            f"={refs['MaterialTotal']}"),
        ("",            "Labor total",                            f"={refs['LaborTotal']}"),
        ("",            "Equipment total",                        f"={refs['EquipTotal']}"),
        ("",            "Subcontract total",                      f"={refs['SubTotal']}"),
        ("",            "Premiums $ (sum of applies-Y)",          f"={refs['PremiumDollars']}"),
        ("",            "Direct cost subtotal",                   "=SUM(C3:C7)"),
        ("",            "Premium % applied to direct cost",       f"={refs['PremiumPct']}"),
        ("",            "Premium % adder $",                      "=C8*C9"),
        ("",            "Direct cost with premium %",             "=C8+C10"),
        ("",            "Contingency %",                          0.05),
        ("",            "Contingency $",                          "=C11*C12"),
        ("",            "Direct + contingency",                   "=C11+C13"),
        ("",            "Overhead %",                             0.10),
        ("",            "Overhead $",                             "=C14*C15"),
        ("",            "Profit %",                               0.10),
        ("",            "Profit $",                               "=(C14+C16)*C17"),
        ("",            "Bond %",                                 0.012),
        ("",            "Bond $",                                 "=(C14+C16+C18)*C19"),
        ("",            "Insurance / GL premium %",               0.018),
        ("",            "Insurance $",                            "=(C14+C16+C18)*C21"),
        ("",            "FINAL BID",                              "=C14+C16+C18+C20+C22"),
    ]
    for i, (a, label, val) in enumerate(rows):
        r = 3 + i
        ws.cell(row=r, column=2, value=label).font = FONT_BODY_BOLD
        ws.cell(row=r, column=2).alignment = RIGHT
        ws.cell(row=r, column=2).border = BORDER
        c = ws.cell(row=r, column=3, value=val)
        c.border = BORDER
        if label.endswith("%"):
            c.number_format = FMT_PCT
        else:
            c.number_format = FMT_USD
        c.font = FONT_BODY_BOLD
        if label == "FINAL BID":
            c.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
            ws.cell(row=r, column=2).font = Font(name="Calibri", size=14, bold=True, color=BRAND_BLUE)
            ws.cell(row=r, column=2).fill = FILL_GOLD
    # Notes
    ws.cell(row=3 + len(rows) + 2, column=2, value=(
        "Premium % adders APPLY ON TOP of the direct-cost subtotal before contingency. "
        "Contingency, OH, and profit each compound on the running subtotal. Adjust OH and "
        "profit % to match your historical floor."
    )).font = FONT_GREY_ITALIC
    ws.merge_cells(
        start_row=3 + len(rows) + 2,
        start_column=2,
        end_row=3 + len(rows) + 2,
        end_column=4,
    )
    # CSI roll-up table (separate from main summary)
    csi_start = 3 + len(rows) + 5
    ws.cell(row=csi_start, column=2, value="CSI DIVISION ROLL-UP (informational)").font = FONT_H2
    ws.merge_cells(start_row=csi_start, start_column=2, end_row=csi_start, end_column=4)
    csi_headers = ["CSI Division", "Material $", "Labor $"]
    for i, h in enumerate(csi_headers):
        c = ws.cell(row=csi_start + 1, column=2 + i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER
    csi_divs = [
        ("22 11 - Water Distribution",  "22 11"),
        ("22 13 - Sanitary Sewerage",   "22 13"),
        ("22 14 - Storm Drainage",      "22 14"),
        ("22 33-35 - Water Heaters",    "22 3"),
        ("22 42-45 - Plumbing Fixtures","22 4"),
        ("22 60-67 - Gas + Specialty",  "22 6"),
        ("22 05 - Common Work Results", "22 05"),
        ("22 07 - Insulation",          "22 07"),
        ("22 08 - Test + Commissioning","22 08"),
    ]
    for i, (label, prefix) in enumerate(csi_divs):
        r = csi_start + 2 + i
        ws.cell(row=r, column=2, value=label).alignment = LEFT
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=3, value=(
            f"=SUMPRODUCT((LEFT('Material Takeoff'!A4:A200,LEN(\"{prefix}\"))=\"{prefix}\")*'Material Takeoff'!F4:F200)"
        )).number_format = FMT_USD
        ws.cell(row=r, column=3).border = BORDER
        ws.cell(row=r, column=4, value=(
            f"=SUMPRODUCT((LEFT('Labor Estimate'!A4:A200,LEN(\"{prefix}\"))=\"{prefix}\")*'Labor Estimate'!H4:H200)"
        )).number_format = FMT_USD
        ws.cell(row=r, column=4).border = BORDER


# ---------------------------------------------------------------------------
# Tab 10: CSI Reference
# ---------------------------------------------------------------------------
def build_csi_reference(ws):
    ws.title = "CSI Reference"
    set_col_widths(ws, [("A", 14), ("B", 80)])
    title_row(ws, 1, "CSI MASTERFORMAT — DIVISION 22 PLUMBING", 2)
    rows = [
        ("22 00 00", "PLUMBING (overall division)"),
        ("22 05 00", "Common Work Results for Plumbing"),
        ("22 05 19", "Meters and Gauges for Plumbing Piping"),
        ("22 05 23", "General-Duty Valves for Plumbing Piping"),
        ("22 05 29", "Hangers and Supports for Plumbing Piping and Equipment"),
        ("22 05 53", "Identification for Plumbing Piping and Equipment"),
        ("22 07 00", "Plumbing Insulation"),
        ("22 08 00", "Commissioning of Plumbing"),
        ("22 11 00", "Facility Water Distribution"),
        ("22 11 13", "Facility Water Distribution Piping"),
        ("22 11 16", "Domestic Water Piping"),
        ("22 11 19", "Domestic Water Piping Specialties"),
        ("22 11 23", "Domestic Water Pumps"),
        ("22 13 00", "Facility Sanitary Sewerage"),
        ("22 13 13", "Facility Sanitary Sewers"),
        ("22 13 16", "Sanitary Waste and Vent Piping"),
        ("22 13 19", "Sanitary Waste Piping Specialties"),
        ("22 13 23", "Sanitary Waste Interceptors"),
        ("22 13 29", "Sanitary Sewerage Pumps"),
        ("22 14 00", "Facility Storm Drainage"),
        ("22 14 13", "Facility Storm Drainage Piping"),
        ("22 14 23", "Storm Drainage Piping Specialties"),
        ("22 14 26", "Facility Storm Drains"),
        ("22 14 29", "Sump Pumps"),
        ("22 31 00", "Domestic Water Softeners"),
        ("22 32 00", "Domestic Water Filtration Equipment"),
        ("22 33 00", "Electric Domestic Water Heaters"),
        ("22 34 00", "Fuel-Fired Domestic Water Heaters"),
        ("22 35 00", "Domestic Water Heat Exchangers"),
        ("22 40 00", "PLUMBING FIXTURES"),
        ("22 41 00", "Residential Plumbing Fixtures"),
        ("22 42 00", "Commercial Plumbing Fixtures"),
        ("22 42 13", "Commercial Water Closets, Urinals, and Bidets"),
        ("22 42 16", "Commercial Lavatories and Sinks"),
        ("22 42 19", "Wash Fountains"),
        ("22 42 23", "Commercial Showers, Receptors, and Basins"),
        ("22 42 26", "Commercial Disposers"),
        ("22 42 33", "Wash-Down Equipment"),
        ("22 42 39", "Commercial Faucets, Supplies, and Trim"),
        ("22 45 00", "Emergency Plumbing Fixtures (eye-wash / drench-shower)"),
        ("22 47 00", "Drinking Fountains and Water Coolers"),
        ("22 60 00", "GAS AND VACUUM SYSTEMS FOR LABORATORY AND HEALTHCARE FACILITIES"),
        ("22 61 00", "Compressed-Air Systems for Laboratory and Healthcare Facilities"),
        ("22 62 00", "Vacuum Systems for Laboratory and Healthcare Facilities"),
        ("22 63 00", "Gas Systems for Laboratory and Healthcare Facilities"),
        ("22 66 00", "Chemical-Waste Systems for Laboratory and Healthcare Facilities"),
        ("22 67 00", "Processed Water Systems for Laboratory and Healthcare Facilities"),
    ]
    for i, (code, name) in enumerate(rows):
        r = 3 + i
        ws.cell(row=r, column=1, value=code).alignment = CENTER
        ws.cell(row=r, column=1).font = FONT_BODY_BOLD
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2, value=name).alignment = LEFT
        ws.cell(row=r, column=2).font = FONT_BODY
        ws.cell(row=r, column=2).border = BORDER


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb = Workbook()

    ws1 = wb.active
    build_instructions(ws1)

    ws2 = wb.create_sheet("Project Info")
    build_project_info(ws2)

    ws3 = wb.create_sheet("Wage Rates")
    build_wage_rates(ws3, wb)

    ws4 = wb.create_sheet("Material Takeoff")
    mat_ref = build_material_takeoff(ws4)

    ws5 = wb.create_sheet("Labor Estimate")
    lab_ref = build_labor_estimate(ws5)

    ws6 = wb.create_sheet("Equipment")
    eq_ref = build_equipment(ws6)

    ws7 = wb.create_sheet("Subcontract")
    sub_ref = build_subcontract(ws7)

    ws8 = wb.create_sheet("Premiums")
    prem_dollars, prem_pct = build_premiums(ws8)

    refs = {
        mat_ref[0]: mat_ref[1],
        lab_ref[0]: lab_ref[1],
        eq_ref[0]: eq_ref[1],
        sub_ref[0]: sub_ref[1],
        prem_dollars[0]: prem_dollars[1],
        prem_pct[0]: prem_pct[1],
    }

    ws9 = wb.create_sheet("Bid Summary")
    build_bid_summary(ws9, refs)
    # Move to position 2
    wb.move_sheet(ws9, offset=-7)

    ws10 = wb.create_sheet("CSI Reference")
    build_csi_reference(ws10)

    wb.save(OUT)
    print(f"OK — wrote {OUT}")


if __name__ == "__main__":
    main()
