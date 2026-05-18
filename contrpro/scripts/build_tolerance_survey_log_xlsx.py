#!/usr/bin/env python3
"""
Build ContrPro Tolerance Survey Log (XLSX) — Steel Erection trade pack.

Companion to plumb-and-true-tolerance-guide.html. Captures the per-column
plumb survey, anchor-rod survey, floor-elevation survey, and out-of-tolerance
deviation tracker required for AISC §7.13 compliance demonstration.

Tabs:
  1. Instructions
  2. Project Info
  3. Anchor Rod Survey          (per anchor — set-out vs as-found + variance)
  4. Column Plumb Log           (per column per survey event — deviation X and Y)
  5. Floor Elevation Log        (per floor + corner — beam-top elevation vs theoretical)
  6. Deviation / RFI Tracker    (out-of-tolerance items requiring RFI to EOR)
  7. Daily Survey Summary       (auto-aggregated counts + status)
  8. Tolerance Reference        (hidden — §7.13 limits driving the OK/Out flags)

Run:
    /Users/home/charles/.venv/bin/python3 \\
        /Users/home/charles/contrpro/scripts/build_tolerance_survey_log_xlsx.py

Output:
    /Users/home/charles/contrpro/files/packages/complete/steel-erection/Tolerance_Survey_Log.xlsx
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

BRAND_BLUE = "1E3A5F"
BRAND_BLUE_LIGHT = "D6E0EC"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
YELLOW_FILL = "FFEB9C"
YELLOW_FONT = "9C5700"
GREY_TEXT = "808080"

FILL_HEADER = PatternFill("solid", fgColor=BRAND_BLUE)
FILL_SUBHEADER = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_SUMMARY_LABEL = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW_FILL)

FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=BRAND_BLUE)
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=BRAND_BLUE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_BIG_NUMBER = Font(name="Calibri", size=18, bold=True, color=BRAND_BLUE)
FONT_GREY_ITALIC = Font(name="Calibri", size=11, italic=True, color=GREY_TEXT)

THIN = Side(border_style="thin", color="B0B7BF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_INT = "0"
FMT_NUM2 = "0.00"
FMT_NUM3 = "0.000"
FMT_IN = '0.000"\\""'
FMT_FT = '#,##0.0" ft"'
FMT_DATE = "yyyy-mm-dd"


def set_col_widths(ws, widths):
    for col, w in widths:
        ws.column_dimensions[col].width = w


def style_header_row(ws, row, cols, start_col=1):
    for c in range(start_col, start_col + cols):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


def apply_body_style(ws, row, cols, start_col=1):
    for c in range(start_col, start_col + cols):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_BODY
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.border = BORDER


# ---------------------------------------------------------------------------
# Tab 1: Instructions
# ---------------------------------------------------------------------------

def build_instructions(ws):
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 110)])

    ws["A1"] = "TOLERANCE SURVEY LOG — Instructions"
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 36

    rows = [
        "",
        "WHAT THIS WORKBOOK IS",
        "Daily survey + deviation log for AISC §7.13 tolerance compliance. Pair with the Plumb-and-True "
        "Tolerance Guide HTML. The workbook auto-flags OK / Out-of-Tolerance based on §7.13 limits.",
        "",
        "WORKFLOW",
        "1. Project Info: project, GC, erector, surveyor, EOR, project-spec tolerance (default §7.13 OR "
        "tighter per Contract).",
        "2. Anchor Rod Survey: log every anchor — theoretical and as-found positions. Variance auto-"
        "computes. Out-of-tolerance items auto-flag.",
        "3. Column Plumb Log: log every column at every plumb-up event (initial, bay plumb-up, top-out). "
        "Enter offset at top of column in X and Y from theoretical plumb. Compare to 1:500 limit (or "
        "project-specific limit).",
        "4. Floor Elevation Log: log beam-top elevations vs theoretical. Multi-point per floor (typically "
        "4 corners + center).",
        "5. Deviation / RFI Tracker: any out-of-tolerance item that requires EOR review gets a row here. "
        "Track RFI #, EOR response, resolution, re-survey confirmation.",
        "6. Daily Survey Summary: auto-totals — surveys completed today, deviations open, RFIs pending.",
        "",
        "AISC §7.13 KEY LIMITS (built into the workbook)",
        "  Column plumb: 1:500 of column length (with per-length table on Tolerance Reference tab)",
        "  Anchor rod position (single): ±1/8\"",
        "  Anchor rod pattern position: ±1/4\"",
        "  Anchor rod stickup: ±1/8\"",
        "  Beam top elevation: ±1/8\"",
        "  Floor-to-floor height: ±1/4\"",
        "",
        "TIGHTER PROJECT-SPEC TOLERANCES",
        "If the Contract Documents specify tighter tolerances (AESS, hospital, MRI, etc.), override the "
        "default on Project Info. The workbook will apply the tighter limit when flagging out-of-"
        "tolerance items.",
        "",
        "SURVEY EVENT TYPES",
        "  - Initial: column just set, before next beam frames to it",
        "  - Bay plumb-up: after 2-bay placement, before next bays",
        "  - Top-out: after last column set on this elevation",
        "  - Pre-cladding: as-built before curtainwall sub mobilizes",
        "  - Final / Substantial Completion: archive record",
        "",
        "DEVIATION RESOLUTION",
        "Any out-of-tolerance column is either (a) re-plumbed via diagonal bracing (if column still "
        "adjustable) and re-surveyed, OR (b) RFI'd to EOR. The Deviation Tracker handles both paths.",
        "",
        "AS-BUILT SIGN-OFF",
        "At Substantial Completion, the workbook (plus the EOR's acceptance letter) is the evidence file "
        "for the cladding mobilization gate and the retainage reduction request.",
    ]

    for i, text in enumerate(rows, start=2):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if text and text == text.upper() and len(text) < 70:
            cell.font = FONT_H2
        else:
            cell.font = FONT_BODY


# ---------------------------------------------------------------------------
# Tab 2: Project Info
# ---------------------------------------------------------------------------

def build_project_info(ws):
    ws.title = "Project Info"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 42), ("C", 36)])

    ws["B1"] = "TOLERANCE SURVEY — PROJECT INFO"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:C1")
    ws.row_dimensions[1].height = 32

    info_rows = [
        ("Project Name", ""),
        ("Project Address", ""),
        ("Owner", ""),
        ("Controlling Contractor (GC)", ""),
        ("Steel Erector", ""),
        ("EOR (Structural Engineer of Record)", ""),
        ("Surveyor (in-house or sub)", ""),
        ("Total Station Make / Model", ""),
        ("Total Station Calibration Date", ""),
        ("Benchmark Reference Monument 1 (location)", ""),
        ("Benchmark Reference Monument 2 (location)", ""),
        ("", ""),
        ("PROJECT-SPEC TOLERANCE OVERRIDES (leave default = §7.13 if not tightened)", ""),
        ("Column plumb ratio (default 500 → 1:500)", 500),
        ("Anchor rod single position tolerance (in)", 0.125),
        ("Anchor rod pattern position tolerance (in)", 0.250),
        ("Anchor rod stickup tolerance (in)", 0.125),
        ("Beam top elevation tolerance (in)", 0.125),
        ("Floor-to-floor height tolerance (in)", 0.250),
        ("Exterior column drift envelope (in, at corners)", 1.000),
    ]

    name_map = {
        "ProjectName": "$C$3",
        "Erector": "$C$7",
        "EOR": "$C$8",
        "PlumbRatio": "$C$16",
        "AnchorSingleTol": "$C$17",
        "AnchorPatternTol": "$C$18",
        "AnchorStickupTol": "$C$19",
        "BeamElevTol": "$C$20",
        "FloorHeightTol": "$C$21",
        "DriftEnvelope": "$C$22",
    }
    for nm, ref in name_map.items():
        ws.parent.defined_names[nm] = DefinedName(name=nm, attr_text=f"'Project Info'!{ref}")

    for i, (label, val) in enumerate(info_rows, start=3):
        is_subhead = label and label == label.upper() and not label.startswith(("Project ", "Owner", "Controlling", "Steel ", "EOR ", "Surveyor", "Total ", "Benchmark"))
        if is_subhead:
            ws.cell(row=i, column=2, value=label).font = FONT_H2
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
            continue
        if label:
            ws.cell(row=i, column=2, value=label).font = FONT_BODY_BOLD
            ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
            ws.cell(row=i, column=2).border = BORDER
            c = ws.cell(row=i, column=3, value=val)
            c.font = FONT_BODY
            c.border = BORDER
            if isinstance(val, float):
                c.number_format = FMT_NUM3
            if "Date" in label:
                c.number_format = FMT_DATE


# ---------------------------------------------------------------------------
# Tab 3: Anchor Rod Survey
# ---------------------------------------------------------------------------

ANCHOR_COLS = [
    ("A", "Column ID", 12),
    ("B", "Anchor Rod #", 11),
    ("C", "Theo X (ft)", 12),
    ("D", "Theo Y (ft)", 12),
    ("E", "As-Found X (ft)", 14),
    ("F", "As-Found Y (ft)", 14),
    ("G", "ΔX (in)", 10),
    ("H", "ΔY (in)", 10),
    ("I", "Radial Dev (in)", 13),
    ("J", "Theo Stickup (in)", 14),
    ("K", "As-Found Stickup (in)", 16),
    ("L", "Stickup Dev (in)", 13),
    ("M", "Single-Rod Position OK?", 16),
    ("N", "Stickup OK?", 12),
    ("O", "Notes / RFI", 22),
]


def build_anchor_survey(ws):
    ws.title = "Anchor Rod Survey"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in ANCHOR_COLS])

    ws["A1"] = "ANCHOR ROD SURVEY"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:O1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Pre-erection verification of anchor rods. ΔX/ΔY are computed in inches; radial deviation is the "
        "absolute distance from theoretical. Single-Rod tolerance default = ±1/8\" (per AISC §7.5.1)."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:O2")

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(ANCHOR_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(ANCHOR_COLS))

    DATA_ROWS = 100
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        for c in (3, 4, 5, 6):
            ws.cell(row=r, column=c).number_format = FMT_NUM3
        for c in (7, 8, 9, 10, 11, 12):
            ws.cell(row=r, column=c).number_format = FMT_NUM3
        # ΔX = (As-Found X - Theo X) * 12  → inches
        ws.cell(row=r, column=7, value=f'=IFERROR((E{r}-C{r})*12,"")').number_format = FMT_NUM3
        ws.cell(row=r, column=8, value=f'=IFERROR((F{r}-D{r})*12,"")').number_format = FMT_NUM3
        # Radial deviation
        ws.cell(
            row=r,
            column=9,
            value=f'=IFERROR(SQRT(G{r}^2+H{r}^2),"")',
        ).number_format = FMT_NUM3
        # Stickup deviation
        ws.cell(row=r, column=12, value=f'=IFERROR(K{r}-J{r},"")').number_format = FMT_NUM3
        # Position OK
        ws.cell(
            row=r,
            column=13,
            value=(
                f'=IF(OR(G{r}="",H{r}=""),"",'
                f'IF(AND(ABS(G{r})<=AnchorSingleTol,ABS(H{r})<=AnchorSingleTol),"OK","OUT"))'
            ),
        )
        # Stickup OK
        ws.cell(
            row=r,
            column=14,
            value=(
                f'=IF(L{r}="","",'
                f'IF(ABS(L{r})<=AnchorStickupTol,"OK","OUT"))'
            ),
        )
        apply_body_style(ws, r, len(ANCHOR_COLS))

    # Conditional format Position OK
    for col_letter in ("M", "N"):
        ws.conditional_formatting.add(
            f"{col_letter}{HEADER_ROW+1}:{col_letter}{HEADER_ROW+DATA_ROWS}",
            FormulaRule(formula=[f'{col_letter}{HEADER_ROW+1}="OK"'], fill=FILL_GREEN, font=Font(bold=True, color=GREEN_FONT)),
        )
        ws.conditional_formatting.add(
            f"{col_letter}{HEADER_ROW+1}:{col_letter}{HEADER_ROW+DATA_ROWS}",
            FormulaRule(formula=[f'{col_letter}{HEADER_ROW+1}="OUT"'], fill=FILL_RED, font=Font(bold=True, color=RED_FONT)),
        )


# ---------------------------------------------------------------------------
# Tab 4: Column Plumb Log
# ---------------------------------------------------------------------------

PLUMB_COLS = [
    ("A", "Survey Date", 12),
    ("B", "Event Type", 18),
    ("C", "Column ID", 12),
    ("D", "Gridline", 10),
    ("E", "Column Length (ft)", 13),
    ("F", "Offset X at Top (in)", 16),
    ("G", "Offset Y at Top (in)", 16),
    ("H", "Total Radial Dev (in)", 14),
    ("I", "Allowable (1:Ratio of length, in)", 16),
    ("J", "Plumb OK?", 11),
    ("K", "Surveyor", 16),
    ("L", "Adjusted Y/N", 11),
    ("M", "Re-Survey Date", 13),
    ("N", "Notes / RFI", 22),
]


def build_plumb_log(ws):
    ws.title = "Column Plumb Log"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in PLUMB_COLS])

    ws["A1"] = "COLUMN PLUMB LOG"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:N1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Offsets measured at top of column from theoretical plumb vertical. Column Length is from top of "
        "base plate to top of column (or to top of splice on multi-story). Allowable = ColumnLength * 12 / "
        "PlumbRatio (default 1:500)."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:N2")
    ws.row_dimensions[2].height = 30

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(PLUMB_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(PLUMB_COLS))

    DATA_ROWS = 200
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        ws.cell(row=r, column=1).number_format = FMT_DATE
        ws.cell(row=r, column=5).number_format = FMT_FT
        for c in (6, 7, 8, 9):
            ws.cell(row=r, column=c).number_format = FMT_NUM3
        # Total radial
        ws.cell(
            row=r,
            column=8,
            value=f'=IFERROR(SQRT(F{r}^2+G{r}^2),"")',
        ).number_format = FMT_NUM3
        # Allowable in inches = (length_ft * 12) / PlumbRatio
        ws.cell(
            row=r,
            column=9,
            value=f'=IFERROR((E{r}*12)/PlumbRatio,"")',
        ).number_format = FMT_NUM3
        # OK?
        ws.cell(
            row=r,
            column=10,
            value=(
                f'=IF(OR(H{r}="",I{r}=""),"",'
                f'IF(H{r}<=I{r},"OK","OUT"))'
            ),
        )
        ws.cell(row=r, column=13).number_format = FMT_DATE
        apply_body_style(ws, r, len(PLUMB_COLS))

    dv_event = DataValidation(
        type="list",
        formula1='"Initial,Bay Plumb-Up,Top-Out,Pre-Cladding,Final / Substantial Completion"',
        allow_blank=True,
    )
    dv_event.add(f"B{HEADER_ROW+1}:B{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_event)

    dv_yn = DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=True)
    dv_yn.add(f"L{HEADER_ROW+1}:L{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_yn)

    # Conditional formatting
    ws.conditional_formatting.add(
        f"J{HEADER_ROW+1}:J{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'J{HEADER_ROW+1}="OK"'], fill=FILL_GREEN, font=Font(bold=True, color=GREEN_FONT)),
    )
    ws.conditional_formatting.add(
        f"J{HEADER_ROW+1}:J{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'J{HEADER_ROW+1}="OUT"'], fill=FILL_RED, font=Font(bold=True, color=RED_FONT)),
    )


# ---------------------------------------------------------------------------
# Tab 5: Floor Elevation Log
# ---------------------------------------------------------------------------

FLOOR_COLS = [
    ("A", "Survey Date", 12),
    ("B", "Floor Level", 12),
    ("C", "Survey Point ID", 14),
    ("D", "Location", 18),
    ("E", "Theo Elevation (ft)", 14),
    ("F", "As-Found Elevation (ft)", 16),
    ("G", "Deviation (in)", 12),
    ("H", "Beam Top OK?", 13),
    ("I", "Adjacent Beam Match (in)", 16),
    ("J", "Beam Match OK?", 13),
    ("K", "Surveyor", 16),
    ("L", "Notes / RFI", 22),
]


def build_floor_log(ws):
    ws.title = "Floor Elevation Log"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in FLOOR_COLS])

    ws["A1"] = "FLOOR / BEAM ELEVATION LOG"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:L1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Beam-top elevations per floor. Recommended survey points: each corner + center. Theoretical "
        "from S-series drawings. Deviation auto-flags against ±1/8\" beam-top tolerance (default §7.13.1.4)."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:L2")

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(FLOOR_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(FLOOR_COLS))

    DATA_ROWS = 150
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        ws.cell(row=r, column=1).number_format = FMT_DATE
        for c in (5, 6):
            ws.cell(row=r, column=c).number_format = FMT_NUM3
        # Deviation in inches = (as-found - theo) * 12
        ws.cell(row=r, column=7, value=f'=IFERROR((F{r}-E{r})*12,"")').number_format = FMT_NUM3
        # OK?
        ws.cell(
            row=r,
            column=8,
            value=f'=IF(G{r}="","",IF(ABS(G{r})<=BeamElevTol,"OK","OUT"))',
        )
        ws.cell(row=r, column=9).number_format = FMT_NUM3
        # Beam match OK
        ws.cell(
            row=r,
            column=10,
            value=f'=IF(I{r}="","",IF(ABS(I{r})<=BeamElevTol,"OK","OUT"))',
        )
        apply_body_style(ws, r, len(FLOOR_COLS))

    # Conditional formatting
    for col_letter in ("H", "J"):
        ws.conditional_formatting.add(
            f"{col_letter}{HEADER_ROW+1}:{col_letter}{HEADER_ROW+DATA_ROWS}",
            FormulaRule(formula=[f'{col_letter}{HEADER_ROW+1}="OK"'], fill=FILL_GREEN, font=Font(bold=True, color=GREEN_FONT)),
        )
        ws.conditional_formatting.add(
            f"{col_letter}{HEADER_ROW+1}:{col_letter}{HEADER_ROW+DATA_ROWS}",
            FormulaRule(formula=[f'{col_letter}{HEADER_ROW+1}="OUT"'], fill=FILL_RED, font=Font(bold=True, color=RED_FONT)),
        )


# ---------------------------------------------------------------------------
# Tab 6: Deviation / RFI Tracker
# ---------------------------------------------------------------------------

DEV_COLS = [
    ("A", "Dev #", 8),
    ("B", "Discovered Date", 13),
    ("C", "Survey Reference", 18),
    ("D", "Item Type", 14),
    ("E", "Location / Column ID", 18),
    ("F", "Description", 32),
    ("G", "Measured", 12),
    ("H", "Allowable", 12),
    ("I", "RFI #", 10),
    ("J", "Sent to EOR (Date)", 14),
    ("K", "EOR Response Date", 14),
    ("L", "EOR Disposition", 26),
    ("M", "Resolution Date", 14),
    ("N", "Resolved Status", 14),
    ("O", "Notes", 22),
]


def build_deviation_tracker(ws):
    ws.title = "Deviation RFI Tracker"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [(c, w) for c, _, w in DEV_COLS])

    ws["A1"] = "DEVIATION / RFI TRACKER"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:O1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "One row per out-of-tolerance item requiring EOR review. Tracks the RFI through to resolution."
    )
    ws["A2"].font = FONT_GREY_ITALIC
    ws.merge_cells("A2:O2")

    HEADER_ROW = 4
    for i, (col, header, _) in enumerate(DEV_COLS, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=header)
    style_header_row(ws, HEADER_ROW, len(DEV_COLS))

    DATA_ROWS = 50
    for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + DATA_ROWS):
        ws.cell(row=r, column=1).number_format = FMT_INT
        for c in (2, 10, 11, 13):
            ws.cell(row=r, column=c).number_format = FMT_DATE
        apply_body_style(ws, r, len(DEV_COLS))

    dv_type = DataValidation(
        type="list",
        formula1='"Anchor Rod,Column Plumb,Beam Elevation,Floor Height,Drift,Member Damage,Other"',
        allow_blank=True,
    )
    dv_type.add(f"D{HEADER_ROW+1}:D{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_type)

    dv_disp = DataValidation(
        type="list",
        formula1='"Accept as-is,Field correct,Re-erect,Reinforce,Replace member,Pending"',
        allow_blank=True,
    )
    dv_disp.add(f"L{HEADER_ROW+1}:L{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_disp)

    dv_status = DataValidation(
        type="list",
        formula1='"Open,RFI Sent,EOR Review,Action Pending,Resolved,Closed"',
        allow_blank=True,
    )
    dv_status.add(f"N{HEADER_ROW+1}:N{HEADER_ROW+DATA_ROWS}")
    ws.add_data_validation(dv_status)

    # Conditional formatting
    ws.conditional_formatting.add(
        f"N{HEADER_ROW+1}:N{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'OR(N{HEADER_ROW+1}="Resolved",N{HEADER_ROW+1}="Closed")'], fill=FILL_GREEN, font=Font(bold=True, color=GREEN_FONT)),
    )
    ws.conditional_formatting.add(
        f"N{HEADER_ROW+1}:N{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'OR(N{HEADER_ROW+1}="Open",N{HEADER_ROW+1}="RFI Sent")'], fill=FILL_RED, font=Font(bold=True, color=RED_FONT)),
    )
    ws.conditional_formatting.add(
        f"N{HEADER_ROW+1}:N{HEADER_ROW+DATA_ROWS}",
        FormulaRule(formula=[f'OR(N{HEADER_ROW+1}="EOR Review",N{HEADER_ROW+1}="Action Pending")'], fill=FILL_YELLOW, font=Font(color=YELLOW_FONT)),
    )


# ---------------------------------------------------------------------------
# Tab 7: Daily Survey Summary
# ---------------------------------------------------------------------------

def build_daily_summary(ws):
    ws.title = "Daily Summary"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 42), ("C", 18)])

    ws["B1"] = "DAILY SURVEY SUMMARY"
    ws["B1"].font = FONT_TITLE
    ws.merge_cells("B1:C1")
    ws.row_dimensions[1].height = 32

    ws["B2"] = "Auto-aggregated. Use for monthly billing back-up + as-built sign-off package."
    ws["B2"].font = FONT_GREY_ITALIC
    ws.merge_cells("B2:C2")

    rows = [
        ("ANCHOR ROD SURVEY", ""),
        ("Total anchor rods surveyed", '=COUNTA(\'Anchor Rod Survey\'!B5:B104)'),
        ("Anchor rods position OK", '=COUNTIF(\'Anchor Rod Survey\'!M5:M104,"OK")'),
        ("Anchor rods position OUT", '=COUNTIF(\'Anchor Rod Survey\'!M5:M104,"OUT")'),
        ("Anchor rods stickup OK", '=COUNTIF(\'Anchor Rod Survey\'!N5:N104,"OK")'),
        ("Anchor rods stickup OUT", '=COUNTIF(\'Anchor Rod Survey\'!N5:N104,"OUT")'),
        ("", ""),
        ("COLUMN PLUMB", ""),
        ("Total plumb surveys", '=COUNTA(\'Column Plumb Log\'!C5:C204)'),
        ("Columns plumb OK", '=COUNTIF(\'Column Plumb Log\'!J5:J204,"OK")'),
        ("Columns plumb OUT", '=COUNTIF(\'Column Plumb Log\'!J5:J204,"OUT")'),
        ("", ""),
        ("FLOOR ELEVATION", ""),
        ("Total elevation surveys", '=COUNTA(\'Floor Elevation Log\'!C5:C154)'),
        ("Beam top OK", '=COUNTIF(\'Floor Elevation Log\'!H5:H154,"OK")'),
        ("Beam top OUT", '=COUNTIF(\'Floor Elevation Log\'!H5:H154,"OUT")'),
        ("", ""),
        ("DEVIATION TRACKER", ""),
        ("Total deviations logged", '=COUNTA(\'Deviation RFI Tracker\'!A5:A54)'),
        ("Deviations Open", '=COUNTIF(\'Deviation RFI Tracker\'!N5:N54,"Open")'),
        ("Deviations RFI Sent", '=COUNTIF(\'Deviation RFI Tracker\'!N5:N54,"RFI Sent")'),
        ("Deviations EOR Review", '=COUNTIF(\'Deviation RFI Tracker\'!N5:N54,"EOR Review")'),
        ("Deviations Resolved + Closed", '=COUNTIF(\'Deviation RFI Tracker\'!N5:N54,"Resolved")+COUNTIF(\'Deviation RFI Tracker\'!N5:N54,"Closed")'),
        ("", ""),
        ("READINESS FOR CLADDING / RETAINAGE", ""),
        ("Open deviations blocking sign-off", '=COUNTIF(\'Deviation RFI Tracker\'!N5:N54,"Open")+COUNTIF(\'Deviation RFI Tracker\'!N5:N54,"RFI Sent")+COUNTIF(\'Deviation RFI Tracker\'!N5:N54,"EOR Review")+COUNTIF(\'Deviation RFI Tracker\'!N5:N54,"Action Pending")'),
    ]

    start_row = 4
    for i, (label, formula) in enumerate(rows, start=start_row):
        is_subhead = label and label == label.upper() and not formula
        if is_subhead:
            ws.cell(row=i, column=2, value=label).font = FONT_H2
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
            continue
        if label:
            ws.cell(row=i, column=2, value=label).font = FONT_BODY_BOLD
            ws.cell(row=i, column=2).fill = FILL_SUMMARY_LABEL
            ws.cell(row=i, column=2).border = BORDER
            c = ws.cell(row=i, column=3, value=formula)
            c.font = FONT_BIG_NUMBER if "Total" in label or "blocking" in label.lower() else FONT_BODY_BOLD
            c.number_format = FMT_INT
            c.border = BORDER


# ---------------------------------------------------------------------------
# Tab 8: Tolerance Reference (hidden)
# ---------------------------------------------------------------------------

def build_tolerance_reference(ws):
    ws.title = "Tolerance Reference"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 28), ("B", 22), ("C", 22)])

    ws["A1"] = "AISC §7.13 — Tolerance Limits Reference"
    ws["A1"].font = FONT_H2
    ws.merge_cells("A1:C1")

    headers = ["Item", "Default Limit", "AISC §"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=h)
        ws.cell(row=3, column=i).font = FONT_HEADER
        ws.cell(row=3, column=i).fill = FILL_HEADER
        ws.cell(row=3, column=i).alignment = Alignment(horizontal="center")
        ws.cell(row=3, column=i).border = BORDER

    refs = [
        ("Anchor rod single position", '±0.125" (1/8")', "§7.5.1"),
        ("Anchor rod pattern position", '±0.250" (1/4")', "§7.5.1"),
        ("Anchor rod stickup", '±0.125" (1/8")', "§7.5.2"),
        ("Anchor rod plumb (vertical)", "1:40", "§7.5.2"),
        ("Column plumb", "1:500 of length", "§7.13.1.2"),
        ("Beam top elevation", '±0.125" (1/8")', "§7.13.1.4"),
        ("Floor-to-floor height", '±0.250" (1/4")', "§7.13.1.4"),
        ("Building drift (multi-story)", "1:500 from base, exterior columns", "§7.13.1.3"),
        ("Member straightness (mill)", "Per ASTM A6", "ASTM A6"),
        ("Member camber preservation", "Per shop drawings", "Project spec"),
    ]
    for i, (item, limit, sec) in enumerate(refs, start=4):
        ws.cell(row=i, column=1, value=item).font = FONT_BODY
        ws.cell(row=i, column=2, value=limit).font = FONT_BODY
        ws.cell(row=i, column=3, value=sec).font = FONT_BODY
        for col in (1, 2, 3):
            ws.cell(row=i, column=col).border = BORDER

    ws.sheet_state = "hidden"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    wb = Workbook()
    build_instructions(wb.active)
    build_project_info(wb.create_sheet())
    build_anchor_survey(wb.create_sheet())
    build_plumb_log(wb.create_sheet())
    build_floor_log(wb.create_sheet())
    build_deviation_tracker(wb.create_sheet())
    build_daily_summary(wb.create_sheet())
    build_tolerance_reference(wb.create_sheet())

    out_dir = "/Users/home/charles/contrpro/files/packages/complete/steel-erection"
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "Tolerance_Survey_Log.xlsx")
    wb.save(out_path)
    sz = os.path.getsize(out_path)
    print(f"Wrote {out_path} ({sz//1024} KB)")


if __name__ == "__main__":
    main()
